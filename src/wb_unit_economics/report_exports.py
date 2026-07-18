from __future__ import annotations

import csv
import hashlib
import shutil
import subprocess
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from wb_unit_economics.client_report import (
    ClientReportModel,
    build_client_analytical_markdown,
    render_client_report_html,
)
from wb_unit_economics.document_exports import markdown_sha256, render_markdown_docx

EXPORT_SHEETS = {
    "unitRows": "Юнит экономика",
    "liquidityRows": "Ликвидность МД",
    "monthly": "Динамика",
    "expenses": "Расходы WB",
    "returns": "Возвраты",
    "lostSales": "Упущенные продажи",
    "documentReconciliation": "Сверка документов 1С",
    "reconciliationMonthly": "Сверка с 1С ОПиУ",
    "taxInputReconciliation": "Сверка входящего НДС",
}

FULL_EXCEL_SHEETS: list[tuple[str, str]] = [
    ("Дашборд", "dashboard"),
    ("Сводка", "summary"),
    ("Сводка по организациям", "organizationSummary"),
    ("Сводка по кабинетам WB", "cabinetSummary"),
    ("Юнит экономика", "unitRows"),
    ("Ликвидность МД", "liquidityRows"),
    ("Товары", "productSummary"),
    ("Динамика", "monthly"),
    ("Сводный отчет WB", "wbReportSummary"),
    ("WB поля отчета", "unitRows"),
    ("Сверка по отчетам WB", "documentReconciliation"),
    ("Сверка с 1С", "reconciliationMonthly"),
    ("Сверка документов 1С", "documentReconciliation"),
    ("Сверка с 1С ОПиУ", "reconciliationMonthly"),
    ("Валовая прибыль 1С", "reconciliationMonthly"),
    ("Сверка услуг WB", "expenses"),
    ("Расшифровка услуг 1С", "documentReconciliation"),
    ("Распределение расходов", "expenses"),
    ("Товары по отчетам 1С", "documentReconciliation"),
    ("Расходы WB", "expenses"),
    ("Возвраты", "returns"),
    ("Упущенные продажи", "lostSales"),
    ("Себестоимость 1С", "costRows"),
    ("Маппинг", "mappingRows"),
    ("Ошибки данных", "errorRows"),
    ("Сверка входящего НДС", "taxInputReconciliation"),
]

OZON_EXPORT_SHEETS = {
    "ozonSummaryRows": "Сводная Ozon",
    "ozonUnitRows": "Юнит экономика Ozon",
    "ozonServiceChargeRows": "Начисления услуг Ozon",
    "ozonArticleSkuRows": "Статьи по SKU",
    "ozonReconciliationRows": "Сверка Ozon 1C",
}

SHEET_COLUMNS: dict[str, list[str]] = {
    "unitRows": [
        "product",
        "articleWb",
        "article1c",
        "barcode",
        "nmId",
        "scheme",
        "organization",
        "cabinet",
        "week",
        "month",
        "sales",
        "returns",
        "netQty",
        "returnRate",
        "revenueBeforeSpp",
        "spp",
        "sppRate",
        "revenue",
        "pnlRevenue",
        "revenueWithoutVat",
        "cost",
        "afterCost",
        "commission",
        "afterCommission",
        "logistics",
        "afterLogistics",
        "storage",
        "afterStorage",
        "acceptance",
        "afterAcceptance",
        "promotion",
        "afterPromotion",
        "penalties",
        "afterPenalties",
        "acquiring",
        "beforeVatAdjustment",
        "pnlVatAdjustment",
        "profitBeforeTax",
        "includedTaxes",
        "profit",
        "margin",
        "unitProfit",
        "vat",
        "vatOutput",
        "vatInput",
        "vatInputFromWb",
        "vatInputFrom1c",
        "vatInputFromImportScenario",
        "vatInputFromWbScenario",
        "vatInputDifference",
        "vatInputCompleteness",
        "inputVatMode",
        "vatInputConfirmed",
        "vatPayable",
        "usn",
        "incomeTaxKind",
        "incomeTaxBase",
        "incomeTax",
        "incomeTaxIncluded",
        "taxMethod",
        "taxProfileSource",
        "taxCompleteness",
        "status",
        "statusReason",
        "sppStatus",
        "lossClass",
        "lossDriver",
        "accountingPeriodDate",
        "accountingPeriodSource",
        "documentReport",
        "wbReportId",
        "wbReportDate",
    ],
    "liquidityRows": [
        "month",
        "organization",
        "cabinet",
        "product",
        "nmId",
        "articleWb",
        "article1c",
        "barcode",
        "scheme",
        "sales",
        "returns",
        "netQty",
        "returnRate",
        "revenue",
        "cost",
        "md1Markup",
        "commission",
        "md2AfterCommission",
        "storage",
        "md3AfterStorage",
        "logistics",
        "acceptance",
        "md4AfterLogisticsAcceptance",
        "promotion",
        "md5AfterPromotion",
        "penalties",
        "acquiring",
        "md6BeforeTax",
        "vat",
        "usn",
        "profit",
        "margin",
        "unitProfit",
        "liquidityStatus",
        "liquidityDriver",
        "status",
        "statusReason",
        "sppStatus",
    ],
    "monthly": [
        "month",
        "status",
        "sales",
        "returns",
        "return_rate",
        "revenue",
        "profit",
        "margin",
    ],
    "expenses": ["expense", "amount", "share"],
    "returns": [
        "week",
        "month",
        "organization",
        "cabinet",
        "product",
        "nmId",
        "articleWb",
        "article1c",
        "barcode",
        "sales",
        "returns",
        "returnRate",
        "returnAmount",
        "profit",
        "status",
        "driver",
        "returnReason",
    ],
    "lostSales": [
        "cabinet",
        "product",
        "nmId",
        "articleWb",
        "article1c",
        "barcode",
        "periodDays",
        "zeroStockDays",
        "criticalStockDays",
        "onecStock",
        "onecWarehouses",
        "sales",
        "avgDailySales",
        "lostUnits",
        "lostRevenue",
        "lostProfit",
        "profitPerSale",
        "note",
        "sourceStatus",
    ],
    "documentReconciliation": [
        "status",
        "payoutStatus",
        "periodStatus",
        "documentReport",
        "salesPeriod",
        "expectedDocumentDate",
        "documentType",
        "cabinet",
        "organization",
        "summaryReportId",
        "weeklySalesReportId",
        "weeklyBuyoutReportId",
        "wbReportIds",
        "onecDocuments",
        "onecDocumentTypes",
        "onecDocumentDates",
        "wbSalesQuantity",
        "wbReturnQuantity",
        "wbNetQuantity",
        "onecSalesQuantity",
        "onecReturnQuantity",
        "onecNetQuantity",
        "salesQuantityDelta",
        "returnQuantityDelta",
        "netQuantityDelta",
        "wbQuantity",
        "onecQuantity",
        "quantityDelta",
        "wbAmount",
        "onecAmount",
        "amountDelta",
        "buyoutRetailAmountSum",
        "buyoutForPaySum",
        "buyoutBankPaymentSum",
        "onecExpenseInvoiceAmount",
        "buyoutRetailDelta",
        "buyoutForPayDelta",
        "buyoutBankDelta",
        "pdfBankPayment",
        "wbForPaySum",
        "onecSettlementTotal",
        "settlementDelta",
        "onecSourceRows",
        "comment",
    ],
    "reconciliationMonthly": [
        "month",
        "wb_quantity",
        "onec_quantity",
        "quantity_delta",
        "wb_cogs",
        "onec_cogs",
        "cogs_delta",
        "wb_mp_expenses",
        "onec_mp_expenses",
        "mp_expenses_delta",
        "comment",
    ],
    "ozonSummaryRows": [
        "articleId",
        "label",
        "group",
        "amount",
        "effectAmount",
        "share",
        "taxSystem",
        "taxProfileSource",
        "taxCompleteness",
        "sourceLabels",
        "expenseBasis",
        "attributionType",
        "periodExpenseAmount",
        "skuAttributedExpenseAmount",
        "unattributedExpenseAmount",
        "allocatedUnattributedExpenseAmount",
        "overAttributedExpenseAmount",
        "periodExpenseDeltaAmount",
        "roundingDeltaAmount",
        "costQualityStatus",
        "revenueCoveragePct",
        "quantityCoveragePct",
        "unmappedRevenueRowCount",
        "ambiguousRevenueRowCount",
        "missingCostCount",
        "anomalyCount",
        "insufficientHistoryCount",
        "estimatedCostImpact",
        "materialityThresholdAmount",
        "materialityThresholdMode",
        "materialityThresholdMinAmount",
        "materialityThresholdMaxAmount",
        "martAverageUnitCost",
        "direct1cAverageUnitCost",
        "direct1cDeviationPct",
        "status",
        "actionText",
    ],
    "ozonUnitRows": [
        "periodStart",
        "periodEnd",
        "offerId",
        "productId",
        "sku",
        "barcode",
        "productName",
        "quantity",
        "onecRevenue",
        "unitCost",
        "cogs",
        "ozonCommission",
        "ozonServices",
        "ozonPartnerServices",
        "ozonLogistics",
        "ozonStorage",
        "ozonOtherExpenses",
        "ozonExpenses",
        "skuAttributedExpenseAmount",
        "periodUnattributedExpenseAmount",
        "expenseBasis",
        "expenseAttributionType",
        "expenseAllocationBasis",
        "profitBeforeTax",
        "marginBeforeTax",
        "vatOutput",
        "vatInput",
        "vatPayable",
        "revenueTax",
        "incomeTax",
        "profitBeforeIncomeTax",
        "profitAfterTax",
        "marginAfterTax",
        "taxSystem",
        "taxProfileSource",
        "taxCompleteness",
        "profit",
        "margin",
        "onecItemId",
        "onecName",
        "qualityStatus",
        "costQualityStatus",
        "referenceUnitCost",
        "unitCostDeviationPct",
        "estimatedCostImpact",
        "costQualityReason",
        "expenseStatus",
        "problemReason",
        "actionText",
    ],
    "ozonServiceChargeRows": [
        "articleId",
        "label",
        "sourceLabel",
        "offerId",
        "sku",
        "productName",
        "onecName",
        "amount",
        "effectAmount",
        "allocationShare",
        "expenseBasis",
        "attributionType",
        "periodExpenseAmount",
        "skuAttributedExpenseAmount",
        "unattributedExpenseAmount",
        "basis",
        "status",
        "note",
    ],
    "ozonArticleSkuRows": [
        "articleId",
        "label",
        "offerId",
        "sku",
        "barcode",
        "productName",
        "onecName",
        "amount",
        "effectAmount",
        "allocationShare",
        "expenseBasis",
        "attributionType",
        "periodExpenseAmount",
        "skuAttributedExpenseAmount",
        "unattributedExpenseAmount",
        "includedInSkuProfit",
        "basis",
        "status",
    ],
    "ozonReconciliationRows": [
        "kind",
        "label",
        "parentLabel",
        "ozonAmount",
        "onecAmount",
        "deltaAmount",
        "periodExpenseAmount",
        "skuAttributedExpenseAmount",
        "unattributedExpenseAmount",
        "allocatedUnattributedExpenseAmount",
        "overAttributedExpenseAmount",
        "periodExpenseDeltaAmount",
        "roundingDeltaAmount",
        "expenseBasis",
        "attributionType",
        "includedInExpense",
        "includedInSkuProfit",
        "note",
    ],
    "taxInputReconciliation": [
        "week",
        "weekEnd",
        "cabinet",
        "organization",
        "vatInputFromWb",
        "vatInputFrom1c",
        "vatInputDifference",
        "vatInputCompleteness",
        "sourceRowCount",
    ],
}

COLUMN_LABELS = {
    "week": "Неделя",
    "accountingPeriodDate": "Учетная дата 1С",
    "accountingPeriodSource": "Источник учетной даты",
    "month": "Месяц",
    "status": "Статус",
    "documentReport": "Документ-отчет",
    "wbReportId": "Номер отчета WB",
    "wbReportDate": "Дата отчета WB",
    "organization": "Организация 1С",
    "cabinet": "Кабинет WB",
    "product": "Товар",
    "nmId": "nmId WB",
    "articleWb": "Артикул WB",
    "article1c": "Артикул 1С",
    "barcode": "Баркод",
    "scheme": "Схема продажи",
    "sales": "Продажи, шт",
    "returns": "Возвраты, шт",
    "netQty": "Чистое кол-во",
    "returnRate": "% возвратов",
    "return_rate": "% возвратов",
    "revenueBeforeSpp": "Выручка до СПП",
    "spp": "СПП",
    "sppRate": "% СПП",
    "revenue": "Выручка после СПП",
    "vat": "НДС к уплате",
    "vatOutput": "Исходящий НДС",
    "vatInput": "Входящий НДС",
    "vatInputFromWb": "НДС входящий WB",
    "vatInputFrom1c": "НДС входящий 1С",
    "vatInputFromImportScenario": "Расчётный НДС импорта",
    "vatInputFromWbScenario": "Расчётный НДС услуг WB",
    "vatInputDifference": "Расхождение НДС",
    "vatInputCompleteness": "Полнота НДС",
    "inputVatMode": "Режим входящего НДС",
    "vatInputConfirmed": "Входящий НДС подтверждён",
    "vatPayable": "НДС к уплате",
    "revenueWithoutVat": "Выручка без НДС",
    "pnlRevenue": "Выручка для расчета прибыли",
    "cost": "Себестоимость 1С",
    "afterCost": "Остаток после себестоимости",
    "md1Markup": "МД1 Наценка",
    "commission": "Комиссия WB",
    "afterCommission": "Остаток после комиссии",
    "md2AfterCommission": "МД2 после комиссии",
    "logistics": "Логистика WB",
    "afterLogistics": "Остаток после логистики",
    "storage": "Хранение WB",
    "afterStorage": "Остаток после хранения",
    "md3AfterStorage": "МД3 после хранения",
    "acceptance": "Приемка WB",
    "afterAcceptance": "Остаток после приемки",
    "md4AfterLogisticsAcceptance": "МД4 после логистики и приемки",
    "promotion": "Продвижение WB",
    "afterPromotion": "Остаток после продвижения",
    "md5AfterPromotion": "МД5 после продвижения",
    "penalties": "Штрафы/доплаты WB",
    "afterPenalties": "Остаток после штрафов/доплат",
    "acquiring": "Эквайринг WB",
    "beforeVatAdjustment": "Остаток до корректировки НДС услуг",
    "pnlVatAdjustment": "Корректировка P&L по НДС услуг WB",
    "md6BeforeTax": "МД6 до НДФЛ",
    "usn": "Налог с выручки/НДФЛ",
    "incomeTaxKind": "Вид НДФЛ",
    "incomeTaxBase": "База НДФЛ",
    "incomeTax": "НДФЛ",
    "incomeTaxIncluded": "НДФЛ включен",
    "profitBeforeTax": "Управленческая прибыль WB",
    "includedTaxes": "Налоги, включенные в итог",
    "marginBeforeTax": "Маржа до налогов",
    "profitBeforeIncomeTax": "Управленческая прибыль WB",
    "profitAfterTax": "Прибыль до налогов",
    "marginAfterTax": "Маржинальность до налогов",
    "revenueTax": "Налог с выручки",
    "taxSystem": "Налоговый режим",
    "profit": "Управленческая прибыль WB",
    "margin": "Маржа без НДС",
    "unitProfit": "Упр. прибыль на шт",
    "taxMethod": "Налоговый метод",
    "taxProfileSource": "Источник налогового профиля",
    "taxCompleteness": "Полнота налогового расчета",
    "weekEnd": "Конец недели",
    "sourceRowCount": "Строк источника",
    "statusReason": "Причина статуса",
    "sppStatus": "Статус СПП",
    "lossClass": "Класс убытка",
    "lossDriver": "Драйвер убытка",
    "liquidityStatus": "Статус ликвидности",
    "liquidityDriver": "Драйвер ликвидности",
    "expense": "Статья расходов",
    "amount": "Сумма",
    "share": "Доля от выручки",
    "returnAmount": "Сумма возвратов",
    "driver": "Драйвер",
    "returnReason": "Причина возврата",
    "periodDays": "Дни в периоде",
    "zeroStockDays": "Дни без остатка WB",
    "criticalStockDays": "Дни критического остатка",
    "onecStock": "Остаток 1С на складах",
    "onecWarehouses": "Склады 1С с остатком",
    "avgDailySales": "Среднедневные продажи",
    "lostUnits": "Потенциально упущенные штуки",
    "lostRevenue": "Потенциально упущенная выручка",
    "lostProfit": "Потенциально упущенная прибыль",
    "profitPerSale": "Прибыль на продажу",
    "note": "Управленческий вывод",
    "sourceStatus": "Статус источника",
    "payoutStatus": "Статус выплаты",
    "periodStatus": "Статус периода",
    "salesPeriod": "Период продаж",
    "expectedDocumentDate": "Ожидаемая дата документа",
    "documentType": "Тип документа",
    "summaryReportId": "Номер сводного отчета",
    "weeklySalesReportId": "Номер недельного отчета продаж",
    "weeklyBuyoutReportId": "Номер недельного отчета выкупов",
    "wbReportIds": "Номера отчетов WB",
    "onecDocuments": "Документы 1С",
    "onecDocumentTypes": "Типы документов 1С",
    "onecDocumentDates": "Даты документов 1С",
    "wbSalesQuantity": "Продажи WB, шт",
    "wbReturnQuantity": "Возвраты WB, шт",
    "wbNetQuantity": "Чистое кол-во WB",
    "onecSalesQuantity": "Продажи 1С, шт",
    "onecReturnQuantity": "Возвраты 1С, шт",
    "onecNetQuantity": "Чистое кол-во 1С",
    "salesQuantityDelta": "Отклонение продаж, шт",
    "returnQuantityDelta": "Отклонение возвратов, шт",
    "netQuantityDelta": "Отклонение чистого кол-ва",
    "wbQuantity": "Количество WB",
    "onecQuantity": "Количество 1С",
    "quantityDelta": "Отклонение количества",
    "wbAmount": "Сумма WB",
    "onecAmount": "Сумма 1С",
    "amountDelta": "Отклонение суммы",
    "buyoutRetailAmountSum": "Выкупы WB: розничная сумма",
    "buyoutForPaySum": "Выкупы WB: к перечислению",
    "buyoutBankPaymentSum": "Выкупы WB: банковский платеж",
    "onecExpenseInvoiceAmount": "Акт 1С по услугам",
    "buyoutRetailDelta": "Отклонение розничной суммы выкупов",
    "buyoutForPayDelta": "Отклонение к перечислению",
    "buyoutBankDelta": "Отклонение банковского платежа",
    "pdfBankPayment": "Банковский платеж из PDF",
    "wbForPaySum": "WB к перечислению",
    "onecSettlementTotal": "Взаиморасчеты 1С",
    "settlementDelta": "Отклонение взаиморасчетов",
    "onecSourceRows": "Строки источника 1С",
    "comment": "Комментарий",
    "wb_quantity": "Количество WB",
    "onec_quantity": "Количество 1С",
    "quantity_delta": "Отклонение количества",
    "wb_cogs": "Себестоимость WB",
    "onec_cogs": "Себестоимость 1С",
    "cogs_delta": "Отклонение себестоимости",
    "wb_mp_expenses": "Расходы WB",
    "onec_mp_expenses": "Расходы 1С",
    "mp_expenses_delta": "Отклонение расходов",
    "articleId": "Код статьи",
    "label": "Статья",
    "group": "Группа",
    "effectAmount": "Влияние на прибыль",
    "sourceLabels": "Источники",
    "actionText": "Действие",
    "periodStart": "Начало периода",
    "periodEnd": "Конец периода",
    "offerId": "Offer ID",
    "productId": "Product ID",
    "sku": "Ozon SKU",
    "productName": "Товар Ozon",
    "quantity": "Количество",
    "onecRevenue": "Выручка 1С",
    "unitCost": "Себестоимость единицы 1С",
    "cogs": "Себестоимость 1С",
    "ozonCommission": "Комиссия Ozon",
    "ozonServices": "Услуги Ozon",
    "ozonPartnerServices": "Услуги партнеров / перевыставление",
    "ozonLogistics": "Логистика Ozon",
    "ozonStorage": "Хранение / размещение Ozon",
    "ozonOtherExpenses": "Прочие расходы Ozon",
    "ozonExpenses": "Расходы Ozon",
    "onecItemId": "Ключ номенклатуры 1С",
    "onecName": "Номенклатура 1С",
    "qualityStatus": "Статус расчета",
    "costQualityStatus": "Качество себестоимости",
    "referenceUnitCost": "Референсная стоимость единицы",
    "unitCostDeviationPct": "Отклонение от референса",
    "estimatedCostImpact": "Предполагаемое влияние на COGS",
    "costQualityReason": "Причина качества себестоимости",
    "revenueCoveragePct": "Покрытие себестоимостью по выручке",
    "quantityCoveragePct": "Покрытие себестоимостью по количеству",
    "unmappedRevenueRowCount": "Строк выручки без сопоставления",
    "ambiguousRevenueRowCount": "Строк выручки с неоднозначным сопоставлением",
    "missingCostCount": "Строк без себестоимости",
    "anomalyCount": "Аномальных SKU",
    "insufficientHistoryCount": "SKU с недостаточной историей",
    "materialityThresholdAmount": "Порог существенности",
    "materialityThresholdMode": "Режим порога существенности",
    "materialityThresholdMinAmount": "Минимальный месячный порог",
    "materialityThresholdMaxAmount": "Максимальный месячный порог",
    "martAverageUnitCost": "Средняя стоимость mart",
    "direct1cAverageUnitCost": "Средняя стоимость регистра 1С",
    "direct1cDeviationPct": "Отклонение mart от регистра 1С",
    "expenseStatus": "Статус расходов",
    "problemReason": "Причина статуса",
    "sourceLabel": "Источник",
    "allocationShare": "Доля распределения",
    "basis": "База расчета",
    "expenseBasis": "База расхода",
    "expenseAttributionType": "Тип атрибуции",
    "expenseAllocationBasis": "База распределения",
    "attributionType": "Тип атрибуции",
    "periodExpenseAmount": "Mutual settlement",
    "skuAttributedExpenseAmount": "Расходы из Ozon detail",
    "periodUnattributedExpenseAmount": "Остаток периода",
    "unattributedExpenseAmount": "Остаток периода",
    "allocatedUnattributedExpenseAmount": "Распределенный остаток",
    "overAttributedExpenseAmount": "Ozon detail больше mutual settlement",
    "periodExpenseDeltaAmount": "Дельта периода",
    "roundingDeltaAmount": "Округление",
    "includedInSkuProfit": "Включено в прибыль SKU",
    "kind": "Тип строки",
    "parentLabel": "Связанная строка",
    "ozonAmount": "Ozon API",
    "deltaAmount": "Дельта",
    "includedInExpense": "Входит в расходы",
}

SHEET_COLUMN_LABELS = {
    ("unitRows", "profitBeforeTax"): (
        "Управленческая прибыль WB до включенных налогов"
    ),
    ("unitRows", "profit"): "Итог после включенных налогов",
    ("liquidityRows", "profit"): "Упр. прибыль",
    ("liquidityRows", "status"): "Статус данных",
    ("monthly", "status"): "Статус месяца",
    ("monthly", "profit"): "Управленческая прибыль WB",
    ("returns", "status"): "Статус данных",
    ("documentReconciliation", "status"): "Статус сверки",
    ("ozonReconciliationRows", "onecAmount"): "1C контроль",
}

STATUS_LABELS = {
    "": "",
    "ok": "ОК",
    "OK": "ОК",
    "ready": "Готово",
    "needs_review": "Нужна проверка",
    "partial_source": "Неполный источник",
    "complete": "Полное",
    "warning": "Предупреждение",
    "partial_period": "Неполный период",
    "source_coverage_gap": "Недостаточное покрытие источников",
    "blocked": "Заблокировано",
    "blocked_low_disk": "Заблокировано: мало места на диске",
    "failed": "Ошибка",
    "needs_configuration": "Нужна настройка",
    "no_data": "Нет данных",
    "missing_source": "Нет источника",
    "missing_stock_history": "Нет истории остатков WB",
    "transport_or_schema_error": "Ошибка транспорта или схемы",
    "task_failed": "Задача завершилась ошибкой",
    "task_timeout": "Истекло время ожидания задачи",
    "db_first_report_marts": "DB-first витрины отчета",
    "insufficient_history": "Недостаточно истории",
    "partial_provider_window_no_extrapolation": (
        "Доступный период WB, без экстраполяции"
    ),
    "unit_cost_outlier": "Аномальная стоимость",
    "nonpositive_unit_cost": "Неположительная стоимость",
    "missing_cost": "Нет себестоимости",
    "within_reference_range": "В пределах референса",
    "not_evaluated": "Не проверено",
    "not_applicable": "Не применимо",
}

STATUS_FIELDS = {
    "status",
    "sourceStatus",
    "payoutStatus",
    "periodStatus",
    "costQualityStatus",
    "costQualityReason",
}

PERCENT_FIELDS = {
    "returnRate",
    "return_rate",
    "sppRate",
    "margin",
    "marginBeforeTax",
    "marginAfterTax",
    "share",
    "revenueCoveragePct",
    "quantityCoveragePct",
    "unitCostDeviationPct",
    "direct1cDeviationPct",
}
MONEY_FIELDS = {
    "revenueBeforeSpp",
    "spp",
    "revenue",
    "vat",
    "vatOutput",
    "vatInput",
    "vatPayable",
    "revenueTax",
    "incomeTax",
    "revenueWithoutVat",
    "pnlRevenue",
    "cost",
    "afterCost",
    "md1Markup",
    "commission",
    "afterCommission",
    "md2AfterCommission",
    "logistics",
    "afterLogistics",
    "storage",
    "afterStorage",
    "md3AfterStorage",
    "acceptance",
    "afterAcceptance",
    "md4AfterLogisticsAcceptance",
    "promotion",
    "afterPromotion",
    "md5AfterPromotion",
    "penalties",
    "afterPenalties",
    "acquiring",
    "beforeVatAdjustment",
    "pnlVatAdjustment",
    "md6BeforeTax",
    "usn",
    "profitBeforeTax",
    "includedTaxes",
    "profitBeforeIncomeTax",
    "profitAfterTax",
    "profit",
    "unitProfit",
    "amount",
    "returnAmount",
    "lostRevenue",
    "lostProfit",
    "profitPerSale",
    "wbAmount",
    "onecAmount",
    "amountDelta",
    "buyoutRetailAmountSum",
    "buyoutForPaySum",
    "buyoutBankPaymentSum",
    "onecExpenseInvoiceAmount",
    "buyoutRetailDelta",
    "buyoutForPayDelta",
    "buyoutBankDelta",
    "pdfBankPayment",
    "wbForPaySum",
    "onecSettlementTotal",
    "settlementDelta",
    "wb_cogs",
    "onec_cogs",
    "cogs_delta",
    "wb_mp_expenses",
    "onec_mp_expenses",
    "mp_expenses_delta",
    "effectAmount",
    "onecRevenue",
    "unitCost",
    "referenceUnitCost",
    "estimatedCostImpact",
    "materialityThresholdAmount",
    "martAverageUnitCost",
    "direct1cAverageUnitCost",
    "cogs",
    "ozonCommission",
    "ozonServices",
    "ozonPartnerServices",
    "ozonLogistics",
    "ozonStorage",
    "ozonOtherExpenses",
    "ozonExpenses",
    "skuAttributedExpenseAmount",
    "periodUnattributedExpenseAmount",
    "ozonAmount",
    "deltaAmount",
    "periodExpenseAmount",
    "unattributedExpenseAmount",
    "allocatedUnattributedExpenseAmount",
    "overAttributedExpenseAmount",
    "periodExpenseDeltaAmount",
    "roundingDeltaAmount",
}


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def artifact_record(path: Path, status: str = "ready") -> dict[str, Any]:
    return {
        "path": str(path.resolve()),
        "hash": file_sha256(path) if path.exists() else "",
        "byte_size": path.stat().st_size if path.exists() else 0,
        "status": status,
    }


def _aggregate_unit_rows(
    rows: list[dict[str, Any]],
    group_fields: tuple[tuple[str, str], ...],
) -> list[dict[str, Any]]:
    totals: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        key = tuple(str(row.get(source) or "") for source, _label in group_fields)
        target = totals.setdefault(
            key,
            {label: key[index] for index, (_source, label) in enumerate(group_fields)},
        )
        for source, label in (
            ("sales", "Продажи, шт"),
            ("returns", "Возвраты, шт"),
            ("revenue", "Выручка с НДС"),
            ("revenueWithoutVat", "Выручка без НДС"),
            ("profitBeforeTax", "Управленческая прибыль WB"),
            ("profit", "Прибыль до налогов"),
        ):
            target[label] = float(target.get(label) or 0) + float(row.get(source) or 0)
    return list(totals.values())


def _dashboard_rows(summary: dict[str, Any]) -> list[dict[str, Any]]:
    kpis = summary.get("kpis") or {}
    tax_context = summary.get("taxContext") or {}
    readiness = summary.get("readiness") or {}
    labels = (
        ("Выручка с НДС", "revenueWithVat"),
        ("Выручка без НДС", "revenueWithoutVat"),
        ("Исходящий НДС", "vatOutput"),
        ("Налог УСН", "revenueTax"),
        ("Всего налогов", "totalTax"),
        ("Управленческая прибыль WB", "profitBeforeTax"),
        ("Прибыль до налогов", "profitAfterTax"),
    )
    rows = [
        {"Показатель": label, "Значение": kpis.get(key)} for label, key in labels
    ]
    rows.extend(
        [
            {
                "Показатель": "Налоговый профиль",
                "Значение": tax_context.get("taxSystem") or "Не подтверждён",
                "Комментарий": tax_context.get("message") or "",
            },
            {
                "Показатель": "Налоговый мост",
                "Значение": (
                    "Сходится" if kpis.get("taxBridgeCalculated") else "Не подтверждён"
                ),
            },
            {
                "Показатель": "Готовность",
                "Значение": readiness.get("status") or "needs_review",
                "Комментарий": readiness.get("nextAction") or "",
            },
        ]
    )
    return rows


def _full_excel_rows(
    summary: dict[str, Any],
    sheet_name: str,
    source_key: str,
) -> list[dict[str, Any]]:
    unit_rows = [
        _ensure_unit_profit_bridge(dict(item))
        for item in _safe_rows(summary.get("unitRows"))
    ]
    if source_key == "unitRows":
        rows = unit_rows
    elif source_key in summary:
        rows = [dict(item) for item in _safe_rows(summary.get(source_key))]
    elif source_key in {"dashboard", "summary"}:
        rows = _dashboard_rows(summary)
    elif source_key == "organizationSummary":
        rows = _aggregate_unit_rows(unit_rows, (("organization", "Организация 1С"),))
    elif source_key == "cabinetSummary":
        rows = _aggregate_unit_rows(unit_rows, (("cabinet", "Кабинет WB"),))
    elif source_key == "productSummary":
        rows = _aggregate_unit_rows(
            unit_rows,
            (
                ("product", "Товар"),
                ("articleWb", "Артикул WB"),
                ("article1c", "Артикул 1С"),
            ),
        )
    elif source_key == "wbReportSummary":
        rows = _aggregate_unit_rows(
            unit_rows,
            (("wbReportId", "ID отчёта WB"), ("documentReport", "Отчёт WB")),
        )
    elif source_key == "costRows":
        rows = [
            {
                "Товар": row.get("product"),
                "Организация 1С": row.get("organization"),
                "Артикул 1С": row.get("article1c"),
                "Штрихкод": row.get("barcode"),
                "Себестоимость 1С": row.get("cost"),
                "Статус": row.get("status"),
                "Причина": row.get("statusReason"),
            }
            for row in unit_rows
        ]
    elif source_key == "mappingRows":
        rows = [
            {
                "Товар": row.get("product"),
                "Кабинет WB": row.get("cabinet"),
                "Организация 1С": row.get("organization"),
                "Артикул WB": row.get("articleWb"),
                "Артикул 1С": row.get("article1c"),
                "Штрихкод": row.get("barcode"),
                "Статус": row.get("status"),
                "Причина": row.get("statusReason"),
            }
            for row in unit_rows
        ]
    elif source_key == "errorRows":
        rows = [row for row in unit_rows if str(row.get("status") or "") != "ОК"]
    else:
        rows = []
    if sheet_name == "Упущенные продажи":
        coverage = summary.get("lostSalesCoverage")
        if isinstance(coverage, dict) and coverage.get("calculated") is not True:
            message = coverage.get("message") or (
                "Не рассчитано: история остатков покрывает "
                f"{coverage.get('coveredDays', 0)} из "
                f"{coverage.get('totalDays', 0)} дней"
            )
            rows.insert(0, {"product": message, "sourceStatus": "insufficient_history"})
        elif isinstance(coverage, dict) and coverage.get("fullCoverage") is not True:
            rows.insert(
                0,
                {
                    "product": coverage.get("message")
                    or "Расчёт выполнен только за доступный период истории остатков.",
                    "sourceStatus": "partial_provider_window_no_extrapolation",
                },
            )
    return rows


def _ensure_unit_profit_bridge(row: dict[str, Any]) -> dict[str, Any]:
    """Add export-only cumulative P&L columns to persisted DB-first rows."""
    required = (
        "revenue",
        "cost",
        "commission",
        "logistics",
        "storage",
        "acceptance",
        "promotion",
        "penalties",
        "acquiring",
        "profitBeforeTax",
        "profit",
    )
    if any(row.get(field) in (None, "") for field in required):
        return row
    if (
        row.get("pnlVatMode") == "without_vat_for_osno"
        and row.get("revenueWithoutVat") in (None, "")
    ):
        return row

    def amount(field: str) -> float:
        return float(row.get(field) or 0)

    pnl_revenue = (
        amount("revenueWithoutVat")
        if row.get("pnlVatMode") == "without_vat_for_osno"
        else amount("revenue")
    )
    after_cost = pnl_revenue - amount("cost")
    after_commission = after_cost - amount("commission")
    after_logistics = after_commission - amount("logistics")
    after_storage = after_logistics - amount("storage")
    after_acceptance = after_storage - amount("acceptance")
    after_promotion = after_acceptance - amount("promotion")
    after_penalties = after_promotion - amount("penalties")
    before_vat_adjustment = after_penalties - amount("acquiring")
    profit_before_tax = amount("profitBeforeTax")
    profit = amount("profit")
    bridge = {
        "pnlRevenue": pnl_revenue,
        "afterCost": after_cost,
        "afterCommission": after_commission,
        "afterLogistics": after_logistics,
        "afterStorage": after_storage,
        "afterAcceptance": after_acceptance,
        "afterPromotion": after_promotion,
        "afterPenalties": after_penalties,
        "beforeVatAdjustment": before_vat_adjustment,
        "pnlVatAdjustment": profit_before_tax - before_vat_adjustment,
        "includedTaxes": profit_before_tax - profit,
    }
    for field, value in bridge.items():
        row.setdefault(field, round(value, 2))
    return row


def write_excel_from_marts(summary: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    _write_readme(readme, summary)
    for sheet_name, key in FULL_EXCEL_SHEETS:
        _write_rows_sheet(
            workbook.create_sheet(sheet_name),
            _full_excel_rows(summary, sheet_name, key),
            sheet_key=key,
        )
    _write_methodology(workbook.create_sheet("Методика"), summary)
    workbook.active = workbook.sheetnames.index("Дашборд")
    workbook.save(output_path)
    return output_path


def write_ozon_diagnostics_excel(
    diagnostics: dict[str, Any],
    output_path: Path,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = OZON_EXPORT_SHEETS["ozonSummaryRows"]
    _write_rows_sheet(
        first_sheet,
        _ozon_summary_rows(diagnostics),
        sheet_key="ozonSummaryRows",
    )
    for key, sheet_name in list(OZON_EXPORT_SHEETS.items())[1:]:
        _write_rows_sheet(
            workbook.create_sheet(sheet_name),
            _ozon_export_rows(diagnostics, key),
            sheet_key=key,
        )
    _write_ozon_methodology(workbook.create_sheet("Методика"), diagnostics)
    workbook.save(output_path)
    return output_path


def _ozon_export_rows(diagnostics: dict[str, Any], key: str) -> list[dict[str, Any]]:
    mart = diagnostics.get("ozonMart") or {}
    if key == "ozonUnitRows":
        return [dict(item) for item in _safe_rows(mart.get("rows"))]
    if key == "ozonServiceChargeRows":
        return [
            dict(item)
            for item in _safe_rows(mart.get("articleDrilldown"))
            if item.get("includedInSkuProfit")
        ]
    if key == "ozonArticleSkuRows":
        return [
            dict(item)
            for item in _safe_rows(mart.get("articleDrilldown"))
            if item.get("includedInSkuProfit")
        ]
    if key == "ozonReconciliationRows":
        rows = [
            dict(item)
            for item in _safe_rows(mart.get("articleDrilldown"))
            if item.get("kind") == "period_expense_control"
        ]
        rows.extend(
            dict(item)
            for item in _safe_rows(
                (diagnostics.get("expenseReconciliation") or {}).get("articleRows")
            )
        )
        return rows
    return []


def _ozon_summary_rows(diagnostics: dict[str, Any]) -> list[dict[str, Any]]:
    mart = diagnostics.get("ozonMart") or {}
    totals = mart.get("totals") or {}
    attribution = mart.get("expenseAttribution") or {}
    cost_quality = mart.get("costQuality") or {}
    revenue = _number_or_none(totals.get("onecRevenue"))
    expense_status = str(
        (diagnostics.get("expenseReconciliation") or {}).get("status") or ""
    )
    rows: list[dict[str, Any]] = []
    for item in _safe_rows(mart.get("articleRows")):
        effect = _number_or_none(item.get("effectAmount"))
        group = str(item.get("group") or "")
        needs_review = (
            expense_status
            and expense_status != "matched"
            and group not in {"revenue", "cogs", "result"}
        )
    if cost_quality:
        quality_status = str(cost_quality.get("status") or "complete")
        rows.append(
            {
                "articleId": "cost_quality",
                "label": "Контроль качества себестоимости",
                "group": "cogs",
                "costQualityStatus": quality_status,
                "revenueCoveragePct": cost_quality.get("revenueCoveragePct"),
                "quantityCoveragePct": cost_quality.get("quantityCoveragePct"),
                "unmappedRevenueRowCount": cost_quality.get(
                    "unmappedRevenueRowCount"
                ),
                "ambiguousRevenueRowCount": cost_quality.get(
                    "ambiguousRevenueRowCount"
                ),
                "missingCostCount": cost_quality.get("missingCostCount"),
                "anomalyCount": cost_quality.get("anomalyCount"),
                "insufficientHistoryCount": cost_quality.get(
                    "insufficientHistoryCount"
                ),
                "estimatedCostImpact": cost_quality.get("estimatedImpactAmount"),
                "materialityThresholdAmount": cost_quality.get(
                    "materialityThresholdAmount"
                ),
                "materialityThresholdMode": cost_quality.get(
                    "materialityThresholdMode"
                ),
                "materialityThresholdMinAmount": cost_quality.get(
                    "materialityThresholdMinAmount"
                ),
                "materialityThresholdMaxAmount": cost_quality.get(
                    "materialityThresholdMaxAmount"
                ),
                "martAverageUnitCost": cost_quality.get("martAverageUnitCost"),
                "direct1cAverageUnitCost": cost_quality.get(
                    "direct1cAverageUnitCost"
                ),
                "direct1cDeviationPct": cost_quality.get("direct1cDeviationPct"),
                "status": "ready" if quality_status == "complete" else "needs_review",
                "actionText": (
                    "Действие не требуется."
                    if quality_status == "complete"
                    else (
                        "Проверить предупреждения по SKU; исправления выполняет "
                        "аналитик."
                    )
                ),
            }
        )
        rows.append(
            {
                "articleId": item.get("articleId"),
                "label": item.get("label"),
                "group": group,
                "amount": item.get("amount"),
                "effectAmount": item.get("effectAmount"),
                "share": abs(effect) / revenue
                if effect is not None and revenue and revenue > 0
                else None,
                "sourceLabels": item.get("sourceLabels") or [],
                "expenseBasis": attribution.get("basis")
                or totals.get("expenseBasis"),
                "attributionType": attribution.get("status") or "",
                "periodExpenseAmount": attribution.get("periodExpenseAmount"),
                "skuAttributedExpenseAmount": attribution.get(
                    "skuAttributedExpenseAmount"
                ),
                "unattributedExpenseAmount": attribution.get(
                    "unattributedExpenseAmount"
                ),
                "allocatedUnattributedExpenseAmount": attribution.get(
                    "allocatedUnattributedExpenseAmount"
                ),
                "overAttributedExpenseAmount": attribution.get(
                    "overAttributedExpenseAmount"
                ),
                "periodExpenseDeltaAmount": attribution.get(
                    "periodExpenseDeltaAmount"
                ),
                "roundingDeltaAmount": attribution.get("roundingDeltaAmount"),
                "status": "needs_review" if needs_review else "ready",
                "actionText": (
                    "Сверить строку в листе «Сверка Ozon 1C»."
                    if needs_review
                    else "Действие не требуется."
                ),
            }
        )
    closed_totals = mart.get("closedPeriodTotals") or {}
    tax_rows = (
        ("vat_output", "Исходящий НДС", "vatOutput", "tax_bridge", None),
        ("vat_input", "Входящий НДС", "vatInput", "tax_bridge", None),
        ("vat_payable", "НДС к уплате", "vatPayable", "tax", "expense"),
        ("revenue_tax", "Налог с выручки", "revenueTax", "tax", "expense"),
        ("income_tax", "НДФЛ / налог на доход", "incomeTax", "tax", "expense"),
        (
            "profit_after_tax",
            "Прибыль до налогов",
            "profitAfterTax",
            "result",
            "result",
        ),
    )
    for article_id, label, field, group, effect_kind in tax_rows:
        value = totals.get(field)
        effect = (
            -value
            if effect_kind == "expense" and value is not None
            else value
            if effect_kind == "result"
            else None
        )
        rows.append(
            {
                "articleId": article_id,
                "label": label,
                "group": group,
                "amount": value,
                "effectAmount": effect,
                "status": "ready" if value is not None else "needs_review",
                "taxSystem": totals.get("taxSystem") or "",
                "taxProfileSource": totals.get("taxProfileSource") or "missing",
                "taxCompleteness": totals.get("taxCompleteness") or "not_calculated",
                "actionText": (
                    "Действие не требуется."
                    if value is not None
                    else (
                        "Обновить настройки налогообложения из 1С и проверить "
                        "полноту налоговой базы."
                    )
                ),
            }
        )
    excluded_open = _safe_rows(mart.get("excludedOpenPeriods"))
    excluded_incomplete = _safe_rows(mart.get("excludedIncompletePeriods"))
    if excluded_open or excluded_incomplete:
        rows.append(
            {
                "articleId": "closed_period_profit",
                "label": "Прибыль до налогов закрытых периодов",
                "group": "result",
                "amount": closed_totals.get("profitBeforeTax"),
                "effectAmount": closed_totals.get("profitBeforeTax"),
                "status": "needs_review",
                "actionText": (
                    "Общий итог скрыт: есть незакрытый месяц."
                    if excluded_open
                    else (
                        "Общий итог скрыт: закрытый месяц исключен из-за неполных "
                        "данных."
                    )
                ),
            }
        )
    if not rows:
        rows.append(
            {
                "articleId": "status",
                "label": "Ozon diagnostics",
                "group": "status",
                "status": diagnostics.get("status") or "not_started",
                "actionText": diagnostics.get("message") or "",
            }
        )
    return rows


def _write_ozon_methodology(sheet: Any, diagnostics: dict[str, Any]) -> None:
    latest = diagnostics.get("latestRun") or {}
    rows = [
        ("Правило", "Значение"),
        ("Назначение", "Staff-only Excel для проверки Ozon unit economics"),
        (
            "Публикация",
            "Не создает ReportRun и не заменяет клиентский WB Excel MVP",
        ),
        ("Выручка", "1C отчет комиссионера / регистр продаж Ozon"),
        (
            "Себестоимость",
            "Signed-движения 1C по организации, номенклатуре и календарному месяцу",
        ),
        (
            "Контроль себестоимости",
            "Медиана до 3 прошлых закрытых месяцев; аномалия ниже 50% или выше "
            "200% блокирует месяц при влиянии от max(100 000 руб.; 0,5% выручки)",
        ),
        (
            "Расходы Ozon",
            "SKU-поля Ozon detail первичны; mutual settlement - контроль периода",
        ),
        (
            "Fallback распределения",
            "Только положительный остаток mutual settlement минус SKU-detail "
            "распределяется по доле 1C-выручки",
        ),
        (
            "Отрицательный остаток",
            "Не распределяется; показывается как Ozon detail больше mutual settlement",
        ),
        ("Cash-flow", "Денежный контроль, не база P&L"),
        (
            "Налоговый профиль",
            "Явный профиль 1C, затем аудируемое временное исключение, затем missing",
        ),
        (
            "Незакрытый месяц",
            "Общая прибыль скрывается; closedPeriodTotals выводится отдельно",
        ),
        (
            "Закрытый неполный месяц",
            "Исключается из closedPeriodTotals и выводится в excludedIncompletePeriods",
        ),
        ("Legacy P&L", "Не используется; API-блок pnl помечен deprecated"),
        (
            "1C без пары в Ozon",
            "Показывается в сверке и не распределяется в прибыль SKU",
        ),
        ("Источник данных", latest.get("snapshotSetId") or ""),
        ("Статус", diagnostics.get("status") or ""),
        ("Сообщение", diagnostics.get("message") or ""),
    ]
    for index, row in enumerate(rows, start=1):
        sheet.cell(index, 1, row[0])
        sheet.cell(index, 2, row[1])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 90


def _safe_rows(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _number_or_none(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def write_csv_marts(summary: dict[str, Any], output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for key in (
        "unitRows",
        "lostSales",
        "reconciliationMonthly",
        "documentReconciliation",
    ):
        path = output_dir / f"{key}.csv"
        _write_csv(path, summary.get(key, []))
        paths.append(path)
    readme_path = output_dir / "README.md"
    readme_path.write_text(build_markdown_summary(summary), encoding="utf-8")
    paths.append(readme_path)
    return paths


def write_html_summary(summary: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    model = ClientReportModel.from_payload(summary)
    markdown = build_client_analytical_markdown(model)
    output_path.write_text(
        render_client_report_html(
            markdown,
            title=str(model.meta.get("title") or "Аналитический отчёт WB"),
        ),
        encoding="utf-8",
    )
    return output_path


def write_markdown_summary(summary: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(build_markdown_summary(summary), encoding="utf-8")
    return output_path


def write_docx_summary(summary: dict[str, Any], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    model = ClientReportModel.from_payload(summary)
    markdown = build_client_analytical_markdown(model)
    return render_markdown_docx(
        markdown,
        output_path,
        branded=False,
        landscape=False,
        cover_subtitle=str(model.meta.get("reportPeriod") or model.meta.get("period")),
        source_sha256=markdown_sha256(markdown),
    )


def convert_docx_to_pdf(docx_path: Path) -> tuple[Path | None, str, str]:
    converter = shutil.which("libreoffice") or shutil.which("soffice")
    if converter is None:
        return None, "unavailable", "LibreOffice/soffice не найден."
    output_dir = docx_path.parent
    expected = docx_path.with_suffix(".pdf")
    try:
        result = subprocess.run(
            [
                converter,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(docx_path),
            ],
            capture_output=True,
            text=True,
            check=False,
            timeout=120,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        return None, "failed", str(exc)
    if result.returncode != 0 or not expected.exists():
        return None, "failed", "DOCX создан, PDF не сформирован."
    return expected, "ready", "PDF сформирован."


def build_markdown_summary(summary: dict[str, Any]) -> str:
    return build_client_analytical_markdown(summary)


def _write_readme(sheet: Any, summary: dict[str, Any]) -> None:
    meta = summary.get("meta", {})
    rows = [
        ("Источник", _localized_source(meta.get("source", "DB report marts"))),
        ("Клиент", meta.get("client", "")),
        ("Период", meta.get("period", "")),
        ("Период отчета", meta.get("reportPeriod", meta.get("period", ""))),
        ("Покрытие источников", meta.get("sourceCoverage", "")),
        ("Статус готовности", _readiness_label(summary)),
        ("Версия методики", meta.get("methodologyVersion", "")),
        ("Происхождение данных", _localized_status(meta.get("lineageType", ""))),
        ("Строк юнит-экономики", len(summary.get("unitRows", []))),
        (
            "Финальный показатель",
            "Управленческая прибыль WB после включенных налогов; "
            "это не полная чистая прибыль бизнеса.",
        ),
        ("Строк упущенных продаж", len(summary.get("lostSales", []))),
    ]
    for index, row in enumerate(rows, start=1):
        sheet.cell(index, 1, row[0])
        sheet.cell(index, 2, row[1])
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 80


def _write_rows_sheet(
    sheet: Any, rows: list[dict[str, Any]], *, sheet_key: str
) -> None:
    headers = _sheet_headers(sheet_key, rows)
    if not headers:
        sheet.cell(1, 1, "Нет строк")
        return
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column, field in enumerate(headers, start=1):
        cell = sheet.cell(1, column, _column_label(sheet_key, field))
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row_index, row in enumerate(rows, start=2):
        for column, field in enumerate(headers, start=1):
            cell = sheet.cell(row_index, column, _localized_cell_value(field, row))
            if field in PERCENT_FIELDS and cell.value not in (None, ""):
                cell.number_format = "0.00%"
            elif field in MONEY_FIELDS and cell.value not in (None, ""):
                cell.number_format = '#,##0.00" ₽"'
    for column in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(1, column).column_letter].width = 18
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _write_methodology(sheet: Any, summary: dict[str, Any]) -> None:
    rows = [
        ("Правило", "Значение"),
        ("Источник правды", "Опубликованная расчетная БД"),
        ("Excel", "Только экспорт из расчетных витрин отчета"),
        ("Web-кабинет", "Читает опубликованный report_id из БД"),
        ("Исходные снимки", "Не публикуются через клиентское API"),
        (
            "Происхождение данных",
            _localized_status(summary.get("meta", {}).get("lineageType", "")),
        ),
    ]
    for index, row in enumerate(rows, start=1):
        sheet.cell(index, 1, row[0])
        sheet.cell(index, 2, row[1])


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    headers = sorted({key for row in rows for key in row})
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _readiness_label(summary: dict[str, Any]) -> str:
    readiness = summary.get("readiness", {})
    status = _localized_status(readiness.get("status", "")).strip()
    label = str(readiness.get("label") or "").strip()
    if status and label:
        if status == label:
            return status
        return f"{status} ({label})"
    return status or label


def _sheet_headers(sheet_key: str, rows: list[dict[str, Any]]) -> list[str]:
    if sheet_key in SHEET_COLUMNS:
        return SHEET_COLUMNS[sheet_key]
    return sorted({key for row in rows for key in row})


def _column_label(sheet_key: str, field: str) -> str:
    return SHEET_COLUMN_LABELS.get((sheet_key, field), COLUMN_LABELS.get(field, field))


def _localized_cell_value(field: str, row: dict[str, Any]) -> Any:
    value = row.get(field)
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    if field in STATUS_FIELDS:
        return _localized_status(value)
    return value


def _localized_status(value: object) -> str:
    text = str(value or "").strip()
    return STATUS_LABELS.get(text, text)


def _localized_source(value: object) -> str:
    text = str(value or "").strip()
    if text == "DB report marts":
        return "Расчетные витрины отчета"
    return _localized_status(text)
