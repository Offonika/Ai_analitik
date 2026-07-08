from __future__ import annotations

import csv
import hashlib
import json
import os
import time
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook

OZON_API_BASE_URL = "https://api-seller.ozon.ru"
OZON_CASH_FLOW_ENDPOINT = "/v1/finance/cash-flow-statement/list"
OZON_PRODUCT_REPORT_ENDPOINT = "/v1/report/products/create"
OZON_REPORT_INFO_ENDPOINT = "/v1/report/info"
OZON_STOCK_ON_WAREHOUSES_ENDPOINT = "/v2/analytics/stock_on_warehouses"
OZON_RETURNS_REPORT_ENDPOINT = "/v2/report/returns/create"
OZON_REALIZATION_ENDPOINT = "/v2/finance/realization"
OZON_REALIZATION_POSTING_ENDPOINT = "/v1/finance/realization/posting"
OZON_PRODUCTS_BUYOUT_ENDPOINT = "/v1/finance/products/buyout"
OZON_B2B_SALES_JSON_ENDPOINT = "/v1/finance/document-b2b-sales/json"
OZON_MUTUAL_SETTLEMENT_ENDPOINT = "/v1/finance/mutual-settlement"
OZON_REPORT_INFO_MAX_POLLS = 12
OZON_REPORT_INFO_MIN_DELAY_SECONDS = 2.0


@dataclass(frozen=True)
class OzonSellerAccount:
    seller_account_id: str
    account_name: str
    client_id: str = field(repr=False)
    api_key: str = field(repr=False)


@dataclass(frozen=True)
class OzonSettings:
    accounts: tuple[OzonSellerAccount, ...]
    timeout_seconds: float = 45.0
    base_url: str = OZON_API_BASE_URL

    @classmethod
    def from_env_file(
        cls,
        env_file: Path = Path(".env"),
        *,
        max_accounts: int = 10,
    ) -> OzonSettings:
        values = _load_env_values(env_file)
        values.update(os.environ)
        accounts: list[OzonSellerAccount] = []
        for index in range(1, max_accounts + 1):
            client_id = values.get(f"OZON_ACCOUNT_{index}_CLIENT_ID", "").strip()
            api_key = values.get(f"OZON_ACCOUNT_{index}_API_KEY", "").strip()
            if not client_id and not api_key:
                continue
            if not client_id or not api_key:
                raise OzonConfigError(
                    f"OZON_ACCOUNT_{index}_CLIENT_ID and "
                    f"OZON_ACCOUNT_{index}_API_KEY must be configured together"
                )
            seller_account_id = (
                values.get(f"OZON_ACCOUNT_{index}_ID", "").strip()
                or f"OZON_ACCOUNT_{index}"
            )
            account_name = values.get(f"OZON_ACCOUNT_{index}_NAME", "").strip()
            accounts.append(
                OzonSellerAccount(
                    seller_account_id=seller_account_id,
                    account_name=account_name or seller_account_id,
                    client_id=client_id,
                    api_key=api_key,
                )
            )
        if not accounts:
            raise OzonConfigError("No OZON_ACCOUNT_* credentials configured")

        timeout_value = values.get("OZON_TIMEOUT_SECONDS", "").strip()
        base_url = values.get("OZON_API_BASE_URL", "").strip() or OZON_API_BASE_URL
        return cls(
            accounts=tuple(accounts),
            timeout_seconds=float(timeout_value) if timeout_value else 45.0,
            base_url=base_url.rstrip("/"),
        )


@dataclass(frozen=True)
class OzonPageResult:
    source_type: str
    seller_account_id: str
    account_name: str
    page_index: int
    ok: bool
    status: str
    row_count: int
    raw_payload_hash: str = ""
    output_path: Path | None = None
    status_code: int | None = None
    error: str = ""
    report_code: str = ""
    source_endpoint: str = ""


class OzonConfigError(ValueError):
    pass


class OzonClient:
    def __init__(
        self,
        account: OzonSellerAccount,
        *,
        base_url: str = OZON_API_BASE_URL,
        timeout_seconds: float = 45.0,
        transport: httpx.BaseTransport | None = None,
    ) -> None:
        self._base_url = base_url.rstrip("/")
        self._client = httpx.Client(
            headers={
                "Client-Id": account.client_id,
                "Api-Key": account.api_key,
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
            timeout=timeout_seconds,
            follow_redirects=True,
            transport=transport,
        )

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> OzonClient:
        return self

    def __exit__(self, *args: object) -> None:
        self.close()

    def post(self, endpoint: str, payload: dict[str, Any]) -> httpx.Response:
        return self._client.post(f"{self._base_url}{endpoint}", json=payload)


def ozon_settings_from_secret(
    secret: str,
    *,
    default_name: str = "Ozon API",
    default_seller_account_id: str = "OZON_ACCOUNT",
) -> OzonSettings:
    raw = secret.strip()
    if not raw:
        raise OzonConfigError("ozon_secret_empty")
    accounts_payload: list[dict[str, Any]]
    if raw.startswith("{") or raw.startswith("["):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise OzonConfigError("ozon_secret_json_invalid") from exc
        if isinstance(parsed, list):
            accounts_payload = [item for item in parsed if isinstance(item, dict)]
        elif isinstance(parsed, dict):
            nested = parsed.get("accounts")
            if isinstance(nested, list):
                accounts_payload = [item for item in nested if isinstance(item, dict)]
            else:
                accounts_payload = [parsed]
        else:
            raise OzonConfigError("ozon_secret_json_must_be_object")
    else:
        accounts_payload = [_parse_key_value_secret(raw)]

    accounts: list[OzonSellerAccount] = []
    fallback_id = _safe_account_id(default_seller_account_id)
    for index, item in enumerate(accounts_payload, start=1):
        client_id = _first_text(item, "clientId", "client_id", "CLIENT_ID")
        api_key = _first_text(item, "apiKey", "api_key", "token", "API_KEY")
        if not client_id or not api_key:
            continue
        seller_account_id = (
            _first_text(item, "sellerAccountId", "seller_account_id", "id")
            or (fallback_id if len(accounts_payload) == 1 else f"{fallback_id}_{index}")
        )
        account_name = _first_text(item, "accountName", "account_name", "name")
        accounts.append(
            OzonSellerAccount(
                seller_account_id=seller_account_id,
                account_name=account_name or default_name,
                client_id=client_id,
                api_key=api_key,
            )
        )
    if not accounts:
        raise OzonConfigError("ozon_secret_missing_client_id_or_api_key")
    return OzonSettings(accounts=tuple(accounts))


def export_ozon_cash_flow(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    page_size: int = 1000,
    max_pages: int = 50,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_paged_endpoint(
        settings,
        output_dir,
        source_type="ozon_finance_cash_flow",
        endpoint=OZON_CASH_FLOW_ENDPOINT,
        period_start=period_start,
        period_end=period_end,
        page_size=page_size,
        max_pages=max_pages,
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=lambda page, page_size: {
            "page": page,
            "page_size": page_size,
            "date": {
                "from": _date_time_start(period_start),
                "to": _date_time_end(period_end),
            },
            "with_details": True,
        },
        row_extractor=_extract_cash_flow_rows,
    )


def export_ozon_stock_on_warehouses(
    settings: OzonSettings,
    output_dir: Path,
    *,
    limit: int = 1000,
    max_pages: int = 50,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_offset_endpoint(
        settings,
        output_dir,
        source_type="ozon_stock_on_warehouses",
        endpoint=OZON_STOCK_ON_WAREHOUSES_ENDPOINT,
        limit=limit,
        max_pages=max_pages,
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=lambda offset, limit: {"limit": limit, "offset": offset},
        row_extractor=lambda payload: _nested_list(payload, "result", "rows"),
    )


def export_ozon_realization(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    months = _months_between(period_start, period_end)
    return _export_monthly_endpoint(
        settings,
        output_dir,
        source_type="ozon_realization",
        endpoint=OZON_REALIZATION_ENDPOINT,
        months=months,
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=_realization_month_payload,
        row_extractor=lambda payload: _nested_list(payload, "result", "rows"),
    )


def export_ozon_realization_posting(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    max_pages: int = 50,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    months = _months_between(period_start, period_end)
    return _export_monthly_paged_endpoint(
        settings,
        output_dir,
        source_type="ozon_realization_posting",
        endpoint=OZON_REALIZATION_POSTING_ENDPOINT,
        months=months,
        max_pages=max_pages,
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=_realization_posting_payload,
        row_extractor=_extract_generic_rows,
    )


def export_ozon_products_buyout(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_period_chunks_endpoint(
        settings,
        output_dir,
        source_type="ozon_products_buyout",
        endpoint=OZON_PRODUCTS_BUYOUT_ENDPOINT,
        periods=_month_ranges_between(period_start, period_end),
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=_plain_date_range_payload,
        row_extractor=_extract_generic_rows,
    )


def export_ozon_b2b_sales_json(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_monthly_endpoint(
        settings,
        output_dir,
        source_type="ozon_b2b_sales_json",
        endpoint=OZON_B2B_SALES_JSON_ENDPOINT,
        months=_months_between(period_start, period_end),
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=_b2b_month_payload,
        row_extractor=_extract_generic_rows,
    )


def export_ozon_mutual_settlement(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_monthly_report_task(
        settings,
        output_dir,
        source_type="ozon_mutual_settlement",
        endpoint=OZON_MUTUAL_SETTLEMENT_ENDPOINT,
        months=_months_between(period_start, period_end),
        request_delay_seconds=request_delay_seconds,
        transport=transport,
        payload_factory=lambda month: {"date": month},
    )


def export_ozon_products_report(
    settings: OzonSettings,
    output_dir: Path,
    *,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_report_task(
        settings,
        output_dir,
        source_type="ozon_products_report",
        endpoint=OZON_PRODUCT_REPORT_ENDPOINT,
        payload={},
        request_delay_seconds=request_delay_seconds,
        transport=transport,
    )


def export_ozon_returns_report(
    settings: OzonSettings,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    request_delay_seconds: float = 0.0,
    transport: httpx.BaseTransport | None = None,
) -> list[OzonPageResult]:
    return _export_report_task(
        settings,
        output_dir,
        source_type="ozon_returns_report",
        endpoint=OZON_RETURNS_REPORT_ENDPOINT,
        payload={
            "filter": {
                "date": {
                    "from": _date_time_start(period_start),
                    "to": _date_time_end(period_end),
                }
            }
        },
        request_delay_seconds=request_delay_seconds,
        transport=transport,
    )


def _export_single_endpoint(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    payload: dict[str, Any],
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    row_extractor: Any,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            result = _post_and_store(
                client,
                account,
                output_dir,
                source_type=source_type,
                endpoint=endpoint,
                page_index=1,
                payload=payload,
                row_extractor=row_extractor,
            )
            results.append(result)
            _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_period_chunks_endpoint(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    periods: Iterable[tuple[date, date]],
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    payload_factory: Any,
    row_extractor: Any,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            for page_index, (chunk_start, chunk_end) in enumerate(periods, start=1):
                result = _post_and_store(
                    client,
                    account,
                    output_dir,
                    source_type=source_type,
                    endpoint=endpoint,
                    page_index=page_index,
                    payload=payload_factory(chunk_start, chunk_end),
                    row_extractor=row_extractor,
                    file_suffix=(
                        f"{chunk_start.isoformat()}_{chunk_end.isoformat()}"
                    ),
                )
                results.append(result)
                _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_paged_endpoint(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    period_start: date,
    period_end: date,
    page_size: int,
    max_pages: int,
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    payload_factory: Any,
    row_extractor: Any,
) -> list[OzonPageResult]:
    del period_start, period_end
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            for page in range(1, max_pages + 1):
                result = _post_and_store(
                    client,
                    account,
                    output_dir,
                    source_type=source_type,
                    endpoint=endpoint,
                    page_index=page,
                    payload=payload_factory(page, page_size),
                    row_extractor=row_extractor,
                )
                results.append(result)
                if not result.ok or result.row_count < page_size:
                    break
                _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_monthly_paged_endpoint(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    months: Iterable[str],
    max_pages: int,
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    payload_factory: Any,
    row_extractor: Any,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            page_index = 1
            for month in months:
                month_payload_hashes: set[str] = set()
                for page in range(1, max_pages + 1):
                    result = _post_and_store(
                        client,
                        account,
                        output_dir,
                        source_type=source_type,
                        endpoint=endpoint,
                        page_index=page_index,
                        payload=payload_factory(month, page),
                        row_extractor=row_extractor,
                        file_suffix=f"{month}_page_{page:04d}",
                    )
                    if result.ok and result.raw_payload_hash:
                        if result.raw_payload_hash in month_payload_hashes:
                            if result.output_path is not None:
                                result.output_path.unlink(missing_ok=True)
                            break
                        month_payload_hashes.add(result.raw_payload_hash)
                    results.append(result)
                    page_index += 1
                    if not result.ok or result.row_count == 0:
                        break
                    _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_offset_endpoint(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    limit: int,
    max_pages: int,
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    payload_factory: Any,
    row_extractor: Any,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            offset = 0
            for page in range(1, max_pages + 1):
                result = _post_and_store(
                    client,
                    account,
                    output_dir,
                    source_type=source_type,
                    endpoint=endpoint,
                    page_index=page,
                    payload=payload_factory(offset, limit),
                    row_extractor=row_extractor,
                )
                results.append(result)
                if not result.ok or result.row_count < limit:
                    break
                offset += limit
                _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_monthly_endpoint(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    months: Iterable[str],
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    payload_factory: Any,
    row_extractor: Any,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            for page, month in enumerate(months, start=1):
                result = _post_and_store(
                    client,
                    account,
                    output_dir,
                    source_type=source_type,
                    endpoint=endpoint,
                    page_index=page,
                    payload=payload_factory(month),
                    row_extractor=row_extractor,
                    file_suffix=month,
                )
                results.append(result)
                _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_report_task(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    payload: dict[str, Any],
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            created = _post_and_store(
                client,
                account,
                output_dir,
                source_type=source_type,
                endpoint=endpoint,
                page_index=1,
                payload=payload,
                row_extractor=lambda item: [item.get("result", item)],
            )
            results.append(created)
            code = created.report_code
            if code:
                for poll_index in range(1, OZON_REPORT_INFO_MAX_POLLS + 1):
                    _sleep_between_report_info_polls(
                        poll_index,
                        request_delay_seconds,
                    )
                    info = _post_and_store(
                        client,
                        account,
                        output_dir,
                        source_type=f"{source_type}_info",
                        endpoint=OZON_REPORT_INFO_ENDPOINT,
                        page_index=poll_index,
                        payload={"code": code},
                        row_extractor=lambda item: [item.get("result", item)],
                        file_suffix=f"info_{poll_index:02d}",
                    )
                    results.append(info)
                    file_url = _report_file_url(info.output_path)
                    if file_url:
                        results.append(
                            _download_report_file(
                                account,
                                output_dir,
                                source_type=source_type,
                                file_url=file_url,
                                report_code=code,
                                timeout_seconds=settings.timeout_seconds,
                                transport=transport,
                            )
                        )
                        break
                    if not info.ok:
                        break
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _export_monthly_report_task(
    settings: OzonSettings,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    months: Iterable[str],
    request_delay_seconds: float,
    transport: httpx.BaseTransport | None,
    payload_factory: Any,
) -> list[OzonPageResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[OzonPageResult] = []
    for account in settings.accounts:
        with OzonClient(
            account,
            base_url=settings.base_url,
            timeout_seconds=settings.timeout_seconds,
            transport=transport,
        ) as client:
            page_index = 1
            for month in months:
                created = _post_and_store(
                    client,
                    account,
                    output_dir,
                    source_type=source_type,
                    endpoint=endpoint,
                    page_index=page_index,
                    payload=payload_factory(month),
                    row_extractor=lambda item: [item.get("result", item)],
                    file_suffix=month,
                )
                results.append(created)
                page_index += 1
                code = created.report_code
                if code:
                    for poll_index in range(1, OZON_REPORT_INFO_MAX_POLLS + 1):
                        _sleep_between_report_info_polls(
                            poll_index,
                            request_delay_seconds,
                        )
                        info = _post_and_store(
                            client,
                            account,
                            output_dir,
                            source_type=f"{source_type}_info",
                            endpoint=OZON_REPORT_INFO_ENDPOINT,
                            page_index=page_index,
                            payload={"code": code},
                            row_extractor=lambda item: [item.get("result", item)],
                            file_suffix=f"{month}_info_{poll_index:02d}",
                        )
                        results.append(info)
                        page_index += 1
                        file_url = _report_file_url(info.output_path)
                        if file_url:
                            results.append(
                                _download_report_file(
                                    account,
                                    output_dir,
                                    source_type=source_type,
                                    file_url=file_url,
                                    report_code=code,
                                    timeout_seconds=settings.timeout_seconds,
                                    transport=transport,
                                    page_index=page_index,
                                    file_suffix=month,
                                )
                            )
                            page_index += 1
                            break
                        if not info.ok:
                            break
                _sleep_between_requests(request_delay_seconds)
    _write_manifest(output_dir, source_type, endpoint, results)
    return results


def _post_and_store(
    client: OzonClient,
    account: OzonSellerAccount,
    output_dir: Path,
    *,
    source_type: str,
    endpoint: str,
    page_index: int,
    payload: dict[str, Any],
    row_extractor: Any,
    file_suffix: str = "",
) -> OzonPageResult:
    try:
        response = client.post(endpoint, payload)
    except httpx.HTTPError as exc:
        return OzonPageResult(
            source_type=source_type,
            seller_account_id=account.seller_account_id,
            account_name=account.account_name,
            page_index=page_index,
            ok=False,
            status="transport_error",
            row_count=0,
            error=exc.__class__.__name__,
            source_endpoint=endpoint,
        )
    output_path: Path | None = None
    row_count = 0
    payload_hash = ""
    report_code = ""
    try:
        body = response.json()
    except ValueError:
        body = {"rawText": response.text}
    expected_empty = _expected_empty_response(
        source_type=source_type,
        response=response,
        payload=body,
    )
    status = "empty_expected" if expected_empty else _status_from_response(response)
    if response.status_code < 400:
        rows = row_extractor(body)
        row_count = len(rows)
        report_code = _report_code(body)
        payload_hash = _hash_payload(body)
        suffix = f"_{file_suffix}" if file_suffix else f"_page_{page_index:04d}"
        output_path = output_dir / (
            f"{account.seller_account_id}_{source_type}{suffix}.raw.json"
        )
        output_path.write_text(
            json.dumps(body, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return OzonPageResult(
        source_type=source_type,
        seller_account_id=account.seller_account_id,
        account_name=account.account_name,
        page_index=page_index,
        ok=response.status_code < 400 or expected_empty,
        status=status,
        row_count=row_count,
        raw_payload_hash=payload_hash,
        output_path=output_path,
        status_code=response.status_code,
        error="" if response.status_code < 400 else _error_message(body),
        report_code=report_code,
        source_endpoint=endpoint,
    )


def _report_file_url(output_path: Path | None) -> str:
    if output_path is None:
        return ""
    try:
        payload = json.loads(output_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return ""
    result = payload.get("result") if isinstance(payload, dict) else {}
    if not isinstance(result, dict):
        return ""
    file_url = result.get("file")
    return str(file_url).strip() if file_url else ""


def _download_report_file(
    account: OzonSellerAccount,
    output_dir: Path,
    *,
    source_type: str,
    file_url: str,
    report_code: str,
    timeout_seconds: float,
    transport: httpx.BaseTransport | None,
    page_index: int = 2,
    file_suffix: str = "",
) -> OzonPageResult:
    try:
        with httpx.Client(
            timeout=timeout_seconds,
            follow_redirects=True,
            transport=transport,
        ) as client:
            response = client.get(file_url)
    except httpx.HTTPError as exc:
        return OzonPageResult(
            source_type=f"{source_type}_file",
            seller_account_id=account.seller_account_id,
            account_name=account.account_name,
            page_index=page_index,
            ok=False,
            status="transport_error",
            row_count=0,
            error=exc.__class__.__name__,
            report_code=report_code,
            source_endpoint="report_file",
        )

    output_path = None
    row_count = 0
    payload_hash = ""
    if response.status_code < 400:
        content = response.content
        payload_hash = hashlib.sha256(content).hexdigest()
        suffix = _report_file_suffix(file_url, response, content)
        marker = f"_{file_suffix}" if file_suffix else ""
        output_path = output_dir / (
            f"{account.seller_account_id}_{source_type}{marker}_file.raw{suffix}"
        )
        output_path.write_bytes(content)
        row_count = (
            len(_xlsx_rows_from_bytes(content))
            if suffix == ".xlsx"
            else len(_tabular_rows(response.text))
        )

    return OzonPageResult(
        source_type=f"{source_type}_file",
        seller_account_id=account.seller_account_id,
        account_name=account.account_name,
        page_index=page_index,
        ok=response.status_code < 400,
        status=_status_from_response(response),
        row_count=row_count,
        raw_payload_hash=payload_hash,
        output_path=output_path,
        status_code=response.status_code,
        error="" if response.status_code < 400 else response.text[:240],
        report_code=report_code,
        source_endpoint="report_file",
    )


def _report_file_suffix(
    file_url: str,
    response: httpx.Response,
    content: bytes,
) -> str:
    path = file_url.split("?", 1)[0].lower()
    content_type = response.headers.get("content-type", "").lower()
    if path.endswith(".xlsx") or "spreadsheetml" in content_type:
        return ".xlsx"
    if content.startswith(b"PK\x03\x04"):
        return ".xlsx"
    if path.endswith(".tsv"):
        return ".tsv"
    if path.endswith(".csv"):
        return ".csv"
    if "tab-separated" in content_type:
        return ".tsv"
    return ".csv"


def _sleep_between_report_info_polls(
    poll_index: int,
    request_delay_seconds: float,
) -> None:
    if poll_index <= 1:
        _sleep_between_requests(request_delay_seconds)
        return
    _sleep_between_requests(
        max(request_delay_seconds, OZON_REPORT_INFO_MIN_DELAY_SECONDS)
    )


def _tabular_rows(text: str) -> list[dict[str, str]]:
    stripped = text.lstrip("\ufeff")
    if not stripped.strip():
        return []
    reader = csv.DictReader(
        StringIO(stripped, newline=""),
        delimiter=_tabular_delimiter(stripped),
    )
    return [
        {
            str(key or "").strip(): str(value or "").strip()
            for key, value in row.items()
        }
        for row in reader
    ]


def _xlsx_rows_from_bytes(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return _rows_from_table_values(rows)


def _rows_from_table_values(rows: list[tuple[Any, ...]]) -> list[dict[str, str]]:
    normalized_rows = [
        [_cell_text(value) for value in values]
        for values in rows
        if any(_cell_text(value) for value in values)
    ]
    if not normalized_rows:
        return []
    header_index = max(
        range(len(normalized_rows)),
        key=lambda index: (sum(1 for value in normalized_rows[index] if value), -index),
    )
    header_values = normalized_rows[header_index]
    header: list[str] = [
        value if value else f"column_{index}"
        for index, value in enumerate(header_values, start=1)
    ]
    data_rows: list[dict[str, str]] = []
    for text_values in normalized_rows[header_index + 1 :]:
        data_rows.append(
            {
                header[index]: text_values[index]
                for index in range(min(len(header), len(text_values)))
            }
        )
    return data_rows


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _tabular_delimiter(text: str) -> str:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        return "\t" if "\t" in sample else ";"
    return dialect.delimiter


def _status_from_response(response: httpx.Response) -> str:
    if response.status_code in {401, 403}:
        return "access_error"
    if response.status_code == 429:
        return "rate_limited"
    if response.status_code >= 400:
        return "http_error"
    return "ok"


def _expected_empty_response(
    *,
    source_type: str,
    response: httpx.Response,
    payload: dict[str, Any],
) -> bool:
    if response.status_code != 404:
        return False
    message = _error_message(payload).strip().lower()
    if source_type in {"ozon_realization", "ozon_realization_posting"}:
        return message == "report was not found"
    if source_type == "ozon_mutual_settlement":
        return "finance document not found" in message
    return False


def _extract_cash_flow_rows(payload: dict[str, Any]) -> list[Any]:
    result = payload.get("result")
    if isinstance(result, dict):
        items = result.get("items")
        if isinstance(items, list):
            return items
        details = result.get("details")
        if isinstance(details, list):
            return details
        if isinstance(details, dict):
            return [details]
        return [result]
    if isinstance(result, list):
        return result
    return []


def _extract_generic_rows(payload: dict[str, Any]) -> list[Any]:
    result = payload.get("result")
    if isinstance(result, list):
        return result
    if isinstance(result, dict):
        for key in (
            "items",
            "rows",
            "data",
            "documents",
            "operations",
            "invoices",
            "products",
        ):
            value = result.get(key)
            if isinstance(value, list):
                return value
        details = result.get("details")
        if isinstance(details, dict):
            return [details]
        return [result]
    for key in (
        "items",
        "rows",
        "data",
        "documents",
        "operations",
        "invoices",
        "products",
    ):
        value = payload.get(key)
        if isinstance(value, list):
            return value
    return []


def _nested_list(payload: dict[str, Any], *path: str) -> list[Any]:
    current: Any = payload
    for key in path:
        if not isinstance(current, dict):
            return []
        current = current.get(key)
    return current if isinstance(current, list) else []


def _report_code(payload: dict[str, Any]) -> str:
    result = payload.get("result")
    if isinstance(result, dict):
        return str(result.get("code") or "").strip()
    return str(payload.get("code") or "").strip()


def _error_message(payload: dict[str, Any]) -> str:
    for key in ("message", "error", "detail"):
        value = payload.get(key)
        if value:
            return str(value)[:500]
    return ""


def _write_manifest(
    output_dir: Path,
    source_type: str,
    endpoint: str,
    results: list[OzonPageResult],
) -> None:
    manifest = {
        "source": source_type,
        "sourceEndpoint": endpoint,
        "loadedAt": datetime.now().isoformat(),
        "results": [
            {
                "sellerAccountId": item.seller_account_id,
                "accountName": item.account_name,
                "pageIndex": item.page_index,
                "status": item.status,
                "ok": item.ok,
                "rowCount": item.row_count,
                "statusCode": item.status_code,
                "rawPayloadHash": item.raw_payload_hash,
                "outputFile": item.output_path.name if item.output_path else None,
                "reportCode": item.report_code,
                "error": item.error,
            }
            for item in results
        ],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _sleep_between_requests(seconds: float) -> None:
    if seconds > 0:
        time.sleep(seconds)


def _months_between(period_start: date, period_end: date) -> list[str]:
    months: list[str] = []
    year, month = period_start.year, period_start.month
    while (year, month) <= (period_end.year, period_end.month):
        months.append(f"{year:04d}-{month:02d}")
        month += 1
        if month > 12:
            year += 1
            month = 1
    return months


def _month_ranges_between(
    period_start: date,
    period_end: date,
) -> list[tuple[date, date]]:
    ranges: list[tuple[date, date]] = []
    year, month = period_start.year, period_start.month
    while (year, month) <= (period_end.year, period_end.month):
        chunk_start = max(period_start, date(year, month, 1))
        next_month = (
            date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
        )
        chunk_end = min(period_end, next_month - timedelta(days=1))
        ranges.append((chunk_start, chunk_end))
        month += 1
        if month > 12:
            year += 1
            month = 1
    return ranges


def _realization_month_payload(month_value: str) -> dict[str, int]:
    year_text, month_text = month_value.split("-", maxsplit=1)
    return {"month": int(month_text), "year": int(year_text)}


def _realization_posting_payload(month_value: str, page: int) -> dict[str, int]:
    payload = _realization_month_payload(month_value)
    payload["page"] = page
    return payload


def _b2b_month_payload(month_value: str) -> dict[str, str]:
    return {"date": month_value}


def _date_range_payload(period_start: date, period_end: date) -> dict[str, Any]:
    return {
        "date": {
            "from": _date_time_start(period_start),
            "to": _date_time_end(period_end),
        }
    }


def _plain_date_range_payload(period_start: date, period_end: date) -> dict[str, str]:
    return {
        "date_from": period_start.isoformat(),
        "date_to": period_end.isoformat(),
    }


def _date_time_start(value: date) -> str:
    return f"{value.isoformat()}T00:00:00Z"


def _date_time_end(value: date) -> str:
    return f"{value.isoformat()}T23:59:59Z"


def _hash_payload(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _parse_key_value_secret(raw: str) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for item in raw.replace("\n", ";").split(";"):
        if not item.strip():
            continue
        key, sep, value = item.partition("=")
        if not sep:
            raise OzonConfigError("ozon_secret_key_value_expected")
        result[key.strip()] = value.strip()
    return result


def _first_text(values: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = values.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _safe_account_id(value: str) -> str:
    normalized = "".join(
        char if char.isalnum() else "_" for char in value.strip().upper()
    ).strip("_")
    return normalized or "OZON_ACCOUNT"


def _load_env_values(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    values: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values
