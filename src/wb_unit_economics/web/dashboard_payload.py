from __future__ import annotations

import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts import build_client_demo_dashboard as demo  # noqa: E402

PERIOD_START = date(2026, 3, 1)
PERIOD_END = date(2026, 6, 17)
REPORT_PERIOD_LABEL = "01.03.2026 - 17.06.2026"
REPORT_PERIOD_TEXT = "март, апрель, май, июнь; июнь неполный, по 17.06.2026"
RETURN_REASON_LIMITATION = "Причина возврата не передается текущими источниками"
DATE_PATTERN = re.compile(r"\d{4}-\d{2}-\d{2}|\d{2}\.\d{2}\.\d{4}")
RU_MONTH_NUMBERS = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}


def date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return demo._text(value)


def parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = demo._text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def read_readme(workbook: Any) -> dict[str, Any]:
    if "README" not in workbook.sheetnames:
        return {}
    result: dict[str, Any] = {}
    for row in workbook["README"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            result[str(row[0])] = row[1]
    return result


def period_label_from_value(value: Any, fallback: str) -> str:
    dates = [
        parsed
        for parsed in (
            parse_date_value(match) for match in DATE_PATTERN.findall(demo._text(value))
        )
        if parsed is not None
    ]
    if len(dates) >= 2:
        return f"{dates[0]:%d.%m.%Y} - {dates[1]:%d.%m.%Y}"
    return fallback


def period_boundaries_from_label(value: str) -> tuple[str, str]:
    dates = [
        parsed
        for parsed in (parse_date_value(match) for match in DATE_PATTERN.findall(value))
        if parsed is not None
    ]
    if len(dates) >= 2:
        return dates[0].isoformat(), dates[1].isoformat()
    return "", ""


def _month_start_from_label(value: str) -> date | None:
    match = re.match(r"^([А-Яа-яЁё]+)\s+(\d{4})", demo._text(value))
    if match is None:
        return None
    month = RU_MONTH_NUMBERS.get(match.group(1).casefold())
    if month is None:
        return None
    return date(int(match.group(2)), month, 1)


def _month_label_sort_key(value: str) -> tuple[date, str]:
    return (_month_start_from_label(value) or date.max, value)


def monthly_rows_with_dates(
    rows: list[dict[str, Any]], report_period: str
) -> list[dict[str, Any]]:
    start_text, end_text = period_boundaries_from_label(report_period)
    period_start = parse_date_value(start_text)
    period_end = parse_date_value(end_text)
    result = []
    for source in rows:
        row = dict(source)
        month_start = _month_start_from_label(demo._text(row.get("month")))
        if month_start is None:
            result.append(row)
            continue
        next_month = (
            date(month_start.year + 1, 1, 1)
            if month_start.month == 12
            else date(month_start.year, month_start.month + 1, 1)
        )
        month_end = next_month - timedelta(days=1)
        days_in_month = (next_month - month_start).days
        covered_start = max(month_start, period_start or month_start)
        covered_end = min(month_end, period_end or month_end)
        days_elapsed = max(0, (covered_end - covered_start).days + 1)
        row.update(
            {
                "monthStart": month_start.isoformat(),
                "isPartial": days_elapsed < days_in_month,
                "daysElapsed": days_elapsed,
                "daysInMonth": days_in_month,
            }
        )
        result.append(row)
    return sorted(result, key=lambda item: item.get("monthStart") or "9999")


def lost_sales_coverage(workbook: Any, report_period: str) -> dict[str, Any]:
    _, end_text = period_boundaries_from_label(report_period)
    start_text, _ = period_boundaries_from_label(report_period)
    start = parse_date_value(start_text)
    end = parse_date_value(end_text)
    total_days = (end - start).days + 1 if start and end else 0
    covered_days = 0
    if "Упущенные продажи" in workbook.sheetnames:
        sheet = workbook["Упущенные продажи"]
        for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
            if demo._text(row[0]) != "Покрытие истории остатков":
                continue
            match = re.search(r"(\d+)\s+из\s+(\d+)", demo._text(row[1]))
            if match:
                covered_days = int(match.group(1))
                total_days = int(match.group(2))
            break
    calculated = bool(total_days and covered_days == total_days)
    return {
        "status": "complete" if calculated else "incomplete",
        "calculated": calculated,
        "coveredDays": covered_days,
        "totalDays": total_days,
        "message": (
            "Покрытие истории остатков полное."
            if calculated
            else (
                "Не рассчитано: история остатков покрывает "
                f"{covered_days} из {total_days} дней."
            )
        ),
        "accounts": [],
    }


def analysis_period_text(value: Any, fallback: str) -> str:
    text = demo._text(value)
    if text.startswith("Период анализа:"):
        text = text.removeprefix("Период анализа:").strip()
    return text or fallback


def first_present(row: dict[str, Any], *headers: str) -> Any:
    for header in headers:
        value = row.get(header)
        if value is not None and value != "":
            return value
    return None


def sales_period_bounds(value: Any) -> tuple[date | None, date | None]:
    text = demo._text(value)
    left, _, right = text.partition(" - ")
    return parse_date_value(left), parse_date_value(right)


def format_ru_date(value: date | None) -> str:
    return value.strftime("%d.%m.%Y") if value else ""


def document_report_label(
    document_type: str,
    sales_period_start: date | None,
    sales_period_end: date | None,
) -> str:
    if not document_type or not sales_period_start or not sales_period_end:
        return ""
    return (
        f"{document_type} · "
        f"{format_ru_date(sales_period_start)}-{format_ru_date(sales_period_end)} · "
        f"закрытие {format_ru_date(sales_period_end)}"
    )


def loss_details(row: dict[str, Any]) -> tuple[str, str]:
    status = demo._text(row.get("Статус данных"))
    profit = demo._row_profit(row)
    if status and status != "ОК":
        return "Нужна проверка данных", status
    if profit >= 0:
        return "Прибыльный / нейтральный", "Маржинальный доход не отрицательный"
    sales = demo._num(row.get("Продажи, шт"))
    returns = demo._num(row.get("Возвраты, шт"))
    return_rate = demo._safe_div(returns, sales) or 0.0
    tax_method = demo._text(
        row.get("Налоговый метод") or row.get("Налоговый режим/ставка")
    )
    pnl_vat_mode = demo._text(row.get("Режим P&L НДС"))
    if not pnl_vat_mode and "ОСНО" in tax_method:
        pnl_vat_mode = "without_vat_for_osno"
    tax_factor = demo._num(
        row.get("Налог с выручки/НДФЛ")
        or row.get("Налог с выручки")
        or row.get("УСН 1%")
    )
    if pnl_vat_mode != "without_vat_for_osno":
        tax_factor += demo._num(
            row.get("НДС к уплате") or row.get("НДС") or row.get("НДС 5%")
        )
    factors = {
        "Высокая себестоимость": demo._num(row.get("Себестоимость 1С")),
        "Высокая логистика WB": demo._num(row.get("Логистика WB")),
        "Высокая комиссия WB": demo._num(row.get("Комиссия WB")),
        "Высокое хранение WB": demo._num(row.get("Хранение WB")),
        "WB Продвижение": demo._num(row.get("Продвижение WB")),
        "Штрафы/удержания WB": demo._num(row.get("Штрафы/доплаты WB")),
        "Эквайринг WB": demo._num(row.get("Эквайринг WB")),
        "Налоги": tax_factor,
    }
    if return_rate >= 0.18:
        factors["Возвраты + логистика"] = demo._num(
            row.get("Сумма возвратов")
        ) + demo._num(row.get("Логистика WB"))
    driver = max(factors.items(), key=lambda item: item[1])[0]
    if driver == "Высокая себестоимость":
        return "Высокая закупка / недостаточная наценка", driver
    if driver == "Возвраты + логистика":
        return "Возвраты + логистика", driver
    return "Прочие расходы", driver


def unit_rows(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(demo._read_unit_rows(workbook), start=1):
        month = demo._row_month(row.get("Неделя"))
        loss_class, loss_driver = loss_details(row)
        sales = demo._num(row.get("Продажи, шт"))
        returns = demo._num(row.get("Возвраты, шт"))
        net_qty = demo._num(row.get("Чистое кол-во")) or (sales - returns)
        revenue_before_spp = demo._row_revenue_before_spp(row)
        revenue = demo._row_revenue(row)
        profit = demo._row_profit(row)
        rows.append(
            {
                "id": f"unit-{index}",
                "week": date_text(row.get("Неделя")),
                "month": month,
                "documentReport": demo._text(row.get("Документ-отчет")),
                "wbReportId": demo._text(row.get("Номер отчета WB")),
                "wbReportDate": date_text(row.get("Дата отчета WB")),
                "organization": demo._text(row.get("Организация 1С")),
                "cabinet": demo._text(row.get("Кабинет WB")),
                "product": demo._text(row.get("Товар")),
                "nmId": demo._text(row.get("nmId WB")),
                "articleWb": demo._text(row.get("Артикул WB")),
                "article1c": demo._text(row.get("Артикул 1С")),
                "barcode": demo._text(row.get("Баркод")),
                "scheme": demo._text(row.get("Схема продажи")) or "Не указано",
                "sales": round(sales),
                "returns": round(returns),
                "netQty": round(net_qty),
                "returnRate": demo._safe_div(returns, sales),
                "revenueBeforeSpp": round(revenue_before_spp, 2),
                "spp": demo._round(row.get("СПП")),
                "sppRate": demo._round(row.get("% СПП"), 4),
                "revenue": round(revenue, 2),
                "vat": demo._round(
                    row.get("НДС к уплате") or row.get("НДС") or row.get("НДС 5%")
                ),
                "vatOutput": demo._round(row.get("Исходящий НДС")),
                "vatInput": demo._round(row.get("Входящий НДС")),
                "vatInputFromWb": demo._round(row.get("НДС входящий WB")),
                "vatInputFrom1c": demo._round(row.get("НДС входящий 1С")),
                "vatInputDifference": demo._round(row.get("Расхождение НДС")),
                "vatInputCompleteness": demo._text(row.get("Полнота НДС")),
                "vatPayable": demo._round(row.get("НДС к уплате")),
                "revenueWithoutVat": demo._round(row.get("Выручка без НДС")),
                "cost": demo._round(row.get("Себестоимость 1С")),
                "commission": demo._round(row.get("Комиссия WB")),
                "logistics": demo._round(row.get("Логистика WB")),
                "storage": demo._round(row.get("Хранение WB")),
                "acceptance": demo._round(row.get("Приемка WB")),
                "promotion": demo._round(row.get("Продвижение WB")),
                "penalties": demo._round(row.get("Штрафы/доплаты WB")),
                "acquiring": demo._round(row.get("Эквайринг WB")),
                "usn": demo._round(
                    row.get("Налог с выручки/НДФЛ")
                    or row.get("Налог с выручки")
                    or row.get("УСН 1%")
                ),
                "incomeTaxBase": demo._round(row.get("База НДФЛ")),
                "incomeTax": demo._round(row.get("НДФЛ")),
                "incomeTaxIncluded": bool(row.get("НДФЛ включен") or False),
                "profitBeforeTax": demo._round(
                    row.get("Маржинальный доход WB до налогов")
                ),
                "profit": round(profit, 2),
                "margin": demo._round(
                    first_present(
                        row,
                        "Маржа WB без НДС",
                        "Маржа WB после налогов",
                    ),
                    4,
                ),
                "unitProfit": demo._round(
                    first_present(
                        row,
                        "Управленческая прибыль WB на шт",
                        "Маржинальный доход WB после налогов на шт",
                    )
                ),
                "taxMethod": demo._text(
                    row.get("Налоговый метод") or row.get("Налоговый режим/ставка")
                ),
                "taxProfileSource": demo._text(
                    row.get("Источник налогового профиля")
                ),
                "taxCompleteness": demo._text(row.get("Полнота налогового расчета")),
                "pnlVatMode": demo._text(row.get("Режим P&L НДС"))
                or (
                    "without_vat_for_osno"
                    if "ОСНО"
                    in demo._text(
                        row.get("Налоговый метод") or row.get("Налоговый режим/ставка")
                    )
                    else ""
                ),
                "status": demo._text(row.get("Статус данных")) or "Не указан",
                "statusReason": demo._text(row.get("Причина статуса")),
                "sppStatus": demo._text(row.get("Статус СПП")),
                "lossClass": loss_class,
                "lossDriver": loss_driver,
            }
        )
    return rows


def lost_sales_rows(workbook: Any) -> list[dict[str, Any]]:
    rows = []
    for index, row in enumerate(demo._read_lost_sales(workbook), start=1):
        rows.append(
            {
                "id": f"lost-{index}",
                "cabinet": row["cabinet"],
                "product": row["product"],
                "article1c": row["article_1c"],
                "barcode": row["barcode"],
                "zeroStockDays": row["zero_stock_days"],
                "onecStock": row.get("onec_stock", 0),
                "onecWarehouses": row.get("onec_warehouses", ""),
                "sales": row["sales"],
                "lostUnits": row["lost_units"],
                "lostRevenue": row["lost_revenue"],
                "lostContributionMargin": row["lost_profit"],
                "lostProfit": row["lost_profit"],
                "note": row["note"],
            }
        )
    return rows


def document_reconciliation_rows(workbook: Any) -> list[dict[str, Any]]:
    sheet_name = "Сверка документов 1С"
    if sheet_name not in workbook.sheetnames:
        return []
    try:
        header_row = demo._find_header_row(workbook, sheet_name, "Статус сверки")
    except ValueError:
        return []
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(
        demo._read_table(workbook, sheet_name, header_row=header_row).rows,
        start=1,
    ):
        status = demo._text(row.get("Статус сверки"))
        if not status:
            break
        sales_period = demo._text(row.get("Период продаж"))
        sales_start, sales_end = sales_period_bounds(sales_period)
        expected_document_date = parse_date_value(row.get("Ожидаемая дата документа"))
        document_type = demo._text(row.get("Тип документа WB/1С"))
        rows.append(
            {
                "id": f"document-reconciliation-{index}",
                "status": status,
                "payoutStatus": demo._text(row.get("Статус выплаты")),
                "periodStatus": demo._text(row.get("Статус периода")),
                "documentReport": document_report_label(
                    document_type,
                    sales_start,
                    sales_end,
                ),
                "salesPeriod": sales_period,
                "salesPeriodStart": sales_start.isoformat() if sales_start else "",
                "salesPeriodEnd": sales_end.isoformat() if sales_end else "",
                "expectedDocumentDate": (
                    expected_document_date.isoformat()
                    if expected_document_date
                    else ""
                ),
                "documentType": document_type,
                "cabinet": demo._text(row.get("Кабинет WB")),
                "organization": demo._text(row.get("Организация 1С")),
                "summaryReportId": demo._text(row.get("Номер отчета WB (сводный)")),
                "weeklySalesReportId": demo._text(row.get("WB отчет продаж")),
                "weeklyBuyoutReportId": demo._text(row.get("WB отчет выкупов")),
                "wbReportIds": demo._text(row.get("WB reportId в пакете")),
                "onecDocuments": demo._text(row.get("Документы 1С")),
                "onecDocumentTypes": demo._text(row.get("Типы документов 1С")),
                "onecDocumentDates": demo._text(row.get("Даты документов 1С")),
                "wbSalesQuantity": demo._round(row.get("WB продажи"), 4),
                "wbReturnQuantity": demo._round(row.get("WB возвраты"), 4),
                "wbNetQuantity": demo._round(row.get("WB чистое"), 4),
                "onecSalesQuantity": demo._round(row.get("1С продажи"), 4),
                "onecReturnQuantity": demo._round(row.get("1С возвраты"), 4),
                "onecNetQuantity": demo._round(row.get("1С чистое"), 4),
                "salesQuantityDelta": demo._round(row.get("Дельта продажи"), 4),
                "returnQuantityDelta": demo._round(row.get("Дельта возвраты"), 4),
                "netQuantityDelta": demo._round(row.get("Дельта чистое"), 4),
                "wbQuantity": demo._round(row.get("WB количество для 1С"), 4),
                "onecQuantity": demo._round(row.get("1С количество"), 4),
                "quantityDelta": demo._round(row.get("Дельта количество"), 4),
                "wbAmount": demo._round(row.get("WB сумма документа")),
                "onecAmount": demo._round(row.get("1С сумма документа")),
                "amountDelta": demo._round(row.get("Дельта сумма")),
                "buyoutRetailAmountSum": demo._round(
                    row.get("WB выкуп: retailAmountSum")
                ),
                "buyoutForPaySum": demo._round(row.get("WB выкуп: forPaySum")),
                "buyoutBankPaymentSum": demo._round(
                    row.get("WB выкуп: bankPaymentSum")
                ),
                "onecExpenseInvoiceAmount": demo._round(
                    row.get("1С расходная накладная")
                ),
                "buyoutRetailDelta": demo._round(row.get("Δ выкуп retail")),
                "buyoutForPayDelta": demo._round(row.get("Δ выкуп к перечислению")),
                "buyoutBankDelta": demo._round(row.get("Δ выкуп банк")),
                "pdfBankPayment": demo._round(row.get("PDF 8. К перечислению")),
                "wbForPaySum": demo._round(row.get("WB к перечислению (forPaySum)")),
                "onecSettlementTotal": demo._round(
                    first_present(
                        row,
                        "1С оборот взаиморасчетов",
                        "1С итого взаиморасчетов",
                    )
                ),
                "settlementDelta": demo._round(
                    first_present(row, "Дельта к обороту 1С", "Дельта к перечислению")
                ),
                "onecSourceRows": demo._round(row.get("Строк регистра 1С"), 0),
                "comment": demo._text(row.get("Комментарий")),
            }
        )
    return rows


def options(
    rows: list[dict[str, Any]],
    *,
    document_reconciliation: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    def unique(key: str) -> list[str]:
        return sorted({demo._text(row.get(key)) for row in rows if row.get(key)})

    document_reconciliation = document_reconciliation or []
    row_months = {row["month"] for row in rows}
    period_dates = sorted(
        period_date
        for row in rows
        if (period_date := _row_filter_period_date(row)) is not None
    )
    return {
        "months": sorted(row_months, key=_month_label_sort_key),
        "periodStart": period_dates[0].isoformat() if period_dates else "",
        "periodEnd": period_dates[-1].isoformat() if period_dates else "",
        "cabinets": unique("cabinet"),
        "organizations": unique("organization"),
        "schemes": unique("scheme"),
        "statuses": sorted(
            {
                *unique("status"),
                *(
                    demo._text(row.get("status"))
                    for row in document_reconciliation
                    if row.get("status")
                ),
            }
        ),
        "lossClasses": unique("lossClass"),
        "documentReports": sorted(
            {
                *unique("documentReport"),
                *(
                    demo._text(row.get("documentReport"))
                    for row in document_reconciliation
                    if row.get("documentReport")
                ),
            }
        ),
    }


def _row_filter_period_date(row: dict[str, Any]) -> date | None:
    week = parse_date_value(row.get("week"))
    if week is not None:
        return week + timedelta(days=6)
    accounting_period_date = parse_date_value(row.get("accountingPeriodDate"))
    if accounting_period_date is not None:
        return accounting_period_date
    return parse_date_value(row.get("wbReportDate"))


def build_dashboard_payload(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing = demo.VISIBLE_SHEETS.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"В Excel отсутствуют листы: {', '.join(sorted(missing))}")
    readme = read_readme(workbook)
    report_period = period_label_from_value(
        readme.get("Период отчета") or readme.get("Период"),
        REPORT_PERIOD_LABEL,
    )
    source_coverage = period_label_from_value(
        readme.get("Покрытие источников"),
        "",
    )
    source_coverage_start, source_coverage_end = period_boundaries_from_label(
        source_coverage
    )
    rows = unit_rows(workbook)
    document_reconciliation = document_reconciliation_rows(workbook)
    reconciliation, reconciliation_monthly = demo._read_onec_opiu_reconciliation(
        workbook
    )
    return {
        "meta": {
            "title": "Кабинет юнит-экономики WB",
            "client": "Шумейко и Партнеры",
            "period": report_period,
            "reportPeriod": report_period,
            "periodText": analysis_period_text(
                readme.get("Период анализа"),
                REPORT_PERIOD_TEXT,
            ),
            "periodStatus": demo._text(readme.get("Статус периода"))
            or "предварительный: июнь неполный",
            "sourceCoverage": source_coverage,
            "sourceCoverageStart": source_coverage_start,
            "sourceCoverageEnd": source_coverage_end,
            "methodologyVersion": demo._text(readme.get("Версия методики"))
            or "Excel MVP / 2026-06-19",
            "generatedAt": datetime.now().strftime("%d.%m.%Y %H:%M"),
            "sourceWorkbook": workbook_path.name,
            "returnReasonLimitation": RETURN_REASON_LIMITATION,
        },
        "options": options(rows, document_reconciliation=document_reconciliation),
        "monthly": monthly_rows_with_dates(
            demo._read_monthly_from_sheet(workbook),
            report_period,
        ),
        "expenses": demo._read_expenses_from_sheet(workbook),
        "unitRows": rows,
        "returns": [],
        "lostSales": lost_sales_rows(workbook),
        "lostSalesCoverage": lost_sales_coverage(workbook, report_period),
        "reconciliation": reconciliation,
        "reconciliationMonthly": reconciliation_monthly,
        "documentReconciliation": document_reconciliation,
        "taxInputReconciliation": tax_input_reconciliation_rows(rows),
    }


def tax_input_reconciliation_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (
            demo._text(row.get("week")),
            demo._text(row.get("cabinet")),
            demo._text(row.get("organization")),
        )
        bucket = buckets.setdefault(
            key,
            {
                "week": key[0],
                "weekEnd": "",
                "cabinet": key[1],
                "organization": key[2],
                "vatInputFromWb": 0.0,
                "vatInputFromWbCharges": 0.0,
                "vatInputFromWbReversals": 0.0,
                "vatInputFrom1c": 0.0,
                "vatInputFrom1cCharges": 0.0,
                "vatInputFrom1cReversals": 0.0,
                "sourceRowCount": 0,
                "statuses": set(),
            },
        )
        wb_value = float(row.get("vatInputFromWb") or 0)
        onec_value = float(row.get("vatInputFrom1c") or 0)
        bucket["vatInputFromWb"] += wb_value
        bucket["vatInputFrom1c"] += onec_value
        bucket["sourceRowCount"] += 1
        status = demo._text(row.get("vatInputCompleteness"))
        if status:
            bucket["statuses"].add(status)
    result = []
    for bucket in buckets.values():
        statuses = bucket.pop("statuses")
        bucket["vatInputFromWbCharges"] = max(bucket["vatInputFromWb"], 0)
        bucket["vatInputFromWbReversals"] = min(bucket["vatInputFromWb"], 0)
        bucket["vatInputFrom1cCharges"] = max(bucket["vatInputFrom1c"], 0)
        bucket["vatInputFrom1cReversals"] = min(bucket["vatInputFrom1c"], 0)
        bucket["vatInputDifference"] = round(
            bucket["vatInputFrom1c"] - bucket["vatInputFromWb"],
            2,
        )
        onec_has_documents = bool(
            bucket["vatInputFrom1cCharges"] or bucket["vatInputFrom1cReversals"]
        )
        bucket["vatInputCompleteness"] = (
            _worse_tax_input_status(statuses) if onec_has_documents else "missing"
        )
        bucket["wbEvidenceStatus"] = (
            "confirmed" if bucket["vatInputFromWb"] else "missing"
        )
        bucket["onecEvidenceStatus"] = (
            "confirmed" if onec_has_documents else "missing"
        )
        bucket["vatDeductionMode"] = "unknown"
        bucket["wbSource"] = "WB weekly realization report"
        bucket["onecSource"] = (
            "1C confirming documents" if onec_has_documents else "missing"
        )
        result.append(bucket)
    result.sort(
        key=lambda item: (
            abs(float(item["vatInputDifference"])),
            item["week"],
        ),
        reverse=True,
    )
    for index, bucket in enumerate(result, start=1):
        bucket["id"] = f"tax-input-reconciliation-{index}"
    return result


def _worse_tax_input_status(statuses: set[str]) -> str:
    priority = {"mismatch": 30, "partial": 20, "missing": 10, "confirmed": 0}
    if not statuses:
        return "missing"
    return max(statuses, key=lambda status: priority.get(status, 0))
