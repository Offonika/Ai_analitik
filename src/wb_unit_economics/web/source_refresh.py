from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import time
import zipfile
from calendar import monthrange
from collections import defaultdict
from collections.abc import Callable, Iterable, Mapping
from contextlib import suppress
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from scripts.build_excel_mvp_from_snapshots import build_excel_mvp_from_args
from wb_unit_economics.calculation import METHODOLOGY_VERSION, week_bounds
from wb_unit_economics.contracts import (
    MarketplaceFinanceDailyFact as MarketplaceFinanceDailyFactContract,
)
from wb_unit_economics.contracts import (
    SkuMapping,
    WbSalesReportSummaryRow,
)
from wb_unit_economics.logistics_analysis import (
    LOGISTICS_FACTORS_METHODOLOGY_VERSION,
    LOGISTICS_MEASUREMENTS_METHODOLOGY_VERSION,
    LOGISTICS_ROUTES_METHODOLOGY_VERSION,
    LOGISTICS_TARIFFS_METHODOLOGY_VERSION,
    LogisticsAnalysisResult,
    LogisticsInputDiagnostics,
    LogisticsSourceRow,
    UnitEconomicsSlice,
    build_dimension_rows,
    build_logistics_analysis,
    build_measurement_rows,
    build_route_rows,
    build_tariff_rows,
    source_row_from_payload,
)
from wb_unit_economics.onec_odata import (
    ACCOUNTING_REPORT_SAMPLE_COLLECTIONS,
    DEFAULT_SAMPLE_COLLECTIONS,
    GROSS_PROFIT_SAMPLE_COLLECTIONS,
    INPUT_VAT_SAMPLE_COLLECTIONS,
    SERVICE_SAMPLE_COLLECTIONS,
    TAX_PROFILE_SAMPLE_COLLECTIONS,
    OnecODataConfigError,
    OnecODataMetadataCheckResult,
    OnecODataSettings,
    OnecSampleExportResult,
    check_onec_odata_metadata_with_retry,
    export_onec_accounting_balance_and_turnovers,
    export_onec_accounting_recordtype_balances,
    export_onec_samples,
)
from wb_unit_economics.ozon import (
    OzonConfigError,
    OzonPageResult,
    OzonSettings,
    export_ozon_b2b_sales_json,
    export_ozon_cash_flow,
    export_ozon_mutual_settlement,
    export_ozon_products_buyout,
    export_ozon_products_report,
    export_ozon_realization,
    export_ozon_realization_posting,
    export_ozon_returns_report,
    export_ozon_stock_on_warehouses,
    ozon_settings_from_secret,
)
from wb_unit_economics.ozon_mart import (
    _iter_realization_items,
    _realization_amount,
    _realization_expenses,
    _realization_quantity,
)
from wb_unit_economics.return_reason_analysis import build_return_reason_analysis
from wb_unit_economics.source_integrity import (
    RawIntegrityError,
    iter_json_array,
    verify_raw_directory,
)
from wb_unit_economics.wb_content import (
    WbContentSettings,
    WbProductCardsPageResult,
    export_wb_product_cards,
)
from wb_unit_economics.wb_content import (
    WbSellerAccount as WbContentSellerAccount,
)
from wb_unit_economics.wb_documents import (
    WbDocumentExportResult,
    export_wb_documents,
    load_wb_document_export_results,
)
from wb_unit_economics.wb_finance import (
    WbFinanceConfigError,
    WbFinancePageResult,
    WbFinanceSellerAccount,
    WbFinanceSettings,
    WbSalesReportListPageResult,
    export_wb_finance,
    export_wb_sales_report_list,
    load_wb_finance_export_results,
    load_wb_sales_report_summary_rows,
    resume_wb_finance_export,
    wb_finance_export_is_complete,
)
from wb_unit_economics.wb_goods_return import (
    GoodsReturnSourceRow,
    WbGoodsReturnClient,
    WbGoodsReturnExportResult,
    export_wb_goods_return,
    normalize_goods_return_source_row,
)
from wb_unit_economics.wb_measurements import (
    WbMeasurementExportResult,
    WbMeasurementsClient,
    export_wb_measurement_penalties,
    export_wb_warehouse_measurements,
)
from wb_unit_economics.wb_return_claims import (
    ClaimSourceRow,
    WbReturnClaimsClient,
    WbReturnClaimsExportResult,
    export_wb_return_claims,
    normalize_claim_source_row,
)
from wb_unit_economics.wb_stocks import (
    WbStockExportResult,
    export_wb_stock_history_daily,
)
from wb_unit_economics.wb_supplier_sales import (
    WbSupplierSalesClient,
    WbSupplierSalesExportResult,
    export_wb_supplier_sales,
)
from wb_unit_economics.wb_tariffs import (
    WbTariffsClient,
    WbTariffsExportResult,
    build_tariff_snapshot_dates,
    export_wb_tariffs,
)
from wb_unit_economics.web import integrations, mapping_service, repository, security
from wb_unit_economics.web.dashboard_payload import build_dashboard_payload
from wb_unit_economics.web.models import (
    ClientCompany,
    MarketplaceOperationFact,
    ReportLogisticsAnalysisContext,
    ReportLogisticsTariffRow,
    ReportRun,
    ReportUnitRow,
    SourceRefreshCollection,
    SourceRefreshRun,
    SourceRefreshStageEvent,
    SourceRefreshTask,
    SourceSnapshotRow,
    Tenant,
    TenantIntegration,
    User,
    WbCabinet,
)
from wb_unit_economics.web.models import (
    MarketplaceFinanceDailyFact as MarketplaceFinanceDailyFactModel,
)
from wb_unit_economics.web.report_kinds import (
    ACCOUNTING_REPORT_KINDS,
    MONTH_CLOSE_CONTROL,
)
from wb_unit_economics.web.reports.evidence import (
    AccountingEvidenceSource,
    materialize_accounting_evidence,
)
from wb_unit_economics.web.settings import WebSettings

SOURCE_REFRESH_MODES = {
    "daily",
    "incremental",
    "weekly",
    "full",
    "onec-only",
    "ozon-only",
}


def default_period_for_mode(
    settings: WebSettings,
    mode: str,
) -> tuple[date, date]:
    if mode not in SOURCE_REFRESH_MODES:
        raise SourceRefreshConfigError("unsupported source refresh mode")
    configured_start = date.fromisoformat(settings.source_refresh_period_start)
    configured_end = settings.source_refresh_period_end.strip()
    if configured_end:
        period_end = date.fromisoformat(configured_end)
    else:
        period_end = datetime.now(tz=MOSCOW_TZ).date() - timedelta(days=1)
    if mode == "daily":
        rolling_start = period_end - timedelta(
            days=max(1, settings.source_refresh_rolling_window_days) - 1
        )
        return max(configured_start, rolling_start), period_end
    return configured_start, period_end


def _incremental_yesterday() -> date:
    return datetime.now(tz=MOSCOW_TZ).date() - timedelta(days=1)


SOURCE_REFRESH_RESUME_MODES = {"auto", "never"}
ONEC_RESUME_MODES = {
    "daily",
    "incremental",
    "weekly",
    "full",
    "onec-only",
    "ozon-only",
}
WB_REQUIRED_MODES = {"daily", "incremental", "weekly", "full"}
OZON_REQUIRED_MODES = {"ozon-only"}
OZON_OPTIONAL_MODES = {"daily", "weekly", "full"}
OZON_TYPED_FILE_AUTHORITATIVE_TYPES = {
    "ozon_finance_cash_flow",
    "ozon_realization",
    "ozon_realization_posting",
    "ozon_mutual_settlement",
    "ozon_products_buyout",
    "ozon_b2b_sales_json",
    "ozon_products_report",
    "ozon_stock_on_warehouses",
    "ozon_returns_report",
}
CREDENTIAL_SOURCES = {"tenant", "env"}
SOURCE_SNAPSHOT_ROW_CHUNK_SIZE = 1000
ONEC_DATABASE_ROW_PERSIST_MAX_BYTES = 25 * 1024 * 1024
READY_INTEGRATION_STATUSES = {"configured", "check_ok"}
ONEC_REFRESH_COLLECTIONS = (
    *DEFAULT_SAMPLE_COLLECTIONS,
    *TAX_PROFILE_SAMPLE_COLLECTIONS,
    *INPUT_VAT_SAMPLE_COLLECTIONS,
    *GROSS_PROFIT_SAMPLE_COLLECTIONS,
    *SERVICE_SAMPLE_COLLECTIONS,
)
MANDATORY_ONEC_COLLECTION_IDS = {
    "nomenclature",
    "organizations",
    "barcodes",
    "sales_register",
}
PUBLICATION_REQUIRED_ONEC_COLLECTION_IDS = {"commissioner_reports"}
MANDATORY_OK_STATUSES = {"loaded", "empty_expected"}
OPTIONAL_OK_STATUSES = {"loaded", "empty_expected"}
REVIEW_STATUSES = {"needs_review", "stale", "partial_source"}
WB_FINANCE_REFRESH_ROLES = {"finance_reports", "full_readonly"}
WB_STOCK_HISTORY_REFRESH_ROLES = {
    "finance_reports",
    "stocks_analytics",
    "full_readonly",
}
OZON_REFRESH_ROLES = {
    "finance_reports",
    "products_catalog",
    "stocks_analytics",
    "returns_reports",
    "full_readonly",
}
MOSCOW_TZ = ZoneInfo("Europe/Moscow")
WB_FINANCE_MAX_PAGES_ISSUE = "max_pages_reached_with_next_rrd_id"
WB_REPORT_LIST_MAX_PAGES_ISSUE = "max_pages_reached_with_full_report_list_page"
WB_PRODUCT_CARDS_MAX_PAGES_ISSUE = "max_pages_reached_with_full_cards_page"
WB_GOODS_RETURN_MAX_DAYS = 31


def _default_wb_tariffs_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
) -> list[WbTariffsExportResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    factor_snapshot_date = datetime.now(tz=MOSCOW_TZ).date()
    snapshot_dates = build_tariff_snapshot_dates(
        period_start,
        period_end,
        factor_snapshot_date=factor_snapshot_date,
    )
    ordered_dates = (factor_snapshot_date,) + tuple(
        item for item in snapshot_dates if item != factor_snapshot_date
    )
    results: list[WbTariffsExportResult] = []
    for account in accounts:
        client = WbTariffsClient(api_key=account.api_key)
        seller_account_id = str(account.seller_account_id)
        file_prefix = hashlib.sha256(seller_account_id.encode("utf-8")).hexdigest()[:12]
        rate_limited = False
        for index, target_date in enumerate(ordered_dates):
            if rate_limited:
                results.append(
                    WbTariffsExportResult(
                        ok=False,
                        seller_account_id=seller_account_id,
                        target_date=target_date,
                        status_code=429,
                        error="RateLimitSkipped",
                    )
                )
                continue
            result = export_wb_tariffs(
                client,
                output_dir,
                target_date=target_date,
                seller_account_id=seller_account_id,
                file_prefix=file_prefix,
            )
            results.append(result)
            rate_limited = result.status_code == 429
            if index + 1 < len(ordered_dates) and not rate_limited:
                time.sleep(1.05)
    manifest = {
        "source": "wb_tariffs",
        "factorSnapshotDate": factor_snapshot_date.isoformat(),
        "requestedWeekStarts": [
            item.isoformat() for item in snapshot_dates if item != factor_snapshot_date
        ],
        "results": [_wb_tariffs_result_payload(item) for item in results],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return results


def _default_wb_goods_return_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
) -> list[WbGoodsReturnExportResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    date_from = max(
        period_start, period_end - timedelta(days=WB_GOODS_RETURN_MAX_DAYS - 1)
    )
    results: list[WbGoodsReturnExportResult] = []
    for account in accounts:
        seller_account_id = str(account.seller_account_id)
        file_prefix = hashlib.sha256(seller_account_id.encode("utf-8")).hexdigest()[:12]
        client = WbGoodsReturnClient(api_key=account.api_key)
        results.append(
            export_wb_goods_return(
                client,
                output_dir,
                date_from=date_from,
                date_to=period_end,
                seller_account_id=seller_account_id,
                account_name=str(account.account_name),
                file_prefix=file_prefix,
            )
        )
    manifest = {
        "source": "wb_goods_return",
        "periodStart": period_start.isoformat(),
        "periodEnd": period_end.isoformat(),
        "coverageStart": date_from.isoformat(),
        "coverageEnd": period_end.isoformat(),
        "results": [_wb_goods_return_result_payload(item) for item in results],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return results


def _default_wb_return_claims_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
) -> list[WbReturnClaimsExportResult]:
    del period_start, period_end
    output_dir.mkdir(parents=True, exist_ok=True)
    as_of = datetime.now(tz=MOSCOW_TZ).date()
    results: list[WbReturnClaimsExportResult] = []
    for account in accounts:
        seller_account_id = str(account.seller_account_id)
        file_prefix = hashlib.sha256(seller_account_id.encode("utf-8")).hexdigest()[:12]
        client = WbReturnClaimsClient(api_key=account.api_key)
        results.append(
            export_wb_return_claims(
                client,
                output_dir,
                as_of=as_of,
                seller_account_id=seller_account_id,
                account_name=str(account.account_name),
                file_prefix=file_prefix,
            )
        )
    manifest = {
        "source": "wb_return_claims",
        "coverageStart": (
            min(
                (
                    item.coverage_start
                    for item in results
                    if item.coverage_start is not None
                ),
                default=None,
            ).isoformat()
            if results and any(item.coverage_start is not None for item in results)
            else ""
        ),
        "coverageEnd": (
            max(
                (
                    item.coverage_end
                    for item in results
                    if item.coverage_end is not None
                ),
                default=None,
            ).isoformat()
            if results and any(item.coverage_end is not None for item in results)
            else ""
        ),
        "results": [_wb_return_claims_result_payload(item) for item in results],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return results


def _default_wb_supplier_sales_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
) -> list[WbSupplierSalesExportResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    factor_snapshot_date = datetime.now(tz=MOSCOW_TZ).date()
    earliest_available = factor_snapshot_date - timedelta(days=89)
    has_provider_overlap = (
        period_start <= factor_snapshot_date and period_end >= earliest_available
    )
    date_from = max(period_start, earliest_available) if has_provider_overlap else None
    coverage_end = min(period_end, factor_snapshot_date) if date_from else None
    results: list[WbSupplierSalesExportResult] = []
    for account in accounts:
        seller_account_id = str(account.seller_account_id)
        if date_from is None:
            results.append(
                WbSupplierSalesExportResult(
                    ok=False,
                    seller_account_id=seller_account_id,
                    account_name=str(account.account_name),
                    error="ProviderWindowUnavailable",
                )
            )
            continue
        client = WbSupplierSalesClient(api_key=account.api_key)
        file_prefix = hashlib.sha256(seller_account_id.encode("utf-8")).hexdigest()[:12]
        results.append(
            export_wb_supplier_sales(
                client,
                output_dir,
                date_from=date_from,
                seller_account_id=seller_account_id,
                account_name=str(account.account_name),
                file_prefix=file_prefix,
            )
        )
    manifest = {
        "source": "wb_supplier_sales",
        "factorSnapshotDate": factor_snapshot_date.isoformat(),
        "coverageStart": date_from.isoformat() if date_from else "",
        "coverageEnd": coverage_end.isoformat() if coverage_end else "",
        "results": [_wb_supplier_sales_result_payload(item) for item in results],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return results


def _default_wb_measurements_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
    source_type: str,
) -> list[WbMeasurementExportResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    start_at = datetime.combine(
        period_start,
        datetime.min.time(),
        tzinfo=MOSCOW_TZ,
    ).astimezone(UTC)
    end_at = datetime.combine(
        period_end + timedelta(days=1),
        datetime.min.time(),
        tzinfo=MOSCOW_TZ,
    ).astimezone(UTC) - timedelta(microseconds=1)
    factor_snapshot_at = datetime.now(tz=UTC)
    results: list[WbMeasurementExportResult] = []
    for account in accounts:
        seller_account_id = str(account.seller_account_id)
        file_prefix = hashlib.sha256(seller_account_id.encode("utf-8")).hexdigest()[:12]
        client = WbMeasurementsClient(api_key=account.api_key)
        exporter = (
            export_wb_measurement_penalties
            if source_type == "wb_measurement_penalties"
            else export_wb_warehouse_measurements
        )
        results.append(
            exporter(
                client,
                output_dir,
                date_from=start_at,
                date_to=end_at,
                seller_account_id=seller_account_id,
                account_name=str(account.account_name),
                file_prefix=file_prefix,
            )
        )
    manifest = {
        "source": source_type,
        "periodStart": period_start.isoformat(),
        "periodEnd": period_end.isoformat(),
        "coverageStart": period_start.isoformat(),
        "coverageEnd": period_end.isoformat(),
        "factorSnapshotAt": factor_snapshot_at.isoformat(),
        "results": [_wb_measurement_result_payload(item) for item in results],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return results


def _default_wb_measurement_penalties_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
) -> list[WbMeasurementExportResult]:
    return _default_wb_measurements_exporter(
        accounts,
        output_dir,
        period_start=period_start,
        period_end=period_end,
        source_type="wb_measurement_penalties",
    )


def _default_wb_warehouse_measurements_exporter(
    accounts: Any,
    output_dir: Path,
    *,
    period_start: date,
    period_end: date,
) -> list[WbMeasurementExportResult]:
    return _default_wb_measurements_exporter(
        accounts,
        output_dir,
        period_start=period_start,
        period_end=period_end,
        source_type="wb_warehouse_measurements",
    )


class SourceRefreshDisabledError(RuntimeError):
    pass


class SourceRefreshBusyError(RuntimeError):
    pass


class SourceRefreshConfigError(RuntimeError):
    pass


_RETRYABLE_SOURCE_HTTP_STATUSES = {408, 425, 429, 500, 502, 503, 504}
_PERMANENT_SOURCE_HTTP_STATUSES = {400, 401, 403, 404, 409, 422}
_TRANSIENT_SOURCE_ERROR_MARKERS = (
    "connecterror",
    "connecttimeout",
    "networkerror",
    "pooltimeout",
    "readerror",
    "readtimeout",
    "remoteprotocolerror",
    "service_unavailable",
    "temporary",
    "timeout",
    "transport",
    "writeerror",
    "writetimeout",
)
_PERMANENT_SOURCE_ERROR_MARKERS = (
    "authorization",
    "credential",
    "forbidden",
    "mapping",
    "partial_source",
    "unauthorized",
)


def _source_refresh_exception_is_transient(exc: Exception) -> bool:
    response = getattr(exc, "response", None)
    status_code = getattr(response, "status_code", None)
    if isinstance(status_code, int):
        return status_code in _RETRYABLE_SOURCE_HTTP_STATUSES
    if isinstance(exc, OSError):
        return True
    class_names = " ".join(
        item.__name__.lower() for item in exc.__class__.__mro__
    )
    return any(marker in class_names for marker in _TRANSIENT_SOURCE_ERROR_MARKERS)


def _collection_failure_is_transient(
    collection: SourceRefreshCollection,
) -> bool:
    if collection.status != "failed":
        return False
    payload = collection.payload if isinstance(collection.payload, Mapping) else {}
    nodes: list[Mapping[str, Any]] = [payload]
    results = payload.get("results")
    if isinstance(results, list):
        nodes.extend(item for item in results if isinstance(item, Mapping))
    status_codes: set[int] = set()
    retryable = False
    errors = [str(collection.error_message or "")]
    for node in nodes:
        value = node.get("statusCode", node.get("httpStatus"))
        try:
            if value is not None:
                status_codes.add(int(value))
        except (TypeError, ValueError):
            pass
        retryable = retryable or node.get("retryable") is True
        errors.extend(
            str(node.get(key) or "")
            for key in ("error", "errorType", "message")
        )
    if status_codes & _PERMANENT_SOURCE_HTTP_STATUSES:
        return False
    error_text = " ".join(errors).lower()
    if any(marker in error_text for marker in _PERMANENT_SOURCE_ERROR_MARKERS):
        return False
    if status_codes & _RETRYABLE_SOURCE_HTTP_STATUSES or retryable:
        return True
    return any(marker in error_text for marker in _TRANSIENT_SOURCE_ERROR_MARKERS)


def _metadata_failure_is_transient(
    result: OnecODataMetadataCheckResult,
) -> bool:
    if result.status_code in _PERMANENT_SOURCE_HTTP_STATUSES:
        return False
    if result.status_code in _RETRYABLE_SOURCE_HTTP_STATUSES:
        return True
    error = str(result.error or "").lower()
    if any(marker in error for marker in _PERMANENT_SOURCE_ERROR_MARKERS):
        return False
    return any(marker in error for marker in _TRANSIENT_SOURCE_ERROR_MARKERS)


@dataclass(frozen=True)
class SourceCredentials:
    wb_settings: WbFinanceSettings | None
    onec_settings: OnecODataSettings | None
    ozon_settings: OzonSettings | None
    wb_cabinet_ids: dict[str, str]
    ozon_cabinet_ids: dict[str, str]
    issues: tuple[dict[str, Any], ...]
    optional_issues: tuple[dict[str, Any], ...] = ()


@dataclass(frozen=True)
class CollectorContext:
    db: Session
    refresh_run: SourceRefreshRun
    credentials: SourceCredentials
    root_dir: Path
    period_start: date
    period_end: date
    mode: str


@dataclass(frozen=True)
class CollectorResult:
    collection: SourceRefreshCollection | None = None
    output_dir: Path | None = None


@dataclass(frozen=True)
class SourceCollector:
    source_type: str
    label: str
    required: bool
    modes: frozenset[str]
    roles: frozenset[str]
    collect: Callable[[SourceRefreshService, CollectorContext], CollectorResult]

    def supports(self, mode: str) -> bool:
        return mode in self.modes


@dataclass(frozen=True)
class CollectorOutputs:
    output_dirs: dict[str, Path]
    mapping_collection: SourceRefreshCollection


@dataclass(frozen=True)
class PersistedPipelineInputs:
    source_report: ReportRun | None
    base_refresh_run: SourceRefreshRun | None
    mapping_collection: SourceRefreshCollection
    wb_finance_dir: Path | None
    wb_report_list_dir: Path | None
    wb_cards_dir: Path | None
    wb_stock_history_dir: Path | None
    onec_dir: Path | None
    composite_rebuild: bool


class SourceRefreshService:
    def __init__(
        self,
        settings: WebSettings,
        *,
        wb_finance_exporter: Callable[..., list[WbFinancePageResult]] = (
            export_wb_finance
        ),
        wb_report_list_exporter: Callable[
            ..., list[WbSalesReportListPageResult]
        ] = export_wb_sales_report_list,
        wb_documents_exporter: Callable[..., list[WbDocumentExportResult]] = (
            export_wb_documents
        ),
        wb_product_cards_exporter: Callable[
            ..., list[WbProductCardsPageResult]
        ] = export_wb_product_cards,
        wb_stock_history_exporter: Callable[
            ..., list[WbStockExportResult]
        ] = export_wb_stock_history_daily,
        wb_tariffs_exporter: Callable[
            ..., list[WbTariffsExportResult]
        ] = _default_wb_tariffs_exporter,
        wb_goods_return_exporter: Callable[
            ..., list[WbGoodsReturnExportResult]
        ] = _default_wb_goods_return_exporter,
        wb_return_claims_exporter: Callable[
            ..., list[WbReturnClaimsExportResult]
        ] = _default_wb_return_claims_exporter,
        wb_supplier_sales_exporter: Callable[
            ..., list[WbSupplierSalesExportResult]
        ] = _default_wb_supplier_sales_exporter,
        wb_measurement_penalties_exporter: Callable[
            ..., list[WbMeasurementExportResult]
        ] = _default_wb_measurement_penalties_exporter,
        wb_warehouse_measurements_exporter: Callable[
            ..., list[WbMeasurementExportResult]
        ] = _default_wb_warehouse_measurements_exporter,
        ozon_cash_flow_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_cash_flow
        ),
        ozon_realization_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_realization
        ),
        ozon_realization_posting_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_realization_posting
        ),
        ozon_products_buyout_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_products_buyout
        ),
        ozon_b2b_sales_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_b2b_sales_json
        ),
        ozon_mutual_settlement_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_mutual_settlement
        ),
        ozon_products_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_products_report
        ),
        ozon_stocks_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_stock_on_warehouses
        ),
        ozon_returns_exporter: Callable[..., list[OzonPageResult]] = (
            export_ozon_returns_report
        ),
        onec_exporter: Callable[..., list[OnecSampleExportResult]] = (
            export_onec_samples
        ),
        onec_accounting_balance_exporter: Callable[
            ..., OnecSampleExportResult
        ] = export_onec_accounting_balance_and_turnovers,
        onec_accounting_recordtype_exporter: Callable[
            ..., OnecSampleExportResult
        ] = export_onec_accounting_recordtype_balances,
        onec_metadata_checker: Callable[
            [OnecODataSettings], OnecODataMetadataCheckResult
        ]
        | None = None,
        workbook_builder: Callable[[argparse.Namespace], Any] = (
            build_excel_mvp_from_args
        ),
        dashboard_payload_builder: Callable[[Path], dict[str, Any]] = (
            build_dashboard_payload
        ),
    ) -> None:
        self.settings = settings
        self._wb_finance_exporter = wb_finance_exporter
        self._wb_report_list_exporter = wb_report_list_exporter
        self._wb_documents_exporter = wb_documents_exporter
        self._wb_product_cards_exporter = wb_product_cards_exporter
        self._wb_stock_history_exporter = wb_stock_history_exporter
        self._wb_tariffs_exporter = wb_tariffs_exporter
        self._wb_goods_return_exporter = wb_goods_return_exporter
        self._wb_return_claims_exporter = wb_return_claims_exporter
        self._wb_supplier_sales_exporter = wb_supplier_sales_exporter
        self._wb_measurement_penalties_exporter = wb_measurement_penalties_exporter
        self._wb_warehouse_measurements_exporter = wb_warehouse_measurements_exporter
        self._ozon_cash_flow_exporter = ozon_cash_flow_exporter
        self._ozon_realization_exporter = ozon_realization_exporter
        self._ozon_realization_posting_exporter = ozon_realization_posting_exporter
        self._ozon_products_buyout_exporter = ozon_products_buyout_exporter
        self._ozon_b2b_sales_exporter = ozon_b2b_sales_exporter
        self._ozon_mutual_settlement_exporter = ozon_mutual_settlement_exporter
        self._ozon_products_exporter = ozon_products_exporter
        self._ozon_stocks_exporter = ozon_stocks_exporter
        self._ozon_returns_exporter = ozon_returns_exporter
        self._onec_exporter = onec_exporter
        self._onec_accounting_balance_exporter = onec_accounting_balance_exporter
        self._onec_accounting_recordtype_exporter = onec_accounting_recordtype_exporter
        self._onec_metadata_checker = (
            onec_metadata_checker or check_onec_odata_metadata_with_retry
        )
        self._workbook_builder = workbook_builder
        self._dashboard_payload_builder = dashboard_payload_builder

    def run(
        self,
        db: Session,
        *,
        tenant_id: str,
        client_id: str | None = None,
        mode: str,
        credential_source: str = "tenant",
        dry_run: bool = False,
        user: User | None = None,
        source_report: ReportRun | None = None,
        reason: str = "",
        period_start: date | None = None,
        period_end: date | None = None,
        resume_mode: str = "auto",
        resume_from_run_id: str | None = None,
    ) -> dict[str, Any]:
        refresh_run = self._create_refresh_run(
            db,
            tenant_id=tenant_id,
            client_id=client_id,
            mode=mode,
            credential_source=credential_source,
            dry_run=dry_run,
            user=user,
            source_report=source_report,
            reason=reason,
            period_start=period_start,
            period_end=period_end,
            resume_mode=resume_mode,
            resume_from_run_id=resume_from_run_id,
        )
        if isinstance(refresh_run, dict):
            return refresh_run
        return self.run_existing(db, refresh_run.id)

    def enqueue(
        self,
        db: Session,
        *,
        tenant_id: str,
        client_id: str | None = None,
        mode: str,
        credential_source: str = "tenant",
        user: User | None = None,
        source_report: ReportRun | None = None,
        reason: str = "",
        period_start: date | None = None,
        period_end: date | None = None,
        resume_mode: str = "auto",
        resume_from_run_id: str | None = None,
    ) -> dict[str, Any]:
        refresh_run = self._create_refresh_run(
            db,
            tenant_id=tenant_id,
            client_id=client_id,
            mode=mode,
            credential_source=credential_source,
            dry_run=False,
            user=user,
            source_report=source_report,
            reason=reason,
            period_start=period_start,
            period_end=period_end,
            resume_mode=resume_mode,
            resume_from_run_id=resume_from_run_id,
        )
        if isinstance(refresh_run, dict):
            return refresh_run
        return repository.source_refresh_run_payload(refresh_run)

    def run_existing(
        self,
        db: Session,
        refresh_run_id: str,
        *,
        worker_id: str = "",
        stop_after_sources: bool = False,
    ) -> dict[str, Any]:
        refresh_run = db.scalar(
            select(SourceRefreshRun)
            .where(SourceRefreshRun.id == refresh_run_id)
            .with_for_update()
        )
        if refresh_run is None:
            raise SourceRefreshConfigError(
                f"source refresh run not found: {refresh_run_id}"
            )
        if refresh_run.finished_at is not None:
            return repository.source_refresh_run_payload(refresh_run)
        if (
            refresh_run.status != "queued"
            and worker_id
            and refresh_run.worker_id
            and refresh_run.worker_id != worker_id
        ):
            raise SourceRefreshBusyError(
                "source refresh run is already owned by another worker"
            )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            worker_id=worker_id or refresh_run.worker_id,
            heartbeat_at=security.utcnow(),
        )
        db.commit()
        user = (
            db.get(User, refresh_run.requested_by_user_id)
            if refresh_run.requested_by_user_id
            else None
        )
        if refresh_run.mode == "report-generation":
            return self._execute_accounting_report_generation(
                db,
                refresh_run,
                user=user,
            )
        return self._execute_run(
            db,
            refresh_run,
            user=user,
            stop_after_sources=stop_after_sources,
        )

    def run_split_materialize_task(
        self,
        db: Session,
        task: SourceRefreshTask,
        *,
        worker_id: str,
    ) -> dict[str, Any]:
        """Materialize typed facts from persisted raw collections only."""
        refresh_run = db.get(SourceRefreshRun, task.refresh_run_id)
        if refresh_run is None:
            raise SourceRefreshConfigError("source refresh run not found")
        if (
            task.task_type != "materialize_facts"
            or task.status != "running"
            or task.worker_id != worker_id
        ):
            raise SourceRefreshBusyError("materialize task is not owned by worker")
        db.info["source_refresh_split_pipeline"] = True
        repository.update_source_refresh_run(
            db,
            refresh_run,
            worker_id=worker_id,
            heartbeat_at=security.utcnow(),
        )
        event = repository.begin_source_refresh_stage(
            db,
            refresh_run,
            stage="materialize_facts",
            task=task,
        )
        db.commit()
        try:
            inputs = self._persisted_pipeline_inputs(db, refresh_run)
            if (
                refresh_run.mode in {"daily", "incremental", "weekly", "full"}
                and self.settings.marketplace_daily_facts_enabled
                and inputs.wb_finance_dir is not None
                and inputs.onec_dir is not None
            ):
                self._materialize_wb_daily_facts(
                    db,
                    refresh_run,
                    wb_finance_dir=inputs.wb_finance_dir,
                    onec_dir=inputs.onec_dir,
                    wb_report_list_dir=inputs.wb_report_list_dir,
                    wb_cards_dir=inputs.wb_cards_dir,
                    wb_stock_history_dir=inputs.wb_stock_history_dir,
                )
            if refresh_run.mode == "incremental":
                coverage_issue = self._daily_facts_coverage_issue(
                    db,
                    tenant_id=refresh_run.tenant_id,
                    client_id=refresh_run.client_id,
                    period_start=refresh_run.period_start,
                    period_end=refresh_run.period_end,
                )
                if coverage_issue:
                    raise SourceRefreshConfigError(coverage_issue)
            now = security.utcnow()
            stage_metrics = repository.source_refresh_stage_completion_metrics(
                db,
                refresh_run,
                stage="materialize_facts",
            )
            repository.complete_source_refresh_task(
                db,
                task,
                metrics=stage_metrics,
                finished_at=now,
            )
            repository.finish_source_refresh_stage(
                db,
                event,
                status="succeeded",
                metrics=stage_metrics,
                finished_at=now,
            )
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="rebuilding",
                worker_id="",
                heartbeat_at=now,
            )
            db.commit()
            return repository.source_refresh_run_payload(refresh_run)
        except Exception as exc:
            self._fail_split_task(
                db,
                task_id=task.id,
                event_id=event.id,
                safe_error_code="materialize_facts_failed",
                exc=exc,
            )
            refresh_run = db.get(SourceRefreshRun, task.refresh_run_id)
            if refresh_run is None:
                raise
            return repository.source_refresh_run_payload(refresh_run)
        finally:
            db.info.pop("source_refresh_split_pipeline", None)

    def run_split_build_report_task(
        self,
        db: Session,
        task: SourceRefreshTask,
        *,
        worker_id: str,
    ) -> dict[str, Any]:
        """Build DB marts and a staff draft without running any export."""
        refresh_run = db.get(SourceRefreshRun, task.refresh_run_id)
        if refresh_run is None:
            raise SourceRefreshConfigError("source refresh run not found")
        if (
            task.task_type != "build_report"
            or task.status != "running"
            or task.worker_id != worker_id
        ):
            raise SourceRefreshBusyError("build task is not owned by worker")
        if not self.settings.db_first_reports_enabled:
            raise SourceRefreshConfigError("split pipeline requires DB-first reports")
        db.info["source_refresh_split_pipeline"] = True
        repository.update_source_refresh_run(
            db,
            refresh_run,
            worker_id=worker_id,
            heartbeat_at=security.utcnow(),
        )
        event = repository.begin_source_refresh_stage(
            db,
            refresh_run,
            stage="build_report",
            task=task,
        )
        db.commit()
        try:
            inputs = self._persisted_pipeline_inputs(db, refresh_run)
            contributing_runs: list[SourceRefreshRun] = []
            if refresh_run.mode == "incremental":
                contributing_runs = self._daily_fact_contributing_runs(
                    db,
                    refresh_run,
                )
            wb_summary_rows = (
                self._incremental_wb_summary_rows(
                    refresh_run,
                    base_refresh_run=inputs.base_refresh_run,
                    current_report_list_dir=inputs.wb_report_list_dir,
                )
                if refresh_run.mode == "incremental"
                else []
            )
            finance_collection = next(
                (
                    item
                    for item in refresh_run.collections
                    if item.source_type == "wb_finance_detail"
                ),
                None,
            )
            facts_materialized = bool(
                finance_collection is not None
                and ((finance_collection.payload or {}).get("dailyFacts") or {}).get(
                    "status"
                )
                == "materialized"
            )
            wb_daily_facts = (
                self._daily_facts_for_report(
                    db,
                    refresh_run,
                    wb_summary_rows=wb_summary_rows,
                )
                if facts_materialized
                else None
            )
            report_snapshot_set_id = self._report_snapshot_set_id(
                refresh_run,
                base_refresh_run=(
                    inputs.base_refresh_run
                    if inputs.composite_rebuild
                    or refresh_run.mode == "incremental"
                    else None
                ),
                contributing_runs=contributing_runs,
            )
            report, excel_path = self._build_db_first_report(
                db,
                refresh_run,
                source_report=inputs.source_report,
                wb_finance_dir=inputs.wb_finance_dir,
                onec_dir=inputs.onec_dir,
                wb_report_list_dir=inputs.wb_report_list_dir,
                wb_cards_dir=inputs.wb_cards_dir,
                wb_stock_history_dir=inputs.wb_stock_history_dir,
                source_snapshot_set_id=report_snapshot_set_id,
                base_refresh_run=(
                    inputs.base_refresh_run
                    if inputs.composite_rebuild
                    or refresh_run.mode == "incremental"
                    else None
                ),
                contributing_runs=contributing_runs,
                wb_daily_facts=wb_daily_facts,
                wb_summary_rows=wb_summary_rows,
            )
            primary_document_refresh = (
                inputs.base_refresh_run
                if inputs.composite_rebuild and inputs.base_refresh_run is not None
                else refresh_run
            )
            primary_document_scope = repository.apply_wb_buyout_primary_documents(
                db,
                report,
                primary_document_refresh,
                source_runs=(
                    contributing_runs
                    if refresh_run.mode == "incremental"
                    else ()
                ),
            )
            self._attach_source_loads(
                db,
                report,
                refresh_run,
                contributing_runs=contributing_runs,
            )
            mapping_report_scope = repository.reconcile_report_mapping_source_load(
                db,
                report,
            )
            refresh_run.new_report_run_id = report.id
            repository.link_source_refresh_tasks_to_report(db, refresh_run, report)
            report_row_count = int(
                db.scalar(
                    select(func.count())
                    .select_from(ReportUnitRow)
                    .where(ReportUnitRow.report_run_id == report.id)
                )
                or 0
            )
            now = security.utcnow()
            repository.complete_source_refresh_task(
                db,
                task,
                metrics={"rowCount": report_row_count},
                finished_at=now,
            )
            repository.finish_source_refresh_stage(
                db,
                event,
                status="succeeded",
                metrics={"rowCount": report_row_count},
                finished_at=now,
            )
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="rebuilding",
                worker_id="",
                new_report_run_id=report.id,
                workbook_path="",
                heartbeat_at=now,
            )
            repository.audit(
                db,
                action="source_refresh_report_marts_created",
                user=None,
                tenant_id=refresh_run.tenant_id,
                entity_type="source_refresh_run",
                entity_id=refresh_run.id,
                payload={
                    "newReportRunId": report.id,
                    "snapshotSetId": report_snapshot_set_id,
                    "mappingReportScope": mapping_report_scope,
                    "buyoutPrimaryDocumentScope": primary_document_scope,
                    "automaticPublication": False,
                    "excelPending": True,
                },
            )
            db.commit()
            payload = repository.source_refresh_run_payload(refresh_run)
            payload["pendingWorkbookPath"] = str(excel_path.name)
            return payload
        except Exception as exc:
            self._fail_split_task(
                db,
                task_id=task.id,
                event_id=event.id,
                safe_error_code="build_report_failed",
                exc=exc,
            )
            refresh_run = db.get(SourceRefreshRun, task.refresh_run_id)
            if refresh_run is None:
                raise
            return repository.source_refresh_run_payload(refresh_run)
        finally:
            db.info.pop("source_refresh_split_pipeline", None)

    def _persisted_pipeline_inputs(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
    ) -> PersistedPipelineInputs:
        root_dir = Path(refresh_run.root_dir).resolve()
        allowed_root = self.settings.source_refresh_root_path.resolve()
        if (
            not refresh_run.root_dir
            or not root_dir.is_relative_to(allowed_root)
            or not root_dir.is_dir()
        ):
            raise SourceRefreshConfigError("source refresh raw root is unavailable")

        def subdirectory(name: str) -> Path | None:
            candidate = root_dir / name
            if (
                candidate.is_dir()
                and not candidate.is_symlink()
                and candidate.resolve().is_relative_to(root_dir)
            ):
                return candidate.resolve()
            return None

        mapping_collection = next(
            (
                item
                for item in reversed(refresh_run.collections)
                if item.source_type == "sku_mapping"
            ),
            None,
        )
        if mapping_collection is None:
            raise SourceRefreshConfigError("sku mapping collection is unavailable")
        source_report = (
            db.get(ReportRun, refresh_run.source_report_run_id)
            if refresh_run.source_report_run_id
            else None
        )
        base_refresh_run = (
            db.get(SourceRefreshRun, refresh_run.base_source_refresh_run_id)
            if refresh_run.base_source_refresh_run_id
            else None
        )
        composite_rebuild = bool(
            refresh_run.mode == "onec-only"
            and source_report is not None
            and base_refresh_run is not None
        )
        wb_finance_dir = subdirectory("wb_finance")
        wb_report_list_dir = subdirectory("wb_sales_report_list")
        wb_cards_dir = subdirectory("wb_product_cards")
        wb_stock_history_dir = subdirectory("wb_stock_history_daily")
        if composite_rebuild and base_refresh_run is not None:
            wb_finance_dir = self._required_collection_raw_dir(
                base_refresh_run,
                "wb_finance_detail",
            )
            wb_cards_dir = self._required_collection_raw_dir(
                base_refresh_run,
                "wb_product_cards",
            )
            wb_report_list_dir = self._optional_collection_raw_dir(
                base_refresh_run,
                "wb_sales_report_list",
            )
            wb_stock_history_dir = self._optional_collection_raw_dir(
                base_refresh_run,
                "wb_stock_history_daily",
            )
        return PersistedPipelineInputs(
            source_report=source_report,
            base_refresh_run=base_refresh_run,
            mapping_collection=mapping_collection,
            wb_finance_dir=wb_finance_dir,
            wb_report_list_dir=wb_report_list_dir,
            wb_cards_dir=wb_cards_dir,
            wb_stock_history_dir=wb_stock_history_dir,
            onec_dir=subdirectory("onec"),
            composite_rebuild=composite_rebuild,
        )

    def _fail_split_task(
        self,
        db: Session,
        *,
        task_id: str,
        event_id: int,
        safe_error_code: str,
        exc: Exception,
    ) -> None:
        with suppress(Exception):
            db.rollback()
        task = db.get(SourceRefreshTask, task_id)
        if task is None:
            return
        event = db.get(SourceRefreshStageEvent, event_id)
        now = security.utcnow()
        if task.status == "running":
            repository.fail_source_refresh_task(
                db,
                task,
                safe_error_code=safe_error_code,
                safe_error_message=_safe_error(exc),
                transient=isinstance(exc, OSError),
                failed_at=now,
            )
        if event is not None and event.status == "running":
            repository.finish_source_refresh_stage(
                db,
                event,
                status="failed",
                safe_error_code=safe_error_code,
                finished_at=now,
            )
        refresh_run = db.get(SourceRefreshRun, task.refresh_run_id)
        if refresh_run is not None:
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="failed",
                worker_id="",
                failure_code=safe_error_code,
                error_message=_safe_error(exc),
                finished_at=now,
            )
        db.commit()

    def _execute_accounting_report_generation(
        self,
        db: Session,
        generation: SourceRefreshRun,
        *,
        user: User | None,
    ) -> dict[str, Any]:
        if generation.target_report_kind not in ACCOUNTING_REPORT_KINDS:
            raise SourceRefreshConfigError("unsupported report generation kind")
        root_dir = (
            self.settings.source_refresh_root_path / generation.snapshot_set_id
        ).resolve()
        try:
            repository.update_source_refresh_run(
                db,
                generation,
                status="running",
                started_at=generation.started_at or security.utcnow(),
                heartbeat_at=security.utcnow(),
                root_dir=str(root_dir),
            )
            generation.generation_stage = "refreshing_sources"
            generation.failure_code = ""
            _commit_source_refresh_progress(db)
            disk_issue = self._low_disk_issue()
            if disk_issue is not None:
                raise SourceRefreshConfigError(disk_issue["error_message"])
            credentials = self._credentials(
                db,
                tenant_id=generation.tenant_id,
                credential_source=generation.credential_source,
                mode=generation.mode,
            )
            for issue in (*credentials.issues, *credentials.optional_issues):
                repository.add_source_refresh_collection(
                    db,
                    generation,
                    source_type=issue["source_type"],
                    source_label=issue["source_label"],
                    required=issue["required"],
                    status=issue["status"],
                    error_message=issue.get("error_message", ""),
                    payload=issue.get("payload", {}),
                    organization_id=generation.organization_id,
                )
            if credentials.issues or credentials.onec_settings is None:
                raise SourceRefreshConfigError("onec_readonly_not_ready")
            metadata_result = self._record_onec_metadata_check(
                db,
                generation,
                credentials.onec_settings,
            )
            _commit_source_refresh_progress(db)
            if not metadata_result.ok:
                raise SourceRefreshConfigError("onec_odata_metadata_unavailable")

            output_dir = root_dir / "onec_accounting"
            evidence_start = generation.source_window_start or generation.period_start
            accounting_collections = tuple(
                collection
                for collection in ACCOUNTING_REPORT_SAMPLE_COLLECTIONS
                if collection.sample_id != "accounting_register_records"
            )
            results = self._onec_exporter(
                credentials.onec_settings,
                accounting_collections,
                output_dir,
                top=max(self._onec_page_size(), 5000),
                # Period-local accounting collections can start well after the
                # generic refresh cap. They retain every fetched GET page as
                # raw evidence but materialize only the requested period.
                max_pages=max(self._onec_max_pages(), 1000),
                period_start=evidence_start,
                period_end=generation.period_end,
                source_identity=hashlib.sha256(
                    credentials.onec_settings.base_url.encode("utf-8")
                ).hexdigest(),
            )
            if generation.target_report_kind == MONTH_CLOSE_CONTROL:
                balance_result = self._onec_accounting_balance_exporter(
                    credentials.onec_settings,
                    output_dir,
                    period_start=generation.period_start,
                    period_end=generation.period_end,
                )
                results = [
                    *results,
                    balance_result,
                ]
                if not balance_result.ok or balance_result.row_count == 0:
                    results = [
                        *results,
                        self._onec_accounting_recordtype_exporter(
                            credentials.onec_settings,
                            output_dir,
                            period_start=generation.period_start,
                            period_end=generation.period_end,
                            page_size=max(
                                self._onec_page_size(),
                                self.settings.accounting_recordtype_page_size,
                            ),
                            max_pages=max(self._onec_max_pages(), 1000),
                        ),
                    ]
            self._record_onec(db, generation, output_dir, results)
            repository.sync_organization_tax_profiles(db, generation, user=user)
            repository.validate_source_snapshot_duplicates(db, generation)
            generation.generation_stage = "materializing_evidence"
            generation.heartbeat_at = security.utcnow()
            _commit_source_refresh_progress(db)

            sources = self._accounting_evidence_sources(db, generation, root_dir)
            if not any(
                source.status in {"loaded", "ready", "complete", "partial_source"}
                for source in sources.values()
            ):
                raise LookupError("accounting evidence sources are unavailable")
            evidence = materialize_accounting_evidence(
                report_kind=generation.target_report_kind,
                organization_id=str(generation.organization_id or ""),
                period_start=generation.period_start,
                period_end=generation.period_end,
                refresh_run_id=generation.id,
                sources=sources,
            )
            if generation.target_report_kind == MONTH_CLOSE_CONTROL and not (
                evidence.get("osvBalanceAndTurnovers", {}).get("rows")
                or evidence.get("osvRecordTypeFallback", {}).get("rows")
            ):
                raise LookupError("accounting osv evidence is unavailable")
            evidence_source_type = f"{generation.target_report_kind}_evidence"
            repository.add_source_refresh_collection(
                db,
                generation,
                source_type=evidence_source_type,
                source_label="Normalized accounting evidence v2",
                required=True,
                status="loaded",
                snapshot_hash=str(evidence["evidenceSha256"]),
                row_count=1,
                payload={
                    "contractVersion": evidence["contractVersion"],
                    "organizationId": generation.organization_id,
                    "payloadSha256": evidence["evidenceSha256"],
                    "sourceSnapshotIds": sorted(
                        {
                            source.snapshot_id
                            for source in sources.values()
                            if source.snapshot_id
                        }
                    ),
                    "normalizedEvidence": evidence,
                },
                organization_id=generation.organization_id,
            )
            generation.generation_stage = "building_report"
            generation.heartbeat_at = security.utcnow()
            _commit_source_refresh_progress(db)
            repository.complete_accounting_report_generation(
                db,
                generation=generation,
                user=user,
            )
            _commit_source_refresh_progress(db)
            return repository.generation_run_payload(generation)
        except Exception as exc:
            safe_error = _safe_error(exc)
            with suppress(Exception):
                db.rollback()
            generation = db.get(SourceRefreshRun, generation.id) or generation
            repository.update_source_refresh_run(
                db,
                generation,
                status="failed",
                failure_code=(
                    "accounting_evidence_missing"
                    if isinstance(exc, LookupError)
                    else "report_generation_failed"
                ),
                error_message=safe_error,
                finished_at=security.utcnow(),
            )
            generation.generation_stage = "failed"
            repository.audit(
                db,
                action="report_generation_failed",
                user=user,
                tenant_id=generation.tenant_id,
                entity_type="source_refresh_run",
                entity_id=generation.id,
                payload={
                    "reportKind": generation.target_report_kind,
                    "organizationId": generation.organization_id,
                    "errorType": exc.__class__.__name__,
                },
            )
            _commit_source_refresh_progress(db)
            return repository.generation_run_payload(generation)

    def _accounting_evidence_sources(
        self,
        db: Session,
        generation: SourceRefreshRun,
        root_dir: Path,
    ) -> dict[str, AccountingEvidenceSource]:
        result: dict[str, AccountingEvidenceSource] = {}
        collections = list(
            db.scalars(
                select(SourceRefreshCollection).where(
                    SourceRefreshCollection.refresh_run_id == generation.id,
                    SourceRefreshCollection.source_type.like("onec_%"),
                    SourceRefreshCollection.source_type.notin_(
                        {"onec_odata_metadata", "onec_tax_profiles"}
                    ),
                )
            )
        )
        for collection in collections:
            rows = [
                row.row_payload or {}
                for row in db.scalars(
                    select(SourceSnapshotRow)
                    .where(
                        SourceSnapshotRow.refresh_run_id == generation.id,
                        SourceSnapshotRow.collection_id == collection.id,
                    )
                    .order_by(SourceSnapshotRow.row_number)
                )
            ]
            if not rows:
                rows = self._read_onec_collection_rows(
                    collection.raw_path,
                    allowed_root=root_dir,
                )
            result[collection.source_type] = AccountingEvidenceSource(
                source_type=collection.source_type,
                status=collection.status,
                snapshot_id=collection.snapshot_hash or str(collection.id),
                rows=tuple(row for row in rows if isinstance(row, Mapping)),
            )
        return result

    @staticmethod
    def _read_onec_collection_rows(
        raw_path: str,
        *,
        allowed_root: Path,
    ) -> list[dict[str, Any]]:
        if not raw_path:
            return []
        path = Path(raw_path).resolve()
        root = allowed_root.resolve()
        if not path.is_relative_to(root) or not path.is_file():
            return []
        try:
            if path.stat().st_size > 100 * 1024 * 1024:
                return []
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return []
        if not isinstance(payload, dict):
            return []
        return [row for row in _extract_onec_rows(payload) if isinstance(row, dict)]

    def _create_refresh_run(
        self,
        db: Session,
        *,
        tenant_id: str,
        client_id: str | None,
        mode: str,
        credential_source: str,
        dry_run: bool,
        user: User | None,
        source_report: ReportRun | None,
        reason: str,
        period_start: date | None,
        period_end: date | None,
        resume_mode: str,
        resume_from_run_id: str | None,
    ) -> SourceRefreshRun | dict[str, Any]:
        mode = mode.strip()
        credential_source = credential_source.strip()
        resume_mode = resume_mode.strip().lower()
        if mode not in SOURCE_REFRESH_MODES:
            raise SourceRefreshConfigError(f"unsupported source refresh mode: {mode}")
        if credential_source not in CREDENTIAL_SOURCES:
            raise SourceRefreshConfigError(
                f"unsupported credential source: {credential_source}"
            )
        if resume_mode not in SOURCE_REFRESH_RESUME_MODES:
            raise SourceRefreshConfigError(
                f"unsupported source refresh resume mode: {resume_mode}"
            )
        if resume_mode == "never" and resume_from_run_id:
            raise SourceRefreshConfigError(
                "resume_from_run_id cannot be used with resume_mode=never"
            )
        if not self.settings.external_integrations_enabled and not dry_run:
            raise SourceRefreshDisabledError(
                "Внешние интеграции отключены для этого контура."
            )
        if not self.settings.source_refresh_enabled and not dry_run:
            raise SourceRefreshDisabledError(
                "Обновление источников выключено в настройках сервиса."
            )
        if db.get(Tenant, tenant_id) is None:
            raise SourceRefreshConfigError(f"tenant not found: {tenant_id}")
        if source_report is not None and source_report.tenant_id != tenant_id:
            raise PermissionError("source report tenant mismatch")

        resolved_client_id = client_id or (
            source_report.client_id
            if source_report
            else repository.client_id_for_tenant(tenant_id)
        )

        base_source_refresh_run: SourceRefreshRun | None = None
        source_window_start: date | None = None
        source_window_end: date | None = None
        if mode == "incremental":
            if not self.settings.source_refresh_incremental_enabled and not dry_run:
                raise SourceRefreshDisabledError(
                    "Инкрементальное обновление выключено feature flag."
                )
            if not self.settings.marketplace_daily_facts_enabled and not dry_run:
                raise SourceRefreshConfigError(
                    "incremental requires SHUMEYKO_MARKETPLACE_DAILY_FACTS_ENABLED=true"
                )
            if not self.settings.db_first_reports_enabled and not dry_run:
                raise SourceRefreshConfigError(
                    "incremental requires SHUMEYKO_DB_FIRST_REPORTS_ENABLED=true"
                )
            source_report = db.scalar(
                select(ReportRun)
                .where(
                    ReportRun.tenant_id == tenant_id,
                    ReportRun.client_id == resolved_client_id,
                    ReportRun.publication_status == "published",
                    ReportRun.is_current.is_(True),
                )
                .order_by(ReportRun.generated_at.desc())
            )
            if source_report is None:
                raise SourceRefreshConfigError(
                    "incremental requires the current published report"
                )
            # The composite report always extends the current published report
            # through yesterday. Request dates must not silently change its base.
            period_start = source_report.period_start
            period_end = _incremental_yesterday()
            source_window_end = period_end
            source_window_start = max(
                period_start,
                period_end
                - timedelta(
                    days=max(
                        1,
                        self.settings.source_refresh_incremental_window_days,
                    )
                    - 1
                ),
            )
            base_source_refresh_run = self.find_incremental_base_refresh(
                db,
                source_report,
                source_window_start=source_window_start,
                credential_source=credential_source,
            )
        if source_report is not None and mode == "onec-only":
            period_start = period_start or source_report.period_start
            period_end = period_end or source_report.period_end
            base_source_refresh_run = self.find_reusable_full_wb_refresh(
                db,
                source_report,
                credential_source=credential_source,
            )
            if base_source_refresh_run is None:
                # A report may only be rebuilt from a complete immutable WB base.
                # If it no longer exists, fetch a new full read-only snapshot.
                mode = "full"

        default_period_start, default_period_end = self.default_period_for_mode(mode)
        period_start = period_start or default_period_start
        period_end = period_end or default_period_end
        source_window_start = source_window_start or period_start
        source_window_end = source_window_end or period_end
        if period_start > period_end:
            raise SourceRefreshConfigError(
                "source refresh period_start must not be after period_end"
            )
        snapshot_set_id = self._snapshot_set_id(mode)
        resumed_from_run = self._resolve_resume_run(
            db,
            tenant_id=tenant_id,
            client_id=resolved_client_id,
            mode=mode,
            credential_source=credential_source,
            period_start=period_start,
            period_end=period_end,
            resume_mode=resume_mode,
            resume_from_run_id=resume_from_run_id,
            dry_run=dry_run,
        )
        if not dry_run:
            conflict = repository.active_conflicting_source_refresh_run(
                db,
                tenant_id=tenant_id,
                mode=mode,
                client_id=resolved_client_id,
            )
            if conflict is not None:
                return self._create_blocked_run(
                    db,
                    tenant_id=tenant_id,
                    client_id=resolved_client_id,
                    mode=mode,
                    credential_source=credential_source,
                    dry_run=dry_run,
                    snapshot_set_id=snapshot_set_id,
                    period_start=period_start,
                    period_end=period_end,
                    source_window_start=source_window_start,
                    source_window_end=source_window_end,
                    user=user,
                    source_report=source_report,
                    base_source_refresh_run=base_source_refresh_run,
                    blocked_by_run=conflict,
                    reason=reason,
                    status="blocked_active_refresh",
                    error_message=(
                        "Conflicting source refresh is active: "
                        f"{conflict.id} ({conflict.mode})."
                    ),
                )
        try:
            refresh_run = repository.create_source_refresh_run(
                db,
                tenant_id=tenant_id,
                mode=mode,
                credential_source=credential_source,
                dry_run=dry_run,
                snapshot_set_id=snapshot_set_id,
                period_start=period_start,
                period_end=period_end,
                source_window_start=source_window_start,
                source_window_end=source_window_end,
                client_id=resolved_client_id,
                user=user,
                source_report=source_report,
                resumed_from_run=resumed_from_run,
                base_source_refresh_run=base_source_refresh_run,
                reason=reason,
            )
        except ValueError as exc:
            raise SourceRefreshBusyError(str(exc)) from exc
        db.flush()
        return refresh_run

    def find_reusable_full_wb_refresh(
        self,
        db: Session,
        source_report: ReportRun,
        *,
        credential_source: str = "tenant",
    ) -> SourceRefreshRun | None:
        candidates = list(
            db.scalars(
                select(SourceRefreshRun)
                .where(
                    SourceRefreshRun.tenant_id == source_report.tenant_id,
                    SourceRefreshRun.client_id == source_report.client_id,
                    SourceRefreshRun.mode == "full",
                    SourceRefreshRun.credential_source == credential_source,
                    SourceRefreshRun.period_start <= source_report.period_start,
                    SourceRefreshRun.period_end >= source_report.period_end,
                    SourceRefreshRun.finished_at.is_not(None),
                    SourceRefreshRun.status.in_(
                        {"report_created", "needs_review", "source_loaded"}
                    ),
                )
                .order_by(SourceRefreshRun.created_at.desc())
            )
        )
        for candidate in candidates:
            if self._full_wb_refresh_is_reusable(db, candidate, source_report):
                return candidate
        return None

    def find_incremental_base_refresh(
        self,
        db: Session,
        source_report: ReportRun,
        *,
        source_window_start: date,
        credential_source: str = "tenant",
    ) -> SourceRefreshRun | None:
        required_coverage_end = source_window_start - timedelta(days=1)
        candidates = list(
            db.scalars(
                select(SourceRefreshRun)
                .where(
                    SourceRefreshRun.tenant_id == source_report.tenant_id,
                    SourceRefreshRun.client_id == source_report.client_id,
                    SourceRefreshRun.mode == "full",
                    SourceRefreshRun.credential_source == credential_source,
                    SourceRefreshRun.period_start <= source_report.period_start,
                    SourceRefreshRun.period_end
                    >= min(required_coverage_end, source_report.period_end),
                    SourceRefreshRun.finished_at.is_not(None),
                    SourceRefreshRun.status.in_(
                        {"report_created", "needs_review", "source_loaded"}
                    ),
                )
                .order_by(SourceRefreshRun.created_at.desc())
            )
        )
        allowed_root = self.settings.source_refresh_root_path.resolve()
        for candidate in candidates:
            if not candidate.root_dir:
                continue
            candidate_root = Path(candidate.root_dir).resolve()
            if (
                not candidate_root.is_relative_to(allowed_root)
                or not candidate_root.is_dir()
            ):
                continue
            finance = next(
                (
                    item
                    for item in candidate.collections
                    if item.source_type == "wb_finance_detail"
                ),
                None,
            )
            cards = next(
                (
                    item
                    for item in candidate.collections
                    if item.source_type == "wb_product_cards"
                ),
                None,
            )
            if finance is None or cards is None:
                continue
            daily_facts = dict((finance.payload or {}).get("dailyFacts") or {})
            if (
                finance.status not in MANDATORY_OK_STATUSES
                or cards.status not in MANDATORY_OK_STATUSES
                or daily_facts.get("status") != "materialized"
                or (daily_facts.get("parity") or {}).get("status") != "aggregate_only"
                or (daily_facts.get("persistedParity") or {}).get("status") != "matched"
            ):
                continue
            if not self._wb_collection_manifest_is_complete(finance):
                continue
            if (
                self._daily_facts_coverage_issue(
                    db,
                    tenant_id=source_report.tenant_id,
                    client_id=source_report.client_id,
                    period_start=source_report.period_start,
                    period_end=required_coverage_end,
                )
                is None
            ):
                return candidate
        return None

    def _daily_fact_contributing_runs(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
    ) -> list[SourceRefreshRun]:
        run_ids = list(
            db.scalars(
                select(MarketplaceFinanceDailyFactModel.source_refresh_run_id)
                .where(
                    MarketplaceFinanceDailyFactModel.tenant_id == refresh_run.tenant_id,
                    MarketplaceFinanceDailyFactModel.client_id == refresh_run.client_id,
                    MarketplaceFinanceDailyFactModel.marketplace == "wb",
                    MarketplaceFinanceDailyFactModel.fact_date
                    >= refresh_run.period_start,
                    MarketplaceFinanceDailyFactModel.fact_date
                    <= refresh_run.period_end,
                )
                .distinct()
            )
        )
        return [
            item
            for item in db.scalars(
                select(SourceRefreshRun).where(SourceRefreshRun.id.in_(run_ids))
            )
            if item is not None
        ]

    def _daily_facts_coverage_issue(
        self,
        db: Session,
        *,
        tenant_id: str,
        client_id: str,
        period_start: date,
        period_end: date,
    ) -> str | None:
        if period_end < period_start:
            return None
        fact_conditions = (
            MarketplaceFinanceDailyFactModel.tenant_id == tenant_id,
            MarketplaceFinanceDailyFactModel.client_id == client_id,
            MarketplaceFinanceDailyFactModel.marketplace == "wb",
            MarketplaceFinanceDailyFactModel.fact_date >= period_start,
            MarketplaceFinanceDailyFactModel.fact_date <= period_end,
        )
        if (
            db.scalar(
                select(MarketplaceFinanceDailyFactModel.id)
                .where(*fact_conditions)
                .limit(1)
            )
            is None
        ):
            return "daily_facts_empty"
        if (
            db.scalar(
                select(MarketplaceFinanceDailyFactModel.id)
                .where(
                    *fact_conditions,
                    MarketplaceFinanceDailyFactModel.is_partial_source.is_(True),
                )
                .limit(1)
            )
            is not None
        ):
            return "daily_facts_partial_source"
        run_ids = list(
            db.scalars(
                select(MarketplaceFinanceDailyFactModel.source_refresh_run_id)
                .where(*fact_conditions)
                .distinct()
            )
        )
        runs = list(
            db.scalars(select(SourceRefreshRun).where(SourceRefreshRun.id.in_(run_ids)))
        )
        intervals: list[tuple[date, date]] = []
        for item in runs:
            coverage_start = item.source_window_start or item.period_start
            coverage_end = item.source_window_end or item.period_end
            finance_collection = next(
                (
                    collection
                    for collection in item.collections
                    if collection.source_type == "wb_finance_detail"
                ),
                None,
            )
            finance_payload = (
                dict(finance_collection.payload or {})
                if finance_collection is not None
                else {}
            )
            try:
                coverage_start = date.fromisoformat(
                    str(finance_payload["sourceCoverageStart"])
                )
                coverage_end = date.fromisoformat(
                    str(finance_payload["sourceCoverageEnd"])
                )
            except (KeyError, TypeError, ValueError):
                pass
            intervals.append(
                (
                    max(period_start, coverage_start),
                    min(period_end, coverage_end),
                )
            )
        intervals.sort()
        cursor = period_start
        for start, end in intervals:
            if end < cursor:
                continue
            if start > cursor:
                return f"daily_facts_coverage_gap:{cursor.isoformat()}"
            cursor = max(cursor, end + timedelta(days=1))
            if cursor > period_end:
                return None
        return f"daily_facts_coverage_gap:{cursor.isoformat()}"

    def _full_wb_refresh_is_reusable(
        self,
        db: Session,
        candidate: SourceRefreshRun,
        source_report: ReportRun,
    ) -> bool:
        if not candidate.root_dir:
            return False
        root_dir = Path(candidate.root_dir).resolve()
        allowed_root = self.settings.source_refresh_root_path.resolve()
        if not root_dir.is_relative_to(allowed_root) or not root_dir.is_dir():
            return False
        collections = {
            item.source_type: item
            for item in candidate.collections
            if item.source_type.startswith("wb_")
        }
        finance = collections.get("wb_finance_detail")
        required_wb = [item for item in collections.values() if item.required]
        if finance is None or not required_wb:
            return False
        coverage = finance.payload or {}
        try:
            coverage_start = date.fromisoformat(str(coverage["sourceCoverageStart"]))
            coverage_end = date.fromisoformat(str(coverage["sourceCoverageEnd"]))
        except (KeyError, TypeError, ValueError):
            return False
        if (
            coverage_start > source_report.period_start
            or coverage_end < source_report.period_end
        ):
            return False
        if not all(
            item.status in MANDATORY_OK_STATUSES
            and self._collection_raw_dir(item) is not None
            and self._wb_collection_manifest_is_complete(item)
            for item in required_wb
        ):
            return False
        active_cabinet_ids = {
            item.id
            for item in db.scalars(
                select(WbCabinet).where(
                    WbCabinet.tenant_id == source_report.tenant_id,
                    WbCabinet.client_id == source_report.client_id,
                    WbCabinet.status == "active",
                    WbCabinet.provider == "wb_api",
                )
            )
        }
        snapshot_cabinet_ids = {
            str(item.get("wbCabinetId") or "").strip()
            for item in (finance.payload or {}).get("results", [])
            if isinstance(item, dict) and str(item.get("wbCabinetId") or "").strip()
        }
        return not active_cabinet_ids or active_cabinet_ids.issubset(
            snapshot_cabinet_ids
        )

    def _wb_collection_manifest_is_complete(
        self,
        collection: SourceRefreshCollection,
    ) -> bool:
        raw_dir = self._collection_raw_dir(collection)
        if raw_dir is None:
            return False
        manifest_path = raw_dir / "manifest.json"
        if (
            not manifest_path.is_file()
            or manifest_path.stat().st_size > 5 * 1024 * 1024
        ):
            return False
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, TypeError, ValueError, json.JSONDecodeError):
            return False
        results = manifest.get("results") if isinstance(manifest, dict) else None
        if not isinstance(results, list) or not results:
            return False
        account_results: dict[str, list[dict[str, Any]]] = {}
        for item in results:
            if not isinstance(item, dict):
                return False
            account_id = str(item.get("seller_account_id") or "").strip()
            if not account_id:
                return False
            account_results.setdefault(account_id, []).append(item)
            for key in ("output_file", "flat_output_file"):
                output_name = str(item.get(key) or "").strip()
                if output_name and (
                    Path(output_name).name != output_name
                    or not (raw_dir / output_name).is_file()
                ):
                    return False
        if collection.source_type == "wb_finance_detail":
            return all(
                str(items[-1].get("status") or "") == "no_data"
                for items in account_results.values()
            )
        return all(item.get("ok") is True for item in results)

    def _collection_raw_dir(
        self,
        collection: SourceRefreshCollection,
    ) -> Path | None:
        if not collection.raw_path:
            return None
        path = Path(collection.raw_path).resolve()
        allowed_root = self.settings.source_refresh_root_path.resolve()
        if not path.is_relative_to(allowed_root):
            return None
        if path.is_file():
            path = path.parent
        return path if path.is_dir() else None

    def _execute_run(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        user: User | None,
        stop_after_sources: bool = False,
    ) -> dict[str, Any]:
        tenant_id = refresh_run.tenant_id
        mode = refresh_run.mode
        credential_source = refresh_run.credential_source
        dry_run = refresh_run.dry_run
        period_start = refresh_run.period_start
        period_end = refresh_run.period_end
        source_window_start = refresh_run.source_window_start or period_start
        source_window_end = refresh_run.source_window_end or period_end
        source_report = (
            db.get(ReportRun, refresh_run.source_report_run_id)
            if refresh_run.source_report_run_id
            else None
        )
        base_refresh_run = (
            db.get(SourceRefreshRun, refresh_run.base_source_refresh_run_id)
            if refresh_run.base_source_refresh_run_id
            else None
        )
        root_dir = (
            self.settings.source_refresh_root_path / refresh_run.snapshot_set_id
        ).resolve()
        db.info["source_refresh_run_id"] = refresh_run.id
        try:
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="running",
                error_message="",
                failure_code="",
                started_at=refresh_run.started_at or security.utcnow(),
                heartbeat_at=security.utcnow(),
                root_dir=str(root_dir),
            )
            _commit_source_refresh_progress(db)
            if (
                self.settings.logistics_analysis_enabled
                and not self.settings.db_first_reports_enabled
            ):
                repository.update_source_refresh_run(
                    db,
                    refresh_run,
                    failure_code="logistics_requires_db_first",
                )
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status="needs_configuration",
                    error_message=(
                        "Logistics analysis requires "
                        "SHUMEYKO_DB_FIRST_REPORTS_ENABLED=true."
                    ),
                )
            if not dry_run:
                disk_issue = self._low_disk_issue()
                if disk_issue is not None:
                    return self._finish_without_report(
                        db,
                        refresh_run,
                        status="blocked_low_disk",
                        error_message=disk_issue["error_message"],
                    )
            if mode == "incremental" and (
                source_report is None or base_refresh_run is None
            ):
                repository.update_source_refresh_run(
                    db,
                    refresh_run,
                    failure_code="incremental_base_unavailable",
                )
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status="needs_full_refresh",
                    error_message=(
                        "Compatible full daily-facts base is unavailable; "
                        "run a full refresh."
                    ),
                )
            credentials = self._credentials(
                db,
                tenant_id=tenant_id,
                credential_source=credential_source,
                mode=mode,
            )
            for issue in (*credentials.issues, *credentials.optional_issues):
                repository.add_source_refresh_collection(
                    db,
                    refresh_run,
                    source_type=issue["source_type"],
                    source_label=issue["source_label"],
                    required=issue["required"],
                    status=issue["status"],
                    error_message=issue.get("error_message", ""),
                    payload=issue.get("payload", {}),
                )
            outputs = self._run_collectors(
                CollectorContext(
                    db=db,
                    refresh_run=refresh_run,
                    credentials=credentials,
                    root_dir=root_dir,
                    period_start=source_window_start,
                    period_end=source_window_end,
                    mode=mode,
                ),
                include_external=False,
            )
            mapping_collection = outputs.mapping_collection
            if credentials.issues:
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status="needs_configuration",
                    error_message=(
                        "Source credentials are not ready for scheduled refresh."
                    ),
                )
            if dry_run:
                if self._mandatory_failed(refresh_run):
                    return self._finish_without_report(
                        db,
                        refresh_run,
                        status="failed",
                        error_message="Mandatory source refresh collection failed.",
                    )
                if self._needs_review(refresh_run, mapping_collection):
                    return self._finish_without_report(
                        db,
                        refresh_run,
                        status="needs_review",
                    )
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status="dry_run_ready",
                )

            if credentials.onec_settings is not None:
                metadata_result = self._record_onec_metadata_check(
                    db,
                    refresh_run,
                    credentials.onec_settings,
                )
                _commit_source_refresh_progress(db)
                if not metadata_result.ok:
                    return self._finish_without_report(
                        db,
                        refresh_run,
                        status="failed",
                        failure_code="onec_odata_metadata_unavailable",
                        error_message=(
                            "onec_odata_metadata_unavailable: "
                            f"{metadata_result.error or 'unknown_error'}"
                        ),
                        retryable=_metadata_failure_is_transient(metadata_result),
                    )

            outputs = self._run_collectors(
                CollectorContext(
                    db=db,
                    refresh_run=refresh_run,
                    credentials=credentials,
                    root_dir=root_dir,
                    period_start=source_window_start,
                    period_end=source_window_end,
                    mode=mode,
                ),
                include_external=True,
            )
            if self._mandatory_failed(refresh_run):
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status="failed",
                    failure_code="mandatory_source_failed",
                    error_message="Mandatory source refresh collection failed.",
                    retryable=self._mandatory_failure_is_transient(refresh_run),
                )
            ozon_promotion_succeeded = self._promote_ozon_typed_facts(
                db,
                refresh_run,
                ozon_cabinet_ids=credentials.ozon_cabinet_ids,
            )
            if not ozon_promotion_succeeded and mode == "ozon-only":
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status="failed",
                    error_message=(
                        "Ozon typed-facts promotion failed; previous facts "
                        "were preserved."
                    ),
                )
            self._rebuild_mapping_service(db, refresh_run, mapping_collection, user)
            repository.sync_organization_tax_profiles(
                db,
                refresh_run,
                user=user,
            )
            repository.validate_source_snapshot_duplicates(db, refresh_run)
            _commit_source_refresh_progress(db)
            wb_finance_dir = outputs.output_dirs.get("wb_finance_detail")
            wb_report_list_dir = outputs.output_dirs.get("wb_sales_report_list")
            wb_cards_dir = outputs.output_dirs.get("wb_product_cards")
            wb_stock_history_dir = outputs.output_dirs.get("wb_stock_history_daily")
            onec_dir = outputs.output_dirs.get("onec_odata")
            composite_rebuild = bool(
                mode == "onec-only"
                and source_report is not None
                and base_refresh_run is not None
            )
            if composite_rebuild:
                wb_finance_dir = self._required_collection_raw_dir(
                    base_refresh_run,
                    "wb_finance_detail",
                )
                wb_cards_dir = self._required_collection_raw_dir(
                    base_refresh_run,
                    "wb_product_cards",
                )
                wb_report_list_dir = self._optional_collection_raw_dir(
                    base_refresh_run,
                    "wb_sales_report_list",
                )
                wb_stock_history_dir = self._optional_collection_raw_dir(
                    base_refresh_run,
                    "wb_stock_history_daily",
                )

            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="source_loaded",
            )
            _commit_source_refresh_progress(db)
            if stop_after_sources:
                return repository.source_refresh_run_payload(refresh_run)
            if (
                mode in {"daily", "incremental"}
                and self.settings.marketplace_daily_facts_enabled
                and wb_finance_dir is not None
                and onec_dir is not None
            ):
                self._materialize_wb_daily_facts(
                    db,
                    refresh_run,
                    wb_finance_dir=wb_finance_dir,
                    onec_dir=onec_dir,
                    wb_report_list_dir=wb_report_list_dir,
                    wb_cards_dir=wb_cards_dir,
                    wb_stock_history_dir=wb_stock_history_dir,
                )
            if mode == "incremental":
                coverage_issue = self._daily_facts_coverage_issue(
                    db,
                    tenant_id=refresh_run.tenant_id,
                    client_id=refresh_run.client_id,
                    period_start=refresh_run.period_start,
                    period_end=refresh_run.period_end,
                )
                if coverage_issue:
                    repository.update_source_refresh_run(
                        db,
                        refresh_run,
                        failure_code="incremental_daily_facts_coverage_gap",
                    )
                    return self._finish_without_report(
                        db,
                        refresh_run,
                        status="needs_full_refresh",
                        error_message=coverage_issue,
                    )
            if mode == "ozon-only":
                status = (
                    "needs_review"
                    if self._needs_review(refresh_run, mapping_collection)
                    else "source_loaded"
                )
                self._finish_without_report(
                    db,
                    refresh_run,
                    status=status,
                )
                source_blocker = repository.ozon_draft_source_blocker(
                    db,
                    refresh_run,
                )
                if source_blocker:
                    repository.audit(
                        db,
                        action="ozon_draft_report_skipped",
                        user=user,
                        tenant_id=refresh_run.tenant_id,
                        entity_type="source_refresh_run",
                        entity_id=refresh_run.id,
                        payload={"reason": source_blocker},
                    )
                    _commit_source_refresh_progress(db)
                    return repository.source_refresh_run_payload(refresh_run)
                repository.materialize_ozon_draft_report(
                    db,
                    refresh_run,
                    user=user,
                )
                return repository.source_refresh_run_payload(refresh_run)
            if mode == "daily" or (mode == "onec-only" and not composite_rebuild):
                status = (
                    "needs_review"
                    if self._needs_review(refresh_run, mapping_collection)
                    else "source_loaded"
                )
                return self._finish_without_report(
                    db,
                    refresh_run,
                    status=status,
                )

            repository.update_source_refresh_run(db, refresh_run, status="rebuilding")
            _commit_source_refresh_progress(db)
            contributing_runs: list[SourceRefreshRun] = []
            wb_daily_facts: list[MarketplaceFinanceDailyFactContract] | None = None
            wb_summary_rows: list[WbSalesReportSummaryRow] | None = None
            if mode == "incremental":
                contributing_runs = self._daily_fact_contributing_runs(db, refresh_run)
                wb_summary_rows = self._incremental_wb_summary_rows(
                    refresh_run,
                    base_refresh_run=base_refresh_run,
                    current_report_list_dir=wb_report_list_dir,
                )
                wb_daily_facts = self._daily_facts_for_report(
                    db,
                    refresh_run,
                    wb_summary_rows=wb_summary_rows,
                )
            report_snapshot_set_id = self._report_snapshot_set_id(
                refresh_run,
                base_refresh_run=(
                    base_refresh_run
                    if composite_rebuild or mode == "incremental"
                    else None
                ),
                contributing_runs=contributing_runs,
            )
            if self.settings.db_first_reports_enabled:
                new_report, workbook_path = self._build_db_first_report(
                    db,
                    refresh_run,
                    source_report=source_report,
                    wb_finance_dir=wb_finance_dir,
                    onec_dir=onec_dir,
                    wb_report_list_dir=wb_report_list_dir,
                    wb_cards_dir=wb_cards_dir,
                    wb_stock_history_dir=wb_stock_history_dir,
                    source_snapshot_set_id=report_snapshot_set_id,
                    base_refresh_run=(
                        base_refresh_run
                        if composite_rebuild or mode == "incremental"
                        else None
                    ),
                    contributing_runs=contributing_runs,
                    wb_daily_facts=wb_daily_facts,
                    wb_summary_rows=wb_summary_rows,
                )
            else:
                workbook_path = self._build_workbook(
                    db,
                    refresh_run,
                    wb_finance_dir=wb_finance_dir,
                    onec_dir=onec_dir,
                    wb_report_list_dir=wb_report_list_dir,
                    wb_stock_history_dir=wb_stock_history_dir,
                    raw_refresh_run=(
                        base_refresh_run if composite_rebuild else refresh_run
                    ),
                )
                new_report = repository.import_dashboard_payload(
                    db,
                    self._dashboard_payload_builder(workbook_path),
                    tenant_id=tenant_id,
                    tenant_name=self._tenant_name(db, tenant_id),
                    report_id=self._new_report_id(source_report, refresh_run),
                    source_workbook_path=str(workbook_path),
                    lineage_type="legacy_excel_import",
                    publication_status="draft",
                    publish=False,
                    source_snapshot_set_id=report_snapshot_set_id,
                )
            primary_document_refresh = (
                base_refresh_run
                if composite_rebuild and base_refresh_run is not None
                else refresh_run
            )
            primary_document_scope = repository.apply_wb_buyout_primary_documents(
                db,
                new_report,
                primary_document_refresh,
                source_runs=(contributing_runs if mode == "incremental" else ()),
            )
            _commit_source_refresh_progress(db)
            self._attach_source_loads(
                db,
                new_report,
                refresh_run,
                contributing_runs=contributing_runs,
            )
            mapping_report_scope = repository.reconcile_report_mapping_source_load(
                db, new_report
            )
            logistics_context = db.get(
                ReportLogisticsAnalysisContext,
                new_report.id,
            )
            logistics_needs_review = bool(
                new_report.logistics_analysis_required
                and (
                    logistics_context is None
                    or logistics_context.data_status != "ready"
                )
            )
            final_status = (
                "needs_review"
                if self._needs_review(
                    refresh_run,
                    mapping_collection,
                    mapping_report_ready=(
                        mapping_report_scope["mappingIssueRows"] == 0
                    ),
                )
                or logistics_needs_review
                else "report_created"
            )
            refresh_run.new_report_run_id = new_report.id
            repository.link_source_refresh_tasks_to_report(db, refresh_run, new_report)
            db.commit()
            if self.settings.db_first_reports_enabled:
                self._export_db_first_report_excel(
                    db,
                    refresh_run,
                    new_report,
                    excel_path=workbook_path,
                )
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status=final_status,
                new_report_run_id=new_report.id,
                workbook_path=str(workbook_path),
                finished_at=security.utcnow(),
            )
            repository.audit(
                db,
                action="source_refresh_report_created",
                user=user,
                tenant_id=tenant_id,
                entity_type="source_refresh_run",
                entity_id=refresh_run.id,
                payload={
                    "mode": mode,
                    "new_report_run_id": new_report.id,
                    "status": final_status,
                    "snapshotSetId": report_snapshot_set_id,
                    "baseSourceRefreshRunId": (
                        base_refresh_run.id if composite_rebuild else None
                    ),
                    "mappingReportScope": mapping_report_scope,
                    "buyoutPrimaryDocumentScope": primary_document_scope,
                },
            )
            publication_blockers = repository.report_publication_blockers(
                db, new_report
            )
            if publication_blockers:
                repository.update_source_refresh_run(
                    db,
                    refresh_run,
                    status="needs_review",
                    finished_at=security.utcnow(),
                )
                repository.audit(
                    db,
                    action="source_refresh_report_kept_draft",
                    user=user,
                    tenant_id=tenant_id,
                    entity_type="source_refresh_run",
                    entity_id=refresh_run.id,
                    payload={
                        "new_report_run_id": new_report.id,
                        "blockers": publication_blockers,
                    },
                )
                payload = repository.source_refresh_run_payload(refresh_run)
                _commit_source_refresh_progress(db)
                return payload
            repository.audit(
                db,
                action="source_refresh_report_kept_draft",
                user=user,
                tenant_id=tenant_id,
                entity_type="source_refresh_run",
                entity_id=refresh_run.id,
                payload={
                    "new_report_run_id": new_report.id,
                    "blockers": [],
                    "reason": "awaiting_staff_financial_acceptance",
                },
            )
            payload = repository.source_refresh_run_payload(refresh_run)
            _commit_source_refresh_progress(db)
            return payload
        except Exception as exc:
            safe_error = _safe_error(exc)
            with suppress(Exception):
                db.rollback()
            refresh_run = db.get(SourceRefreshRun, refresh_run.id) or refresh_run
            if (
                db.info.get("source_refresh_split_pipeline")
                and _source_refresh_exception_is_transient(exc)
                and repository.requeue_transient_source_refresh_task(
                    db,
                    refresh_run,
                    task_type="collect_sources",
                    safe_error_code="collect_sources_transport_failed",
                    safe_error_message=safe_error,
                )
            ):
                repository.audit(
                    db,
                    action="source_refresh_task_retry_scheduled",
                    user=user,
                    tenant_id=tenant_id,
                    entity_type="source_refresh_run",
                    entity_id=refresh_run.id,
                    payload={
                        "taskType": "collect_sources",
                        "errorType": exc.__class__.__name__,
                    },
                )
                _commit_source_refresh_progress(db)
                return repository.source_refresh_run_payload(refresh_run)
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="failed",
                failure_code="collect_sources_failed",
                error_message=safe_error,
                finished_at=security.utcnow(),
            )
            repository.audit(
                db,
                action="source_refresh_failed",
                user=user,
                tenant_id=tenant_id,
                entity_type="source_refresh_run",
                entity_id=refresh_run.id,
                payload={
                    "mode": mode,
                    "errorType": exc.__class__.__name__,
                },
            )
            _commit_source_refresh_progress(db)
            with suppress(Exception):
                self._prune_failed_snapshot_directories(db, refresh_run)
            return repository.source_refresh_run_payload(refresh_run)

    def _create_blocked_run(
        self,
        db: Session,
        *,
        tenant_id: str,
        client_id: str | None,
        mode: str,
        credential_source: str,
        dry_run: bool,
        snapshot_set_id: str,
        period_start: date,
        period_end: date,
        source_window_start: date,
        source_window_end: date,
        user: User | None,
        source_report: ReportRun | None,
        base_source_refresh_run: SourceRefreshRun | None,
        blocked_by_run: SourceRefreshRun | None,
        reason: str,
        status: str,
        error_message: str,
    ) -> dict[str, Any]:
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id=tenant_id,
            mode=mode,
            credential_source=credential_source,
            dry_run=dry_run,
            snapshot_set_id=snapshot_set_id,
            period_start=period_start,
            period_end=period_end,
            source_window_start=source_window_start,
            source_window_end=source_window_end,
            client_id=client_id,
            user=user,
            source_report=source_report,
            base_source_refresh_run=base_source_refresh_run,
            blocked_by_run=blocked_by_run,
            reason=reason,
            enforce_active_check=False,
        )
        return self._finish_without_report(
            db,
            refresh_run,
            status=status,
            error_message=error_message,
        )

    def _run_collectors(
        self,
        context: CollectorContext,
        *,
        include_external: bool,
    ) -> CollectorOutputs:
        output_dirs: dict[str, Path] = {}
        mapping_collection: SourceRefreshCollection | None = None
        for collector in self._collector_plan(context.mode):
            if include_external == (collector.source_type == "sku_mapping"):
                continue
            collector_context = context
            if context.mode == "incremental" and collector.source_type in {
                "onec_odata",
                "wb_stock_history_daily",
            }:
                collector_context = CollectorContext(
                    db=context.db,
                    refresh_run=context.refresh_run,
                    credentials=context.credentials,
                    root_dir=context.root_dir,
                    period_start=context.refresh_run.period_start,
                    period_end=context.refresh_run.period_end,
                    mode=context.mode,
                )
            result = collector.collect(self, collector_context)
            if result.output_dir is not None:
                output_dirs[collector.source_type] = result.output_dir
            if collector.source_type == "sku_mapping":
                mapping_collection = result.collection
            _commit_source_refresh_progress(context.db)
        if mapping_collection is None and not include_external:
            raise RuntimeError("mapping collector did not create a collection")
        return CollectorOutputs(
            output_dirs=output_dirs,
            mapping_collection=mapping_collection or context.refresh_run.collections[0],
        )

    def _collector_plan(self, mode: str) -> tuple[SourceCollector, ...]:
        collectors = (
            SourceCollector(
                source_type="sku_mapping",
                label="WB ↔ 1C mapping",
                required=True,
                modes=frozenset(SOURCE_REFRESH_MODES),
                roles=frozenset(),
                collect=_collect_mapping_source,
            ),
            SourceCollector(
                source_type="wb_finance_detail",
                label="WB Finance sales report details",
                required=True,
                modes=frozenset({"daily", "incremental", "weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_finance,
            ),
            SourceCollector(
                source_type="wb_sales_report_list",
                label="WB Finance sales report list",
                required=False,
                modes=frozenset({"incremental", "weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_report_list,
            ),
            SourceCollector(
                source_type="wb_redeem_notifications",
                label="WB primary redeem notifications",
                required=False,
                modes=frozenset({"incremental", "weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_redeem_notifications,
            ),
            SourceCollector(
                source_type="wb_product_cards",
                label="WB product cards",
                required=True,
                modes=frozenset({"daily", "incremental", "weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_product_cards,
            ),
            SourceCollector(
                source_type="wb_stock_history_daily",
                label="WB daily stock history",
                required=False,
                modes=frozenset({"incremental", "weekly", "full"}),
                roles=frozenset(WB_STOCK_HISTORY_REFRESH_ROLES),
                collect=_collect_wb_stock_history,
            ),
            SourceCollector(
                source_type="wb_tariffs",
                label="WB box/pallet tariffs",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_tariffs,
            ),
            SourceCollector(
                source_type="wb_goods_return",
                label="WB goods return reasons",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_goods_return,
            ),
            SourceCollector(
                source_type="wb_return_claims",
                label="WB buyer return claims",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_return_claims,
            ),
            SourceCollector(
                source_type="wb_supplier_sales",
                label="WB supplier sales (warehouse & direction)",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_supplier_sales,
            ),
            SourceCollector(
                source_type="wb_measurement_penalties",
                label="WB dimension measurement retentions",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_measurement_penalties,
            ),
            SourceCollector(
                source_type="wb_warehouse_measurements",
                label="WB warehouse measurements",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset(WB_FINANCE_REFRESH_ROLES),
                collect=_collect_wb_warehouse_measurements,
            ),
            SourceCollector(
                source_type="ozon_finance_cash_flow",
                label="Ozon financial cash-flow statement",
                required=False,
                modes=frozenset({"daily", "weekly", "full", "ozon-only"}),
                roles=frozenset({"finance_reports", "full_readonly"}),
                collect=_collect_ozon_cash_flow,
            ),
            SourceCollector(
                source_type="ozon_realization",
                label="Ozon realization report",
                required=False,
                modes=frozenset({"weekly", "full", "ozon-only"}),
                roles=frozenset({"finance_reports", "full_readonly"}),
                collect=_collect_ozon_realization,
            ),
            SourceCollector(
                source_type="ozon_mutual_settlement",
                label="Ozon mutual settlement report",
                required=False,
                modes=frozenset({"weekly", "full", "ozon-only"}),
                roles=frozenset({"finance_reports", "full_readonly"}),
                collect=_collect_ozon_mutual_settlement,
            ),
            SourceCollector(
                source_type="ozon_realization_posting",
                label="Ozon realization posting report",
                required=False,
                modes=frozenset({"weekly", "full", "ozon-only"}),
                roles=frozenset({"finance_reports", "full_readonly"}),
                collect=_collect_ozon_realization_posting,
            ),
            SourceCollector(
                source_type="ozon_products_buyout",
                label="Ozon products buyout report",
                required=False,
                modes=frozenset({"weekly", "full", "ozon-only"}),
                roles=frozenset({"finance_reports", "full_readonly"}),
                collect=_collect_ozon_products_buyout,
            ),
            SourceCollector(
                source_type="ozon_b2b_sales_json",
                label="Ozon B2B sales JSON",
                required=False,
                modes=frozenset({"weekly", "full", "ozon-only"}),
                roles=frozenset({"finance_reports", "full_readonly"}),
                collect=_collect_ozon_b2b_sales,
            ),
            SourceCollector(
                source_type="ozon_products_report",
                label="Ozon products report",
                required=False,
                modes=frozenset({"weekly", "full", "ozon-only"}),
                roles=frozenset({"products_catalog", "full_readonly"}),
                collect=_collect_ozon_products,
            ),
            SourceCollector(
                source_type="ozon_stock_on_warehouses",
                label="Ozon stock on warehouses",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset({"stocks_analytics", "full_readonly"}),
                collect=_collect_ozon_stocks,
            ),
            SourceCollector(
                source_type="ozon_returns_report",
                label="Ozon returns report",
                required=False,
                modes=frozenset({"weekly", "full"}),
                roles=frozenset({"returns_reports", "full_readonly"}),
                collect=_collect_ozon_returns,
            ),
            SourceCollector(
                source_type="onec_odata",
                label="1С OData collections",
                required=True,
                modes=frozenset(SOURCE_REFRESH_MODES),
                roles=frozenset(
                    {"cost_documents", "stocks_warehouses", "full_readonly"}
                ),
                collect=_collect_onec_odata,
            ),
        )
        return tuple(item for item in collectors if item.supports(mode))

    def _low_disk_issue(self) -> dict[str, Any] | None:
        min_free_gb = max(0.0, float(self.settings.source_refresh_min_free_gb))
        if min_free_gb <= 0:
            return None
        probe_path = _existing_path_for_disk_check(
            self.settings.source_refresh_root_path
        )
        usage = shutil.disk_usage(probe_path)
        free_gb = usage.free / (1024**3)
        if free_gb >= min_free_gb:
            return None
        return {
            "error_message": (
                "source_refresh_low_disk: "
                f"free={free_gb:.2f}GiB required={min_free_gb:.2f}GiB"
            ),
            "free_gb": free_gb,
            "required_gb": min_free_gb,
        }

    def _credentials(
        self,
        db: Session,
        *,
        tenant_id: str,
        credential_source: str,
        mode: str,
    ) -> SourceCredentials:
        if credential_source == "env":
            return self._env_credentials(mode)
        return self._tenant_credentials(db, tenant_id=tenant_id, mode=mode)

    def _env_credentials(self, mode: str) -> SourceCredentials:
        issues: list[dict[str, Any]] = []
        optional_issues: list[dict[str, Any]] = []
        wb_settings = None
        ozon_settings = None
        onec_settings = None
        if mode in WB_REQUIRED_MODES:
            try:
                wb_settings = WbFinanceSettings.from_env_file()
            except WbFinanceConfigError as exc:
                issues.append(_credential_issue("wb_api", True, str(exc)))
        if mode in OZON_REQUIRED_MODES:
            try:
                ozon_settings = OzonSettings.from_env_file()
            except OzonConfigError as exc:
                issues.append(_credential_issue("ozon_api", True, str(exc)))
        elif mode in OZON_OPTIONAL_MODES and _env_has_ozon_credentials():
            try:
                ozon_settings = OzonSettings.from_env_file()
            except OzonConfigError as exc:
                optional_issues.append(_credential_issue("ozon_api", False, str(exc)))
        try:
            onec_settings = OnecODataSettings.from_env_file()
        except OnecODataConfigError as exc:
            issues.append(_credential_issue("onec_readonly", True, str(exc)))
        return SourceCredentials(
            wb_settings,
            onec_settings,
            ozon_settings,
            {},
            {},
            tuple(issues),
            tuple(optional_issues),
        )

    def _tenant_credentials(
        self,
        db: Session,
        *,
        tenant_id: str,
        mode: str,
    ) -> SourceCredentials:
        issues: list[dict[str, Any]] = []
        optional_issues: list[dict[str, Any]] = []
        wb_settings = None
        wb_cabinet_ids: dict[str, str] = {}
        ozon_settings = None
        ozon_cabinet_ids: dict[str, str] = {}
        onec_settings = None
        if mode in WB_REQUIRED_MODES:
            wb_integrations = _tenant_integrations_by_base(db, tenant_id, "wb_api")
            if not wb_integrations:
                issues.append(_credential_issue("wb_api", True, "not_configured"))
            else:
                wb_settings, wb_cabinet_ids = self._wb_settings_from_integrations(
                    db,
                    wb_integrations,
                    issues,
                )
        if mode in OZON_REQUIRED_MODES | OZON_OPTIONAL_MODES:
            ozon_integrations = _tenant_integrations_by_base(db, tenant_id, "ozon_api")
            ozon_required = mode in OZON_REQUIRED_MODES
            if not ozon_integrations and ozon_required:
                issues.append(_credential_issue("ozon_api", True, "not_configured"))
            elif ozon_integrations:
                ozon_settings, ozon_cabinet_ids = self._ozon_settings_from_integrations(
                    db,
                    ozon_integrations,
                    issues if ozon_required else optional_issues,
                    required=ozon_required,
                )
        integration = _tenant_integration(db, tenant_id, "onec_readonly")
        if integration is None:
            issues.append(_credential_issue("onec_readonly", True, "not_configured"))
        else:
            onec_settings = self._onec_settings_from_integration(integration, issues)
        return SourceCredentials(
            wb_settings,
            onec_settings,
            ozon_settings,
            wb_cabinet_ids,
            ozon_cabinet_ids,
            tuple(issues),
            tuple(optional_issues),
        )

    def _wb_settings_from_integrations(
        self,
        db: Session,
        integrations_list: list[TenantIntegration],
        issues: list[dict[str, Any]],
    ) -> tuple[WbFinanceSettings | None, dict[str, str]]:
        accounts: list[WbFinanceSellerAccount] = []
        wb_cabinet_ids: dict[str, str] = {}
        skipped_roles: list[str] = []
        finance_index = 0
        for integration in integrations_list:
            if integration.status == "disabled":
                continue
            role = str(
                (integration.config_payload or {}).get("connectionRole") or ""
            ).strip()
            if role and role not in WB_FINANCE_REFRESH_ROLES:
                skipped_roles.append(integration.provider)
                continue
            finance_index += 1
            settings = self._wb_settings_from_integration(
                integration,
                issues,
                default_account_index=finance_index,
            )
            if settings is not None:
                accounts.extend(settings.accounts)
                cabinet = _ensure_wb_cabinet_for_integration(db, integration)
                if cabinet is not None:
                    for account in settings.accounts:
                        wb_cabinet_ids[account.seller_account_id] = cabinet.id
        if accounts:
            return WbFinanceSettings(accounts=tuple(accounts)), wb_cabinet_ids
        if not any(item["source_type"].startswith("wb_api") for item in issues):
            issues.append(
                _credential_issue(
                    "wb_api",
                    True,
                    "no_finance_report_integrations",
                    payload={"skippedProviders": skipped_roles},
                )
            )
        return None, wb_cabinet_ids

    def _wb_settings_from_integration(
        self,
        integration: TenantIntegration,
        issues: list[dict[str, Any]],
        *,
        default_account_index: int = 1,
    ) -> WbFinanceSettings | None:
        secret = self._encrypted_secret_or_issue(integration, issues)
        if not secret:
            return None
        try:
            return wb_finance_settings_from_secret(
                secret,
                default_name=_integration_account_name(integration),
                default_seller_account_id=f"WB_ACCOUNT_{default_account_index}",
            )
        except integrations.IntegrationSecretError as exc:
            issues.append(_credential_issue(integration.provider, True, str(exc)))
            return None

    def _ozon_settings_from_integrations(
        self,
        db: Session,
        integrations_list: list[TenantIntegration],
        issues: list[dict[str, Any]],
        *,
        required: bool = False,
    ) -> tuple[OzonSettings | None, dict[str, str]]:
        accounts = []
        ozon_cabinet_ids: dict[str, str] = {}
        skipped_roles: list[str] = []
        for integration in integrations_list:
            if integration.status == "disabled":
                continue
            role = str(
                (integration.config_payload or {}).get("connectionRole") or ""
            ).strip()
            if role and role not in OZON_REFRESH_ROLES:
                skipped_roles.append(integration.provider)
                continue
            settings = self._ozon_settings_from_integration(
                integration,
                issues,
                required=required,
            )
            if settings is not None:
                accounts.extend(settings.accounts)
                cabinet = _ensure_wb_cabinet_for_integration(db, integration)
                if cabinet is not None:
                    for account in settings.accounts:
                        ozon_cabinet_ids[account.seller_account_id] = cabinet.id
        if accounts:
            return OzonSettings(accounts=tuple(accounts)), ozon_cabinet_ids
        if skipped_roles:
            issues.append(
                _credential_issue(
                    "ozon_api",
                    required,
                    "no_matching_ozon_refresh_roles",
                    payload={"skippedProviders": skipped_roles},
                )
            )
        return None, ozon_cabinet_ids

    def _ozon_settings_from_integration(
        self,
        integration: TenantIntegration,
        issues: list[dict[str, Any]],
        *,
        required: bool = False,
    ) -> OzonSettings | None:
        secret = self._encrypted_secret_or_issue(integration, issues, required=required)
        if not secret:
            return None
        try:
            return ozon_settings_from_secret(
                secret,
                default_name=_integration_account_name(integration),
                default_seller_account_id=_safe_ozon_account_id(integration.provider),
            )
        except OzonConfigError as exc:
            issues.append(_credential_issue(integration.provider, required, str(exc)))
            return None

    def _onec_settings_from_integration(
        self,
        integration: TenantIntegration,
        issues: list[dict[str, Any]],
    ) -> OnecODataSettings | None:
        secret = self._encrypted_secret_or_issue(integration, issues)
        if not secret:
            return None
        try:
            return integrations.onec_odata_settings_from_secret(secret)
        except integrations.IntegrationSecretError as exc:
            issues.append(_credential_issue("onec_readonly", True, str(exc)))
            return None

    def _encrypted_secret_or_issue(
        self,
        integration: TenantIntegration,
        issues: list[dict[str, Any]],
        *,
        required: bool = True,
    ) -> str:
        payload = integration.config_payload or {}
        if integration.status == "disabled":
            issues.append(
                _credential_issue(
                    integration.provider,
                    required,
                    "integration_disabled",
                )
            )
            return ""
        if integration.status not in READY_INTEGRATION_STATUSES:
            issues.append(
                _credential_issue(
                    integration.provider,
                    required,
                    "integration_not_runtime_ready",
                    payload={
                        "status": integration.status,
                        "lastCheckedAt": (
                            integration.last_checked_at.isoformat()
                            if integration.last_checked_at
                            else ""
                        ),
                    },
                )
            )
            return ""
        if payload.get("storage") != "encrypted":
            issues.append(
                _credential_issue(
                    integration.provider,
                    required,
                    "secret_storage_is_not_encrypted",
                    payload={"storageMode": payload.get("storage", "hash_only")},
                )
            )
            return ""
        try:
            return integrations.decrypt_secret(self.settings, payload)
        except integrations.IntegrationSecretError as exc:
            issues.append(_credential_issue(integration.provider, required, str(exc)))
            return ""

    def _record_mapping_source(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
    ) -> SourceRefreshCollection:
        mapping_dir = self.settings.source_refresh_mapping_path
        status, snapshot_hash, item_count, error_message, payload = (
            mapping_service.inspect_mapping_service(
                db,
                tenant_id=refresh_run.tenant_id,
                client_id=refresh_run.client_id,
                stale_after_days=self._mapping_stale_days(),
            )
        )
        file_status, file_hash, file_count, file_error, file_payload = (
            inspect_mapping_source(
                mapping_dir,
                stale_after_days=self._mapping_stale_days(),
            )
        )
        payload = {
            **payload,
            "legacyFileSource": {
                "status": file_status,
                "snapshotHash": file_hash,
                "fileCount": file_count,
                "errorMessage": file_error,
                **file_payload,
            },
        }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="sku_mapping",
            source_label="Marketplace ↔ 1C mapping service",
            required=True,
            status=status,
            snapshot_hash=snapshot_hash,
            row_count=item_count,
            raw_path="mapping_service",
            error_message=error_message,
            payload=payload,
        )
        if file_status in {"loaded", "stale"}:
            _persist_mapping_rows(db, collection, mapping_dir)
        return collection

    def _rebuild_mapping_service(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        mapping_collection: SourceRefreshCollection,
        user: User | None,
    ) -> None:
        result = mapping_service.rebuild_candidates(
            db,
            tenant_id=refresh_run.tenant_id,
            client_id=refresh_run.client_id,
            user=user,
            refresh_run_id=refresh_run.id,
        )
        import_result = mapping_service.import_mapping_directory(
            db,
            tenant_id=refresh_run.tenant_id,
            client_id=refresh_run.client_id,
            mapping_dir=self.settings.source_refresh_mapping_path,
            user=user,
        )
        status, snapshot_hash, item_count, error_message, payload = (
            mapping_service.inspect_mapping_service(
                db,
                tenant_id=refresh_run.tenant_id,
                client_id=refresh_run.client_id,
                stale_after_days=self._mapping_stale_days(),
            )
        )
        legacy_file_source = (mapping_collection.payload or {}).get(
            "legacyFileSource",
            {},
        )
        mapping_collection.status = status
        mapping_collection.snapshot_hash = snapshot_hash
        mapping_collection.row_count = item_count
        mapping_collection.error_message = error_message
        mapping_collection.payload = {
            **payload,
            "rebuild": result,
            "mappingFileImport": import_result,
            "legacyFileSource": legacy_file_source,
        }
        mapping_collection.loaded_at = security.utcnow()
        db.flush()

    def _record_wb_finance(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbFinancePageResult],
        *,
        wb_cabinet_ids: dict[str, str],
        source_coverage_start: date,
        source_coverage_end: date,
    ) -> None:
        result_items = list(results)
        payload_items = [
            _wb_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        max_pages = self._wb_max_pages()
        max_pages_exhausted = _mark_wb_finance_max_pages_issue(
            result_items,
            payload_items,
            max_pages=max_pages,
        )
        row_count = sum(int(item.get("rowCount", 0)) for item in payload_items)
        persist_row_limit = max(
            0,
            int(self.settings.source_refresh_wb_persist_row_limit),
        )
        files_only = self.settings.source_refresh_raw_db_mode == "files_only"
        skip_row_persistence = files_only or row_count > persist_row_limit
        payload: dict[str, Any] = {
            "results": payload_items,
            "sourceCoverageStart": source_coverage_start.isoformat(),
            "sourceCoverageEnd": source_coverage_end.isoformat(),
        }
        if files_only:
            payload["rowPersistence"] = {
                "status": "file_authoritative",
                "rawFilesAuthoritative": True,
            }
        elif skip_row_persistence:
            payload["rowPersistence"] = {
                "status": "skipped_large_snapshot",
                "limit": persist_row_limit,
                "rawFilesAuthoritative": True,
            }
        error_message = ""
        if max_pages_exhausted:
            error_message = (
                "WB finance reached max_pages with a next rrd_id; "
                "source snapshot is incomplete and cannot be published safely."
            )
            payload.update(
                {
                    "completenessIssue": WB_FINANCE_MAX_PAGES_ISSUE,
                    "maxPages": max_pages,
                    "safeMessage": error_message,
                }
            )
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_finance_detail",
            source_label="WB Finance sales report details",
            required=True,
            status=_aggregate_status(payload_items, required=True),
            snapshot_hash=_hash_payload(payload_items),
            row_count=row_count,
            raw_path=str(output_dir),
            error_message=error_message,
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if max_pages_exhausted or skip_row_persistence:
            return
        _persist_wb_finance_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )

    def _record_wb_report_list(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbSalesReportListPageResult],
        *,
        wb_cabinet_ids: dict[str, str],
    ) -> None:
        result_items = list(results)
        payload_items = [
            _wb_report_list_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        max_pages_exhausted = _page_limit_exhausted(
            result_items,
            max_pages=10,
            page_limit=1000,
            row_count_attribute="row_count",
        )
        payload: dict[str, Any] = {"results": payload_items}
        if max_pages_exhausted:
            payload["completenessIssue"] = WB_REPORT_LIST_MAX_PAGES_ISSUE
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_sales_report_list",
            source_label="WB Finance sales report list",
            required=False,
            status=(
                "partial_source"
                if max_pages_exhausted
                else _aggregate_status(payload_items, required=False)
            ),
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(int(item.get("rowCount", 0)) for item in payload_items),
            raw_path=str(output_dir),
            error_message=(
                "WB report list reached max_pages on a full page; "
                "source may be incomplete."
                if max_pages_exhausted
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return
        _persist_wb_report_list_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )

    def _record_wb_redeem_notifications(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbDocumentExportResult],
        *,
        wb_cabinet_ids: dict[str, str],
    ) -> SourceRefreshCollection:
        result_items = list(results)
        payload_items = [
            _wb_document_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        document_rows = _wb_redeem_notification_rows(
            output_dir,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )
        summary_file = output_dir / "redeem_notifications.summary.json"
        summary_file.write_text(
            json.dumps(document_rows, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        summary_hash = _hash_payload(document_rows)
        integrity_results = [
            {
                "sellerAccountId": "WB_DOCUMENTS",
                "pageIndex": 1,
                "status": "loaded" if document_rows else "empty_expected",
                "ok": True,
                "rowCount": len(document_rows),
                "statusCode": 200,
                "rawPayloadHash": summary_hash,
                "outputFile": summary_file.name,
            }
        ]
        manifest_path = output_dir / "manifest.json"
        try:
            provider_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, TypeError, ValueError, json.JSONDecodeError):
            provider_manifest = {}
        if not isinstance(provider_manifest, dict):
            provider_manifest = {}
        provider_manifest["provider_results"] = provider_manifest.get(
            "provider_results",
            provider_manifest.get("results", []),
        )
        provider_manifest["results"] = integrity_results
        manifest_path.write_text(
            json.dumps(provider_manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        payload: dict[str, Any] = {
            "results": integrity_results,
            "accounts": payload_items,
            "parsedDocuments": len(document_rows),
            "documentsApiRequired": True,
        }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_redeem_notifications",
            source_label="WB primary redeem notifications",
            required=False,
            status=_aggregate_status(payload_items, required=False),
            snapshot_hash=_hash_payload(integrity_results),
            row_count=len(document_rows),
            raw_path=str(output_dir),
            error_message="; ".join(item.error for item in result_items if item.error),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        _persist_wb_redeem_notification_rows(db, collection, document_rows)
        return collection

    def _record_wb_product_cards(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbProductCardsPageResult],
        *,
        wb_cabinet_ids: dict[str, str],
    ) -> None:
        result_items = list(results)
        payload_items = [
            _wb_product_cards_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        max_pages_exhausted = _page_limit_exhausted(
            result_items,
            max_pages=self._wb_max_pages(),
            page_limit=min(self._wb_limit(), 100),
            row_count_attribute="card_count",
            group_attributes=("seller_account_id", "cards_source"),
        )
        payload: dict[str, Any] = {"results": payload_items}
        if max_pages_exhausted:
            payload["completenessIssue"] = WB_PRODUCT_CARDS_MAX_PAGES_ISSUE
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_product_cards",
            source_label="WB product cards",
            required=True,
            status=(
                "partial_source"
                if max_pages_exhausted
                else _aggregate_status(payload_items, required=True)
            ),
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(int(item.get("rowCount", 0)) for item in payload_items),
            raw_path=str(output_dir),
            error_message=(
                "WB product cards reached max_pages on a full page; "
                "source may be incomplete."
                if max_pages_exhausted
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return
        _persist_wb_product_card_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )

    def _record_wb_tariffs(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbTariffsExportResult],
        *,
        wb_cabinet_ids: dict[str, str],
        period_start: date,
        period_end: date,
    ) -> SourceRefreshCollection:
        result_items = list(results)
        payload_items = [
            _wb_tariffs_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        factor_snapshot_date = datetime.now(tz=MOSCOW_TZ).date()
        manifest_path = output_dir / "manifest.json"
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, TypeError, ValueError):
            manifest = {}
        if isinstance(manifest, Mapping):
            try:
                factor_snapshot_date = date.fromisoformat(
                    str(manifest.get("factorSnapshotDate") or "")
                )
            except ValueError:
                factor_snapshot_date = datetime.now(tz=MOSCOW_TZ).date()
        successful = [item for item in result_items if item.ok]
        failed = [item for item in result_items if not item.ok]
        if successful and failed:
            status = "partial_source"
        elif successful:
            status = "loaded"
        else:
            status = "needs_review"
        payload: dict[str, Any] = {
            "results": payload_items,
            "periodStart": period_start.isoformat(),
            "periodEnd": period_end.isoformat(),
            "factorSnapshotDate": factor_snapshot_date.isoformat(),
        }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_tariffs",
            source_label="WB box/pallet tariffs",
            required=False,
            status=status,
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(int(item.get("rowCount") or 0) for item in payload_items),
            raw_path=str(output_dir),
            error_message=(
                "Some tariff dates are unavailable."
                if successful and failed
                else "Tariff source is unavailable."
                if failed
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return collection
        _persist_wb_tariff_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )
        return collection

    def _record_wb_goods_return(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbGoodsReturnExportResult],
        *,
        wb_cabinet_ids: dict[str, str],
        period_start: date,
        period_end: date,
    ) -> SourceRefreshCollection:
        result_items = list(results)
        payload_items = [
            _wb_goods_return_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        manifest: Mapping[str, Any] = {}
        try:
            loaded = json.loads(
                (output_dir / "manifest.json").read_text(encoding="utf-8")
            )
            if isinstance(loaded, Mapping):
                manifest = loaded
        except (OSError, UnicodeError, TypeError, ValueError):
            pass
        successful = [item for item in result_items if item.ok]
        failed = [item for item in result_items if not item.ok]
        status = (
            "partial_source"
            if successful and failed
            else "loaded"
            if successful
            else "needs_review"
        )
        payload: dict[str, Any] = {
            "results": payload_items,
            "periodStart": period_start.isoformat(),
            "periodEnd": period_end.isoformat(),
            "coverageStart": str(manifest.get("coverageStart") or ""),
            "coverageEnd": str(manifest.get("coverageEnd") or ""),
        }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_goods_return",
            source_label="WB goods return reasons",
            required=False,
            status=status,
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(item.row_count for item in successful),
            raw_path=str(output_dir),
            error_message=(
                "Some WB goods-return sources are unavailable."
                if successful and failed
                else "WB goods-return source is unavailable."
                if failed
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return collection
        _persist_wb_goods_return_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )
        return collection

    def _record_wb_return_claims(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbReturnClaimsExportResult],
        *,
        wb_cabinet_ids: dict[str, str],
    ) -> SourceRefreshCollection:
        result_items = list(results)
        payload_items = [
            _wb_return_claims_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        manifest: Mapping[str, Any] = {}
        try:
            loaded = json.loads(
                (output_dir / "manifest.json").read_text(encoding="utf-8")
            )
            if isinstance(loaded, Mapping):
                manifest = loaded
        except (OSError, UnicodeError, TypeError, ValueError):
            pass
        successful = [item for item in result_items if item.ok]
        failed = [item for item in result_items if not item.ok]
        status = (
            "partial_source"
            if successful and failed
            else "loaded"
            if successful
            else "needs_review"
        )
        payload: dict[str, Any] = {
            "results": payload_items,
            "coverageStart": str(manifest.get("coverageStart") or ""),
            "coverageEnd": str(manifest.get("coverageEnd") or ""),
        }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_return_claims",
            source_label="WB buyer return claims",
            required=False,
            status=status,
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(item.row_count for item in successful),
            raw_path=str(output_dir),
            error_message=(
                "Some WB return-claims sources are unavailable."
                if successful and failed
                else "WB return-claims source is unavailable."
                if failed
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return collection
        _persist_wb_return_claim_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )
        return collection

    def _record_wb_supplier_sales(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbSupplierSalesExportResult],
        *,
        wb_cabinet_ids: dict[str, str],
        period_start: date,
        period_end: date,
    ) -> SourceRefreshCollection:
        result_items = list(results)
        payload_items = [
            _wb_supplier_sales_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        manifest: Mapping[str, Any] = {}
        try:
            loaded = json.loads(
                (output_dir / "manifest.json").read_text(encoding="utf-8")
            )
            if isinstance(loaded, Mapping):
                manifest = loaded
        except (OSError, UnicodeError, TypeError, ValueError):
            pass
        successful = [item for item in result_items if item.ok]
        failed = [item for item in result_items if not item.ok]
        status = (
            "partial_source"
            if successful and failed
            else "loaded"
            if successful
            else "needs_review"
        )
        payload: dict[str, Any] = {
            "results": payload_items,
            "periodStart": period_start.isoformat(),
            "periodEnd": period_end.isoformat(),
            "coverageStart": str(manifest.get("coverageStart") or ""),
            "coverageEnd": str(manifest.get("coverageEnd") or ""),
            "factorSnapshotDate": str(manifest.get("factorSnapshotDate") or ""),
        }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_supplier_sales",
            source_label="WB supplier sales routes",
            required=False,
            status=status,
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(item.row_count for item in result_items if item.ok),
            raw_path=str(output_dir),
            error_message=(
                "Some supplier-sales routes are unavailable."
                if successful and failed
                else "Supplier-sales routes are unavailable."
                if failed
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return collection
        _persist_wb_supplier_sales_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )
        return collection

    def _record_wb_measurements(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbMeasurementExportResult],
        *,
        source_type: str,
        wb_cabinet_ids: dict[str, str],
        period_start: date,
        period_end: date,
    ) -> SourceRefreshCollection:
        if source_type not in {
            "wb_measurement_penalties",
            "wb_warehouse_measurements",
        }:
            raise ValueError("invalid WB measurement source type")
        result_items = list(results)
        payload_items = [
            _wb_measurement_result_payload(
                item,
                wb_cabinet_id=wb_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        manifest: Mapping[str, Any] = {}
        try:
            loaded = json.loads(
                (output_dir / "manifest.json").read_text(encoding="utf-8")
            )
            if isinstance(loaded, Mapping):
                manifest = loaded
        except (OSError, UnicodeError, TypeError, ValueError):
            pass
        successful = [item for item in result_items if item.ok]
        failed = [item for item in result_items if not item.ok]
        status = (
            "partial_source"
            if successful and failed
            else "loaded"
            if successful
            else "needs_review"
        )
        payload: dict[str, Any] = {
            "results": payload_items,
            "periodStart": period_start.isoformat(),
            "periodEnd": period_end.isoformat(),
            "coverageStart": str(manifest.get("coverageStart") or ""),
            "coverageEnd": str(manifest.get("coverageEnd") or ""),
            "factorSnapshotAt": str(manifest.get("factorSnapshotAt") or ""),
        }
        label = (
            "WB dimension measurement retentions"
            if source_type == "wb_measurement_penalties"
            else "WB warehouse measurements"
        )
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type=source_type,
            source_label=label,
            required=False,
            status=status,
            snapshot_hash=_hash_payload(payload_items),
            row_count=sum(item.row_count for item in successful),
            raw_path=str(output_dir),
            error_message=(
                "Some WB measurement sources are unavailable."
                if successful and failed
                else "WB measurement source is unavailable."
                if failed
                else ""
            ),
            payload=payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if self.settings.source_refresh_raw_db_mode == "files_only":
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "file_authoritative",
                    "rawFilesAuthoritative": True,
                },
            }
            db.flush()
            return collection
        _persist_wb_measurement_rows(
            db,
            collection,
            result_items,
            wb_cabinet_ids=wb_cabinet_ids,
        )
        return collection

    def _record_wb_stock_history(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[WbStockExportResult],
        *,
        wb_cabinet_ids: dict[str, str],
        period_start: date,
        period_end: date,
        actual_period_start: date | None = None,
        actual_period_end: date | None = None,
        provider_window_available: bool = True,
    ) -> SourceRefreshCollection:
        if provider_window_available:
            actual_period_start = actual_period_start or period_start
            actual_period_end = actual_period_end or period_end
        total_days = (period_end - period_start).days + 1
        actual_days = (
            (actual_period_end - actual_period_start).days + 1
            if actual_period_start is not None and actual_period_end is not None
            else 0
        )
        accounts: list[dict[str, Any]] = []
        for item in results:
            source_status = (
                "missing_scope" if item.status == "access_error" else item.status
            )
            covered_days = (
                _stock_history_zip_covered_days(
                    item.output_path,
                    period_start=period_start,
                    period_end=period_end,
                )
                if item.ok and item.output_path is not None
                else 0
            )
            provider_window_calculated = bool(
                provider_window_available
                and actual_days > 0
                and item.ok
                and covered_days == actual_days
            )
            full_coverage = bool(
                provider_window_calculated and actual_days == total_days
            )
            calculated = provider_window_calculated
            status = "complete" if full_coverage else source_status
            if provider_window_calculated and not full_coverage:
                status = "partial_provider_window"
            elif item.ok and not calculated:
                status = (
                    "partial_provider_window"
                    if actual_period_start > period_start
                    else "incomplete"
                )
            accounts.append(
                {
                    "sellerAccountId": item.seller_account_id,
                    "wbCabinetId": wb_cabinet_ids.get(item.seller_account_id, ""),
                    "cabinet": item.account_name,
                    "status": status,
                    "coveredDays": covered_days,
                    "totalDays": total_days,
                    "calculated": calculated,
                    "providerWindowCalculated": provider_window_calculated,
                    "fullCoverage": full_coverage,
                    "calculationPeriodStart": (
                        actual_period_start.isoformat()
                        if actual_period_start is not None
                        else None
                    ),
                    "calculationPeriodEnd": (
                        actual_period_end.isoformat()
                        if actual_period_end is not None
                        else None
                    ),
                    "extrapolated": False,
                    "statusCode": item.status_code,
                    "error": item.error,
                }
            )
        all_complete = bool(accounts) and all(
            bool(item["calculated"]) for item in accounts
        )
        provider_window_complete = bool(accounts) and all(
            bool(item["providerWindowCalculated"]) for item in accounts
        )
        full_coverage = provider_window_complete and all(
            bool(item["fullCoverage"]) for item in accounts
        )
        collection_status = "loaded" if all_complete else "needs_review"
        error_message = ""
        if not all_complete:
            error_message = (
                "WB stock history does not completely cover the common provider "
                "window or Seller Analytics scope is missing; lost contribution "
                "margin is not calculated."
            )
        return repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_stock_history_daily",
            source_label="WB daily stock history",
            required=False,
            status=collection_status,
            snapshot_hash=_hash_payload(accounts),
            row_count=len(accounts),
            raw_path=str(output_dir),
            error_message=error_message,
            payload={
                "periodStart": period_start.isoformat(),
                "periodEnd": period_end.isoformat(),
                "actualPeriodStart": (
                    actual_period_start.isoformat()
                    if actual_period_start is not None
                    else None
                ),
                "actualPeriodEnd": (
                    actual_period_end.isoformat()
                    if actual_period_end is not None
                    else None
                ),
                "stockType": "wb",
                "calculated": all_complete,
                "providerWindowCalculated": provider_window_complete,
                "fullCoverage": full_coverage,
                "calculationPeriodStart": (
                    actual_period_start.isoformat()
                    if actual_period_start is not None
                    else None
                ),
                "calculationPeriodEnd": (
                    actual_period_end.isoformat()
                    if actual_period_end is not None
                    else None
                ),
                "calculationContextVersion": "lost-sales-filter-v1",
                "extrapolated": False,
                "accounts": accounts,
            },
        )

    def _record_ozon_results(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[OzonPageResult],
        *,
        source_type: str,
        source_label: str,
        ozon_cabinet_ids: dict[str, str],
        required: bool = False,
    ) -> None:
        result_items = list(results)
        payload_items = [
            _ozon_result_payload(
                item,
                ozon_cabinet_id=ozon_cabinet_ids.get(item.seller_account_id, ""),
            )
            for item in result_items
        ]
        row_count = _ozon_collection_row_count(result_items)
        files_only = (
            self.settings.source_refresh_raw_db_mode == "files_only"
            and self.settings.marketplace_daily_facts_enabled
            and self.settings.source_refresh_ozon_typed_facts_enabled
            and self.settings.source_refresh_ozon_files_only_enabled
            and source_type in OZON_TYPED_FILE_AUTHORITATIVE_TYPES
        )
        qualification_run_id = ""
        if files_only:
            seller_account_ids = {
                item.seller_account_id
                for item in result_items
                if item.seller_account_id
            }
            qualification_run_ids = {
                qualified_source_type: self._ozon_qualification_run_id(
                    db,
                    refresh_run,
                    source_type=qualified_source_type,
                    seller_account_ids=seller_account_ids,
                )
                for qualified_source_type in OZON_TYPED_FILE_AUTHORITATIVE_TYPES
            }
            missing_qualifications = sorted(
                item for item, run_id in qualification_run_ids.items() if not run_id
            )
            if missing_qualifications:
                raise SourceRefreshConfigError(
                    "full legacy qualification is required for "
                    + ", ".join(missing_qualifications)
                )
            qualification_run_id = qualification_run_ids[source_type]
        collection_payload: dict[str, Any] = {
            "marketplace": "ozon",
            "results": payload_items,
        }
        if files_only:
            collection_payload["rowPersistence"] = {
                "status": "file_authoritative",
                "rawFilesAuthoritative": True,
                "qualificationRunId": qualification_run_id,
            }
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type=source_type,
            source_label=source_label,
            required=required,
            status=_ozon_collection_status(
                result_items,
                payload_items,
                row_count=row_count,
                required=required,
            ),
            snapshot_hash=_hash_payload(payload_items),
            row_count=row_count,
            raw_path=str(output_dir),
            payload=collection_payload,
        )
        _attach_collection_raw_integrity(
            collection,
            source_root=self.settings.source_refresh_root_path,
        )
        if files_only and (
            ((collection.payload or {}).get("rawIntegrity") or {}).get("status")
            != "verified"
        ):
            raise SourceRefreshConfigError(
                f"verified raw files are required for {source_type} files-only mode"
            )
        if not files_only:
            _persist_ozon_rows(
                db,
                collection,
                result_items,
                ozon_cabinet_ids=ozon_cabinet_ids,
            )

    def _ozon_qualification_run_id(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        source_type: str,
        seller_account_ids: set[str] | None = None,
    ) -> str:
        candidates = list(
            db.scalars(
                select(SourceRefreshCollection)
                .where(
                    SourceRefreshCollection.tenant_id == refresh_run.tenant_id,
                    SourceRefreshCollection.client_id == refresh_run.client_id,
                    SourceRefreshCollection.source_type == source_type,
                    SourceRefreshCollection.refresh_run_id != refresh_run.id,
                )
                .order_by(SourceRefreshCollection.loaded_at.desc())
            )
        )
        for candidate in candidates:
            if (
                candidate.status not in MANDATORY_OK_STATUSES
                or candidate.refresh_run.finished_at is None
                or candidate.refresh_run.status
                not in repository.CALCULABLE_OZON_REFRESH_STATUSES
                or candidate.refresh_run.created_at >= refresh_run.created_at
            ):
                continue
            payload = dict(candidate.payload or {})
            if (payload.get("rowPersistence") or {}).get("status") == (
                "file_authoritative"
            ):
                continue
            candidate_sellers = {
                str(item.get("sellerAccountId") or "")
                for item in payload.get("results", [])
                if isinstance(item, dict) and item.get("sellerAccountId")
            }
            if (
                seller_account_ids is not None
                and candidate_sellers != seller_account_ids
            ):
                continue
            typed_parity = dict(payload.get("typedParity") or {})
            if (
                typed_parity.get("status") == "matched"
                and (typed_parity.get("diagnosticsParity") or {}).get("status")
                == "matched"
                and (typed_parity.get("persistenceParity") or {}).get("status")
                == "matched"
                and (typed_parity.get("legacyFileParity") or {}).get("status")
                == "matched"
                and (typed_parity.get("sourceCoverage") or {}).get("status")
                == "matched"
                and (payload.get("rawIntegrity") or {}).get("status") == "verified"
            ):
                return candidate.refresh_run_id
        return ""

    def _promote_ozon_typed_facts(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        ozon_cabinet_ids: dict[str, str],
    ) -> bool:
        if not (
            self.settings.marketplace_daily_facts_enabled
            and self.settings.source_refresh_ozon_typed_facts_enabled
        ):
            return True
        collections = [
            item
            for item in refresh_run.collections
            if item.source_type in OZON_TYPED_FILE_AUTHORITATIVE_TYPES
        ]
        if not collections:
            return True
        invalid = [
            item
            for item in collections
            if item.status not in MANDATORY_OK_STATUSES
            or ((item.payload or {}).get("rawIntegrity") or {}).get("status")
            != "verified"
        ]
        if invalid:
            self._mark_ozon_promotion_failed(
                db,
                collections,
                reason="source_collection_not_promotable",
            )
            return False

        collection_ids = [item.id for item in collections]
        try:
            with db.begin_nested():
                for collection in collections:
                    results, collection_cabinet_ids = _ozon_results_from_collection(
                        collection,
                        source_root=self.settings.source_refresh_root_path,
                    )
                    _materialize_ozon_typed_collection(
                        db,
                        refresh_run,
                        collection,
                        results,
                        ozon_cabinet_ids={
                            **ozon_cabinet_ids,
                            **collection_cabinet_ids,
                        },
                    )
                    typed_parity = (collection.payload or {}).get("typedParity") or {}
                    if typed_parity.get("status") != "pending_diagnostics":
                        raise ValueError("ozon_typed_staging_parity_failed")
                daily_facts = _ozon_daily_facts_for_run(db, refresh_run)
                daily_count = repository.replace_marketplace_finance_daily_facts(
                    db,
                    refresh_run,
                    daily_facts,
                    marketplace="ozon",
                    cabinet_ids=ozon_cabinet_ids,
                )
                for collection in collections:
                    collection.payload = {
                        **(collection.payload or {}),
                        "operationFacts": {
                            **((collection.payload or {}).get("operationFacts") or {}),
                            "dailyRowCount": daily_count,
                            "promotionScope": "all_ozon_sources",
                        },
                    }
                db.flush()
        except Exception as exc:
            current_collections = [
                item
                for collection_id in collection_ids
                if (item := db.get(SourceRefreshCollection, collection_id)) is not None
            ]
            self._mark_ozon_promotion_failed(
                db,
                current_collections,
                reason=f"promotion_rolled_back:{exc.__class__.__name__}",
            )
            return False
        return True

    @staticmethod
    def _mark_ozon_promotion_failed(
        db: Session,
        collections: Iterable[SourceRefreshCollection],
        *,
        reason: str,
    ) -> None:
        for collection in collections:
            collection.status = "needs_review"
            collection.payload = {
                **(collection.payload or {}),
                "typedParity": {
                    "status": "promotion_failed",
                    "reason": reason,
                    "previousFactsPreserved": True,
                },
            }
        db.flush()

    def _record_onec(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        output_dir: Path,
        results: Iterable[OnecSampleExportResult],
    ) -> None:
        for item in results:
            required = item.sample_id in MANDATORY_ONEC_COLLECTION_IDS
            publication_required = (
                item.sample_id in PUBLICATION_REQUIRED_ONEC_COLLECTION_IDS
            )
            status = _onec_status(item, required=required)
            data_quality = _onec_financial_table_quality(item)
            if data_quality["status"] == "partial_source":
                status = "partial_source"
            collection = repository.add_source_refresh_collection(
                db,
                refresh_run,
                source_type=f"onec_{item.sample_id}",
                source_label=item.collection_name,
                required=required,
                publication_required=publication_required,
                status=status,
                snapshot_hash=item.raw_payload_hash,
                row_count=item.row_count,
                raw_path=str(item.output_path or item.checkpoint_path or output_dir),
                error_message=item.error,
                payload={
                    "sampleId": item.sample_id,
                    "statusCode": item.status_code,
                    "pageCount": item.page_count,
                    "checkpointPath": str(item.checkpoint_path or ""),
                    "retryable": item.retryable,
                    "nextCursor": item.next_cursor,
                    "reusedPageCount": item.reused_page_count,
                    "effectivePageSize": item.effective_page_size,
                    "detailMode": item.detail_mode,
                    "publicationRequired": publication_required,
                    "dataQuality": data_quality,
                },
            )
            _persist_onec_rows(db, collection, item)

    def _record_onec_metadata_check(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        onec_settings: OnecODataSettings,
    ) -> OnecODataMetadataCheckResult:
        result = self._onec_metadata_checker(onec_settings)
        checked_at = security.utcnow().isoformat()
        status = "loaded" if result.ok else "failed"
        message = (
            "Метаданные 1С OData доступны в режиме только для чтения."
            if result.ok
            else "Метаданные 1С OData недоступны для автоматического обновления."
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_odata_metadata",
            source_label="Метаданные 1С OData",
            required=True,
            status=status,
            error_message=result.error,
            payload={
                "endpointCategory": "odata_metadata",
                "statusCode": result.status_code,
                "metadataValid": result.ok,
                "attemptCount": result.attempt_count,
                "timeoutSeconds": result.timeout_seconds,
                "checkedAt": checked_at,
            },
        )
        if refresh_run.credential_source == "tenant":
            repository.record_tenant_integration_runtime_check(
                db,
                tenant_id=refresh_run.tenant_id,
                provider="onec_readonly",
                status="check_ok" if result.ok else "check_failed",
                message=message,
                check_payload={
                    "checkMode": "source_refresh_preflight",
                    "endpointCategory": "odata_metadata",
                    "httpStatus": result.status_code,
                    "metadataValid": result.ok,
                    "errorType": result.error,
                    "attemptCount": result.attempt_count,
                    "timeoutSeconds": result.timeout_seconds,
                },
            )
        return result

    def _finish_without_report(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        status: str,
        error_message: str = "",
        failure_code: str = "",
        retryable: bool = False,
    ) -> dict[str, Any]:
        if (
            status == "failed"
            and retryable
            and db.info.get("source_refresh_split_pipeline")
            and repository.requeue_transient_source_refresh_task(
                db,
                refresh_run,
                task_type="collect_sources",
                safe_error_code=failure_code or "collect_sources_transport_failed",
                safe_error_message=error_message,
            )
        ):
            repository.audit(
                db,
                action="source_refresh_task_retry_scheduled",
                user=None,
                tenant_id=refresh_run.tenant_id,
                entity_type="source_refresh_run",
                entity_id=refresh_run.id,
                payload={
                    "taskType": "collect_sources",
                    "failureCode": failure_code
                    or "collect_sources_transport_failed",
                },
            )
            return repository.source_refresh_run_payload(refresh_run)
        update_fields: dict[str, Any] = {}
        if failure_code:
            update_fields["failure_code"] = failure_code
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status=status,
            error_message=error_message,
            finished_at=security.utcnow(),
            **update_fields,
        )
        if status == "failed":
            with suppress(Exception):
                self._prune_failed_snapshot_directories(db, refresh_run)
        return repository.source_refresh_run_payload(refresh_run)

    def _prune_failed_snapshot_directories(
        self,
        db: Session,
        current_run: SourceRefreshRun,
    ) -> list[Path]:
        keep = max(2, int(self.settings.source_refresh_failed_snapshot_keep))
        failed_runs = list(
            db.scalars(
                select(SourceRefreshRun)
                .where(
                    SourceRefreshRun.tenant_id == current_run.tenant_id,
                    SourceRefreshRun.status == "failed",
                    SourceRefreshRun.finished_at.is_not(None),
                )
                .order_by(
                    SourceRefreshRun.created_at.desc(),
                    SourceRefreshRun.id.desc(),
                )
            )
        )
        protected_snapshot_ids = {
            str(value)
            for value in db.scalars(
                select(ReportRun.source_snapshot_set_id).where(
                    ReportRun.tenant_id == current_run.tenant_id,
                    ReportRun.publication_status == "published",
                    ReportRun.source_snapshot_set_id != "",
                )
            )
            if value
        }
        source_root = self.settings.source_refresh_root_path.resolve()
        removed: list[Path] = []
        for failed_run in failed_runs[keep:]:
            snapshot_set_id = failed_run.snapshot_set_id
            if (
                failed_run.new_report_run_id
                or snapshot_set_id in protected_snapshot_ids
                or not snapshot_set_id
                or Path(snapshot_set_id).name != snapshot_set_id
            ):
                continue
            candidate = source_root / snapshot_set_id
            if candidate.parent != source_root or candidate.is_symlink():
                continue
            if candidate.is_dir():
                shutil.rmtree(candidate)
                removed.append(candidate)
        return removed

    def _build_workbook(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        wb_finance_dir: Path | None,
        onec_dir: Path | None,
        wb_report_list_dir: Path | None,
        wb_stock_history_dir: Path | None,
        raw_refresh_run: SourceRefreshRun | None = None,
    ) -> Path:
        integrity_refresh_run = raw_refresh_run or refresh_run
        if wb_finance_dir is not None:
            _reverify_collection_raw_integrity(
                integrity_refresh_run,
                source_type="wb_finance_detail",
                raw_path=wb_finance_dir,
                source_root=self.settings.source_refresh_root_path,
            )
        if wb_report_list_dir is not None:
            _reverify_collection_raw_integrity(
                integrity_refresh_run,
                source_type="wb_sales_report_list",
                raw_path=wb_report_list_dir,
                source_root=self.settings.source_refresh_root_path,
            )
        output_dir = (
            self.settings.export_root_path / "source_refresh" / refresh_run.id
        ).resolve()
        allowed = self.settings.export_root_path.resolve()
        if output_dir != allowed and allowed not in output_dir.parents:
            raise ValueError("source-refresh workbook path is outside reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "shumeyko_wb_excel_mvp.xlsx"
        tax_profiles = repository.tax_profiles_for_source_refresh(db, refresh_run)
        sku_mappings = self._calculation_sku_mappings(db, refresh_run)
        args = argparse.Namespace(
            client_id=(
                refresh_run.client_id
                or repository.client_id_for_tenant(refresh_run.tenant_id)
            ),
            wb_finance_dir=wb_finance_dir,
            wb_finance_source="files",
            postgres_db_name="shumeyko_wb_unit_economics",
            postgres_host="",
            postgres_port=55433,
            postgres_user="",
            postgres_snapshot_id=None,
            mapping_source="files",
            mapping_snapshot_id=None,
            cost_source="files",
            cost_snapshot_id=None,
            wb_cards_dir=None,
            onec_dir=onec_dir,
            onec_marketplace_mapping_dir=self.settings.source_refresh_mapping_path,
            sales_register_dir=onec_dir,
            onec_services_dir=onec_dir,
            wb_report_list_dir=wb_report_list_dir,
            wb_paid_storage_dir=None,
            wb_promotion_stats_dir=None,
            wb_stock_history_dir=wb_stock_history_dir,
            onec_stock_dir=onec_dir,
            onec_opiu_dir=onec_dir,
            onec_opiu_config=None,
            output=output_path,
            report_period_start=refresh_run.period_start,
            report_period_end=refresh_run.period_end,
            cost_amount_field="Сумма",
            sales_cost_amount_field="СебестоимостьБезНДС",
            tax_profiles=tax_profiles,
            sku_mappings=sku_mappings,
        )
        self._workbook_builder(args)
        if not output_path.exists():
            raise ValueError("source refresh workbook was not created")
        return output_path

    def _calculation_sku_mappings(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
    ) -> list[SkuMapping]:
        exported = mapping_service.export_sku_mapping(
            db,
            tenant_id=refresh_run.tenant_id,
            client_id=refresh_run.client_id,
        )
        return [
            SkuMapping.model_validate(item)
            for item in exported.get("skuMappingRows", [])
        ]

    def _materialize_wb_daily_facts(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        wb_finance_dir: Path,
        onec_dir: Path,
        wb_report_list_dir: Path | None,
        wb_cards_dir: Path | None,
        wb_stock_history_dir: Path | None,
    ) -> None:
        from scripts.rebuild_report_from_sources import build_db_first_payload

        _reverify_collection_raw_integrity(
            refresh_run,
            source_type="wb_finance_detail",
            raw_path=wb_finance_dir,
            source_root=self.settings.source_refresh_root_path,
        )

        replacement_summary_rows = self._incremental_wb_summary_rows(
            refresh_run,
            base_refresh_run=None,
            current_report_list_dir=wb_report_list_dir,
        )
        coverage_start = refresh_run.source_window_start or refresh_run.period_start
        coverage_end = refresh_run.source_window_end or refresh_run.period_end
        materialization_boundary_dates = [
            coverage_start,
            coverage_end,
            *(item.date_from for item in replacement_summary_rows),
            *(item.date_to for item in replacement_summary_rows),
            *(item.create_date for item in replacement_summary_rows),
        ]
        materialization_period_start = week_bounds(min(materialization_boundary_dates))[
            0
        ]
        materialization_period_end = week_bounds(max(materialization_boundary_dates))[1]

        args = argparse.Namespace(
            client_id=refresh_run.client_id,
            wb_finance_dir=wb_finance_dir,
            wb_finance_source="files-stream",
            wb_sales_report_summary_rows=replacement_summary_rows,
            stream_cache_dir=(
                Path("data/.cache/source_refresh_stream") / refresh_run.id
            ),
            keep_stream_cache=False,
            marketplace_daily_facts_enabled=True,
            postgres_db_name="shumeyko_wb_unit_economics",
            postgres_host="",
            postgres_port=55433,
            postgres_user="",
            postgres_snapshot_id=None,
            mapping_source="files",
            mapping_snapshot_id=None,
            cost_source="files",
            cost_snapshot_id=None,
            wb_cards_dir=wb_cards_dir,
            onec_dir=onec_dir,
            onec_marketplace_mapping_dir=self.settings.source_refresh_mapping_path,
            sales_register_dir=onec_dir,
            onec_services_dir=onec_dir,
            wb_report_list_dir=wb_report_list_dir,
            wb_paid_storage_dir=None,
            wb_promotion_stats_dir=None,
            wb_stock_history_dir=wb_stock_history_dir,
            onec_stock_dir=onec_dir,
            onec_opiu_dir=onec_dir,
            onec_opiu_config=None,
            report_period_start=materialization_period_start,
            report_period_end=materialization_period_end,
            cost_amount_field="Сумма",
            sales_cost_amount_field="auto",
            source_refresh_run_id=refresh_run.id,
            tenant_name=self._client_name(db, refresh_run.tenant_id),
            sku_mappings=self._calculation_sku_mappings(db, refresh_run),
        )
        tax_profiles = repository.tax_profiles_for_source_refresh(db, refresh_run)
        input_vat_policies = repository.input_vat_policies_for_source_refresh(
            db, refresh_run
        )
        db.commit()
        build = build_db_first_payload(
            args,
            tax_profiles=tax_profiles,
            input_vat_policies=input_vat_policies,
        )
        legacy_row_count = repository.source_snapshot_row_count_for_run(
            db,
            refresh_run_id=refresh_run.id,
            source_type="wb_finance_detail",
        )
        calculation_parity = {
            "status": "not_run_no_legacy_rows",
            "legacyRowCount": legacy_row_count,
        }
        if legacy_row_count > 0:
            source_parity = _legacy_wb_source_parity(
                db,
                refresh_run,
                wb_finance_dir=wb_finance_dir,
            )
            legacy_args = argparse.Namespace(**vars(args))
            legacy_args.wb_finance_source = "files"
            legacy_build = build_db_first_payload(
                legacy_args,
                tax_profiles=tax_profiles,
                input_vat_policies=input_vat_policies,
            )
            calculation_parity = _full_wb_calculation_parity(
                legacy_build,
                build,
                legacy_row_count=legacy_row_count,
                source_parity=source_parity,
            )
            parity_path = Path(refresh_run.root_dir) / "parity" / "wb-calculation.json"
            parity_path.parent.mkdir(parents=True, exist_ok=True)
            parity_path.write_text(
                json.dumps(
                    calculation_parity,
                    ensure_ascii=False,
                    indent=2,
                    sort_keys=True,
                ),
                encoding="utf-8",
            )
            calculation_parity["artifactPath"] = str(parity_path)
        self._save_onec_cost_snapshots(db, refresh_run, build)
        self._save_wb_daily_facts(
            db,
            refresh_run,
            build,
            calculation_parity=calculation_parity,
            replacement_summary_rows=replacement_summary_rows,
        )
        _commit_source_refresh_progress(db)

    def _save_wb_daily_facts(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        build: dict[str, Any],
        *,
        calculation_parity: dict[str, Any] | None = None,
        replacement_summary_rows: Iterable[WbSalesReportSummaryRow] | None = None,
    ) -> None:
        source_daily_facts = build.pop("daily_facts", [])
        all_daily_facts = (
            source_daily_facts
            if isinstance(source_daily_facts, list)
            else list(source_daily_facts)
        )
        parity = _wb_daily_fact_parity(build, all_daily_facts)
        coverage_start = refresh_run.source_window_start or refresh_run.period_start
        coverage_end = refresh_run.source_window_end or refresh_run.period_end
        replacement_report_keys = (
            {
                (
                    str(item.seller_account_id).strip(),
                    str(item.marketplace_report_id).strip(),
                )
                for item in all_daily_facts
                if str(item.seller_account_id).strip()
                and str(item.marketplace_report_id).strip()
            }
            if replacement_summary_rows is None
            else {
                (
                    str(item.seller_account_id).strip(),
                    str(item.report_id).strip(),
                )
                for item in replacement_summary_rows
                if str(item.seller_account_id).strip() and str(item.report_id).strip()
            }
        )
        daily_facts = [
            item
            for item in all_daily_facts
            if coverage_start <= item.fact_date <= coverage_end
            or (
                item.seller_account_id,
                item.marketplace_report_id,
            )
            in replacement_report_keys
        ]
        daily_fact_count = repository.replace_marketplace_finance_daily_facts(
            db,
            refresh_run,
            daily_facts,
            marketplace="wb",
            cabinet_ids=_marketplace_cabinet_ids(
                refresh_run,
                source_type="wb_finance_detail",
            ),
            coverage_start=coverage_start,
            coverage_end=coverage_end,
            report_keys=replacement_report_keys,
        )
        persisted_parity = _persisted_daily_facts_parity(
            db,
            refresh_run,
            daily_facts,
        )
        finance_collection = next(
            (
                item
                for item in refresh_run.collections
                if item.source_type == "wb_finance_detail"
            ),
            None,
        )
        if finance_collection is None:
            return
        finance_collection.payload = {
            **(finance_collection.payload or {}),
            "dailyFacts": {
                "status": "materialized",
                "rowCount": daily_fact_count,
                "periodStart": (
                    min(item.fact_date for item in daily_facts).isoformat()
                    if daily_facts
                    else None
                ),
                "periodEnd": (
                    max(item.fact_date for item in daily_facts).isoformat()
                    if daily_facts
                    else None
                ),
                "parity": parity,
                "persistedParity": persisted_parity,
            },
            "calculationParity": calculation_parity or {"status": "not_run"},
        }
        db.flush()

    @staticmethod
    def _save_onec_cost_snapshots(
        db: Session,
        refresh_run: SourceRefreshRun,
        build: dict[str, Any],
    ) -> None:
        snapshots = build.pop("cost_snapshots", [])
        count = repository.replace_onec_unf_cost_snapshots(
            db,
            refresh_run,
            snapshots,
        )
        collection = next(
            (
                item
                for item in refresh_run.collections
                if item.source_type == "onec_sales_register"
            ),
            None,
        )
        if collection is not None:
            collection.payload = {
                **(collection.payload or {}),
                "typedCostSnapshots": {
                    "status": "materialized",
                    "rowCount": count,
                    "contract": "onec_unf_cost_snapshot",
                },
            }
        db.flush()

    def _daily_facts_for_report(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        wb_summary_rows: Iterable[WbSalesReportSummaryRow] = (),
    ) -> list[MarketplaceFinanceDailyFactContract]:
        report_keys = sorted(
            {
                (
                    str(item.seller_account_id).strip(),
                    str(item.report_id).strip(),
                )
                for item in wb_summary_rows
                if str(item.seller_account_id).strip() and str(item.report_id).strip()
            }
        )
        fact_period_start = week_bounds(refresh_run.period_start)[0]
        fact_period_end = week_bounds(refresh_run.period_end)[1]
        report_scope = and_(
            MarketplaceFinanceDailyFactModel.fact_date >= fact_period_start,
            MarketplaceFinanceDailyFactModel.fact_date <= fact_period_end,
        )
        if report_keys:
            report_scope = or_(
                report_scope,
                *(
                    and_(
                        MarketplaceFinanceDailyFactModel.seller_account_id
                        == seller_account_id,
                        MarketplaceFinanceDailyFactModel.marketplace_report_id
                        == report_id,
                    )
                    for seller_account_id, report_id in report_keys
                ),
            )
        rows = list(
            db.scalars(
                select(MarketplaceFinanceDailyFactModel)
                .where(
                    MarketplaceFinanceDailyFactModel.tenant_id == refresh_run.tenant_id,
                    MarketplaceFinanceDailyFactModel.client_id == refresh_run.client_id,
                    MarketplaceFinanceDailyFactModel.marketplace == "wb",
                    report_scope,
                )
                .order_by(
                    MarketplaceFinanceDailyFactModel.fact_date,
                    MarketplaceFinanceDailyFactModel.grain_hash,
                )
            )
        )
        if not rows:
            raise SourceRefreshConfigError(
                "incremental daily-facts report input is empty"
            )
        field_names = MarketplaceFinanceDailyFactContract.model_fields
        return [
            MarketplaceFinanceDailyFactContract.model_validate(
                {name: getattr(row, name) for name in field_names}
            )
            for row in rows
        ]

    def _incremental_wb_summary_rows(
        self,
        refresh_run: SourceRefreshRun,
        *,
        base_refresh_run: SourceRefreshRun | None,
        current_report_list_dir: Path | None,
    ) -> list[WbSalesReportSummaryRow]:
        rows_by_key: dict[tuple[str, str, int | None], WbSalesReportSummaryRow] = {}
        if base_refresh_run is not None:
            base_dir = self._optional_collection_raw_dir(
                base_refresh_run,
                "wb_sales_report_list",
            )
            if base_dir is not None:
                _reverify_collection_raw_integrity(
                    base_refresh_run,
                    source_type="wb_sales_report_list",
                    raw_path=base_dir,
                    source_root=self.settings.source_refresh_root_path,
                )
                for row in load_wb_sales_report_summary_rows(
                    base_dir,
                    client_id=refresh_run.client_id,
                ):
                    if row.date_to < (
                        refresh_run.source_window_start or refresh_run.period_start
                    ):
                        rows_by_key[
                            (row.seller_account_id, row.report_id, row.report_type)
                        ] = row
        if current_report_list_dir is not None:
            _reverify_collection_raw_integrity(
                refresh_run,
                source_type="wb_sales_report_list",
                raw_path=current_report_list_dir,
                source_root=self.settings.source_refresh_root_path,
            )
            for row in load_wb_sales_report_summary_rows(
                current_report_list_dir,
                client_id=refresh_run.client_id,
            ):
                rows_by_key[(row.seller_account_id, row.report_id, row.report_type)] = (
                    row
                )
        return sorted(
            rows_by_key.values(),
            key=lambda row: (
                row.date_to,
                row.seller_account_id,
                row.report_id,
                row.report_type or 0,
            ),
        )

    def _build_db_first_report(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        *,
        source_report: ReportRun | None,
        wb_finance_dir: Path | None,
        onec_dir: Path | None,
        wb_report_list_dir: Path | None,
        wb_cards_dir: Path | None,
        wb_stock_history_dir: Path | None,
        source_snapshot_set_id: str,
        base_refresh_run: SourceRefreshRun | None,
        contributing_runs: Iterable[SourceRefreshRun] = (),
        wb_daily_facts: list[MarketplaceFinanceDailyFactContract] | None = None,
        wb_summary_rows: list[WbSalesReportSummaryRow] | None = None,
    ) -> tuple[ReportRun, Path]:
        from scripts.rebuild_report_from_sources import (
            _validate_marts,
            build_db_first_payload,
        )

        integrity_refresh_run = (
            refresh_run
            if wb_daily_facts is not None
            else base_refresh_run or refresh_run
        )
        if wb_finance_dir is not None:
            _reverify_collection_raw_integrity(
                integrity_refresh_run,
                source_type="wb_finance_detail",
                raw_path=wb_finance_dir,
                source_root=self.settings.source_refresh_root_path,
            )
        if wb_report_list_dir is not None:
            _reverify_collection_raw_integrity(
                integrity_refresh_run,
                source_type="wb_sales_report_list",
                raw_path=wb_report_list_dir,
                source_root=self.settings.source_refresh_root_path,
            )

        output_dir = (
            self.settings.export_root_path / "source_refresh" / refresh_run.id
        ).resolve()
        allowed = self.settings.export_root_path.resolve()
        if output_dir != allowed and allowed not in output_dir.parents:
            raise ValueError("source-refresh artifact path is outside reports")
        output_dir.mkdir(parents=True, exist_ok=True)
        excel_path = output_dir / "shumeyko_wb_excel_mvp.xlsx"
        client_name = self._client_name(db, refresh_run.tenant_id)
        sku_mappings = self._calculation_sku_mappings(db, refresh_run)
        args = argparse.Namespace(
            client_id=(
                refresh_run.client_id
                or (
                    source_report.client_id
                    if source_report
                    else repository.client_id_for_tenant(refresh_run.tenant_id)
                )
            ),
            wb_finance_dir=wb_finance_dir,
            wb_finance_source=(
                "daily-facts" if wb_daily_facts is not None else "files-stream"
            ),
            wb_daily_facts=wb_daily_facts,
            wb_sales_report_summary_rows=wb_summary_rows,
            stream_cache_dir=Path("data/.cache/source_refresh_stream") / refresh_run.id,
            keep_stream_cache=False,
            marketplace_daily_facts_enabled=(
                self.settings.marketplace_daily_facts_enabled and wb_daily_facts is None
            ),
            postgres_db_name="shumeyko_wb_unit_economics",
            postgres_host="",
            postgres_port=55433,
            postgres_user="",
            postgres_snapshot_id=None,
            mapping_source="files",
            mapping_snapshot_id=None,
            cost_source="files",
            cost_snapshot_id=None,
            wb_cards_dir=wb_cards_dir,
            onec_dir=onec_dir,
            onec_marketplace_mapping_dir=self.settings.source_refresh_mapping_path,
            sales_register_dir=onec_dir,
            onec_services_dir=onec_dir,
            wb_report_list_dir=wb_report_list_dir,
            wb_paid_storage_dir=None,
            wb_promotion_stats_dir=None,
            wb_stock_history_dir=wb_stock_history_dir,
            onec_stock_dir=onec_dir,
            onec_opiu_dir=onec_dir,
            onec_opiu_config=None,
            report_period_start=refresh_run.period_start,
            report_period_end=refresh_run.period_end,
            cost_amount_field="Сумма",
            sales_cost_amount_field="auto",
            source_refresh_run_id=refresh_run.id,
            tenant_name=client_name,
            sku_mappings=sku_mappings,
        )
        tax_profiles = repository.tax_profiles_for_source_refresh(
            db,
            refresh_run,
        )
        input_vat_policies = repository.input_vat_policies_for_source_refresh(
            db,
            refresh_run,
        )
        tax_collection = next(
            (
                item
                for item in refresh_run.collections
                if item.source_type == "onec_tax_profiles"
            ),
            None,
        )
        expected_tax_profiles = (
            int((tax_collection.payload or {}).get("profileCount") or 0)
            if tax_collection is not None
            else 0
        )
        if expected_tax_profiles and len(tax_profiles) < expected_tax_profiles:
            raise SourceRefreshConfigError(
                "confirmed 1C tax profiles were not resolved for report rebuild"
            )
        db.commit()
        build = build_db_first_payload(
            args,
            tax_profiles=tax_profiles,
            input_vat_policies=input_vat_policies,
        )
        self._save_onec_cost_snapshots(db, refresh_run, build)
        if self.settings.marketplace_daily_facts_enabled and wb_daily_facts is None:
            self._save_wb_daily_facts(
                db,
                refresh_run,
                build,
                replacement_summary_rows=wb_summary_rows,
            )
        report = repository.save_report_marts(
            db,
            build["payload"],
            tenant_id=refresh_run.tenant_id,
            tenant_name=self._tenant_name(db, refresh_run.tenant_id),
            report_id=self._new_report_id(source_report, refresh_run),
            publication_status="draft",
            publish=False,
            source_snapshot_set_id=source_snapshot_set_id,
        )
        repository.replace_source_loads_from_refresh(
            db,
            report,
            refresh_run,
            base_refresh_run=base_refresh_run,
            contributing_runs=contributing_runs,
        )
        if self.settings.logistics_analysis_enabled:
            report.logistics_analysis_required = True
            logistics_result = _build_and_persist_logistics_analysis(
                db,
                report,
                primary_refresh_run=refresh_run,
                base_refresh_run=base_refresh_run,
                contributing_runs=contributing_runs,
            )
            if self.settings.logistics_factors_enabled:
                _build_and_persist_logistics_dimensions(
                    db,
                    report,
                    logistics_result=logistics_result,
                    primary_refresh_run=refresh_run,
                    base_refresh_run=base_refresh_run,
                    contributing_runs=contributing_runs,
                )
                if self.settings.logistics_measurements_enabled:
                    _build_and_persist_logistics_measurements(
                        db,
                        report,
                        logistics_result=logistics_result,
                        primary_refresh_run=refresh_run,
                        base_refresh_run=base_refresh_run,
                        contributing_runs=contributing_runs,
                    )
                if self.settings.logistics_tariffs_enabled:
                    _build_and_persist_logistics_tariffs(
                        db,
                        report,
                        logistics_result=logistics_result,
                        primary_refresh_run=refresh_run,
                        base_refresh_run=base_refresh_run,
                        contributing_runs=contributing_runs,
                    )
                if self.settings.logistics_routes_enabled:
                    _build_and_persist_logistics_routes(
                        db,
                        report,
                        logistics_result=logistics_result,
                        primary_refresh_run=refresh_run,
                        base_refresh_run=base_refresh_run,
                        contributing_runs=contributing_runs,
                    )
                if self.settings.logistics_return_reasons_enabled:
                    _build_and_persist_logistics_return_reasons(
                        db,
                        report,
                        logistics_result=logistics_result,
                        primary_refresh_run=refresh_run,
                        base_refresh_run=base_refresh_run,
                        contributing_runs=contributing_runs,
                    )
        _validate_marts(build["payload"])
        db.commit()
        return report, excel_path

    def _export_db_first_report_excel(
        self,
        db: Session,
        refresh_run: SourceRefreshRun,
        report: ReportRun,
        *,
        excel_path: Path,
    ) -> None:
        from scripts.export_report_artifacts import export_report_artifacts

        output_dir = excel_path.parent
        artifact_payload = repository.report_full_payload(db, report)
        artifact_payload.pop("unitRows", None)
        repository.transition_source_refresh_stage(
            db,
            refresh_run,
            stage="export_excel",
        )
        db.commit()
        records = export_report_artifacts(
            artifact_payload,
            report_id=report.id,
            output_dir=output_dir,
            excel_path=excel_path,
            excel=True,
            docx=False,
            pdf=False,
            html=False,
            csv=False,
            unit_rows_factory=lambda: repository.iter_report_unit_row_payloads(
                db,
                report,
                page_size=1_000,
            ),
        )
        for artifact_type, record in records:
            repository.record_report_artifact(
                db,
                report,
                artifact_type=artifact_type,
                path=record["path"],
                sha256=record["hash"],
                byte_size=record["byte_size"],
                status=record["status"],
            )
        db.flush()

    def _attach_source_loads(
        self,
        db: Session,
        report: ReportRun,
        refresh_run: SourceRefreshRun,
        *,
        contributing_runs: Iterable[SourceRefreshRun] = (),
    ) -> None:
        base_refresh_run = (
            db.get(SourceRefreshRun, refresh_run.base_source_refresh_run_id)
            if refresh_run.base_source_refresh_run_id
            else None
        )
        repository.replace_source_loads_from_refresh(
            db,
            report,
            refresh_run,
            base_refresh_run=base_refresh_run,
            contributing_runs=contributing_runs,
        )

    def _required_collection_raw_dir(
        self,
        refresh_run: SourceRefreshRun,
        source_type: str,
    ) -> Path:
        path = self._optional_collection_raw_dir(refresh_run, source_type)
        if path is None:
            raise SourceRefreshConfigError(
                f"reusable WB snapshot has no readable {source_type} directory"
            )
        return path

    def _optional_collection_raw_dir(
        self,
        refresh_run: SourceRefreshRun,
        source_type: str,
    ) -> Path | None:
        collection = next(
            (
                item
                for item in refresh_run.collections
                if item.source_type == source_type
                and item.status in MANDATORY_OK_STATUSES
            ),
            None,
        )
        return self._collection_raw_dir(collection) if collection else None

    def _report_snapshot_set_id(
        self,
        refresh_run: SourceRefreshRun,
        *,
        base_refresh_run: SourceRefreshRun | None,
        contributing_runs: Iterable[SourceRefreshRun] = (),
    ) -> str:
        contributors = sorted(
            {
                item
                for item in contributing_runs
                if item.id != refresh_run.id
                and item.id != getattr(base_refresh_run, "id", None)
            },
            key=lambda item: (item.created_at, item.id),
        )
        if base_refresh_run is None and not contributors:
            return refresh_run.snapshot_set_id
        lineage = [f"methodology:{METHODOLOGY_VERSION}"]

        def add_run(role: str, run: SourceRefreshRun) -> None:
            lineage.append(f"{role}:run:{run.snapshot_set_id}")
            lineage.extend(
                f"{role}:source:{item.source_type}:{item.snapshot_hash}"
                for item in sorted(
                    run.collections,
                    key=lambda collection: (
                        collection.source_type,
                        collection.wb_cabinet_id,
                        collection.id,
                    ),
                )
                if item.status in MANDATORY_OK_STATUSES | REVIEW_STATUSES
                and item.snapshot_hash
            )

        if base_refresh_run is not None:
            add_run("base", base_refresh_run)
        for contributor in contributors:
            add_run("overlay", contributor)
        add_run("current", refresh_run)
        digest = hashlib.sha256("\n".join(lineage).encode()).hexdigest()[:20]
        return f"composite-{digest}"

    def _mandatory_failed(self, refresh_run: SourceRefreshRun) -> bool:
        return any(
            item.required
            and (
                item.status not in MANDATORY_OK_STATUSES | REVIEW_STATUSES
                or (
                    (
                        item.source_type == "wb_finance_detail"
                        or item.source_type.startswith("onec_")
                    )
                    and item.status == "partial_source"
                )
            )
            for item in refresh_run.collections
        )

    def _mandatory_failure_is_transient(
        self,
        refresh_run: SourceRefreshRun,
    ) -> bool:
        failed = [
            item
            for item in refresh_run.collections
            if item.required
            and (
                item.status not in MANDATORY_OK_STATUSES | REVIEW_STATUSES
                or (
                    (
                        item.source_type == "wb_finance_detail"
                        or item.source_type.startswith("onec_")
                    )
                    and item.status == "partial_source"
                )
            )
        ]
        return bool(failed) and all(
            _collection_failure_is_transient(item) for item in failed
        )

    def _needs_review(
        self,
        refresh_run: SourceRefreshRun,
        mapping_collection: SourceRefreshCollection,
        *,
        mapping_report_ready: bool = False,
    ) -> bool:
        if mapping_collection.status == "stale":
            return True
        return any(
            (
                (not item.required and item.status not in OPTIONAL_OK_STATUSES)
                or (item.required and item.status in REVIEW_STATUSES)
            )
            for item in refresh_run.collections
            if not (mapping_report_ready and item.id == mapping_collection.id)
        )

    def _resolve_resume_run(
        self,
        db: Session,
        *,
        tenant_id: str,
        client_id: str,
        mode: str,
        credential_source: str,
        period_start: date,
        period_end: date,
        resume_mode: str,
        resume_from_run_id: str | None,
        dry_run: bool,
    ) -> SourceRefreshRun | None:
        if dry_run or resume_mode == "never":
            return None
        if mode not in ONEC_RESUME_MODES:
            if resume_from_run_id:
                raise SourceRefreshConfigError(
                    "explicit resume is available only for modes with 1C OData"
                )
            return None
        checkpoint_available = (
            self._run_has_onec_checkpoint
            if mode == "ozon-only"
            else self._run_has_resume_checkpoint
        )

        candidate: SourceRefreshRun | None
        if resume_from_run_id:
            candidate = db.get(SourceRefreshRun, resume_from_run_id)
            if candidate is None:
                raise SourceRefreshConfigError(
                    f"resume source refresh not found: {resume_from_run_id}"
                )
        else:
            candidates = list(
                db.scalars(
                    select(SourceRefreshRun)
                    .where(
                        SourceRefreshRun.tenant_id == tenant_id,
                        SourceRefreshRun.client_id == client_id,
                        SourceRefreshRun.credential_source == credential_source,
                        SourceRefreshRun.period_start == period_start,
                        SourceRefreshRun.period_end == period_end,
                        SourceRefreshRun.mode.in_(ONEC_RESUME_MODES),
                        SourceRefreshRun.finished_at.is_not(None),
                        SourceRefreshRun.status.in_({"failed", "needs_review"}),
                    )
                    .order_by(SourceRefreshRun.created_at.desc())
                    .limit(20)
                )
            )
            candidate = next(
                (item for item in candidates if checkpoint_available(item)),
                None,
            )
            if candidate is None:
                return None

        compatible = (
            candidate.tenant_id == tenant_id
            and candidate.client_id == client_id
            and candidate.credential_source == credential_source
            and candidate.period_start == period_start
            and candidate.period_end == period_end
            and candidate.mode in ONEC_RESUME_MODES
            and candidate.finished_at is not None
            and checkpoint_available(candidate)
        )
        root_dir = Path(candidate.root_dir).resolve() if candidate.root_dir else None
        allowed_root = self.settings.source_refresh_root_path.resolve()
        has_safe_resume_dir = bool(
            root_dir
            and root_dir.is_relative_to(allowed_root)
            and (
                (root_dir / "onec").is_dir()
                if mode == "ozon-only"
                else (
                    (root_dir / "onec").is_dir()
                    or (root_dir / "wb_finance" / "manifest.json").is_file()
                )
            )
        )
        if not compatible or not has_safe_resume_dir:
            if resume_from_run_id:
                raise SourceRefreshConfigError(
                    "resume source refresh is incompatible or has no safe checkpoint"
                )
            return None
        return candidate

    def _run_has_resume_checkpoint(self, refresh_run: SourceRefreshRun) -> bool:
        if self._run_has_onec_checkpoint(refresh_run):
            return True
        if not refresh_run.root_dir:
            return False
        root_dir = Path(refresh_run.root_dir).resolve()
        allowed_root = self.settings.source_refresh_root_path.resolve()
        return bool(
            root_dir.is_relative_to(allowed_root)
            and (root_dir / "wb_finance" / "manifest.json").is_file()
        )

    def _run_has_onec_checkpoint(self, refresh_run: SourceRefreshRun) -> bool:
        if not refresh_run.root_dir:
            return False
        root_dir = Path(refresh_run.root_dir).resolve()
        allowed_root = self.settings.source_refresh_root_path.resolve()
        onec_dir = root_dir / "onec"
        if not root_dir.is_relative_to(allowed_root) or not onec_dir.is_dir():
            return False
        return any(path.is_file() for path in onec_dir.glob("*/manifest.json"))

    def default_period_for_mode(self, mode: str) -> tuple[date, date]:
        return default_period_for_mode(self.settings, mode)

    def _snapshot_set_id(self, mode: str) -> str:
        stamp = datetime.now(tz=MOSCOW_TZ).strftime("%Y%m%d-%H%M%S")
        return f"{mode}-{stamp}"

    def _new_report_id(
        self,
        source_report: ReportRun | None,
        refresh_run: SourceRefreshRun,
    ) -> str:
        prefix = source_report.id if source_report else refresh_run.tenant_id
        stamp = security.utcnow().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_source_refresh_{stamp}"

    def _tenant_name(self, db: Session, tenant_id: str) -> str:
        tenant = db.get(repository.Tenant, tenant_id)
        return tenant.name if tenant is not None else tenant_id

    def _client_name(self, db: Session, tenant_id: str) -> str:
        client = db.scalar(
            select(repository.Client).where(repository.Client.tenant_id == tenant_id)
        )
        if client is not None and client.name:
            return client.name
        return self._tenant_name(db, tenant_id)

    def _onec_page_size(self) -> int:
        return max(1, min(int(self.settings.source_refresh_onec_page_size), 100000))

    def _onec_max_pages(self) -> int:
        return max(1, min(int(self.settings.source_refresh_onec_max_pages), 10000))

    def _wb_limit(self) -> int:
        return max(1, min(int(self.settings.source_refresh_wb_limit), 100000))

    def _wb_max_pages(self) -> int:
        return max(1, min(int(self.settings.source_refresh_wb_max_pages), 10000))

    def _wb_delay_seconds(self) -> float:
        return max(0.0, float(self.settings.source_refresh_wb_request_delay_seconds))

    def _wb_content_delay_seconds(self) -> float:
        return max(
            0.0,
            float(self.settings.source_refresh_wb_content_request_delay_seconds),
        )

    def _ozon_page_size(self) -> int:
        return max(1, min(int(self.settings.source_refresh_ozon_page_size), 1000))

    def _ozon_max_pages(self) -> int:
        return max(1, min(int(self.settings.source_refresh_ozon_max_pages), 10000))

    def _ozon_delay_seconds(self) -> float:
        return max(0.0, float(self.settings.source_refresh_ozon_request_delay_seconds))

    def _ozon_report_poll_timeout_seconds(self) -> float:
        return max(
            0.001,
            float(self.settings.source_refresh_ozon_report_poll_timeout_seconds),
        )

    def _ozon_report_poll_interval_seconds(self) -> float:
        return max(
            0.001,
            float(self.settings.source_refresh_ozon_report_poll_interval_seconds),
        )

    def _mapping_stale_days(self) -> int:
        return max(1, int(self.settings.source_refresh_mapping_stale_days))


def _collect_mapping_source(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    collection = service._record_mapping_source(context.db, context.refresh_run)
    return CollectorResult(collection=collection)


def _collect_wb_finance(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_finance"
    source_coverage_start = context.period_start - timedelta(
        days=context.period_start.weekday()
    )
    resume_dir = _safe_resume_subdirectory(service, context, "wb_finance")
    if resume_dir is not None and (resume_dir / "manifest.json").is_file():
        _copy_resume_directory(resume_dir, output_dir)
        results = load_wb_finance_export_results(output_dir)
        if not wb_finance_export_is_complete(results, context.credentials.wb_settings):
            resume_wb_finance_export(
                context.credentials.wb_settings,
                output_dir,
                max_pages=service._wb_max_pages(),
                request_delay_seconds=service._wb_delay_seconds(),
            )
            results = load_wb_finance_export_results(output_dir)
    else:
        results = service._wb_finance_exporter(
            context.credentials.wb_settings,
            output_dir,
            period_start=source_coverage_start,
            period_end=context.period_end,
            limit=service._wb_limit(),
            max_pages=service._wb_max_pages(),
            request_delay_seconds=service._wb_delay_seconds(),
            period="daily",
        )
    service._record_wb_finance(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
        source_coverage_start=source_coverage_start,
        source_coverage_end=context.period_end,
    )
    return CollectorResult(output_dir=output_dir)


def _safe_resume_subdirectory(
    service: SourceRefreshService,
    context: CollectorContext,
    name: str,
) -> Path | None:
    if not context.refresh_run.resumed_from_run_id:
        return None
    previous = context.db.get(SourceRefreshRun, context.refresh_run.resumed_from_run_id)
    if previous is None or not previous.root_dir:
        return None
    candidate = Path(previous.root_dir).resolve() / name
    allowed_root = service.settings.source_refresh_root_path.resolve()
    if not candidate.is_relative_to(allowed_root) or not candidate.is_dir():
        return None
    return candidate


def _copy_resume_directory(source: Path, destination: Path) -> None:
    """Copy an immutable checkpoint tree, preferring hard links."""
    destination.mkdir(parents=True, exist_ok=True)
    for source_path in source.rglob("*"):
        if source_path.is_symlink():
            continue
        destination_path = destination / source_path.relative_to(source)
        if source_path.is_dir():
            destination_path.mkdir(parents=True, exist_ok=True)
            continue
        if not source_path.is_file():
            continue
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        if destination_path.exists():
            destination_path.unlink()
        try:
            os.link(source_path, destination_path)
        except OSError:
            shutil.copy2(source_path, destination_path)


def _collect_wb_report_list(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_sales_report_list"
    results = service._wb_report_list_exporter(
        context.credentials.wb_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        limit=1000,
        max_pages=10,
        request_delay_seconds=service._wb_delay_seconds(),
        period="weekly",
    )
    service._record_wb_report_list(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_wb_redeem_notifications(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_redeem_notifications"
    resume_dir = _safe_resume_subdirectory(
        service,
        context,
        "wb_redeem_notifications",
    )
    if resume_dir is not None and (resume_dir / "manifest.json").is_file():
        _copy_resume_directory(resume_dir, output_dir)
        results = load_wb_document_export_results(output_dir)
    else:
        results = service._wb_documents_exporter(
            settings=context.credentials.wb_settings,
            output_dir=output_dir,
            period_start=context.period_start,
            period_end=context.period_end,
            category_keywords=("выкуп", "redeem-notification"),
            download=True,
        )
    collection = service._record_wb_redeem_notifications(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_wb_product_cards(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_product_cards"
    content_settings = WbContentSettings(
        accounts=tuple(
            WbContentSellerAccount(
                seller_account_id=item.seller_account_id,
                account_name=item.account_name,
                api_key=item.api_key,
            )
            for item in context.credentials.wb_settings.accounts
        ),
        timeout_seconds=30.0,
    )
    results = service._wb_product_cards_exporter(
        content_settings,
        output_dir,
        limit=min(service._wb_limit(), 100),
        max_pages=service._wb_max_pages(),
        request_delay_seconds=service._wb_content_delay_seconds(),
    )
    service._record_wb_product_cards(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_wb_stock_history(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_stock_history_daily"
    provider_period = _stock_history_provider_period(
        context.period_start,
        context.period_end,
    )
    if provider_period is None:
        output_dir.mkdir(parents=True, exist_ok=True)
        results = [
            WbStockExportResult(
                seller_account_id=item.seller_account_id,
                account_name=item.account_name,
                source="wb_stock_history_daily_csv",
                ok=False,
                status="outside_provider_window",
                row_count=0,
                error="Requested period is outside the WB three-month history window.",
            )
            for item in context.credentials.wb_settings.accounts
        ]
        collection = service._record_wb_stock_history(
            context.db,
            context.refresh_run,
            output_dir,
            results,
            wb_cabinet_ids=context.credentials.wb_cabinet_ids,
            period_start=context.period_start,
            period_end=context.period_end,
            actual_period_start=None,
            actual_period_end=None,
            provider_window_available=False,
        )
        return CollectorResult(collection=collection, output_dir=output_dir)
    provider_period_start, provider_period_end = provider_period
    results = service._wb_stock_history_exporter(
        context.credentials.wb_settings,
        output_dir,
        period_start=provider_period_start,
        period_end=provider_period_end,
        stock_type="wb",
    )
    collection = service._record_wb_stock_history(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
        period_start=context.period_start,
        period_end=context.period_end,
        actual_period_start=provider_period_start,
        actual_period_end=provider_period_end,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_wb_tariffs(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_tariffs"
    results = service._wb_tariffs_exporter(
        context.credentials.wb_settings.accounts,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    collection = service._record_wb_tariffs(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_wb_goods_return(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_goods_return"
    results = service._wb_goods_return_exporter(
        context.credentials.wb_settings.accounts,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    collection = service._record_wb_goods_return(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_wb_return_claims(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_return_claims"
    results = service._wb_return_claims_exporter(
        context.credentials.wb_settings.accounts,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    collection = service._record_wb_return_claims(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_wb_supplier_sales(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "wb_supplier_sales"
    results = service._wb_supplier_sales_exporter(
        context.credentials.wb_settings.accounts,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    collection = service._record_wb_supplier_sales(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_wb_measurement_penalties(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    return _collect_wb_measurement_source(
        service,
        context,
        source_type="wb_measurement_penalties",
        exporter=service._wb_measurement_penalties_exporter,
    )


def _collect_wb_warehouse_measurements(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    return _collect_wb_measurement_source(
        service,
        context,
        source_type="wb_warehouse_measurements",
        exporter=service._wb_warehouse_measurements_exporter,
    )


def _collect_wb_measurement_source(
    service: SourceRefreshService,
    context: CollectorContext,
    *,
    source_type: str,
    exporter: Callable[..., list[WbMeasurementExportResult]],
) -> CollectorResult:
    if context.credentials.wb_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / source_type
    results = exporter(
        context.credentials.wb_settings.accounts,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    collection = service._record_wb_measurements(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type=source_type,
        wb_cabinet_ids=context.credentials.wb_cabinet_ids,
        period_start=context.period_start,
        period_end=context.period_end,
    )
    return CollectorResult(collection=collection, output_dir=output_dir)


def _collect_ozon_cash_flow(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_finance_cash_flow"
    results = service._ozon_cash_flow_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        page_size=service._ozon_page_size(),
        max_pages=service._ozon_max_pages(),
        request_delay_seconds=service._ozon_delay_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_finance_cash_flow",
        source_label="Ozon financial cash-flow statement",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
        required=context.mode == "ozon-only",
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_realization(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_realization"
    results = service._ozon_realization_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        request_delay_seconds=service._ozon_delay_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_realization",
        source_label="Ozon realization report",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_mutual_settlement(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_mutual_settlement"
    results = service._ozon_mutual_settlement_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        request_delay_seconds=service._ozon_delay_seconds(),
        report_poll_timeout_seconds=service._ozon_report_poll_timeout_seconds(),
        report_poll_interval_seconds=service._ozon_report_poll_interval_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_mutual_settlement",
        source_label="Ozon mutual settlement report",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_realization_posting(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_realization_posting"
    results = service._ozon_realization_posting_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        max_pages=service._ozon_max_pages(),
        request_delay_seconds=service._ozon_delay_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_realization_posting",
        source_label="Ozon realization posting report",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_products_buyout(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_products_buyout"
    results = service._ozon_products_buyout_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        request_delay_seconds=service._ozon_delay_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_products_buyout",
        source_label="Ozon products buyout report",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_b2b_sales(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_b2b_sales_json"
    results = service._ozon_b2b_sales_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        request_delay_seconds=service._ozon_delay_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_b2b_sales_json",
        source_label="Ozon B2B sales JSON",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_products(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_products_report"
    results = service._ozon_products_exporter(
        context.credentials.ozon_settings,
        output_dir,
        request_delay_seconds=service._ozon_delay_seconds(),
        report_poll_timeout_seconds=service._ozon_report_poll_timeout_seconds(),
        report_poll_interval_seconds=service._ozon_report_poll_interval_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_products_report",
        source_label="Ozon products report",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_stocks(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_stock_on_warehouses"
    results = service._ozon_stocks_exporter(
        context.credentials.ozon_settings,
        output_dir,
        limit=service._ozon_page_size(),
        max_pages=service._ozon_max_pages(),
        request_delay_seconds=service._ozon_delay_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_stock_on_warehouses",
        source_label="Ozon stock on warehouses",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_ozon_returns(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.ozon_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "ozon_returns_report"
    results = service._ozon_returns_exporter(
        context.credentials.ozon_settings,
        output_dir,
        period_start=context.period_start,
        period_end=context.period_end,
        request_delay_seconds=service._ozon_delay_seconds(),
        report_poll_timeout_seconds=service._ozon_report_poll_timeout_seconds(),
        report_poll_interval_seconds=service._ozon_report_poll_interval_seconds(),
    )
    service._record_ozon_results(
        context.db,
        context.refresh_run,
        output_dir,
        results,
        source_type="ozon_returns_report",
        source_label="Ozon returns report",
        ozon_cabinet_ids=context.credentials.ozon_cabinet_ids,
    )
    return CollectorResult(output_dir=output_dir)


def _collect_onec_odata(
    service: SourceRefreshService,
    context: CollectorContext,
) -> CollectorResult:
    if context.credentials.onec_settings is None:
        return CollectorResult()
    output_dir = context.root_dir / "onec"
    resume_from_dir: Path | None = None
    if context.refresh_run.resumed_from_run_id:
        resumed_from = context.db.get(
            SourceRefreshRun, context.refresh_run.resumed_from_run_id
        )
        if resumed_from is not None and resumed_from.root_dir:
            candidate = Path(resumed_from.root_dir).resolve() / "onec"
            allowed_root = service.settings.source_refresh_root_path.resolve()
            if candidate.is_relative_to(allowed_root) and candidate.is_dir():
                resume_from_dir = candidate
    source_identity = hashlib.sha256(
        context.credentials.onec_settings.base_url.encode("utf-8")
    ).hexdigest()
    _commit_source_refresh_progress(context.db)
    results = service._onec_exporter(
        context.credentials.onec_settings,
        ONEC_REFRESH_COLLECTIONS,
        output_dir,
        top=service._onec_page_size(),
        max_pages=service._onec_max_pages(),
        period_start=context.period_start,
        period_end=context.period_end,
        resume_from_dir=resume_from_dir,
        source_identity=source_identity,
    )
    service._record_onec(context.db, context.refresh_run, output_dir, results)
    return CollectorResult(output_dir=output_dir)


def _existing_path_for_disk_check(path: Path) -> Path:
    current = path.resolve()
    while not current.exists() and current.parent != current:
        current = current.parent
    return current


def wb_finance_settings_from_secret(
    secret: str,
    *,
    default_name: str = "Wildberries API",
    default_seller_account_id: str = "WB_ACCOUNT",
) -> WbFinanceSettings:
    raw = secret.strip()
    if not raw:
        raise integrations.IntegrationSecretError("wb_secret_empty")
    accounts_payload: list[dict[str, Any]]
    if raw.startswith("{") or raw.startswith("["):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise integrations.IntegrationSecretError("wb_secret_json_invalid") from exc
        if isinstance(parsed, list):
            accounts_payload = [item for item in parsed if isinstance(item, dict)]
        elif isinstance(parsed, dict):
            nested = parsed.get("accounts")
            if isinstance(nested, list):
                accounts_payload = [item for item in nested if isinstance(item, dict)]
            else:
                accounts_payload = _accounts_from_wb_env_style_object(parsed)
                if not accounts_payload:
                    accounts_payload = [parsed]
        else:
            raise integrations.IntegrationSecretError("wb_secret_json_must_be_object")
    else:
        accounts_payload = [{"apiKey": raw, "accountName": default_name}]

    accounts: list[WbFinanceSellerAccount] = []
    fallback_id = _safe_wb_account_id(default_seller_account_id)
    for index, item in enumerate(accounts_payload, start=1):
        api_key = _first_text(
            item,
            "apiKey",
            "api_key",
            "token",
            "authorization",
            f"WB_ACCOUNT_{index}_API_KEY",
        )
        if not api_key:
            continue
        seller_account_id = _first_text(
            item,
            "sellerAccountId",
            "seller_account_id",
            "id",
            f"WB_ACCOUNT_{index}_ID",
        ) or (fallback_id if len(accounts_payload) == 1 else f"{fallback_id}_{index}")
        account_name = (
            _first_text(
                item,
                "accountName",
                "account_name",
                "name",
                f"WB_ACCOUNT_{index}_NAME",
            )
            or seller_account_id
        )
        accounts.append(
            WbFinanceSellerAccount(
                seller_account_id=seller_account_id,
                account_name=account_name,
                api_key=api_key,
            )
        )
    if not accounts:
        raise integrations.IntegrationSecretError("wb_secret_missing_api_key")
    return WbFinanceSettings(accounts=tuple(accounts))


def inspect_mapping_source(
    mapping_dir: Path,
    *,
    stale_after_days: int,
) -> tuple[str, str, int, str, dict[str, Any]]:
    if not mapping_dir.exists():
        return (
            "failed",
            "",
            0,
            "mapping_source_missing",
            {"mappingDir": str(mapping_dir)},
        )
    files = sorted(path for path in mapping_dir.rglob("*") if path.is_file())
    if not files:
        return (
            "failed",
            "",
            0,
            "mapping_source_empty",
            {"mappingDir": str(mapping_dir)},
        )
    digest = hashlib.sha256()
    newest_mtime = 0.0
    for path in files:
        stat = path.stat()
        newest_mtime = max(newest_mtime, stat.st_mtime)
        digest.update(str(path.relative_to(mapping_dir)).encode("utf-8"))
        digest.update(str(stat.st_size).encode("ascii"))
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
    newest_at = datetime.fromtimestamp(newest_mtime, tz=MOSCOW_TZ)
    age_days = (datetime.now(tz=MOSCOW_TZ) - newest_at).days
    status = "stale" if age_days > stale_after_days else "loaded"
    return (
        status,
        digest.hexdigest(),
        len(files),
        "",
        {
            "mappingDir": str(mapping_dir),
            "fileCount": len(files),
            "newestMtime": newest_at.isoformat(),
            "ageDays": age_days,
            "staleAfterDays": stale_after_days,
        },
    )


def _tenant_integration(
    db: Session,
    tenant_id: str,
    provider: str,
) -> TenantIntegration | None:
    return db.scalar(
        select(TenantIntegration).where(
            TenantIntegration.tenant_id == tenant_id,
            TenantIntegration.provider == provider,
        )
    )


def _tenant_integrations_by_base(
    db: Session,
    tenant_id: str,
    provider_base: str,
) -> list[TenantIntegration]:
    return [
        item
        for item in db.scalars(
            select(TenantIntegration)
            .where(TenantIntegration.tenant_id == tenant_id)
            .order_by(TenantIntegration.provider)
        )
        if repository.integration_provider_base(item.provider) == provider_base
    ]


def _integration_account_name(integration: TenantIntegration) -> str:
    payload = integration.config_payload or {}
    return (
        str(payload.get("cabinetName") or "").strip()
        or str(payload.get("organizationName") or "").strip()
        or integration.label
        or integration.provider
    )


def _ensure_wb_cabinet_for_integration(
    db: Session,
    integration: TenantIntegration,
):
    payload = dict(integration.config_payload or {})
    account_name = _integration_account_name(integration)
    tenant = db.get(Tenant, integration.tenant_id)
    client = repository.ensure_client_for_tenant(
        db,
        tenant_id=integration.tenant_id,
        name=tenant.name if tenant else integration.tenant_id,
    )
    cabinet = db.scalar(
        select(WbCabinet)
        .where(
            WbCabinet.client_id == client.id,
            WbCabinet.display_name == account_name,
        )
        .order_by((WbCabinet.provider == "").desc(), WbCabinet.id)
    )
    if cabinet is not None:
        company_id = cabinet.client_company_id or str(
            payload.get("clientCompanyId") or ""
        )
        if integration.provider and not cabinet.provider:
            cabinet.provider = integration.provider
            cabinet.updated_at = security.utcnow()
    else:
        company = repository.ensure_client_company(
            db,
            tenant_id=integration.tenant_id,
            client_id=client.id,
            display_name=str(payload.get("organizationName") or "").strip(),
        )
        company_id = (
            company.id if company else str(payload.get("clientCompanyId") or "")
        )
        cabinet = repository.ensure_wb_cabinet(
            db,
            tenant_id=integration.tenant_id,
            client_id=client.id,
            display_name=account_name,
            cabinet_key=str(payload.get("connectionKey") or integration.provider),
            provider=integration.provider,
            client_company_id=company_id,
        )
    if cabinet is None:
        return None
    if payload.get("clientId") != client.id or payload.get("wbCabinetId") != cabinet.id:
        payload["clientId"] = client.id
        payload["clientCompanyId"] = company_id
        payload["wbCabinetId"] = cabinet.id
        integration.config_payload = payload
        integration.updated_at = security.utcnow()
        db.flush()
    return cabinet


def _safe_wb_account_id(value: str) -> str:
    normalized = "".join(
        char if char.isalnum() else "_" for char in value.strip().upper()
    ).strip("_")
    return normalized or "WB_ACCOUNT"


def _safe_ozon_account_id(value: str) -> str:
    normalized = "".join(
        char if char.isalnum() else "_" for char in value.strip().upper()
    ).strip("_")
    return normalized or "OZON_ACCOUNT"


def _env_has_ozon_credentials(env_file: Path = Path(".env")) -> bool:
    env_keys = set(os.environ)
    if any(
        key.startswith("OZON_ACCOUNT_") and key.endswith(("_CLIENT_ID", "_API_KEY"))
        for key in env_keys
    ):
        return True
    if not env_file.exists():
        return False
    try:
        for line in env_file.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if (
                stripped.startswith("OZON_ACCOUNT_")
                and ("_CLIENT_ID=" in stripped or "_API_KEY=" in stripped)
                and stripped.partition("=")[2].strip().strip('"').strip("'")
            ):
                return True
    except OSError:
        return False
    return False


def _accounts_from_wb_env_style_object(parsed: dict[str, Any]) -> list[dict[str, Any]]:
    accounts: list[dict[str, Any]] = []
    for index in range(1, 11):
        api_key = _first_text(parsed, f"WB_ACCOUNT_{index}_API_KEY")
        if not api_key:
            continue
        accounts.append(
            {
                "apiKey": api_key,
                "sellerAccountId": f"WB_ACCOUNT_{index}",
                "accountName": _first_text(parsed, f"WB_ACCOUNT_{index}_NAME"),
            }
        )
    return accounts


def _first_text(values: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = values.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _credential_issue(
    provider: str,
    required: bool,
    message: str,
    *,
    payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    provider_base = repository.integration_provider_base(provider)
    labels = {
        "wb_api": "Wildberries API",
        "ozon_api": "Ozon Seller API",
        "onec_readonly": "1С — только чтение",
    }
    return {
        "source_type": provider,
        "source_label": labels.get(provider_base, provider),
        "required": required,
        "status": "needs_configuration",
        "error_message": message,
        "payload": payload or {},
    }


def _wb_result_payload(
    item: WbFinancePageResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "pageIndex": item.page_index,
        "status": _wb_status(item),
        "sourceStatus": item.status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "rrdIdStart": item.rrd_id_start,
        "rrdIdNext": item.rrd_id_next,
        "statusCode": item.status_code,
        "rawPayloadHash": item.raw_payload_hash,
        "outputFile": item.output_path.name if item.output_path else None,
        "error": item.error,
    }


def _mark_wb_finance_max_pages_issue(
    results: list[WbFinancePageResult],
    payload_items: list[dict[str, Any]],
    *,
    max_pages: int,
) -> bool:
    exhausted = False
    for item, payload in zip(results, payload_items, strict=False):
        if not _wb_finance_has_next_page_after_limit(item, max_pages=max_pages):
            continue
        exhausted = True
        payload["status"] = "partial_source"
        payload["maxPagesReached"] = True
        payload["completenessIssue"] = WB_FINANCE_MAX_PAGES_ISSUE
    return exhausted


def _wb_finance_has_next_page_after_limit(
    item: WbFinancePageResult,
    *,
    max_pages: int,
) -> bool:
    return (
        item.ok
        and item.status == "ok"
        and item.page_index >= max_pages
        and item.rrd_id_next is not None
        and item.rrd_id_next != item.rrd_id_start
    )


def _wb_report_list_payload(
    item: WbSalesReportListPageResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "pageIndex": item.page_index,
        "status": _wb_status(item),
        "sourceStatus": item.status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "statusCode": item.status_code,
        "offset": item.offset,
        "rawPayloadHash": item.raw_payload_hash,
        "outputFile": item.output_path.name if item.output_path else None,
        "error": item.error,
    }


def _wb_document_result_payload(
    item: WbDocumentExportResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "status": _wb_document_status(item),
        "sourceStatus": item.status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "downloadedCount": item.downloaded_count,
        "statusCode": item.status_code,
        "outputFile": item.output_file or None,
        "error": item.error,
    }


def _wb_redeem_notification_rows(
    output_dir: Path,
    results: Iterable[WbDocumentExportResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in results:
        if not result.ok or not result.output_file:
            continue
        manifest_path = (output_dir / result.output_file).resolve()
        if not manifest_path.is_relative_to(output_dir.resolve()):
            continue
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except (OSError, TypeError, ValueError, json.JSONDecodeError):
            continue
        if not isinstance(manifest, list):
            continue
        for item in manifest:
            if not isinstance(item, dict):
                continue
            download = item.get("download")
            summary = download.get("summary") if isinstance(download, dict) else None
            if not isinstance(summary, dict) or summary.get("status") != "parsed":
                continue
            report_id = str(summary.get("reportId") or "").strip()
            if not report_id:
                continue
            rows.append(
                {
                    "reportId": report_id,
                    "documentName": str(item.get("name") or ""),
                    "category": str(item.get("category") or ""),
                    "creationTime": str(item.get("creationTime") or ""),
                    "quantity": str(summary.get("quantity") or ""),
                    "purchaseAmount": str(summary.get("purchaseAmount") or ""),
                    "vatAmount": (
                        str(summary.get("vatAmount"))
                        if summary.get("vatAmount") is not None
                        else None
                    ),
                    "summaryStatus": "parsed",
                    "sellerAccountId": result.seller_account_id,
                    "accountName": result.account_name,
                    "wbCabinetId": wb_cabinet_ids.get(
                        result.seller_account_id,
                        "",
                    ),
                    "rawDocumentHash": str(download.get("sha256") or ""),
                }
            )
    return rows


def _wb_product_cards_payload(
    item: WbProductCardsPageResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "cardsSource": item.cards_source,
        "pageIndex": item.page_index,
        "status": _wb_product_cards_status(item),
        "ok": item.ok,
        "rowCount": item.flat_row_count,
        "cardCount": item.card_count,
        "statusCode": item.status_code,
        "rawPayloadHash": item.raw_payload_hash,
        "flatPayloadHash": item.flat_payload_hash,
        "outputFile": item.output_path.name if item.output_path else None,
        "flatOutputFile": item.flat_output_path.name if item.flat_output_path else None,
        "error": item.error,
    }


def _wb_tariffs_result_payload(
    item: WbTariffsExportResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    row_count = item.box_row_count + item.pallet_row_count
    if item.ok and row_count > 0:
        status = "loaded"
    elif item.ok:
        status = "empty_unexpected"
    elif item.status_code in {401, 403}:
        status = "auth_failed"
    elif item.status_code == 429:
        status = "rate_limited"
    else:
        status = "failed"
    target_date = item.target_date
    return {
        "sellerAccountId": item.seller_account_id,
        "wbCabinetId": wb_cabinet_id,
        "targetDate": target_date.isoformat() if target_date else None,
        "pageIndex": int(target_date.strftime("%Y%m%d")) if target_date else 0,
        "status": status,
        "ok": item.ok,
        "rowCount": row_count,
        "boxRowCount": item.box_row_count,
        "palletRowCount": item.pallet_row_count,
        "statusCode": item.status_code,
        "rawPayloadHash": item.raw_payload_hash,
        "flatPayloadHash": item.flat_payload_hash,
        "outputFile": (item.raw_output_path.name if item.raw_output_path else None),
        "flatOutputFile": (
            item.flat_output_path.name if item.flat_output_path else None
        ),
        "error": item.error,
    }


def _wb_goods_return_result_payload(
    item: WbGoodsReturnExportResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    if item.ok:
        status = "loaded" if item.row_count else "empty_expected"
    elif item.status_code in {401, 403}:
        status = "auth_failed"
    elif item.status_code == 429:
        status = "rate_limited"
    else:
        status = "failed"
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "pageIndex": 1,
        "status": status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "statusCode": item.status_code,
        "coverageStart": (
            item.coverage_start.isoformat() if item.coverage_start else ""
        ),
        "coverageEnd": item.coverage_end.isoformat() if item.coverage_end else "",
        "rawPayloadHash": item.raw_payload_hash,
        "flatPayloadHash": item.flat_payload_hash,
        "outputFile": (item.raw_output_path.name if item.raw_output_path else None),
        "flatOutputFile": (
            item.flat_output_path.name if item.flat_output_path else None
        ),
        "error": item.error,
    }


def _wb_return_claims_result_payload(
    item: WbReturnClaimsExportResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "pageIndex": 1,
        "status": item.source_state,
        "activeStatus": item.active_state,
        "archiveStatus": item.archive_state,
        "ok": item.ok,
        "rowCount": item.row_count,
        "statusCode": item.status_code,
        "coverageStart": (
            item.coverage_start.isoformat() if item.coverage_start else ""
        ),
        "coverageEnd": item.coverage_end.isoformat() if item.coverage_end else "",
        "rawPayloadHash": item.raw_payload_hash,
        "flatPayloadHash": item.flat_payload_hash,
        "outputFile": (item.raw_output_path.name if item.raw_output_path else None),
        "flatOutputFile": (
            item.flat_output_path.name if item.flat_output_path else None
        ),
        "error": item.error,
    }


def _wb_supplier_sales_result_payload(
    item: WbSupplierSalesExportResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    if item.ok:
        status = "loaded" if item.row_count else "empty_expected"
    elif item.status_code in {401, 403}:
        status = "auth_failed"
    elif item.status_code == 429:
        status = "rate_limited"
    elif item.error == "ProviderWindowUnavailable":
        status = "outside_provider_window"
    else:
        status = "failed"
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "pageIndex": 1,
        "status": status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "statusCode": item.status_code,
        "rawPayloadHash": item.raw_payload_hash,
        "flatPayloadHash": item.flat_payload_hash,
        "outputFile": (item.raw_output_path.name if item.raw_output_path else None),
        "flatOutputFile": (
            item.flat_output_path.name if item.flat_output_path else None
        ),
        "error": item.error,
    }


def _wb_measurement_result_payload(
    item: WbMeasurementExportResult,
    *,
    wb_cabinet_id: str = "",
) -> dict[str, Any]:
    if item.ok:
        status = "loaded" if item.row_count else "empty_expected"
    elif item.status_code in {401, 403}:
        status = "auth_failed"
    elif item.status_code == 429:
        status = "rate_limited"
    else:
        status = "failed"
    return {
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": wb_cabinet_id,
        "pageIndex": 1,
        "status": status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "providerTotal": item.provider_total,
        "statusCode": item.status_code,
        "rawPayloadHash": item.raw_payload_hash,
        "flatPayloadHash": item.flat_payload_hash,
        "outputFile": (item.raw_output_path.name if item.raw_output_path else None),
        "flatOutputFile": (
            item.flat_output_path.name if item.flat_output_path else None
        ),
        "error": item.error,
    }


def _ozon_result_payload(
    item: OzonPageResult,
    *,
    ozon_cabinet_id: str = "",
) -> dict[str, Any]:
    return {
        "marketplace": "ozon",
        "sourceType": item.source_type,
        "sellerAccountId": item.seller_account_id,
        "accountName": item.account_name,
        "wbCabinetId": ozon_cabinet_id,
        "pageIndex": item.page_index,
        "status": _ozon_status(item),
        "sourceStatus": item.status,
        "ok": item.ok,
        "rowCount": item.row_count,
        "statusCode": item.status_code,
        "sourceEndpoint": item.source_endpoint,
        "rawPayloadHash": item.raw_payload_hash,
        "rawContentSha256": item.raw_content_sha256,
        "outputFile": item.output_path.name if item.output_path else None,
        "reportCode": item.report_code,
        "reportStatus": item.report_status,
        "pageLimitExhausted": item.page_limit_exhausted,
        "error": item.error,
    }


def _marketplace_cabinet_ids(
    refresh_run: SourceRefreshRun,
    *,
    source_type: str,
) -> dict[str, str]:
    result: dict[str, str] = {}
    for collection in refresh_run.collections:
        if collection.source_type != source_type:
            continue
        payload = collection.payload or {}
        for item in payload.get("results", []):
            if not isinstance(item, dict):
                continue
            seller_account_id = str(item.get("sellerAccountId") or "").strip()
            cabinet_id = str(item.get("wbCabinetId") or "").strip()
            if seller_account_id:
                result[seller_account_id] = cabinet_id
    return result


def _wb_daily_fact_parity(
    build: dict[str, Any],
    daily_facts: list[Any],
) -> dict[str, Any]:
    report = build.get("report")
    if report is None:
        raise SourceRefreshConfigError("daily facts parity report is missing")
    checks = {
        "quantity": (
            sum((item.quantity for item in daily_facts), Decimal("0")),
            sum((item.quantity for item in report.rows), Decimal("0")),
        ),
        "netRevenue": (
            sum((item.net_revenue for item in daily_facts), Decimal("0")),
            sum((item.net_revenue for item in report.rows), Decimal("0")),
        ),
        "commission": (
            sum((item.wb_commission for item in daily_facts), Decimal("0")),
            sum((item.wb_commission for item in report.rows), Decimal("0")),
        ),
        "logistics": (
            sum((item.logistics for item in daily_facts), Decimal("0")),
            sum((item.logistics for item in report.rows), Decimal("0")),
        ),
        "storage": (
            sum((item.storage for item in daily_facts), Decimal("0")),
            sum((item.storage for item in report.rows), Decimal("0")),
        ),
        "acceptance": (
            sum((item.acceptance for item in daily_facts), Decimal("0")),
            sum((item.acceptance for item in report.rows), Decimal("0")),
        ),
        "promotion": (
            sum(
                (item.marketplace_promotion for item in daily_facts),
                Decimal("0"),
            ),
            sum((item.wb_promotion for item in report.rows), Decimal("0")),
        ),
        "penalties": (
            sum(
                (item.penalties_and_holdbacks for item in daily_facts),
                Decimal("0"),
            ),
            sum(
                (item.penalties_and_holdbacks for item in report.rows),
                Decimal("0"),
            ),
        ),
        "acquiring": (
            sum((item.acquiring for item in daily_facts), Decimal("0")),
            sum((item.acquiring for item in report.rows), Decimal("0")),
        ),
        "cogs": (
            sum((item.cogs for item in daily_facts), Decimal("0")),
            sum(
                (item.cogs_from_1c_with_extra_costs for item in report.rows),
                Decimal("0"),
            ),
        ),
        "vatInputMarketplace": (
            sum(
                (item.vat_input_from_marketplace for item in daily_facts),
                Decimal("0"),
            ),
            sum((item.vat_input_from_wb for item in report.rows), Decimal("0")),
        ),
        "vatInput1c": (
            sum((item.vat_input_from_1c for item in daily_facts), Decimal("0")),
            sum((item.vat_input_from_1c for item in report.rows), Decimal("0")),
        ),
    }
    differences = {key: left - right for key, (left, right) in checks.items()}
    mismatches = [key for key, difference in differences.items() if difference != 0]
    source_row_count = sum(item.source_row_count for item in daily_facts)
    if source_row_count != int(build.get("wb_rows") or 0):
        mismatches.append("sourceRowCount")
    if mismatches:
        cogs_detail = (
            ""
            if "cogs" not in mismatches
            else f" (cogsDifference={differences['cogs']})"
        )
        raise SourceRefreshConfigError(
            "daily facts parity mismatch: " + ",".join(sorted(mismatches)) + cogs_detail
        )
    return {
        "status": "aggregate_only",
        "sourceRowCount": source_row_count,
        "dailyFactRows": len(daily_facts),
        "reportRows": len(report.rows),
        "checks": sorted(checks),
        "differences": {
            key: str(value) for key, value in differences.items() if value != 0
        },
        "roundingTolerance": {"cogs": "0.00"},
    }


def _full_wb_calculation_parity(
    legacy_build: dict[str, Any],
    streamed_build: dict[str, Any],
    *,
    legacy_row_count: int,
    source_parity: dict[str, Any] | None = None,
) -> dict[str, Any]:
    sections = {
        "reportRowsAndTaxes": (
            legacy_build.get("report"),
            streamed_build.get("report"),
        ),
        "kpiAndReconciliation": (
            legacy_build.get("payload"),
            streamed_build.get("payload"),
        ),
        "generatedDailyFacts": (
            legacy_build.get("daily_facts", []),
            streamed_build.get("daily_facts", []),
        ),
    }
    mismatches: list[str] = []
    digests: dict[str, dict[str, str]] = {}
    for name, (legacy, streamed) in sections.items():
        legacy_value = _canonical_parity_value(legacy)
        streamed_value = _canonical_parity_value(streamed)
        legacy_digest = _hash_payload(legacy_value)
        streamed_digest = _hash_payload(streamed_value)
        digests[name] = {
            "legacy": legacy_digest,
            "streamed": streamed_digest,
        }
        if legacy_digest != streamed_digest:
            mismatches.append(name)
    legacy_source_rows = int(
        legacy_build.get("wb_source_rows", legacy_build.get("wb_rows")) or 0
    )
    streamed_source_rows = int(
        streamed_build.get("wb_source_rows", streamed_build.get("wb_rows")) or 0
    )
    legacy_report_period_rows = int(
        legacy_build.get("wb_report_period_rows", legacy_build.get("wb_rows")) or 0
    )
    streamed_report_period_rows = int(
        streamed_build.get(
            "wb_report_period_rows",
            streamed_build.get("wb_rows"),
        )
        or 0
    )
    if legacy_source_rows != streamed_source_rows:
        mismatches.append("calculationSourceRowCount")
    if legacy_report_period_rows != streamed_report_period_rows:
        mismatches.append("calculationReportPeriodRowCount")
    if source_parity is not None and source_parity.get("status") != "matched":
        mismatches.append("legacyDbVsFiles")
    return {
        "status": "matched" if not mismatches else "mismatch",
        "legacyRowCount": legacy_row_count,
        "legacySourceRows": legacy_source_rows,
        "streamedSourceRows": streamed_source_rows,
        "legacyReportPeriodRows": legacy_report_period_rows,
        "streamedReportPeriodRows": streamed_report_period_rows,
        "sourceParity": source_parity or {"status": "not_run"},
        "digests": digests,
        "mismatches": sorted(set(mismatches)),
    }


def _legacy_wb_source_parity(
    db: Session,
    refresh_run: SourceRefreshRun,
    *,
    wb_finance_dir: Path,
) -> dict[str, Any]:
    database_hashes = sorted(
        str(value)
        for value in db.scalars(
            select(SourceSnapshotRow.raw_payload_hash).where(
                SourceSnapshotRow.refresh_run_id == refresh_run.id,
                SourceSnapshotRow.source_type == "wb_finance_detail",
            )
        )
    )
    collection = next(
        (
            item
            for item in refresh_run.collections
            if item.source_type == "wb_finance_detail"
        ),
        None,
    )
    if collection is None:
        return {"status": "mismatch", "mismatches": ["collectionMissing"]}
    file_hashes: list[str] = []
    for result in (collection.payload or {}).get("results", []):
        if not isinstance(result, dict):
            continue
        output_name = str(result.get("outputFile") or "").strip()
        if not output_name:
            continue
        for row in _read_json_list(wb_finance_dir / output_name):
            file_hashes.append(_hash_payload(row))
    file_hashes.sort()
    database_digest = _hash_payload(database_hashes)
    file_digest = _hash_payload(file_hashes)
    matched = (
        len(database_hashes) == len(file_hashes) and database_digest == file_digest
    )
    return {
        "status": "matched" if matched else "mismatch",
        "databaseRows": len(database_hashes),
        "fileRows": len(file_hashes),
        "databaseDigest": database_digest,
        "fileDigest": file_digest,
        "mismatches": [] if matched else ["sourceRows"],
    }


def _persisted_daily_facts_parity(
    db: Session,
    refresh_run: SourceRefreshRun,
    generated_facts: list[MarketplaceFinanceDailyFactContract],
) -> dict[str, Any]:
    persisted = list(
        db.scalars(
            select(MarketplaceFinanceDailyFactModel).where(
                MarketplaceFinanceDailyFactModel.tenant_id == refresh_run.tenant_id,
                MarketplaceFinanceDailyFactModel.client_id == refresh_run.client_id,
                MarketplaceFinanceDailyFactModel.marketplace == "wb",
                MarketplaceFinanceDailyFactModel.source_refresh_run_id
                == refresh_run.id,
            )
        )
    )
    field_names = tuple(MarketplaceFinanceDailyFactContract.model_fields)
    expected = [
        {name: getattr(item, name) for name in field_names} for item in generated_facts
    ]
    actual = [{name: getattr(item, name) for name in field_names} for item in persisted]
    expected_value = sorted(
        (_canonical_parity_value(item) for item in expected),
        key=_hash_payload,
    )
    actual_value = sorted(
        (_canonical_parity_value(item) for item in actual),
        key=_hash_payload,
    )
    expected_digest = _hash_payload(expected_value)
    actual_digest = _hash_payload(actual_value)
    matched = len(expected) == len(actual) and expected_digest == actual_digest
    return {
        "status": "matched" if matched else "mismatch",
        "expectedRows": len(expected),
        "persistedRows": len(actual),
        "expectedDigest": expected_digest,
        "persistedDigest": actual_digest,
        "mismatches": [] if matched else ["dailyFacts"],
    }


def _canonical_parity_value(value: Any) -> Any:
    if hasattr(value, "model_dump"):
        return _canonical_parity_value(value.model_dump(mode="python"))
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {
            str(key): _canonical_parity_value(item)
            for key, item in sorted(value.items(), key=lambda pair: str(pair[0]))
            if str(key) not in {"generatedAt", "generated_at"}
        }
    if isinstance(value, (list, tuple)):
        return [_canonical_parity_value(item) for item in value]
    if isinstance(value, set):
        return sorted(_canonical_parity_value(item) for item in value)
    if hasattr(value, "value"):
        return _canonical_parity_value(value.value)
    return value


def _ozon_status(item: OzonPageResult) -> str:
    if item.ok and item.status == "ok" and item.row_count > 0:
        return "loaded"
    if item.ok and item.status == "ok" and item.row_count == 0:
        return "empty_expected"
    if item.ok and item.status == "empty_expected":
        return "empty_expected"
    if item.status == "access_error":
        return "auth_failed"
    if item.status == "rate_limited":
        return "rate_limited"
    if item.status == "transport_error":
        return "schema_error"
    if item.status in {
        "page_limit_exhausted",
        "report_failed",
        "report_success_without_file",
        "report_timeout",
    }:
        return "partial_source"
    return "failed" if not item.ok else "loaded"


def _ozon_collection_row_count(results: list[OzonPageResult]) -> int:
    has_report_info = any(item.source_type.endswith("_info") for item in results)
    if has_report_info:
        return sum(
            item.row_count for item in results if item.source_type.endswith("_file")
        )
    return sum(item.row_count for item in results)


def _ozon_collection_status(
    results: list[OzonPageResult],
    payload_items: list[dict[str, Any]],
    *,
    row_count: int,
    required: bool,
) -> str:
    has_report_info = any(item.source_type.endswith("_info") for item in results)
    has_file_result = any(item.source_type.endswith("_file") for item in results)
    if has_report_info and (not has_file_result or row_count == 0):
        return "failed" if required else "needs_review"
    return _aggregate_status(payload_items, required=required)


def _wb_status(item: WbFinancePageResult | WbSalesReportListPageResult) -> str:
    if item.ok and item.status == "ok" and item.row_count > 0:
        return "loaded"
    if item.ok and item.status == "ok" and item.row_count == 0:
        return "empty_unexpected"
    if item.ok and item.status == "no_data":
        return "empty_unexpected" if item.page_index == 1 else "empty_expected"
    if item.status in {"access_error"}:
        return "auth_failed"
    if item.status == "rate_limited":
        return "rate_limited"
    if item.status == "transport_or_schema_error":
        return "schema_error"
    return "failed" if not item.ok else "loaded"


def _wb_product_cards_status(item: WbProductCardsPageResult) -> str:
    if item.ok and item.flat_row_count > 0:
        return "loaded"
    if item.ok and item.card_count == 0:
        return "empty_unexpected" if item.page_index == 1 else "empty_expected"
    if item.status_code in {401, 403}:
        return "auth_failed"
    if item.status_code == 429:
        return "rate_limited"
    return "failed" if not item.ok else "loaded"


def _wb_document_status(item: WbDocumentExportResult) -> str:
    if item.ok and item.row_count > 0 and item.downloaded_count > 0:
        return "loaded"
    if item.ok and item.row_count == 0:
        return "empty_expected"
    if item.status_code in {401, 403}:
        return "auth_failed"
    if item.status_code == 429:
        return "rate_limited"
    return "failed" if not item.ok else "needs_review"


def _aggregate_status(items: list[dict[str, Any]], *, required: bool) -> str:
    if not items:
        return "failed" if required else "needs_review"
    statuses = {str(item.get("status")) for item in items}
    if statuses <= MANDATORY_OK_STATUSES:
        return "loaded"
    if required and statuses - MANDATORY_OK_STATUSES:
        return sorted(statuses - MANDATORY_OK_STATUSES)[0]
    return "needs_review"


def _page_limit_exhausted(
    results: Iterable[Any],
    *,
    max_pages: int,
    page_limit: int,
    row_count_attribute: str,
    group_attributes: tuple[str, ...] = ("seller_account_id",),
) -> bool:
    latest_by_group: dict[tuple[str, ...], Any] = {}
    for item in results:
        group = tuple(
            str(getattr(item, attribute, "")) for attribute in group_attributes
        )
        previous = latest_by_group.get(group)
        if previous is None or int(item.page_index) > int(previous.page_index):
            latest_by_group[group] = item
    return any(
        bool(getattr(item, "ok", False))
        and int(getattr(item, "page_index", 0)) >= max_pages
        and int(getattr(item, row_count_attribute, 0)) >= page_limit
        for item in latest_by_group.values()
    )


def _onec_status(item: OnecSampleExportResult, *, required: bool) -> str:
    if item.status == "partial_source" or (not item.ok and item.page_count > 0):
        return "partial_source"
    if not item.ok:
        if not required and item.status_code == 404:
            return "empty_expected"
        if item.status_code in {401, 403}:
            return "auth_failed"
        if item.status_code == 429:
            return "rate_limited"
        return "failed"
    if item.row_count == 0:
        return "empty_unexpected" if required else "empty_expected"
    return "loaded"


def _onec_financial_table_quality(
    item: OnecSampleExportResult,
) -> dict[str, Any]:
    """Reject a commissioner snapshot that contains headers but no SKU details."""
    if item.sample_id != "commissioner_reports":
        return {"status": "not_applicable"}
    if not item.ok or item.row_count == 0 or item.output_path is None:
        return {"status": "not_checked"}
    try:
        rows = _extract_onec_rows(_read_json_object(item.output_path))
    except (OSError, ValueError, TypeError) as exc:
        return {
            "status": "partial_source",
            "code": "commissioner_financial_tables_unreadable",
            "error": exc.__class__.__name__,
        }
    table_names = ("Запасы", "ЗапасыВозвраты")
    rows_with_tables = 0
    rows_with_organization = 0
    financial_line_count = 0
    for row in rows:
        if str(row.get("Организация_Key") or "").strip():
            rows_with_organization += 1
        has_table = False
        for table_name in table_names:
            table = row.get(table_name)
            if isinstance(table, list):
                has_table = True
                financial_line_count += sum(
                    1 for value in table if isinstance(value, dict)
                )
        if has_table:
            rows_with_tables += 1
    if not rows_with_tables or not financial_line_count:
        return {
            "status": "partial_source",
            "code": "commissioner_financial_tables_missing",
            "headerRows": len(rows),
            "rowsWithTables": rows_with_tables,
            "financialLineCount": financial_line_count,
        }
    if not rows_with_organization:
        return {
            "status": "partial_source",
            "code": "commissioner_organization_missing",
            "headerRows": len(rows),
            "rowsWithTables": rows_with_tables,
            "rowsWithOrganization": rows_with_organization,
            "financialLineCount": financial_line_count,
        }
    return {
        "status": "loaded",
        "headerRows": len(rows),
        "rowsWithTables": rows_with_tables,
        "rowsWithOrganization": rows_with_organization,
        "financialLineCount": financial_line_count,
    }


def _commit_source_refresh_progress(db: Session) -> None:
    refresh_run_id = str(getattr(db, "info", {}).get("source_refresh_run_id") or "")
    if refresh_run_id:
        refresh_run = db.get(SourceRefreshRun, refresh_run_id)
        if refresh_run is not None and refresh_run.finished_at is None:
            repository.update_source_refresh_run(
                db,
                refresh_run,
                heartbeat_at=security.utcnow(),
            )
    db.commit()


def source_refresh_progress_payload(
    refresh_run: SourceRefreshRun,
    *,
    source_root: Path,
) -> dict[str, Any]:
    status = refresh_run.status
    terminal = refresh_run.finished_at is not None
    if terminal:
        stage = "complete"
        current_source = "Завершено"
    elif status == "queued":
        stage = "queued"
        current_source = "Очередь"
    elif status == "rebuilding":
        stage = "rebuilding"
        current_source = "Расчёт отчёта"
    elif status == "source_loaded":
        stage = "mapping"
        current_source = "Проверка сопоставления"
    elif refresh_run.mode == "onec-only":
        stage = "onec"
        current_source = "1С"
    elif refresh_run.mode == "ozon-only":
        stage = "onec"
        current_source = "Ozon и 1С"
    else:
        stage = "wb_finance"
        current_source = "WB"

    root = (
        Path(refresh_run.root_dir).resolve()
        if refresh_run.root_dir
        else (source_root / refresh_run.snapshot_set_id).resolve()
    )
    allowed_root = source_root.resolve()
    last_activity_at: datetime | None = None
    manifest: dict[str, Any] = {}
    manifest_path = root / "wb_finance" / "manifest.json"
    if (
        root.is_relative_to(allowed_root)
        and manifest_path.is_file()
        and manifest_path.stat().st_size <= 5 * 1024 * 1024
    ):
        with suppress(OSError, ValueError, TypeError, json.JSONDecodeError):
            loaded = json.loads(manifest_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                manifest = loaded

    if root.is_relative_to(allowed_root) and root.is_dir():
        activity_paths = [manifest_path]
        with suppress(OSError):
            activity_paths.extend(item for item in root.iterdir() if item.is_dir())
        onec_root = root / "onec"
        if onec_root.is_dir():
            with suppress(OSError):
                activity_paths.extend(
                    item for item in onec_root.iterdir() if item.is_dir()
                )
        for activity_path in activity_paths:
            with suppress(OSError):
                modified_at = datetime.fromtimestamp(
                    activity_path.stat().st_mtime,
                    tz=ZoneInfo("UTC"),
                )
                if last_activity_at is None or modified_at > last_activity_at:
                    last_activity_at = modified_at

    results = [item for item in manifest.get("results", []) if isinstance(item, dict)]
    account_results: dict[str, list[dict[str, Any]]] = {}
    bytes_written = 0
    for item in results:
        account_id = str(item.get("seller_account_id") or "").strip()
        if account_id:
            account_results.setdefault(account_id, []).append(item)
        output_name = str(item.get("output_file") or "").strip()
        if not output_name or Path(output_name).name != output_name:
            continue
        output_path = manifest_path.parent / output_name
        with suppress(OSError):
            if output_path.is_file() and output_path.resolve().is_relative_to(
                allowed_root
            ):
                bytes_written += output_path.stat().st_size

    wb_accounts_completed = sum(
        1
        for items in account_results.values()
        if items and str(items[-1].get("status") or "") == "no_data"
    )
    pages_loaded = sum(
        1 for item in results if str(item.get("status") or "") != "no_data"
    )
    rows_loaded = sum(max(0, int(item.get("row_count") or 0)) for item in results)
    completed_statuses = {
        "loaded",
        "empty_expected",
        "needs_review",
        "partial_source",
        "failed",
        "auth_failed",
        "rate_limited",
    }
    completed_sources = sum(
        1 for item in refresh_run.collections if item.status in completed_statuses
    )
    if not terminal and stage == "wb_finance" and (root / "onec").is_dir():
        stage = "onec"
        current_source = "1С"
    return {
        "stage": stage,
        "currentSource": current_source,
        "completedSources": completed_sources,
        "totalSources": len(refresh_run.collections),
        "wbAccountsCompleted": wb_accounts_completed,
        "wbAccountsTotal": len(account_results),
        "pagesLoaded": pages_loaded,
        "rowsLoaded": rows_loaded,
        "bytesWritten": bytes_written,
        "lastActivityAt": (
            last_activity_at.isoformat() if last_activity_at is not None else None
        ),
    }


def _persist_wb_finance_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbFinancePageResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            for local_index, row in enumerate(_read_json_list(result.output_path), 1):
                source_row_id = _first_row_id(row, "rrdId", "srid", "orderUid")
                if not source_row_id:
                    source_row_id = f"{result.page_index}:{local_index}"
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row),
                        "row_payload": row,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": wb_cabinet_ids.get(
                            result.seller_account_id,
                            "",
                        ),
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


@dataclass(frozen=True)
class _LogisticsSnapshotCandidate:
    snapshot_id: str
    refresh_run_id: str
    tenant_id: str
    client_id: str
    wb_cabinet_id: str
    source_type: str
    source_row_id: str
    loaded_at: datetime
    collection_id: int
    row_number: int
    role: str
    source_row: LogisticsSourceRow
    source_hash: str
    stable_identity: bool


@dataclass(frozen=True)
class _LogisticsSnapshotInput:
    snapshot_id: str
    refresh_run_id: str
    tenant_id: str
    client_id: str
    wb_cabinet_id: str
    source_type: str
    source_row_id: str
    row_number: int
    raw_payload_hash: str
    row_payload: Any
    loaded_at: datetime
    collection_id: int


def _file_authoritative_wb_collection(
    run: SourceRefreshRun,
) -> SourceRefreshCollection | None:
    collection = next(
        (item for item in run.collections if item.source_type == "wb_finance_detail"),
        None,
    )
    if collection is None:
        return None
    persistence = (collection.payload or {}).get("rowPersistence") or {}
    if persistence.get("status") not in {
        "file_authoritative",
        "skipped_large_snapshot",
    }:
        return None
    if persistence.get("rawFilesAuthoritative") is not True:
        return None
    return collection


def _iter_file_authoritative_logistics_inputs(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
) -> Iterable[_LogisticsSnapshotInput]:
    payload = collection.payload or {}
    if (payload.get("rawIntegrity") or {}).get("status") != "verified":
        raise SourceRefreshConfigError(
            "file-authoritative WB finance raw integrity is not verified"
        )
    results = payload.get("results")
    if not isinstance(results, list):
        raise SourceRefreshConfigError(
            "file-authoritative WB finance results are missing"
        )
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    try:
        verify_raw_directory(
            raw_dir,
            source_type="wb_finance_detail",
            source_root=source_root,
            collection_results=[item for item in results if isinstance(item, Mapping)],
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except RawIntegrityError as exc:
        raise SourceRefreshConfigError(
            "file-authoritative WB finance raw integrity changed"
        ) from exc
    row_number = 1
    for result in results:
        if not isinstance(result, Mapping):
            raise SourceRefreshConfigError(
                "file-authoritative WB finance result is not an object"
            )
        output_name = str(result.get("outputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise SourceRefreshConfigError(
                "file-authoritative WB finance output path is unsafe"
            )
        output_path = (raw_dir / output_name).resolve()
        resolved_raw_dir = raw_dir.resolve()
        if not output_path.is_relative_to(resolved_raw_dir):
            raise SourceRefreshConfigError(
                "file-authoritative WB finance output path is unsafe"
            )
        page_index = int(result.get("pageIndex") or 0)
        wb_cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for local_index, row in enumerate(iter_json_array(output_path), 1):
            source_row_id = (
                _first_row_id(row, "rrdId", "srid", "orderUid")
                if isinstance(row, dict)
                else ""
            )
            if not source_row_id:
                source_row_id = f"{page_index}:{local_index}"
            canonical_hash = _hash_payload(row)
            yield _LogisticsSnapshotInput(
                snapshot_id=f"file:{collection.id}:{row_number}",
                refresh_run_id=refresh_run.id,
                tenant_id=collection.tenant_id,
                client_id=collection.client_id,
                wb_cabinet_id=wb_cabinet_id,
                source_type=collection.source_type,
                source_row_id=source_row_id,
                row_number=row_number,
                raw_payload_hash=canonical_hash,
                row_payload=row,
                loaded_at=collection.loaded_at,
                collection_id=collection.id,
            )
            row_number += 1
    if row_number - 1 != collection.row_count:
        raise SourceRefreshConfigError(
            "file-authoritative WB finance row count changed"
        )


def _build_and_persist_logistics_analysis(
    db: Session,
    report: ReportRun,
    *,
    primary_refresh_run: SourceRefreshRun | None = None,
    base_refresh_run: SourceRefreshRun | None = None,
    contributing_runs: Iterable[SourceRefreshRun] = (),
    refresh_runs: Iterable[SourceRefreshRun] = (),
) -> Any:
    # `refresh_runs` remains as a compatibility bridge for existing internal
    # callers. Production passes explicit lineage roles so revision ownership
    # is deterministic.
    legacy_runs = [item for item in refresh_runs if item is not None]
    if primary_refresh_run is None and legacy_runs:
        primary_refresh_run = legacy_runs[0]
        contributing_runs = (*legacy_runs[1:], *contributing_runs)
    roles: dict[str, tuple[SourceRefreshRun, str]] = {}

    def add_run(run: SourceRefreshRun | None, role: str) -> None:
        if run is not None and run.id not in roles:
            roles[run.id] = (run, role)

    add_run(primary_refresh_run, "current")
    add_run(base_refresh_run, "base")
    for run in contributing_runs:
        add_run(run, "contributor")

    source_rows, diagnostics = _select_logistics_source_rows(
        db,
        report,
        roles=roles,
        primary_refresh_run=primary_refresh_run,
        base_refresh_run=base_refresh_run,
    )

    unit_rows = []
    database_scope_mismatches = 0
    for row in db.scalars(
        select(ReportUnitRow).where(ReportUnitRow.report_run_id == report.id)
    ):
        if not _report_unit_row_scope_matches(db, report, row):
            database_scope_mismatches += 1
        financial_date = row.week or row.accounting_period_date
        week_start = (
            financial_date - timedelta(days=financial_date.weekday())
            if financial_date is not None
            else None
        )
        unit_rows.append(
            UnitEconomicsSlice(
                tenant_id=report.tenant_id,
                client_id=row.client_id,
                financial_week_start=week_start,
                wb_cabinet_id=row.wb_cabinet_id,
                client_company_id=row.client_company_id,
                scheme=row.scheme,
                nm_id=row.nm_id,
                sku=row.barcode,
                vendor_code=row.article_wb,
                product=row.product,
                revenue=Decimal(row.revenue),
                profit_before_tax=Decimal(row.profit_before_tax),
                logistics=_report_logistics_decimal(row.logistics),
                source_row_id=row.row_uid,
            )
        )
    result = build_logistics_analysis(
        source_rows,
        unit_rows,
        report_period_start=report.period_start,
        report_period_end=report.period_end,
        expected_tenant_id=report.tenant_id,
        expected_client_id=report.client_id,
        input_diagnostics=LogisticsInputDiagnostics(
            invalid_source_payload_shape_count=(
                diagnostics.invalid_source_payload_shape_count
            ),
            source_identity_error_count=diagnostics.source_identity_error_count,
            source_revision_conflict_count=(diagnostics.source_revision_conflict_count),
            source_revision_discarded_count=(
                diagnostics.source_revision_discarded_count
            ),
            scope_mismatch_count=(
                diagnostics.scope_mismatch_count + database_scope_mismatches
            ),
            blocking_reasons=tuple(
                dict.fromkeys(
                    (
                        *diagnostics.blocking_reasons,
                        *(
                            ("tenant_scope_mismatch",)
                            if database_scope_mismatches
                            else ()
                        ),
                    )
                )
            ),
            lineage_records=diagnostics.lineage_records,
        ),
    )
    repository.replace_report_logistics_analysis(db, report, result)
    return result


@dataclass(frozen=True)
class _DimensionSnapshotSelection:
    card_rows: tuple[dict[str, Any], ...] = ()
    source_snapshot_hash: str = ""
    source_loaded_at: datetime | None = None
    source_row_count: int = 0
    blocking_reasons: tuple[str, ...] = ()
    review_reasons: tuple[str, ...] = ()


def _build_and_persist_logistics_dimensions(
    db: Session,
    report: ReportRun,
    *,
    logistics_result: LogisticsAnalysisResult,
    primary_refresh_run: SourceRefreshRun | None = None,
    base_refresh_run: SourceRefreshRun | None = None,
    contributing_runs: Iterable[SourceRefreshRun] = (),
) -> None:
    roles: list[tuple[int, SourceRefreshRun]] = []
    if primary_refresh_run is not None:
        roles.append((0, primary_refresh_run))
    if base_refresh_run is not None and all(
        run.id != base_refresh_run.id for _, run in roles
    ):
        roles.append((1, base_refresh_run))
    for run in contributing_runs:
        if all(existing.id != run.id for _, existing in roles):
            roles.append((2, run))
    selection = _select_dimension_snapshot(db, report, roles=roles)
    blocking = list(selection.blocking_reasons)
    review = list(selection.review_reasons)
    if logistics_result.context.data_status == "blocked":
        blocking.append("logistics_analysis_blocked")
    rows = (
        build_dimension_rows(logistics_result.sku_rows, selection.card_rows)
        if not blocking
        else []
    )
    missing = sum(row["data_quality_status"] == "missing_dimensions" for row in rows)
    invalid = sum(
        row["data_quality_status"] in {"invalid_dimensions", "identity_conflict"}
        for row in rows
    )
    conflicting = sum(
        row["data_quality_status"] == "conflicting_dimensions" for row in rows
    )
    matched = sum(row["evidence_type"] == "fact" for row in rows)
    signals = sum(row.get("dimensions_valid") is False for row in rows)
    if missing:
        review.append("dimension_values_missing")
    if invalid:
        review.append("dimension_values_invalid")
    if conflicting:
        review.append("dimension_values_conflicting")
    data_status = "blocked" if blocking else "partial" if review else "ready"
    input_hash = _hash_payload(
        {
            "factorMethodologyVersion": LOGISTICS_FACTORS_METHODOLOGY_VERSION,
            "sourceSnapshotHash": selection.source_snapshot_hash,
            "cardSourceHashes": sorted(
                str(row.get("source_hash") or "") for row in selection.card_rows
            ),
            "skuSourceHashes": sorted(
                row.source_hash_digest for row in logistics_result.sku_rows
            ),
            "blockingReasons": sorted(set(blocking)),
            "reviewReasons": sorted(set(review)),
        }
    )
    repository.replace_report_logistics_dimension_analysis(
        db,
        report,
        context={
            "tenant_id": report.tenant_id,
            "client_id": report.client_id,
            "factor_methodology_version": LOGISTICS_FACTORS_METHODOLOGY_VERSION,
            "data_status": data_status,
            "input_hash": input_hash,
            "source_snapshot_hash": selection.source_snapshot_hash,
            "source_loaded_at": selection.source_loaded_at,
            "source_row_count": selection.source_row_count,
            "dimension_row_count": len(rows),
            "matched_product_count": matched,
            "missing_product_count": missing,
            "invalid_product_count": invalid,
            "conflicting_product_count": conflicting,
            "signal_product_count": signals,
            "blocking_reasons": sorted(set(blocking)),
            "review_reasons": sorted(set(review)),
            "created_at": datetime.now(tz=ZoneInfo("UTC")),
        },
        rows=rows,
    )


def _select_dimension_snapshot(
    db: Session,
    report: ReportRun,
    *,
    roles: Iterable[tuple[int, SourceRefreshRun]],
) -> _DimensionSnapshotSelection:
    candidates: list[tuple[int, SourceRefreshRun, SourceRefreshCollection]] = []
    for priority, run in roles:
        for collection in run.collections:
            if collection.source_type == "wb_product_cards":
                candidates.append((priority, run, collection))
    if not candidates:
        return _DimensionSnapshotSelection(review_reasons=("dimension_source_missing",))
    selected_priority = min(item[0] for item in candidates)
    selected = [item for item in candidates if item[0] == selected_priority]
    snapshot_hashes = {item[2].snapshot_hash for item in selected}
    if len(snapshot_hashes) != 1:
        return _DimensionSnapshotSelection(
            blocking_reasons=("dimension_source_revision_conflict",)
        )
    _priority, run, collection = sorted(
        selected,
        key=lambda item: (
            _dimension_loaded_at_timestamp(item[2].loaded_at),
            item[1].id,
            item[2].id,
        ),
        reverse=True,
    )[0]
    if (
        collection.tenant_id != report.tenant_id
        or collection.client_id != report.client_id
    ):
        return _DimensionSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=("dimension_source_scope_mismatch",),
        )
    rows = list(
        db.scalars(
            select(SourceSnapshotRow)
            .where(
                SourceSnapshotRow.refresh_run_id == run.id,
                SourceSnapshotRow.collection_id == collection.id,
                SourceSnapshotRow.source_type == "wb_product_cards",
            )
            .order_by(SourceSnapshotRow.row_number)
        )
    )
    collection_payload = collection.payload or {}
    collection_results = collection_payload.get("results")
    if not isinstance(collection_results, list) or not all(
        isinstance(item, Mapping) for item in collection_results
    ):
        return _DimensionSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("dimension_source_manifest_invalid",),
        )
    if _hash_payload(collection_results) != collection.snapshot_hash:
        return _DimensionSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("dimension_source_snapshot_hash_mismatch",),
        )
    try:
        declared_row_count = sum(
            int(item.get("rowCount") or item.get("flat_row_count") or 0)
            for item in collection_results
        )
    except (TypeError, ValueError):
        declared_row_count = -1
    if declared_row_count != collection.row_count:
        return _DimensionSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("dimension_source_manifest_row_count_mismatch",),
        )
    persistence = collection_payload.get("rowPersistence") or {}
    file_authoritative = (
        persistence.get("status") in {"file_authoritative", "skipped_large_snapshot"}
        and persistence.get("rawFilesAuthoritative") is True
    )
    if file_authoritative and rows:
        return _DimensionSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("dimension_source_storage_ambiguity",),
        )
    blocking: list[str] = []
    card_rows: list[dict[str, Any]] = []
    if file_authoritative:
        try:
            card_rows.extend(
                _iter_file_authoritative_dimension_rows(
                    collection,
                    refresh_run=run,
                )
            )
        except (OSError, TypeError, ValueError, RawIntegrityError):
            blocking.append("dimension_file_snapshot_invalid")
    else:
        if len(rows) != collection.row_count:
            blocking.append("dimension_database_row_count_mismatch")
        for snapshot in rows:
            payload = snapshot.row_payload
            if not isinstance(payload, Mapping):
                blocking.append("dimension_source_payload_invalid")
                continue
            if snapshot.raw_payload_hash != _hash_payload(payload):
                blocking.append("dimension_source_payload_hash_mismatch")
                continue
            card_rows.append(
                _dimension_card_input(payload, wb_cabinet_id=snapshot.wb_cabinet_id)
            )
    blocking.extend(_dimension_card_scope_errors(db, report, card_rows))
    review: list[str] = []
    if collection.status == "partial_source":
        review.append("dimension_source_partial")
    if collection.status not in {"loaded", "partial_source"}:
        review.append("dimension_source_unavailable")
    return _DimensionSnapshotSelection(
        card_rows=tuple(card_rows) if not blocking else (),
        source_snapshot_hash=collection.snapshot_hash,
        source_loaded_at=collection.loaded_at,
        source_row_count=len(card_rows),
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        review_reasons=tuple(dict.fromkeys(review)),
    )


def _dimension_loaded_at_timestamp(value: datetime) -> float:
    if value.tzinfo is None:
        value = value.replace(tzinfo=ZoneInfo("UTC"))
    return value.astimezone(ZoneInfo("UTC")).timestamp()


def _iter_file_authoritative_dimension_rows(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
) -> Iterable[dict[str, Any]]:
    payload = collection.payload or {}
    if (payload.get("rawIntegrity") or {}).get("status") != "verified":
        raise RawIntegrityError("product card raw integrity is not verified")
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("product card results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type="wb_product_cards",
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    count = 0
    for result in results:
        if not isinstance(result, Mapping):
            raise RawIntegrityError("product card result is not an object")
        output_name = str(result.get("flatOutputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise RawIntegrityError("product card flat output path is unsafe")
        output_path = (raw_dir / output_name).resolve()
        if not output_path.is_relative_to(raw_dir.resolve()):
            raise RawIntegrityError("product card flat output path is unsafe")
        wb_cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for row in iter_json_array(output_path):
            if not isinstance(row, Mapping):
                raise RawIntegrityError("product card row is not an object")
            count += 1
            yield _dimension_card_input(row, wb_cabinet_id=wb_cabinet_id)
    if count != collection.row_count:
        raise RawIntegrityError("product card row count changed")


def _dimension_card_input(
    payload: Mapping[str, Any],
    *,
    wb_cabinet_id: str,
) -> dict[str, Any]:
    normalized = {
        "wb_cabinet_id": str(wb_cabinet_id or "").strip(),
        "nm_id": str(payload.get("nm_id") or "").strip(),
        "length_cm": payload.get("length_cm"),
        "width_cm": payload.get("width_cm"),
        "height_cm": payload.get("height_cm"),
        "weight_brutto_kg": payload.get("weight_brutto_kg"),
        "dimensions_valid": payload.get("dimensions_valid"),
    }
    normalized["source_hash"] = _hash_payload(normalized)
    return normalized


def _dimension_card_scope_errors(
    db: Session,
    report: ReportRun,
    rows: Iterable[Mapping[str, Any]],
) -> list[str]:
    cabinet_ids = {str(row.get("wb_cabinet_id") or "") for row in rows}
    if "" in cabinet_ids:
        return ["dimension_source_cabinet_missing"]
    cabinets = {
        item.id: item
        for item in db.scalars(select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids)))
    }
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return ["dimension_source_scope_mismatch"]
        company = db.get(ClientCompany, cabinet.client_company_id)
        if (
            company is None
            or company.tenant_id != report.tenant_id
            or company.client_id != report.client_id
        ):
            return ["dimension_source_scope_mismatch"]
    return []


@dataclass(frozen=True)
class _GoodsReturnSnapshotSelection:
    source_rows: tuple[GoodsReturnSourceRow, ...] = ()
    source_snapshot_hash: str = ""
    source_loaded_at: datetime | None = None
    source_coverage_start: date | None = None
    source_coverage_end: date | None = None
    source_row_count: int = 0
    blocking_reasons: tuple[str, ...] = ()
    review_reasons: tuple[str, ...] = ()


def _select_goods_return_snapshot(
    db: Session,
    report: ReportRun,
    *,
    roles: Iterable[tuple[int, SourceRefreshRun]],
) -> _GoodsReturnSnapshotSelection:
    candidates: list[tuple[int, SourceRefreshRun, SourceRefreshCollection]] = []
    for priority, run in roles:
        for collection in run.collections:
            if collection.source_type == "wb_goods_return":
                candidates.append((priority, run, collection))
    if not candidates:
        return _GoodsReturnSnapshotSelection(
            review_reasons=("goods_return_source_missing",)
        )
    selected_priority = min(item[0] for item in candidates)
    selected = [item for item in candidates if item[0] == selected_priority]
    if len({item[2].snapshot_hash for item in selected}) != 1:
        return _GoodsReturnSnapshotSelection(
            blocking_reasons=("goods_return_source_revision_conflict",)
        )
    _priority, run, collection = sorted(
        selected,
        key=lambda item: (
            _dimension_loaded_at_timestamp(item[2].loaded_at),
            item[1].id,
            item[2].id,
        ),
        reverse=True,
    )[0]
    if (
        collection.tenant_id != report.tenant_id
        or collection.client_id != report.client_id
    ):
        return _GoodsReturnSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=("goods_return_source_scope_mismatch",),
        )
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list) or not all(
        isinstance(item, Mapping) for item in results
    ):
        return _GoodsReturnSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("goods_return_source_manifest_invalid",),
        )
    if _hash_payload(results) != collection.snapshot_hash:
        return _GoodsReturnSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("goods_return_source_snapshot_hash_mismatch",),
        )
    try:
        declared_count = sum(int(item.get("rowCount") or 0) for item in results)
    except (TypeError, ValueError):
        declared_count = -1
    if declared_count != collection.row_count:
        return _GoodsReturnSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("goods_return_source_manifest_row_count_mismatch",),
        )
    manifest_scope = _goods_return_manifest_scope_errors(db, report, results)
    if manifest_scope:
        return _GoodsReturnSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=tuple(manifest_scope),
        )
    snapshots = list(
        db.scalars(
            select(SourceSnapshotRow)
            .where(
                SourceSnapshotRow.refresh_run_id == run.id,
                SourceSnapshotRow.collection_id == collection.id,
                SourceSnapshotRow.source_type == "wb_goods_return",
            )
            .order_by(SourceSnapshotRow.row_number)
        )
    )
    persistence = payload.get("rowPersistence") or {}
    file_authoritative = (
        persistence.get("status") in {"file_authoritative", "skipped_large_snapshot"}
        and persistence.get("rawFilesAuthoritative") is True
    )
    if file_authoritative and snapshots:
        return _GoodsReturnSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("goods_return_source_storage_ambiguity",),
        )
    blocking: list[str] = []
    source_rows: list[GoodsReturnSourceRow] = []
    if file_authoritative:
        try:
            source_rows.extend(
                _iter_file_authoritative_goods_return_rows(
                    collection,
                    refresh_run=run,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                )
            )
        except (OSError, TypeError, ValueError, RawIntegrityError):
            blocking.append("goods_return_file_snapshot_invalid")
    else:
        if len(snapshots) != collection.row_count:
            blocking.append("goods_return_database_row_count_mismatch")
        for snapshot in snapshots:
            row_payload = snapshot.row_payload
            if not isinstance(row_payload, Mapping):
                blocking.append("goods_return_source_payload_invalid")
                continue
            if snapshot.raw_payload_hash != _hash_payload(row_payload):
                blocking.append("goods_return_source_payload_hash_mismatch")
                continue
            source_rows.append(
                normalize_goods_return_source_row(
                    row_payload,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                    wb_cabinet_id=snapshot.wb_cabinet_id,
                )
            )
    try:
        if (payload.get("rawIntegrity") or {}).get("status") != "verified":
            raise RawIntegrityError("goods-return raw integrity is not verified")
        raw_dir = Path(collection.raw_path)
        source_root = Path(run.root_dir) if run.root_dir else raw_dir
        verify_raw_directory(
            raw_dir,
            source_type="wb_goods_return",
            source_root=source_root,
            collection_results=[item for item in results if isinstance(item, Mapping)],
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except (OSError, TypeError, ValueError, RawIntegrityError):
        blocking.append("goods_return_raw_snapshot_invalid")
    blocking.extend(_goods_return_scope_errors(db, report, source_rows))
    coverage_start = _safe_iso_date(payload.get("coverageStart"))
    coverage_end = _safe_iso_date(payload.get("coverageEnd"))
    review: list[str] = []
    if coverage_start is None or coverage_end is None:
        review.append("goods_return_source_coverage_missing")
    elif coverage_start > report.period_start or coverage_end < report.period_end:
        review.append("goods_return_source_period_partial")
    if any(row.validation_errors for row in source_rows):
        review.append("goods_return_source_identity_invalid")
    if collection.status == "partial_source":
        review.append("goods_return_source_partial")
    if collection.status not in {"loaded", "partial_source"}:
        review.append("goods_return_source_unavailable")
    return _GoodsReturnSnapshotSelection(
        source_rows=tuple(source_rows) if not blocking else (),
        source_snapshot_hash=collection.snapshot_hash,
        source_loaded_at=collection.loaded_at,
        source_coverage_start=coverage_start,
        source_coverage_end=coverage_end,
        source_row_count=len(source_rows),
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        review_reasons=tuple(dict.fromkeys(review)),
    )


def _iter_file_authoritative_goods_return_rows(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
    tenant_id: str,
    client_id: str,
) -> Iterable[GoodsReturnSourceRow]:
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("goods-return results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type="wb_goods_return",
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    count = 0
    for result in results:
        if not isinstance(result, Mapping):
            raise RawIntegrityError("goods-return result is not an object")
        output_name = str(result.get("flatOutputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise RawIntegrityError("goods-return flat output path is unsafe")
        output_path = (raw_dir / output_name).resolve()
        if not output_path.is_relative_to(raw_dir.resolve()):
            raise RawIntegrityError("goods-return flat output path is unsafe")
        rows = json.loads(output_path.read_text(encoding="utf-8"))
        if not isinstance(rows, list):
            raise RawIntegrityError("goods-return flat rows are not a list")
        cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for row in rows:
            if not isinstance(row, Mapping):
                raise RawIntegrityError("goods-return row is not an object")
            count += 1
            yield normalize_goods_return_source_row(
                row,
                tenant_id=tenant_id,
                client_id=client_id,
                wb_cabinet_id=cabinet_id,
            )
    if count != collection.row_count:
        raise RawIntegrityError("goods-return row count changed")


def _goods_return_manifest_scope_errors(
    db: Session,
    report: ReportRun,
    results: Iterable[Mapping[str, Any]],
) -> list[str]:
    cabinet_ids = {str(item.get("wbCabinetId") or "").strip() for item in results}
    if "" in cabinet_ids:
        return ["goods_return_source_cabinet_missing"]
    return _goods_return_cabinet_scope_errors(db, report, cabinet_ids)


def _goods_return_scope_errors(
    db: Session,
    report: ReportRun,
    rows: Iterable[GoodsReturnSourceRow],
) -> list[str]:
    values = list(rows)
    if any(
        row.tenant_id != report.tenant_id or row.client_id != report.client_id
        for row in values
    ):
        return ["goods_return_source_scope_mismatch"]
    cabinet_ids = {row.wb_cabinet_id for row in values}
    if "" in cabinet_ids:
        return ["goods_return_source_cabinet_missing"]
    return _goods_return_cabinet_scope_errors(db, report, cabinet_ids)


def _goods_return_cabinet_scope_errors(
    db: Session,
    report: ReportRun,
    cabinet_ids: set[str],
) -> list[str]:
    cabinets = (
        {
            item.id: item
            for item in db.scalars(
                select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids))
            )
        }
        if cabinet_ids
        else {}
    )
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return ["goods_return_source_scope_mismatch"]
        company = db.get(ClientCompany, cabinet.client_company_id)
        if (
            company is None
            or company.tenant_id != report.tenant_id
            or company.client_id != report.client_id
        ):
            return ["goods_return_source_scope_mismatch"]
    return []


@dataclass(frozen=True)
class _ReturnClaimsSnapshotSelection:
    source_rows: tuple[ClaimSourceRow, ...] = ()
    cabinet_states: tuple[tuple[str, str], ...] = ()
    source_state: str = "unavailable"
    source_snapshot_hash: str = ""
    source_loaded_at: datetime | None = None
    source_coverage_start: date | None = None
    source_coverage_end: date | None = None
    source_row_count: int = 0
    blocking_reasons: tuple[str, ...] = ()
    review_reasons: tuple[str, ...] = ()


def _select_return_claims_snapshot(
    db: Session,
    report: ReportRun,
    *,
    roles: Iterable[tuple[int, SourceRefreshRun]],
) -> _ReturnClaimsSnapshotSelection:
    candidates: list[tuple[int, SourceRefreshRun, SourceRefreshCollection]] = []
    for priority, run in roles:
        for collection in run.collections:
            if collection.source_type == "wb_return_claims":
                candidates.append((priority, run, collection))
    if not candidates:
        return _ReturnClaimsSnapshotSelection(
            review_reasons=("return_claims_source_missing",)
        )
    selected_priority = min(item[0] for item in candidates)
    selected = [item for item in candidates if item[0] == selected_priority]
    if len({item[2].snapshot_hash for item in selected}) != 1:
        return _ReturnClaimsSnapshotSelection(
            blocking_reasons=("return_claims_source_revision_conflict",)
        )
    _priority, run, collection = sorted(
        selected,
        key=lambda item: (
            _dimension_loaded_at_timestamp(item[2].loaded_at),
            item[1].id,
            item[2].id,
        ),
        reverse=True,
    )[0]
    if (
        collection.tenant_id != report.tenant_id
        or collection.client_id != report.client_id
    ):
        return _ReturnClaimsSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=("return_claims_source_scope_mismatch",),
        )
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list) or not all(
        isinstance(item, Mapping) for item in results
    ):
        return _ReturnClaimsSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("return_claims_source_manifest_invalid",),
        )
    if _hash_payload(results) != collection.snapshot_hash:
        return _ReturnClaimsSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("return_claims_source_snapshot_hash_mismatch",),
        )
    try:
        declared_count = sum(int(item.get("rowCount") or 0) for item in results)
    except (TypeError, ValueError):
        declared_count = -1
    if declared_count != collection.row_count:
        return _ReturnClaimsSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("return_claims_source_manifest_row_count_mismatch",),
        )
    manifest_scope = _return_claims_manifest_scope_errors(db, report, results)
    if manifest_scope:
        return _ReturnClaimsSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=tuple(manifest_scope),
        )
    snapshots = list(
        db.scalars(
            select(SourceSnapshotRow)
            .where(
                SourceSnapshotRow.refresh_run_id == run.id,
                SourceSnapshotRow.collection_id == collection.id,
                SourceSnapshotRow.source_type == "wb_return_claims",
            )
            .order_by(SourceSnapshotRow.row_number)
        )
    )
    persistence = payload.get("rowPersistence") or {}
    file_authoritative = (
        persistence.get("status") in {"file_authoritative", "skipped_large_snapshot"}
        and persistence.get("rawFilesAuthoritative") is True
    )
    if file_authoritative and snapshots:
        return _ReturnClaimsSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("return_claims_source_storage_ambiguity",),
        )
    blocking: list[str] = []
    source_rows: list[ClaimSourceRow] = []
    if file_authoritative:
        try:
            source_rows.extend(
                _iter_file_authoritative_return_claim_rows(
                    collection,
                    refresh_run=run,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                )
            )
        except (OSError, TypeError, ValueError, RawIntegrityError):
            blocking.append("return_claims_file_snapshot_invalid")
    else:
        if len(snapshots) != collection.row_count:
            blocking.append("return_claims_database_row_count_mismatch")
        for snapshot in snapshots:
            row_payload = snapshot.row_payload
            if not isinstance(row_payload, Mapping):
                blocking.append("return_claims_source_payload_invalid")
                continue
            if snapshot.raw_payload_hash != _hash_payload(row_payload):
                blocking.append("return_claims_source_payload_hash_mismatch")
                continue
            source_rows.append(
                normalize_claim_source_row(
                    row_payload,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                    wb_cabinet_id=snapshot.wb_cabinet_id,
                )
            )
    try:
        if (payload.get("rawIntegrity") or {}).get("status") != "verified":
            raise RawIntegrityError("return-claims raw integrity is not verified")
        raw_dir = Path(collection.raw_path)
        source_root = Path(run.root_dir) if run.root_dir else raw_dir
        verify_raw_directory(
            raw_dir,
            source_type="wb_return_claims",
            source_root=source_root,
            collection_results=[item for item in results if isinstance(item, Mapping)],
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except (OSError, TypeError, ValueError, RawIntegrityError):
        blocking.append("return_claims_raw_snapshot_invalid")
    blocking.extend(_return_claims_scope_errors(db, report, source_rows))
    coverage_start = _safe_iso_date(payload.get("coverageStart"))
    coverage_end = _safe_iso_date(payload.get("coverageEnd"))
    cabinet_states = tuple(
        sorted(
            (
                str(item.get("wbCabinetId") or "").strip(),
                str(item.get("status") or "unavailable").strip(),
            )
            for item in results
        )
    )
    distinct_states = {state for _cabinet_id, state in cabinet_states}
    source_state = (
        next(iter(distinct_states)) if len(distinct_states) == 1 else "partial"
    )
    review: list[str] = []
    if coverage_start is None or coverage_end is None:
        review.append("return_claims_source_coverage_missing")
    elif coverage_start > report.period_start or coverage_end < report.period_end:
        review.append("return_claims_source_period_partial")
    if any(row.validation_errors for row in source_rows):
        review.append("return_claims_source_identity_invalid")
    if source_state == "confirmed_empty":
        review.append("return_claims_source_empty")
    elif source_state in {"access_denied", "paid_scope_required"}:
        review.append("return_claims_source_access_denied")
    elif source_state not in {"confirmed_nonempty"}:
        review.append("return_claims_source_unavailable")
    if collection.status == "partial_source":
        review.append("return_claims_source_partial")
    return _ReturnClaimsSnapshotSelection(
        source_rows=tuple(source_rows) if not blocking else (),
        cabinet_states=cabinet_states,
        source_state=source_state,
        source_snapshot_hash=collection.snapshot_hash,
        source_loaded_at=collection.loaded_at,
        source_coverage_start=coverage_start,
        source_coverage_end=coverage_end,
        source_row_count=len(source_rows),
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        review_reasons=tuple(dict.fromkeys(review)),
    )


def _iter_file_authoritative_return_claim_rows(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
    tenant_id: str,
    client_id: str,
) -> Iterable[ClaimSourceRow]:
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("return-claims results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type="wb_return_claims",
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    count = 0
    for result in results:
        if not isinstance(result, Mapping):
            raise RawIntegrityError("return-claims result is not an object")
        output_name = str(result.get("flatOutputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise RawIntegrityError("return-claims flat output path is unsafe")
        output_path = (raw_dir / output_name).resolve()
        if not output_path.is_relative_to(raw_dir.resolve()):
            raise RawIntegrityError("return-claims flat output path is unsafe")
        rows = json.loads(output_path.read_text(encoding="utf-8"))
        if not isinstance(rows, list):
            raise RawIntegrityError("return-claims flat rows are not a list")
        cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for row in rows:
            if not isinstance(row, Mapping):
                raise RawIntegrityError("return-claims row is not an object")
            count += 1
            yield normalize_claim_source_row(
                row,
                tenant_id=tenant_id,
                client_id=client_id,
                wb_cabinet_id=cabinet_id,
            )
    if count != collection.row_count:
        raise RawIntegrityError("return-claims row count changed")


def _return_claims_manifest_scope_errors(
    db: Session,
    report: ReportRun,
    results: Iterable[Mapping[str, Any]],
) -> list[str]:
    cabinet_ids = {str(item.get("wbCabinetId") or "").strip() for item in results}
    if "" in cabinet_ids:
        return ["return_claims_source_cabinet_missing"]
    return _return_claims_cabinet_scope_errors(db, report, cabinet_ids)


def _return_claims_scope_errors(
    db: Session,
    report: ReportRun,
    rows: Iterable[ClaimSourceRow],
) -> list[str]:
    values = list(rows)
    if any(
        row.tenant_id != report.tenant_id or row.client_id != report.client_id
        for row in values
    ):
        return ["return_claims_source_scope_mismatch"]
    cabinet_ids = {row.wb_cabinet_id for row in values}
    if "" in cabinet_ids:
        return ["return_claims_source_cabinet_missing"]
    return _return_claims_cabinet_scope_errors(db, report, cabinet_ids)


def _return_claims_cabinet_scope_errors(
    db: Session,
    report: ReportRun,
    cabinet_ids: set[str],
) -> list[str]:
    cabinets = (
        {
            item.id: item
            for item in db.scalars(
                select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids))
            )
        }
        if cabinet_ids
        else {}
    )
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return ["return_claims_source_scope_mismatch"]
        company = db.get(ClientCompany, cabinet.client_company_id)
        if (
            company is None
            or company.tenant_id != report.tenant_id
            or company.client_id != report.client_id
        ):
            return ["return_claims_source_scope_mismatch"]
    return []


def _build_and_persist_logistics_return_reasons(
    db: Session,
    report: ReportRun,
    *,
    logistics_result: LogisticsAnalysisResult,
    primary_refresh_run: SourceRefreshRun | None = None,
    base_refresh_run: SourceRefreshRun | None = None,
    contributing_runs: Iterable[SourceRefreshRun] = (),
) -> None:
    """Build R-3 from immutable Finance, goods-return and claims snapshots."""

    role_list: list[tuple[int, SourceRefreshRun]] = []
    role_map: dict[str, tuple[SourceRefreshRun, str]] = {}

    def add_role(
        priority: int,
        run: SourceRefreshRun | None,
        role: str,
    ) -> None:
        if run is None or run.id in role_map:
            return
        role_list.append((priority, run))
        role_map[run.id] = (run, role)

    add_role(0, primary_refresh_run, "current")
    add_role(1, base_refresh_run, "base")
    for run in contributing_runs:
        add_role(2, run, "contributor")

    goods_return = _select_goods_return_snapshot(db, report, roles=role_list)
    claims = _select_return_claims_snapshot(db, report, roles=role_list)
    finance_rows, finance_diagnostics = _select_logistics_source_rows(
        db,
        report,
        roles=role_map,
        primary_refresh_run=primary_refresh_run,
        base_refresh_run=base_refresh_run,
    )
    result = build_return_reason_analysis(
        finance_rows,
        logistics_result.order_rows,
        goods_return.source_rows,
        claims.source_rows,
        goods_return_snapshot_hash=goods_return.source_snapshot_hash,
        claims_snapshot_hash=claims.source_snapshot_hash,
        goods_return_coverage_start=goods_return.source_coverage_start,
        goods_return_coverage_end=goods_return.source_coverage_end,
        claims_coverage_start=claims.source_coverage_start,
        claims_coverage_end=claims.source_coverage_end,
        claims_source_status=claims.source_state,
        blocking_reasons=tuple(
            dict.fromkeys(
                (
                    *logistics_result.context.blocking_reasons,
                    *finance_diagnostics.blocking_reasons,
                )
            )
        ),
        review_reasons=logistics_result.context.review_reasons,
        goods_return_blocking_reasons=goods_return.blocking_reasons,
        goods_return_review_reasons=goods_return.review_reasons,
        claims_blocking_reasons=claims.blocking_reasons,
        claims_review_reasons=claims.review_reasons,
    )
    repository.replace_report_logistics_return_reason_analysis(
        db,
        report,
        result,
        goods_return_source_loaded_at=goods_return.source_loaded_at,
        claims_source_loaded_at=claims.source_loaded_at,
    )


@dataclass(frozen=True)
class _MeasurementSnapshotSelection:
    source_type: str
    measurement_rows: tuple[dict[str, Any], ...] = ()
    source_snapshot_hash: str = ""
    source_loaded_at: datetime | None = None
    factor_snapshot_at: datetime | None = None
    source_coverage_start: date | None = None
    source_coverage_end: date | None = None
    source_row_count: int = 0
    provider_event_count: int = 0
    complete: bool = False
    blocking_reasons: tuple[str, ...] = ()
    review_reasons: tuple[str, ...] = ()


def _build_and_persist_logistics_measurements(
    db: Session,
    report: ReportRun,
    *,
    logistics_result: LogisticsAnalysisResult,
    primary_refresh_run: SourceRefreshRun | None = None,
    base_refresh_run: SourceRefreshRun | None = None,
    contributing_runs: Iterable[SourceRefreshRun] = (),
) -> None:
    roles: list[tuple[int, SourceRefreshRun]] = []
    if primary_refresh_run is not None:
        roles.append((0, primary_refresh_run))
    if base_refresh_run is not None and all(
        run.id != base_refresh_run.id for _, run in roles
    ):
        roles.append((1, base_refresh_run))
    for run in contributing_runs:
        if all(existing.id != run.id for _, existing in roles):
            roles.append((2, run))

    penalty = _select_measurement_snapshot(
        db,
        report,
        roles=roles,
        source_type="wb_measurement_penalties",
    )
    warehouse = _select_measurement_snapshot(
        db,
        report,
        roles=roles,
        source_type="wb_warehouse_measurements",
    )
    selections = (penalty, warehouse)
    blocking = [
        reason for selection in selections for reason in selection.blocking_reasons
    ]
    review = [reason for selection in selections for reason in selection.review_reasons]
    if logistics_result.context.data_status == "blocked":
        blocking.append("logistics_analysis_blocked")
    rows = (
        build_measurement_rows(
            logistics_result.sku_rows,
            penalty.measurement_rows,
            warehouse.measurement_rows,
        )
        if not blocking
        else []
    )

    unmatched = sum(row["coverage_status"] == "unmatched_product" for row in rows)
    ambiguous = sum(row["coverage_status"] == "ambiguous_product_scope" for row in rows)
    invalid = sum(row["coverage_status"] == "invalid_measurement" for row in rows)
    conflicting = sum(
        row["coverage_status"] == "conflicting_measurement" for row in rows
    )
    matched = sum(bool(row.get("product_ref")) for row in rows)
    penalty_events = sum(
        (row.get("penalty_amount") or Decimal("0")) > 0 for row in rows
    )
    reversal_events = sum(
        (row.get("reversal_amount") or Decimal("0")) > 0 for row in rows
    )
    warehouse_only = sum(
        row.get("event_kind") == "warehouse_measurement" for row in rows
    )
    for count, reason in (
        (unmatched, "measurement_product_unmatched"),
        (ambiguous, "measurement_product_scope_ambiguous"),
        (invalid, "measurement_values_invalid"),
        (conflicting, "measurement_values_conflicting"),
    ):
        if count:
            review.append(reason)
    complete_endpoints = sum(selection.complete for selection in selections)
    unavailable_endpoints = len(selections) - complete_endpoints
    if unavailable_endpoints:
        review.append("measurement_endpoint_incomplete")
    data_status = "blocked" if blocking else "partial" if review else "ready"
    factor_times = [
        selection.factor_snapshot_at
        for selection in selections
        if selection.factor_snapshot_at is not None
    ]
    factor_snapshot_at = max(factor_times) if factor_times else None
    scoped_products = {
        (row.wb_cabinet_id, row.nm_id)
        for row in logistics_result.sku_rows
        if row.wb_cabinet_id and row.nm_id
    }
    products_with_events = {
        str(row.get("product_ref")) for row in rows if row.get("product_ref")
    }
    input_hash = _hash_payload(
        {
            "factorMethodologyVersion": LOGISTICS_MEASUREMENTS_METHODOLOGY_VERSION,
            "penaltySnapshotHash": penalty.source_snapshot_hash,
            "warehouseSnapshotHash": warehouse.source_snapshot_hash,
            "penaltySourceHashes": sorted(
                str(row.get("source_hash") or "") for row in penalty.measurement_rows
            ),
            "warehouseSourceHashes": sorted(
                str(row.get("source_hash") or "") for row in warehouse.measurement_rows
            ),
            "skuSourceHashes": sorted(
                row.source_hash_digest for row in logistics_result.sku_rows
            ),
            "blockingReasons": sorted(set(blocking)),
            "reviewReasons": sorted(set(review)),
        }
    )
    repository.replace_report_logistics_measurement_analysis(
        db,
        report,
        context={
            "tenant_id": report.tenant_id,
            "client_id": report.client_id,
            "factor_methodology_version": (LOGISTICS_MEASUREMENTS_METHODOLOGY_VERSION),
            "data_status": data_status,
            "input_hash": input_hash,
            "penalty_source_snapshot_hash": penalty.source_snapshot_hash,
            "warehouse_source_snapshot_hash": warehouse.source_snapshot_hash,
            "penalty_source_loaded_at": penalty.source_loaded_at,
            "warehouse_source_loaded_at": warehouse.source_loaded_at,
            "factor_snapshot_at": factor_snapshot_at,
            "source_coverage_start": (
                report.period_start
                if any(selection.source_coverage_start for selection in selections)
                else None
            ),
            "source_coverage_end": (
                report.period_end
                if any(selection.source_coverage_end for selection in selections)
                else None
            ),
            "expected_endpoint_count": len(selections),
            "complete_endpoint_count": complete_endpoints,
            "unavailable_endpoint_count": unavailable_endpoints,
            "source_event_count": sum(
                selection.source_row_count for selection in selections
            ),
            "provider_event_count": sum(
                selection.provider_event_count for selection in selections
            ),
            "measurement_row_count": len(rows),
            "scoped_product_count": len(scoped_products),
            "product_with_event_count": len(products_with_events),
            "matched_event_count": matched,
            "unmatched_event_count": unmatched,
            "ambiguous_event_count": ambiguous,
            "invalid_event_count": invalid,
            "conflicting_event_count": conflicting,
            "penalty_event_count": penalty_events,
            "reversal_event_count": reversal_events,
            "warehouse_only_event_count": warehouse_only,
            "blocking_reasons": sorted(set(blocking)),
            "review_reasons": sorted(set(review)),
            "created_at": datetime.now(tz=UTC),
        },
        rows=rows,
    )


def _select_measurement_snapshot(
    db: Session,
    report: ReportRun,
    *,
    roles: Iterable[tuple[int, SourceRefreshRun]],
    source_type: str,
) -> _MeasurementSnapshotSelection:
    if source_type not in {
        "wb_measurement_penalties",
        "wb_warehouse_measurements",
    }:
        raise ValueError("invalid measurement source type")
    code = (
        "measurement_penalties"
        if source_type == "wb_measurement_penalties"
        else "warehouse_measurements"
    )
    candidates: list[tuple[int, SourceRefreshRun, SourceRefreshCollection]] = []
    for priority, run in roles:
        for collection in run.collections:
            if collection.source_type == source_type:
                candidates.append((priority, run, collection))
    if not candidates:
        return _MeasurementSnapshotSelection(
            source_type=source_type,
            review_reasons=(f"{code}_source_missing",),
        )
    selected_priority = min(item[0] for item in candidates)
    selected = [item for item in candidates if item[0] == selected_priority]
    if len({item[2].snapshot_hash for item in selected}) != 1:
        return _MeasurementSnapshotSelection(
            source_type=source_type,
            blocking_reasons=(f"{code}_source_revision_conflict",),
        )
    _priority, run, collection = sorted(
        selected,
        key=lambda item: (
            _dimension_loaded_at_timestamp(item[2].loaded_at),
            item[1].id,
            item[2].id,
        ),
        reverse=True,
    )[0]
    if (
        collection.tenant_id != report.tenant_id
        or collection.client_id != report.client_id
    ):
        return _MeasurementSnapshotSelection(
            source_type=source_type,
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=(f"{code}_source_scope_mismatch",),
        )
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list) or not all(
        isinstance(item, Mapping) for item in results
    ):
        return _MeasurementSnapshotSelection(
            source_type=source_type,
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=(f"{code}_source_manifest_invalid",),
        )
    blocking: list[str] = []
    review: list[str] = []
    if _hash_payload(results) != collection.snapshot_hash:
        blocking.append(f"{code}_source_snapshot_hash_mismatch")
    try:
        declared_count = sum(int(item.get("rowCount") or 0) for item in results)
        provider_total = sum(
            int(item.get("providerTotal") or 0)
            for item in results
            if item.get("ok") is True
        )
    except (TypeError, ValueError):
        declared_count = -1
        provider_total = -1
    if declared_count != collection.row_count:
        blocking.append(f"{code}_source_manifest_row_count_mismatch")
    for item in results:
        row_count = item.get("rowCount")
        item_total = item.get("providerTotal")
        if item.get("ok") is True and (
            not isinstance(row_count, int)
            or isinstance(row_count, bool)
            or not isinstance(item_total, int)
            or isinstance(item_total, bool)
            or item_total != row_count
        ):
            blocking.append(f"{code}_source_provider_total_mismatch")
        if item.get("ok") is not True:
            review.append(f"{code}_source_unavailable")
    blocking.extend(_measurement_manifest_scope_errors(db, report, results, code=code))

    period_start = _safe_iso_date(payload.get("periodStart"))
    period_end = _safe_iso_date(payload.get("periodEnd"))
    coverage_start = _safe_iso_date(payload.get("coverageStart"))
    coverage_end = _safe_iso_date(payload.get("coverageEnd"))
    if (
        period_start != report.period_start
        or period_end != report.period_end
        or coverage_start != report.period_start
        or coverage_end != report.period_end
    ):
        blocking.append(f"{code}_source_window_mismatch")
    factor_snapshot_at = _safe_aware_datetime(payload.get("factorSnapshotAt"))
    if factor_snapshot_at is None:
        blocking.append(f"{code}_factor_snapshot_missing")

    snapshots = list(
        db.scalars(
            select(SourceSnapshotRow)
            .where(
                SourceSnapshotRow.refresh_run_id == run.id,
                SourceSnapshotRow.collection_id == collection.id,
                SourceSnapshotRow.source_type == source_type,
            )
            .order_by(SourceSnapshotRow.row_number)
        )
    )
    persistence = payload.get("rowPersistence") or {}
    file_authoritative = (
        persistence.get("status") in {"file_authoritative", "skipped_large_snapshot"}
        and persistence.get("rawFilesAuthoritative") is True
    )
    if file_authoritative and snapshots:
        blocking.append(f"{code}_source_storage_ambiguity")
    measurement_rows: list[dict[str, Any]] = []
    if file_authoritative:
        try:
            measurement_rows.extend(
                _iter_file_authoritative_measurement_rows(
                    collection,
                    refresh_run=run,
                    source_type=source_type,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                )
            )
        except (OSError, TypeError, ValueError, RawIntegrityError):
            blocking.append(f"{code}_file_snapshot_invalid")
    else:
        if len(snapshots) != collection.row_count:
            blocking.append(f"{code}_database_row_count_mismatch")
        for snapshot in snapshots:
            row_payload = snapshot.row_payload
            if not isinstance(row_payload, Mapping):
                blocking.append(f"{code}_source_payload_invalid")
                continue
            if snapshot.raw_payload_hash != _hash_payload(row_payload):
                blocking.append(f"{code}_source_payload_hash_mismatch")
                continue
            if row_payload.get("measurement_source_type") != source_type:
                blocking.append(f"{code}_source_type_mismatch")
                continue
            measurement_rows.append(
                _measurement_input(
                    row_payload,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                    wb_cabinet_id=snapshot.wb_cabinet_id,
                )
            )
    try:
        _verify_measurement_manifest(
            collection,
            refresh_run=run,
            source_type=source_type,
            expected_metadata={
                "periodStart": payload.get("periodStart"),
                "periodEnd": payload.get("periodEnd"),
                "coverageStart": payload.get("coverageStart"),
                "coverageEnd": payload.get("coverageEnd"),
                "factorSnapshotAt": payload.get("factorSnapshotAt"),
            },
        )
    except (OSError, TypeError, ValueError, RawIntegrityError):
        blocking.append(f"{code}_raw_snapshot_invalid")
    blocking.extend(
        _measurement_row_scope_errors(db, report, measurement_rows, code=code)
    )
    if collection.status == "partial_source":
        review.append(f"{code}_source_partial")
    if collection.status not in {"loaded", "partial_source"}:
        review.append(f"{code}_source_unavailable")
    complete = (
        collection.status == "loaded"
        and bool(results)
        and all(item.get("ok") is True for item in results)
        and not blocking
    )
    return _MeasurementSnapshotSelection(
        source_type=source_type,
        measurement_rows=tuple(measurement_rows) if not blocking else (),
        source_snapshot_hash=collection.snapshot_hash,
        source_loaded_at=collection.loaded_at,
        factor_snapshot_at=factor_snapshot_at,
        source_coverage_start=coverage_start,
        source_coverage_end=coverage_end,
        source_row_count=len(measurement_rows),
        provider_event_count=max(provider_total, 0),
        complete=complete,
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        review_reasons=tuple(dict.fromkeys(review)),
    )


def _safe_aware_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = f"{text[:-1]}+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return None
    return parsed.astimezone(UTC)


def _measurement_manifest_scope_errors(
    db: Session,
    report: ReportRun,
    results: Iterable[Mapping[str, Any]],
    *,
    code: str,
) -> list[str]:
    cabinet_ids = {str(item.get("wbCabinetId") or "").strip() for item in results}
    if "" in cabinet_ids:
        return [f"{code}_source_cabinet_missing"]
    cabinets = (
        {
            item.id: item
            for item in db.scalars(
                select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids))
            )
        }
        if cabinet_ids
        else {}
    )
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return [f"{code}_source_scope_mismatch"]
    return []


def _measurement_row_scope_errors(
    db: Session,
    report: ReportRun,
    rows: Iterable[Mapping[str, Any]],
    *,
    code: str,
) -> list[str]:
    cabinet_ids = {str(item.get("wb_cabinet_id") or "").strip() for item in rows}
    if "" in cabinet_ids:
        return [f"{code}_source_cabinet_missing"]
    cabinets = (
        {
            item.id: item
            for item in db.scalars(
                select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids))
            )
        }
        if cabinet_ids
        else {}
    )
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return [f"{code}_source_scope_mismatch"]
    return []


def _verify_measurement_manifest(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
    source_type: str,
    expected_metadata: Mapping[str, Any],
) -> None:
    payload = collection.payload or {}
    if (payload.get("rawIntegrity") or {}).get("status") != "verified":
        raise RawIntegrityError("measurement raw integrity is not verified")
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("measurement results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type=source_type,
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    manifest = json.loads((raw_dir / "manifest.json").read_text(encoding="utf-8"))
    if not isinstance(manifest, Mapping) or manifest.get("source") != source_type:
        raise RawIntegrityError("measurement manifest source mismatch")
    for field, expected in expected_metadata.items():
        if manifest.get(field) != expected:
            raise RawIntegrityError("measurement manifest metadata mismatch")


def _iter_file_authoritative_measurement_rows(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
    source_type: str,
    tenant_id: str,
    client_id: str,
) -> Iterable[dict[str, Any]]:
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("measurement results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type=source_type,
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    count = 0
    for result in results:
        if not isinstance(result, Mapping):
            raise RawIntegrityError("measurement result is not an object")
        output_name = str(result.get("flatOutputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise RawIntegrityError("measurement flat output path is unsafe")
        output_path = (raw_dir / output_name).resolve()
        if not output_path.is_relative_to(raw_dir.resolve()):
            raise RawIntegrityError("measurement flat output path is unsafe")
        rows = json.loads(output_path.read_text(encoding="utf-8"))
        if not isinstance(rows, list):
            raise RawIntegrityError("measurement flat rows are not a list")
        cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for row in rows:
            if not isinstance(row, Mapping):
                raise RawIntegrityError("measurement row is not an object")
            count += 1
            yield _measurement_input(
                row,
                tenant_id=tenant_id,
                client_id=client_id,
                wb_cabinet_id=cabinet_id,
            )
    if count != collection.row_count:
        raise RawIntegrityError("measurement row count changed")


def _measurement_input(
    payload: Mapping[str, Any],
    *,
    tenant_id: str,
    client_id: str,
    wb_cabinet_id: str,
) -> dict[str, Any]:
    keys = (
        "nm_id",
        "dim_id",
        "prc_over",
        "volume",
        "width",
        "length",
        "height",
        "volume_sup",
        "width_sup",
        "length_sup",
        "height_sup",
        "dt_bonus",
        "is_valid",
        "is_valid_dt",
        "penalty_amount",
        "reversal_amount",
        "dt",
    )
    normalized = {
        "tenant_id": tenant_id,
        "client_id": client_id,
        "wb_cabinet_id": str(wb_cabinet_id or "").strip(),
        **{key: payload.get(key) for key in keys},
    }
    normalized["source_hash"] = _hash_payload(normalized)
    return normalized


@dataclass(frozen=True)
class _TariffSnapshotSelection:
    tariff_rows: tuple[dict[str, Any], ...] = ()
    source_snapshot_hash: str = ""
    source_loaded_at: datetime | None = None
    factor_snapshot_date: date | None = None
    source_row_count: int = 0
    blocking_reasons: tuple[str, ...] = ()
    review_reasons: tuple[str, ...] = ()


def _build_and_persist_logistics_tariffs(
    db: Session,
    report: ReportRun,
    *,
    logistics_result: LogisticsAnalysisResult,
    primary_refresh_run: SourceRefreshRun | None = None,
    base_refresh_run: SourceRefreshRun | None = None,
    contributing_runs: Iterable[SourceRefreshRun] = (),
) -> None:
    roles: list[tuple[int, SourceRefreshRun]] = []
    if primary_refresh_run is not None:
        roles.append((0, primary_refresh_run))
    if base_refresh_run is not None and all(
        run.id != base_refresh_run.id for _, run in roles
    ):
        roles.append((1, base_refresh_run))
    for run in contributing_runs:
        if all(existing.id != run.id for _, existing in roles):
            roles.append((2, run))
    selection = _select_tariff_snapshot(db, report, roles=roles)
    blocking = list(selection.blocking_reasons)
    review = list(selection.review_reasons)
    if logistics_result.context.data_status == "blocked":
        blocking.append("logistics_analysis_blocked")
    rows = (
        build_tariff_rows(
            logistics_result.sku_rows,
            selection.tariff_rows,
            factor_snapshot_date=selection.factor_snapshot_date,
        )
        if not blocking
        else []
    )
    expected_points = len(
        {
            (
                row.tenant_id,
                row.client_id,
                row.wb_cabinet_id,
                row.client_company_id,
                row.scheme,
                row.financial_week_start,
                tariff_type,
            )
            for row in logistics_result.sku_rows
            for tariff_type in ("box", "pallet")
        }
    )
    point_evidence: dict[tuple[Any, ...], set[str]] = defaultdict(set)
    unavailable_points: set[tuple[Any, ...]] = set()
    for row in rows:
        point = (
            row["wb_cabinet_id"],
            row["client_company_id"],
            row["scheme"],
            row["financial_week_start"],
            row["tariff_type"],
        )
        if row["coverage_status"] == "ready":
            point_evidence[point].add(str(row["evidence_type"]))
        else:
            unavailable_points.add(point)
    factual_points = sum("fact" in values for values in point_evidence.values())
    estimated_points = sum(
        "fact" not in values and "estimate" in values
        for values in point_evidence.values()
    )
    unavailable_count = max(
        expected_points - factual_points - estimated_points,
        len(unavailable_points - set(point_evidence)),
    )
    invalid = sum(row["coverage_status"] == "invalid_tariff" for row in rows)
    conflicting = sum(row["coverage_status"] == "conflicting_tariff" for row in rows)
    warehouses = len({row["warehouse"] for row in rows if row["warehouse"]})
    if estimated_points:
        review.append("tariff_archive_estimate_used")
    if unavailable_count:
        review.append("tariff_values_unavailable")
    if invalid:
        review.append("tariff_values_invalid")
    if conflicting:
        review.append("tariff_values_conflicting")
    data_status = "blocked" if blocking else "partial" if review else "ready"
    input_hash = _hash_payload(
        {
            "factorMethodologyVersion": LOGISTICS_TARIFFS_METHODOLOGY_VERSION,
            "sourceSnapshotHash": selection.source_snapshot_hash,
            "factorSnapshotDate": selection.factor_snapshot_date,
            "tariffSourceHashes": sorted(
                str(row.get("source_hash") or "") for row in selection.tariff_rows
            ),
            "skuSourceHashes": sorted(
                row.source_hash_digest for row in logistics_result.sku_rows
            ),
            "blockingReasons": sorted(set(blocking)),
            "reviewReasons": sorted(set(review)),
        }
    )
    repository.replace_report_logistics_tariff_analysis(
        db,
        report,
        context={
            "tenant_id": report.tenant_id,
            "client_id": report.client_id,
            "factor_methodology_version": LOGISTICS_TARIFFS_METHODOLOGY_VERSION,
            "data_status": data_status,
            "input_hash": input_hash,
            "source_snapshot_hash": selection.source_snapshot_hash,
            "source_loaded_at": selection.source_loaded_at,
            "factor_snapshot_date": selection.factor_snapshot_date,
            "source_row_count": selection.source_row_count,
            "tariff_row_count": len(rows),
            "expected_point_count": expected_points,
            "factual_point_count": factual_points,
            "estimated_point_count": estimated_points,
            "unavailable_point_count": unavailable_count,
            "invalid_row_count": invalid,
            "conflicting_row_count": conflicting,
            "warehouse_count": warehouses,
            "blocking_reasons": sorted(set(blocking)),
            "review_reasons": sorted(set(review)),
            "created_at": datetime.now(tz=ZoneInfo("UTC")),
        },
        rows=rows,
    )


def _select_tariff_snapshot(
    db: Session,
    report: ReportRun,
    *,
    roles: Iterable[tuple[int, SourceRefreshRun]],
) -> _TariffSnapshotSelection:
    candidates: list[tuple[int, SourceRefreshRun, SourceRefreshCollection]] = []
    for priority, run in roles:
        for collection in run.collections:
            if collection.source_type == "wb_tariffs":
                candidates.append((priority, run, collection))
    if not candidates:
        return _TariffSnapshotSelection(review_reasons=("tariff_source_missing",))
    selected_priority = min(item[0] for item in candidates)
    selected = [item for item in candidates if item[0] == selected_priority]
    snapshot_hashes = {item[2].snapshot_hash for item in selected}
    if len(snapshot_hashes) != 1:
        return _TariffSnapshotSelection(
            blocking_reasons=("tariff_source_revision_conflict",)
        )
    _priority, run, collection = sorted(
        selected,
        key=lambda item: (
            _dimension_loaded_at_timestamp(item[2].loaded_at),
            item[1].id,
            item[2].id,
        ),
        reverse=True,
    )[0]
    if (
        collection.tenant_id != report.tenant_id
        or collection.client_id != report.client_id
    ):
        return _TariffSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=("tariff_source_scope_mismatch",),
        )
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list) or not all(
        isinstance(item, Mapping) for item in results
    ):
        return _TariffSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("tariff_source_manifest_invalid",),
        )
    if _hash_payload(results) != collection.snapshot_hash:
        return _TariffSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("tariff_source_snapshot_hash_mismatch",),
        )
    try:
        declared_row_count = sum(int(item.get("rowCount") or 0) for item in results)
    except (TypeError, ValueError):
        declared_row_count = -1
    if declared_row_count != collection.row_count:
        return _TariffSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("tariff_source_manifest_row_count_mismatch",),
        )
    try:
        factor_snapshot_date = date.fromisoformat(
            str(payload.get("factorSnapshotDate") or "")
        )
    except ValueError:
        factor_snapshot_date = None
    snapshots = list(
        db.scalars(
            select(SourceSnapshotRow)
            .where(
                SourceSnapshotRow.refresh_run_id == run.id,
                SourceSnapshotRow.collection_id == collection.id,
                SourceSnapshotRow.source_type == "wb_tariffs",
            )
            .order_by(SourceSnapshotRow.row_number)
        )
    )
    persistence = payload.get("rowPersistence") or {}
    file_authoritative = (
        persistence.get("status") in {"file_authoritative", "skipped_large_snapshot"}
        and persistence.get("rawFilesAuthoritative") is True
    )
    if file_authoritative and snapshots:
        return _TariffSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            factor_snapshot_date=factor_snapshot_date,
            source_row_count=collection.row_count,
            blocking_reasons=("tariff_source_storage_ambiguity",),
        )
    blocking: list[str] = []
    tariff_rows: list[dict[str, Any]] = []
    if file_authoritative:
        try:
            tariff_rows.extend(
                _iter_file_authoritative_tariff_rows(collection, refresh_run=run)
            )
        except (OSError, TypeError, ValueError, RawIntegrityError):
            blocking.append("tariff_file_snapshot_invalid")
    else:
        if len(snapshots) != collection.row_count:
            blocking.append("tariff_database_row_count_mismatch")
        for snapshot in snapshots:
            row_payload = snapshot.row_payload
            if not isinstance(row_payload, Mapping):
                blocking.append("tariff_source_payload_invalid")
                continue
            if snapshot.raw_payload_hash != _hash_payload(row_payload):
                blocking.append("tariff_source_payload_hash_mismatch")
                continue
            tariff_rows.append(
                _tariff_input(row_payload, wb_cabinet_id=snapshot.wb_cabinet_id)
            )
    try:
        if (payload.get("rawIntegrity") or {}).get("status") != "verified":
            raise RawIntegrityError("tariff raw integrity is not verified")
        raw_dir = Path(collection.raw_path)
        source_root = Path(run.root_dir) if run.root_dir else raw_dir
        verify_raw_directory(
            raw_dir,
            source_type="wb_tariffs",
            source_root=source_root,
            collection_results=[item for item in results if isinstance(item, Mapping)],
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except (OSError, TypeError, ValueError, RawIntegrityError):
        blocking.append("tariff_raw_snapshot_invalid")
    blocking.extend(_tariff_scope_errors(db, report, tariff_rows))
    review: list[str] = []
    if factor_snapshot_date is None:
        review.append("tariff_factor_snapshot_date_missing")
    if collection.status == "partial_source":
        review.append("tariff_source_partial")
    if collection.status not in {"loaded", "partial_source"}:
        review.append("tariff_source_unavailable")
    return _TariffSnapshotSelection(
        tariff_rows=tuple(tariff_rows) if not blocking else (),
        source_snapshot_hash=collection.snapshot_hash,
        source_loaded_at=collection.loaded_at,
        factor_snapshot_date=factor_snapshot_date,
        source_row_count=len(tariff_rows),
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        review_reasons=tuple(dict.fromkeys(review)),
    )


def _iter_file_authoritative_tariff_rows(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
) -> Iterable[dict[str, Any]]:
    payload = collection.payload or {}
    if (payload.get("rawIntegrity") or {}).get("status") != "verified":
        raise RawIntegrityError("tariff raw integrity is not verified")
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("tariff results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type="wb_tariffs",
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    count = 0
    for result in results:
        if not isinstance(result, Mapping):
            raise RawIntegrityError("tariff result is not an object")
        output_name = str(result.get("flatOutputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise RawIntegrityError("tariff flat output path is unsafe")
        output_path = (raw_dir / output_name).resolve()
        if not output_path.is_relative_to(raw_dir.resolve()):
            raise RawIntegrityError("tariff flat output path is unsafe")
        flat = json.loads(output_path.read_text(encoding="utf-8"))
        if not isinstance(flat, Mapping):
            raise RawIntegrityError("tariff flat payload is not an object")
        wb_cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for tariff_type in ("box", "pallet"):
            rows = flat.get(tariff_type)
            if not isinstance(rows, list):
                raise RawIntegrityError("tariff flat rows are not a list")
            for row in rows:
                if not isinstance(row, Mapping):
                    raise RawIntegrityError("tariff row is not an object")
                count += 1
                yield _tariff_input(
                    {**row, "tariff_type": tariff_type},
                    wb_cabinet_id=wb_cabinet_id,
                )
    if count != collection.row_count:
        raise RawIntegrityError("tariff row count changed")


def _tariff_input(
    payload: Mapping[str, Any],
    *,
    wb_cabinet_id: str,
) -> dict[str, Any]:
    keys = (
        "requested_date",
        "source_tariff_date",
        "tariff_type",
        "warehouse_name",
        "geo_name",
        "dt_next_box",
        "dt_next_pallet",
        "dt_till_max",
        "box_delivery_base",
        "box_delivery_liter",
        "box_delivery_coef_expr",
        "box_delivery_marketplace_base",
        "box_delivery_marketplace_liter",
        "box_delivery_marketplace_coef_expr",
        "box_storage_base",
        "box_storage_liter",
        "box_storage_coef_expr",
        "pallet_delivery_expr",
        "pallet_delivery_value_base",
        "pallet_delivery_value_liter",
        "pallet_storage_expr",
        "pallet_storage_value_expr",
    )
    normalized = {
        "wb_cabinet_id": str(wb_cabinet_id or "").strip(),
        **{key: payload.get(key) for key in keys},
    }
    normalized["requested_date"] = normalized.get("requested_date") or normalized.get(
        "source_tariff_date"
    )
    normalized["source_hash"] = _hash_payload(normalized)
    return normalized


def _tariff_scope_errors(
    db: Session,
    report: ReportRun,
    rows: Iterable[Mapping[str, Any]],
) -> list[str]:
    cabinet_ids = {str(row.get("wb_cabinet_id") or "") for row in rows}
    if "" in cabinet_ids:
        return ["tariff_source_cabinet_missing"]
    cabinets = (
        {
            item.id: item
            for item in db.scalars(
                select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids))
            )
        }
        if cabinet_ids
        else {}
    )
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return ["tariff_source_scope_mismatch"]
        company = db.get(ClientCompany, cabinet.client_company_id)
        if (
            company is None
            or company.tenant_id != report.tenant_id
            or company.client_id != report.client_id
        ):
            return ["tariff_source_scope_mismatch"]
    return []


@dataclass(frozen=True)
class _RouteSnapshotSelection:
    route_rows: tuple[dict[str, Any], ...] = ()
    source_snapshot_hash: str = ""
    source_loaded_at: datetime | None = None
    source_coverage_start: date | None = None
    source_coverage_end: date | None = None
    source_row_count: int = 0
    blocking_reasons: tuple[str, ...] = ()
    review_reasons: tuple[str, ...] = ()


def _build_and_persist_logistics_routes(
    db: Session,
    report: ReportRun,
    *,
    logistics_result: LogisticsAnalysisResult,
    primary_refresh_run: SourceRefreshRun | None = None,
    base_refresh_run: SourceRefreshRun | None = None,
    contributing_runs: Iterable[SourceRefreshRun] = (),
) -> None:
    roles: list[tuple[int, SourceRefreshRun]] = []
    if primary_refresh_run is not None:
        roles.append((0, primary_refresh_run))
    if base_refresh_run is not None and all(
        run.id != base_refresh_run.id for _, run in roles
    ):
        roles.append((1, base_refresh_run))
    for run in contributing_runs:
        if all(existing.id != run.id for _, existing in roles):
            roles.append((2, run))
    selection = _select_route_snapshot(db, report, roles=roles)
    blocking = list(selection.blocking_reasons)
    review = list(selection.review_reasons)
    if logistics_result.context.data_status == "blocked":
        blocking.append("logistics_analysis_blocked")

    tariff_rows = [
        {
            "wb_cabinet_id": row.wb_cabinet_id,
            "client_company_id": row.client_company_id,
            "scheme": row.scheme,
            "financial_week_start": row.financial_week_start,
            "tariff_type": row.tariff_type,
            "warehouse": row.warehouse,
            "delivery_coefficient_pct": row.delivery_coefficient_pct,
            "evidence_type": row.evidence_type,
            "coverage_status": row.coverage_status,
            "source_hash_digest": row.source_hash_digest,
        }
        for row in db.scalars(
            select(ReportLogisticsTariffRow).where(
                ReportLogisticsTariffRow.report_run_id == report.id
            )
        )
    ]
    rows = (
        build_route_rows(
            logistics_result.order_rows,
            selection.route_rows,
            tariff_rows,
        )
        if not blocking
        else []
    )
    total_logistics = sum(
        (row.logistics_total for row in logistics_result.order_rows),
        Decimal("0"),
    )
    route_logistics = sum(
        (Decimal(row["logistics_total"]) for row in rows),
        Decimal("0"),
    )
    reconciliation_delta = route_logistics - total_logistics
    if not blocking and abs(reconciliation_delta) > Decimal("0.01"):
        blocking.append("route_logistics_reconciliation_failed")
        rows = []

    by_chain: dict[str, str] = {}
    for row in rows:
        chain_key = str(row.get("chain_key") or "")
        status = str(row.get("coverage_status") or "")
        if chain_key:
            by_chain[chain_key] = status
    total_chains = len({row.chain_key for row in logistics_result.order_rows})
    matched = sum(status == "ready" for status in by_chain.values())
    conflicting = sum(status == "conflicting_route" for status in by_chain.values())
    missing = max(0, total_chains - matched - conflicting)
    linked_logistics = sum(
        (
            Decimal(row["logistics_total"])
            for row in rows
            if row.get("coverage_status") == "ready"
        ),
        Decimal("0"),
    )
    warehouses = len(
        {
            str(row.get("warehouse") or "")
            for row in rows
            if row.get("warehouse_status") == "ready"
        }
    )
    destinations = len(
        {
            str(row.get("destination") or "")
            for row in rows
            if row.get("destination_status") == "ready"
        }
    )
    if missing:
        review.append("route_values_missing")
    if conflicting:
        review.append("route_values_conflicting")
    data_status = "blocked" if blocking else "partial" if review else "ready"
    input_hash = _hash_payload(
        {
            "factorMethodologyVersion": LOGISTICS_ROUTES_METHODOLOGY_VERSION,
            "sourceSnapshotHash": selection.source_snapshot_hash,
            "routeSourceHashes": sorted(
                str(row.get("source_hash") or "") for row in selection.route_rows
            ),
            "orderSourceHashes": sorted(
                row.source_hash_digest for row in logistics_result.order_rows
            ),
            "tariffSourceHashes": sorted(
                str(row.get("source_hash_digest") or "") for row in tariff_rows
            ),
            "blockingReasons": sorted(set(blocking)),
            "reviewReasons": sorted(set(review)),
        }
    )
    repository.replace_report_logistics_route_analysis(
        db,
        report,
        context={
            "tenant_id": report.tenant_id,
            "client_id": report.client_id,
            "factor_methodology_version": LOGISTICS_ROUTES_METHODOLOGY_VERSION,
            "data_status": data_status,
            "input_hash": input_hash,
            "source_snapshot_hash": selection.source_snapshot_hash,
            "source_loaded_at": selection.source_loaded_at,
            "source_coverage_start": selection.source_coverage_start,
            "source_coverage_end": selection.source_coverage_end,
            "source_row_count": selection.source_row_count,
            "route_row_count": len(rows),
            "total_chain_count": total_chains,
            "matched_chain_count": matched,
            "missing_chain_count": missing,
            "conflicting_chain_count": conflicting,
            "warehouse_count": warehouses,
            "destination_count": destinations,
            "total_logistics": total_logistics,
            "linked_logistics": linked_logistics,
            "reconciliation_delta": reconciliation_delta,
            "blocking_reasons": sorted(set(blocking)),
            "review_reasons": sorted(set(review)),
            "created_at": datetime.now(tz=ZoneInfo("UTC")),
        },
        rows=rows,
    )


def _select_route_snapshot(
    db: Session,
    report: ReportRun,
    *,
    roles: Iterable[tuple[int, SourceRefreshRun]],
) -> _RouteSnapshotSelection:
    candidates: list[tuple[int, SourceRefreshRun, SourceRefreshCollection]] = []
    for priority, run in roles:
        for collection in run.collections:
            if collection.source_type == "wb_supplier_sales":
                candidates.append((priority, run, collection))
    if not candidates:
        return _RouteSnapshotSelection(review_reasons=("route_source_missing",))
    selected_priority = min(item[0] for item in candidates)
    selected = [item for item in candidates if item[0] == selected_priority]
    if len({item[2].snapshot_hash for item in selected}) != 1:
        return _RouteSnapshotSelection(
            blocking_reasons=("route_source_revision_conflict",)
        )
    _priority, run, collection = sorted(
        selected,
        key=lambda item: (
            _dimension_loaded_at_timestamp(item[2].loaded_at),
            item[1].id,
            item[2].id,
        ),
        reverse=True,
    )[0]
    if (
        collection.tenant_id != report.tenant_id
        or collection.client_id != report.client_id
    ):
        return _RouteSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            blocking_reasons=("route_source_scope_mismatch",),
        )
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list) or not all(
        isinstance(item, Mapping) for item in results
    ):
        return _RouteSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("route_source_manifest_invalid",),
        )
    if _hash_payload(results) != collection.snapshot_hash:
        return _RouteSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("route_source_snapshot_hash_mismatch",),
        )
    try:
        declared_count = sum(int(item.get("rowCount") or 0) for item in results)
    except (TypeError, ValueError):
        declared_count = -1
    if declared_count != collection.row_count:
        return _RouteSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("route_source_manifest_row_count_mismatch",),
        )
    manifest_scope = _route_manifest_scope_errors(db, report, results)
    if manifest_scope:
        return _RouteSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=tuple(manifest_scope),
        )
    snapshots = list(
        db.scalars(
            select(SourceSnapshotRow)
            .where(
                SourceSnapshotRow.refresh_run_id == run.id,
                SourceSnapshotRow.collection_id == collection.id,
                SourceSnapshotRow.source_type == "wb_supplier_sales",
            )
            .order_by(SourceSnapshotRow.row_number)
        )
    )
    persistence = payload.get("rowPersistence") or {}
    file_authoritative = (
        persistence.get("status") in {"file_authoritative", "skipped_large_snapshot"}
        and persistence.get("rawFilesAuthoritative") is True
    )
    if file_authoritative and snapshots:
        return _RouteSnapshotSelection(
            source_snapshot_hash=collection.snapshot_hash,
            source_loaded_at=collection.loaded_at,
            source_row_count=collection.row_count,
            blocking_reasons=("route_source_storage_ambiguity",),
        )
    blocking: list[str] = []
    route_rows: list[dict[str, Any]] = []
    if file_authoritative:
        try:
            route_rows.extend(
                _iter_file_authoritative_route_rows(
                    collection,
                    refresh_run=run,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                )
            )
        except (OSError, TypeError, ValueError, RawIntegrityError):
            blocking.append("route_file_snapshot_invalid")
    else:
        if len(snapshots) != collection.row_count:
            blocking.append("route_database_row_count_mismatch")
        for snapshot in snapshots:
            row_payload = snapshot.row_payload
            if not isinstance(row_payload, Mapping):
                blocking.append("route_source_payload_invalid")
                continue
            if snapshot.raw_payload_hash != _hash_payload(row_payload):
                blocking.append("route_source_payload_hash_mismatch")
                continue
            route_rows.append(
                _route_input(
                    row_payload,
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                    wb_cabinet_id=snapshot.wb_cabinet_id,
                )
            )
    try:
        if (payload.get("rawIntegrity") or {}).get("status") != "verified":
            raise RawIntegrityError("route raw integrity is not verified")
        raw_dir = Path(collection.raw_path)
        source_root = Path(run.root_dir) if run.root_dir else raw_dir
        verify_raw_directory(
            raw_dir,
            source_type="wb_supplier_sales",
            source_root=source_root,
            collection_results=[item for item in results if isinstance(item, Mapping)],
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except (OSError, TypeError, ValueError, RawIntegrityError):
        blocking.append("route_raw_snapshot_invalid")
    blocking.extend(_route_scope_errors(db, report, route_rows))
    coverage_start = _safe_iso_date(payload.get("coverageStart"))
    coverage_end = _safe_iso_date(payload.get("coverageEnd"))
    review: list[str] = []
    if coverage_start is None or coverage_end is None:
        review.append("route_source_coverage_missing")
    elif coverage_start > report.period_start or coverage_end < report.period_end:
        review.append("route_source_period_partial")
    if collection.status == "partial_source":
        review.append("route_source_partial")
    if collection.status not in {"loaded", "partial_source"}:
        review.append("route_source_unavailable")
    return _RouteSnapshotSelection(
        route_rows=tuple(route_rows) if not blocking else (),
        source_snapshot_hash=collection.snapshot_hash,
        source_loaded_at=collection.loaded_at,
        source_coverage_start=coverage_start,
        source_coverage_end=coverage_end,
        source_row_count=len(route_rows),
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        review_reasons=tuple(dict.fromkeys(review)),
    )


def _iter_file_authoritative_route_rows(
    collection: SourceRefreshCollection,
    *,
    refresh_run: SourceRefreshRun,
    tenant_id: str,
    client_id: str,
) -> Iterable[dict[str, Any]]:
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list):
        raise RawIntegrityError("route results are missing")
    raw_dir = Path(collection.raw_path)
    source_root = Path(refresh_run.root_dir) if refresh_run.root_dir else raw_dir
    verify_raw_directory(
        raw_dir,
        source_type="wb_supplier_sales",
        source_root=source_root,
        collection_results=[item for item in results if isinstance(item, Mapping)],
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    count = 0
    for result in results:
        if not isinstance(result, Mapping):
            raise RawIntegrityError("route result is not an object")
        output_name = str(result.get("flatOutputFile") or "").strip()
        if not output_name:
            continue
        if Path(output_name).name != output_name:
            raise RawIntegrityError("route flat output path is unsafe")
        output_path = (raw_dir / output_name).resolve()
        if not output_path.is_relative_to(raw_dir.resolve()):
            raise RawIntegrityError("route flat output path is unsafe")
        rows = json.loads(output_path.read_text(encoding="utf-8"))
        if not isinstance(rows, list):
            raise RawIntegrityError("route flat rows are not a list")
        cabinet_id = str(result.get("wbCabinetId") or "").strip()
        for row in rows:
            if not isinstance(row, Mapping):
                raise RawIntegrityError("route row is not an object")
            count += 1
            yield _route_input(
                row,
                tenant_id=tenant_id,
                client_id=client_id,
                wb_cabinet_id=cabinet_id,
            )
    if count != collection.row_count:
        raise RawIntegrityError("route row count changed")


def _route_input(
    payload: Mapping[str, Any],
    *,
    tenant_id: str,
    client_id: str,
    wb_cabinet_id: str,
) -> dict[str, Any]:
    keys = (
        "srid",
        "nm_id",
        "warehouse_name",
        "country_name",
        "oblast_okrug_name",
        "region_name",
        "sale_date",
        "last_change_date",
    )
    normalized = {
        "tenant_id": tenant_id.strip(),
        "client_id": client_id.strip(),
        "wb_cabinet_id": wb_cabinet_id.strip(),
        **{key: payload.get(key) for key in keys},
    }
    normalized["source_hash"] = _hash_payload(normalized)
    return normalized


def _safe_iso_date(value: Any) -> date | None:
    try:
        return date.fromisoformat(str(value or ""))
    except ValueError:
        return None


def _route_manifest_scope_errors(
    db: Session,
    report: ReportRun,
    results: Iterable[Mapping[str, Any]],
) -> list[str]:
    cabinet_ids = {str(item.get("wbCabinetId") or "").strip() for item in results}
    if "" in cabinet_ids:
        return ["route_source_cabinet_missing"]
    return _route_cabinet_scope_errors(db, report, cabinet_ids)


def _route_scope_errors(
    db: Session,
    report: ReportRun,
    rows: Iterable[Mapping[str, Any]],
) -> list[str]:
    cabinet_ids = {str(row.get("wb_cabinet_id") or "").strip() for row in rows}
    if "" in cabinet_ids:
        return ["route_source_cabinet_missing"]
    return _route_cabinet_scope_errors(db, report, cabinet_ids)


def _route_cabinet_scope_errors(
    db: Session,
    report: ReportRun,
    cabinet_ids: set[str],
) -> list[str]:
    cabinets = (
        {
            item.id: item
            for item in db.scalars(
                select(WbCabinet).where(WbCabinet.id.in_(cabinet_ids))
            )
        }
        if cabinet_ids
        else {}
    )
    for cabinet_id in cabinet_ids:
        cabinet = cabinets.get(cabinet_id)
        if (
            cabinet is None
            or cabinet.tenant_id != report.tenant_id
            or cabinet.client_id != report.client_id
        ):
            return ["route_source_scope_mismatch"]
        company = db.get(ClientCompany, cabinet.client_company_id)
        if (
            company is None
            or company.tenant_id != report.tenant_id
            or company.client_id != report.client_id
        ):
            return ["route_source_scope_mismatch"]
    return []


def _select_logistics_source_rows(
    db: Session,
    report: ReportRun,
    *,
    roles: Mapping[str, tuple[SourceRefreshRun, str]],
    primary_refresh_run: SourceRefreshRun | None,
    base_refresh_run: SourceRefreshRun | None,
) -> tuple[list[LogisticsSourceRow], LogisticsInputDiagnostics]:
    if not roles:
        return [], LogisticsInputDiagnostics()
    run_ids = sorted(roles)
    candidates: list[_LogisticsSnapshotCandidate] = []
    lineage: list[dict[str, Any]] = []
    invalid_payload_shapes = 0
    source_identity_errors = 0
    source_revision_conflicts = 0
    source_revision_discarded = 0
    scope_mismatches = 0
    blocking: list[str] = []
    multiple_runs = len(run_ids) > 1
    database_row_counts = {
        str(run_id): int(row_count)
        for run_id, row_count in db.execute(
            select(
                SourceSnapshotRow.refresh_run_id,
                func.count(SourceSnapshotRow.id),
            )
            .where(
                SourceSnapshotRow.refresh_run_id.in_(run_ids),
                SourceSnapshotRow.source_type == "wb_finance_detail",
            )
            .group_by(SourceSnapshotRow.refresh_run_id)
        )
    }
    file_collections: dict[str, tuple[SourceRefreshRun, SourceRefreshCollection]] = {}
    database_run_ids: list[str] = []
    for run_id in run_ids:
        run, _role = roles[run_id]
        collection = _file_authoritative_wb_collection(run)
        if collection is None:
            database_run_ids.append(run_id)
            continue
        if int(database_row_counts.get(run_id, 0)):
            blocking.append("source_storage_ambiguity")
            continue
        file_collections[run_id] = (run, collection)
    snapshot_rows = db.execute(
        select(
            SourceSnapshotRow.id.label("snapshot_id"),
            SourceSnapshotRow.refresh_run_id,
            SourceSnapshotRow.tenant_id,
            SourceSnapshotRow.client_id,
            SourceSnapshotRow.wb_cabinet_id,
            SourceSnapshotRow.source_type,
            SourceSnapshotRow.source_row_id,
            SourceSnapshotRow.row_number,
            SourceSnapshotRow.raw_payload_hash,
            SourceSnapshotRow.row_payload,
            SourceSnapshotRow.loaded_at,
            SourceSnapshotRow.collection_id,
        )
        .where(
            SourceSnapshotRow.refresh_run_id.in_(database_run_ids),
            SourceSnapshotRow.source_type == "wb_finance_detail",
        )
        .order_by(
            SourceSnapshotRow.loaded_at,
            SourceSnapshotRow.refresh_run_id,
            SourceSnapshotRow.collection_id,
            SourceSnapshotRow.row_number,
        )
        .execution_options(stream_results=True, yield_per=1_000)
    ).yield_per(1_000)

    def snapshot_inputs() -> Iterable[_LogisticsSnapshotInput]:
        for snapshot in snapshot_rows:
            yield _LogisticsSnapshotInput(
                snapshot_id=str(snapshot.snapshot_id),
                refresh_run_id=str(snapshot.refresh_run_id),
                tenant_id=str(snapshot.tenant_id or ""),
                client_id=str(snapshot.client_id or ""),
                wb_cabinet_id=str(snapshot.wb_cabinet_id or ""),
                source_type=str(snapshot.source_type or ""),
                source_row_id=str(snapshot.source_row_id or ""),
                row_number=int(snapshot.row_number),
                raw_payload_hash=str(snapshot.raw_payload_hash or ""),
                row_payload=snapshot.row_payload,
                loaded_at=snapshot.loaded_at,
                collection_id=int(snapshot.collection_id),
            )
        for run_id in sorted(file_collections):
            run, collection = file_collections[run_id]
            try:
                yield from _iter_file_authoritative_logistics_inputs(
                    collection,
                    refresh_run=run,
                )
            except (OSError, TypeError, ValueError, SourceRefreshConfigError):
                blocking.append("file_authoritative_snapshot_invalid")

    for snapshot in snapshot_inputs():
        _refresh_run, role = roles[snapshot.refresh_run_id]
        payload = snapshot.row_payload
        stored_source_hash = str(snapshot.raw_payload_hash or "")
        canonical_source_hash = _hash_payload(payload)
        source_hash_matches = stored_source_hash == canonical_source_hash
        if not source_hash_matches:
            blocking.append("source_payload_hash_mismatch")
        base_record = {
            "runId": snapshot.refresh_run_id,
            "role": role,
            "sourceRowId": str(snapshot.source_row_id or ""),
            "sourceHash": canonical_source_hash,
            "storedSourceHash": stored_source_hash,
            "sourceHashStatus": "ready" if source_hash_matches else "invalid",
            "cabinetId": str(snapshot.wb_cabinet_id or ""),
            "rowNumber": int(snapshot.row_number),
        }
        if not isinstance(payload, Mapping):
            invalid_payload_shapes += 1
            blocking.append("invalid_source_payload_shape")
            lineage.append(
                {
                    **base_record,
                    "selection": "invalid",
                    "reason": "invalid_source_payload_shape",
                    "payloadType": type(payload).__name__,
                }
            )
            continue
        wb_cabinet_id = str(snapshot.wb_cabinet_id or "").strip()
        cabinet = db.get(WbCabinet, wb_cabinet_id)
        company = (
            db.get(ClientCompany, cabinet.client_company_id)
            if cabinet is not None and cabinet.client_company_id
            else None
        )
        cabinet_scope_ready = bool(
            cabinet is not None
            and cabinet.tenant_id == report.tenant_id
            and cabinet.client_id == report.client_id
        )
        company_scope_ready = bool(
            company is not None
            and company.tenant_id == report.tenant_id
            and company.client_id == report.client_id
        )
        if not cabinet_scope_ready or not company_scope_ready:
            scope_mismatches += 1
            blocking.append("tenant_scope_mismatch")
        client_company_id = (
            str(cabinet.client_company_id or "") if cabinet_scope_ready else ""
        )
        source_hash = canonical_source_hash
        source_row = source_row_from_payload(
            payload,
            tenant_id=str(snapshot.tenant_id or ""),
            client_id=str(snapshot.client_id or ""),
            wb_cabinet_id=wb_cabinet_id,
            client_company_id=client_company_id,
            source_row_id=snapshot.source_row_id,
            source_hash=source_hash,
            fallback_date=report.period_start,
        )
        provider_row_id = _first_row_id(payload, "rrdId", "rrd_id")
        snapshot_row_id = str(snapshot.source_row_id or "").strip()
        stable_identity = bool(
            provider_row_id and snapshot_row_id and provider_row_id == snapshot_row_id
        )
        if not snapshot_row_id or (multiple_runs and not provider_row_id):
            source_identity_errors += 1
            blocking.append("source_identity_missing")
        elif provider_row_id and provider_row_id != snapshot_row_id:
            source_identity_errors += 1
            blocking.append("source_identity_mismatch")
        candidates.append(
            _LogisticsSnapshotCandidate(
                snapshot_id=str(snapshot.snapshot_id),
                refresh_run_id=str(snapshot.refresh_run_id),
                tenant_id=str(snapshot.tenant_id or ""),
                client_id=str(snapshot.client_id or ""),
                wb_cabinet_id=wb_cabinet_id,
                source_type=str(snapshot.source_type or ""),
                source_row_id=snapshot_row_id,
                loaded_at=snapshot.loaded_at,
                collection_id=int(snapshot.collection_id),
                row_number=int(snapshot.row_number),
                role=role,
                source_row=source_row,
                source_hash=source_hash,
                stable_identity=stable_identity,
            )
        )

    daily_owners: dict[date, set[str]] = {}
    for fact_date, run_id in db.execute(
        select(
            MarketplaceFinanceDailyFactModel.fact_date,
            MarketplaceFinanceDailyFactModel.source_refresh_run_id,
        ).where(
            MarketplaceFinanceDailyFactModel.tenant_id == report.tenant_id,
            MarketplaceFinanceDailyFactModel.client_id == report.client_id,
            MarketplaceFinanceDailyFactModel.marketplace == "wb",
            MarketplaceFinanceDailyFactModel.fact_date >= report.period_start,
            MarketplaceFinanceDailyFactModel.fact_date <= report.period_end,
            MarketplaceFinanceDailyFactModel.source_refresh_run_id.in_(run_ids),
        )
    ):
        daily_owners.setdefault(fact_date, set()).add(str(run_id))

    candidates_by_date: dict[date, set[str]] = {}
    for candidate in candidates:
        financial_date = candidate.source_row.financial_date
        if financial_date is not None:
            candidates_by_date.setdefault(financial_date, set()).add(
                candidate.refresh_run_id
            )

    selected_candidates: list[_LogisticsSnapshotCandidate] = []
    owner_conflict_dates: set[date] = set()
    for candidate in candidates:
        financial_date = candidate.source_row.financial_date
        record = {
            "runId": candidate.refresh_run_id,
            "role": candidate.role,
            "sourceRowId": candidate.source_row_id,
            "sourceHash": candidate.source_hash,
            "cabinetId": candidate.wb_cabinet_id,
            "financialDate": (
                financial_date.isoformat() if financial_date is not None else None
            ),
        }
        if financial_date is not None and not (
            report.period_start <= financial_date <= report.period_end
        ):
            lineage.append({**record, "selection": "outside_report"})
            continue
        owner_id, owner_conflict = _logistics_owner_for_date(
            financial_date,
            candidate_run_ids=candidates_by_date.get(financial_date, set()),
            daily_owner_ids=daily_owners.get(financial_date, set()),
            primary_refresh_run=primary_refresh_run,
            base_refresh_run=base_refresh_run,
            roles=roles,
        )
        if owner_conflict and financial_date is not None:
            owner_conflict_dates.add(financial_date)
        if candidate.refresh_run_id != owner_id:
            source_revision_discarded += 1
            lineage.append(
                {
                    **record,
                    "selection": "discarded",
                    "reason": "lower_precedence_revision",
                    "ownerRunId": owner_id,
                }
            )
            continue
        selected_candidates.append(candidate)
        lineage.append({**record, "selection": "candidate"})
    if owner_conflict_dates:
        source_revision_conflicts += len(owner_conflict_dates)
        blocking.append("source_window_overlap_conflict")

    selected: list[LogisticsSourceRow] = []
    by_identity: dict[tuple[str, ...], list[_LogisticsSnapshotCandidate]] = {}
    for candidate in selected_candidates:
        identity = (
            candidate.tenant_id,
            candidate.client_id,
            candidate.wb_cabinet_id,
            candidate.source_type,
            candidate.source_row_id,
        )
        if not candidate.stable_identity:
            identity = (
                *identity,
                candidate.refresh_run_id,
                str(candidate.snapshot_id),
            )
        by_identity.setdefault(identity, []).append(candidate)
    for identity in sorted(by_identity):
        revisions = sorted(
            by_identity[identity],
            key=lambda item: (
                item.loaded_at,
                item.refresh_run_id,
                item.collection_id,
                item.row_number,
            ),
        )
        hashes = {item.source_hash for item in revisions}
        if len(hashes) > 1:
            source_revision_conflicts += 1
            blocking.append("source_revision_conflict")
        source_revision_discarded += max(0, len(revisions) - 1)
        selected.append(revisions[-1].source_row)

    diagnostics = LogisticsInputDiagnostics(
        invalid_source_payload_shape_count=invalid_payload_shapes,
        source_identity_error_count=source_identity_errors,
        source_revision_conflict_count=source_revision_conflicts,
        source_revision_discarded_count=source_revision_discarded,
        scope_mismatch_count=scope_mismatches,
        blocking_reasons=tuple(dict.fromkeys(blocking)),
        lineage_records=tuple(lineage),
    )
    return sorted(
        selected,
        key=lambda row: (
            row.financial_date or date.min,
            row.wb_cabinet_id,
            row.source_row_id,
            row.source_hash,
        ),
    ), diagnostics


def _logistics_owner_for_date(
    financial_date: date | None,
    *,
    candidate_run_ids: set[str],
    daily_owner_ids: set[str],
    primary_refresh_run: SourceRefreshRun | None,
    base_refresh_run: SourceRefreshRun | None,
    roles: Mapping[str, tuple[SourceRefreshRun, str]],
) -> tuple[str, bool]:
    if financial_date is not None and primary_refresh_run is not None:
        start = (
            primary_refresh_run.source_window_start or primary_refresh_run.period_start
        )
        end = primary_refresh_run.source_window_end or primary_refresh_run.period_end
        if start <= financial_date <= end:
            return primary_refresh_run.id, False
    if daily_owner_ids:
        owners = sorted(daily_owner_ids)
        return owners[-1], len(owners) > 1
    if financial_date is not None and base_refresh_run is not None:
        start = base_refresh_run.source_window_start or base_refresh_run.period_start
        end = base_refresh_run.source_window_end or base_refresh_run.period_end
        if start <= financial_date <= end:
            return base_refresh_run.id, False
    covering = []
    if financial_date is not None:
        for run_id, (run, _role) in roles.items():
            start = run.source_window_start or run.period_start
            end = run.source_window_end or run.period_end
            if start <= financial_date <= end and run_id in candidate_run_ids:
                covering.append(run_id)
    if covering:
        owners = sorted(set(covering))
        return owners[-1], len(owners) > 1
    if candidate_run_ids:
        owners = sorted(candidate_run_ids)
        return owners[-1], len(owners) > 1
    if primary_refresh_run is not None:
        return primary_refresh_run.id, False
    return sorted(roles)[-1], len(roles) > 1


def _report_unit_row_scope_matches(
    db: Session,
    report: ReportRun,
    row: ReportUnitRow,
) -> bool:
    cabinet = db.get(WbCabinet, row.wb_cabinet_id)
    company = db.get(ClientCompany, row.client_company_id)
    return bool(
        cabinet is not None
        and cabinet.tenant_id == report.tenant_id
        and cabinet.client_id == report.client_id
        and cabinet.client_company_id == row.client_company_id
        and company is not None
        and company.tenant_id == report.tenant_id
        and company.client_id == report.client_id
    )


def _report_logistics_decimal(value: Any) -> Decimal | None:
    try:
        return Decimal(value)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _persist_wb_report_list_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbSalesReportListPageResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            for local_index, row in enumerate(_read_json_list(result.output_path), 1):
                source_row_id = _first_row_id(row, "reportId")
                if not source_row_id:
                    source_row_id = f"{result.page_index}:{local_index}"
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row),
                        "row_payload": row,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": wb_cabinet_ids.get(
                            result.seller_account_id,
                            "",
                        ),
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_wb_redeem_notification_rows(
    db: Session,
    collection: SourceRefreshCollection,
    rows: Iterable[dict[str, Any]],
) -> None:
    batch: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, 1):
        report_id = str(row.get("reportId") or "").strip()
        batch.append(
            {
                "row_number": row_number,
                "raw_payload_hash": _hash_payload(row),
                "row_payload": row,
                "source_row_id": report_id or str(row_number),
                "wb_cabinet_id": str(row.get("wbCabinetId") or ""),
                "loaded_at": collection.loaded_at,
            }
        )
        _flush_snapshot_batch(db, collection, batch)
    _flush_snapshot_batch(db, collection, batch, force=True)


def _persist_wb_product_card_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbProductCardsPageResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            if result.flat_output_path is None:
                continue
            for local_index, row in enumerate(
                _read_json_list(result.flat_output_path),
                1,
            ):
                source_row_id = _first_row_id(
                    row,
                    "nm_id",
                    "barcode",
                    "vendor_code",
                    "chrt_id",
                )
                if not source_row_id:
                    source_row_id = f"{result.page_index}:{local_index}"
                row_payload = {
                    **row,
                    "marketplace": "wb",
                    "seller_account_id": result.seller_account_id,
                    "source_page_index": result.page_index,
                    "source_cards_source": result.cards_source,
                    "source_output_file": result.flat_output_path.name,
                }
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row_payload),
                        "row_payload": row_payload,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": wb_cabinet_ids.get(
                            result.seller_account_id,
                            "",
                        ),
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_wb_tariff_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbTariffsExportResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            if not result.ok or result.flat_output_path is None:
                continue
            payload = json.loads(result.flat_output_path.read_text(encoding="utf-8"))
            if not isinstance(payload, Mapping):
                raise ValueError("tariff flat payload must be an object")
            for tariff_type in ("box", "pallet"):
                rows = payload.get(tariff_type)
                if not isinstance(rows, list):
                    raise ValueError("tariff flat rows must be a list")
                for local_index, row in enumerate(rows, 1):
                    if not isinstance(row, Mapping):
                        raise ValueError("tariff flat row must be an object")
                    row_payload = {
                        **row,
                        "marketplace": "wb",
                        "tariff_type": tariff_type,
                        "source_tariff_date": (
                            result.target_date.isoformat() if result.target_date else ""
                        ),
                        "source_output_file": result.flat_output_path.name,
                    }
                    warehouse = str(row.get("warehouse_name") or "").strip()
                    source_row_id = ":".join(
                        (
                            row_payload["source_tariff_date"],
                            tariff_type,
                            warehouse,
                            str(local_index),
                        )
                    )
                    batch.append(
                        {
                            "row_number": row_number,
                            "raw_payload_hash": _hash_payload(row_payload),
                            "row_payload": row_payload,
                            "source_row_id": source_row_id,
                            "wb_cabinet_id": wb_cabinet_ids.get(
                                result.seller_account_id,
                                "",
                            ),
                            "loaded_at": collection.loaded_at,
                        }
                    )
                    row_number += 1
                    _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_wb_goods_return_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbGoodsReturnExportResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            if not result.ok or result.flat_output_path is None:
                continue
            rows = json.loads(result.flat_output_path.read_text(encoding="utf-8"))
            if not isinstance(rows, list):
                raise ValueError("goods-return flat rows must be a list")
            cabinet_id = wb_cabinet_ids.get(result.seller_account_id, "")
            for local_index, row in enumerate(rows, 1):
                if not isinstance(row, Mapping):
                    raise ValueError("goods-return flat row must be an object")
                row_payload = {
                    **row,
                    "marketplace": "wb",
                    "source_output_file": result.flat_output_path.name,
                }
                source_row_id = _hash_payload(
                    {
                        "wbCabinetId": cabinet_id,
                        "srid": row.get("srid"),
                        "nmId": row.get("nm_id"),
                        "rowNumber": local_index,
                    }
                )
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row_payload),
                        "row_payload": row_payload,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": cabinet_id,
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_wb_return_claim_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbReturnClaimsExportResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            if not result.ok or result.flat_output_path is None:
                continue
            rows = json.loads(result.flat_output_path.read_text(encoding="utf-8"))
            if not isinstance(rows, list):
                raise ValueError("return-claims flat rows must be a list")
            cabinet_id = wb_cabinet_ids.get(result.seller_account_id, "")
            for local_index, row in enumerate(rows, 1):
                if not isinstance(row, Mapping):
                    raise ValueError("return-claims flat row must be an object")
                forbidden = {
                    "id",
                    "user_comment",
                    "wb_comment",
                    "origin_id_info",
                    "photos",
                    "photo",
                    "videos",
                    "video",
                    "actions",
                }
                if forbidden.intersection(row):
                    raise ValueError("return-claims flat row contains raw fields")
                row_payload = {
                    **row,
                    "marketplace": "wb",
                    "source_output_file": result.flat_output_path.name,
                }
                source_row_id = _hash_payload(
                    {
                        "wbCabinetId": cabinet_id,
                        "srid": row.get("srid"),
                        "nmId": row.get("nm_id"),
                        "isArchive": row.get("is_archive"),
                        "rowNumber": local_index,
                    }
                )
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row_payload),
                        "row_payload": row_payload,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": cabinet_id,
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_wb_supplier_sales_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbSupplierSalesExportResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            if not result.ok or result.flat_output_path is None:
                continue
            rows = json.loads(result.flat_output_path.read_text(encoding="utf-8"))
            if not isinstance(rows, list):
                raise ValueError("supplier-sales flat rows must be a list")
            for local_index, row in enumerate(rows, 1):
                if not isinstance(row, Mapping):
                    raise ValueError("supplier-sales flat row must be an object")
                row_payload = {
                    **row,
                    "marketplace": "wb",
                    "source_output_file": result.flat_output_path.name,
                }
                source_row_id = ":".join(
                    (
                        str(row.get("srid") or ""),
                        str(row.get("nm_id") or ""),
                        str(local_index),
                    )
                )
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row_payload),
                        "row_payload": row_payload,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": wb_cabinet_ids.get(
                            result.seller_account_id,
                            "",
                        ),
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_wb_measurement_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[WbMeasurementExportResult],
    *,
    wb_cabinet_ids: dict[str, str],
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for result in results:
            if not result.ok or result.flat_output_path is None:
                continue
            rows = json.loads(result.flat_output_path.read_text(encoding="utf-8"))
            if not isinstance(rows, list):
                raise ValueError("measurement flat rows must be a list")
            for local_index, row in enumerate(rows, 1):
                if not isinstance(row, Mapping):
                    raise ValueError("measurement flat row must be an object")
                row_payload = {
                    **row,
                    "marketplace": "wb",
                    "measurement_source_type": result.source_type,
                    "source_output_file": result.flat_output_path.name,
                }
                source_row_id = ":".join(
                    (
                        str(row.get("dim_id") or ""),
                        str(row.get("nm_id") or ""),
                        str(local_index),
                    )
                )
                batch.append(
                    {
                        "row_number": row_number,
                        "raw_payload_hash": _hash_payload(row_payload),
                        "row_payload": row_payload,
                        "source_row_id": source_row_id,
                        "wb_cabinet_id": wb_cabinet_ids.get(
                            result.seller_account_id,
                            "",
                        ),
                        "loaded_at": collection.loaded_at,
                    }
                )
                row_number += 1
                _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError, json.JSONDecodeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_ozon_rows(
    db: Session,
    collection: SourceRefreshCollection,
    results: Iterable[OzonPageResult],
    *,
    ozon_cabinet_ids: dict[str, str],
) -> None:
    try:
        batch: list[dict[str, Any]] = []
        for value in _iter_ozon_snapshot_row_values(
            collection,
            results,
            ozon_cabinet_ids=ozon_cabinet_ids,
        ):
            batch.append(value)
            _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _iter_ozon_snapshot_row_values(
    collection: SourceRefreshCollection,
    results: Iterable[OzonPageResult],
    *,
    ozon_cabinet_ids: dict[str, str],
) -> Iterable[dict[str, Any]]:
    row_number = 1
    for result in results:
        if _is_ozon_report_control_result(result):
            continue
        for local_index, row in enumerate(_read_ozon_rows(result.output_path), 1):
            source_row_id = (
                _first_row_id(
                    row,
                    "id",
                    "operation_id",
                    "posting_number",
                    "product_id",
                    "offer_id",
                    "sku",
                    "code",
                    "ID товара",
                    "Идентификатор товара",
                    "Артикул",
                    "Артикул продавца",
                    "SKU",
                    "Штрихкод",
                    "Баркод",
                )
                or f"{result.source_type}:{result.page_index}:{local_index}"
            )
            row_payload = {
                **row,
                "marketplace": "ozon",
                "seller_account_id": result.seller_account_id,
                "source_endpoint": result.source_endpoint,
                "source_page_index": result.page_index,
                "source_output_file": (
                    result.output_path.name if result.output_path else ""
                ),
            }
            yield {
                "row_number": row_number,
                "raw_payload_hash": _hash_payload(row_payload),
                "row_payload": row_payload,
                "source_row_id": source_row_id,
                "wb_cabinet_id": ozon_cabinet_ids.get(
                    result.seller_account_id,
                    "",
                ),
                "loaded_at": collection.loaded_at,
            }
            row_number += 1


def _ozon_results_from_collection(
    collection: SourceRefreshCollection,
    *,
    source_root: Path,
) -> tuple[list[OzonPageResult], dict[str, str]]:
    payload = dict(collection.payload or {})
    raw_results = payload.get("results")
    if not isinstance(raw_results, list):
        raise RawIntegrityError("collection results are missing")
    verified = verify_raw_directory(
        Path(collection.raw_path),
        source_type=collection.source_type,
        source_root=source_root,
        collection_results=raw_results,
        collection_row_count=collection.row_count,
        collection_snapshot_hash=collection.snapshot_hash,
    )
    collection.payload = {**payload, "rawIntegrity": verified.as_payload()}
    raw_dir = Path(collection.raw_path).resolve()
    results: list[OzonPageResult] = []
    cabinet_ids: dict[str, str] = {}
    for item in raw_results:
        if not isinstance(item, dict):
            raise RawIntegrityError("collection result is invalid")
        seller_account_id = str(item.get("sellerAccountId") or "")
        cabinet_id = str(item.get("wbCabinetId") or "")
        if seller_account_id:
            cabinet_ids[seller_account_id] = cabinet_id
        output_name = str(item.get("outputFile") or "")
        if output_name and Path(output_name).name != output_name:
            raise RawIntegrityError("collection output path is unsafe")
        source_endpoint = str(item.get("sourceEndpoint") or "")
        source_type = str(item.get("sourceType") or "")
        if not source_type:
            if source_endpoint == "/v1/report/info":
                source_type = f"{collection.source_type}_info"
            elif source_endpoint == "report_file":
                source_type = f"{collection.source_type}_file"
            else:
                source_type = collection.source_type
        results.append(
            OzonPageResult(
                source_type=source_type,
                seller_account_id=seller_account_id,
                account_name=str(item.get("accountName") or ""),
                page_index=int(item.get("pageIndex") or 0),
                ok=bool(item.get("ok")),
                status=str(item.get("sourceStatus") or item.get("status") or ""),
                row_count=int(item.get("rowCount") or 0),
                raw_payload_hash=str(item.get("rawPayloadHash") or ""),
                raw_content_sha256=str(item.get("rawContentSha256") or ""),
                output_path=(raw_dir / output_name) if output_name else None,
                status_code=(
                    int(item["statusCode"])
                    if item.get("statusCode") is not None
                    else None
                ),
                error=str(item.get("error") or ""),
                report_code=str(item.get("reportCode") or ""),
                report_status=str(item.get("reportStatus") or ""),
                page_limit_exhausted=bool(item.get("pageLimitExhausted")),
                source_endpoint=source_endpoint,
            )
        )
    return results, cabinet_ids


def _materialize_ozon_typed_collection(
    db: Session,
    refresh_run: SourceRefreshRun,
    collection: SourceRefreshCollection,
    results: Iterable[OzonPageResult],
    *,
    ozon_cabinet_ids: dict[str, str],
) -> None:
    payload = collection.payload or {}
    raw_integrity = payload.get("rawIntegrity") or {}
    if raw_integrity.get("status") != "verified":
        collection.payload = {
            **payload,
            "typedParity": {
                "status": "blocked_raw_integrity",
                "reason": str(raw_integrity.get("error") or "not_verified"),
            },
        }
        db.flush()
        raise ValueError("ozon_raw_integrity_not_verified")
    result_items = list(results)
    file_operation_rows = list(
        _iter_ozon_operation_facts(
            result_items,
            ozon_cabinet_ids=ozon_cabinet_ids,
            client_id=refresh_run.client_id,
        )
    )
    operation_rows = _merge_ozon_operation_facts(file_operation_rows)
    legacy_operation_rows = list(
        _iter_legacy_ozon_operation_facts(
            db,
            collection,
            ozon_cabinet_ids=ozon_cabinet_ids,
            client_id=refresh_run.client_id,
        )
    )
    legacy_parity = _ozon_legacy_file_parity(
        file_operation_rows,
        legacy_operation_rows,
    )
    qualification_run_id = str(
        (payload.get("rowPersistence") or {}).get("qualificationRunId") or ""
    )
    if legacy_parity.get("status") == "not_run_no_legacy_rows" and qualification_run_id:
        legacy_parity = {
            "status": "qualified_file_reference",
            "legacyRows": 0,
            "qualificationRunId": qualification_run_id,
            "fileDigest": _ozon_operation_rows_digest(operation_rows),
        }
    operation_count = repository.replace_marketplace_operation_facts(
        db,
        collection,
        operation_rows,
    )
    persistence_parity = repository.marketplace_operation_facts_parity(
        db,
        collection,
        operation_rows,
    )
    reconstructed_source_rows = sum(
        1
        for _value in _iter_ozon_snapshot_row_values(
            collection,
            result_items,
            ozon_cabinet_ids=ozon_cabinet_ids,
        )
    )
    source_coverage = {
        "status": (
            "matched"
            if collection.status in MANDATORY_OK_STATUSES
            and reconstructed_source_rows == collection.row_count
            and len(file_operation_rows) >= reconstructed_source_rows
            else "mismatch"
        ),
        "expectedRows": collection.row_count,
        "reconstructedRows": reconstructed_source_rows,
        "normalizedRows": len(file_operation_rows),
        "collectionStatus": collection.status,
    }
    parity_ready = (
        persistence_parity.get("status") == "matched"
        and legacy_parity.get("status")
        in {"matched", "not_run_no_legacy_rows", "qualified_file_reference"}
        and source_coverage["status"] == "matched"
    )
    typed_parity = {
        **persistence_parity,
        "status": "pending_diagnostics" if parity_ready else "mismatch",
        "sourceRows": collection.row_count,
        "operationRowsBeforeMerge": len(file_operation_rows),
        "operationRowsAfterMerge": len(operation_rows),
        "persistenceParity": persistence_parity,
        "legacyFileParity": legacy_parity,
        "sourceCoverage": source_coverage,
    }
    collection.payload = {
        **(collection.payload or {}),
        "operationFacts": {
            "status": "materialized",
            "rowCount": operation_count,
        },
        "typedParity": typed_parity,
    }
    db.flush()


def _iter_ozon_operation_facts(
    results: Iterable[OzonPageResult],
    *,
    ozon_cabinet_ids: dict[str, str],
    client_id: str = "",
) -> Iterable[dict[str, Any]]:
    for result in results:
        if _is_ozon_report_control_result(result):
            continue
        for local_index, row in enumerate(_read_ozon_rows(result.output_path), 1):
            yield from _ozon_operation_facts_for_row(
                result,
                row,
                local_index=local_index,
                ozon_cabinet_ids=ozon_cabinet_ids,
                client_id=client_id,
            )


def _iter_legacy_ozon_operation_facts(
    db: Session,
    collection: SourceRefreshCollection,
    *,
    ozon_cabinet_ids: dict[str, str],
    client_id: str = "",
) -> Iterable[dict[str, Any]]:
    rows = db.scalars(
        select(SourceSnapshotRow)
        .where(SourceSnapshotRow.collection_id == collection.id)
        .order_by(SourceSnapshotRow.row_number)
    )
    augmentation_keys = {
        "marketplace",
        "seller_account_id",
        "source_endpoint",
        "source_page_index",
        "source_output_file",
    }
    for row in rows:
        payload = dict(row.row_payload or {})
        seller_account_id = str(payload.get("seller_account_id") or "")
        source_endpoint = str(payload.get("source_endpoint") or "")
        if (
            collection.source_type == "ozon_mutual_settlement"
            and source_endpoint != "report_file"
        ):
            continue
        page_index = int(payload.get("source_page_index") or 0)
        source_payload = {
            key: value for key, value in payload.items() if key not in augmentation_keys
        }
        result = OzonPageResult(
            source_type=collection.source_type,
            seller_account_id=seller_account_id,
            account_name="",
            page_index=page_index,
            ok=True,
            status="ok",
            row_count=1,
            source_endpoint=source_endpoint,
        )
        yield from _ozon_operation_facts_for_row(
            result,
            source_payload,
            local_index=row.row_number,
            ozon_cabinet_ids=ozon_cabinet_ids,
            client_id=client_id,
        )


def _ozon_operation_facts_for_row(
    result: OzonPageResult,
    row: dict[str, Any],
    *,
    local_index: int,
    ozon_cabinet_ids: dict[str, str],
    client_id: str = "",
) -> Iterable[dict[str, Any]]:
    items = _ozon_operation_items(result, row)
    for item in items:
        fact = _ozon_operation_fact(
            result,
            row=row,
            item=item,
            local_index=local_index,
            ozon_cabinet_ids=ozon_cabinet_ids,
            client_id=client_id,
        )
        if result.source_type.removesuffix("_file") == "ozon_finance_cash_flow":
            yield from _ozon_cash_flow_operation_facts(fact, item)
            continue
        if not result.source_type.startswith("ozon_realization"):
            yield fact
            continue
        expense_fields = {
            "commission": Decimal(fact["commission"]),
            "service_amount": Decimal(fact["service_amount"]),
            "logistics": Decimal(fact["logistics"]),
            "storage": Decimal(fact["storage"]),
            "promotion": Decimal(fact["promotion"]),
            "compensation": Decimal(fact["compensation"]),
            "other_amount": Decimal(fact["other_amount"]),
        }
        product_fact = {
            **fact,
            "service_key": "product",
            "service_name": "",
            **{field: Decimal("0") for field in expense_fields},
        }
        product_fact["source_key"] = _ozon_fact_source_key(product_fact)
        yield product_fact
        for service_key, amount in expense_fields.items():
            if amount == 0:
                continue
            service_fact = {
                **fact,
                "service_key": service_key,
                "service_name": service_key,
                "quantity": Decimal("0"),
                "amount": Decimal("0"),
                "price": Decimal("0"),
                "income": Decimal("0"),
                "expense": Decimal("0"),
                **{field: Decimal("0") for field in expense_fields},
            }
            service_fact[service_key] = amount
            service_fact["source_key"] = _ozon_fact_source_key(service_fact)
            yield service_fact


def _ozon_cash_flow_operation_facts(
    fact: dict[str, Any],
    item: dict[str, Any],
) -> Iterable[dict[str, Any]]:
    summary = {
        **fact,
        "service_key": "cash_flow_summary",
        "service_name": "cash_flow_summary",
    }
    summary["source_key"] = _ozon_fact_source_key(summary)
    yield summary
    for category in (
        "delivery",
        "return",
        "loan",
        "invoice_transfer",
        "rfbs",
        "services",
        "others",
    ):
        payload = item.get(category)
        if not isinstance(payload, dict):
            continue
        total = _decimal_value(payload, "total", "amount", "price")
        category_fact = {
            **fact,
            "service_key": f"cash_flow_category:{category}",
            "service_name": category,
            "operation_type": "cash_flow_category",
            "quantity": Decimal("0"),
            "amount": total,
            "price": Decimal("0"),
            "income": Decimal("0"),
            "expense": Decimal("0"),
            "commission": Decimal("0"),
            "service_amount": Decimal("0"),
            "logistics": Decimal("0"),
            "storage": Decimal("0"),
            "promotion": Decimal("0"),
            "compensation": Decimal("0"),
            "other_amount": Decimal("0"),
            "raw_payload_hash": _hash_payload(payload),
            "_stable_payload_hash": _ozon_stable_payload_hash(payload),
        }
        category_fact["source_key"] = _ozon_fact_source_key(category_fact)
        yield category_fact
        lines = payload.get("items")
        if not isinstance(lines, list):
            continue
        for line in lines:
            if not isinstance(line, dict):
                continue
            line_name = _text_value(line, "name", "operation_type", "type")
            line_fact = {
                **category_fact,
                "service_key": (
                    f"cash_flow_item:{category}:{_ozon_stable_payload_hash(line)}"
                ),
                "service_name": line_name or category,
                "operation_type": "cash_flow_item",
                "amount": _decimal_value(line, "price", "amount", "total"),
                "raw_payload_hash": _hash_payload(line),
                "_stable_payload_hash": _ozon_stable_payload_hash(line),
            }
            line_fact["source_key"] = _ozon_fact_source_key(line_fact)
            yield line_fact


def _ozon_legacy_file_parity(
    file_rows: list[dict[str, Any]],
    legacy_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    if not legacy_rows:
        return {"status": "not_run_no_legacy_rows", "legacyRows": 0}
    file_values = _merge_ozon_operation_facts(file_rows)
    legacy_values = _merge_ozon_operation_facts(legacy_rows)
    file_digest = _ozon_operation_rows_digest(file_values)
    legacy_digest = _ozon_operation_rows_digest(legacy_values)
    matched = len(file_values) == len(legacy_values) and file_digest == legacy_digest
    return {
        "status": "matched" if matched else "mismatch",
        "fileRows": len(file_rows),
        "legacyRows": len(legacy_rows),
        "fileFacts": len(file_values),
        "legacyFacts": len(legacy_values),
        "fileDigest": file_digest,
        "legacyDigest": legacy_digest,
        "mismatches": [] if matched else ["normalizedOperations"],
    }


def _ozon_operation_rows_digest(rows: list[dict[str, Any]]) -> str:
    values = [
        {
            key: value
            for key, value in row.items()
            if not key.startswith("_")
            and key not in {"source_row_id", "source_row_number"}
        }
        for row in rows
    ]
    values.sort(key=lambda item: str(item.get("source_key") or ""))
    return _hash_payload(values)


def _ozon_operation_items(
    result: OzonPageResult,
    row: dict[str, Any],
) -> list[dict[str, Any]]:
    source_type = result.source_type.removesuffix("_file")
    if source_type == "ozon_finance_cash_flow":
        cash_flows = row.get("cash_flows")
        if not isinstance(cash_flows, list):
            return [row]
        details = row.get("details")
        detail_by_period = (
            {
                str(item.get("period") or ""): item
                for item in details
                if isinstance(item, dict)
            }
            if isinstance(details, list)
            else {}
        )
        return [
            {
                **item,
                **detail_by_period.get(str(item.get("period") or ""), {}),
            }
            for item in cash_flows
            if isinstance(item, dict)
        ]
    if source_type in {"ozon_realization", "ozon_realization_posting"}:
        normalized = _iter_realization_items(row)
        order = row.get("order")
        if not isinstance(order, dict):
            return normalized
        return [{**item, **order} for item in normalized]
    if source_type == "ozon_products_buyout":
        products = row.get("products")
        if isinstance(products, list):
            return [item for item in products if isinstance(item, dict)]
    if source_type == "ozon_b2b_sales_json":
        invoices = row.get("invoices")
        if not isinstance(invoices, list):
            return [row]
        result_items: list[dict[str, Any]] = []
        for invoice in invoices:
            if not isinstance(invoice, dict):
                continue
            info = invoice.get("info")
            invoice_info = info if isinstance(info, dict) else {}
            operations = invoice.get("operations")
            if not isinstance(operations, list) or not operations:
                result_items.append({**invoice, **invoice_info})
                continue
            for operation in operations:
                if isinstance(operation, dict):
                    result_items.append({**invoice, **invoice_info, **operation})
        return result_items
    return [row]


def _ozon_daily_facts_for_run(
    db: Session,
    refresh_run: SourceRefreshRun,
) -> list[MarketplaceFinanceDailyFactContract]:
    rows = list(
        db.scalars(
            select(MarketplaceOperationFact).where(
                MarketplaceOperationFact.tenant_id == refresh_run.tenant_id,
                MarketplaceOperationFact.client_id == refresh_run.client_id,
                MarketplaceOperationFact.marketplace == "ozon",
                MarketplaceOperationFact.source_refresh_run_id == refresh_run.id,
            )
        )
    )
    grouped: dict[tuple[object, ...], dict[str, Any]] = {}
    for row in rows:
        if row.operation_date is None:
            continue
        if not (
            refresh_run.period_start <= row.operation_date <= refresh_run.period_end
        ):
            continue
        operation_group = row.operation_type or row.service_key or "unknown"
        key = (
            row.seller_account_id,
            row.operation_date,
            row.source_type,
            row.product_id,
            row.offer_id,
            row.sku,
            row.barcode,
            operation_group,
        )
        bucket = grouped.setdefault(
            key,
            {
                "quantity": Decimal("0"),
                "amount": Decimal("0"),
                "commission": Decimal("0"),
                "logistics": Decimal("0"),
                "storage": Decimal("0"),
                "promotion": Decimal("0"),
                "other": Decimal("0"),
                "hashes": [],
                "partial": False,
            },
        )
        bucket["quantity"] += Decimal(row.quantity)
        bucket["amount"] += Decimal(row.income or row.amount)
        bucket["commission"] += Decimal(row.commission)
        bucket["logistics"] += Decimal(row.logistics)
        bucket["storage"] += Decimal(row.storage)
        bucket["promotion"] += Decimal(row.promotion)
        bucket["other"] += Decimal(row.other_amount) + Decimal(row.expense)
        if row.raw_payload_hash not in bucket["hashes"]:
            bucket["hashes"].append(row.raw_payload_hash)
        bucket["partial"] = bool(bucket["partial"] or row.is_partial_source)
    facts: list[MarketplaceFinanceDailyFactContract] = []
    for key, bucket in grouped.items():
        (
            seller_account_id,
            fact_date,
            source_type,
            product_id,
            offer_id,
            sku,
            barcode,
            operation_group,
        ) = key
        quantity = Decimal(bucket["quantity"])
        try:
            nm_id = int(str(product_id)) if str(product_id) else None
        except ValueError:
            nm_id = None
        source_hash_digest = hashlib.sha256(
            "\0".join(sorted(str(value) for value in bucket["hashes"])).encode("utf-8")
        ).hexdigest()
        facts.append(
            MarketplaceFinanceDailyFactContract(
                client_id=refresh_run.client_id,
                seller_account_id=str(seller_account_id),
                organization_id="",
                fact_date=fact_date,
                marketplace_report_id=str(source_type),
                document_kind=str(source_type),
                nm_id=nm_id,
                vendor_code=str(offer_id or sku),
                barcode=str(barcode),
                onec_item_id="",
                sales_model="",
                operation_group=str(operation_group),
                sales_quantity=max(quantity, Decimal("0")),
                return_quantity=abs(min(quantity, Decimal("0"))),
                quantity=quantity,
                net_revenue=Decimal(bucket["amount"]),
                wb_commission=Decimal(bucket["commission"]),
                logistics=Decimal(bucket["logistics"]),
                storage=Decimal(bucket["storage"]),
                marketplace_promotion=Decimal(bucket["promotion"]),
                penalties_and_holdbacks=Decimal(bucket["other"]),
                source_row_count=len(bucket["hashes"]),
                source_hash_digest=source_hash_digest,
                is_partial_source=bool(bucket["partial"]),
                methodology_version="ozon-typed-daily-v1",
            )
        )
    return facts


def _merge_ozon_operation_facts(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    numeric_fields = (
        "quantity",
        "amount",
        "price",
        "income",
        "expense",
        "debit_amount",
        "credit_amount",
        "commission",
        "service_amount",
        "logistics",
        "storage",
        "promotion",
        "compensation",
        "other_amount",
    )
    fallback_occurrences: dict[str, int] = {}
    for row in rows:
        key = str(row["source_key"])
        if _ozon_fact_uses_fallback(row):
            fallback_occurrences[key] = fallback_occurrences.get(key, 0) + 1
            key = _hash_payload(
                {
                    "businessKey": key,
                    "duplicateOrdinal": fallback_occurrences[key],
                }
            )
            row = {**row, "source_key": key}
        current = result.get(key)
        if current is None:
            result[key] = dict(row)
            continue
        for field in numeric_fields:
            current[field] = Decimal(current.get(field) or 0) + Decimal(
                row.get(field) or 0
            )
        current["source_row_number"] = min(
            int(current.get("source_row_number") or 0),
            int(row.get("source_row_number") or 0),
        )
        current["is_partial_source"] = bool(
            current.get("is_partial_source") or row.get("is_partial_source")
        )
        current["raw_payload_hash"] = _hash_payload(
            sorted(
                {
                    str(current.get("raw_payload_hash") or ""),
                    str(row.get("raw_payload_hash") or ""),
                }
            )
        )
    return list(result.values())


def _ozon_fact_uses_fallback(row: dict[str, Any]) -> bool:
    return bool(row.get("_preserve_occurrences")) or not (
        str(row.get("operation_id") or "") or str(row.get("posting_number") or "")
    )


def _ozon_operation_fact(
    result: OzonPageResult,
    *,
    row: dict[str, Any],
    item: dict[str, Any],
    local_index: int,
    ozon_cabinet_ids: dict[str, str],
    client_id: str = "",
) -> dict[str, Any]:
    source_row_id = _first_row_id(
        row,
        "id",
        "operation_id",
        "posting_number",
        "product_id",
        "offer_id",
        "sku",
        "code",
        "ID товара",
        "Идентификатор товара",
        "Артикул",
        "Артикул продавца",
        "SKU",
    ) or (
        f"{result.source_type.removesuffix('_file')}:{result.page_index}:{local_index}"
    )
    operation_id = _text_value(
        item, "operation_id", "operationId", "id", "Документ"
    ) or _text_value(row, "operation_id", "operationId", "id")
    posting_number = _text_value(
        item, "posting_number", "postingNumber", "Номер отправления"
    ) or _text_value(row, "posting_number", "postingNumber", "Номер отправления")
    product_id = _text_value(
        item,
        "product_id",
        "productId",
        "Ozon Product ID",
        "ID товара",
        "Идентификатор товара",
    )
    offer_id = _text_value(item, "offer_id", "offerId", "Артикул продавца", "Артикул")
    sku = _text_value(item, "sku", "SKU", "ozon_sku", "Ozon SKU")
    barcode = _text_value(
        item,
        "barcode",
        "Barcode",
        "Штрихкод",
        "Штрихкод (Серийный номер / EAN)",
        "Баркод",
    )
    product_name = _text_value(
        item,
        "product_name",
        "Product name",
        "name",
        "Название товара",
        "Наименование",
    )
    operation_type = _text_value(
        item,
        "operation_type",
        "operation_type_name",
        "operationType",
        "type",
        "Тип операции",
        "Название операции",
        "Наименование",
    ) or _text_value(
        row,
        "operation_type",
        "operation_type_name",
        "operationType",
        "type",
        "Тип операции",
        "Название операции",
    )
    service_key = _text_value(
        item,
        "service_key",
        "serviceKey",
        "service_code",
        "serviceCode",
        "service_name",
        "serviceName",
        "Название услуги",
    ) or ("aggregate" if result.source_type == "ozon_realization" else "operation")
    service_name = _text_value(
        item,
        "service_name",
        "serviceName",
        "Название услуги",
    )
    source_type = result.source_type.removesuffix("_file")
    preserve_occurrences = False
    if source_type == "ozon_mutual_settlement":
        service_name = operation_type or service_name
        service_key = f"mutual:{_ozon_stable_payload_hash(item)}"
        preserve_occurrences = True
    is_realization_source = result.source_type.startswith("ozon_realization")
    is_cash_flow_source = (
        result.source_type.removesuffix("_file") == "ozon_finance_cash_flow"
    )
    cash_period_start, cash_period_end = (
        _ozon_cash_flow_period(item) if is_cash_flow_source else (None, None)
    )
    if is_cash_flow_source and cash_period_start is not None:
        operation_id = (
            f"{cash_period_start.isoformat()}|"
            f"{(cash_period_end or cash_period_start).isoformat()}"
        )
    expenses, expenses_loaded = (
        _realization_expenses(item) if is_realization_source else ({}, False)
    )
    amount = (
        _realization_amount(item)
        if is_realization_source
        else _decimal_value(item, "orders_amount")
        - abs(_decimal_value(item, "returns_amount"))
        if is_cash_flow_source
        else _decimal_value(item, "amount", "accruals_for_sale", "Итого", "Сумма")
    )
    fact = {
        "client_id": client_id,
        "wb_cabinet_id": ozon_cabinet_ids.get(result.seller_account_id, ""),
        "seller_account_id": result.seller_account_id,
        "source_type": source_type,
        "source_key": "",
        "source_row_id": source_row_id,
        "source_row_number": local_index,
        "operation_id": operation_id,
        "posting_number": posting_number,
        "product_id": product_id,
        "offer_id": offer_id,
        "sku": sku,
        "service_key": service_key,
        "service_name": service_name,
        "barcode": barcode,
        "product_name": product_name,
        "operation_type": operation_type
        or ("cash_flow" if is_cash_flow_source else ""),
        "operation_date": cash_period_start
        or _date_value(
            item,
            "operation_date",
            "operationDate",
            "accrual_date",
            "sale_date",
            "created_date",
            "date",
            "Дата операции",
            "period",
        )
        or _date_value(row, "operation_date", "operationDate", "date"),
        "quantity": (
            _realization_quantity(item)
            if is_realization_source
            else _decimal_value(item, "quantity", "qty", "Количество")
        ),
        "amount": amount or Decimal("0"),
        "price": _decimal_value(
            item,
            "price",
            "seller_price_per_instance",
            "buyout_price",
            "Цена",
        ),
        "income": (
            _decimal_value(item, "orders_amount")
            if is_cash_flow_source
            else _decimal_value(
                item,
                "income",
                "revenue",
                "accruals_for_sale",
                "Начислено",
            )
        ),
        "expense": (
            abs(_decimal_value(item, "returns_amount"))
            if is_cash_flow_source
            else _decimal_value(
                item,
                "expense",
                "sale_commission",
                "delivery_charge",
                "return_delivery_charge",
                "Расход",
            )
        ),
        "debit_amount": _decimal_value(
            item,
            "debit_amount",
            "Сумма дебиторской задолженности, RUR",
        ),
        "credit_amount": _decimal_value(
            item,
            "credit_amount",
            "Сумма кредиторской задолженности, RUR",
        ),
        "commission": (
            _decimal_value(item, "commission_amount")
            if is_cash_flow_source
            else expenses.get("commission", Decimal("0"))
        ),
        "service_amount": (
            _decimal_value(item, "services_amount")
            if is_cash_flow_source
            else expenses.get("services", Decimal("0"))
        ),
        "logistics": (
            _decimal_value(item, "item_delivery_and_return_amount")
            if is_cash_flow_source
            else expenses.get("logistics", Decimal("0"))
        ),
        "storage": expenses.get("storage", Decimal("0")),
        "promotion": expenses.get("promotion", Decimal("0")),
        "compensation": expenses.get("compensation", Decimal("0")),
        "other_amount": expenses.get("other", Decimal("0"))
        + expenses.get("partner_services", Decimal("0")),
        "expenses_loaded": expenses_loaded,
        "is_partial_source": not (result.ok and result.status == "ok")
        or (is_realization_source and amount is None),
        "currency": _text_value(item, "currency", "currency_code", "Валюта") or "RUB",
        "source_endpoint": result.source_endpoint,
        "raw_payload_hash": _hash_payload(item),
        "_stable_payload_hash": _ozon_stable_payload_hash({"row": row, "item": item}),
        "_preserve_occurrences": preserve_occurrences,
    }
    fact["source_key"] = _ozon_fact_source_key(fact)
    return fact


def _ozon_fact_source_key(fact: dict[str, Any]) -> str:
    operation_id = str(fact.get("operation_id") or "")
    posting_number = str(fact.get("posting_number") or "")
    return _hash_payload(
        {
            "clientId": str(fact.get("client_id") or ""),
            "cabinetId": str(fact.get("wb_cabinet_id") or ""),
            "sellerAccountId": str(fact.get("seller_account_id") or ""),
            "sourceType": str(fact.get("source_type") or ""),
            "operationId": operation_id,
            "postingNumber": posting_number,
            "productId": str(fact.get("product_id") or ""),
            "offerId": str(fact.get("offer_id") or ""),
            "sku": str(fact.get("sku") or ""),
            "serviceKey": str(fact.get("service_key") or ""),
            "operationDate": (
                fact["operation_date"].isoformat()
                if isinstance(fact.get("operation_date"), date)
                else str(fact.get("operation_date") or "")
            ),
            "fallbackRawPayloadHash": (
                str(fact.get("_stable_payload_hash") or "")
                if not operation_id and not posting_number
                else ""
            ),
        }
    )


def _ozon_stable_payload_hash(payload: dict[str, Any]) -> str:
    return _hash_payload(
        {
            key: value
            for key, value in payload.items()
            if key.casefold().replace("_", "")
            not in {"rownumber", "sourcepagenumber", "sourcepageindex"}
        }
    )


def _ozon_cash_flow_period(
    payload: dict[str, Any],
) -> tuple[date | None, date | None]:
    period = payload.get("period")
    if isinstance(period, dict):
        period_start = _date_value(period, "begin", "start", "from")
        period_end = _date_value(period, "end", "to") or period_start
        return period_start, period_end
    period_start = _date_value(payload, "period")
    return period_start, period_start


def _text_value(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _decimal_value(row: dict[str, Any], *keys: str) -> Decimal:
    value = _text_value(row, *keys).replace(" ", "").replace(",", ".")
    if not value:
        return Decimal("0")
    try:
        return Decimal(value)
    except InvalidOperation:
        return Decimal("0")


def _date_value(row: dict[str, Any], *keys: str) -> date | None:
    value = _text_value(row, *keys)
    if not value:
        return None
    if len(value) == 7:
        with suppress(ValueError):
            return date.fromisoformat(f"{value}-01")
    with suppress(ValueError):
        return date.fromisoformat(value[:10])
    for pattern in ("%d.%m.%Y", "%d-%m-%Y"):
        with suppress(ValueError):
            return datetime.strptime(value[:10], pattern).date()
    return None


def _is_ozon_report_control_result(result: OzonPageResult) -> bool:
    if result.source_type.endswith("_info"):
        return True
    if result.source_type.endswith("_file"):
        return False
    if result.source_type == "ozon_mutual_settlement":
        return True
    endpoint = str(result.source_endpoint or "")
    return endpoint.startswith("/v1/report/") or endpoint == "/v1/report/info"


def _persist_onec_rows(
    db: Session,
    collection: SourceRefreshCollection,
    result: OnecSampleExportResult,
) -> None:
    if result.output_path is None:
        return
    try:
        byte_size = result.output_path.stat().st_size
    except OSError as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)
        return
    if byte_size > ONEC_DATABASE_ROW_PERSIST_MAX_BYTES:
        if result.sample_id == "commissioner_reports":
            try:
                rows = _extract_onec_rows(_read_json_object(result.output_path))
                compact_rows = [_compact_onec_commissioner_row(row) for row in rows]
                _persist_onec_snapshot_rows(
                    db,
                    collection,
                    result,
                    compact_rows,
                )
            except (OSError, ValueError, TypeError) as exc:
                _mark_raw_row_persistence_failure(db, collection, exc)
                return
            collection.payload = {
                **(collection.payload or {}),
                "rowPersistence": {
                    "status": "compacted_large_snapshot",
                    "limitBytes": ONEC_DATABASE_ROW_PERSIST_MAX_BYTES,
                    "byteSize": byte_size,
                    "rawFilesAuthoritative": True,
                    "persistedDocumentRows": len(compact_rows),
                },
            }
            db.flush()
            return
        collection.payload = {
            **(collection.payload or {}),
            "rowPersistence": {
                "status": "skipped_large_snapshot",
                "limitBytes": ONEC_DATABASE_ROW_PERSIST_MAX_BYTES,
                "byteSize": byte_size,
                "rawFilesAuthoritative": True,
            },
        }
        db.flush()
        return
    try:
        rows = _extract_onec_rows(_read_json_object(result.output_path))
        _persist_onec_snapshot_rows(db, collection, result, rows)
    except (OSError, ValueError, TypeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _persist_onec_snapshot_rows(
    db: Session,
    collection: SourceRefreshCollection,
    result: OnecSampleExportResult,
    rows: list[dict[str, Any]],
) -> None:
    batch: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, 1):
        source_row_id = _first_row_id(row, "Ref_Key", "LineNumber", "НомерСтроки")
        if not source_row_id:
            source_row_id = f"{result.sample_id}:{row_number}"
        batch.append(
            {
                "row_number": row_number,
                "raw_payload_hash": _hash_payload(row),
                "row_payload": row,
                "source_row_id": source_row_id,
                "loaded_at": collection.loaded_at,
            }
        )
        _flush_snapshot_batch(db, collection, batch)
    _flush_snapshot_batch(db, collection, batch, force=True)


def _compact_onec_commissioner_row(row: dict[str, Any]) -> dict[str, Any]:
    header_fields = (
        "Ref_Key",
        "Date",
        "Number",
        "Posted",
        "DeletionMark",
        "Комментарий",
        "НомерВходящегоДокумента",
        "Организация_Key",
        "Контрагент_Key",
        "СуммаДокумента",
        "СуммаДокументаВозврат",
        "СуммаДокументаСУчетомВознаграждения",
        "СуммаВознаграждения",
    )
    line_fields = (
        "Номенклатура_Key",
        "ХарактеристикаНоменклатуры_Key",
        "Количество",
        "Цена",
        "Сумма",
        "Всего",
        "СуммаНДС",
        "СтавкаНДС",
    )
    compact = {key: row[key] for key in header_fields if key in row}
    for table_name in ("Запасы", "ЗапасыВозвраты"):
        table = row.get(table_name)
        compact[table_name] = (
            [
                {key: item[key] for key in line_fields if key in item}
                for item in table
                if isinstance(item, dict)
            ]
            if isinstance(table, list)
            else []
        )
    return compact


def _persist_mapping_rows(
    db: Session,
    collection: SourceRefreshCollection,
    mapping_dir: Path,
) -> None:
    try:
        row_number = 1
        batch: list[dict[str, Any]] = []
        for path in sorted(item for item in mapping_dir.rglob("*") if item.is_file()):
            stat = path.stat()
            relative_path = str(path.relative_to(mapping_dir))
            row = {
                "path": relative_path,
                "name": path.name,
                "size": stat.st_size,
                "mtime": datetime.fromtimestamp(
                    stat.st_mtime,
                    tz=MOSCOW_TZ,
                ).isoformat(),
                "sha256": _file_sha256(path),
            }
            batch.append(
                {
                    "row_number": row_number,
                    "raw_payload_hash": _hash_payload(row),
                    "row_payload": row,
                    "source_row_id": relative_path,
                    "loaded_at": collection.loaded_at,
                }
            )
            row_number += 1
            _flush_snapshot_batch(db, collection, batch)
        _flush_snapshot_batch(db, collection, batch, force=True)
    except (OSError, ValueError, TypeError) as exc:
        _mark_raw_row_persistence_failure(db, collection, exc)


def _flush_snapshot_batch(
    db: Session,
    collection: SourceRefreshCollection,
    batch: list[dict[str, Any]],
    *,
    force: bool = False,
) -> None:
    if not batch or (not force and len(batch) < SOURCE_SNAPSHOT_ROW_CHUNK_SIZE):
        return
    repository.add_source_snapshot_rows(db, collection, batch)
    batch.clear()
    _commit_source_refresh_progress(db)


def _read_json_list(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError("raw JSON payload is not a list")
    return [item for item in payload if isinstance(item, dict)]


def _read_json_object(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("raw JSON payload is not an object")
    return payload


def _read_ozon_rows(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    if _path_has_xlsx_signature(path):
        return _read_ozon_xlsx_rows(path)
    if path.suffix.lower() == ".xlsx":
        return _read_ozon_xlsx_rows(path)
    if path.suffix.lower() in {".csv", ".tsv", ".txt"}:
        return _read_ozon_tabular_rows(path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    result = payload.get("result")
    if isinstance(result, list):
        return [item for item in result if isinstance(item, dict)]
    if isinstance(result, dict):
        cash_flows = result.get("cash_flows")
        if isinstance(cash_flows, list):
            details = result.get("details")
            detail_rows = (
                [item for item in details if isinstance(item, dict)]
                if isinstance(details, list)
                else [details]
                if isinstance(details, dict)
                else []
            )
            details_by_period = {
                _hash_payload(item.get("period")): item for item in detail_rows
            }
            return [
                {
                    **item,
                    **details_by_period.get(_hash_payload(item.get("period")), {}),
                }
                for item in cash_flows
                if isinstance(item, dict)
            ]
        for key in ("items", "rows", "data"):
            value = result.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
        details = result.get("details")
        if isinstance(details, dict):
            return [details]
        return [result]
    for key in ("items", "rows", "data"):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return [payload]


def _read_ozon_tabular_rows(path: Path) -> list[dict[str, Any]]:
    text = _read_text_with_encoding_fallback(path)
    delimiter = _tabular_delimiter(text)
    reader = csv.DictReader(StringIO(text, newline=""), delimiter=delimiter)
    return [
        {str(key or "").strip(): str(value or "").strip() for key, value in row.items()}
        for row in reader
    ]


def _read_ozon_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(
        BytesIO(path.read_bytes()),
        read_only=True,
        data_only=True,
    )
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return _rows_from_table_values(rows)


def _rows_from_table_values(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    normalized_rows = [
        [_cell_text(value) for value in values]
        for values in rows
        if any(_cell_text(value) for value in values)
    ]
    if not normalized_rows:
        return []
    header_index = max(
        range(len(normalized_rows)),
        key=lambda index: (sum(1 for value in normalized_rows[index] if value), -index),
    )
    header_values = normalized_rows[header_index]
    header: list[str] = [
        value if value else f"column_{index}"
        for index, value in enumerate(header_values, start=1)
    ]
    data_rows: list[dict[str, Any]] = []
    for text_values in normalized_rows[header_index + 1 :]:
        data_rows.append(
            {
                header[index]: text_values[index]
                for index in range(min(len(header), len(text_values)))
            }
        )
    return data_rows


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _path_has_xlsx_signature(path: Path) -> bool:
    try:
        with path.open("rb") as handle:
            return handle.read(4) == b"PK\x03\x04"
    except OSError:
        return False


def _read_text_with_encoding_fallback(path: Path) -> str:
    content = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-16", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _tabular_delimiter(text: str) -> str:
    sample = text[:4096]
    candidates = {
        "\t": sample.count("\t"),
        ";": sample.count(";"),
        ",": sample.count(","),
    }
    return max(candidates, key=candidates.get)


def _extract_onec_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    value = payload.get("value")
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    data = payload.get("d")
    if isinstance(data, dict):
        results = data.get("results")
        if isinstance(results, list):
            return [item for item in results if isinstance(item, dict)]
    if isinstance(data, list):
        return [item for item in data if isinstance(item, dict)]
    return []


def _first_row_id(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _stock_history_zip_covered_days(
    path: Path | None,
    *,
    period_start: date,
    period_end: date,
) -> int:
    if path is None or not path.exists():
        return 0
    expected = {
        date.fromordinal(day)
        for day in range(period_start.toordinal(), period_end.toordinal() + 1)
    }
    covered: set[date] = set()
    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                if name.endswith("/"):
                    continue
                text = archive.read(name).decode("utf-8-sig", errors="replace")
                reader = csv.reader(StringIO(text))
                headers = next(reader, [])
                for header in headers:
                    try:
                        parsed = datetime.strptime(header.strip(), "%d.%m.%Y").date()
                    except ValueError:
                        continue
                    if parsed in expected:
                        covered.add(parsed)
    except (OSError, zipfile.BadZipFile):
        return 0
    return len(covered)


def _subtract_calendar_months(value: date, months: int) -> date:
    absolute_month = value.year * 12 + value.month - 1 - max(0, months)
    year, zero_based_month = divmod(absolute_month, 12)
    month = zero_based_month + 1
    day = min(value.day, monthrange(year, month)[1])
    return date(year, month, day)


def _moscow_today() -> date:
    return datetime.now(tz=MOSCOW_TZ).date()


def _stock_history_provider_period(
    period_start: date,
    period_end: date,
    *,
    today: date | None = None,
) -> tuple[date, date] | None:
    earliest_available = _subtract_calendar_months(today or _moscow_today(), 3)
    actual_start = max(period_start, earliest_available)
    if actual_start > period_end:
        return None
    return actual_start, period_end


def _mark_raw_row_persistence_failure(
    db: Session,
    collection: SourceRefreshCollection,
    exc: Exception,
) -> None:
    collection.status = "failed" if collection.required else "needs_review"
    collection.error_message = f"raw_row_persistence_failed:{exc.__class__.__name__}"
    payload = dict(collection.payload or {})
    payload["rawRowPersistence"] = {
        "status": "failed",
        "errorType": exc.__class__.__name__,
    }
    collection.payload = payload
    db.flush()


def _hash_payload(payload: Any) -> str:
    content = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


def _attach_collection_raw_integrity(
    collection: SourceRefreshCollection,
    *,
    source_root: Path,
) -> None:
    payload = collection.payload or {}
    results = payload.get("results")
    if not isinstance(results, list):
        raise SourceRefreshConfigError(
            f"raw integrity results are missing for {collection.source_type}"
        )
    try:
        verified = verify_raw_directory(
            Path(collection.raw_path),
            source_type=collection.source_type,
            source_root=source_root,
            collection_results=results,
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except RawIntegrityError as exc:
        collection.payload = {
            **payload,
            "rawIntegrity": {
                "status": "failed",
                "error": str(exc),
            },
        }
        return
    collection.payload = {
        **payload,
        "rawIntegrity": verified.as_payload(),
    }


def _reverify_collection_raw_integrity(
    refresh_run: SourceRefreshRun,
    *,
    source_type: str,
    raw_path: Path,
    source_root: Path,
) -> None:
    collection = next(
        (
            item
            for item in refresh_run.collections
            if item.source_type == source_type
            and Path(item.raw_path).resolve() == raw_path.resolve()
        ),
        None,
    )
    if collection is None:
        raise SourceRefreshConfigError(
            f"registered raw collection is missing for {source_type}"
        )
    payload = collection.payload or {}
    if (payload.get("rawIntegrity") or {}).get("status") != "verified":
        return
    results = payload.get("results")
    if not isinstance(results, list):
        raise SourceRefreshConfigError(
            f"registered raw results are missing for {source_type}"
        )
    try:
        verified = verify_raw_directory(
            raw_path,
            source_type=source_type,
            source_root=source_root,
            collection_results=results,
            collection_row_count=collection.row_count,
            collection_snapshot_hash=collection.snapshot_hash,
        )
    except RawIntegrityError as exc:
        raise SourceRefreshConfigError(
            f"raw integrity changed before rebuild for {source_type}: {exc}"
        ) from exc
    collection.payload = {
        **payload,
        "rawIntegrity": verified.as_payload(),
    }


def _safe_error(exc: Exception) -> str:
    if isinstance(
        exc,
        (
            SourceRefreshConfigError,
            WbFinanceConfigError,
            OnecODataConfigError,
            integrations.IntegrationSecretError,
        ),
    ):
        return str(exc)
    message = _safe_error_message(str(exc))
    if not message:
        return exc.__class__.__name__
    return f"{exc.__class__.__name__}: {message}"


def _safe_error_message(message: str) -> str:
    safe = message.strip().replace("\n", " ")
    if not safe:
        return ""
    safe = re.sub(
        r"(?i)(authorization|api[_-]?key|token|password|secret)=([^\s&]+)",
        r"\1=<redacted>",
        safe,
    )
    safe = re.sub(
        r"(?i)(bearer|basic)\s+[a-z0-9._~+/=-]+",
        r"\1 <redacted>",
        safe,
    )
    safe = re.sub(r"[A-Za-z0-9_-]{32,}", "<redacted>", safe)
    return safe[:500]
