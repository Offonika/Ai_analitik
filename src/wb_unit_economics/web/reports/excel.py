from __future__ import annotations

import json
import os
import re
from collections.abc import Iterable, Mapping
from contextlib import suppress
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

MONTH_CLOSE_SHEETS = (
    "Сводка закрытия",
    "Покрытие регламента",
    "ОСВ",
    "ЕНС и налоги",
    "НДС",
    "Банк",
    "Ручные операции",
    "Подтверждения",
    "Риски и дозапросы",
    "Источники и статус",
)
TAX_LOAD_SHEETS = (
    "Обзор",
    "Расчёт УСН",
    "Налоги",
    "График платежей",
    "НДС",
    "ЕНС",
    "Источники и статус",
    "Дозапросы",
)

TAX_LOAD_FIELD_LABELS = {
    "clientName": "Клиент",
    "organizationName": "Организация 1С",
    "reportId": "ID отчёта",
    "tenantId": "ID контура",
    "clientId": "ID клиента",
    "reportKind": "Вид отчёта",
    "selectedMonth": "Выбранный месяц",
    "organizationId": "ID организации 1С",
    "periodStart": "Начало отчётного периода",
    "periodEnd": "Окончание отчётного периода",
    "ytdStart": "Начало периода с начала года",
    "ytdEnd": "Окончание периода с начала года",
    "taxSystem": "Налоговый режим",
    "profileStatus": "Статус налогового профиля",
    "revenueTaxRate": "Ставка налога с выручки, %",
    "accountantApprovalStatus": "Подтверждение бухгалтера",
    "methodologyVersion": "Версия методики",
    "generatedAt": "Дата формирования",
    "publicationStatus": "Статус публикации",
    "sourceRefreshRunId": "ID загрузки источников",
    "sourceSnapshotSetId": "ID набора снимков",
    "evidenceSha256": "SHA-256 подтверждающих данных",
    "metricKind": "Показатель",
    "numeratorKind": "Состав числителя",
    "numeratorValue": "Уплаченные собственные налоги",
    "denominatorKind": "Состав знаменателя ФНС",
    "denominatorValue": "Официальный доходный знаменатель",
    "fnsTaxBurdenRatio": "Налоговая нагрузка по методике ФНС, %",
    "calculationPeriodKind": "Период расчёта",
    "methodologyStatus": "Статус методики",
    "comparisonStatus": "Статус сравнения",
    "benchmarkYear": "Год ориентира",
    "benchmarkValue": "Значение ориентира, %",
    "usnIncomeDenominatorKind": "Состав знаменателя УСН",
    "usnIncomeValue": "Доход УСН без НДС",
    "usnIncomeTaxBurden": "Управленческая нагрузка УСН, %",
    "usnIncomeStatus": "Статус показателя УСН",
    "businessStatus": "Статус отчёта",
    "contractVersion": "Версия контракта",
    "payloadSha256": "SHA-256 отчёта",
    "taxCode": "Код налога",
    "taxName": "Налог",
    "periodKind": "Период",
    "taxBase": "Налоговая база",
    "accrued": "Начислено",
    "paid": "Уплачено",
    "balance": "Сальдо",
    "dueDate": "Срок уплаты",
    "valueStatus": "Статус значения",
    "evidenceStatus": "Статус подтверждения",
    "sourceKind": "Источник",
    "issueCode": "Код замечания",
    "paymentKind": "Вид платежа",
    "includedInFnsTaxBurden": "Включён в нагрузку ФНС",
    "exclusionReason": "Причина исключения",
    "amount": "Сумма",
    "confirmationStatus": "Статус подтверждения",
    "status": "Статус",
    "outputVat": "Начисленный НДС",
    "inputVat": "Входящий НДС",
    "payableVat": "НДС к уплате",
    "asOfDate": "Дата состояния",
    "snapshotId": "ID снимка",
    "code": "Код замечания",
    "severity": "Важность",
    "section": "Раздел",
    "message": "Что найдено",
    "nextAction": "Что сделать",
}

TAX_LOAD_VALUE_LABELS = {
    "tax_load": "Налоговая нагрузка",
    "accountant_review_required": "Нужна проверка бухгалтера",
    "preliminary": "Предварительный",
    "confirmed": "Подтверждено",
    "loaded": "Загружено",
    "ready": "Готово",
    "complete": "Завершено",
    "missing": "Нет данных",
    "partial": "Загружено частично",
    "partial_source": "Источник неполный",
    "empty_expected": "Ожидаемо нет данных",
    "source_gap": "Недостаточно данных",
    "informational": "Справочно",
    "warning": "Нужно проверить",
    "error": "Ошибка",
    "critical": "Критично",
    "draft": "Черновик",
    "published": "Опубликован",
    "not_confirmed": "Не подтверждено",
    "unconfirmed": "Не подтверждено",
    "not_applicable": "Не применяется",
    "management_reference": "Управленческий ориентир",
    "pending_methodology_confirmation": "Ожидает подтверждения методики",
    "preliminary_ytd": "Предварительно, с начала года",
    "ytd": "С начала года",
    "year_to_date": "С начала года",
    "month": "За месяц",
    "monthly": "За месяц",
    "report_month": "За месяц",
    "fns_tax_risk": "Налоговая нагрузка по методике ФНС",
    "paid_taxes_excluding_agents_and_insurance_contributions": (
        "Фактически уплаченные собственные налоги без агентских платежей "
        "и страховых взносов"
    ),
    "financial_result_income_excluding_participation_income": (
        "Доход по отчёту о финансовых результатах без доходов от участия"
    ),
    "usn_income_receipts_excluding_vat": "Доход УСН из поступлений без НДС",
    "own_tax": "Собственный налог",
    "agent_ndfl": "НДФЛ налогового агента",
    "agent_profit_tax_dividends": "Налог налогового агента с дивидендов",
    "insurance_contribution": "Страховые взносы",
    "unclassified": "Не классифицировано",
    "agent_payment": "Агентский платёж",
    "osno": "ОСНО",
    "usn_income": "УСН «Доходы»",
    "usn_income_expense": "УСН «Доходы минус расходы»",
    "усн доходы": "УСН «Доходы»",
    "усн доходы минус расходы": "УСН «Доходы минус расходы»",
}

TAX_LOAD_SOURCE_LABELS = {
    "onec_tax": "Налоговый учёт 1С",
    "onec_osv": "ОСВ 1С",
    "onec_bank": "Банк в 1С",
    "onec_accounting_bank_in": "Банковские поступления 1С",
    "onec_accounting_bank_out": "Банковские списания 1С",
    "onec_accounting_chart": "План счетов 1С",
    "onec_accounting_ens": "Единый налоговый счёт 1С",
    "onec_accounting_ens_sanctions": "Санкции по ЕНС в 1С",
    "onec_accounting_manual_operations": "Ручные операции 1С",
    "onec_accounting_month_close_docs": "Документы закрытия месяца 1С",
    "onec_accounting_purchase_corrections": "Корректировки приобретений 1С",
    "onec_accounting_register_corrections": "Корректировки регистров 1С",
    "onec_accounting_sales_corrections": "Корректировки продаж 1С",
    "onec_accounting_taxes": "Расчёты по налогам 1С",
    "onec_accounting_taxes_on_ens": "Платежи по налогам на ЕНС в 1С",
    "onec_kudir": "Книга учёта доходов и расходов 1С",
    "onec_official_financial_results": "Отчёт о финансовых результатах 1С",
    "onec_organizations": "Организации 1С",
    "onec_tax_accrual_lines": "Строки начислений налогов 1С",
    "onec_tax_accruals": "Начисления налогов 1С",
    "onec_tax_kinds": "Виды налогов 1С",
    "onec_tax_registrations": "Налоговые регистрации 1С",
    "onec_tax_register": "Налоговый регистр 1С",
    "onec_tax_special_regime_notifications": (
        "Уведомления по специальным налоговым режимам 1С"
    ),
    "onec_vat_books": "Книги НДС 1С",
    "onec_vat_purchase_book": "Книга покупок 1С",
    "onec_vat_sales_book": "Книга продаж 1С",
    "accountant_confirmation": "Подтверждение бухгалтера",
}

TAX_LOAD_DATE_FIELDS = {
    "periodStart",
    "periodEnd",
    "ytdStart",
    "ytdEnd",
    "generatedAt",
    "dueDate",
    "asOfDate",
}
TAX_LOAD_ENUM_FIELDS = {
    "reportKind",
    "publicationStatus",
    "businessStatus",
    "metricKind",
    "numeratorKind",
    "denominatorKind",
    "calculationPeriodKind",
    "methodologyStatus",
    "comparisonStatus",
    "usnIncomeDenominatorKind",
    "usnIncomeStatus",
    "periodKind",
    "valueStatus",
    "evidenceStatus",
    "paymentKind",
    "exclusionReason",
    "confirmationStatus",
    "status",
    "severity",
    "taxSystem",
    "profileStatus",
    "accountantApprovalStatus",
}

TAX_LOAD_CURRENCY_FIELDS = {
    "numeratorValue",
    "denominatorValue",
    "usnIncomeValue",
    "taxBase",
    "accrued",
    "paid",
    "balance",
    "amount",
    "outputVat",
    "inputVat",
    "payableVat",
}
TAX_LOAD_PERCENT_FIELDS = {
    "fnsTaxBurdenRatio",
    "benchmarkValue",
    "usnIncomeTaxBurden",
}
TAX_LOAD_FRACTION_PERCENT_FIELDS = {"revenueTaxRate"}
TAX_LOAD_OVERVIEW_FIELDS = (
    "clientName",
    "organizationName",
    "reportKind",
    "selectedMonth",
    "calculationPeriodKind",
    "taxSystem",
    "profileStatus",
    "revenueTaxRate",
    "periodStart",
    "periodEnd",
    "ytdStart",
    "ytdEnd",
    "numeratorValue",
    "denominatorValue",
    "fnsTaxBurdenRatio",
    "usnIncomeValue",
    "usnIncomeTaxBurden",
    "usnIncomeStatus",
    "methodologyStatus",
    "businessStatus",
    "accountantApprovalStatus",
    "generatedAt",
)
TAX_LOAD_ROW_FIELDS = {
    "Налоги": (
        "taxName",
        "periodKind",
        "taxBase",
        "accrued",
        "paid",
        "balance",
        "dueDate",
        "includedInFnsTaxBurden",
        "exclusionReason",
        "evidenceStatus",
        "sourceKind",
    ),
    "График платежей": (
        "taxName",
        "dueDate",
        "amount",
        "confirmationStatus",
    ),
    "Источники и статус": (
        "sourceKind",
        "periodStart",
        "periodEnd",
        "status",
    ),
    "Дозапросы": ("severity", "section", "message", "nextAction"),
}
TAX_LOAD_SUMMARY_FIELDS = {
    "НДС": ("status", "outputVat", "inputVat", "payableVat", "sourceKind"),
    "ЕНС": ("status", "balance", "asOfDate"),
}
TAX_LOAD_TABLE_NAMES = {
    "Обзор": "TaxLoadOverview",
    "Расчёт УСН": "TaxLoadUsnCalculation",
    "Налоги": "TaxLoadTaxes",
    "График платежей": "TaxLoadSchedule",
    "НДС": "TaxLoadVat",
    "ЕНС": "TaxLoadEns",
    "Источники и статус": "TaxLoadSources",
    "Дозапросы": "TaxLoadIssues",
}

TAX_LOAD_CURRENCY_FORMAT = "#,##0.00 [$₽-419]"
TAX_LOAD_PERCENT_FORMAT = '0.00 " %"'
TAX_LOAD_FRACTION_PERCENT_FORMAT = "0.00%"
TAX_LOAD_DATE_FORMAT = "DD.MM.YYYY"
TAX_LOAD_DATETIME_FORMAT = "DD.MM.YYYY HH:MM"
FORMULA_PREFIXES = ("=", "+", "-", "@")
RUSSIAN_MONTHS = (
    "",
    "Январь",
    "Февраль",
    "Март",
    "Апрель",
    "Май",
    "Июнь",
    "Июль",
    "Август",
    "Сентябрь",
    "Октябрь",
    "Ноябрь",
    "Декабрь",
)

HEADER_FILL = PatternFill("solid", fgColor="0F6B78")
LABEL_FILL = PatternFill("solid", fgColor="E8F3F5")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LABEL_FONT = Font(bold=True, color="17324D")
THIN_BORDER = Border(
    left=Side(style="thin", color="D5E1E5"),
    right=Side(style="thin", color="D5E1E5"),
    top=Side(style="thin", color="D5E1E5"),
    bottom=Side(style="thin", color="D5E1E5"),
)
USN_TOTAL_FILL = PatternFill("solid", fgColor="FCE4C7")
USN_INCOME_FILL = PatternFill("solid", fgColor="FFF200")
USN_SOURCE_FILL = PatternFill("solid", fgColor="E2F0D9")


def _cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _append_safe_row(sheet: Any, values: Iterable[Any]) -> None:
    sheet.append(list(values))
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str) and cell.value.lstrip().startswith(
            FORMULA_PREFIXES
        ):
            cell.data_type = "s"


def _write_mapping(sheet: Any, value: Mapping[str, Any]) -> None:
    _append_safe_row(sheet, ["Показатель", "Значение"])
    for key, item in value.items():
        _append_safe_row(sheet, [key, _cell(item)])


def _write_rows(sheet: Any, rows: Iterable[Mapping[str, Any]]) -> None:
    normalized = list(rows)
    if not normalized:
        _append_safe_row(sheet, ["Нет подтвержденных данных"])
        return
    headers: list[str] = []
    for row in normalized:
        for key in row:
            if key not in headers:
                headers.append(key)
    _append_safe_row(sheet, headers)
    for row in normalized:
        _append_safe_row(sheet, [_cell(row.get(key)) for key in headers])


def _tax_load_field_label(key: str) -> str:
    try:
        return TAX_LOAD_FIELD_LABELS[key]
    except KeyError as exc:
        raise ValueError(f"tax-load Excel label is missing for field {key!r}") from exc


def _tax_load_date(value: Any) -> date | datetime | None:
    if isinstance(value, datetime):
        parsed = value
        has_time = True
    elif isinstance(value, date):
        return value if value.year >= 1900 else None
    elif isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            return None
        has_time = "T" in value or " " in value
    else:
        return None
    if parsed.year < 1900:
        return None
    if has_time:
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone(UTC).replace(tzinfo=None)
        return parsed
    return parsed.date()


def _tax_load_month_label(value: Any) -> str:
    parsed = _tax_load_date(value)
    if parsed is None:
        return "Не указано"
    return f"{RUSSIAN_MONTHS[parsed.month]} {parsed.year}"


def _tax_load_user_text(value: Any) -> str:
    text = str(value or "")
    for source_kind, label in sorted(
        TAX_LOAD_SOURCE_LABELS.items(), key=lambda item: len(item[0]), reverse=True
    ):
        text = re.sub(rf"\b{re.escape(source_kind)}\b", label, text, flags=re.I)
    text = re.sub(r"\bread-only\b", "только для чтения", text, flags=re.I)
    return re.sub(
        r"\bRecordType\s+fallback\b",
        "резервный источник",
        text,
        flags=re.I,
    )


def _tax_load_decimal(value: Any) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    normalized = str(value).replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except (InvalidOperation, TypeError, ValueError):
        return None


def _tax_load_cell(key: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if key in TAX_LOAD_DATE_FIELDS:
        return _tax_load_date(value) or "Не указано"
    if (
        key in TAX_LOAD_CURRENCY_FIELDS
        or key in TAX_LOAD_PERCENT_FIELDS
        or key in TAX_LOAD_FRACTION_PERCENT_FIELDS
    ):
        return _tax_load_decimal(value)
    if key == "sourceKind":
        normalized = str(value).strip().casefold()
        if normalized in TAX_LOAD_SOURCE_LABELS:
            return TAX_LOAD_SOURCE_LABELS[normalized]
        return "Не определено"
    if key in TAX_LOAD_ENUM_FIELDS:
        normalized = str(value).strip().casefold()
        return TAX_LOAD_VALUE_LABELS.get(normalized, "Не определено")
    if key in {"taxName", "section", "message", "nextAction"}:
        return _tax_load_user_text(value)
    return _cell(value)


def _format_tax_load_cell(cell: Any, key: str) -> None:
    if key in TAX_LOAD_CURRENCY_FIELDS:
        cell.number_format = TAX_LOAD_CURRENCY_FORMAT
    elif key in TAX_LOAD_FRACTION_PERCENT_FIELDS:
        cell.number_format = TAX_LOAD_FRACTION_PERCENT_FORMAT
    elif key in TAX_LOAD_PERCENT_FIELDS:
        cell.number_format = TAX_LOAD_PERCENT_FORMAT
    elif key == "generatedAt":
        cell.number_format = TAX_LOAD_DATETIME_FORMAT
    elif key in TAX_LOAD_DATE_FIELDS:
        cell.number_format = TAX_LOAD_DATE_FORMAT


def _write_tax_load_mapping(
    sheet: Any,
    value: Mapping[str, Any],
    fields: Iterable[str],
) -> None:
    _append_safe_row(sheet, ["Показатель", "Значение"])
    for key in fields:
        _append_safe_row(
            sheet,
            [_tax_load_field_label(key), _tax_load_cell(key, value.get(key))],
        )
        _format_tax_load_cell(sheet.cell(row=sheet.max_row, column=2), key)


def _write_tax_load_rows(
    sheet: Any,
    rows: Iterable[Mapping[str, Any]],
    fields: Iterable[str],
) -> None:
    headers = tuple(fields)
    normalized = list(rows)
    _append_safe_row(sheet, [_tax_load_field_label(key) for key in headers])
    if not normalized:
        _append_safe_row(
            sheet,
            ["Нет подтверждённых данных", *([None] * (len(headers) - 1))],
        )
        return
    for row in normalized:
        _append_safe_row(
            sheet,
            [_tax_load_cell(key, row.get(key)) for key in headers],
        )
        for column, key in enumerate(headers, start=1):
            _format_tax_load_cell(sheet.cell(row=sheet.max_row, column=column), key)


def _usn_period_columns(ytd_end: Any) -> list[dict[str, Any]]:
    parsed = _tax_load_date(ytd_end)
    if parsed is None:
        return [{"kind": "total", "cutoff": 12, "label": "Итого с начала года"}]
    result: list[dict[str, Any]] = []
    total_labels = {
        3: "Итого за I квартал",
        6: "Итого за полугодие",
        9: "Итого за 9 месяцев",
        12: "Итого за год",
    }
    for month in range(1, parsed.month + 1):
        result.append(
            {"kind": "month", "month": month, "label": RUSSIAN_MONTHS[month]}
        )
        if month in total_labels:
            result.append(
                {"kind": "total", "cutoff": month, "label": total_labels[month]}
            )
    if result[-1]["kind"] != "total":
        result.append(
            {
                "kind": "total",
                "cutoff": parsed.month,
                "label": "Итого с начала года",
            }
        )
    result[-1]["isFinal"] = True
    return result


def _usn_month_values(rows: Any) -> dict[int, Decimal]:
    result: dict[int, Decimal] = {}
    if not isinstance(rows, list):
        return result
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        month_text = str(row.get("month") or "")
        value = _tax_load_decimal(row.get("value"))
        status = str(row.get("status") or "").strip().casefold()
        if value is None or status not in {"loaded", "ready", "complete", "confirmed"}:
            continue
        try:
            month = int(month_text[5:7])
        except (TypeError, ValueError):
            continue
        if 1 <= month <= 12:
            result[month] = value
    return result


def _usn_row_values(
    columns: list[dict[str, Any]],
    monthly: Mapping[int, Decimal],
    *,
    final_value: Any,
    require_complete_months: bool,
) -> list[Any]:
    result: list[Any] = []
    final_decimal = _tax_load_decimal(final_value)
    for column in columns:
        if column["kind"] == "month":
            result.append(monthly.get(int(column["month"])))
            continue
        cutoff = int(column["cutoff"])
        available = [
            monthly[month] for month in range(1, cutoff + 1) if month in monthly
        ]
        complete = len(available) == cutoff
        result.append(
            sum(available, Decimal("0"))
            if available and (complete or not require_complete_months)
            else None
        )
    if columns and columns[-1].get("isFinal") and final_decimal is not None:
        result[-1] = final_decimal
    return result


def _write_usn_calculation(sheet: Any, payload: Mapping[str, Any]) -> None:
    detail = dict(payload.get("usnDetail") or {})
    if detail.get("status") == "not_applicable":
        _append_safe_row(sheet, ["Показатель", "Значение"])
        _append_safe_row(
            sheet,
            ["Статус", "Расчёт применяется только для УСН «Доходы»"],
        )
        return
    columns = _usn_period_columns(payload.get("ytdEnd"))
    _append_safe_row(sheet, ["Показатель", *(item["label"] for item in columns)])
    summary = dict(payload.get("taxLoadSummary") or {})
    profile = dict(payload.get("taxProfile") or {})
    income_ytd = detail.get("incomeYtd") or summary.get("usnIncomeValue")
    income_monthly = _usn_month_values(detail.get("monthlyIncome"))
    payment_monthly = _usn_month_values(detail.get("monthlyTaxPayments"))
    income_values = _usn_row_values(
        columns,
        income_monthly,
        final_value=income_ytd,
        require_complete_months=True,
    )
    payment_values = _usn_row_values(
        columns,
        payment_monthly,
        final_value=detail.get("paidTaxYtd"),
        require_complete_months=False,
    )
    def last_only(value: Any) -> list[Any]:
        return [None] * (len(columns) - 1) + [value]

    rows = (
        ("Доход по КУДиР 1С", income_values, "currency"),
        ("Итого доход без НДС", income_values, "currency"),
        (
            "Ставка УСН",
            last_only(
                _tax_load_decimal(
                    detail.get("revenueTaxRate") or profile.get("revenueTaxRate")
                )
            ),
            "rate",
        ),
        (
            "Исчислено УСН с начала года",
            last_only(_tax_load_decimal(detail.get("calculatedTaxYtd"))),
            "currency",
        ),
        ("Уплачено УСН", payment_values, "currency"),
        (
            "К доплате / переплата УСН",
            last_only(_tax_load_decimal(detail.get("taxPayable"))),
            "currency",
        ),
        (
            "Срок уплаты",
            last_only(_tax_load_date(detail.get("dueDate")) or "Не указано"),
            "date",
        ),
        (
            "Статус данных",
            last_only(_tax_load_cell("status", detail.get("status") or "source_gap")),
            "text",
        ),
    )
    if not detail.get("monthlyIncome"):
        rows += (
            (
                "Помесячная детализация",
                last_only("Сформируйте отчёт повторно для заполнения месяцев"),
                "text",
            ),
        )
    for label, values, value_kind in rows:
        _append_safe_row(sheet, [label, *values])
        for column in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=sheet.max_row, column=column)
            if value_kind == "currency":
                cell.number_format = TAX_LOAD_CURRENCY_FORMAT
            elif value_kind == "rate":
                cell.number_format = TAX_LOAD_FRACTION_PERCENT_FORMAT
            elif value_kind == "date" and isinstance(cell.value, (date, datetime)):
                cell.number_format = TAX_LOAD_DATE_FORMAT


def _style(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[column[0].column_letter].width = max(width, 12)


def _style_tax_load(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_title_rows = "1:1"
        sheet.sheet_properties.tabColor = "0F6B78"
        sheet.row_dimensions[1].height = 28

        for cell in sheet[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = THIN_BORDER
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        if sheet.title in {"Обзор", "НДС", "ЕНС"}:
            for row in range(2, sheet.max_row + 1):
                label = sheet.cell(row=row, column=1)
                label.font = LABEL_FONT
                label.fill = LABEL_FILL

        for column_index in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            values = [
                str(sheet.cell(row=row, column=column_index).value or "")
                for row in range(1, min(sheet.max_row, 200) + 1)
            ]
            header = str(sheet.cell(row=1, column=column_index).value or "")
            maximum = 58 if header in {"Что найдено", "Что сделать"} else 38
            minimum = 22 if header in {"Налог", "Источник"} else 12
            content_width = max((len(item) for item in values), default=0) + 2
            width = min(max(content_width, minimum), maximum)
            sheet.column_dimensions[column_letter].width = width

        if sheet.title == "Расчёт УСН":
            sheet.column_dimensions["A"].width = 34
            for column_index in range(2, sheet.max_column + 1):
                header = str(sheet.cell(row=1, column=column_index).value or "")
                if header.startswith("Итого"):
                    for row_index in range(1, sheet.max_row + 1):
                        sheet.cell(row=row_index, column=column_index).fill = (
                            USN_TOTAL_FILL
                        )
                    sheet.cell(row=1, column=column_index).font = LABEL_FONT
            for row_index in range(2, sheet.max_row + 1):
                label = str(sheet.cell(row=row_index, column=1).value or "")
                sheet.cell(row=row_index, column=1).font = LABEL_FONT
                if label == "Доход по КУДиР 1С":
                    for cell in sheet[row_index]:
                        cell.fill = USN_SOURCE_FILL
                elif label in {
                    "Итого доход без НДС",
                    "К доплате / переплата УСН",
                }:
                    for cell in sheet[row_index]:
                        cell.fill = USN_INCOME_FILL
                        cell.font = LABEL_FONT

        if sheet.max_row >= 2 and sheet.max_column >= 2:
            reference = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            table = Table(
                displayName=TAX_LOAD_TABLE_NAMES[sheet.title],
                ref=reference,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)


def _save_workbook_atomic(workbook: Workbook, output_path: Path) -> None:
    temporary_path: Path | None = None
    try:
        with NamedTemporaryFile(
            dir=output_path.parent,
            prefix=f".{output_path.stem}-",
            suffix=".xlsx",
            delete=False,
        ) as handle:
            temporary_path = Path(handle.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
        temporary_path = None
    finally:
        if temporary_path is not None:
            with suppress(FileNotFoundError):
                temporary_path.unlink()


def write_scenario_excel(
    payload: Mapping[str, Any],
    payload_sha256: str,
    output_path: Path,
    *,
    export_context: Mapping[str, Any] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    report_kind = payload.get("reportKind")
    if report_kind == "month_close_control":
        for title in MONTH_CLOSE_SHEETS:
            workbook.create_sheet(title)
        summary = {
            **dict(payload.get("meta") or {}),
            "businessRecommendation": payload.get("businessRecommendation"),
            "contractVersion": payload.get("contractVersion"),
            "payloadSha256": payload_sha256,
        }
        _write_mapping(workbook["Сводка закрытия"], summary)
        _write_rows(workbook["Покрытие регламента"], payload.get("controls") or [])
        osv_sheet = workbook["ОСВ"]
        _write_mapping(osv_sheet, payload.get("osvSummary") or {})
        osv_sheet.append([])
        _write_rows(osv_sheet, payload.get("osvRows") or [])
        _write_mapping(
            workbook["ЕНС и налоги"],
            {**(payload.get("ensSummary") or {}), **(payload.get("taxSummary") or {})},
        )
        _write_mapping(workbook["НДС"], payload.get("vatSummary") or {})
        _write_mapping(workbook["Банк"], payload.get("bankSummary") or {})
        _write_mapping(
            workbook["Ручные операции"],
            payload.get("manualOperationsSummary") or {},
        )
        _write_rows(workbook["Подтверждения"], payload.get("confirmations") or [])
        _write_rows(workbook["Риски и дозапросы"], payload.get("issues") or [])
        _write_rows(workbook["Источники и статус"], payload.get("sourceCoverage") or [])
    elif report_kind == "tax_load":
        for title in TAX_LOAD_SHEETS:
            workbook.create_sheet(title)
        meta = dict(payload.get("meta") or {})
        profile = dict(payload.get("taxProfile") or {})
        summary = dict(payload.get("taxLoadSummary") or {})
        approval = payload.get("accountantApproval")
        context = dict(export_context or {})
        overview = {
            "clientName": context.get("clientName") or "Не указано",
            "organizationName": context.get("organizationName") or "Не указано",
            "reportKind": meta.get("reportKind") or payload.get("reportKind"),
            "selectedMonth": _tax_load_month_label(meta.get("periodStart")),
            "calculationPeriodKind": summary.get("calculationPeriodKind"),
            "taxSystem": profile.get("taxSystem"),
            "profileStatus": profile.get("profileStatus"),
            "revenueTaxRate": profile.get("revenueTaxRate"),
            "periodStart": meta.get("periodStart"),
            "periodEnd": meta.get("periodEnd"),
            "ytdStart": payload.get("ytdStart"),
            "ytdEnd": payload.get("ytdEnd"),
            "numeratorValue": summary.get("numeratorValue"),
            "denominatorValue": summary.get("denominatorValue"),
            "fnsTaxBurdenRatio": summary.get("fnsTaxBurdenRatio"),
            "usnIncomeValue": summary.get("usnIncomeValue"),
            "usnIncomeTaxBurden": summary.get("usnIncomeTaxBurden"),
            "usnIncomeStatus": summary.get("usnIncomeStatus"),
            "methodologyStatus": summary.get("methodologyStatus"),
            "businessStatus": payload.get("businessStatus"),
            "accountantApprovalStatus": (
                approval.get("status") or "confirmed"
                if isinstance(approval, Mapping)
                else "not_confirmed"
            ),
            "generatedAt": meta.get("generatedAt"),
            "methodologyVersion": meta.get("methodologyVersion"),
            "reportId": meta.get("reportId"),
        }
        _write_tax_load_mapping(workbook["Обзор"], overview, TAX_LOAD_OVERVIEW_FIELDS)
        _write_usn_calculation(workbook["Расчёт УСН"], payload)
        row_payloads = {
            "Налоги": payload.get("taxRows") or [],
            "График платежей": payload.get("paymentSchedule") or [],
            "Источники и статус": payload.get("sourceCoverage") or [],
            "Дозапросы": payload.get("issues") or [],
        }
        for title, fields in TAX_LOAD_ROW_FIELDS.items():
            _write_tax_load_rows(workbook[title], row_payloads[title], fields)
        summary_payloads = {
            "НДС": payload.get("vatSummary") or {},
            "ЕНС": payload.get("ensSummary") or {},
        }
        for title, fields in TAX_LOAD_SUMMARY_FIELDS.items():
            _write_tax_load_mapping(workbook[title], summary_payloads[title], fields)
    else:
        raise ValueError("unsupported scenario report kind")
    if report_kind == "tax_load":
        workbook.properties.title = "Налоговая нагрузка"
        workbook.properties.subject = "Налоговый отчёт с начала года"
        workbook.properties.creator = "Шумейко и партнёры"
        workbook.properties.identifier = str(
            (payload.get("meta") or {}).get("reportId") or ""
        )
        workbook.properties.version = str(
            (payload.get("meta") or {}).get("methodologyVersion") or ""
        )
        workbook.properties.language = "ru-RU"
        _style_tax_load(workbook)
    else:
        _style(workbook)
    try:
        _save_workbook_atomic(workbook, output_path)
    finally:
        workbook.close()
    return output_path
