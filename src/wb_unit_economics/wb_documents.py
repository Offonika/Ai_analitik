from __future__ import annotations

import base64
import hashlib
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile
from zoneinfo import ZoneInfo

import httpx
from openpyxl import load_workbook

from wb_unit_economics.wb_finance import WbFinanceSellerAccount, WbFinanceSettings

MOSCOW_TZ = ZoneInfo("Europe/Moscow")
DOCUMENT_CATEGORIES_ENDPOINT = (
    "https://documents-api.wildberries.ru/api/v1/documents/categories"
)
DOCUMENT_LIST_ENDPOINT = "https://documents-api.wildberries.ru/api/v1/documents/list"
DOCUMENT_DOWNLOAD_ENDPOINT = (
    "https://documents-api.wildberries.ru/api/v1/documents/download"
)
DEFAULT_DOCUMENT_CATEGORY_KEYWORDS = (
    "отчет",
    "отчёт",
    "комиссион",
    "выкуп",
    "уведомление",
    "упд",
    "акт",
    "услуг",
    "реализац",
)


@dataclass(frozen=True)
class WbDocumentExportResult:
    seller_account_id: str
    account_name: str
    ok: bool
    status: str
    row_count: int
    downloaded_count: int = 0
    output_file: str = ""
    status_code: int | None = None
    error: str = ""


class WbDocumentsClient:
    """Read-only client for WB primary accounting documents."""

    def __init__(
        self,
        account: WbFinanceSellerAccount,
        *,
        timeout_seconds: float = 45.0,
        transport: httpx.BaseTransport | None = None,
    ) -> None:
        self._client = httpx.Client(
            headers={"Authorization": account.api_key},
            timeout=timeout_seconds,
            follow_redirects=True,
            transport=transport,
        )

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> WbDocumentsClient:
        return self

    def __exit__(self, *_exc: object) -> None:
        self.close()

    def fetch_categories(
        self,
        *,
        locale: str = "ru",
    ) -> tuple[list[dict[str, Any]], int]:
        response = self._client.get(
            DOCUMENT_CATEGORIES_ENDPOINT,
            params={"locale": locale},
        )
        response.raise_for_status()
        payload = response.json()
        data = payload.get("data", payload) if isinstance(payload, dict) else payload
        categories = data.get("categories", data) if isinstance(data, dict) else data
        if not isinstance(categories, list):
            raise ValueError("Unexpected WB documents categories payload")
        return (
            [item for item in categories if isinstance(item, dict)],
            response.status_code,
        )

    def fetch_documents_page(
        self,
        *,
        begin_time: date,
        end_time: date,
        locale: str = "ru",
        limit: int = 50,
        offset: int = 0,
        category: str | None = None,
    ) -> tuple[list[dict[str, Any]], dict[str, Any], int]:
        params: dict[str, object] = {
            "locale": locale,
            "beginTime": begin_time.isoformat(),
            "endTime": end_time.isoformat(),
            "sort": "date",
            "order": "asc",
            "limit": limit,
            "offset": offset,
        }
        if category:
            params["category"] = category
        response = self._client.get(DOCUMENT_LIST_ENDPOINT, params=params)
        response.raise_for_status()
        payload = response.json()
        data = payload.get("data", payload) if isinstance(payload, dict) else payload
        documents = data.get("documents", data) if isinstance(data, dict) else data
        if not isinstance(documents, list):
            raise ValueError("Unexpected WB documents list payload")
        return (
            [item for item in documents if isinstance(item, dict)],
            payload if isinstance(payload, dict) else {"data": payload},
            response.status_code,
        )

    def download_document(
        self,
        *,
        service_name: str,
        extension: str,
    ) -> tuple[bytes, dict[str, Any], int]:
        response = self._client.get(
            DOCUMENT_DOWNLOAD_ENDPOINT,
            params={"serviceName": service_name, "extension": extension},
        )
        response.raise_for_status()
        payload = response.json()
        data = payload.get("data", payload) if isinstance(payload, dict) else payload
        if not isinstance(data, dict) or not data.get("document"):
            raise ValueError("Unexpected WB document download payload")
        document = base64.b64decode(str(data["document"]))
        metadata = {
            "fileName": data.get("fileName"),
            "extension": data.get("extension", extension),
        }
        return document, metadata, response.status_code


def export_wb_documents(
    *,
    settings: WbFinanceSettings,
    output_dir: Path,
    period_start: date,
    period_end: date,
    category_keywords: tuple[str, ...] = DEFAULT_DOCUMENT_CATEGORY_KEYWORDS,
    download: bool = False,
    locale: str = "ru",
    transport: httpx.BaseTransport | None = None,
) -> list[WbDocumentExportResult]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[WbDocumentExportResult] = []
    for account in settings.accounts:
        account_dir = output_dir / _safe_filename(account.seller_account_id)
        account_dir.mkdir(parents=True, exist_ok=True)
        try:
            with WbDocumentsClient(
                account,
                timeout_seconds=settings.timeout_seconds,
                transport=transport,
            ) as client:
                rows, raw_payloads = _fetch_all_document_rows(
                    client,
                    period_start=period_start,
                    period_end=period_end,
                    locale=locale,
                )
                selected = [
                    row
                    for row in rows
                    if _document_matches_keywords(row, category_keywords)
                ]
                downloads = (
                    _download_selected_documents(client, selected, account_dir)
                    if download
                    else []
                )
        except httpx.HTTPStatusError as exc:
            results.append(
                WbDocumentExportResult(
                    seller_account_id=account.seller_account_id,
                    account_name=account.account_name,
                    ok=False,
                    status="http_error",
                    row_count=0,
                    status_code=exc.response.status_code,
                    error=f"WB Documents HTTP {exc.response.status_code}",
                )
            )
            continue
        except Exception as exc:
            results.append(
                WbDocumentExportResult(
                    seller_account_id=account.seller_account_id,
                    account_name=account.account_name,
                    ok=False,
                    status="error",
                    row_count=0,
                    error=str(exc),
                )
            )
            continue
        raw_file = account_dir / "documents_pages.raw.json"
        raw_file.write_text(
            json.dumps(raw_payloads, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        manifest_rows = [
            _document_manifest_row(row, downloads_by_service=downloads)
            for row in selected
        ]
        metadata_file = account_dir / "documents_manifest.json"
        metadata_file.write_text(
            json.dumps(manifest_rows, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        results.append(
            WbDocumentExportResult(
                seller_account_id=account.seller_account_id,
                account_name=account.account_name,
                ok=True,
                status="ok",
                row_count=len(selected),
                downloaded_count=sum(
                    1 for item in manifest_rows if item.get("download")
                ),
                output_file=str(metadata_file.relative_to(output_dir)),
                status_code=200,
            )
        )
    _write_export_manifest(
        output_dir=output_dir,
        period_start=period_start,
        period_end=period_end,
        category_keywords=category_keywords,
        download=download,
        results=results,
    )
    return results


def load_wb_document_export_results(
    output_dir: Path,
) -> list[WbDocumentExportResult]:
    """Restore safe provider results from an immutable documents snapshot."""
    manifest_path = output_dir / "manifest.json"
    payload = json.loads(manifest_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Unexpected WB documents manifest payload")
    rows = payload.get("provider_results", payload.get("results", []))
    if not isinstance(rows, list):
        raise ValueError("Unexpected WB documents manifest results")
    results: list[WbDocumentExportResult] = []
    for row in rows:
        if not isinstance(row, dict) or not row.get("seller_account_id"):
            continue
        results.append(
            WbDocumentExportResult(
                seller_account_id=str(row["seller_account_id"]),
                account_name=str(row.get("account_name") or ""),
                ok=bool(row.get("ok")),
                status=str(row.get("status") or "error"),
                row_count=int(row.get("row_count") or 0),
                downloaded_count=int(row.get("downloaded_count") or 0),
                output_file=str(row.get("output_file") or ""),
                status_code=(
                    int(row["status_code"])
                    if row.get("status_code") is not None
                    else None
                ),
                error=str(row.get("error") or ""),
            )
        )
    return results


def default_wb_documents_output_dir(now: datetime | None = None) -> Path:
    timestamp = (now or datetime.now(tz=MOSCOW_TZ)).strftime("%Y%m%dT%H%M%S")
    return Path("data/wb_documents") / timestamp


def _fetch_all_document_rows(
    client: WbDocumentsClient,
    *,
    period_start: date,
    period_end: date,
    locale: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    raw_payloads: list[dict[str, Any]] = []
    offset = 0
    limit = 50
    while True:
        page_rows, payload, _status_code = client.fetch_documents_page(
            begin_time=period_start,
            end_time=period_end,
            locale=locale,
            limit=limit,
            offset=offset,
        )
        raw_payloads.append(payload)
        rows.extend(page_rows)
        if len(page_rows) < limit:
            break
        offset += limit
        time.sleep(0.2)
    return rows, raw_payloads


def _download_selected_documents(
    client: WbDocumentsClient,
    rows: list[dict[str, Any]],
    account_dir: Path,
) -> dict[str, dict[str, Any]]:
    document_dir = account_dir / "documents"
    document_dir.mkdir(parents=True, exist_ok=True)
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        service_name = str(row.get("serviceName") or "")
        extensions = (
            row.get("extensions") if isinstance(row.get("extensions"), list) else []
        )
        if not service_name or not extensions:
            continue
        extension = str(extensions[0])
        content, metadata, _status_code = _download_document_with_retry(
            client,
            service_name=service_name,
            extension=extension,
        )
        filename = _safe_filename(
            str(metadata.get("fileName") or f"{service_name}.{extension}")
        )
        if not filename.endswith(f".{extension}"):
            filename = f"{filename}.{extension}"
        path = document_dir / filename
        path.write_bytes(content)
        result[service_name] = {
            "path": str(path.relative_to(account_dir)),
            "sha256": hashlib.sha256(content).hexdigest(),
            "size_bytes": len(content),
            "extension": extension,
            "summary": _redeem_notification_summary(
                content,
                extension=extension,
                service_name=service_name,
            ),
        }
        time.sleep(0.2)
    return result


def _download_document_with_retry(
    client: WbDocumentsClient,
    *,
    service_name: str,
    extension: str,
) -> tuple[bytes, dict[str, Any], int]:
    for attempt in range(4):
        try:
            return client.download_document(
                service_name=service_name,
                extension=extension,
            )
        except httpx.HTTPStatusError as exc:
            if exc.response.status_code != 429 or attempt == 3:
                raise
            retry_after = exc.response.headers.get("Retry-After", "")
            try:
                delay = max(1.0, float(retry_after))
            except ValueError:
                delay = 10.0
            time.sleep(delay)
    raise RuntimeError("WB document download retry exhausted")


def _redeem_notification_summary(
    content: bytes,
    *,
    extension: str,
    service_name: str,
) -> dict[str, Any]:
    report_id_match = re.search(r"redeem-notification-(\d+)", service_name)
    base: dict[str, Any] = {
        "reportId": report_id_match.group(1) if report_id_match else "",
        "status": "unsupported",
        "quantity": None,
        "purchaseAmount": None,
        "vatAmount": None,
    }
    if extension.casefold() != "zip":
        return base
    try:
        with ZipFile(BytesIO(content)) as archive:
            workbook_name = next(
                (
                    name
                    for name in archive.namelist()
                    if name.casefold().endswith(".xlsx")
                ),
                "",
            )
            if not workbook_name:
                return {**base, "status": "xlsx_missing"}
            workbook = load_workbook(
                BytesIO(archive.read(workbook_name)),
                read_only=True,
                data_only=True,
            )
            worksheet = workbook.active
            for values in worksheet.iter_rows(values_only=True):
                populated = [value for value in values if value not in (None, "")]
                if not populated or not str(populated[0]).strip().casefold().startswith(
                    "итого"
                ):
                    continue
                if len(populated) < 3:
                    return {**base, "status": "total_row_incomplete"}
                return {
                    **base,
                    "status": "parsed",
                    "quantity": _decimal_text(populated[1]),
                    "purchaseAmount": _decimal_text(populated[2]),
                    "vatAmount": (
                        _decimal_text(populated[4])
                        if len(populated) > 4
                        and str(populated[4]).strip() not in {"-", "–", "—"}
                        else None
                    ),
                    "workbookName": Path(workbook_name).name,
                }
            return {**base, "status": "total_row_missing"}
    except (BadZipFile, InvalidOperation, OSError, ValueError):
        return {**base, "status": "parse_error"}


def _decimal_text(value: Any) -> str:
    normalized = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".")
    return format(Decimal(normalized), "f")


def _document_matches_keywords(
    row: dict[str, Any],
    keywords: tuple[str, ...],
) -> bool:
    haystack = " ".join(
        str(row.get(key) or "") for key in ("name", "category", "serviceName")
    ).lower()
    return any(keyword.lower() in haystack for keyword in keywords)


def _document_manifest_row(
    row: dict[str, Any],
    *,
    downloads_by_service: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    service_name = str(row.get("serviceName") or "")
    return {
        "serviceName": service_name,
        "name": row.get("name"),
        "category": row.get("category"),
        "creationTime": row.get("creationTime"),
        "extensions": row.get("extensions"),
        "viewed": row.get("viewed"),
        "download": downloads_by_service.get(service_name),
    }


def _write_export_manifest(
    *,
    output_dir: Path,
    period_start: date,
    period_end: date,
    category_keywords: tuple[str, ...],
    download: bool,
    results: list[WbDocumentExportResult],
) -> None:
    manifest = {
        "source": "WB Documents API",
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "category_keywords": list(category_keywords),
        "download": download,
        "generated_at": datetime.now(tz=MOSCOW_TZ).isoformat(),
        "results": [asdict(result) for result in results],
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-zА-Яа-я0-9._-]+", "_", value).strip("._") or "document"
