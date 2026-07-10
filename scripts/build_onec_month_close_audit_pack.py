# ruff: noqa: E501
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from collections.abc import Mapping
from contextlib import suppress
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.month_close_cabinet_settings import (  # noqa: E402
    CabinetOnecLookup,
    CabinetOnecSettingsError,
    load_onec_settings_from_cabinet,
)
from wb_unit_economics.onec_odata import (  # noqa: E402
    BASE_URL_KEYS,
    PASSWORD_KEYS,
    TIMEOUT_KEYS,
    USERNAME_KEYS,
    VERIFY_SSL_KEYS,
    OnecODataClient,
    OnecODataConfigError,
    OnecODataSettings,
    extract_odata_rows,
    raw_payload_hash,
)

MOSCOW_TZ = ZoneInfo("Europe/Moscow")
DEFAULT_PERIOD_START = date(2026, 5, 1)
DEFAULT_PERIOD_END = date(2026, 6, 1)


@dataclass(frozen=True)
class CollectionSpec:
    sample_id: str
    collection_name: str
    purpose: str
    date_field: str | None = None
    page_size: int = 1000
    max_pages: int = 100
    stop_after_period: bool = True


@dataclass
class CollectionResult:
    sample_id: str
    collection_name: str
    purpose: str
    ok: bool
    period_rows: list[dict[str, Any]] = field(default_factory=list)
    scanned_rows: int = 0
    page_count: int = 0
    output_file: str | None = None
    raw_payload_hash: str = ""
    status_code: int | None = None
    error: str = ""
    min_date: str | None = None
    max_date: str | None = None
    has_period_rows: bool = False


@dataclass
class VirtualProbe:
    name: str
    ok: bool
    status_code: int | None = None
    row_count: int = 0
    fields: list[str] = field(default_factory=list)
    error: str = ""


@dataclass(frozen=True)
class OsvSourceInfo:
    source_id: str
    summary: str
    balance_sheet_name: str
    sample_sheet_name: str


BALANCE_AND_TURNOVERS_SOURCE = OsvSourceInfo(
    source_id="balance_and_turnovers",
    summary="Построена по AccountingRegister_Управленческий/BalanceAndTurnovers.",
    balance_sheet_name="ОСВ виртуальная",
    sample_sheet_name="ОСВ источник sample",
)
RECORDTYPE_SOURCE = OsvSourceInfo(
    source_id="recordtype",
    summary="Построена реконструкция по AccountingRegister_Управленческий_RecordType.",
    balance_sheet_name="ОСВ реконструкция",
    sample_sheet_name="Проводки май sample",
)


def main() -> int:
    args = _parse_args()
    period_start = args.period_start
    period_end = args.period_end
    output_dir = args.output_dir or _default_data_dir()
    report_path = args.report_path or _default_report_path(period_start)

    try:
        settings = _load_settings_from_args(args)
    except (CabinetOnecSettingsError, OnecODataConfigError) as exc:
        print(str(exc), file=sys.stderr)
        return 2

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    period_start_dt = datetime.combine(period_start, time.min)
    period_end_dt = datetime.combine(period_end, time.min)
    generated_at = datetime.now(tz=MOSCOW_TZ).isoformat()

    with OnecODataClient(settings) as client:
        service_names = _fetch_service_names(settings)
        chart_result = _fetch_collection(
            client,
            CollectionSpec(
                "chart_of_accounts",
                "ChartOfAccounts_Управленческий",
                "План счетов для расшифровки ОСВ.",
                page_size=500,
                max_pages=10,
                stop_after_period=False,
            ),
            output_dir,
            period_start_dt,
            period_end_dt,
        )
        org_result = _fetch_collection(
            client,
            CollectionSpec(
                "organizations",
                "Catalog_Организации",
                "Организации для расшифровки ключей 1С.",
                page_size=500,
                max_pages=10,
                stop_after_period=False,
            ),
            output_dir,
            period_start_dt,
            period_end_dt,
        )
        tax_type_result = _fetch_collection(
            client,
            CollectionSpec(
                "tax_types",
                "Catalog_ВидыНалогов",
                "Справочник видов налогов.",
                page_size=500,
                max_pages=10,
                stop_after_period=False,
            ),
            output_dir,
            period_start_dt,
            period_end_dt,
        )

        account_lookup = _lookup_by_key(chart_result.period_rows)
        org_lookup = _lookup_by_key(org_result.period_rows)
        tax_lookup = _lookup_by_key(tax_type_result.period_rows)
        virtual_probes = _probe_virtual_accounting_functions(
            settings=settings,
            period_start=period_start_dt,
            period_end=period_end_dt,
        )
        (
            osv_source,
            osv_collection_results,
            account_balances,
            accounting_period_records,
        ) = _fetch_osv_source(
            client=client,
            settings=settings,
            output_dir=output_dir,
            period_start=period_start_dt,
            period_end=period_end_dt,
            account_lookup=account_lookup,
            virtual_probes=virtual_probes,
            virtual_page_size=args.osv_page_size,
            virtual_max_pages=args.osv_max_pages,
            record_page_size=args.accounting_page_size,
            record_max_pages=args.accounting_max_pages,
            record_tail_max_pages=args.accounting_tail_max_pages,
            record_start_skip=args.accounting_start_skip,
        )

        collection_results = [
            chart_result,
            org_result,
            tax_type_result,
            *osv_collection_results,
        ]
        for spec in _period_collection_specs():
            collection_results.append(
                _fetch_collection(
                    client, spec, output_dir, period_start_dt, period_end_dt
                )
            )

    rows_by_id = {item.sample_id: item.period_rows for item in collection_results}
    coverage_rows = _build_coverage_rows(
        collection_results, virtual_probes, service_names
    )
    month_close_rows = _decorate_rows(rows_by_id["month_close_docs"], org_lookup)
    tax_summary_rows = _summarize_tax_rows(rows_by_id["taxes"], tax_lookup, org_lookup)
    ens_rows = _decorate_rows(rows_by_id["ens"], org_lookup)
    vat_rows = _summarize_vat(rows_by_id)
    manual_rows = _summarize_manual_operations(rows_by_id)
    bank_rows = _summarize_bank_rows(rows_by_id, org_lookup)
    risks = _build_risks(
        coverage_rows=coverage_rows,
        account_balances=account_balances,
        rows_by_id=rows_by_id,
        service_names=service_names,
        virtual_probes=virtual_probes,
        osv_source=osv_source,
    )
    requests = _build_accountant_requests(risks)

    _write_workbook(
        path=report_path,
        generated_at=generated_at,
        period_start=period_start,
        period_end=period_end,
        coverage_rows=coverage_rows,
        account_balances=account_balances,
        accounting_period_records=accounting_period_records,
        month_close_rows=month_close_rows,
        tax_summary_rows=tax_summary_rows,
        ens_rows=ens_rows,
        vat_rows=vat_rows,
        manual_rows=manual_rows,
        bank_rows=bank_rows,
        rows_by_id=rows_by_id,
        virtual_probes=virtual_probes,
        osv_source=osv_source,
        risks=risks,
        requests=requests,
    )

    manifest = {
        "generated_at": generated_at,
        "period_start": period_start.isoformat(),
        "period_end_exclusive": period_end.isoformat(),
        "read_boundary": "GET only",
        "report_path": str(report_path),
        "osv_source": osv_source.__dict__,
        "collections": [_collection_manifest(item) for item in collection_results],
        "virtual_probes": [probe.__dict__ for probe in virtual_probes],
        "service_coverage": {
            "AccountingRegister_Управленческий": (
                "AccountingRegister_Управленческий" in service_names
            ),
            "AccountingRegister_Хозрасчет": "AccountingRegister_Хозрасчет"
            in service_names,
            "ChartOfAccounts_Хозрасчет": "ChartOfAccounts_Хозрасчет" in service_names,
        },
        "risk_count": len(risks),
    }
    _write_json(output_dir / "manifest.json", manifest)
    _safe_print(f"1C month-close audit data: {output_dir}")
    _safe_print(f"1C month-close audit workbook: {report_path}")
    _safe_print(f"Coverage rows: {len(coverage_rows)}")
    _safe_print(f"OSV source: {osv_source.source_id}")
    _safe_print(f"OSV rows: {len(account_balances)}")
    _safe_print(f"Risks: {len(risks)}")
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a read-only 1C OData month-close audit pack."
    )
    parser.add_argument(
        "--env-file",
        type=Path,
        default=None,
        help="Optional env file. By default only exported ONEC_ODATA_* variables are used.",
    )
    parser.add_argument(
        "--period-start", type=_parse_date, default=DEFAULT_PERIOD_START
    )
    parser.add_argument("--period-end", type=_parse_date, default=DEFAULT_PERIOD_END)
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--report-path", type=Path, default=None)
    parser.add_argument(
        "--cabinet-client-name",
        default="",
        help="Load encrypted 1C OData settings from web cabinet by client name.",
    )
    parser.add_argument(
        "--tenant-id",
        default="",
        help="Load encrypted 1C OData settings from web cabinet by tenant id.",
    )
    parser.add_argument(
        "--integration-provider",
        default="onec_readonly",
        help="Tenant integration provider id for 1C settings.",
    )
    parser.add_argument(
        "--web-database-url",
        default="",
        help="Optional web cabinet database URL. Defaults to SHUMEYKO_DATABASE_URL or local web settings.",
    )
    parser.add_argument("--accounting-page-size", type=int, default=2000)
    parser.add_argument("--accounting-max-pages", type=int, default=80)
    parser.add_argument("--accounting-tail-max-pages", type=int, default=40)
    parser.add_argument("--osv-page-size", type=int, default=1000)
    parser.add_argument("--osv-max-pages", type=int, default=50)
    parser.add_argument(
        "--accounting-start-skip",
        type=int,
        default=0,
        help="Start reading accounting register from this OData skip offset.",
    )
    return parser.parse_args()


def _safe_print(message: str) -> None:
    try:
        print(message)
    except BrokenPipeError:
        with suppress(OSError):
            sys.stdout.close()


def _load_settings_from_args(args: argparse.Namespace) -> OnecODataSettings:
    if args.cabinet_client_name or args.tenant_id:
        return load_onec_settings_from_cabinet(
            CabinetOnecLookup(
                client_name=args.cabinet_client_name,
                tenant_id=args.tenant_id,
                provider=args.integration_provider,
                database_url=args.web_database_url,
            )
        )
    return _load_settings(args.env_file)


def _load_settings(env_file: Path | None) -> OnecODataSettings:
    if env_file is not None:
        return OnecODataSettings.from_env_file(env_file)
    values = os.environ
    base_url = _first_present(values, BASE_URL_KEYS)
    username = _first_present(values, USERNAME_KEYS)
    password = _first_present(values, PASSWORD_KEYS)
    missing: list[str] = []
    if not base_url:
        missing.append(BASE_URL_KEYS[0])
    if not username:
        missing.append(USERNAME_KEYS[0])
    if not password:
        missing.append(PASSWORD_KEYS[0])
    if missing:
        names = ", ".join(missing)
        raise OnecODataConfigError(
            f"Missing required 1C OData environment variables: {names}"
        )
    timeout_value = _first_present(values, TIMEOUT_KEYS)
    verify_value = _first_present(values, VERIFY_SSL_KEYS)
    return OnecODataSettings(
        base_url=base_url.rstrip("/"),
        username=username,
        password=password,
        timeout_seconds=float(timeout_value) if timeout_value else 30.0,
        verify_ssl=_parse_bool(verify_value, default=True),
    )


def _first_present(values: Mapping[str, str], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = values.get(key, "").strip()
        if value:
            return value
    return ""


def _parse_bool(value: str, *, default: bool) -> bool:
    if not value:
        return default
    return value.strip().lower() not in {"0", "false", "no", "off"}


def _period_collection_specs() -> list[CollectionSpec]:
    return [
        CollectionSpec(
            "month_close_docs",
            "Document_ЗакрытиеМесяца",
            "Документы закрытия месяца.",
            "Date",
            500,
            20,
        ),
        CollectionSpec(
            "taxes",
            "AccumulationRegister_РасчетыПоНалогам_RecordType",
            "Движения расчетов по налогам.",
            "Period",
            1000,
            20,
        ),
        CollectionSpec(
            "ens",
            "AccumulationRegister_РасчетыПоЕдиномуНалоговомуСчету_RecordType",
            "Движения по единому налоговому счету.",
            "Period",
            1000,
            10,
        ),
        CollectionSpec(
            "taxes_on_ens",
            "AccumulationRegister_РасчетыПоНалогамНаЕдиномНалоговомСчете_RecordType",
            "Детализация налогов на ЕНС.",
            "Period",
            1000,
            20,
        ),
        CollectionSpec(
            "ens_sanctions",
            "AccumulationRegister_РасчетыПоСанкциямНаЕдиномНалоговомСчете_RecordType",
            "Санкции на ЕНС.",
            "Period",
            1000,
            20,
        ),
        CollectionSpec(
            "vat_purchase_book",
            "AccumulationRegister_НДСЗаписиКнигиПокупок_RecordType",
            "Книга покупок НДС.",
            "Period",
            1000,
            20,
        ),
        CollectionSpec(
            "vat_sales_book",
            "AccumulationRegister_НДСЗаписиКнигиПродаж_RecordType",
            "Книга продаж НДС.",
            "Period",
            1000,
            20,
        ),
        CollectionSpec(
            "vat_presented",
            "AccumulationRegister_НДСПредъявленный_RecordType",
            "Предъявленный НДС.",
            "Period",
            1000,
            20,
        ),
        CollectionSpec(
            "manual_operation_docs",
            "Document_Операция",
            "Ручные операции.",
            "Date",
            1000,
            30,
        ),
        CollectionSpec(
            "manual_operation_postings",
            "Document_Операция_Проводки",
            "Проводки ручных операций.",
            None,
            200,
            1,
            False,
        ),
        CollectionSpec(
            "register_corrections",
            "Document_КорректировкаРегистров",
            "Корректировки регистров.",
            "Date",
            1000,
            30,
        ),
        CollectionSpec(
            "purchase_corrections",
            "Document_КорректировкаПоступления",
            "Корректировки поступления.",
            "Date",
            1000,
            30,
        ),
        CollectionSpec(
            "sales_corrections",
            "Document_КорректировкаРеализации",
            "Корректировки реализации.",
            "Date",
            1000,
            30,
        ),
        CollectionSpec(
            "bank_in",
            "Document_ПоступлениеНаСчет",
            "Поступления на расчетный счет.",
            "Date",
            1000,
            50,
        ),
        CollectionSpec(
            "bank_out",
            "Document_РасходСоСчета",
            "Расходы с расчетного счета.",
            "Date",
            1000,
            50,
        ),
        CollectionSpec(
            "tax_accrual_docs",
            "Document_НачислениеНалогов",
            "Документы начисления налогов.",
            "Date",
            1000,
            20,
        ),
        CollectionSpec(
            "tax_accrual_lines",
            "Document_НачислениеНалогов_Налоги",
            "Табличная часть начисления налогов.",
            None,
            200,
            1,
            False,
        ),
        CollectionSpec(
            "ens_operation_docs",
            "Document_ОперацияПоЕдиномуНалоговомуСчету",
            "Документы операций по ЕНС.",
            "Date",
            1000,
            20,
        ),
        CollectionSpec(
            "ens_operation_taxes",
            "Document_ОперацияПоЕдиномуНалоговомуСчету_Налоги",
            "Налоги в операциях по ЕНС.",
            None,
            200,
            1,
            False,
        ),
    ]


def _fetch_collection(
    client: OnecODataClient,
    spec: CollectionSpec,
    output_dir: Path,
    period_start: datetime,
    period_end: datetime,
) -> CollectionResult:
    period_rows: list[dict[str, Any]] = []
    scanned_rows = 0
    page_count = 0
    min_date: datetime | None = None
    max_date: datetime | None = None
    found_period = False
    last_status: int | None = None
    try:
        for page_index in range(spec.max_pages):
            payload, status = client.fetch_collection(
                spec.collection_name,
                top=spec.page_size,
                skip=page_index * spec.page_size,
            )
            last_status = status
            rows = [row for row in extract_odata_rows(payload) if isinstance(row, dict)]
            page_count += 1
            scanned_rows += len(rows)
            page_dates = []
            if spec.date_field:
                page_dates = [
                    item
                    for item in (
                        _parse_datetime(row.get(spec.date_field)) for row in rows
                    )
                    if item is not None
                ]
                if page_dates:
                    page_min = min(page_dates)
                    page_max = max(page_dates)
                    min_date = page_min if min_date is None else min(min_date, page_min)
                    max_date = page_max if max_date is None else max(max_date, page_max)
            for row in rows:
                if spec.date_field:
                    row_date = _parse_datetime(row.get(spec.date_field))
                    if row_date and period_start <= row_date < period_end:
                        period_rows.append(row)
                        found_period = True
                else:
                    period_rows.append(row)
            if len(rows) < spec.page_size:
                break
            if (
                spec.date_field
                and spec.stop_after_period
                and found_period
                and page_dates
                and min(page_dates) >= period_end
            ):
                break
    except httpx.HTTPStatusError as exc:
        return CollectionResult(
            sample_id=spec.sample_id,
            collection_name=spec.collection_name,
            purpose=spec.purpose,
            ok=False,
            scanned_rows=scanned_rows,
            page_count=page_count,
            status_code=exc.response.status_code,
            error=f"HTTP {exc.response.status_code}: {_odata_error_message(exc.response.text)}",
        )
    except (httpx.HTTPError, ValueError) as exc:
        return CollectionResult(
            sample_id=spec.sample_id,
            collection_name=spec.collection_name,
            purpose=spec.purpose,
            ok=False,
            scanned_rows=scanned_rows,
            page_count=page_count,
            status_code=last_status,
            error=exc.__class__.__name__,
        )

    payload_to_write = {
        "value": period_rows,
        "_source": {
            "collection_name": spec.collection_name,
            "purpose": spec.purpose,
            "date_field": spec.date_field,
            "period_start": period_start.isoformat(),
            "period_end_exclusive": period_end.isoformat(),
            "scanned_rows": scanned_rows,
            "page_count": page_count,
            "page_size": spec.page_size,
            "max_pages": spec.max_pages,
        },
    }
    output_path = output_dir / f"{spec.sample_id}.raw.json"
    _write_json(output_path, payload_to_write)
    return CollectionResult(
        sample_id=spec.sample_id,
        collection_name=spec.collection_name,
        purpose=spec.purpose,
        ok=True,
        period_rows=period_rows,
        scanned_rows=scanned_rows,
        page_count=page_count,
        output_file=output_path.name,
        raw_payload_hash=raw_payload_hash(payload_to_write),
        status_code=last_status,
        min_date=min_date.isoformat() if min_date else None,
        max_date=max_date.isoformat() if max_date else None,
        has_period_rows=bool(period_rows),
    )


def _fetch_osv_source(
    *,
    client: OnecODataClient,
    settings: OnecODataSettings,
    output_dir: Path,
    period_start: datetime,
    period_end: datetime,
    account_lookup: dict[str, dict[str, Any]],
    virtual_probes: list[VirtualProbe],
    virtual_page_size: int,
    virtual_max_pages: int,
    record_page_size: int,
    record_max_pages: int,
    record_tail_max_pages: int,
    record_start_skip: int,
) -> tuple[OsvSourceInfo, list[CollectionResult], list[dict[str, Any]], list[dict[str, Any]]]:
    collection_results: list[CollectionResult] = []
    if _choose_osv_source(virtual_probes) == BALANCE_AND_TURNOVERS_SOURCE.source_id:
        virtual_result, virtual_balances, virtual_rows = _fetch_balance_and_turnovers(
            settings=settings,
            output_dir=output_dir,
            period_start=period_start,
            period_end=period_end,
            account_lookup=account_lookup,
            page_size=virtual_page_size,
            max_pages=virtual_max_pages,
        )
        collection_results.append(virtual_result)
        if virtual_result.ok and (virtual_balances or not virtual_rows):
            return (
                BALANCE_AND_TURNOVERS_SOURCE,
                collection_results,
                virtual_balances,
                virtual_rows,
            )

    record_result, record_balances, record_rows = _fetch_accounting_records(
        client=client,
        output_dir=output_dir,
        period_start=period_start,
        period_end=period_end,
        account_lookup=account_lookup,
        page_size=record_page_size,
        max_pages=record_max_pages,
        tail_max_pages=record_tail_max_pages,
        start_skip=record_start_skip,
    )
    collection_results.append(record_result)
    return RECORDTYPE_SOURCE, collection_results, record_balances, record_rows


def _choose_osv_source(virtual_probes: list[VirtualProbe]) -> str:
    if any(
        probe.name == "BalanceAndTurnovers" and probe.ok for probe in virtual_probes
    ):
        return BALANCE_AND_TURNOVERS_SOURCE.source_id
    return RECORDTYPE_SOURCE.source_id


def _fetch_balance_and_turnovers(
    *,
    settings: OnecODataSettings,
    output_dir: Path,
    period_start: datetime,
    period_end: datetime,
    account_lookup: dict[str, dict[str, Any]],
    page_size: int,
    max_pages: int,
) -> tuple[CollectionResult, list[dict[str, Any]], list[dict[str, Any]]]:
    collection_name = "AccountingRegister_Управленческий/BalanceAndTurnovers"
    rows: list[dict[str, Any]] = []
    page_count = 0
    last_status: int | None = None
    base = settings.base_url.rstrip("/")
    register = quote("AccountingRegister_Управленческий", safe="")
    function_call = (
        "BalanceAndTurnovers("
        f"StartPeriod=datetime'{period_start:%Y-%m-%dT%H:%M:%S}',"
        f"EndPeriod=datetime'{period_end:%Y-%m-%dT%H:%M:%S}',"
        "AccountCondition='',Condition='',Dimensions='Организация')"
    )
    try:
        with httpx.Client(
            auth=(settings.username, settings.password),
            headers={"Accept": "application/json"},
            timeout=settings.timeout_seconds,
            verify=settings.verify_ssl,
            follow_redirects=True,
        ) as http_client:
            for page_index in range(max_pages):
                response = http_client.get(
                    f"{base}/{register}/{function_call}",
                    params={
                        "$format": "json",
                        "$top": str(page_size),
                        "$skip": str(page_index * page_size),
                    },
                )
                last_status = response.status_code
                response.raise_for_status()
                page_rows = [
                    row
                    for row in extract_odata_rows(response.json())
                    if isinstance(row, dict)
                ]
                rows.extend(page_rows)
                page_count += 1
                if len(page_rows) < page_size:
                    break
    except httpx.HTTPStatusError as exc:
        result = CollectionResult(
            sample_id="accounting_balance_and_turnovers",
            collection_name=collection_name,
            purpose="Виртуальная таблица остатков и оборотов управленческого регистра.",
            ok=False,
            scanned_rows=len(rows),
            page_count=page_count,
            status_code=exc.response.status_code,
            error=f"HTTP {exc.response.status_code}: {_odata_error_message(exc.response.text)}",
        )
        return result, [], []
    except (httpx.HTTPError, ValueError) as exc:
        result = CollectionResult(
            sample_id="accounting_balance_and_turnovers",
            collection_name=collection_name,
            purpose="Виртуальная таблица остатков и оборотов управленческого регистра.",
            ok=False,
            scanned_rows=len(rows),
            page_count=page_count,
            status_code=last_status,
            error=exc.__class__.__name__,
        )
        return result, [], []

    balance_rows = _balance_rows_from_balance_and_turnovers(rows, account_lookup)
    payload_to_write = {
        "value": rows,
        "_source": {
            "collection_name": collection_name,
            "period_start": period_start.isoformat(),
            "period_end_exclusive": period_end.isoformat(),
            "scanned_rows": len(rows),
            "page_count": page_count,
            "page_size": page_size,
            "max_pages": max_pages,
            "source_kind": BALANCE_AND_TURNOVERS_SOURCE.source_id,
        },
    }
    output_path = output_dir / "accounting_balance_and_turnovers.raw.json"
    _write_json(output_path, payload_to_write)
    result = CollectionResult(
        sample_id="accounting_balance_and_turnovers",
        collection_name=collection_name,
        purpose="Виртуальная таблица остатков и оборотов управленческого регистра.",
        ok=True,
        period_rows=rows,
        scanned_rows=len(rows),
        page_count=page_count,
        output_file=output_path.name,
        raw_payload_hash=raw_payload_hash(payload_to_write),
        status_code=last_status,
        has_period_rows=bool(rows),
        error=(
            ""
            if balance_rows or not rows
            else "BalanceAndTurnovers вернул строки, но поля ОСВ не удалось нормализовать; используется fallback RecordType."
        ),
    )
    return result, balance_rows, rows


def _balance_rows_from_balance_and_turnovers(
    rows: list[dict[str, Any]],
    account_lookup: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        account_key = _first_text(
            row,
            (
                "Account_Key",
                "Счет_Key",
                "Account",
                "Счет",
                "AccountDr_Key",
                "AccountCt_Key",
            ),
        )
        account = account_lookup.get(account_key, {})
        account_code = _first_text(
            row,
            ("Account_Code", "Счет_Code", "Code", "КодСчета", "НомерСчета"),
        ) or str(account.get("Code") or "")
        account_name = _first_text(
            row,
            (
                "Account_Description",
                "Счет_Description",
                "Account_Name",
                "Счет_Наименование",
                "НаименованиеСчета",
            ),
        ) or str(account.get("Description") or "")
        item = {
            "account_key": account_key,
            "account_code": account_code,
            "account_name": account_name,
            "opening_debit": _first_amount_by_kind(row, "opening", "debit"),
            "opening_credit": _first_amount_by_kind(row, "opening", "credit"),
            "debit_turnover": _first_amount_by_kind(row, "turnover", "debit"),
            "credit_turnover": _first_amount_by_kind(row, "turnover", "credit"),
            "closing_debit": _first_amount_by_kind(row, "closing", "debit"),
            "closing_credit": _first_amount_by_kind(row, "closing", "credit"),
            "period_row_count": 1,
            "pre_period_row_count": "",
        }
        if not any(
            item[key]
            for key in (
                "opening_debit",
                "opening_credit",
                "debit_turnover",
                "credit_turnover",
                "closing_debit",
                "closing_credit",
            )
        ):
            continue
        item["opening_net"] = item["opening_debit"] - item["opening_credit"]
        item["closing_net"] = item["closing_debit"] - item["closing_credit"]
        item["required_regulation_account"] = _is_required_account(account_code)
        result.append(item)
    result.sort(key=lambda item: _account_sort_key(str(item["account_code"])))
    return result


def _fetch_accounting_records(
    *,
    client: OnecODataClient,
    output_dir: Path,
    period_start: datetime,
    period_end: datetime,
    account_lookup: dict[str, dict[str, Any]],
    page_size: int,
    max_pages: int,
    tail_max_pages: int,
    start_skip: int,
) -> tuple[CollectionResult, list[dict[str, Any]], list[dict[str, Any]]]:
    collection_name = "AccountingRegister_Управленческий_RecordType"
    period_rows: list[dict[str, Any]] = []
    aggregates: dict[str, dict[str, Any]] = {}
    scanned_rows = 0
    page_count = 0
    min_date: datetime | None = None
    max_date: datetime | None = None
    found_period = False
    last_status: int | None = None
    capped_by_max_pages = True
    tail_scan_error = ""
    tail_page_count = 0
    tail_completed = False
    tail_filter_start: datetime | None = None
    tail_filter_error = ""
    tail_method = ""

    def bucket(account_key: str) -> dict[str, Any]:
        account = account_lookup.get(account_key, {})
        if account_key not in aggregates:
            aggregates[account_key] = {
                "account_key": account_key,
                "account_code": str(account.get("Code") or ""),
                "account_name": str(account.get("Description") or ""),
                "opening_net": 0.0,
                "debit_turnover": 0.0,
                "credit_turnover": 0.0,
                "closing_net": 0.0,
                "period_row_count": 0,
                "pre_period_row_count": 0,
            }
        return aggregates[account_key]

    def process_accounting_rows(rows: list[dict[str, Any]]) -> None:
        nonlocal found_period
        for row in rows:
            row_date = _parse_datetime(row.get("Period"))
            if (
                row_date is None
                or row_date >= period_end
                or not _is_active(row.get("Active"))
            ):
                continue
            amount = _as_float(row.get("Сумма"))
            debit_key = str(row.get("AccountDr_Key") or "")
            credit_key = str(row.get("AccountCr_Key") or "")
            in_period = period_start <= row_date < period_end
            if in_period:
                period_rows.append(_with_account_labels(row, account_lookup))
                found_period = True
            if debit_key:
                item = bucket(debit_key)
                item["closing_net"] += amount
                if in_period:
                    item["debit_turnover"] += amount
                    item["period_row_count"] += 1
                else:
                    item["opening_net"] += amount
                    item["pre_period_row_count"] += 1
            if credit_key:
                item = bucket(credit_key)
                item["closing_net"] -= amount
                if in_period:
                    item["credit_turnover"] += amount
                    item["period_row_count"] += 1
                else:
                    item["opening_net"] -= amount
                    item["pre_period_row_count"] += 1

    try:
        for page_index in range(max_pages):
            skip = start_skip + page_index * page_size
            payload, status = client.fetch_collection(
                collection_name,
                top=page_size,
                skip=skip,
            )
            last_status = status
            rows = [row for row in extract_odata_rows(payload) if isinstance(row, dict)]
            page_count += 1
            scanned_rows += len(rows)
            page_dates = [
                item
                for item in (_parse_datetime(row.get("Period")) for row in rows)
                if item is not None
            ]
            if page_dates:
                page_min = min(page_dates)
                page_max = max(page_dates)
                min_date = page_min if min_date is None else min(min_date, page_min)
                max_date = page_max if max_date is None else max(max_date, page_max)
            process_accounting_rows(rows)
            if len(rows) < page_size:
                capped_by_max_pages = False
                break
            if found_period and page_dates and min(page_dates) >= period_end:
                capped_by_max_pages = False
                break
    except httpx.HTTPStatusError as exc:
        result = CollectionResult(
            sample_id="accounting_register_records",
            collection_name=collection_name,
            purpose="Проводки управленческого регистра бухгалтерии.",
            ok=False,
            scanned_rows=scanned_rows,
            page_count=page_count,
            status_code=exc.response.status_code,
            error=f"HTTP {exc.response.status_code}: {_odata_error_message(exc.response.text)}",
        )
        return result, [], []

    def update_page_dates(rows: list[dict[str, Any]]) -> list[datetime]:
        nonlocal min_date, max_date
        page_dates = [
            item
            for item in (_parse_datetime(row.get("Period")) for row in rows)
            if item is not None
        ]
        if page_dates:
            page_min = min(page_dates)
            page_max = max(page_dates)
            min_date = page_min if min_date is None else min(min_date, page_min)
            max_date = page_max if max_date is None else max(max_date, page_max)
        return page_dates

    def fetch_tail_page(
        *,
        skip: int,
        params: dict[str, str] | None = None,
    ) -> tuple[list[dict[str, Any]], list[datetime]]:
        nonlocal last_status, scanned_rows, tail_page_count
        payload, status = client.fetch_collection(
            collection_name,
            top=page_size,
            skip=skip,
            params=params,
        )
        last_status = status
        rows = [row for row in extract_odata_rows(payload) if isinstance(row, dict)]
        tail_page_count += 1
        scanned_rows += len(rows)
        page_dates = update_page_dates(rows)
        process_accounting_rows(rows)
        return rows, page_dates

    if capped_by_max_pages and max_date is not None and max_date < period_end:
        tail_filter_start = max_date
        tail_filter = (
            f"Period gt {_odata_datetime_literal(tail_filter_start)} "
            f"and Period lt {_odata_datetime_literal(period_end)}"
        )
        tail_method = "period_filter"
        try:
            for tail_page_index in range(tail_max_pages):
                rows, _page_dates = fetch_tail_page(
                    skip=tail_page_index * page_size,
                    params={"$filter": tail_filter},
                )
                if len(rows) < page_size:
                    tail_completed = True
                    capped_by_max_pages = False
                    break
        except httpx.HTTPStatusError as exc:
            tail_filter_error = (
                "Хвостовая догрузка регистра по Period не выполнена: "
                f"HTTP {exc.response.status_code}: {_odata_error_message(exc.response.text)}"
            )
            tail_method = "period_filter_failed"

        if not tail_completed and tail_filter_error:
            tail_method = (
                "period_filter_then_skip_continuation"
                if tail_filter_error
                else "skip_continuation"
            )
            continuation_start_skip = start_skip + max_pages * page_size
            try:
                for tail_page_index in range(tail_max_pages):
                    rows, page_dates = fetch_tail_page(
                        skip=continuation_start_skip + tail_page_index * page_size,
                    )
                    if len(rows) < page_size:
                        tail_completed = True
                        capped_by_max_pages = False
                        break
                    if page_dates and min(page_dates) >= period_end:
                        tail_completed = True
                        capped_by_max_pages = False
                        break
            except httpx.HTTPStatusError as exc:
                continuation_error = (
                    "Хвостовая догрузка регистра продолжением pagination не выполнена: "
                    f"HTTP {exc.response.status_code}: {_odata_error_message(exc.response.text)}"
                )
                tail_scan_error = (
                    f"{tail_filter_error} {continuation_error}"
                    if tail_filter_error
                    else continuation_error
                )

            if not tail_completed and not tail_scan_error:
                continuation_error = (
                    "Хвостовая догрузка регистра продолжением pagination достигла "
                    f"лимита tail_max_pages={tail_max_pages}."
                )
                tail_scan_error = (
                    f"{tail_filter_error} {continuation_error}"
                    if tail_filter_error
                    else continuation_error
                )
        elif not tail_completed:
            tail_scan_error = (
                "Хвостовая догрузка регистра по Period достигла "
                f"лимита tail_max_pages={tail_max_pages}."
            )

    balance_rows = []
    for item in aggregates.values():
        opening = float(item["opening_net"])
        closing = float(item["closing_net"])
        row = dict(item)
        row["opening_debit"] = opening if opening > 0 else 0.0
        row["opening_credit"] = -opening if opening < 0 else 0.0
        row["closing_debit"] = closing if closing > 0 else 0.0
        row["closing_credit"] = -closing if closing < 0 else 0.0
        row["required_regulation_account"] = _is_required_account(row["account_code"])
        balance_rows.append(row)
    balance_rows.sort(key=lambda row: _account_sort_key(row["account_code"]))

    payload_to_write = {
        "value": period_rows,
        "_source": {
            "collection_name": collection_name,
            "date_field": "Period",
            "period_start": period_start.isoformat(),
            "period_end_exclusive": period_end.isoformat(),
            "scanned_rows": scanned_rows,
            "page_count": page_count,
            "page_size": page_size,
            "max_pages": max_pages,
            "start_skip": start_skip,
            "tail_max_pages": tail_max_pages,
            "tail_page_count": tail_page_count,
            "tail_method": tail_method,
            "tail_filter_start": (
                tail_filter_start.isoformat() if tail_filter_start else None
            ),
            "tail_filter_end": period_end.isoformat() if tail_filter_start else None,
            "tail_completed": tail_completed,
            "tail_filter_error": tail_filter_error,
            "tail_scan_error": tail_scan_error,
            "opening_balance_source": (
                "Computed from earlier rows read by pagination."
                if start_skip == 0
                else "Partial: earlier rows before start_skip were not scanned."
            ),
            "pagination_status": _accounting_pagination_status(
                capped_by_max_pages=capped_by_max_pages,
                tail_completed=tail_completed,
                tail_scan_error=tail_scan_error,
            ),
        },
    }
    output_path = output_dir / "accounting_register_records.raw.json"
    _write_json(output_path, payload_to_write)
    result = CollectionResult(
        sample_id="accounting_register_records",
        collection_name=collection_name,
        purpose="Проводки управленческого регистра бухгалтерии.",
        ok=True,
        period_rows=period_rows,
        scanned_rows=scanned_rows,
        page_count=page_count,
        output_file=output_path.name,
        raw_payload_hash=raw_payload_hash(payload_to_write),
        status_code=last_status,
        error=(
            _accounting_records_error(
                start_skip=start_skip,
                capped_by_max_pages=capped_by_max_pages,
                max_pages=max_pages,
                max_date=max_date,
                period_start=period_start,
                period_end=period_end,
                tail_scan_error=tail_scan_error,
            )
        ),
        min_date=min_date.isoformat() if min_date else None,
        max_date=max_date.isoformat() if max_date else None,
        has_period_rows=bool(period_rows),
    )
    return result, balance_rows, period_rows


def _probe_virtual_accounting_functions(
    *,
    settings: OnecODataSettings,
    period_start: datetime,
    period_end: datetime,
) -> list[VirtualProbe]:
    probes: list[VirtualProbe] = []
    base = settings.base_url.rstrip("/")
    register = quote("AccountingRegister_Управленческий", safe="")
    variants = [
        (
            "BalanceAndTurnovers",
            "BalanceAndTurnovers("
            f"StartPeriod=datetime'{period_start:%Y-%m-%dT%H:%M:%S}',"
            f"EndPeriod=datetime'{period_end:%Y-%m-%dT%H:%M:%S}',"
            "AccountCondition='',Condition='',Dimensions='Организация')",
        ),
        (
            "Turnovers",
            "Turnovers("
            f"StartPeriod=datetime'{period_start:%Y-%m-%dT%H:%M:%S}',"
            f"EndPeriod=datetime'{period_end:%Y-%m-%dT%H:%M:%S}',"
            "AccountCondition='',BalancedAccountCondition='',"
            "Condition='',Dimensions='Организация')",
        ),
        (
            "DrCrTurnovers",
            "DrCrTurnovers("
            f"StartPeriod=datetime'{period_start:%Y-%m-%dT%H:%M:%S}',"
            f"EndPeriod=datetime'{period_end:%Y-%m-%dT%H:%M:%S}',"
            "AccountCondition='',BalancedAccountCondition='',"
            "Condition='',Dimensions='Организация')",
        ),
    ]
    with httpx.Client(
        auth=(settings.username, settings.password),
        headers={"Accept": "application/json"},
        timeout=settings.timeout_seconds,
        verify=settings.verify_ssl,
        follow_redirects=True,
    ) as client:
        for name, function_call in variants:
            try:
                response = client.get(
                    f"{base}/{register}/{function_call}",
                    params={"$format": "json", "$top": "5"},
                )
                rows: list[Any] = []
                fields: list[str] = []
                if response.status_code == 200:
                    rows = extract_odata_rows(response.json())
                    if rows and isinstance(rows[0], dict):
                        fields = sorted(rows[0].keys())
                probes.append(
                    VirtualProbe(
                        name=name,
                        ok=response.status_code == 200,
                        status_code=response.status_code,
                        row_count=len(rows),
                        fields=fields[:40],
                        error=(
                            ""
                            if response.status_code == 200
                            else _odata_error_message(response.text)
                        ),
                    )
                )
            except (httpx.HTTPError, ValueError) as exc:
                probes.append(
                    VirtualProbe(name=name, ok=False, error=exc.__class__.__name__)
                )
    return probes


def _fetch_service_names(settings: OnecODataSettings) -> set[str]:
    try:
        with httpx.Client(
            auth=(settings.username, settings.password),
            headers={"Accept": "application/json"},
            timeout=settings.timeout_seconds,
            verify=settings.verify_ssl,
            follow_redirects=True,
        ) as client:
            response = client.get(
                settings.base_url.rstrip("/"), params={"$format": "json"}
            )
            response.raise_for_status()
            payload = response.json()
    except (httpx.HTTPError, ValueError):
        return set()
    names: set[str] = set()

    def walk(value: Any) -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                if key in {"name", "url", "title"} and isinstance(item, str):
                    names.add(item)
                walk(item)
        elif isinstance(value, list):
            for item in value:
                walk(item)

    walk(payload)
    return names


def _build_coverage_rows(
    collection_results: list[CollectionResult],
    virtual_probes: list[VirtualProbe],
    service_names: set[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result in collection_results:
        if not result.ok:
            status = "Ошибка доступа"
        elif result.has_period_rows or result.sample_id in {
            "chart_of_accounts",
            "organizations",
            "tax_types",
            "manual_operation_postings",
            "tax_accrual_lines",
            "ens_operation_taxes",
        }:
            status = "Есть данные"
        else:
            status = "Пусто за период"
        rows.append(
            {
                "Блок": _coverage_block(result.sample_id),
                "Источник": result.collection_name,
                "Назначение": result.purpose,
                "Статус": status,
                "Строк за период": len(result.period_rows),
                "Просмотрено строк": result.scanned_rows,
                "Страниц": result.page_count,
                "Комментарий": result.error or "",
            }
        )
    for probe in virtual_probes:
        rows.append(
            {
                "Блок": "ОСВ",
                "Источник": f"AccountingRegister_Управленческий/{probe.name}",
                "Назначение": "Виртуальная таблица остатков/оборотов.",
                "Статус": "Доступна" if probe.ok else "Недоступна",
                "Строк за период": probe.row_count,
                "Просмотрено строк": probe.row_count,
                "Страниц": "",
                "Комментарий": probe.error,
            }
        )
    rows.append(
        {
            "Блок": "БУ/НУ",
            "Источник": "AccountingRegister_Хозрасчет",
            "Назначение": "Классический бухгалтерский регистр, если используется.",
            "Статус": (
                "Есть в service document"
                if "AccountingRegister_Хозрасчет" in service_names
                else "Не найден"
            ),
            "Строк за период": "",
            "Просмотрено строк": "",
            "Страниц": "",
            "Комментарий": "Нужен для полноценной ОСВ БУ/НУ.",
        }
    )
    return rows


def _coverage_block(sample_id: str) -> str:
    if sample_id in {
        "accounting_register_records",
        "accounting_balance_and_turnovers",
        "chart_of_accounts",
    }:
        return "ОСВ"
    if sample_id in {"taxes", "tax_types", "taxes_on_ens", "ens_sanctions"}:
        return "Налоги"
    if sample_id.startswith("ens"):
        return "ЕНС"
    if sample_id.startswith("vat"):
        return "НДС"
    if "operation" in sample_id or "correction" in sample_id:
        return "Ручные операции"
    if sample_id.startswith("bank"):
        return "Банк"
    if sample_id == "month_close_docs":
        return "Закрытие месяца"
    return "Справочник"


def _summarize_tax_rows(
    rows: list[dict[str, Any]],
    tax_lookup: dict[str, dict[str, Any]],
    org_lookup: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for row in rows:
        tax_key = str(row.get("ВидНалога_Key") or "")
        org_key = str(row.get("Организация_Key") or "")
        record_type = str(row.get("RecordType") or "")
        due_date = str(row.get("СрокУплаты") or "")
        key = (tax_key, org_key, record_type, due_date)
        item = groups.setdefault(
            key,
            {
                "Вид налога": _display_name(tax_lookup.get(tax_key), tax_key),
                "Организация": _display_name(org_lookup.get(org_key), org_key),
                "Тип движения": record_type,
                "Срок уплаты": due_date,
                "Сумма": 0.0,
                "Строк": 0,
            },
        )
        item["Сумма"] += _as_float(row.get("Сумма"))
        item["Строк"] += 1
    return sorted(
        groups.values(), key=lambda item: (item["Вид налога"], item["Срок уплаты"])
    )


def _summarize_vat(rows_by_id: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    result = []
    for sample_id, label in [
        ("vat_purchase_book", "Книга покупок"),
        ("vat_sales_book", "Книга продаж"),
        ("vat_presented", "НДС предъявленный"),
    ]:
        rows = rows_by_id.get(sample_id, [])
        totals = _numeric_totals(rows, ("Сумма", "НДС", "Стоимость"))
        result.append(
            {
                "Источник": label,
                "Строк": len(rows),
                "Сумма полей": _format_totals(totals),
                "Комментарий": "" if rows else "Нет строк за период в OData-снимке.",
            }
        )
    return result


def _summarize_manual_operations(
    rows_by_id: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    result = []
    for sample_id, label in [
        ("manual_operation_docs", "Операция"),
        ("register_corrections", "Корректировка регистров"),
        ("purchase_corrections", "Корректировка поступления"),
        ("sales_corrections", "Корректировка реализации"),
    ]:
        rows = rows_by_id.get(sample_id, [])
        result.append(
            {
                "Документ": label,
                "Строк за период": len(rows),
                "Проведенных": sum(1 for row in rows if _is_active(row.get("Posted"))),
                "Помечено на удаление": sum(
                    1 for row in rows if _is_active(row.get("DeletionMark"))
                ),
                "Комментарий": "Проверить вручную в 1С/CRM, если есть строки."
                if rows
                else "",
            }
        )
    return result


def _summarize_bank_rows(
    rows_by_id: dict[str, list[dict[str, Any]]],
    org_lookup: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    result = []
    for sample_id, direction in [("bank_in", "Поступление"), ("bank_out", "Расход")]:
        for row in rows_by_id.get(sample_id, []):
            org_key = str(row.get("Организация_Key") or "")
            result.append(
                {
                    "Направление": direction,
                    "Дата": row.get("Date"),
                    "Номер": row.get("Number"),
                    "Организация": _display_name(org_lookup.get(org_key), org_key),
                    "Проведен": row.get("Posted"),
                    "Сумма": _first_amount(row),
                    "Назначение/Комментарий": row.get("НазначениеПлатежа")
                    or row.get("Комментарий")
                    or "",
                }
            )
    return result


def _decorate_rows(
    rows: list[dict[str, Any]],
    org_lookup: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        item = dict(row)
        org_key = str(
            row.get("Организация_Key") or row.get("ГоловнаяОрганизация_Key") or ""
        )
        item["Организация_Имя"] = _display_name(org_lookup.get(org_key), org_key)
        result.append(item)
    return result


def _build_risks(
    *,
    coverage_rows: list[dict[str, Any]],
    account_balances: list[dict[str, Any]],
    rows_by_id: dict[str, list[dict[str, Any]]],
    service_names: set[str],
    virtual_probes: list[VirtualProbe],
    osv_source: OsvSourceInfo,
) -> list[dict[str, Any]]:
    risks: list[dict[str, Any]] = []
    if "AccountingRegister_Хозрасчет" not in service_names:
        risks.append(
            _risk(
                "Высокий",
                "БУ/НУ ОСВ",
                "В OData не найден AccountingRegister_Хозрасчет.",
                f"Текущий отчет использует управленческий источник: {osv_source.summary} Для регламента нужна бухгалтерская ОСВ БУ/НУ или подтверждение неприменимости.",
            )
        )
    if osv_source.source_id == RECORDTYPE_SOURCE.source_id:
        risks.append(
            _risk(
                "Средний",
                "ОСВ",
                "Виртуальные таблицы остатков/оборотов не вернули успешный ответ.",
                "ОСВ построена из проводок RecordType; это слабее штатного отчета 1С.",
            )
        )
    if any(
        "Быстрый режим" in str(row.get("Комментарий", ""))
        or "Лимит страниц" in str(row.get("Комментарий", ""))
        for row in coverage_rows
    ):
        risks.append(
            _risk(
                "Высокий",
                "ОСВ",
                "Реконструкция ОСВ частичная из-за неполного чтения регистра.",
                "Для финального регламентного закрытия нужна штатная ОСВ 1С или полный проход регистра до периода закрытия.",
            )
        )
    for prefix in ("26", "44"):
        if not any(
            str(row["account_code"]).startswith(prefix) for row in account_balances
        ):
            risks.append(
                _risk(
                    "Средний",
                    f"Счет {prefix}",
                    f"Счет {prefix} не найден в управленческом плане счетов.",
                    "Если регламент требует 26/44, нужен бухгалтерский план счетов или подтверждение неприменимости.",
                )
            )
    for prefix in ("10", "41", "43"):
        negatives = [
            row
            for row in account_balances
            if str(row["account_code"]).startswith(prefix)
            and row["closing_net"] < -0.01
        ]
        if negatives:
            risks.append(
                _risk(
                    "Высокий",
                    f"Счет {prefix}",
                    f"Есть кредитовый конечный остаток по счету {prefix}.",
                    "Проверить отрицательные остатки по складу/номенклатуре в 1С.",
                )
            )
    if not rows_by_id.get("vat_purchase_book"):
        risks.append(
            _risk(
                "Средний",
                "НДС",
                "Книга покупок НДС не дала строк за период.",
                "Проверить: это реально пусто, другой налоговый режим или неполная публикация OData.",
            )
        )
    if not rows_by_id.get("ens"):
        risks.append(
            _risk(
                "Высокий",
                "ЕНС",
                "Нет движений ЕНС за период.",
                "Для регламента нужен контроль ЕНС/68.90 и скрин ЕНС на дату закрытия.",
            )
        )
    if rows_by_id.get("manual_operation_docs") or rows_by_id.get(
        "register_corrections"
    ):
        risks.append(
            _risk(
                "Средний",
                "Ручные операции",
                "Есть ручные операции или корректировки за период.",
                "Проверить, были ли они после закрытия месяца, и приложить журнал операций.",
            )
        )
    if not rows_by_id.get("bank_out"):
        risks.append(
            _risk(
                "Средний",
                "Банк",
                "Нет расходов с расчетного счета за период в OData-снимке.",
                "Сверка факта уплаты налогов может быть неполной без банковских документов.",
            )
        )
    for row in coverage_rows:
        if row["Статус"] == "Ошибка доступа":
            risks.append(
                _risk(
                    "Средний",
                    str(row["Блок"]),
                    f"Ошибка чтения {row['Источник']}.",
                    str(row["Комментарий"]),
                )
            )
    return risks


def _build_accountant_requests(risks: list[dict[str, Any]]) -> list[dict[str, str]]:
    requests = [
        {
            "Что запросить": "Стандартная ОСВ БУ/НУ за май с субсчетами.",
            "Зачем": "Разово сверить онлайн-ОСВ с отчетом 1С; не использовать как регулярный источник.",
        },
        {
            "Что запросить": "Карточка/ОСВ 68.90 и скрин ЕНС на дату закрытия.",
            "Зачем": "Подтвердить ЕНС и 68.90 не только расчетом из OData.",
        },
        {
            "Что запросить": "Скрин отправки отчета по налоговой нагрузке клиенту.",
            "Зачем": "Регламент требует доказательство отправки, OData это не подтверждает.",
        },
        {
            "Что запросить": "Журнал операций с фильтром ручных корректировок за год.",
            "Зачем": "Регламент требует скрин журнала, а не только список документов из OData.",
        },
    ]
    if any(risk["Блок"] == "НДС" for risk in risks):
        requests.append(
            {
                "Что запросить": "Книга покупок/продаж НДС из 1С или подтверждение неприменимости НДС.",
                "Зачем": "В OData книга покупок за май пустая или неполная.",
            }
        )
    return requests


def _write_workbook(
    *,
    path: Path,
    generated_at: str,
    period_start: date,
    period_end: date,
    coverage_rows: list[dict[str, Any]],
    account_balances: list[dict[str, Any]],
    accounting_period_records: list[dict[str, Any]],
    month_close_rows: list[dict[str, Any]],
    tax_summary_rows: list[dict[str, Any]],
    ens_rows: list[dict[str, Any]],
    vat_rows: list[dict[str, Any]],
    manual_rows: list[dict[str, Any]],
    bank_rows: list[dict[str, Any]],
    rows_by_id: dict[str, list[dict[str, Any]]],
    virtual_probes: list[VirtualProbe],
    osv_source: OsvSourceInfo,
    risks: list[dict[str, Any]],
    requests: list[dict[str, str]],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Итог"
    verdict = (
        "Не подтверждено полностью"
        if any(r["Уровень"] == "Высокий" for r in risks)
        else "Частично закрыто через OData"
    )
    _write_key_values(
        ws,
        [
            (
                "Период",
                f"{period_start.isoformat()} - {period_end.isoformat()} (конец не включается)",
            ),
            ("Сформировано", generated_at),
            ("Итог", verdict),
            ("ОСВ", osv_source.summary),
            ("Источник ОСВ", osv_source.source_id),
            (
                "Виртуальная ОСВ",
                "Доступна"
                if any(p.ok for p in virtual_probes)
                else "Недоступна/ошибка вызова",
            ),
            ("БУ/НУ", "Хозрасчетный регистр в service document не найден."),
            ("Рисков", len(risks)),
        ],
    )
    _write_table(wb.create_sheet("Покрытие OData"), coverage_rows)
    _write_table(
        wb.create_sheet(osv_source.balance_sheet_name),
        _balance_report_rows(account_balances),
    )
    _write_table(
        wb.create_sheet("Счет 68"),
        _balance_report_rows(
            [
                row
                for row in account_balances
                if str(row["account_code"]).startswith("68")
            ]
        ),
    )
    _write_table(
        wb.create_sheet("ЕНС 68.90"),
        [
            *_balance_report_rows(
                [
                    row
                    for row in account_balances
                    if str(row["account_code"]).startswith("68.90")
                ]
            ),
            *[
                {
                    "Источник": "Регистр ЕНС",
                    "Дата": row.get("Period"),
                    "Организация": row.get("Организация_Имя"),
                    "Тип движения": row.get("RecordType"),
                    "Сумма": _as_float(row.get("Сумма")),
                    "Документ": row.get("Recorder_Type"),
                }
                for row in ens_rows
            ],
        ],
    )
    _write_table(wb.create_sheet("Налоги"), tax_summary_rows)
    _write_table(wb.create_sheet("НДС"), vat_rows)
    _write_table(
        wb.create_sheet("Закрытие месяца"),
        [_month_close_display(row) for row in month_close_rows],
    )
    _write_table(wb.create_sheet("Ручные операции"), manual_rows)
    _write_table(wb.create_sheet("Банк"), bank_rows)
    _write_table(
        wb.create_sheet("Налоговая нагрузка"),
        _tax_load_rows(rows_by_id, tax_summary_rows, ens_rows, bank_rows),
    )
    _write_table(
        wb.create_sheet(osv_source.sample_sheet_name),
        accounting_period_records[:5000],
    )
    _write_table(wb.create_sheet("Риски"), risks)
    _write_table(wb.create_sheet("Что запросить"), requests)
    _format_workbook(wb)
    wb.save(path)


def _write_key_values(ws: Any, rows: list[tuple[str, Any]]) -> None:
    ws.append(["Показатель", "Значение"])
    for key, value in rows:
        ws.append([key, value])


def _write_table(ws: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws.append(["Статус"])
        ws.append(["Нет строк"])
        return
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([_cell_value(row.get(header)) for header in headers])


def _format_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for idx, column in enumerate(ws.columns, 1):
            max_len = 10
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))
            ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 64)


def _balance_report_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "Счет": row["account_code"],
            "Наименование": row["account_name"],
            "Нач. Дт": round(row["opening_debit"], 2),
            "Нач. Кт": round(row["opening_credit"], 2),
            "Оборот Дт": round(row["debit_turnover"], 2),
            "Оборот Кт": round(row["credit_turnover"], 2),
            "Кон. Дт": round(row["closing_debit"], 2),
            "Кон. Кт": round(row["closing_credit"], 2),
            "Строк периода": row["period_row_count"],
            "Регламентный счет": "Да" if row["required_regulation_account"] else "",
        }
        for row in rows
        if (
            row["period_row_count"]
            or row["opening_debit"]
            or row["opening_credit"]
            or row["closing_debit"]
            or row["closing_credit"]
        )
    ]


def _month_close_display(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "Дата": row.get("Date"),
        "Номер": row.get("Number"),
        "Организация": row.get("Организация_Имя"),
        "Проведен": row.get("Posted"),
        "Пометка удаления": row.get("DeletionMark"),
        "Расчет фактической себестоимости": row.get("РасчетФактическойСебестоимости"),
        "Расчет финансового результата": row.get("РасчетФинансовогоРезультата"),
        "Зачет аванса по ЕНС": row.get("ЗачетАвансаПоЕдиномуНалоговомуСчету"),
        "Комментарий": row.get("Комментарий"),
    }


def _tax_load_rows(
    rows_by_id: dict[str, list[dict[str, Any]]],
    tax_summary_rows: list[dict[str, Any]],
    ens_rows: list[dict[str, Any]],
    bank_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return [
        {
            "Блок регламента": "Начисленные налоги",
            "Статус по OData": "Есть строки" if tax_summary_rows else "Нет строк",
            "Что есть": f"{len(tax_summary_rows)} агрегированных строк",
            "Что еще нужно": "Сверить с отчетом налоговой нагрузки и отправкой клиенту.",
        },
        {
            "Блок регламента": "ЕНС",
            "Статус по OData": "Есть строки" if ens_rows else "Нет строк",
            "Что есть": f"{len(ens_rows)} движений ЕНС",
            "Что еще нужно": "Скрин ЕНС на дату закрытия и карточка/ОСВ 68.90.",
        },
        {
            "Блок регламента": "Факт оплат",
            "Статус по OData": "Есть банк" if bank_rows else "Нет строк банка",
            "Что есть": f"{len(bank_rows)} банковских документов за период",
            "Что еще нужно": "План-факт график налоговых платежей.",
        },
        {
            "Блок регламента": "Проблемные зоны",
            "Статус по OData": "Частично",
            "Что есть": (
                f"ручные операции: {len(rows_by_id.get('manual_operation_docs', []))}; "
                f"корректировки: {len(rows_by_id.get('register_corrections', []))}"
            ),
            "Что еще нужно": "Клиентское сообщение и отметка встречи/обсуждения.",
        },
    ]


def _collection_manifest(result: CollectionResult) -> dict[str, Any]:
    return {
        "sample_id": result.sample_id,
        "collection_name": result.collection_name,
        "purpose": result.purpose,
        "ok": result.ok,
        "period_row_count": len(result.period_rows),
        "scanned_rows": result.scanned_rows,
        "page_count": result.page_count,
        "status_code": result.status_code,
        "output_file": result.output_file,
        "raw_payload_hash": result.raw_payload_hash,
        "min_date": result.min_date,
        "max_date": result.max_date,
        "error": result.error,
    }


def _with_account_labels(
    row: dict[str, Any],
    account_lookup: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    item = dict(row)
    debit = account_lookup.get(str(row.get("AccountDr_Key") or ""), {})
    credit = account_lookup.get(str(row.get("AccountCr_Key") or ""), {})
    item["Дт счет"] = debit.get("Code")
    item["Дт имя"] = debit.get("Description")
    item["Кт счет"] = credit.get("Code")
    item["Кт имя"] = credit.get("Description")
    return item


def _accounting_records_error(
    *,
    start_skip: int,
    capped_by_max_pages: bool,
    max_pages: int,
    max_date: datetime | None,
    period_start: datetime,
    period_end: datetime,
    tail_scan_error: str = "",
) -> str:
    messages = []
    if start_skip:
        messages.append(
            f"Быстрый режим: строки до skip={start_skip} не сканировались, начальное сальдо частичное."
        )
    if tail_scan_error:
        messages.append(tail_scan_error)
    if capped_by_max_pages:
        max_date_text = max_date.isoformat() if max_date else "не определена"
        messages.append(
            "Лимит страниц регистра достигнут "
            f"(max_pages={max_pages}, max_date={max_date_text}) "
            "до/без полного покрытия периода "
            f"{period_start.date().isoformat()} - {period_end.date().isoformat()}."
        )
    return " ".join(messages)


def _accounting_pagination_status(
    *,
    capped_by_max_pages: bool,
    tail_completed: bool,
    tail_scan_error: str,
) -> str:
    if tail_scan_error:
        return "main_capped_tail_failed"
    if tail_completed:
        return "main_capped_tail_completed"
    if capped_by_max_pages:
        return "capped_by_max_pages"
    return "complete_or_period_scanned"


def _odata_datetime_literal(value: datetime) -> str:
    return f"datetime'{value:%Y-%m-%dT%H:%M:%S}'"


def _lookup_by_key(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {str(row.get("Ref_Key")): row for row in rows if row.get("Ref_Key")}


def _display_name(row: dict[str, Any] | None, fallback: str) -> str:
    if not row:
        return fallback
    return str(row.get("Description") or row.get("Наименование") or fallback)


def _first_text(row: Mapping[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.get(key)
        if isinstance(value, (str, int, float)):
            text = str(value).strip()
            if text:
                return text
    return ""


def _first_amount_by_kind(
    row: Mapping[str, Any],
    balance_kind: str,
    side: str,
) -> float:
    exact_aliases = {
        ("opening", "debit"): (
            "СуммаOpeningBalanceDt",
            "СуммаOpeningBalanceDr",
            "AmountOpeningBalanceDt",
            "AmountOpeningBalanceDr",
            "OpeningBalanceDt",
            "OpeningBalanceDr",
            "OpeningDebit",
            "НачальныйОстатокДт",
            "НачальноеСальдоДт",
            "СальдоНачальноеДт",
        ),
        ("opening", "credit"): (
            "СуммаOpeningBalanceCt",
            "СуммаOpeningBalanceCr",
            "AmountOpeningBalanceCt",
            "AmountOpeningBalanceCr",
            "OpeningBalanceCt",
            "OpeningBalanceCr",
            "OpeningCredit",
            "НачальныйОстатокКт",
            "НачальноеСальдоКт",
            "СальдоНачальноеКт",
        ),
        ("turnover", "debit"): (
            "СуммаTurnoverDt",
            "СуммаTurnoverDr",
            "AmountTurnoverDt",
            "AmountTurnoverDr",
            "TurnoverDt",
            "TurnoverDr",
            "DebitTurnover",
            "ОборотДт",
            "ДебетовыйОборот",
        ),
        ("turnover", "credit"): (
            "СуммаTurnoverCt",
            "СуммаTurnoverCr",
            "AmountTurnoverCt",
            "AmountTurnoverCr",
            "TurnoverCt",
            "TurnoverCr",
            "CreditTurnover",
            "ОборотКт",
            "КредитовыйОборот",
        ),
        ("closing", "debit"): (
            "СуммаClosingBalanceDt",
            "СуммаClosingBalanceDr",
            "AmountClosingBalanceDt",
            "AmountClosingBalanceDr",
            "ClosingBalanceDt",
            "ClosingBalanceDr",
            "ClosingDebit",
            "КонечныйОстатокДт",
            "КонечноеСальдоДт",
            "СальдоКонечноеДт",
        ),
        ("closing", "credit"): (
            "СуммаClosingBalanceCt",
            "СуммаClosingBalanceCr",
            "AmountClosingBalanceCt",
            "AmountClosingBalanceCr",
            "ClosingBalanceCt",
            "ClosingBalanceCr",
            "ClosingCredit",
            "КонечныйОстатокКт",
            "КонечноеСальдоКт",
            "СальдоКонечноеКт",
        ),
    }
    for alias in exact_aliases.get((balance_kind, side), ()):
        if alias in row:
            return _as_float(row.get(alias))

    kind_tokens = {
        "opening": ("opening", "initial", "begin", "нач", "вход"),
        "turnover": ("turnover", "оборот"),
        "closing": ("closing", "ending", "end", "кон", "исход"),
    }[balance_kind]
    side_tokens = {
        "debit": ("debit", "debet", "dt", "dr", "дт", "дебет"),
        "credit": ("credit", "kredit", "ct", "cr", "кт", "кредит"),
    }[side]
    for key, value in row.items():
        name = _normalized_field_name(key)
        if any(token in name for token in kind_tokens) and any(
            token in name for token in side_tokens
        ):
            return _as_float(value)
    return 0.0


def _normalized_field_name(value: Any) -> str:
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def _numeric_totals(
    rows: list[dict[str, Any]], needles: tuple[str, ...]
) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for row in rows:
        for key, value in row.items():
            if any(needle.lower() in str(key).lower() for needle in needles):
                amount = _as_float(value)
                if amount:
                    totals[str(key)] += amount
    return dict(sorted(totals.items()))


def _format_totals(totals: dict[str, float]) -> str:
    if not totals:
        return ""
    return "; ".join(f"{key}: {value:,.2f}" for key, value in totals.items())


def _first_amount(row: dict[str, Any]) -> float:
    for key in ("СуммаДокумента", "Сумма", "СуммаПлатежа"):
        if key in row:
            return _as_float(row.get(key))
    for key, value in row.items():
        if "Сумма" in str(key):
            amount = _as_float(value)
            if amount:
                return amount
    return 0.0


def _risk(level: str, block: str, finding: str, action: str) -> dict[str, str]:
    return {"Уровень": level, "Блок": block, "Наблюдение": finding, "Действие": action}


def _is_required_account(code: str) -> bool:
    return any(
        code.startswith(prefix)
        for prefix in ("10", "26", "41", "43", "44", "68", "68.90", "90", "91", "99")
    )


def _account_sort_key(code: str) -> tuple[Any, ...]:
    parts: list[str] = []
    for part in str(code).split("."):
        stripped = part.strip()
        parts.append(f"{int(stripped):08d}" if stripped.isdigit() else stripped)
    return tuple(parts)


def _parse_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _parse_date(value: str) -> date:
    return date.fromisoformat(value)


def _as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _is_active(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "1", "yes", "да"}


def _cell_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return json.dumps(value, ensure_ascii=False, default=str)


def _odata_error_message(text: str) -> str:
    try:
        payload = json.loads(text)
        message = payload.get("odata.error", {}).get("message", {}).get("value")
        if message:
            return str(message).replace("\n", " ")[:500]
    except ValueError:
        pass
    return text.replace("\n", " ")[:500]


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True, default=str),
        encoding="utf-8",
    )


def _default_data_dir() -> Path:
    timestamp = datetime.now(tz=MOSCOW_TZ).strftime("%Y%m%d-%H%M%S")
    return Path("data") / "onec_month_close_audit" / timestamp


def _default_report_path(period_start: date) -> Path:
    return Path("reports") / f"onec_month_close_audit_pack_{period_start:%Y_%m}.xlsx"


if __name__ == "__main__":
    sys.exit(main())
