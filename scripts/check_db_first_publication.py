#!/usr/bin/env python3
"""Check DB-first publication state without exposing secrets or raw rows."""

# ruff: noqa: E402

from __future__ import annotations

import argparse
import csv
import os
import shutil
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.exc import SQLAlchemyError

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from wb_unit_economics.web.database import make_engine, make_session_factory
from wb_unit_economics.web.database import schema_version as current_schema_version
from wb_unit_economics.web.models import (
    ReportArtifact,
    ReportLostSalesRow,
    ReportRun,
    ReportUnitRow,
    SourceRefreshRun,
    TenantIntegration,
)
from wb_unit_economics.web.settings import WebSettings

DEFAULT_EXCEL = ROOT / "reports" / "shumeyko_wb_excel_mvp.xlsx"
DEFAULT_UNIT_CSV = ROOT / "reports" / "db_first" / "csv" / "unitRows.csv"
DEFAULT_LOST_SALES_CSV = ROOT / "reports" / "db_first" / "csv" / "lostSales.csv"
REQUIRED_ARTIFACT_TYPES = {"csv", "docx", "excel", "html", "pdf"}
READY_INTEGRATION_STATUSES = {"configured", "check_ok"}
SHEET_HEADER_MARKERS = {
    "Юнит экономика": "Неделя",
    "Упущенные продажи": "Кабинет WB",
}


def main() -> int:
    args = _parse_args()
    settings = _settings(args)
    issues: list[str] = []
    warnings: list[str] = []
    try:
        engine = make_engine(settings.database_url)
        session_factory = make_session_factory(engine)
        with session_factory() as db:
            report = _report(db, tenant_id=args.tenant, report_id=args.report_id)
            print(f"Database type: {engine.dialect.name}")
            print(f"Schema version: {current_schema_version(engine)}")
            if args.require_postgres and engine.dialect.name != "postgresql":
                issues.append("database is not PostgreSQL")
            if report is None:
                issues.append(f"report not found: {args.report_id}")
            else:
                _check_report(db, report, args, issues, warnings)
            _check_integrations(db, args, issues, warnings)
            _check_source_refresh_disk(settings, warnings)
            _print_latest_refresh(db, tenant_id=args.tenant)
    except SQLAlchemyError as exc:
        print(f"DB-first publication check failed: {exc.__class__.__name__}")
        return 2

    if warnings:
        print("Warnings:")
        for item in warnings:
            print(f"- {item}")
    if issues:
        print("Health: failed")
        for item in issues:
            print(f"- {item}")
        return 1
    print("Health: ok")
    return 0


def _check_report(
    db: Any,
    report: ReportRun,
    args: argparse.Namespace,
    issues: list[str],
    warnings: list[str],
) -> None:
    unit_rows = _count(db, ReportUnitRow, report.id)
    lost_sales_rows = _count(db, ReportLostSalesRow, report.id)
    artifact_rows = db.scalars(
        select(ReportArtifact).where(ReportArtifact.report_run_id == report.id)
    ).all()
    ready_artifacts = [item for item in artifact_rows if item.status == "ready"]
    ready_types = Counter(item.artifact_type for item in ready_artifacts)

    print(f"Report: {report.id}")
    print(f"Publication status: {report.publication_status}")
    print(f"Current: {report.is_current}")
    print(f"Lineage: {report.lineage_type}")
    print(f"Generated at: {report.generated_at.isoformat()}")
    print(f"Unit rows: {unit_rows}")
    print(f"Lost sales rows: {lost_sales_rows}")
    print(f"Ready artifacts: {len(ready_artifacts)}")
    for artifact_type, count in sorted(ready_types.items()):
        print(f"- {artifact_type}: {count}")

    if report.publication_status != args.expected_publication_status:
        issues.append(
            "publication status mismatch: expected "
            f"{args.expected_publication_status}, got {report.publication_status}"
        )
    expected_current = args.expected_current == "true"
    if report.is_current is not expected_current:
        issues.append(
            f"current mismatch: expected {expected_current}, got {report.is_current}"
        )
    if report.lineage_type != "db_first_report_marts":
        issues.append("report lineage is not db_first_report_marts")
    if args.expected_unit_rows is not None and unit_rows != args.expected_unit_rows:
        issues.append(
            f"unitRows mismatch: expected {args.expected_unit_rows}, got {unit_rows}"
        )
    if (
        args.expected_lost_sales_rows is not None
        and lost_sales_rows != args.expected_lost_sales_rows
    ):
        issues.append(
            "lostSales mismatch: "
            f"expected {args.expected_lost_sales_rows}, got {lost_sales_rows}"
        )
    if args.expected_artifacts is not None and len(ready_artifacts) != (
        args.expected_artifacts
    ):
        issues.append(
            "artifact count mismatch: "
            f"expected {args.expected_artifacts}, got {len(ready_artifacts)}"
        )
    missing_types = REQUIRED_ARTIFACT_TYPES - set(ready_types)
    if missing_types:
        joined = ", ".join(sorted(missing_types))
        issues.append(f"missing ready artifact types: {joined}")

    _check_file_counts(
        args,
        db_unit_rows=unit_rows,
        db_lost_sales_rows=lost_sales_rows,
        issues=issues,
        warnings=warnings,
    )


def _check_file_counts(
    args: argparse.Namespace,
    *,
    db_unit_rows: int,
    db_lost_sales_rows: int,
    issues: list[str],
    warnings: list[str],
) -> None:
    if args.skip_file_counts:
        return

    excel_unit_rows, excel_lost_sales_rows = _excel_counts(args.excel_path)
    unit_csv_rows = _csv_data_rows(args.unit_csv_path)
    lost_csv_rows = _csv_data_rows(args.lost_sales_csv_path)
    pairs = [
        ("Excel unitRows", excel_unit_rows, db_unit_rows, args.excel_path),
        ("Excel lostSales", excel_lost_sales_rows, db_lost_sales_rows, args.excel_path),
        ("CSV unitRows", unit_csv_rows, db_unit_rows, args.unit_csv_path),
        ("CSV lostSales", lost_csv_rows, db_lost_sales_rows, args.lost_sales_csv_path),
    ]
    for label, observed, expected, path in pairs:
        if observed is None:
            message = f"{label} file count unavailable: {path}"
            if args.require_files:
                issues.append(message)
            else:
                warnings.append(message)
            continue
        print(f"{label}: {observed}")
        if observed != expected:
            issues.append(f"{label} mismatch: expected {expected}, got {observed}")


def _check_integrations(
    db: Any,
    args: argparse.Namespace,
    issues: list[str],
    warnings: list[str],
) -> None:
    integrations = db.scalars(
        select(TenantIntegration)
        .where(TenantIntegration.tenant_id == args.tenant)
        .order_by(TenantIntegration.provider)
    ).all()
    print(f"Tenant integrations: {len(integrations)}")
    if not integrations:
        message = "tenant integrations are not configured"
        if args.require_integrations:
            issues.append(message)
        else:
            warnings.append(message)
        return
    not_ready: list[str] = []
    for item in integrations:
        payload = item.config_payload or {}
        storage = str(payload.get("storage") or "hash_only")
        print(f"- {item.provider}: status={item.status}, storage={storage}")
        if item.status not in READY_INTEGRATION_STATUSES or storage == "hash_only":
            not_ready.append(f"{item.provider}:{item.status}:{storage}")
    if not_ready:
        message = "tenant integrations are not runtime-ready: " + ", ".join(not_ready)
        if args.require_integrations:
            issues.append(message)
        else:
            warnings.append(message)


def _print_latest_refresh(db: Any, *, tenant_id: str) -> None:
    refresh = db.scalar(
        select(SourceRefreshRun)
        .where(SourceRefreshRun.tenant_id == tenant_id)
        .order_by(SourceRefreshRun.created_at.desc())
    )
    if refresh is None:
        print("Latest source refresh: none")
        return
    print(
        "Latest source refresh: "
        f"{refresh.mode} {refresh.status} {refresh.created_at.isoformat()}"
    )
    if refresh.new_report_run_id:
        print(f"Latest source refresh report: {refresh.new_report_run_id}")


def _check_source_refresh_disk(settings: WebSettings, warnings: list[str]) -> None:
    status_line, warning = _source_refresh_disk_status(settings)
    if status_line:
        print(status_line)
    if warning:
        warnings.append(warning)


def _source_refresh_disk_status(settings: WebSettings) -> tuple[str, str | None]:
    min_free_gb = max(0.0, float(settings.source_refresh_min_free_gb))
    if min_free_gb <= 0:
        return "", None
    source_root = settings.source_refresh_root_path
    probe_path = _existing_path_for_disk_check(source_root)
    free_gb = shutil.disk_usage(probe_path).free / (1024**3)
    status = (
        "Source refresh root free GiB: "
        f"{free_gb:.2f} (required {min_free_gb:.2f})"
    )
    if free_gb >= min_free_gb:
        return status, None
    return (
        status,
        f"source refresh low disk: free={free_gb:.2f}GiB "
        f"required={min_free_gb:.2f}GiB",
    )


def _existing_path_for_disk_check(path: Path) -> Path:
    current = path.resolve()
    while not current.exists() and current.parent != current:
        current = current.parent
    return current


def _count(db: Any, model: type, report_id: str) -> int:
    return int(
        db.scalar(
            select(func.count())
            .select_from(model)
            .where(model.report_run_id == report_id)
        )
        or 0
    )


def _report(db: Any, *, tenant_id: str, report_id: str) -> ReportRun | None:
    return db.scalar(
        select(ReportRun).where(
            ReportRun.tenant_id == tenant_id,
            ReportRun.id == report_id,
        )
    )


def _csv_data_rows(path: Path) -> int | None:
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        row_count = sum(1 for _row in reader)
    return max(0, row_count - 1)


def _excel_counts(path: Path) -> tuple[int | None, int | None]:
    if not path.exists():
        return None, None
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        unit_rows = _sheet_data_rows(workbook, "Юнит экономика")
        lost_sales_rows = _sheet_data_rows(workbook, "Упущенные продажи")
        return unit_rows, lost_sales_rows
    finally:
        workbook.close()


def _sheet_data_rows(workbook: Any, sheet_name: str) -> int | None:
    if sheet_name not in workbook.sheetnames:
        return None
    sheet = workbook[sheet_name]
    header_row = _header_row(sheet, SHEET_HEADER_MARKERS.get(sheet_name))
    if header_row is None:
        return max(0, int(sheet.max_row or 0) - 1)
    data_rows = 0
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if sheet_name == "Упущенные продажи":
            # The exported sheet may prepend a visible coverage warning with an
            # empty cabinet.  It is metadata, not a persisted lost-sales row.
            is_data_row = bool(row and row[0] not in (None, ""))
        else:
            is_data_row = any(cell not in (None, "") for cell in row)
        if is_data_row:
            data_rows += 1
    return data_rows


def _header_row(sheet: Any, marker: str | None) -> int | None:
    if marker is None:
        return None
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        first_cell = row[0] if row else None
        if first_cell == marker:
            return row_index
    return None


def _settings(args: argparse.Namespace) -> WebSettings:
    database_url = args.database_url or os.getenv("SHUMEYKO_DATABASE_URL", "")
    if database_url:
        return WebSettings(_env_file=None, database_url=database_url)
    return WebSettings(_env_file=None)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--database-url", default="")
    parser.add_argument("--tenant", default="shumeyko")
    parser.add_argument("--report-id", required=True)
    parser.add_argument("--expected-unit-rows", type=int, default=None)
    parser.add_argument("--expected-lost-sales-rows", type=int, default=None)
    parser.add_argument("--expected-artifacts", type=int, default=None)
    parser.add_argument(
        "--expected-publication-status",
        choices=("draft", "published", "failed"),
        default="published",
    )
    parser.add_argument(
        "--expected-current",
        choices=("true", "false"),
        default="true",
    )
    parser.add_argument("--excel-path", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--unit-csv-path", type=Path, default=DEFAULT_UNIT_CSV)
    parser.add_argument(
        "--lost-sales-csv-path", type=Path, default=DEFAULT_LOST_SALES_CSV
    )
    parser.add_argument("--skip-file-counts", action="store_true")
    parser.add_argument("--require-files", action="store_true")
    parser.add_argument("--require-postgres", action="store_true")
    parser.add_argument("--require-integrations", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(main())
