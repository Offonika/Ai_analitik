from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DEFAULT_INPUT_DIR = Path("reports/power_query_client_package/marts_csv")
DEFAULT_OUTPUT = Path(
    "reports/Excel-дашборд Power Query WB-1C Шумейко и Партнеры "
    "01.03.2026-17.06.2026.xlsx"
)
PERIOD_TEXT = "01.03.2026 - 17.06.2026"
PERIOD_NOTE = "март, апрель, май, июнь; июнь неполный, по 17.06.2026"


@dataclass(frozen=True)
class Palette:
    navy: str = "0B3A53"
    blue: str = "2F80C4"
    sky: str = "EAF3FB"
    pale: str = "F7FAFC"
    line: str = "A9C9E8"
    green: str = "D9EAD3"
    yellow: str = "FFF2CC"
    red: str = "F4CCCC"
    white: str = "FFFFFF"
    text: str = "1F2933"
    muted: str = "5B6770"


MART_FILES = {
    "kpi": "kpi_period.csv",
    "monthly": "monthly_dynamics.csv",
    "expenses": "expenses.csv",
    "unit": "unit_economics.csv",
    "returns": "returns.csv",
    "lost": "lost_sales.csv",
    "reconciliation": "onec_opiu_reconciliation.csv",
}


def main() -> int:
    args = _parse_args()
    build_power_query_excel_dashboard(input_dir=args.input_dir, output=args.output)
    print(args.output)
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a polished Excel dashboard from Power Query BI marts."
    )
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def build_power_query_excel_dashboard(*, input_dir: Path, output: Path) -> Path:
    marts = {
        name: _read_csv(input_dir / filename) for name, filename in MART_FILES.items()
    }
    workbook = Workbook()
    workbook.remove(workbook.active)
    palette = Palette()
    styles = _styles(palette)

    dashboard = workbook.create_sheet("Дашборд", 0)
    _write_dashboard(dashboard, marts, styles)
    _write_monthly_sheet(workbook.create_sheet("Динамика"), marts["monthly"], styles)
    _write_expenses_sheet(
        workbook.create_sheet("Расходы WB"), marts["expenses"], styles
    )
    _write_returns_sheet(workbook.create_sheet("Возвраты"), marts["returns"], styles)
    _write_lost_sales_sheet(
        workbook.create_sheet("Упущенные продажи"), marts["lost"], styles
    )
    _write_reconciliation_sheet(
        workbook.create_sheet("Сверка с 1С ОПиУ"), marts["reconciliation"], styles
    )
    _write_instruction_sheet(workbook.create_sheet("Инструкция"), styles)
    _write_hidden_data_sheets(workbook, marts, styles)

    workbook.properties.title = "Excel-дашборд WB/1C Power Query"
    workbook.properties.subject = "Юнит-экономика WB, витрины bi, март-июнь 2026"
    workbook.properties.creator = "Шумейко и Партнеры / Codex"
    workbook.properties.keywords = "Power Query, Wildberries, 1C, dashboard"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    validate_dashboard_workbook(output)
    return output


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        rows = [{key: _coerce(value) for key, value in row.items()} for row in reader]
        return list(reader.fieldnames or []), rows


def _coerce(value: str | None) -> Any:
    if value is None:
        return None
    value = value.strip()
    if value == "":
        return None
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return value


def _styles(palette: Palette) -> dict[str, Any]:
    thin = Side(style="thin", color=palette.line)
    return {
        "title": Font(bold=True, size=20, color=palette.navy),
        "subtitle": Font(size=11, color=palette.muted),
        "section": Font(bold=True, size=13, color=palette.navy),
        "header": Font(bold=True, color=palette.white),
        "normal": Font(size=10, color=palette.text),
        "muted": Font(size=9, color=palette.muted),
        "kpi_label": Font(bold=True, size=9, color=palette.muted),
        "kpi_value": Font(bold=True, size=15, color=palette.navy),
        "fill_header": PatternFill("solid", fgColor=palette.navy),
        "fill_light": PatternFill("solid", fgColor=palette.sky),
        "fill_soft": PatternFill("solid", fgColor=palette.pale),
        "fill_good": PatternFill("solid", fgColor=palette.green),
        "fill_warn": PatternFill("solid", fgColor=palette.yellow),
        "fill_bad": PatternFill("solid", fgColor=palette.red),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _write_dashboard(
    ws, marts: dict[str, tuple[list[str], list[dict[str, Any]]]], styles: dict[str, Any]
) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Дашборд юнит-экономики WB/1C"
    ws["A1"].font = styles["title"]
    ws["A2"] = (
        f"Период: {PERIOD_TEXT}; {PERIOD_NOTE}. Обновляется из BI-витрин Power Query."
    )
    ws["A2"].font = styles["subtitle"]
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    kpi = _kpi_dict(marts["kpi"][1])
    lost_rows = marts["lost"][1]
    kpi_cards = [
        ("Выручка после СПП", kpi.get("Выручка после СПП"), "rub"),
        (
            "Управленческая прибыль WB",
            kpi.get("Управленческая прибыль WB")
            or kpi.get("Маржинальный доход WB после налогов"),
            "rub",
        ),
        (
            "Маржа WB без НДС",
            kpi.get("Маржа WB без НДС") or kpi.get("Маржа WB после налогов"),
            "pct",
        ),
        ("Продажи, шт", kpi.get("Продажи"), "int"),
        ("Возвраты, шт", kpi.get("Возвраты"), "int"),
        ("% возвратов", kpi.get("Доля возвратов"), "pct"),
        (
            "SKU с упущенной прибылью",
            len(
                [
                    row
                    for row in lost_rows
                    if _num(row.get("Потенциально упущенная прибыль")) > 0
                ]
            ),
            "int",
        ),
        (
            "Потенциально упущенная прибыль",
            sum(_num(row.get("Потенциально упущенная прибыль")) for row in lost_rows),
            "rub",
        ),
    ]
    _write_kpi_cards(ws, 4, kpi_cards, styles)

    monthly_headers, monthly_rows = marts["monthly"]
    _write_section_title(ws, "A11", "Помесячная динамика", styles)
    monthly_small = [
        [
            row.get("Месяц"),
            row.get("Выручка после СПП"),
            row.get("Расходы WB"),
            row.get("Управленческая прибыль WB")
            or row.get("Маржинальный доход WB после налогов"),
            row.get("Маржа WB без НДС") or row.get("Маржа WB после налогов"),
        ]
        for row in monthly_rows
    ]
    _write_table(
        ws,
        "A12",
        ["Месяц", "Выручка после СПП", "Расходы WB", "Упр. прибыль", "Маржа"],
        monthly_small,
        styles,
        table_name="dash_monthly",
    )
    _format_range(ws, 13, 13 + len(monthly_small) - 1, 2, 4, "rub")
    _format_range(ws, 13, 13 + len(monthly_small) - 1, 5, 5, "pct")
    _add_line_chart(
        ws,
        title="Выручка и маржинальный доход по месяцам",
        anchor="A19",
        min_row=12,
        max_row=12 + len(monthly_small),
        min_col=1,
        max_col=4,
    )

    top_loss = _top_loss_rows(marts["unit"][1], limit=10)
    _write_section_title(ws, "G11", "Топ убыточных товаров", styles)
    _write_table(
        ws,
        "G12",
        ["Товар", "Артикул 1С", "Баркод", "Выручка", "Упр. прибыль", "Причина"],
        [
            [
                row["Товар"],
                row["Артикул 1С"],
                row["Баркод"],
                row["Выручка после СПП"],
                row["Управленческая прибыль WB"],
                row["Причина"],
            ]
            for row in top_loss
        ],
        styles,
        table_name="dash_losses",
    )
    _format_range(ws, 13, 12 + len(top_loss), 10, 11, "rub")

    _write_section_title(ws, "G26", "Качество данных и ограничения", styles)
    quality_rows = [
        [row.get("Строк"), row.get("Значение")]
        for row in marts["kpi"][1]
        if row.get("Статус данных")
    ]
    if not quality_rows:
        quality_rows = [
            ["Ограничение", "Причины возвратов не передаются текущими источниками"]
        ]
    _write_table(
        ws,
        "G27",
        ["Показатель", "Значение"],
        quality_rows[:8],
        styles,
        table_name="dash_quality",
    )

    for col, width in {
        "A": 18,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 14,
        "G": 28,
        "H": 16,
        "I": 18,
        "J": 16,
        "K": 18,
        "L": 36,
    }.items():
        ws.column_dimensions[col].width = width


def _write_monthly_sheet(
    ws, mart: tuple[list[str], list[dict[str, Any]]], styles: dict[str, Any]
) -> None:
    _setup_titled_sheet(ws, "Помесячная динамика", styles)
    headers, rows = mart
    _write_table(
        ws,
        "A4",
        headers,
        [[row.get(header) for header in headers] for row in rows],
        styles,
        table_name="monthly_dynamics",
    )
    _autofit(ws, min_width=12, max_width=28)
    if rows:
        _format_known_columns(ws, headers, 5, 4 + len(rows))
        _add_line_chart(
            ws,
            "Выручка, расходы и маржинальный доход",
            "A12",
            4,
            4 + len(rows),
            1,
            min(12, len(headers)),
        )


def _write_expenses_sheet(
    ws, mart: tuple[list[str], list[dict[str, Any]]], styles: dict[str, Any]
) -> None:
    _setup_titled_sheet(ws, "Структура расходов WB", styles)
    headers, rows = mart
    visible_rows = rows[:12]
    _write_table(
        ws,
        "A4",
        headers,
        [[row.get(header) for header in headers] for row in visible_rows],
        styles,
        table_name="expenses",
    )
    _format_known_columns(ws, headers, 5, 4 + len(visible_rows))
    _autofit(ws, min_width=12, max_width=24)
    if visible_rows:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Расходы: сумма по статьям"
        chart.y_axis.title = "Статья"
        chart.x_axis.title = "Руб."
        data = Reference(ws, min_col=2, min_row=4, max_row=4 + len(visible_rows))
        cats = Reference(ws, min_col=1, min_row=5, max_row=4 + len(visible_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 15
        ws.add_chart(chart, "L4")


def _write_returns_sheet(
    ws, mart: tuple[list[str], list[dict[str, Any]]], styles: dict[str, Any]
) -> None:
    _setup_titled_sheet(ws, "Возвраты", styles)
    _, rows = mart
    by_amount = sorted(
        rows, key=lambda row: _num(row.get("Сумма возвратов")), reverse=True
    )[:20]
    by_rate = sorted(
        [row for row in rows if _num(row.get("Продажи, шт")) >= 5],
        key=lambda row: _num(row.get("% возвратов")),
        reverse=True,
    )[:20]
    headers = [
        "Товар",
        "Артикул 1С",
        "Баркод",
        "Продажи, шт",
        "Возвраты, шт",
        "% возвратов",
        "Сумма возвратов",
        "Маржинальный доход",
    ]
    _write_section_title(ws, "A3", "Топ по сумме возвратов", styles)
    _write_table(
        ws,
        "A4",
        headers,
        [_returns_row(row) for row in by_amount],
        styles,
        table_name="returns_amount",
    )
    _write_section_title(ws, "J3", "Высокий процент возвратов", styles)
    _write_table(
        ws,
        "J4",
        headers,
        [_returns_row(row) for row in by_rate],
        styles,
        table_name="returns_rate",
    )
    _format_range(ws, 5, 4 + len(by_amount), 6, 6, "pct")
    _format_range(ws, 5, 4 + len(by_amount), 7, 8, "rub")
    _format_range(ws, 5, 4 + len(by_rate), 15, 15, "pct")
    _format_range(ws, 5, 4 + len(by_rate), 16, 17, "rub")
    _autofit(ws, min_width=10, max_width=30)


def _write_lost_sales_sheet(
    ws, mart: tuple[list[str], list[dict[str, Any]]], styles: dict[str, Any]
) -> None:
    _setup_titled_sheet(ws, "Упущенные продажи", styles)
    _, rows = mart
    top = sorted(
        rows,
        key=lambda row: _num(row.get("Потенциально упущенная прибыль")),
        reverse=True,
    )[:30]
    headers = [
        "Кабинет WB",
        "Товар",
        "Артикул 1С",
        "Баркод",
        "Дней без остатка WB",
        "Остаток 1С на складах, шт",
        "Склады 1С с остатком",
        "Продажи, шт",
        "Потенциально упущено, шт",
        "Потенциально упущенная выручка",
        "Потенциально упущенная прибыль",
        "Вывод",
    ]
    _write_table(
        ws,
        "A4",
        headers,
        [[row.get(header) for header in headers] for row in top],
        styles,
        table_name="lost_sales",
    )
    _format_range(ws, 5, 4 + len(top), 10, 11, "rub")
    _autofit(ws, min_width=10, max_width=36)
    if top:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Топ по потенциально упущенной прибыли"
        chart.y_axis.title = "Товар"
        chart.x_axis.title = "Руб."
        data = Reference(ws, min_col=11, min_row=4, max_row=min(14, 4 + len(top)))
        cats = Reference(ws, min_col=2, min_row=5, max_row=min(14, 4 + len(top)))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 16
        ws.add_chart(chart, "L4")


def _write_reconciliation_sheet(
    ws, mart: tuple[list[str], list[dict[str, Any]]], styles: dict[str, Any]
) -> None:
    _setup_titled_sheet(ws, "Сверка с 1С/ОПиУ", styles)
    headers, rows = mart
    _write_table(
        ws,
        "A4",
        headers,
        [[row.get(header) for header in headers] for row in rows],
        styles,
        table_name="onec_reconciliation",
    )
    _format_known_columns(ws, headers, 5, 4 + len(rows))
    _autofit(ws, min_width=14, max_width=54)
    ws["A2"] = (
        "Сравниваем не общую выручку ОПиУ, а контрольные блоки: "
        "себестоимость и расходы маркетплейса по месяцам."
    )
    ws["A2"].font = styles["subtitle"]


def _write_instruction_sheet(ws, styles: dict[str, Any]) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Как пользоваться файлом"
    ws["A1"].font = styles["title"]
    rows = [
        ("1", "Открыть файл в Excel и убедиться, что листы дашбордов заполнены."),
        (
            "2",
            "Для живого обновления использовать Power Query-файл "
            "и учетку read-only Postgres.",
        ),
        (
            "3",
            "После обновления витрин в базе пересобрать этот дашборд "
            "или загрузить обновленные таблицы Power Query на листы данных.",
        ),
        ("4", "Пароли, WB-токены, 1С-доступы и raw snapshots в файл не включаются."),
    ]
    _write_table(ws, "A4", ["Шаг", "Действие"], rows, styles, table_name="instruction")
    ws["A11"] = "Период"
    ws["A11"].font = styles["section"]
    ws["B11"] = f"{PERIOD_TEXT}; {PERIOD_NOTE}"
    ws["A13"] = "Важно"
    ws["A13"].font = styles["section"]
    ws["B13"] = (
        "Это клиентский визуальный слой поверх расчетных BI-витрин. "
        "Источник расчета остается в базе и Excel MVP."
    )
    ws["B13"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90


def _write_hidden_data_sheets(
    workbook: Workbook,
    marts: dict[str, tuple[list[str], list[dict[str, Any]]]],
    styles: dict[str, Any],
) -> None:
    for name, (headers, rows) in marts.items():
        ws = workbook.create_sheet(f"data_{name}")
        _setup_sheet(ws)
        _write_table(
            ws,
            "A1",
            headers,
            [[row.get(header) for header in headers] for row in rows],
            styles,
            table_name=f"data_{name}",
        )
        _autofit(ws, min_width=10, max_width=28)
        ws.sheet_state = "hidden"


def _kpi_dict(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {str(row.get("Показатель")): row.get("Значение") for row in rows}


def _top_loss_rows(rows: list[dict[str, Any]], *, limit: int) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, Any, Any], dict[str, Any]] = {}
    for row in rows:
        key = (row.get("Товар"), row.get("Артикул 1С"), row.get("Баркод"))
        target = grouped.setdefault(
            key,
            {
                "Товар": row.get("Товар"),
                "Артикул 1С": row.get("Артикул 1С"),
                "Баркод": row.get("Баркод"),
                "Выручка после СПП": 0.0,
                "Управленческая прибыль WB": 0.0,
                "Причина": row.get("Причина статуса") or row.get("Статус данных"),
            },
        )
        target["Выручка после СПП"] += _num(row.get("Выручка после СПП"))
        target["Управленческая прибыль WB"] += _num(
            row.get("Управленческая прибыль WB")
            or row.get("Маржинальный доход WB после налогов")
        )
        if row.get("Главная причина"):
            target["Причина"] = row.get("Главная причина")
    losses = [
        row
        for row in grouped.values()
        if row["Управленческая прибыль WB"] < 0
    ]
    return sorted(losses, key=lambda row: row["Управленческая прибыль WB"])[
        :limit
    ]


def _returns_row(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("Товар"),
        row.get("Артикул 1С"),
        row.get("Баркод"),
        row.get("Продажи, шт"),
        row.get("Возвраты, шт"),
        row.get("% возвратов"),
        row.get("Сумма возвратов"),
        row.get("Управленческая прибыль WB")
        or row.get("Маржинальный доход WB после налогов"),
    ]


def _num(value: Any) -> float:
    if isinstance(value, int | float):
        return float(value)
    return 0.0


def _write_kpi_cards(
    ws, start_row: int, cards: list[tuple[str, Any, str]], styles: dict[str, Any]
) -> None:
    for index, (label, value, fmt) in enumerate(cards):
        row = start_row + (index // 4) * 3
        col = 1 + (index % 4) * 3
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(
            start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1
        )
        label_cell = ws.cell(row, col, label)
        value_cell = ws.cell(row + 1, col, value)
        label_cell.font = styles["kpi_label"]
        value_cell.font = styles["kpi_value"]
        _set_number_format(value_cell, fmt)
        for r in range(row, row + 2):
            for c in range(col, col + 2):
                cell = ws.cell(r, c)
                cell.fill = styles["fill_light"]
                cell.border = styles["border"]
                cell.alignment = Alignment(wrap_text=True, vertical="center")


def _setup_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"


def _setup_titled_sheet(ws, title: str, styles: dict[str, Any]) -> None:
    _setup_sheet(ws)
    ws["A1"] = title
    ws["A1"].font = styles["title"]
    ws["A2"] = f"Период: {PERIOD_TEXT}; {PERIOD_NOTE}"
    ws["A2"].font = styles["subtitle"]


def _write_section_title(ws, cell_ref: str, title: str, styles: dict[str, Any]) -> None:
    ws[cell_ref] = title
    ws[cell_ref].font = styles["section"]


def _write_table(
    ws,
    start_cell: str,
    headers: list[str],
    rows: list[list[Any]] | list[tuple[Any, ...]],
    styles: dict[str, Any],
    *,
    table_name: str,
) -> None:
    start_col, start_row = _split_cell(start_cell)
    for index, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + index, header)
        cell.font = styles["header"]
        cell.fill = styles["fill_header"]
        cell.border = styles["border"]
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row):
            cell = ws.cell(start_row + row_offset, start_col + col_offset, value)
            cell.font = styles["normal"]
            cell.border = styles["border"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_offset % 2 == 0:
                cell.fill = styles["fill_soft"]
    end_row = start_row + max(len(rows), 1)
    end_col = start_col + len(headers) - 1
    start_ref = f"{get_column_letter(start_col)}{start_row}"
    end_ref = f"{get_column_letter(end_col)}{end_row}"
    ref = f"{start_ref}:{end_ref}"
    table = Table(displayName=_safe_table_name(table_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _split_cell(cell_ref: str) -> tuple[int, int]:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    digits = "".join(ch for ch in cell_ref if ch.isdigit())
    col = 0
    for char in letters.upper():
        col = col * 26 + ord(char) - ord("A") + 1
    return col, int(digits)


def _safe_table_name(name: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in name)


def _format_known_columns(ws, headers: list[str], start_row: int, end_row: int) -> None:
    for index, header in enumerate(headers, start=1):
        lowered = header.casefold()
        if (
            "% " in header
            or header.startswith("%")
            or "маржа" in lowered
            or "доля" in lowered
        ):
            _format_range(ws, start_row, end_row, index, index, "pct")
        elif any(
            marker in lowered
            for marker in (
                "выруч",
                "сумма",
                "расход",
                "доход",
                "приб",
                "себестоим",
                "логист",
                "комисс",
                "спп",
                "ндс",
                "усн",
            )
        ):
            _format_range(ws, start_row, end_row, index, index, "rub")


def _format_range(
    ws, start_row: int, end_row: int, start_col: int, end_col: int, fmt: str
) -> None:
    if end_row < start_row:
        return
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            _set_number_format(cell, fmt)


def _set_number_format(cell, fmt: str) -> None:
    if fmt == "rub":
        cell.number_format = '#,##0" руб."'
    elif fmt == "pct":
        cell.number_format = "0.0%"
    elif fmt == "int":
        cell.number_format = "#,##0"


def _add_line_chart(
    ws, title: str, anchor: str, min_row: int, max_row: int, min_col: int, max_col: int
) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Руб."
    chart.x_axis.title = "Месяц"
    data = Reference(
        ws, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row
    )
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 17
    ws.add_chart(chart, anchor)


def _autofit(ws, *, min_width: int, max_width: int) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        values = [
            len(str(ws.cell(row, col).value or ""))
            for row in range(1, min(ws.max_row, 200) + 1)
        ]
        ws.column_dimensions[letter].width = max(
            min_width, min(max(values, default=0) + 2, max_width)
        )
    for row in range(1, min(ws.max_row, 80) + 1):
        ws.row_dimensions[row].height = 22


def validate_dashboard_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        required = {
            "Дашборд",
            "Динамика",
            "Расходы WB",
            "Возвраты",
            "Упущенные продажи",
            "Сверка с 1С ОПиУ",
            "Инструкция",
        }
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise AssertionError(f"Missing sheets: {sorted(missing)}")
        if workbook["Дашборд"]["A1"].value != "Дашборд юнит-экономики WB/1C":
            raise AssertionError("Dashboard title was not written")
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
