#!/usr/bin/env python3
"""Verify a month-close report against a local standard OSV without raw output."""

from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from wb_unit_economics.web.database import (  # noqa: E402
    make_engine,
    make_session_factory,
)
from wb_unit_economics.web.models import MonthCloseControlReport  # noqa: E402
from wb_unit_economics.web.reports.builders import (  # noqa: E402
    canonical_payload_sha256,
)
from wb_unit_economics.web.reports.excel import (  # noqa: E402
    MONTH_CLOSE_SHEETS,
    write_scenario_excel,
)
from wb_unit_economics.web.settings import WebSettings  # noqa: E402

AMOUNT_FIELDS = (
    "openingDebit",
    "openingCredit",
    "debitTurnover",
    "creditTurnover",
    "closingDebit",
    "closingCredit",
)
ACCOUNT_CODE = re.compile(r"^\s*(\d+(?:\.\d+)*)\b")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report-id", required=True)
    parser.add_argument("--reference-workbook", type=Path, required=True)
    parser.add_argument(
        "--reference-side", choices=("standard", "online"), default="standard"
    )
    parser.add_argument("--excel-output", type=Path, required=True)
    for name in (
        "report-accounts",
        "reference-accounts",
        "common-accounts",
        "exact-accounts",
        "mismatch-accounts",
        "report-only-accounts",
        "reference-only-accounts",
    ):
        parser.add_argument(f"--expected-{name}", type=int)
    return parser.parse_args()


def _accounting_baseline_matches(
    args: argparse.Namespace,
    actual: dict[str, int],
) -> bool:
    return all(
        expected is None or actual[name] == expected
        for name in actual
        if (expected := getattr(args, f"expected_{name}", None)) is not None
    )


def _decimal(value: Any, *, blank_zero: bool = False) -> Decimal | None:
    if value in (None, ""):
        return Decimal("0") if blank_zero else None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _reference_rows(
    path: Path, *, reference_side: str
) -> dict[str, tuple[Decimal | None, ...]]:
    resolved = path.resolve()
    allowed_roots = ((ROOT / "data").resolve(), (ROOT / "reports").resolve())
    if (
        not any(resolved.is_relative_to(root) for root in allowed_roots)
        or not resolved.is_file()
    ):
        raise SystemExit("reference workbook is outside local artifacts")
    workbook = load_workbook(resolved, read_only=True, data_only=True)
    if {"Общие счета", "Только 1С"} <= set(workbook.sheetnames):
        result: dict[str, tuple[Decimal | None, ...]] = {}
        common_slice = slice(3, 9) if reference_side == "standard" else slice(9, 15)
        for row in list(
            workbook["Общие счета"].iter_rows(min_row=2, values_only=True)
        ):
            account_code = str(row[0] or "").strip()
            if account_code:
                result[account_code] = tuple(
                    _decimal(value, blank_zero=True) for value in row[common_slice]
                )
        side_sheet = "Только 1С" if reference_side == "standard" else "Только online"
        for row in workbook[side_sheet].iter_rows(min_row=2, values_only=True):
            account_code = str(row[0] or "").strip()
            if account_code:
                result[account_code] = tuple(
                    _decimal(value, blank_zero=True) for value in row[1:7]
                )
        return result
    sheet = workbook[workbook.sheetnames[0]]
    result: dict[str, tuple[Decimal | None, ...]] = {}
    for row in sheet.iter_rows(values_only=True):
        match = ACCOUNT_CODE.match(str(row[0] or ""))
        if match is None:
            continue
        values = tuple(_decimal(value, blank_zero=True) for value in row[1:7])
        if len(values) == len(AMOUNT_FIELDS):
            result[match.group(1)] = values
    return result


def _report_rows(payload: dict[str, Any]) -> dict[str, tuple[Decimal | None, ...]]:
    result: dict[str, tuple[Decimal | None, ...]] = {}
    for row in payload.get("osvRows") or []:
        account_code = str(row.get("accountCode") or "").strip()
        if account_code:
            result[account_code] = tuple(
                _decimal(row.get(key)) for key in AMOUNT_FIELDS
            )
    return result


def main() -> int:
    args = parse_args()
    settings = WebSettings()
    engine = make_engine(
        settings.database_url,
        statement_timeout_ms=settings.postgres_statement_timeout_ms,
    )
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        stored = db.get(MonthCloseControlReport, args.report_id)
        if stored is None:
            raise SystemExit("month-close report not found")
        payload = dict(stored.payload)
        payload_hash = stored.payload_sha256

    calculated_hash = canonical_payload_sha256(payload)
    report_rows = _report_rows(payload)
    reference_rows = _reference_rows(
        args.reference_workbook, reference_side=args.reference_side
    )
    common = sorted(set(report_rows) & set(reference_rows))
    exact = [code for code in common if report_rows[code] == reference_rows[code]]
    mismatched = [code for code in common if report_rows[code] != reference_rows[code]]
    report_only = sorted(set(report_rows) - set(reference_rows))
    reference_only = sorted(set(reference_rows) - set(report_rows))

    output = args.excel_output.resolve()
    allowed_output = (ROOT / "reports").resolve()
    if not output.is_relative_to(allowed_output):
        raise SystemExit("Excel output is outside reports")
    write_scenario_excel(payload, payload_hash, output)
    workbook = load_workbook(output, read_only=True, data_only=True)
    summary = {
        row[0]: row[1]
        for row in workbook["Сводка закрытия"].iter_rows(values_only=True)
        if row[0]
    }
    excel_parity = (
        tuple(workbook.sheetnames) == MONTH_CLOSE_SHEETS
        and summary.get("reportId") == args.report_id
        and summary.get("payloadSha256") == payload_hash
        and summary.get("contractVersion") == payload.get("contractVersion")
    )
    accounting_counts = {
        "report_accounts": len(report_rows),
        "reference_accounts": len(reference_rows),
        "common_accounts": len(common),
        "exact_accounts": len(exact),
        "mismatch_accounts": len(mismatched),
        "report_only_accounts": len(report_only),
        "reference_only_accounts": len(reference_only),
    }
    accounting_baseline_match = _accounting_baseline_matches(
        args, accounting_counts
    )

    print(f"payload_hash_valid={str(calculated_hash == payload_hash).lower()}")
    print(f"excel_parity={str(excel_parity).lower()}")
    print(f"accounting_baseline_match={str(accounting_baseline_match).lower()}")
    for name, value in accounting_counts.items():
        print(f"{name}={value}")
    return (
        0
        if calculated_hash == payload_hash
        and excel_parity
        and accounting_baseline_match
        else 1
    )


if __name__ == "__main__":
    raise SystemExit(main())
