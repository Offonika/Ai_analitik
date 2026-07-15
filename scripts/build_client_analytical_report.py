from __future__ import annotations

import argparse
import csv
import io
import json
import re
import shutil
import subprocess
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from wb_unit_economics.document_exports import render_markdown_docx

DEFAULT_WORKBOOK = Path("reports/shumeyko_wb_excel_mvp.xlsx")
DEFAULT_LOGO = Path("reports/assets/shumeiko-logo.png")
DEFAULT_BASENAME = "Аналитический отчет по юнит-экономике WB"
DEFAULT_BRANDED_BASENAME = (
    "Фирменный аналитический отчет Шумейко и Партнеры по юнит-экономике WB"
)


@dataclass(frozen=True)
class ClientAnalyticalReportArtifacts:
    markdown_path: Path
    docx_path: Path
    pdf_path: Path | None
    pdf_status: str
    pdf_message: str


def main() -> int:
    args = _parse_args()
    basename = args.basename
    if args.branded and basename == DEFAULT_BASENAME:
        basename = DEFAULT_BRANDED_BASENAME
    artifacts = build_client_analytical_report(
        workbook_path=args.workbook,
        output_dir=args.output_dir,
        basename=basename,
        logo_path=args.logo,
        branded=args.branded,
    )
    print(f"Markdown: {artifacts.markdown_path}")
    print(f"DOCX: {artifacts.docx_path}")
    if artifacts.pdf_path is not None:
        print(f"PDF: {artifacts.pdf_path}")
    else:
        print(f"PDF: {artifacts.pdf_status} - {artifacts.pdf_message}")
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build client analytical Markdown and DOCX report from Excel MVP."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=Path("reports"))
    parser.add_argument("--basename", default=DEFAULT_BASENAME)
    parser.add_argument("--logo", type=Path, default=DEFAULT_LOGO)
    parser.add_argument("--branded", action="store_true")
    return parser.parse_args()


def build_client_analytical_report(
    *,
    workbook_path: Path = DEFAULT_WORKBOOK,
    output_dir: Path = Path("reports"),
    basename: str = DEFAULT_BASENAME,
    logo_path: Path = DEFAULT_LOGO,
    branded: bool = False,
) -> ClientAnalyticalReportArtifacts:
    output_dir.mkdir(parents=True, exist_ok=True)
    markdown_path = output_dir / f"{basename}.md"
    docx_path = output_dir / f"{basename}.docx"

    data = _collect_report_data(workbook_path)
    markdown = _build_markdown(data)
    markdown_path.write_text(markdown, encoding="utf-8")
    resolved_logo_path = logo_path if logo_path.exists() else None
    render_markdown_docx(
        markdown,
        docx_path,
        logo_path=resolved_logo_path,
        branded=branded,
        cover_subtitle=str(data["readme"].get("Период отчета") or ""),
    )
    pdf_path, pdf_status, pdf_message = _build_pdf_from_docx(docx_path)
    return ClientAnalyticalReportArtifacts(
        markdown_path=markdown_path,
        docx_path=docx_path,
        pdf_path=pdf_path,
        pdf_status=pdf_status,
        pdf_message=pdf_message,
    )


def _build_pdf_from_docx(docx_path: Path) -> tuple[Path | None, str, str]:
    converter = shutil.which("libreoffice") or shutil.which("soffice")
    if converter is None:
        return (
            None,
            "unavailable",
            "PDF-конвертер LibreOffice/soffice не установлен на сервере.",
        )
    output_dir = docx_path.parent
    expected_pdf = docx_path.with_suffix(".pdf")
    try:
        result = subprocess.run(
            [
                converter,
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(docx_path),
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=120,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        return None, "failed", f"Не удалось запустить PDF-конвертер: {exc}"
    if result.returncode != 0 or not expected_pdf.exists():
        return (
            None,
            "failed",
            "DOCX сформирован, но PDF-конвертация завершилась ошибкой.",
        )
    return expected_pdf, "ok", "PDF сформирован."


def _collect_report_data(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        readme = _read_readme(workbook["README"])
        unit_rows = _sheet_dicts(workbook["Юнит экономика"])
        return_rows = _sheet_dicts(workbook["Возвраты"])
        lost_sales_rows = (
            _read_lost_sales_rows(workbook["Упущенные продажи"])
            if "Упущенные продажи" in workbook.sheetnames
            else []
        )
        monthly_rows = _read_monthly_rows(workbook["Динамика"])
        expense_rows = _read_expense_structure(workbook["Расходы WB"])
        opiu_rows, opiu_monthly_rows = (
            _read_onec_opiu_reconciliation(workbook["Сверка с 1С ОПиУ"])
            if "Сверка с 1С ОПиУ" in workbook.sheetnames
            else ([], [])
        )
        stock_history = _latest_stock_history_summary()
        onec_stock = _latest_onec_stock_summary()
        return {
            "readme": readme,
            "unit_rows": unit_rows,
            "return_rows": return_rows,
            "monthly_rows": monthly_rows,
            "expense_rows": expense_rows,
            "opiu_rows": opiu_rows,
            "opiu_monthly_rows": opiu_monthly_rows,
            "stock_history": stock_history,
            "onec_stock": onec_stock,
            "metrics": _metrics(unit_rows),
            "status_counts": _status_counts(unit_rows),
            "by_cabinet": _by_cabinet(unit_rows),
            "top_loss": _top_loss_groups(unit_rows),
            "top_returns": _top_return_amounts(return_rows),
            "high_returns": _high_return_rate(return_rows),
            "loss_classes": _loss_class_counts(unit_rows),
            "lost_sales_rows": lost_sales_rows,
            "lost_sales_top": _top_lost_sales(lost_sales_rows),
            "lost_sales_metrics": _lost_sales_metrics(lost_sales_rows),
        }
    finally:
        workbook.close()


def _read_readme(ws: Any) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            result[str(row[0])] = row[1]
    return result


def _headers(ws: Any, row_number: int = 1) -> dict[str, int]:
    return {
        str(cell.value): cell.column - 1
        for cell in ws[row_number]
        if cell.value not in (None, "")
    }


def _sheet_dicts(ws: Any) -> list[dict[str, Any]]:
    headers = _headers(ws)
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        rows.append(
            {
                header: row[index] if index < len(row) else None
                for header, index in headers.items()
            }
        )
    return rows


def _read_monthly_rows(ws: Any) -> list[dict[str, Any]]:
    headers = _headers(ws, row_number=2)
    rows: list[dict[str, Any]] = []
    row_number = 3
    max_col = max(headers.values(), default=0) + 1
    while True:
        values = [ws.cell(row_number, column).value for column in range(1, max_col + 1)]
        if not values[0] or not _is_report_month_label(values[0]):
            break
        rows.append(
            {
                header: values[index] if index < len(values) else None
                for header, index in headers.items()
            }
        )
        row_number += 1
    return rows


def _is_report_month_label(value: Any) -> bool:
    text = str(value or "").strip()
    return text in {
        "Март 2026",
        "Апрель 2026",
        "Май 2026",
        "Июнь 2026 (неполный месяц)",
    }


def _read_onec_opiu_reconciliation(
    ws: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    header_row = None
    for row_number in range(1, min(ws.max_row, 40) + 1):
        if ws.cell(row_number, 1).value == "Показатель":
            header_row = row_number
            break
    if header_row is None:
        return [], []
    headers = _headers(ws, row_number=header_row)
    rows: list[dict[str, Any]] = []
    row_number = header_row + 1
    while row_number <= ws.max_row:
        first = ws.cell(row_number, 1).value
        if not first:
            break
        if first in {
            "Главная сверка по месяцам: количество, себестоимость 1С и расходы МП",
            "Главная сверка по месяцам: себестоимость 1С и расходы МП",
            "Помесячная сверка себестоимости и расходов МП",
            "Месяц",
        }:
            break
        rows.append(
            {
                header: ws.cell(row_number, index + 1).value
                for header, index in headers.items()
            }
        )
        row_number += 1

    monthly_rows: list[dict[str, Any]] = []
    monthly_header_row = None
    for row_number in range(1, min(ws.max_row, 40) + 1):
        if ws.cell(row_number, 1).value == "Месяц":
            monthly_header_row = row_number
            break
    if monthly_header_row:
        monthly_headers = _headers(ws, row_number=monthly_header_row)
        row_number = monthly_header_row + 1
        while row_number <= ws.max_row:
            first = ws.cell(row_number, 1).value
            if not first:
                break
            if first != "Итого" and not _is_report_month_label(first):
                break
            monthly_rows.append(
                {
                    header: ws.cell(row_number, index + 1).value
                    for header, index in monthly_headers.items()
                }
            )
            row_number += 1

    return rows, monthly_rows


def _read_expense_structure(ws: Any) -> list[dict[str, Any]]:
    header_row = None
    for row_number in range(1, min(ws.max_row, 30) + 1):
        if ws.cell(row_number, 1).value == "Статья":
            header_row = row_number
            break
    if header_row is None:
        return []
    headers = _headers(ws, row_number=header_row)
    rows: list[dict[str, Any]] = []
    row_number = header_row + 1
    while True:
        first = ws.cell(row_number, 1).value
        if not first:
            break
        if first == "Детализация по строкам":
            break
        rows.append(
            {
                header: ws.cell(row_number, index + 1).value
                for header, index in headers.items()
            }
        )
        row_number += 1
    return rows


def _read_lost_sales_rows(ws: Any) -> list[dict[str, Any]]:
    header_row = None
    headers: dict[str, int] = {}
    header_scan_limit = min(ws.max_row, 80)
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=header_scan_limit, values_only=True),
        start=1,
    ):
        if row and row[0] == "Кабинет WB":
            header_row = row_number
            headers = {
                str(value): index
                for index, value in enumerate(row)
                if value not in (None, "")
            }
            break
    if header_row is None:
        return []
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        first = row[0] if row else None
        if first in (None, ""):
            break
        rows.append(
            {
                header: row[index] if index < len(row) else None
                for header, index in headers.items()
            }
        )
    return rows


def _latest_stock_history_summary() -> dict[str, Any]:
    export_dir = _latest_dir(Path("data/wb_stock_history_daily"))
    if export_dir is None or not (export_dir / "manifest.json").exists():
        return {"status": "not_loaded"}
    manifest = json.loads((export_dir / "manifest.json").read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    for item in manifest.get("results", []):
        if not isinstance(item, dict) or item.get("status") != "ok":
            continue
        output_file = item.get("output_file")
        if not output_file:
            continue
        zip_path = export_dir / str(output_file)
        csv_rows, date_columns = _count_stock_history_zip(zip_path)
        rows.append(
            {
                "cabinet": item.get("seller_account_id"),
                "csv_rows": csv_rows,
                "date_columns": date_columns,
                "file": output_file,
            }
        )
    return {
        "status": "ok" if rows else "empty",
        "path": str(export_dir),
        "period_start": manifest.get("period_start"),
        "period_end": manifest.get("period_end"),
        "stock_type": manifest.get("stock_type"),
        "rows": rows,
    }


def _latest_onec_stock_summary() -> dict[str, Any]:
    candidates = []
    root = Path("data/onec_samples")
    if not root.exists():
        return {"status": "not_loaded"}
    for manifest_path in root.glob("*/manifest.json"):
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except ValueError:
            continue
        for item in manifest.get("results", []):
            if (
                isinstance(item, dict)
                and item.get("sample_id") == "stock_by_warehouse"
                and item.get("ok") is True
            ):
                candidates.append(
                    (manifest_path.parent.stat().st_mtime, manifest_path, item)
                )
    if not candidates:
        return {"status": "not_loaded"}
    _mtime, manifest_path, item = max(candidates, key=lambda value: value[0])
    return {
        "status": "ok",
        "path": str(manifest_path.parent),
        "row_count": item.get("row_count"),
        "file": item.get("output_file"),
    }


def _latest_dir(root: Path) -> Path | None:
    if not root.exists():
        return None
    dirs = [path for path in root.iterdir() if path.is_dir()]
    if not dirs:
        return None
    return max(dirs, key=lambda path: path.stat().st_mtime)


def _count_stock_history_zip(path: Path) -> tuple[int, int]:
    total_rows = 0
    date_columns = 0
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            data = zf.read(name)
            text = data.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if rows:
                total_rows += max(0, len(rows) - 1)
                if date_columns == 0:
                    date_columns = sum(
                        1 for header in rows[0] if _looks_like_date(header)
                    )
    return total_rows, date_columns


def _looks_like_date(value: object) -> bool:
    return bool(re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", str(value or "")))


def _metrics(rows: list[dict[str, Any]]) -> dict[str, Decimal]:
    result = {
        "revenue_before_spp": Decimal("0"),
        "spp": Decimal("0"),
        "revenue": Decimal("0"),
        "sales": Decimal("0"),
        "returns": Decimal("0"),
        "net_qty": Decimal("0"),
        "profit": Decimal("0"),
        "profit_before": Decimal("0"),
    }
    for row in rows:
        result["revenue_before_spp"] += _num(
            row.get("Выручка до СПП") or row.get("Выручка с НДС")
        )
        result["spp"] += _num(row.get("СПП"))
        result["revenue"] += _row_revenue(row)
        result["sales"] += _num(row.get("Продажи, шт"))
        result["returns"] += _num(row.get("Возвраты, шт"))
        result["net_qty"] += _num(row.get("Чистое кол-во"))
        result["profit"] += _row_profit(row)
        result["profit_before"] += _num(
            row.get("Маржинальный доход WB до налогов")
            or row.get("Прибыль до налогов")
        )
    return result


def _status_counts(rows: list[dict[str, Any]]) -> Counter[str]:
    return Counter(str(row.get("Статус данных")) for row in rows)


def _by_cabinet(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Decimal]] = defaultdict(
        lambda: defaultdict(Decimal)
    )
    for row in rows:
        cabinet = str(row.get("Кабинет WB"))
        grouped[cabinet]["revenue"] += _row_revenue(row)
        grouped[cabinet]["sales"] += _num(row.get("Продажи, шт"))
        grouped[cabinet]["returns"] += _num(row.get("Возвраты, шт"))
        grouped[cabinet]["profit"] += _row_profit(row)
    return [
        {
            "cabinet": cabinet,
            **values,
            "return_rate": _safe_div(values["returns"], values["sales"]),
            "margin": _safe_div(values["profit"], values["revenue"]),
        }
        for cabinet, values in sorted(grouped.items())
    ]


def _top_loss_groups(
    rows: list[dict[str, Any]],
    limit: int = 10,
) -> list[dict[str, Any]]:
    grouped = _group_product_rows(rows)
    losses = [item for item in grouped if item["profit"] < 0]
    losses.sort(
        key=lambda item: (
            item["profit"],
            item["profit_per_unit"] or Decimal("0"),
            -(item["return_rate"] or Decimal("0")),
        )
    )
    return losses[:limit]


def _top_return_amounts(
    rows: list[dict[str, Any]],
    limit: int = 7,
) -> list[dict[str, Any]]:
    grouped = _group_return_rows(rows)
    grouped.sort(key=lambda item: item["return_amount"], reverse=True)
    return grouped[:limit]


def _high_return_rate(
    rows: list[dict[str, Any]],
    limit: int = 7,
) -> list[dict[str, Any]]:
    grouped = [
        item
        for item in _group_return_rows(rows)
        if item["sales"] >= 10 and item["return_rate"] is not None
    ]
    grouped.sort(
        key=lambda item: (
            -(item["return_rate"] or Decimal("0")),
            -item["returns"],
        )
    )
    return grouped[:limit]


def _loss_class_counts(rows: list[dict[str, Any]]) -> Counter[str]:
    return Counter(
        item["loss_class"]
        for item in _group_product_rows(rows)
        if item["profit"] < 0 and item["loss_class"]
    )


def _top_lost_sales(
    rows: list[dict[str, Any]],
    limit: int = 10,
) -> list[dict[str, Any]]:
    candidates = [
        row
        for row in rows
        if _num(row.get("Дней без остатка WB")) > 0
        or _num(row.get("Дней критического остатка")) > 0
    ]
    candidates.sort(
        key=lambda row: (
            _num(row.get("Потенциально упущенная прибыль")),
            _num(row.get("Потенциально упущенная выручка")),
            _num(row.get("Потенциально упущено, шт")),
        ),
        reverse=True,
    )
    return candidates[:limit]


def _lost_sales_metrics(rows: list[dict[str, Any]]) -> dict[str, Decimal]:
    return {
        "sku_with_zero_stock": sum(
            (Decimal("1") for row in rows if _num(row.get("Дней без остатка WB")) > 0),
            Decimal("0"),
        ),
        "lost_units": sum(
            (_num(row.get("Потенциально упущено, шт")) for row in rows),
            Decimal("0"),
        ),
        "lost_revenue": sum(
            (_num(row.get("Потенциально упущенная выручка")) for row in rows),
            Decimal("0"),
        ),
        "lost_profit": sum(
            (_num(row.get("Потенциально упущенная прибыль")) for row in rows),
            Decimal("0"),
        ),
    }


def _group_product_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (
            row.get("Товар"),
            row.get("Артикул 1С"),
            row.get("Баркод"),
            row.get("Кабинет WB"),
        )
        if key not in grouped:
            grouped[key] = {
                "product": row.get("Товар"),
                "article": row.get("Артикул 1С"),
                "barcode": row.get("Баркод"),
                "cabinet": row.get("Кабинет WB"),
                "sales": Decimal("0"),
                "returns": Decimal("0"),
                "net_qty": Decimal("0"),
                "revenue": Decimal("0"),
                "profit": Decimal("0"),
                "return_amount": Decimal("0"),
                "loss_reason": "",
                "loss_class": "",
            }
        bucket = grouped[key]
        bucket["sales"] += _num(row.get("Продажи, шт"))
        bucket["returns"] += _num(row.get("Возвраты, шт"))
        bucket["net_qty"] += _num(row.get("Чистое кол-во"))
        bucket["revenue"] += _row_revenue(row)
        bucket["profit"] += _row_profit(row)
        bucket["return_amount"] += _num(row.get("Сумма возвратов"))
        bucket["loss_reason"] = row.get("Главная причина") or bucket["loss_reason"]
        bucket["loss_class"] = _loss_class(row) or bucket["loss_class"]
    for bucket in grouped.values():
        bucket["return_rate"] = _safe_div(bucket["returns"], bucket["sales"])
        bucket["profit_per_unit"] = _safe_div(bucket["profit"], bucket["net_qty"])
    return list(grouped.values())


def _group_return_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        key = (
            row.get("Товар"),
            row.get("Артикул 1С"),
            row.get("Баркод"),
            row.get("Кабинет WB"),
        )
        if key not in grouped:
            grouped[key] = {
                "product": row.get("Товар"),
                "article": row.get("Артикул 1С"),
                "barcode": row.get("Баркод"),
                "cabinet": row.get("Кабинет WB"),
                "sales": Decimal("0"),
                "returns": Decimal("0"),
                "return_amount": Decimal("0"),
                "profit": Decimal("0"),
            }
        bucket = grouped[key]
        bucket["sales"] += _num(row.get("Продажи, шт"))
        bucket["returns"] += _num(row.get("Возвраты, шт"))
        bucket["return_amount"] += _num(row.get("Сумма возвратов"))
        bucket["profit"] += _row_profit(row)
    for bucket in grouped.values():
        bucket["return_rate"] = _safe_div(bucket["returns"], bucket["sales"])
    return list(grouped.values())


def _loss_class(row: dict[str, Any]) -> str:
    status = str(row.get("Статус данных") or "")
    if status and status != "ОК":
        return "Нужна проверка данных"
    profit = _row_profit(row)
    if profit >= 0:
        return ""
    candidates = {
        "Высокая закупка / недостаточная наценка": _num(
            row.get("Себестоимость 1С")
        ),
        "Возвраты + логистика": (
            abs(_num(row.get("Логистика WB"))) + _num(row.get("Сумма возвратов"))
        ),
        "Прочие расходы": (
            abs(_num(row.get("Комиссия WB")))
            + abs(_num(row.get("Хранение WB")))
            + abs(_num(row.get("Продвижение WB")))
            + abs(_num(row.get("Штрафы/доплаты WB")))
            + abs(_num(row.get("Эквайринг WB")))
            + abs(
                _num(
                    row.get("Налог с выручки/НДФЛ")
                    or row.get("Налог с выручки")
                    or row.get("УСН 1%")
                )
            )
        ),
    }
    return max(candidates.items(), key=lambda item: item[1])[0]


def _build_markdown(data: dict[str, Any]) -> str:
    readme = data["readme"]
    metrics = data["metrics"]
    report_period = str(readme.get("Период отчета") or readme.get("Период") or "")
    source_coverage = str(readme.get("Покрытие источников") or "").strip()
    analysis_note = str(readme.get("Период анализа") or "").strip()
    if not analysis_note:
        analysis_note = f"Период анализа: {report_period or 'не указан'}"
    readiness_status = _client_readiness_status(
        data,
        report_period=report_period,
        source_coverage=source_coverage,
        analysis_note=analysis_note,
    )
    return_rate = _safe_div(metrics["returns"], metrics["sales"])
    margin = _safe_div(metrics["profit"], metrics["revenue"])
    spp_rate = _safe_div(metrics["spp"], metrics["revenue_before_spp"])
    rows = [
        "# Аналитический отчет по юнит-экономике WB",
        "",
        f"Период отчета: {report_period}",
        f"Покрытие источников: {source_coverage or 'не зафиксировано'}",
        analysis_note,
        f"Статус периода: {readme.get('Статус периода')}",
        f"Статус готовности: {readiness_status}",
        f"Дата расчета: {readme.get('Дата расчета')}",
        "",
        "## 1. Ключевые выводы",
        "",
        (
            "Отчет построен за выбранный период отчета. Покрытие источников "
            "показывается отдельно, чтобы сверка с 1С/ОПиУ не смешивала "
            "разные окна и не скрывала неполные загрузки."
        ),
        "",
        (
            "Управленческая прибыль WB за период составляет "
            f"{_money(metrics['profit'])} при выручке после СПП "
            f"{_money(metrics['revenue'])}; маржа WB без НДС "
            f"составляет {_percent(margin)}."
        ),
        "",
        (
            f"Возвраты: {_qty(metrics['returns'])} шт. из "
            f"{_qty(metrics['sales'])} шт. продаж, доля возвратов составляет "
            f"{_percent(return_rate)}."
        ),
        "",
        "## 2. Основные KPI",
        "",
        _markdown_table(
            ["Показатель", "Значение"],
            [
                ["Выручка до СПП", _money(metrics["revenue_before_spp"])],
                ["СПП", _money(metrics["spp"])],
                ["% СПП", _percent(spp_rate)],
                ["Выручка после СПП", _money(metrics["revenue"])],
                ["Продажи", f"{_qty(metrics['sales'])} шт."],
                ["Возвраты", f"{_qty(metrics['returns'])} шт."],
                ["Чистое количество", f"{_qty(metrics['net_qty'])} шт."],
                ["Доля возвратов", _percent(return_rate)],
                ["Маржинальный доход WB до налогов", _money(metrics["profit_before"])],
                ["Управленческая прибыль WB", _money(metrics["profit"])],
                ["Маржа WB без НДС", _percent(margin)],
            ],
        ),
        "",
        "## 3. Сверка с 1С и ОПиУ",
        "",
        (
            "Себестоимость в WB-расчете берется из регистра продаж 1С. "
            "Основная помесячная сверка сравнивает количество и себестоимость "
            "по одинаковой выборке РВБ-документов и по дате документа 1С. "
            "ОПиУ используется только для сверки расходов РВБ."
        ),
        "",
        _markdown_table(
            ["Показатель", "WB-витрина", "1С/ОПиУ", "Дельта", "Комментарий"],
            [
                [
                    row.get("Показатель"),
                    _format_reconciliation_cell(
                        row.get("WB-витрина"), row.get("Показатель")
                    ),
                    _format_reconciliation_cell(
                        row.get("1С/ОПиУ"), row.get("Показатель")
                    ),
                    _format_reconciliation_cell(
                        row.get("Дельта"), row.get("Показатель")
                    ),
                    row.get("Комментарий"),
                ]
                for row in data["opiu_rows"]
            ],
        ),
        "",
        "### Главная сверка по месяцам: количество, себестоимость 1С и расходы МП",
        "",
        _markdown_table(
            [
                "Месяц",
                "WB количество",
                "1С количество",
                "Дельта количества",
                "Себестоимость 1С в WB-расчете",
                "Себестоимость по валовой прибыли 1С",
                "Дельта себестоимости",
                "WB расходы МП",
                "1С расходы МП",
                "Дельта расходов МП",
                "Комментарий",
            ],
            [
                [
                    row.get("Месяц"),
                    _qty(_num(row.get("WB количество"))),
                    _qty(_num(row.get("1С количество"))),
                    _qty(_num(row.get("Дельта количества"))),
                    _money(
                        _num(
                            row.get("Себестоимость 1С в WB-расчете")
                            or row.get("WB себестоимость")
                        )
                    ),
                    _money(
                        _num(
                            row.get("Себестоимость по валовой прибыли 1С")
                            or row.get("1С себестоимость")
                        )
                    ),
                    _money(_num(row.get("Дельта себестоимости"))),
                    _money(_num(row.get("WB расходы МП"))),
                    _money(_num(row.get("1С расходы МП"))),
                    _money(_num(row.get("Дельта расходов МП"))),
                    row.get("Комментарий"),
                ]
                for row in data["opiu_monthly_rows"]
            ],
        ),
        "",
        "## 4. Помесячная динамика",
        "",
        _markdown_table(
            [
                "Месяц",
                "Статус",
                "Продажи",
                "Возвраты",
                "% возвратов",
                "Выручка до СПП",
                "СПП",
                "% СПП",
                "Выручка после СПП",
                "Логистика",
                "Расходы WB",
                "Маржинальный доход WB",
                "Маржа WB",
            ],
            [
                [
                    row.get("Месяц"),
                    row.get("Статус"),
                    _qty(_num(row.get("Продажи, шт"))),
                    _qty(_num(row.get("Возвраты, шт"))),
                    _percent(_num_or_none(row.get("% возвратов"))),
                    _money(_num(row.get("Выручка до СПП"))),
                    _money(_num(row.get("СПП"))),
                    _percent(_num_or_none(row.get("% СПП"))),
                    _money(_num(row.get("Выручка после СПП"))),
                    _money(_num(row.get("Логистика"))),
                    _money(_num(row.get("Расходы WB"))),
                    _money(
                        _num(
                            row.get("Прибыль до НДФЛ")
                            or row.get("Управленческая прибыль WB")
                            or row.get("Маржинальный доход WB после налогов")
                            or row.get("Управленческая прибыль")
                            or row.get("Прибыль после налогов")
                        )
                    ),
                    _percent(
                        _num_or_none(
                            row.get("Маржинальность до НДФЛ")
                            or row.get("Маржа WB без НДС")
                            or row.get("Маржа WB после налогов")
                            or row.get("Маржа без НДС")
                            or row.get("Маржа после налогов")
                        )
                    ),
                ]
                for row in data["monthly_rows"]
            ],
        ),
        "",
        "## 5. Структура расходов в % от выручки после СПП",
        "",
        _markdown_table(
            [
                "Статья",
                "Сумма",
                "% от выручки",
                "Март",
                "Апрель",
                "Май",
                "Июнь",
                "Апрель к Март",
                "Май к Апрель",
                "Июнь к маю",
            ],
            [
                [
                    row.get("Статья"),
                    _money(_num(row.get("Сумма"))),
                    _percent(_num_or_none(row.get("% от выручки"))),
                    _money(_num(row.get("Март"))),
                    _money(_num(row.get("Апрель"))),
                    _money(_num(row.get("Май"))),
                    _money(_num(row.get("Июнь"))),
                    _money(_num(row.get("Апрель к Март"))),
                    _money(_num(row.get("Май к Апрель"))),
                    _money(_num(row.get("Июнь к Май") or row.get("Июнь к маю"))),
                ]
                for row in data["expense_rows"]
            ],
        ),
        "",
        "## 6. Классификация убыточных товаров",
        "",
        _markdown_table(
            ["Класс убытка", "Количество товарных групп"],
            [
                [label, str(count)]
                for label, count in data["loss_classes"].most_common()
            ],
        ),
        "",
        _markdown_table(
            [
                "Товар",
                "Артикул 1С",
                "Баркод",
                "Кабинет",
                "Продажи",
                "Возвраты",
                "% возвратов",
                "Прибыль",
                "Прибыль/шт",
                "Класс",
            ],
            [
                [
                    row["product"],
                    row["article"],
                    row["barcode"],
                    row["cabinet"],
                    _qty(row["sales"]),
                    _qty(row["returns"]),
                    _percent(row["return_rate"]),
                    _money(row["profit"]),
                    _money(row["profit_per_unit"]),
                    row["loss_class"],
                ]
                for row in data["top_loss"]
            ],
        ),
        "",
        "## 7. Возвраты с баркодами",
        "",
        "Причина возврата не передается текущими источниками.",
        "",
        _markdown_table(
            [
                "Товар",
                "Артикул 1С",
                "Баркод",
                "Возвраты",
                "% возвратов",
                "Сумма возвратов",
                "Прибыль",
            ],
            [
                [
                    row["product"],
                    row["article"],
                    row["barcode"],
                    _qty(row["returns"]),
                    _percent(row["return_rate"]),
                    _money(row["return_amount"]),
                    _money(row["profit"]),
                ]
                for row in data["top_returns"]
            ],
        ),
        "",
        "### Высокий % возвратов",
        "",
        _markdown_table(
            [
                "Товар",
                "Артикул 1С",
                "Баркод",
                "Продажи",
                "Возвраты",
                "% возвратов",
                "Сумма возвратов",
                "Прибыль",
            ],
            [
                [
                    row["product"],
                    row["article"],
                    row["barcode"],
                    _qty(row["sales"]),
                    _qty(row["returns"]),
                    _percent(row["return_rate"]),
                    _money(row["return_amount"]),
                    _money(row["profit"]),
                ]
                for row in data["high_returns"]
            ],
        ),
        "",
        "## 8. Предварительная оценка упущенных продаж",
        "",
        _stock_history_text(data),
        "",
        _lost_sales_summary_text(data),
        "",
        _markdown_table(
            [
                "Товар",
                "Артикул 1С",
                "Баркод",
                "Кабинет",
                "Дней без остатка",
                "Продажи",
                "Упущено, шт",
                "Упущенная выручка",
                "Упущенная прибыль",
                "Вывод",
            ],
            [
                [
                    row.get("Товар"),
                    row.get("Артикул 1С"),
                    row.get("Баркод"),
                    row.get("Кабинет WB"),
                    _qty(_num(row.get("Дней без остатка WB"))),
                    _qty(_num(row.get("Продажи, шт"))),
                    _qty(_num(row.get("Потенциально упущено, шт"))),
                    _money(_num(row.get("Потенциально упущенная выручка"))),
                    _money(_num(row.get("Потенциально упущенная прибыль"))),
                    row.get("Вывод"),
                ]
                for row in data["lost_sales_top"]
            ],
        ),
        "",
        (
            "Важно: оценка использует историю остатков WB stockType=wb и продажи "
            "за период. Для финального управленческого решения ее нужно сверить "
            "с 1С остатками комиссионера и календарем поставок."
        ),
        "",
        "## 9. Рекомендации",
        "",
        "1. Закрыть проблемные статусы маппинга и себестоимости перед решениями.",
        (
            "2. По убыточным SKU разделить задачи между закупкой, "
            "маркетплейс-менеджером и аналитиком данных."
        ),
        (
            "3. По листу `Упущенные продажи` разобрать товары с положительной "
            "упущенной прибылью и днями без остатка WB."
        ),
        "4. Не указывать причины возвратов без подтвержденного источника.",
        "",
        "## 10. Что дальше",
        "",
        (
            "1. Сверить расчет упущенных продаж с 1С остатками комиссионера "
            "и планом поставок."
        ),
        (
            "2. Отдельно показать товары, которые числятся в 1С у комиссионера, "
            "но отсутствуют или недоступны на складах WB."
        ),
        (
            "3. Для топа упущенной прибыли проверить минимальные остатки, "
            "график поставок, цену и маржинальность."
        ),
        (
            "4. После согласования методики добавить отдельный регулярный "
            "контроль out-of-stock в web/BI-витрину."
        ),
    ]
    return "\n".join(str(item) for item in rows)


def _stock_history_text(data: dict[str, Any]) -> str:
    stock_history = data.get("stock_history") or {}
    onec_stock = data.get("onec_stock") or {}
    if stock_history.get("status") != "ok":
        return (
            "Исторический stock-report WB пока не подключен к расчетной витрине; "
            "в отчете оставлен план будущего расчета."
        )
    parts = [
        (
            "Исторический stock-report WB за период "
            f"{stock_history.get('period_start')} - {stock_history.get('period_end')} "
            "уже загружен и использован как источник для предварительной "
            "витрины упущенных продаж."
        )
    ]
    for row in stock_history.get("rows", []):
        parts.append(
            f"{row.get('cabinet')}: {row.get('csv_rows')} строк, "
            f"{row.get('date_columns')} дневных колонок."
        )
    if onec_stock.get("status") == "ok":
        parts.append(
            "Дополнительно загружен 1С-срез остатков по складам: "
            f"{onec_stock.get('row_count')} строк для сверки комиссионерских остатков."
        )
    return " ".join(parts)


def _lost_sales_summary_text(data: dict[str, Any]) -> str:
    metrics = data.get("lost_sales_metrics") or {}
    top_rows = data.get("lost_sales_top") or []
    if not top_rows:
        return (
            "В текущем Excel-листе нет товарных строк с днями без остатка WB, "
            "которые можно оценить по продажам периода."
        )
    return (
        "По предварительной витрине найдено "
        f"{_qty(_num(metrics.get('sku_with_zero_stock')))} товарных групп с днями "
        "нулевого остатка WB. Оценочно: "
        f"{_qty(_num(metrics.get('lost_units')))} шт. потенциально упущено, "
        f"{_money(_num(metrics.get('lost_revenue')))} потенциальной выручки и "
        f"{_money(_num(metrics.get('lost_profit')))} потенциальной прибыли. "
        "Ниже топ товарных групп для ручной проверки."
    )


def _client_readiness_status(
    data: dict[str, Any],
    *,
    report_period: str,
    source_coverage: str,
    analysis_note: str,
) -> str:
    if source_coverage and report_period and source_coverage != report_period:
        return "source_coverage_gap"
    statuses = data["status_counts"]
    if statuses.get("Неполный источник"):
        return "partial_source"
    if "непол" in analysis_note.lower():
        return "partial_period"
    if any(status and status != "ОК" for status in statuses):
        return "needs_review"
    return "ready"


def _markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    if not rows:
        rows = [["" for _ in headers]]
    output = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        output.append("| " + " | ".join(str(value or "") for value in row) + " |")
    return "\n".join(output)


def _num(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def _row_revenue(row: dict[str, Any]) -> Decimal:
    return _num(row.get("Выручка после СПП") or row.get("Выручка с НДС"))


def _row_profit(row: dict[str, Any]) -> Decimal:
    return _num(
        row.get("Прибыль до НДФЛ")
        or row.get("Управленческая прибыль WB")
        or row.get("Маржинальный доход WB после налогов")
        or row.get("Управленческая прибыль")
        or row.get("Прибыль после налогов")
    )


def _num_or_none(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    return Decimal(str(value))


def _safe_div(numerator: Decimal, denominator: Decimal) -> Decimal | None:
    if denominator == 0:
        return None
    return numerator / denominator


def _money(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:,.0f} руб.".replace(",", " ")


def _format_reconciliation_cell(value: Any, label: Any = "") -> str:
    if value in (None, ""):
        return ""
    number = _num_or_none(value)
    if number is None:
        return str(value)
    if "%" in str(label):
        return _percent(number)
    return _money(number)


def _qty(value: Decimal) -> str:
    return f"{value:,.0f}".replace(",", " ")


def _percent(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value * Decimal('100'):.1f}%".replace(".", ",")


if __name__ == "__main__":
    raise SystemExit(main())
