from __future__ import annotations

import argparse
import csv
import re
import subprocess
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from scripts.export_power_bi_marts import export_power_bi_marts
except ModuleNotFoundError:  # pragma: no cover - direct script execution path
    from export_power_bi_marts import export_power_bi_marts

DEFAULT_WORKBOOK = Path("reports/shumeyko_wb_excel_mvp.xlsx")
DEFAULT_OUTPUT_DIR = Path("reports/power_query_client_package")
DEFAULT_DATABASE = "shumeyko_wb_unit_economics"
DEFAULT_SCHEMA = "bi"
DEFAULT_PORT = 55433
DEFAULT_PERIOD = "01.03.2026 - 17.06.2026"
TEMPLATE_NAME = (
    "Шаблон Power Query WB-1C Шумейко и Партнеры "
    "01.03.2026-17.06.2026.xlsx"
)


@dataclass(frozen=True)
class Mart:
    name: str
    title: str
    sheet_name: str
    description: str


MARTS: tuple[Mart, ...] = (
    Mart(
        "kpi_period",
        "KPI периода",
        "Дашборд",
        "Главные показатели периода: выручка, СПП, продажи, возвраты, маржа.",
    ),
    Mart(
        "monthly_dynamics",
        "Помесячная динамика",
        "Динамика",
        "Март, апрель, май и неполный июнь с динамикой месяц к месяцу.",
    ),
    Mart(
        "expenses",
        "Структура расходов",
        "Расходы WB",
        "Расходы WB и себестоимость с долями от выручки и динамикой.",
    ),
    Mart(
        "unit_economics",
        "Юнит-экономика товаров",
        "Юнит экономика",
        "Товарная витрина с продажами, возвратами, расходами и маржой.",
    ),
    Mart(
        "returns",
        "Возвраты",
        "Возвраты",
        "Возвраты по товарам, баркоды, суммы возвратов и влияние на прибыль.",
    ),
    Mart(
        "lost_sales",
        "Упущенные продажи",
        "Упущенные продажи",
        "Предварительная витрина товаров с потенциально упущенной прибылью.",
    ),
    Mart(
        "onec_opiu_reconciliation",
        "Сверка с 1С/ОПиУ",
        "Сверка с 1С ОПиУ",
        "Контроль себестоимости и расходов маркетплейса по месяцам.",
    ),
)


def main() -> int:
    args = _parse_args()
    result = build_power_query_client_package(
        workbook_path=args.workbook,
        output_dir=args.output_dir,
        database=args.database,
        schema=args.schema,
        host=args.host,
        port=args.port,
        user=args.user,
        publish_postgres=args.publish_postgres,
    )
    for path in result:
        print(path)
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a safe client Power Query package for WB/1C marts."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--database", default=DEFAULT_DATABASE)
    parser.add_argument("--schema", default=DEFAULT_SCHEMA)
    parser.add_argument("--host", default="")
    parser.add_argument("--port", type=int, default=DEFAULT_PORT)
    parser.add_argument("--user", default="")
    parser.add_argument(
        "--publish-postgres",
        action="store_true",
        help="Publish the generated mart CSV files into PostgreSQL schema tables.",
    )
    return parser.parse_args()


def build_power_query_client_package(
    *,
    workbook_path: Path,
    output_dir: Path,
    database: str,
    schema: str,
    host: str,
    port: int,
    user: str,
    publish_postgres: bool,
) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    marts_dir = output_dir / "marts_csv"
    pq_dir = output_dir / "power_query_m"
    sql_dir = output_dir / "sql"
    marts_dir.mkdir(parents=True, exist_ok=True)
    pq_dir.mkdir(parents=True, exist_ok=True)
    sql_dir.mkdir(parents=True, exist_ok=True)

    mart_paths = export_power_bi_marts(workbook_path, marts_dir)
    mart_headers = _read_mart_headers(mart_paths)
    template_path = output_dir / TEMPLATE_NAME
    _write_excel_template(
        template_path=template_path,
        workbook_path=workbook_path,
        database=database,
        schema=schema,
        host=host,
        port=port,
    )
    pq_paths = _write_power_query_files(
        pq_dir=pq_dir,
        database=database,
        schema=schema,
        host=host,
        port=port,
        mart_headers=mart_headers,
    )
    ddl_path = sql_dir / "publish_power_query_marts.sql"
    _write_publish_sql_contract(ddl_path, schema=schema)
    readme_path = output_dir / "README.md"
    _write_readme(
        readme_path=readme_path,
        template_path=template_path,
        pq_dir=pq_dir,
        marts_dir=marts_dir,
        database=database,
        schema=schema,
        host=host,
        port=port,
    )

    result = [template_path, readme_path, ddl_path, *pq_paths, *mart_paths]
    if publish_postgres:
        published = _publish_marts_to_postgres(
            mart_paths=mart_paths,
            schema=schema,
            database=database,
            host=host,
            port=port,
            user=user,
            sql_dir=sql_dir,
        )
        result.append(published)
    return result


def _write_excel_template(
    *,
    template_path: Path,
    workbook_path: Path,
    database: str,
    schema: str,
    host: str,
    port: int,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Инструкция"

    palette = {
        "navy": "0B3A53",
        "blue": "2F80C4",
        "light": "EAF3FB",
        "soft": "F7FAFC",
        "line": "A9C9E8",
        "warning": "FFF2CC",
        "white": "FFFFFF",
    }
    styles = _styles(palette)
    _write_instruction_sheet(
        workbook["Инструкция"], styles, workbook_path, host, port, database, schema
    )
    _write_parameters_sheet(
        workbook.create_sheet("Параметры подключения"),
        styles,
        host,
        port,
        database,
        schema,
    )
    _write_marts_sheet(workbook.create_sheet("Витрины"), styles, schema)
    _write_queries_sheet(workbook.create_sheet("M-запросы"), styles)
    _write_limitations_sheet(workbook.create_sheet("Ограничения"), styles)

    workbook.properties.title = "Шаблон Power Query WB-1C"
    workbook.properties.subject = "Юнит-экономика WB, подключение к Postgres BI"
    workbook.properties.creator = "Шумейко и Партнеры / Codex"
    workbook.properties.keywords = "Power Query, Wildberries, 1C, unit economics"
    workbook.save(template_path)


def _styles(palette: dict[str, str]) -> dict[str, object]:
    thin = Side(style="thin", color=palette["line"])
    return {
        "title": Font(bold=True, size=18, color=palette["navy"]),
        "section": Font(bold=True, size=12, color=palette["navy"]),
        "header": Font(bold=True, color=palette["white"]),
        "normal": Font(size=11, color="1F2933"),
        "muted": Font(size=10, color="5B6770"),
        "fill_header": PatternFill("solid", fgColor=palette["navy"]),
        "fill_light": PatternFill("solid", fgColor=palette["light"]),
        "fill_soft": PatternFill("solid", fgColor=palette["soft"]),
        "fill_warning": PatternFill("solid", fgColor=palette["warning"]),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _write_instruction_sheet(
    ws,
    styles: dict[str, object],
    source_workbook: Path,
    host: str,
    port: int,
    database: str,
    schema: str,
) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Шаблон обновления юнит-экономики WB через Power Query"
    ws["A1"].font = styles["title"]
    ws["A3"] = (
        "Назначение: клиент открывает файл в Excel, один раз настраивает "
        "подключение к Postgres-витринам и дальше нажимает Данные -> Обновить все."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"].font = styles["normal"]

    rows = [
        ("1", "Открыть вкладку Данные -> Получить данные -> Из базы данных."),
        (
            "2",
            "Выбрать PostgreSQL database. Если коннектор не установлен, "
            "поставить PostgreSQL/Npgsql connector или использовать ODBC.",
        ),
        (
            "3",
            "Создать параметры Power Query: pServer, pDatabase, pSchema. "
            "Готовые тексты лежат в папке power_query_m.",
        ),
        (
            "4",
            "Для каждой витрины вставить соответствующий M-запрос из папки "
            "power_query_m и загрузить результат на отдельный лист.",
        ),
        (
            "5",
            "При первом обновлении Excel попросит логин/пароль к read-only "
            "пользователю базы. Пароль не хранится в этом шаблоне.",
        ),
        (
            "6",
            "Дальше обновление выполняется одной кнопкой: Данные -> Обновить все.",
        ),
    ]
    _write_table(
        ws,
        start_row=6,
        headers=["Шаг", "Что сделать"],
        rows=rows,
        widths=[12, 100],
        styles=styles,
    )

    connection_rows = [
        ("Сервер", f"{host or '<адрес сервера>'}:{port}"),
        ("База", database),
        ("Схема витрин", schema),
        ("Период текущего отчета", DEFAULT_PERIOD),
        ("Источник шаблона", str(source_workbook)),
    ]
    _write_table(
        ws,
        start_row=15,
        headers=["Параметр", "Значение"],
        rows=connection_rows,
        widths=[30, 90],
        styles=styles,
    )

    ws["A23"] = "Важно"
    ws["A23"].font = styles["section"]
    ws["A24"] = (
        "В файл не встроены пароли, токены WB/1C или raw-данные. "
        "Power Query должен читать только готовые таблицы схемы bi."
    )
    ws["A24"].fill = styles["fill_warning"]
    ws["A24"].alignment = Alignment(wrap_text=True, vertical="top")


def _write_parameters_sheet(
    ws,
    styles: dict[str, object],
    host: str,
    port: int,
    database: str,
    schema: str,
) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Параметры для Power Query"
    ws["A1"].font = styles["title"]
    rows = [
        ("pServer", f"{host or '<адрес сервера>'}:{port}", "Текстовый параметр"),
        ("pDatabase", database, "Текстовый параметр"),
        ("pSchema", schema, "Текстовый параметр"),
        (
            "Учетная запись",
            "read-only пользователь Postgres",
            "Выдается отдельно, пароль в файл не записываем",
        ),
        (
            "Доступ",
            "VPN / белый IP / защищенный туннель",
            "Зависит от выбранного способа публикации базы",
        ),
    ]
    _write_table(
        ws,
        start_row=3,
        headers=["Параметр", "Значение", "Комментарий"],
        rows=rows,
        widths=[28, 42, 70],
        styles=styles,
    )


def _write_marts_sheet(ws, styles: dict[str, object], schema: str) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Состав клиентских витрин"
    ws["A1"].font = styles["title"]
    rows = [
        (
            f"{schema}.{mart.name}",
            mart.title,
            mart.sheet_name,
            mart.description,
        )
        for mart in MARTS
    ]
    _write_table(
        ws,
        start_row=3,
        headers=["Таблица в базе", "Название", "Источник Excel", "Что показывает"],
        rows=rows,
        widths=[34, 30, 28, 78],
        styles=styles,
    )


def _write_queries_sheet(ws, styles: dict[str, object]) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Готовые M-запросы"
    ws["A1"].font = styles["title"]
    rows = [
        ("00_pServer.pq", "Параметр сервера Postgres"),
        ("01_pDatabase.pq", "Параметр имени базы"),
        ("02_pSchema.pq", "Параметр схемы BI-витрин"),
        ("03_fxLoadView.pq", "Общая функция загрузки таблицы из схемы bi"),
        *[(f"{mart.name}.pq", mart.title) for mart in MARTS],
    ]
    _write_table(
        ws,
        start_row=3,
        headers=["Файл", "Назначение"],
        rows=rows,
        widths=[34, 90],
        styles=styles,
    )


def _write_limitations_sheet(ws, styles: dict[str, object]) -> None:
    _setup_sheet(ws)
    ws["A1"] = "Ограничения и безопасность"
    ws["A1"].font = styles["title"]
    rows = [
        (
            "Пароли",
            "Не вшиваются в Excel. Клиент вводит учетку Postgres "
            "при первом обновлении.",
        ),
        (
            "Источники",
            "Power Query читает только клиентские витрины bi.*, а не raw snapshots.",
        ),
        (
            "WB/1C",
            "Excel не пишет в WB, 1С или другие системы. "
            "Обновление только читает базу.",
        ),
        (
            "СПП и причины возвратов",
            "Не выдумываются: если источника нет, показывается ограничение источника.",
        ),
        (
            "Публикация наружу",
            "Для клиента нужен защищенный доступ к базе: VPN, "
            "gateway или отдельный хост.",
        ),
    ]
    _write_table(
        ws,
        start_row=3,
        headers=["Тема", "Правило"],
        rows=rows,
        widths=[28, 100],
        styles=styles,
    )


def _setup_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


def _write_table(
    ws,
    *,
    start_row: int,
    headers: list[str],
    rows: Iterable[tuple[object, ...]],
    widths: list[int],
    styles: dict[str, object],
) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.font = styles["header"]
        cell.fill = styles["fill_header"]
        cell.border = styles["border"]
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = styles["border"]
            cell.font = styles["normal"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx % 2 == 0:
                cell.fill = styles["fill_soft"]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_power_query_files(
    *,
    pq_dir: Path,
    database: str,
    schema: str,
    host: str,
    port: int,
    mart_headers: dict[str, list[str]],
) -> list[Path]:
    server = f"{host or '<адрес сервера>'}:{port}"
    files: list[tuple[str, str]] = [
        (
            "00_pServer.pq",
            _parameter_query(server, "Text"),
        ),
        (
            "01_pDatabase.pq",
            _parameter_query(database, "Text"),
        ),
        (
            "02_pSchema.pq",
            _parameter_query(schema, "Text"),
        ),
        (
            "03_fxLoadView.pq",
            "\n".join(
                [
                    "(viewName as text) as table =>",
                    "let",
                    "    Source = PostgreSQL.Database(",
                    "        pServer,",
                    "        pDatabase,",
                    "        [CreateNavigationProperties = false]",
                    "    ),",
                    "    Data = Source{[Schema = pSchema, Item = viewName]}[Data]",
                    "in",
                    "    Data",
                    "",
                ]
            ),
        ),
    ]
    files.extend(
        (f"{mart.name}.pq", _mart_query(mart, mart_headers.get(mart.name, [])))
        for mart in MARTS
    )

    result = []
    for filename, content in files:
        path = pq_dir / filename
        path.write_text(content, encoding="utf-8")
        result.append(path)
    return result


def _parameter_query(value: str, value_type: str) -> str:
    escaped = value.replace('"', '""')
    return (
        f'"{escaped}" meta '
        f'[IsParameterQuery = true, Type = "{value_type}", '
        "IsParameterQueryRequired = true]\n"
    )


def _mart_query(mart: Mart, headers: list[str]) -> str:
    rename_pairs = ",\n        ".join(
        f'{{"{db_column}", "{_m_text(header)}"}}'
        for db_column, header in zip(
            _db_columns_for_headers(headers), headers, strict=True
        )
    )
    if not rename_pairs:
        return "\n".join(
            [
                "let",
                f'    Data = fxLoadView("{mart.name}")',
                "in",
                "    Data",
                "",
            ]
        )
    return "\n".join(
        [
            "let",
            f'    Raw = fxLoadView("{mart.name}"),',
            "    Data = Table.RenameColumns(",
            "        Raw,",
            "        {",
            f"        {rename_pairs}",
            "        },",
            "        MissingField.Ignore",
            "    )",
            "in",
            "    Data",
            "",
        ]
    )


def _m_text(value: str) -> str:
    return value.replace('"', '""')


def _write_publish_sql_contract(path: Path, *, schema: str) -> None:
    lines = [
        "-- Contract for client-facing Power Query marts.",
        "-- The tables are generated from reports/shumeyko_wb_excel_mvp.xlsx.",
        "-- Do not point client BI tools to raw WB/1C snapshots.",
        "-- Table columns are c001, c002, ...; Power Query renames them back",
        "-- to business-readable Russian headers on load.",
        f"CREATE SCHEMA IF NOT EXISTS {_quote_ident(schema)};",
        "",
        "-- Expected tables:",
        *[f"--   {schema}.{mart.name} -- {mart.title}" for mart in MARTS],
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_readme(
    *,
    readme_path: Path,
    template_path: Path,
    pq_dir: Path,
    marts_dir: Path,
    database: str,
    schema: str,
    host: str,
    port: int,
) -> None:
    server = f"{host or '<адрес сервера>'}:{port}"
    lines = [
        "# Клиентский комплект Power Query",
        "",
        "Комплект нужен, чтобы Excel обновлял юнит-экономику из Postgres-витрин.",
        "",
        "## Что внутри",
        "",
        f"- `{template_path.name}` — Excel-шаблон с инструкцией и составом витрин.",
        f"- `{pq_dir.name}/` — готовые M-запросы Power Query.",
        f"- `{marts_dir.name}/` — CSV-копия витрин как промежуточный fallback.",
        "- `sql/` — контракт BI-слоя.",
        "",
        "## Подключение",
        "",
        f"- Сервер: `{server}`",
        f"- База: `{database}`",
        f"- Схема: `{schema}`",
        "- Учетка: отдельный read-only пользователь Postgres.",
        "",
        "Пароль не сохраняется в файлах комплекта. Клиент вводит его в Excel при "
        "первом обновлении.",
        "",
        "## Важно",
        "",
        "- Power Query читает только готовые витрины `bi.*`.",
        "- Raw snapshots, `.env`, WB-токены и 1С-доступы в комплект не входят.",
        "- Если прямой доступ к базе клиенту не открыт, используйте CSV fallback "
        "или Power BI Gateway/VPN.",
        "",
    ]
    readme_path.write_text("\n".join(lines), encoding="utf-8")


def _publish_marts_to_postgres(
    *,
    mart_paths: list[Path],
    schema: str,
    database: str,
    host: str,
    port: int,
    user: str,
    sql_dir: Path,
) -> Path:
    statements: list[str] = [f"CREATE SCHEMA IF NOT EXISTS {_quote_ident(schema)};"]
    for path in mart_paths:
        table = _table_name_from_path(path)
        headers, rows = _read_csv_preview(path)
        db_columns = _db_columns_for_headers(headers)
        column_types = _infer_column_types(headers, rows)
        statements.append(
            _drop_create_sql(schema, table, db_columns, headers, column_types)
        )
        statements.append(
            _copy_command(
                schema=schema,
                table=table,
                columns=db_columns,
                csv_path=path,
            )
        )
        statements.append(
            "COMMENT ON TABLE "
            f"{_quote_ident(schema)}.{_quote_ident(table)} "
            f"IS {_quote_literal('Client Power Query mart from Excel MVP')};"
        )
    statements.append(_refresh_log_sql(schema))
    sql_path = sql_dir / "publish_power_query_marts.generated.sql"
    sql_path.write_text("\n\n".join(statements) + "\n", encoding="utf-8")
    _run_psql(sql_path=sql_path, database=database, host=host, port=port, user=user)
    return sql_path


def _read_mart_headers(mart_paths: list[Path]) -> dict[str, list[str]]:
    result = {}
    for path in mart_paths:
        table = _table_name_from_path(path)
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.reader(file)
            result[table] = next(reader)
    return result


def _db_columns_for_headers(headers: list[str]) -> list[str]:
    return [f"c{index:03d}" for index in range(1, len(headers) + 1)]


def _table_name_from_path(path: Path) -> str:
    name = path.stem.strip().lower()
    if not re.fullmatch(r"[a-z0-9_]+", name):
        raise ValueError(f"Unsafe mart table name: {path.stem}")
    return name


def _read_csv_preview(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.reader(file)
        headers = next(reader)
        rows = [row for _, row in zip(range(5000), reader, strict=False)]
    return headers, rows


def _infer_column_types(headers: list[str], rows: list[list[str]]) -> list[str]:
    width = max([len(headers), *[len(row) for row in rows]], default=0)
    column_types = []
    for col_idx in range(width):
        values = [
            row[col_idx].strip()
            for row in rows
            if col_idx < len(row) and row[col_idx].strip()
        ]
        header = headers[col_idx] if col_idx < len(headers) else ""
        column_types.append(_column_type(header, values))
    return column_types


def _column_type(header: str, values: list[str]) -> str:
    lowered = header.casefold()
    text_markers = (
        "артикул",
        "баркод",
        "barcode",
        "кабинет",
        "комментар",
        "месяц",
        "неделя",
        "огранич",
        "организац",
        "показатель",
        "причин",
        "статус",
        "статья",
        "товар",
    )
    if any(marker in lowered for marker in text_markers):
        return "text"
    if values and all(_is_decimal(value) for value in values):
        return "numeric"
    if values and all(_is_iso_date(value) for value in values):
        return "date"
    return "text"


def _is_decimal(value: str) -> bool:
    try:
        Decimal(value)
    except InvalidOperation:
        return False
    return True


def _is_iso_date(value: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}(?: 00:00:00)?", value))


def _drop_create_sql(
    schema: str,
    table: str,
    columns: list[str],
    headers: list[str],
    column_types: list[str],
) -> str:
    column_sql = []
    comments = []
    for index, column in enumerate(columns):
        column_type = column_types[index] if index < len(column_types) else "text"
        column_sql.append(f"    {_quote_ident(column)} {column_type}")
        header = headers[index] if index < len(headers) else column
        comments.append(
            "COMMENT ON COLUMN "
            f"{_quote_ident(schema)}.{_quote_ident(table)}.{_quote_ident(column)} "
            f"IS {_quote_literal(header)};"
        )
    return "\n".join(
        [
            f"DROP TABLE IF EXISTS {_quote_ident(schema)}.{_quote_ident(table)};",
            f"CREATE TABLE {_quote_ident(schema)}.{_quote_ident(table)} (",
            ",\n".join(column_sql),
            ");",
            *comments,
        ]
    )


def _copy_command(
    *,
    schema: str,
    table: str,
    columns: list[str],
    csv_path: Path,
) -> str:
    column_sql = ", ".join(_quote_ident(column) for column in columns)
    return (
        "\\copy "
        f"{_quote_ident(schema)}.{_quote_ident(table)} ({column_sql}) "
        f"FROM {_quote_literal(str(csv_path.resolve()))} "
        "WITH (FORMAT csv, HEADER true, NULL '', ENCODING 'UTF8');"
    )


def _refresh_log_sql(schema: str) -> str:
    generated_at = datetime.now(UTC).isoformat()
    rows = ",\n".join(
        "("
        f"{_quote_literal(mart.name)}, "
        f"{_quote_literal(mart.title)}, "
        f"{_quote_literal(DEFAULT_PERIOD)}, "
        f"{_quote_literal(generated_at)}"
        ")"
        for mart in MARTS
    )
    return "\n".join(
        [
            f"DROP TABLE IF EXISTS {_quote_ident(schema)}._mart_refresh_log;",
            f"CREATE TABLE {_quote_ident(schema)}._mart_refresh_log (",
            "    mart_name text NOT NULL,",
            "    mart_title text NOT NULL,",
            "    period text NOT NULL,",
            "    published_at timestamptz NOT NULL",
            ");",
            f"INSERT INTO {_quote_ident(schema)}._mart_refresh_log",
            "(mart_name, mart_title, period, published_at)",
            "VALUES",
            rows,
            ";",
        ]
    )


def _run_psql(
    *,
    sql_path: Path,
    database: str,
    host: str,
    port: int,
    user: str,
) -> None:
    cmd = ["psql", "-v", "ON_ERROR_STOP=1", "-p", str(port), "-d", database]
    if host:
        cmd.extend(["-h", host])
    if user:
        cmd.extend(["-U", user])
    cmd.extend(["-f", str(sql_path)])
    subprocess.run(cmd, check=True)


def _quote_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _quote_literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def validate_template(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        required = {
            "Инструкция",
            "Параметры подключения",
            "Витрины",
            "M-запросы",
            "Ограничения",
        }
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise AssertionError(f"Missing sheets: {sorted(missing)}")
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
