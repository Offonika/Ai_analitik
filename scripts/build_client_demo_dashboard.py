#!/usr/bin/env python3
"""Build a portable client demo dashboard from the visible Excel MVP sheets."""

# ruff: noqa: E501

from __future__ import annotations

import argparse
import base64
import json
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_WORKBOOK = Path("reports/shumeyko_wb_excel_mvp.xlsx")
DEFAULT_LOGO = Path("reports/assets/shumeiko-logo.png")
DEFAULT_OUTPUT = Path(
    "reports/Демо-дашборд юнит-экономики WB Шумейко и Партнеры "
    "01.03.2026-17.06.2026.html"
)
PERIOD_START = date(2026, 3, 1)
PERIOD_END = date(2026, 6, 17)
REPORT_PERIOD_LABEL = "01.03.2026 - 17.06.2026"
REPORT_PERIOD_TEXT = "март, апрель, май, июнь; июнь неполный, по 17.06.2026"
RETURN_REASON_LIMITATION = "Причина возврата не передается текущими источниками"
OPIU_MONTH_LABELS = {
    "Март 2026",
    "Апрель 2026",
    "Май 2026",
    "Июнь 2026 (неполный месяц)",
    "Итого",
}

VISIBLE_SHEETS = {
    "Дашборд",
    "Динамика",
    "Расходы WB",
    "Возвраты",
    "Упущенные продажи",
    "Юнит экономика",
    "Сверка с 1С ОПиУ",
}


@dataclass(frozen=True)
class SheetRows:
    headers: list[str]
    rows: list[dict[str, Any]]


def _num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, int | float):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    cleaned = (
        text.replace("\xa0", "")
        .replace(" ", "")
        .replace("%", "")
        .replace("руб.", "")
        .replace(",", ".")
    )
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _maybe_num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    parsed = _num(value)
    return parsed


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_div(numerator: float, denominator: float) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator


def _row_revenue(row: dict[str, Any]) -> float:
    return _num(row.get("Выручка после СПП") or row.get("Выручка с НДС"))


def _row_revenue_before_spp(row: dict[str, Any]) -> float:
    return _num(row.get("Выручка до СПП") or row.get("Выручка с НДС"))


def _row_profit(row: dict[str, Any]) -> float:
    return _num(
        row.get("Прибыль до НДФЛ")
        or row.get("Управленческая прибыль WB")
        or row.get("Маржинальный доход WB после налогов")
        or row.get("Управленческая прибыль")
        or row.get("Прибыль после налогов")
    )


def _round(value: Any, digits: int = 2) -> float | None:
    number = _maybe_num(value)
    if number is None:
        return None
    return round(number, digits)


def _clean_for_json(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 4)
    if isinstance(value, dict):
        return {key: _clean_for_json(val) for key, val in value.items()}
    if isinstance(value, list):
        return [_clean_for_json(item) for item in value]
    return value


def _row_month(value: Any) -> str | None:
    if isinstance(value, datetime | date):
        row_date = value.date() if isinstance(value, datetime) else value
    else:
        text = _text(value)
        if not text:
            return None
        try:
            row_date = datetime.fromisoformat(text[:10]).date()
        except ValueError:
            return None
    if row_date < PERIOD_START or row_date > PERIOD_END:
        return None
    row_date = row_date + timedelta(days=3)
    if row_date.month == 3:
        return "Март 2026"
    if row_date.month == 4:
        return "Апрель 2026"
    if row_date.month == 5:
        return "Май 2026"
    if row_date.month == 6:
        return "Июнь 2026 (неполный месяц)"
    return None


def _read_table(
    workbook: Any,
    sheet_name: str,
    *,
    header_row: int = 1,
    min_data_row: int | None = None,
) -> SheetRows:
    sheet = workbook[sheet_name]
    headers = [_text(cell.value) for cell in sheet[header_row]]
    data_row = min_data_row or header_row + 1
    rows: list[dict[str, Any]] = []
    for excel_row in sheet.iter_rows(min_row=data_row, values_only=True):
        if not any(value is not None and value != "" for value in excel_row):
            continue
        item = {}
        for index, header in enumerate(headers):
            if not header or index >= len(excel_row):
                continue
            item[header] = excel_row[index]
        if item:
            rows.append(item)
    return SheetRows(headers=headers, rows=rows)


def _find_header_row(workbook: Any, sheet_name: str, first_header: str) -> int:
    sheet = workbook[sheet_name]
    for row in range(1, min(sheet.max_row, 80) + 1):
        values = [_text(cell.value) for cell in sheet[row]]
        if first_header in values:
            return row
    raise ValueError(f"Не найден заголовок {first_header!r} на листе {sheet_name!r}")


def _read_dashboard_kpis(workbook: Any) -> dict[str, float]:
    sheet = workbook["Дашборд"]
    kpis: dict[str, float] = {}
    for row in range(1, min(sheet.max_row, 40) + 1):
        label = _text(sheet.cell(row=row, column=1).value)
        if not label:
            continue
        value = sheet.cell(row=row, column=2).value
        if value is not None:
            kpis[label] = _num(value)
    return kpis


def _read_monthly_from_sheet(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in _read_table(workbook, "Динамика", header_row=2).rows:
        month = _text(item.get("Месяц"))
        if not month:
            continue
        if not _is_report_month_label(month):
            break
        rows.append(
            {
                "month": month,
                "status": _text(item.get("Статус")),
                "sales": _round(item.get("Продажи, шт"), 0),
                "returns": _round(item.get("Возвраты, шт"), 0),
                "return_rate": _round(item.get("% возвратов"), 4),
                "revenue_before_spp": _round(item.get("Выручка до СПП")),
                "spp": _round(item.get("СПП")),
                "spp_rate": _round(item.get("% СПП"), 4),
                "revenue": _round(item.get("Выручка после СПП") or item.get("Выручка")),
                "logistics": _round(item.get("Логистика")),
                "wb_expenses": _round(item.get("Расходы WB")),
                "profit": _round(
                    item.get("Прибыль до НДФЛ")
                    or item.get("Управленческая прибыль WB")
                    or item.get("Маржинальный доход WB после налогов")
                    or item.get("Управленческая прибыль")
                    or item.get("Прибыль после налогов")
                ),
                "margin": _round(
                    item.get("Маржинальность до НДФЛ")
                    or item.get("Маржа WB без НДС")
                    or item.get("Маржа WB после налогов")
                    or item.get("Маржа без НДС")
                    or item.get("Маржа после налогов"),
                    4,
                ),
            }
        )
    return rows


def _is_report_month_label(value: Any) -> bool:
    text = _text(value)
    return text in OPIU_MONTH_LABELS - {"Итого"}


def _is_opiu_month_label(value: Any) -> bool:
    return _text(value) in OPIU_MONTH_LABELS


def _read_expenses_from_sheet(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in _read_table(workbook, "Расходы WB", header_row=2).rows:
        label = _text(row.get("Статья"))
        if not label or label == "Детализация по строкам":
            break
        rows.append(
            {
                "expense": label,
                "amount": _round(row.get("Сумма")),
                "share": _round(row.get("% от выручки"), 4),
                "march": _round(row.get("Март")),
                "april": _round(row.get("Апрель")),
                "may": _round(row.get("Май")),
                "june": _round(row.get("Июнь")),
                "april_delta": _round(row.get("Апрель к Март")),
                "may_delta": _round(row.get("Май к Апрель")),
                "june_delta": _round(row.get("Июнь к Май") or row.get("Июнь к маю")),
            }
        )
    return rows


def _read_lost_sales(workbook: Any) -> list[dict[str, Any]]:
    header_row = _find_header_row(workbook, "Упущенные продажи", "Товар")
    data = _read_table(workbook, "Упущенные продажи", header_row=header_row)
    rows: list[dict[str, Any]] = []
    for row in data.rows:
        product = _text(row.get("Товар"))
        if not product:
            continue
        rows.append(
            {
                "product": product,
                "article_1c": _text(row.get("Артикул 1С")),
                "barcode": _text(row.get("Баркод")),
                "cabinet": _text(row.get("Кабинет WB")),
                "zero_stock_days": _round(row.get("Дней без остатка WB"), 0),
                "onec_stock": _round(row.get("Остаток 1С на складах, шт"), 0),
                "onec_warehouses": _text(row.get("Склады 1С с остатком")),
                "sales": _round(row.get("Продажи, шт"), 0),
                "lost_units": _round(row.get("Потенциально упущено, шт")),
                "lost_revenue": _round(
                    row.get("Упущенная выручка")
                    or row.get("Потенциально упущенная выручка")
                ),
                "lost_profit": _round(
                    row.get(
                        "Оценка недополученного маржинального дохода "
                        "до налогов"
                    )
                    or row.get("Упущенная прибыль")
                    or row.get("Потенциально упущенная прибыль")
                ),
                "note": _text(row.get("Ограничение") or row.get("Вывод")),
            }
        )
    return rows


def _read_returns(workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in _read_table(workbook, "Возвраты", header_row=1).rows:
        product = _text(row.get("Товар"))
        if not product:
            continue
        rows.append(
            {
                "week": _text(row.get("Неделя")),
                "cabinet": _text(row.get("Кабинет WB")),
                "product": product,
                "nm_id": _text(row.get("nmId WB")),
                "article_wb": _text(row.get("Артикул WB")),
                "article_1c": _text(row.get("Артикул 1С")),
                "barcode": _text(row.get("Баркод")),
                "sales": _round(row.get("Продажи, шт"), 0),
                "returns": _round(row.get("Возвраты, шт"), 0),
                "return_rate": _round(row.get("% возвратов"), 4),
                "return_amount": _round(row.get("Сумма возвратов")),
                "profit": _round(
                    row.get("Прибыль до НДФЛ")
                    or row.get("Управленческая прибыль WB")
                    or row.get("Маржинальный доход WB после налогов")
                    or row.get("Управленческая прибыль")
                    or row.get("Прибыль после налогов")
                ),
                "status": _text(row.get("Статус данных")),
                "driver": _text(row.get("Главная причина")),
                "return_reason": RETURN_REASON_LIMITATION,
            }
        )
    return rows


def _read_onec_opiu_reconciliation(
    workbook: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sheet_name = "Сверка с 1С ОПиУ"
    header_row = _find_header_row(workbook, sheet_name, "Показатель")
    rows = []
    for row in _read_table(workbook, sheet_name, header_row=header_row).rows:
        indicator = _text(row.get("Показатель"))
        if not indicator:
            break
        if indicator in {
            "Главная сверка по месяцам: количество, себестоимость 1С и расходы МП",
            "Главная сверка по месяцам: себестоимость 1С и расходы МП",
            "Помесячная сверка себестоимости и расходов МП",
        }:
            break
        rows.append(
            {
                "indicator": indicator,
                "wb": row.get("WB-витрина"),
                "onec": row.get("1С/ОПиУ"),
                "delta": row.get("Дельта"),
                "comment": _text(row.get("Комментарий")),
            }
        )

    monthly_rows: list[dict[str, Any]] = []
    try:
        monthly_header_row = _find_header_row(
            workbook,
            sheet_name,
            "Месяц",
        )
    except ValueError:
        return rows, monthly_rows
    for row in _read_table(workbook, sheet_name, header_row=monthly_header_row).rows:
        month = _text(row.get("Месяц"))
        if not month:
            break
        if not _is_opiu_month_label(month):
            break
        monthly_rows.append(
            {
                "month": month,
                "wb_quantity": _round(row.get("WB количество"), 0),
                "onec_quantity": _round(row.get("1С количество"), 0),
                "quantity_delta": _round(row.get("Дельта количества"), 0),
                "wb_cogs": _round(
                    row.get("Себестоимость 1С в WB-расчете")
                    if row.get("Себестоимость 1С в WB-расчете") is not None
                    else row.get("WB себестоимость")
                ),
                "onec_cogs": _round(
                    row.get("Себестоимость по валовой прибыли 1С")
                    if row.get("Себестоимость по валовой прибыли 1С") is not None
                    else row.get("1С себестоимость")
                ),
                "cogs_delta": _round(row.get("Дельта себестоимости")),
                "wb_mp_expenses": _round(row.get("WB расходы МП")),
                "onec_mp_expenses": _round(row.get("1С расходы МП")),
                "mp_expenses_delta": _round(row.get("Дельта расходов МП")),
                "comment": _text(row.get("Комментарий")),
            }
        )
    return rows, monthly_rows


def _read_unit_rows(workbook: Any) -> list[dict[str, Any]]:
    table = _read_table(workbook, "Юнит экономика", header_row=1)
    rows: list[dict[str, Any]] = []
    for row in table.rows:
        if not _text(row.get("Товар")):
            continue
        if _row_month(row.get("Неделя")) is None:
            continue
        rows.append(row)
    return rows


def _article_key(row: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        _text(row.get("Кабинет WB")),
        _text(row.get("Товар")),
        _text(row.get("Артикул 1С")),
        _text(row.get("Баркод")),
        _text(row.get("nmId WB")),
    )


def _loss_driver(values: dict[str, float], status: str) -> tuple[str, str]:
    if status and status != "ОК":
        return "Нужна проверка данных", "Нужна проверка данных"
    return_rate = _safe_div(values["returns"], values["sales"]) or 0.0
    factors = {
        "Высокая себестоимость": values["cost"],
        "Высокая логистика WB": values["logistics"],
        "Высокая комиссия WB": values["commission"],
        "Высокое хранение WB": values["storage"],
        "WB продвижение": values["promotion"],
        "Штрафы/удержания WB": values["penalties"],
        "Эквайринг WB": values["acquiring"],
        "Налоги": values["tax"],
    }
    if return_rate >= 0.18:
        factors["Возвраты + логистика"] = values["return_amount"] + values["logistics"]
    driver = max(factors.items(), key=lambda item: item[1])[0]
    if driver == "Высокая себестоимость":
        return "Высокая закупка / недостаточная наценка", driver
    if driver == "Возвраты + логистика":
        return "Возвраты + логистика", driver
    return "Прочие расходы", driver


def _monthly_from_unit_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    monthly: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in rows:
        month = _row_month(row.get("Неделя"))
        if month is None:
            continue
        bucket = monthly[month]
        bucket["sales"] += _num(row.get("Продажи, шт"))
        bucket["returns"] += _num(row.get("Возвраты, шт"))
        bucket["revenue_before_spp"] += _row_revenue_before_spp(row)
        bucket["spp"] += _num(row.get("СПП"))
        bucket["revenue"] += _row_revenue(row)
        bucket["logistics"] += _num(row.get("Логистика WB"))
        bucket["wb_expenses"] += sum(
            _num(row.get(column))
            for column in (
                "Комиссия WB",
                "Логистика WB",
                "Хранение WB",
                "Приемка WB",
                "Продвижение WB",
                "Штрафы/доплаты WB",
                "Эквайринг WB",
            )
        )
        bucket["profit"] += _row_profit(row)
    ordered = [
        "Март 2026",
        "Апрель 2026",
        "Май 2026",
        "Июнь 2026 (неполный месяц)",
    ]
    result: list[dict[str, Any]] = []
    for month in ordered:
        bucket = monthly.get(month, {})
        sales = bucket.get("sales", 0.0)
        revenue = bucket.get("revenue", 0.0)
        revenue_before_spp = bucket.get("revenue_before_spp", revenue)
        spp = bucket.get("spp", 0.0)
        profit = bucket.get("profit", 0.0)
        result.append(
            {
                "month": month,
                "status": "неполный месяц" if "Июнь" in month else "полный месяц",
                "sales": round(sales),
                "returns": round(bucket.get("returns", 0.0)),
                "return_rate": _safe_div(bucket.get("returns", 0.0), sales),
                "revenue_before_spp": round(revenue_before_spp, 2),
                "spp": round(spp, 2),
                "spp_rate": _safe_div(spp, revenue_before_spp),
                "revenue": round(revenue, 2),
                "logistics": round(bucket.get("logistics", 0.0), 2),
                "wb_expenses": round(bucket.get("wb_expenses", 0.0), 2),
                "profit": round(profit, 2),
                "margin": _safe_div(profit, revenue),
            }
        )
    return result


def _expenses_from_unit_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    labels = [
        ("Себестоимость 1С", ("Себестоимость 1С",)),
        ("Комиссия WB", ("Комиссия WB",)),
        ("Логистика WB", ("Логистика WB",)),
        ("Хранение WB", ("Хранение WB",)),
        ("Приемка WB", ("Приемка WB",)),
        ("WB Продвижение", ("Продвижение WB",)),
        ("Штрафы/доплаты WB", ("Штрафы/доплаты WB",)),
        ("Эквайринг WB", ("Эквайринг WB",)),
        (
            "Налог с выручки/НДФЛ",
            ("Налог с выручки/НДФЛ", "Налог с выручки", "УСН 1%"),
        ),
    ]
    revenue = sum(_row_revenue(row) for row in rows)
    by_month = defaultdict(lambda: defaultdict(float))
    totals = defaultdict(float)
    for row in rows:
        month = _row_month(row.get("Неделя"))
        if month is None:
            continue
        for label, columns in labels:
            value = _num(next((row.get(column) for column in columns if row.get(column)), 0))
            totals[label] += value
            by_month[label][month] += value
    result = []
    for label, _column in labels:
        march = by_month[label].get("Март 2026", 0.0)
        april = by_month[label].get("Апрель 2026", 0.0)
        may = by_month[label].get("Май 2026", 0.0)
        june = by_month[label].get("Июнь 2026 (неполный месяц)", 0.0)
        result.append(
            {
                "expense": label,
                "amount": round(totals[label], 2),
                "share": _safe_div(totals[label], revenue),
                "march": round(march, 2),
                "april": round(april, 2),
                "may": round(may, 2),
                "june": round(june, 2),
                "april_delta": round(april - march, 2),
                "may_delta": round(may - april, 2),
                "june_delta": round(june - may, 2),
            }
        )
    return result


def _top_losses(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = _article_key(row)
        values = groups.setdefault(
            key,
            {
                "cabinet": key[0],
                "product": key[1],
                "article_1c": key[2],
                "barcode": key[3],
                "nm_id": key[4],
                "status_counts": Counter(),
                "sales": 0.0,
                "returns": 0.0,
                "revenue": 0.0,
                "return_amount": 0.0,
                "profit": 0.0,
                "cost": 0.0,
                "commission": 0.0,
                "logistics": 0.0,
                "storage": 0.0,
                "promotion": 0.0,
                "penalties": 0.0,
                "acquiring": 0.0,
                "tax": 0.0,
            },
        )
        values["status_counts"][_text(row.get("Статус данных")) or "Не указан"] += 1
        values["sales"] += _num(row.get("Продажи, шт"))
        values["returns"] += _num(row.get("Возвраты, шт"))
        values["revenue"] += _row_revenue(row)
        values["return_amount"] += _num(row.get("Сумма возвратов"))
        values["profit"] += _row_profit(row)
        values["cost"] += _num(row.get("Себестоимость 1С"))
        values["commission"] += _num(row.get("Комиссия WB"))
        values["logistics"] += _num(row.get("Логистика WB"))
        values["storage"] += _num(row.get("Хранение WB"))
        values["promotion"] += _num(row.get("Продвижение WB"))
        values["penalties"] += _num(row.get("Штрафы/доплаты WB"))
        values["acquiring"] += _num(row.get("Эквайринг WB"))
        tax_method = _text(row.get("Налоговый метод") or row.get("Налоговый режим/ставка"))
        pnl_vat_mode = _text(row.get("Режим P&L НДС"))
        if not pnl_vat_mode and "ОСНО" in tax_method:
            pnl_vat_mode = "without_vat_for_osno"
        if pnl_vat_mode != "without_vat_for_osno":
            values["tax"] += _num(
                row.get("НДС к уплате") or row.get("НДС") or row.get("НДС 5%")
            )
        values["tax"] += _num(
            row.get("Налог с выручки/НДФЛ")
            or row.get("Налог с выручки")
            or row.get("УСН 1%")
        )
    result = []
    for values in groups.values():
        if values["profit"] >= 0:
            continue
        status = values["status_counts"].most_common(1)[0][0]
        loss_class, driver = _loss_driver(values, status)
        sales = values["sales"]
        net_qty = sales - values["returns"]
        result.append(
            {
                "cabinet": values["cabinet"],
                "product": values["product"],
                "article_1c": values["article_1c"],
                "barcode": values["barcode"],
                "sales": round(sales),
                "returns": round(values["returns"]),
                "return_rate": _safe_div(values["returns"], sales),
                "revenue": round(values["revenue"], 2),
                "profit": round(values["profit"], 2),
                "unit_profit": _safe_div(values["profit"], net_qty),
                "loss_class": loss_class,
                "driver": driver,
                "status": status,
            }
        )
    return sorted(
        result,
        key=lambda row: (
            row["profit"],
            row["unit_profit"] if row["unit_profit"] is not None else 0,
            -(row["return_rate"] or 0),
        ),
    )


def _aggregate_kpis(
    rows: list[dict[str, Any]],
    lost_sales: list[dict[str, Any]],
    dashboard_kpis: dict[str, float] | None = None,
) -> dict[str, Any]:
    revenue_before_spp = sum(_row_revenue_before_spp(row) for row in rows)
    spp = sum(_num(row.get("СПП")) for row in rows)
    revenue = sum(_row_revenue(row) for row in rows)
    sales = sum(_num(row.get("Продажи, шт")) for row in rows)
    returns = sum(_num(row.get("Возвраты, шт")) for row in rows)
    profit = sum(_row_profit(row) for row in rows)
    loss_rows = sum(1 for row in rows if _row_profit(row) < 0)
    ok_rows = sum(1 for row in rows if _text(row.get("Статус данных")) == "ОК")
    if dashboard_kpis:
        revenue_before_spp = dashboard_kpis.get("Выручка до СПП", revenue_before_spp)
        spp = dashboard_kpis.get("СПП", spp)
        revenue = dashboard_kpis.get("Выручка после СПП", revenue)
        sales = dashboard_kpis.get("Продажи, шт", sales)
        returns = dashboard_kpis.get("Возвраты, шт", returns)
        profit = dashboard_kpis.get(
            "Прибыль до НДФЛ",
            dashboard_kpis.get(
                "Управленческая прибыль WB",
                dashboard_kpis.get(
                    "Маржинальный доход WB после налогов",
                    dashboard_kpis.get(
                        "Управленческая прибыль",
                        dashboard_kpis.get("Прибыль после налогов", profit),
                    ),
                ),
            ),
        )
    lost_profit_rows = [
        row for row in lost_sales if _num(row.get("lost_profit")) > 0
    ]
    return {
        "revenue_before_spp": round(revenue_before_spp, 2),
        "spp": round(spp, 2),
        "spp_rate": _safe_div(spp, revenue_before_spp),
        "revenue": round(revenue, 2),
        "profit": round(profit, 2),
        "margin": _safe_div(profit, revenue),
        "sales": round(sales),
        "returns": round(returns),
        "net_quantity": round(sales - returns),
        "return_rate": _safe_div(returns, sales),
        "loss_rows": loss_rows,
        "ok_share": _safe_div(ok_rows, len(rows)),
        "lost_products": len(lost_profit_rows),
        "lost_revenue": round(sum(_num(row.get("lost_revenue")) for row in lost_profit_rows), 2),
        "lost_profit": round(sum(_num(row.get("lost_profit")) for row in lost_profit_rows), 2),
        "lost_units": round(sum(_num(row.get("lost_units")) for row in lost_profit_rows), 2),
    }


def _status_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counter = Counter(_text(row.get("Статус данных")) or "Не указан" for row in rows)
    total = sum(counter.values())
    return [
        {"status": status, "rows": count, "share": _safe_div(count, total)}
        for status, count in counter.most_common()
    ]


def _filter_by_cabinet(rows: list[dict[str, Any]], cabinet: str) -> list[dict[str, Any]]:
    return [row for row in rows if _text(row.get("cabinet") or row.get("Кабинет WB")) == cabinet]


def _build_view(
    *,
    name: str,
    unit_rows: list[dict[str, Any]],
    returns: list[dict[str, Any]],
    lost_sales: list[dict[str, Any]],
    monthly_rows: list[dict[str, Any]],
    expenses: list[dict[str, Any]],
    reconciliation: list[dict[str, Any]],
    reconciliation_monthly: list[dict[str, Any]],
    dashboard_kpis: dict[str, float] | None = None,
) -> dict[str, Any]:
    if name != "Все кабинеты":
        returns = _filter_by_cabinet(returns, name)
        lost_sales = _filter_by_cabinet(lost_sales, name)
    loss_rows = _top_losses(unit_rows)
    top_returns = sorted(
        returns,
        key=lambda row: (_num(row["return_amount"]), _num(row["returns"])),
        reverse=True,
    )[:12]
    high_returns = sorted(
        [row for row in returns if _num(row["sales"]) >= 5],
        key=lambda row: (_num(row["return_rate"]), _num(row["return_amount"])),
        reverse=True,
    )[:12]
    top_lost_sales = sorted(
        lost_sales,
        key=lambda row: (_num(row["lost_profit"]), _num(row["lost_revenue"])),
        reverse=True,
    )[:12]
    return {
        "kpis": _aggregate_kpis(unit_rows, lost_sales, dashboard_kpis),
        "monthly": monthly_rows,
        "expenses": expenses,
        "top_losses": loss_rows[:12],
        "top_returns": top_returns,
        "high_returns": high_returns,
        "lost_sales": top_lost_sales,
        "statuses": _status_rows(unit_rows),
        "reconciliation": reconciliation,
        "reconciliation_monthly": reconciliation_monthly,
    }


def collect_dashboard_data(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing = VISIBLE_SHEETS.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"В Excel отсутствуют листы: {', '.join(sorted(missing))}")
    unit_rows = _read_unit_rows(workbook)
    returns = _read_returns(workbook)
    lost_sales = _read_lost_sales(workbook)
    dashboard_kpis = _read_dashboard_kpis(workbook)
    monthly_all = _read_monthly_from_sheet(workbook)
    expenses_all = _read_expenses_from_sheet(workbook)
    reconciliation_rows, reconciliation_monthly_rows = _read_onec_opiu_reconciliation(
        workbook
    )
    cabinets = sorted({_text(row.get("Кабинет WB")) for row in unit_rows if row.get("Кабинет WB")})
    views = {
        "Все кабинеты": _build_view(
            name="Все кабинеты",
            unit_rows=unit_rows,
            returns=returns,
            lost_sales=lost_sales,
            monthly_rows=monthly_all,
            expenses=expenses_all,
            reconciliation=reconciliation_rows,
            reconciliation_monthly=reconciliation_monthly_rows,
            dashboard_kpis=dashboard_kpis,
        )
    }
    for cabinet in cabinets:
        cabinet_unit_rows = [
            row for row in unit_rows if _text(row.get("Кабинет WB")) == cabinet
        ]
        views[cabinet] = _build_view(
            name=cabinet,
            unit_rows=cabinet_unit_rows,
            returns=returns,
            lost_sales=lost_sales,
            monthly_rows=_monthly_from_unit_rows(cabinet_unit_rows),
            expenses=_expenses_from_unit_rows(cabinet_unit_rows),
            reconciliation=reconciliation_rows,
            reconciliation_monthly=reconciliation_monthly_rows,
        )
    return _clean_for_json(
        {
            "meta": {
                "title": "Демо-дашборд юнит-экономики WB",
                "partner": "Шумейко и Партнеры",
                "period": REPORT_PERIOD_LABEL,
                "period_text": REPORT_PERIOD_TEXT,
                "generated_at": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "source": workbook_path.name,
                "limitation_return_reason": RETURN_REASON_LIMITATION,
            },
            "cabinets": ["Все кабинеты", *cabinets],
            "views": views,
        }
    )


def _logo_data_uri(path: Path | None) -> str | None:
    if path is None or not path.exists():
        return None
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def _render_html(payload: dict[str, Any], logo_uri: str | None) -> str:
    data_json = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    logo_html = (
        f'<img src="{escape(logo_uri)}" alt="Шумейко и Партнеры">'
        if logo_uri
        else '<div class="logo-fallback">Шумейко<br>и Партнеры</div>'
    )
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(payload["meta"]["title"])} · {escape(payload["meta"]["period"])}</title>
  <style>
    :root {{
      --ink: #0c3348;
      --ink-soft: #42606e;
      --line: #b9d2e8;
      --bg: #f4f7f9;
      --paper: #ffffff;
      --muted: #eef5fb;
      --accent: #2e7dbf;
      --accent-dark: #12527c;
      --profit: #2c7a57;
      --risk: #b44e3b;
      --warn: #c28b26;
      --shadow: 0 18px 40px rgba(12, 51, 72, 0.10);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--ink);
      background: var(--bg);
      font-family: Arial, Helvetica, sans-serif;
      line-height: 1.45;
    }}
    .shell {{ max-width: 1280px; margin: 0 auto; padding: 24px; }}
    .hero {{
      min-height: 92vh;
      display: grid;
      grid-template-rows: auto 1fr auto;
      gap: 26px;
      padding: 28px;
      background: var(--paper);
      border: 1px solid var(--line);
      box-shadow: var(--shadow);
    }}
    .brand-row {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 18px;
      border-bottom: 3px solid var(--accent);
      padding-bottom: 18px;
    }}
    .brand-row img {{ width: 178px; height: auto; display: block; }}
    .logo-fallback {{
      font-family: Georgia, serif;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 1px;
    }}
    .source-note {{ color: var(--ink-soft); font-size: 14px; text-align: right; }}
    .hero-main {{
      display: grid;
      align-content: center;
      gap: 22px;
    }}
    h1 {{
      margin: 0;
      font-family: Georgia, "Times New Roman", serif;
      font-size: clamp(34px, 5vw, 68px);
      line-height: 1.05;
      letter-spacing: 0;
      max-width: 1060px;
    }}
    .subtitle {{ margin: 0; color: var(--ink-soft); font-size: 20px; }}
    .controls {{
      display: flex;
      align-items: center;
      flex-wrap: wrap;
      gap: 12px;
      padding: 14px;
      background: var(--muted);
      border: 1px solid var(--line);
    }}
    .controls label {{ font-weight: 700; }}
    select, button {{
      min-height: 40px;
      border: 1px solid var(--line);
      background: var(--paper);
      color: var(--ink);
      padding: 8px 12px;
      font: inherit;
    }}
    button {{
      cursor: pointer;
      font-weight: 700;
    }}
    button.active {{
      background: var(--accent);
      color: #fff;
      border-color: var(--accent);
    }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 12px;
    }}
    .kpi {{
      min-height: 112px;
      padding: 16px;
      border: 1px solid var(--line);
      background: var(--paper);
      display: grid;
      align-content: space-between;
      gap: 8px;
    }}
    .kpi .label {{ color: var(--ink-soft); font-size: 13px; }}
    .kpi .value {{ font-weight: 800; font-size: 26px; overflow-wrap: anywhere; }}
    .kpi .hint {{ color: var(--ink-soft); font-size: 12px; }}
    .section {{
      margin-top: 22px;
      padding: 24px;
      background: var(--paper);
      border: 1px solid var(--line);
      box-shadow: 0 8px 20px rgba(12, 51, 72, 0.06);
    }}
    .section.is-hidden {{ display: none; }}
    .section-head {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 18px;
      border-bottom: 2px solid var(--accent);
      padding-bottom: 10px;
      margin-bottom: 18px;
    }}
    h2 {{
      margin: 0;
      font-family: Georgia, "Times New Roman", serif;
      font-size: 30px;
      line-height: 1.15;
      letter-spacing: 0;
    }}
    .caption {{ margin: 6px 0 0; color: var(--ink-soft); }}
    .grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }}
    .month-grid {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; }}
    .month-card {{ border: 1px solid var(--line); padding: 16px; background: #fbfdff; }}
    .month-card h3 {{ margin: 0 0 8px; font-size: 20px; }}
    .metric-row {{
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 12px;
      padding: 7px 0;
      border-bottom: 1px solid #e4eef7;
    }}
    .metric-row:last-child {{ border-bottom: 0; }}
    .bars {{ display: grid; gap: 10px; }}
    .bar-row {{
      display: grid;
      grid-template-columns: minmax(140px, 1.3fr) minmax(140px, 2fr) auto;
      align-items: center;
      gap: 12px;
    }}
    .bar-track {{ height: 12px; background: #e4eef7; overflow: hidden; }}
    .bar-fill {{ height: 100%; background: var(--accent); }}
    .bar-fill.risk {{ background: var(--risk); }}
    .bar-fill.profit {{ background: var(--profit); }}
    .table-wrap {{ overflow-x: auto; border: 1px solid var(--line); }}
    table {{ width: 100%; border-collapse: collapse; min-width: 780px; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid #e4eef7; text-align: left; vertical-align: top; }}
    th {{ background: var(--muted); color: var(--ink); font-size: 13px; }}
    td.num, th.num {{ text-align: right; white-space: nowrap; }}
    .pill {{
      display: inline-block;
      padding: 4px 8px;
      background: var(--muted);
      border: 1px solid var(--line);
      font-size: 12px;
      font-weight: 700;
      color: var(--accent-dark);
    }}
    .negative {{ color: var(--risk); font-weight: 700; }}
    .positive {{ color: var(--profit); font-weight: 700; }}
    .method-list {{ margin: 0; padding-left: 18px; color: var(--ink-soft); }}
    @media (max-width: 920px) {{
      .shell {{ padding: 12px; }}
      .hero {{ min-height: auto; padding: 18px; }}
      .brand-row, .section-head {{ display: grid; }}
      .source-note {{ text-align: left; }}
      .kpi-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .grid-2, .month-grid {{ grid-template-columns: 1fr; }}
      .bar-row {{ grid-template-columns: 1fr; gap: 6px; }}
    }}
    @media (max-width: 560px) {{
      .kpi-grid {{ grid-template-columns: 1fr; }}
      .controls {{ align-items: stretch; }}
      select, button {{ width: 100%; }}
      h1 {{ font-size: 34px; }}
      h2 {{ font-size: 25px; }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <header class="hero">
      <div class="brand-row">
        <div>{logo_html}</div>
        <div class="source-note">
          Источник: Excel MVP<br>
          Период: <strong>{escape(payload["meta"]["period"])}</strong>
        </div>
      </div>
      <div class="hero-main">
        <div>
          <h1>Демо-дашборд юнит-экономики WB</h1>
          <p class="subtitle">{escape(payload["meta"]["period_text"])} · сформировано {escape(payload["meta"]["generated_at"])}</p>
        </div>
        <div class="controls">
          <label for="cabinetSelect">Кабинет WB</label>
          <select id="cabinetSelect"></select>
          <button class="toggle active" data-section="dynamics">Динамика</button>
          <button class="toggle active" data-section="expenses">Расходы</button>
          <button class="toggle active" data-section="losses">Убыточность</button>
          <button class="toggle active" data-section="returns">Возвраты</button>
          <button class="toggle active" data-section="lost">Упущенные</button>
          <button class="toggle active" data-section="reconciliation">Сверка 1С</button>
          <button class="toggle active" data-section="quality">Качество</button>
        </div>
        <div id="kpis" class="kpi-grid"></div>
      </div>
      <p class="caption">Демонстрационный HTML построен из клиентских листов Excel: без raw snapshots, внешних API, CDN и backend.</p>
    </header>
    <main>
      <section id="dynamics" class="section">
        <div class="section-head">
          <div>
            <h2>Помесячная динамика</h2>
            <p class="caption">Июнь помечен как неполный месяц, поэтому сравнение с маем нужно читать осторожно.</p>
          </div>
        </div>
        <div id="monthly" class="month-grid"></div>
      </section>
      <section id="expenses" class="section">
        <div class="section-head">
          <div>
            <h2>Структура расходов</h2>
            <p class="caption">Сумма, доля от выручки и динамика по месяцам для ключевых статей.</p>
          </div>
        </div>
        <div id="expensesBars" class="bars"></div>
      </section>
      <section id="losses" class="section">
        <div class="section-head">
          <div>
            <h2>Убыточные товары</h2>
            <p class="caption">Сортировка по убытку после налогов, затем по прибыли на единицу и доле возвратов.</p>
          </div>
        </div>
        <div id="lossTable" class="table-wrap"></div>
      </section>
      <section id="returns" class="section">
        <div class="section-head">
          <div>
            <h2>Возвраты</h2>
            <p class="caption">Топ по сумме возвратов и высокий процент возвратов. Причины возвратов не выдумываются.</p>
          </div>
        </div>
        <div class="grid-2">
          <div>
            <h3>Топ по сумме возвратов</h3>
            <div id="returnsAmount" class="table-wrap"></div>
          </div>
          <div>
            <h3>Высокий процент возвратов</h3>
            <div id="returnsRate" class="table-wrap"></div>
          </div>
        </div>
      </section>
      <section id="lost" class="section">
        <div class="section-head">
          <div>
            <h2>Упущенные продажи</h2>
            <p class="caption">Предварительная оценка по историческому stock-history WB, требует сверки с 1С остатками и поставками.</p>
          </div>
        </div>
        <div id="lostSales" class="table-wrap"></div>
      </section>
      <section id="reconciliation" class="section">
        <div class="section-head">
          <div>
            <h2>Сверка с 1С</h2>
            <p class="caption">Количество и себестоимость сверяются по одинаковой выборке РВБ-документов и по дате документа 1С. ОПиУ используется только для контроля расходов РВБ.</p>
          </div>
        </div>
        <div id="reconciliationTable" class="table-wrap"></div>
        <h3>Главная сверка по месяцам: количество, себестоимость 1С и расходы МП</h3>
        <div id="reconciliationMonthlyTable" class="table-wrap"></div>
      </section>
      <section id="quality" class="section">
        <div class="section-head">
          <div>
            <h2>Качество данных и методика</h2>
            <p class="caption">Что можно показывать клиенту и где нужны подтвержденные источники.</p>
          </div>
        </div>
        <div class="grid-2">
          <div id="statusTable" class="table-wrap"></div>
          <ul class="method-list">
            <li>Период клиентской витрины: {escape(payload["meta"]["period"])}.</li>
            <li>Клиентский дефолт: март-июнь; июнь неполный, по 17.06.2026.</li>
            <li>Общую выручку и общую себестоимость ОПиУ не используем как товарный контроль WB.</li>
            <li>Основной контроль себестоимости берется из сматченных РВБ-документов отчета «Валовая прибыль 1С».</li>
            <li>Выручка после СПП используется как контрольная строка внутри WB-методики.</li>
            <li>Причина возврата не передается текущими источниками.</li>
            <li>Упущенные продажи считаются как управленческий рейтинг, а не финальный прогноз спроса.</li>
          </ul>
        </div>
      </section>
    </main>
  </div>
  <script id="dashboard-data" type="application/json">{data_json}</script>
  <script>
    const payload = JSON.parse(document.getElementById('dashboard-data').textContent);
    const formatNumber = new Intl.NumberFormat('ru-RU', {{ maximumFractionDigits: 0 }});
    const formatMoney = new Intl.NumberFormat('ru-RU', {{ maximumFractionDigits: 0 }});
    const formatSmall = new Intl.NumberFormat('ru-RU', {{ maximumFractionDigits: 1 }});
    const current = {{ view: 'Все кабинеты' }};

    function money(value) {{
      return `${{formatMoney.format(value || 0)}} ₽`;
    }}

    function number(value) {{
      return formatNumber.format(value || 0);
    }}

    function percent(value) {{
      if (value === null || value === undefined) return 'н/д';
      return `${{formatSmall.format((value || 0) * 100)}}%`;
    }}

    function classByValue(value) {{
      if ((value || 0) < 0) return 'negative';
      if ((value || 0) > 0) return 'positive';
      return '';
    }}

    function shortText(value, max = 44) {{
      const text = String(value || '');
      return text.length > max ? `${{text.slice(0, max - 1)}}…` : text;
    }}

    function renderKpis(view) {{
      const k = view.kpis;
      const cards = [
        ['Выручка до СПП', money(k.revenue_before_spp), payload.meta.period],
        ['СПП', money(k.spp), `Доля ${{percent(k.spp_rate)}}`],
        ['Выручка после СПП', money(k.revenue), 'контрольная строка WB-методики'],
        ['Маржинальный доход WB', money(k.profit), `Маржа ${{percent(k.margin)}}`],
        ['Продажи', `${{number(k.sales)}} шт`, `Чистое количество: ${{number(k.net_quantity)}} шт`],
        ['Возвраты', `${{number(k.returns)}} шт`, `Доля возвратов: ${{percent(k.return_rate)}}`],
        ['Убыточные строки', number(k.loss_rows), 'для детального разбора ниже'],
        ['Строки ОК', percent(k.ok_share), 'качество расчетной базы'],
        ['Товары с упущенной прибылью', number(k.lost_products), 'по stock-history WB'],
        ['Потенциально упущенная прибыль', money(k.lost_profit), `${{formatSmall.format(k.lost_units || 0)}} шт`],
      ];
      document.getElementById('kpis').innerHTML = cards.map(([label, value, hint]) => `
        <article class="kpi">
          <div class="label">${{label}}</div>
          <div class="value">${{value}}</div>
          <div class="hint">${{hint}}</div>
        </article>
      `).join('');
    }}

    function renderMonthly(view) {{
      document.getElementById('monthly').innerHTML = view.monthly.map(row => `
        <article class="month-card">
          <h3>${{row.month}}</h3>
          <span class="pill">${{row.status || ''}}</span>
          <div class="metric-row"><span>Выручка до СПП</span><strong>${{money(row.revenue_before_spp)}}</strong></div>
          <div class="metric-row"><span>СПП</span><strong>${{money(row.spp)}} · ${{percent(row.spp_rate)}}</strong></div>
          <div class="metric-row"><span>Выручка после СПП</span><strong>${{money(row.revenue)}}</strong></div>
          <div class="metric-row"><span>Продажи</span><strong>${{number(row.sales)}} шт</strong></div>
          <div class="metric-row"><span>Возвраты</span><strong>${{number(row.returns)}} шт · ${{percent(row.return_rate)}}</strong></div>
          <div class="metric-row"><span>Логистика</span><strong>${{money(row.logistics)}}</strong></div>
          <div class="metric-row"><span>Расходы WB</span><strong>${{money(row.wb_expenses)}}</strong></div>
          <div class="metric-row"><span>Маржинальный доход WB</span><strong class="${{classByValue(row.profit)}}">${{money(row.profit)}}</strong></div>
          <div class="metric-row"><span>Маржа</span><strong>${{percent(row.margin)}}</strong></div>
        </article>
      `).join('');
    }}

    function renderExpenses(view) {{
      const maxShare = Math.max(...view.expenses.map(row => Math.abs(row.share || 0)), 0.01);
      document.getElementById('expensesBars').innerHTML = view.expenses.map(row => {{
        const width = Math.min(100, Math.abs(row.share || 0) / maxShare * 100);
        return `
          <div class="bar-row">
            <strong>${{row.expense}}</strong>
            <div class="bar-track"><div class="bar-fill" style="width:${{width}}%"></div></div>
            <span>${{money(row.amount)}} · ${{percent(row.share)}} · март ${{money(row.march)}} · апрель ${{money(row.april)}} · май ${{money(row.may)}} · июнь ${{money(row.june)}}</span>
          </div>
        `;
      }}).join('');
    }}

    function table(headers, rows) {{
      const head = headers.map(h => `<th class="${{h.num ? 'num' : ''}}">${{h.label}}</th>`).join('');
      const body = rows.map(row => `
        <tr>${{headers.map(h => `<td class="${{h.num ? 'num' : ''}}">${{h.render(row)}}</td>`).join('')}}</tr>
      `).join('');
      return `<table><thead><tr>${{head}}</tr></thead><tbody>${{body || '<tr><td colspan="' + headers.length + '">Нет строк для выбранного фильтра</td></tr>'}}</tbody></table>`;
    }}

    function renderLosses(view) {{
      const headers = [
        {{ label: 'Товар', render: r => `<strong title="${{r.product}}">${{shortText(r.product)}}</strong><br><span class="caption">${{r.article_1c || ''}} · ${{r.barcode || ''}}</span>` }},
        {{ label: 'Кабинет', render: r => shortText(r.cabinet, 24) }},
        {{ label: 'Продажи', num: true, render: r => number(r.sales) }},
        {{ label: 'Возвраты', num: true, render: r => `${{number(r.returns)}} · ${{percent(r.return_rate)}}` }},
        {{ label: 'Маржин. доход', num: true, render: r => `<span class="${{classByValue(r.profit)}}">${{money(r.profit)}}</span>` }},
        {{ label: 'Доход/шт', num: true, render: r => r.unit_profit === null ? 'н/д' : money(r.unit_profit) }},
        {{ label: 'Класс', render: r => `<span class="pill">${{r.loss_class}}</span>` }},
        {{ label: 'Главная причина', render: r => r.driver }},
      ];
      document.getElementById('lossTable').innerHTML = table(headers, view.top_losses);
    }}

    function renderReturns(view) {{
      const amountHeaders = [
        {{ label: 'Товар', render: r => `<strong title="${{r.product}}">${{shortText(r.product)}}</strong><br><span class="caption">${{r.article_1c || ''}} · ${{r.barcode || ''}}</span>` }},
        {{ label: 'Продажи', num: true, render: r => number(r.sales) }},
        {{ label: 'Возвраты', num: true, render: r => number(r.returns) }},
        {{ label: '%', num: true, render: r => percent(r.return_rate) }},
        {{ label: 'Сумма', num: true, render: r => money(r.return_amount) }},
      ];
      const rateHeaders = [
        {{ label: 'Товар', render: r => `<strong title="${{r.product}}">${{shortText(r.product)}}</strong><br><span class="caption">${{r.article_1c || ''}} · ${{r.barcode || ''}}</span>` }},
        {{ label: 'Продажи', num: true, render: r => number(r.sales) }},
        {{ label: 'Возвраты', num: true, render: r => number(r.returns) }},
        {{ label: '% возвратов', num: true, render: r => percent(r.return_rate) }},
        {{ label: 'Маржин. доход', num: true, render: r => `<span class="${{classByValue(r.profit)}}">${{money(r.profit)}}</span>` }},
      ];
      document.getElementById('returnsAmount').innerHTML = table(amountHeaders, view.top_returns);
      document.getElementById('returnsRate').innerHTML = table(rateHeaders, view.high_returns);
    }}

    function renderLostSales(view) {{
      const headers = [
        {{ label: 'Товар', render: r => `<strong title="${{r.product}}">${{shortText(r.product)}}</strong><br><span class="caption">${{r.article_1c || ''}} · ${{r.barcode || ''}}</span>` }},
        {{ label: 'Кабинет', render: r => shortText(r.cabinet, 24) }},
        {{ label: 'Дней без остатка', num: true, render: r => number(r.zero_stock_days) }},
        {{ label: 'Остаток 1С', num: true, render: r => `<span title="${{r.onec_warehouses || ''}}">${{number(r.onec_stock)}}</span>` }},
        {{ label: 'Упущено, шт', num: true, render: r => formatSmall.format(r.lost_units || 0) }},
        {{ label: 'Упущенная выручка', num: true, render: r => money(r.lost_revenue) }},
        {{ label: 'Упущенная прибыль', num: true, render: r => `<span class="positive">${{money(r.lost_profit)}}</span>` }},
      ];
      document.getElementById('lostSales').innerHTML = table(headers, view.lost_sales);
    }}

    function reconciliationValue(value, indicator) {{
      if (value === null || value === undefined || value === '') return '';
      if (String(indicator || '').includes('%')) return percent(value);
      if (typeof value === 'number') return money(value);
      return String(value);
    }}

    function renderReconciliation(view) {{
      const headers = [
        {{ label: 'Показатель', render: r => `<strong>${{r.indicator}}</strong>` }},
        {{ label: 'WB-витрина', num: true, render: r => reconciliationValue(r.wb, r.indicator) }},
        {{ label: '1С/ОПиУ', num: true, render: r => reconciliationValue(r.onec, r.indicator) }},
        {{ label: 'Дельта', num: true, render: r => reconciliationValue(r.delta, r.indicator) }},
        {{ label: 'Комментарий', render: r => r.comment || '' }},
      ];
      document.getElementById('reconciliationTable').innerHTML = table(headers, view.reconciliation || []);
      const monthlyHeaders = [
        {{ label: 'Месяц', render: r => `<strong>${{r.month}}</strong>` }},
        {{ label: 'WB кол-во', num: true, render: r => number(r.wb_quantity) }},
        {{ label: '1С кол-во', num: true, render: r => number(r.onec_quantity) }},
        {{ label: 'Δ кол-ва', num: true, render: r => `<span class="${{classByValue(r.quantity_delta)}}">${{number(r.quantity_delta)}}</span>` }},
        {{ label: 'Себестоимость 1С в WB-расчете', num: true, render: r => money(r.wb_cogs) }},
        {{ label: 'Себестоимость по валовой прибыли 1С', num: true, render: r => money(r.onec_cogs) }},
        {{ label: 'Δ себестоимости', num: true, render: r => `<span class="${{classByValue(r.cogs_delta)}}">${{money(r.cogs_delta)}}</span>` }},
        {{ label: 'WB расходы МП', num: true, render: r => money(r.wb_mp_expenses) }},
        {{ label: '1С расходы МП', num: true, render: r => money(r.onec_mp_expenses) }},
        {{ label: 'Δ расходов МП', num: true, render: r => `<span class="${{classByValue(r.mp_expenses_delta)}}">${{money(r.mp_expenses_delta)}}</span>` }},
      ];
      document.getElementById('reconciliationMonthlyTable').innerHTML = table(monthlyHeaders, view.reconciliation_monthly || []);
    }}

    function renderQuality(view) {{
      const headers = [
        {{ label: 'Статус данных', render: r => r.status }},
        {{ label: 'Строк', num: true, render: r => number(r.rows) }},
        {{ label: 'Доля', num: true, render: r => percent(r.share) }},
      ];
      document.getElementById('statusTable').innerHTML = table(headers, view.statuses);
    }}

    function render() {{
      const view = payload.views[current.view];
      renderKpis(view);
      renderMonthly(view);
      renderExpenses(view);
      renderLosses(view);
      renderReturns(view);
      renderLostSales(view);
      renderReconciliation(view);
      renderQuality(view);
    }}

    function initControls() {{
      const select = document.getElementById('cabinetSelect');
      select.innerHTML = payload.cabinets.map(name => `<option value="${{name}}">${{name}}</option>`).join('');
      select.addEventListener('change', event => {{
        current.view = event.target.value;
        render();
      }});
      document.querySelectorAll('.toggle').forEach(button => {{
        button.addEventListener('click', () => {{
          const section = document.getElementById(button.dataset.section);
          section.classList.toggle('is-hidden');
          button.classList.toggle('active');
        }});
      }});
    }}

    initControls();
    render();
  </script>
</body>
</html>
"""


def build_dashboard(
    workbook_path: Path = DEFAULT_WORKBOOK,
    output_path: Path = DEFAULT_OUTPUT,
    logo_path: Path | None = DEFAULT_LOGO,
) -> Path:
    payload = collect_dashboard_data(workbook_path)
    logo_uri = _logo_data_uri(logo_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(_render_html(payload, logo_uri), encoding="utf-8")
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a portable client demo dashboard from Excel MVP."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--logo", type=Path, default=DEFAULT_LOGO)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = build_dashboard(args.workbook, args.output, args.logo)
    print(f"HTML dashboard written: {output}")


if __name__ == "__main__":
    main()
