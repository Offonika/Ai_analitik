from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_WORKBOOK = Path("reports/shumeyko_wb_excel_mvp.xlsx")
DEFAULT_OUTPUT_DIR = Path("reports/power_bi_marts")

MARTS = [
    ("kpi_period", "Дашборд", ("Показатель",)),
    ("monthly_dynamics", "Динамика", ("Месяц",)),
    ("expenses", "Расходы WB", ("Статья",)),
    ("unit_economics", "Юнит экономика", ("Неделя",)),
    ("returns", "Возвраты", ("Неделя",)),
    ("lost_sales", "Упущенные продажи", ("Кабинет WB", "Товар")),
    ("onec_opiu_reconciliation", "Сверка с 1С ОПиУ", ("Показатель",)),
]


def main() -> int:
    args = _parse_args()
    paths = export_power_bi_marts(args.workbook, args.output_dir)
    for path in paths:
        print(path)
    return 0


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export client Excel MVP sheets as Power BI/Power Query CSV marts."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def export_power_bi_marts(workbook_path: Path, output_dir: Path) -> list[Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        result = []
        for mart_name, sheet_name, header_candidates in MARTS:
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            header_row = _find_header_row(worksheet, header_candidates)
            if header_row is None:
                continue
            headers, rows = _read_rows(worksheet, header_row)
            output_path = output_dir / f"{mart_name}.csv"
            _write_csv(output_path, headers, rows)
            result.append(output_path)
        _write_readme(output_dir, workbook_path, result)
        return result
    finally:
        workbook.close()


def _find_header_row(worksheet: Any, candidates: tuple[str, ...]) -> int | None:
    for row_number in range(1, min(worksheet.max_row, 80) + 1):
        values = [
            str(cell.value).strip()
            for cell in worksheet[row_number]
            if cell.value not in (None, "")
        ]
        if any(candidate in values for candidate in candidates):
            return row_number
    return None


def _read_rows(worksheet: Any, header_row: int) -> tuple[list[str], list[list[object]]]:
    headers = [
        str(cell.value).strip()
        for cell in worksheet[header_row]
        if cell.value not in (None, "")
    ]
    rows: list[list[object]] = []
    for excel_row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = list(excel_row[: len(headers)])
        if not any(value not in (None, "") for value in values):
            break
        rows.append(values)
    return headers, rows


def _write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(headers)
        writer.writerows(rows)


def _write_readme(output_dir: Path, workbook_path: Path, paths: list[Path]) -> None:
    lines = [
        "# Power BI / Power Query marts",
        "",
        f"Источник: `{workbook_path}`.",
        "",
        "CSV-файлы сформированы из клиентских листов Excel MVP, а не из raw snapshots.",
        "Их можно подключить в Power Query как папку или отдельные CSV-таблицы.",
        "",
        "Файлы:",
        *[f"- `{path.name}`" for path in paths],
        "",
        (
            "Для production Power BI следующий шаг - подключение к расчетным "
            "витринам базы,"
        ),
        "а не к исходным WB/1C raw-данным.",
    ]
    (output_dir / "README.md").write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
