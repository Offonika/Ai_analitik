#!/usr/bin/env python3
"""Attach an unconfirmed local tax schedule to a test-only tax-load canary."""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from wb_unit_economics.web import repository  # noqa: E402
from wb_unit_economics.web.database import (  # noqa: E402
    make_engine,
    make_session_factory,
)
from wb_unit_economics.web.models import ReportRun, TaxLoadReport  # noqa: E402
from wb_unit_economics.web.report_kinds import TAX_LOAD  # noqa: E402
from wb_unit_economics.web.reports.builders import (  # noqa: E402
    build_tax_load_payload,
    canonical_payload_sha256,
)
from wb_unit_economics.web.settings import WebSettings  # noqa: E402

DRAFT_SOURCE_KIND = "manual_tax_load_draft_reference"
DRAFT_ISSUE_CODE = "draft_reference_requires_accountant_review"
SCHEDULE_SHEET_MARKER = "график платеж"
MAX_WORKBOOK_BYTES = 20 * 1024 * 1024


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report-id", required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--workbook-root", type=Path, default=ROOT / "reports")
    parser.add_argument("--apply", action="store_true")
    return parser.parse_args()


def _resolved_workbook(path: Path, allowed_root: Path) -> Path:
    resolved = path.resolve()
    root = allowed_root.resolve()
    if not resolved.is_relative_to(root) or not resolved.is_file():
        raise SystemExit("workbook is outside the allowed local report root")
    if resolved.suffix.lower() != ".xlsx":
        raise SystemExit("workbook must be an xlsx file")
    if resolved.stat().st_size > MAX_WORKBOOK_BYTES:
        raise SystemExit("workbook is too large")
    return resolved


def _decimal_text(value: object) -> str | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        return format(Decimal(str(value)), "f")
    except (InvalidOperation, TypeError, ValueError):
        return None


def _due_date(value: object, *, default_year: int) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?", text)
    if not match:
        return None
    day, month, year = match.groups()
    parsed_year = int(year) if year else default_year
    if parsed_year < 100:
        parsed_year += 2000
    try:
        return date(parsed_year, int(month), int(day)).isoformat()
    except ValueError:
        return None


def _payment_classification(name: str) -> tuple[str, bool, str | None]:
    normalized = name.casefold()
    if "ндфл" in normalized:
        return "agent_ndfl", False, "agent_payment"
    if "взнос" in normalized:
        return "insurance_contribution", False, "insurance_contribution"
    return "own_tax", True, None


def _schedule_sheet(workbook: Any) -> Any:
    matches = [
        sheet
        for sheet in workbook.worksheets
        if SCHEDULE_SHEET_MARKER in sheet.title.casefold()
    ]
    if len(matches) != 1:
        raise SystemExit("workbook must contain exactly one tax payment schedule")
    return matches[0]


def _header_row(sheet: Any) -> int:
    upper = min(sheet.max_row, 50)
    for row_index in range(1, upper):
        current = [sheet.cell(row_index, column).value for column in range(1, 8)]
        next_amount = _decimal_text(sheet.cell(row_index + 1, 2).value)
        if sum(value not in (None, "") for value in current) >= 5 and next_amount:
            return row_index
    raise SystemExit("tax payment schedule header was not found")


def draft_reference_rows(path: Path, *, default_year: int) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = _schedule_sheet(workbook)
    start = _header_row(sheet) + 1
    rows: list[dict[str, Any]] = []
    for row_index in range(start, min(sheet.max_row, start + 100) + 1):
        name = str(sheet.cell(row_index, 1).value or "").strip()
        amount = _decimal_text(sheet.cell(row_index, 2).value)
        if not name and amount is None:
            break
        if not name or amount is None:
            continue
        payment_kind, included, exclusion_reason = _payment_classification(name)
        rows.append(
            {
                "taxCode": f"draft_reference_{len(rows) + 1}",
                "taxName": name,
                "periodKind": "informational_schedule",
                "taxBase": None,
                "accrued": None,
                "paid": None,
                "balance": amount,
                "dueDate": _due_date(
                    sheet.cell(row_index, 3).value,
                    default_year=default_year,
                ),
                "valueStatus": "draft_reference",
                "evidenceStatus": "partial_source",
                "sourceKind": DRAFT_SOURCE_KIND,
                "issueCode": DRAFT_ISSUE_CODE,
                "paymentKind": payment_kind,
                "includedInFnsTaxBurden": included,
                "exclusionReason": exclusion_reason,
            }
        )
    if not rows:
        raise SystemExit("tax payment schedule has no importable rows")
    return rows


def _normalized_evidence(
    report: ReportRun,
    stored_payload: dict[str, Any],
    rows: list[dict[str, Any]],
    source_sha256: str,
) -> dict[str, Any]:
    existing_rows = list(stored_payload.get("taxRows") or [])
    non_reference_rows = [
        row for row in existing_rows if row.get("sourceKind") != DRAFT_SOURCE_KIND
    ]
    if non_reference_rows:
        raise SystemExit("report already contains non-reference tax rows")
    coverage = [
        item
        for item in list(stored_payload.get("sourceCoverage") or [])
        if item.get("sourceKind") != DRAFT_SOURCE_KIND
    ]
    coverage.append(
        {
            "sourceKind": DRAFT_SOURCE_KIND,
            "periodStart": report.period_start.isoformat(),
            "periodEnd": report.period_end.isoformat(),
            "status": "partial_source",
            "snapshotId": source_sha256,
        }
    )
    issues = [
        item
        for item in list(stored_payload.get("issues") or [])
        if item.get("code") not in {DRAFT_ISSUE_CODE, "fns_ratio_source_gap"}
    ]
    issues.append(
        {
            "code": DRAFT_ISSUE_CODE,
            "severity": "warning",
            "section": "График платежей",
            "message": (
                "Информационный график перенесён из локального черновика и "
                "не подтверждает факт начисления или уплаты."
            ),
            "nextAction": "Проверить суммы и сроки ответственным специалистом.",
        }
    )
    evidence = {
        "sourceRefreshRunId": str(
            (stored_payload.get("meta") or {}).get("sourceRefreshRunId") or ""
        ),
        "sourceCoverage": coverage,
        "taxRows": rows,
        "paymentSchedule": [
            {
                "taxCode": row["taxCode"],
                "taxName": row["taxName"],
                "dueDate": row["dueDate"],
                "amount": row["balance"],
                "confirmationStatus": "draft_reference",
                "evidenceStatus": "partial_source",
                "sourceKind": DRAFT_SOURCE_KIND,
                "issueCode": DRAFT_ISSUE_CODE,
            }
            for row in rows
        ],
        "vatSummary": dict(stored_payload.get("vatSummary") or {}),
        "ensSummary": dict(stored_payload.get("ensSummary") or {}),
        "issues": issues,
    }
    evidence["evidenceSha256"] = canonical_payload_sha256(evidence)
    return evidence


def main() -> int:
    args = parse_args()
    settings = WebSettings()
    if settings.runtime_environment != "test":
        raise SystemExit("draft reference attachment is allowed only in test")
    workbook_path = _resolved_workbook(args.workbook, args.workbook_root)
    source_sha256 = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    engine = make_engine(
        settings.database_url,
        statement_timeout_ms=settings.postgres_statement_timeout_ms,
    )
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = db.scalar(
            select(ReportRun)
            .where(ReportRun.id == args.report_id)
            .with_for_update()
        )
        if report is None or report.report_kind != TAX_LOAD:
            raise SystemExit("tax-load report not found")
        rows = draft_reference_rows(
            workbook_path,
            default_year=report.period_end.year,
        )
        stored = db.get(TaxLoadReport, report.id)
        if stored is None:
            raise SystemExit("tax-load payload not found")
        evidence = _normalized_evidence(
            report,
            dict(stored.payload or {}),
            rows,
            source_sha256,
        )
        payload = build_tax_load_payload(
            report,
            tax_profile=dict((stored.payload or {}).get("taxProfile") or {}),
            evidence=evidence,
        )
        payload_sha256 = canonical_payload_sha256(payload)
        if args.apply:
            stored.contract_version = str(payload["contractVersion"])
            stored.payload = payload
            stored.payload_sha256 = payload_sha256
            report.status = str(payload["businessStatus"])
            repository.audit(
                db,
                action="tax_load_draft_reference_attached",
                tenant_id=report.tenant_id,
                entity_type="report_run",
                entity_id=report.id,
                payload={
                    "sourceSha256": source_sha256,
                    "rowsImported": len(rows),
                    "evidenceStatus": "partial_source",
                },
            )
            db.commit()
    print(f"report_id={args.report_id}")
    print(f"rows_imported={len(rows)}")
    print(f"status={'applied' if args.apply else 'dry_run'}")
    print(f"payload_sha256={payload_sha256}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
