from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from scripts.run_report_export_jobs import claim_export_job, execute_export_job
from wb_unit_economics.web import repository
from wb_unit_economics.web.database import init_db, make_engine, make_session_factory
from wb_unit_economics.web.models import ReportExportJob
from wb_unit_economics.web.settings import WebSettings


def test_report_export_job_builds_verified_excel_outside_http(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    settings = WebSettings(
        _env_file=None,
        database_url=str(engine.url),
        allowed_export_root=str(tmp_path / "exports"),
    )
    now = datetime.now(UTC)
    with session_factory() as db:
        report = repository.save_report_marts(
            db,
            _payload(),
            tenant_id="tenant",
            tenant_name="Tenant",
            report_id="report-export",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="snapshot-export",
        )
        job = ReportExportJob(
            id="export-job",
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            report_run_id=report.id,
            task_id=None,
            artifact_id=None,
            export_format="xlsx",
            idempotency_key="export-once",
            request_fingerprint="fingerprint",
            status="queued",
            safe_error_code="",
            created_at=now,
            started_at=None,
            finished_at=None,
            updated_at=now,
        )
        db.add(job)
        db.commit()

        claimed = claim_export_job(db)
        assert claimed is not None
        assert claimed.id == job.id
        execute_export_job(db, claimed, settings)
        db.refresh(claimed)
        artifact = claimed.artifact_id
        artifact_row = repository.report_artifact_path(
            db,
            report,
            "excel",
            settings.export_root_path,
        )

    assert claimed.status == "succeeded"
    assert claimed.finished_at is not None
    assert artifact is not None
    assert artifact_row is not None and artifact_row.is_file()
    workbook = load_workbook(artifact_row, read_only=True, data_only=True)
    try:
        assert "Юнит экономика" in workbook.sheetnames
    finally:
        workbook.close()


def _payload() -> dict:
    return {
        "meta": {
            "client": "Client",
            "clientId": "client",
            "period": "01.07.2026 - 31.07.2026",
            "methodologyVersion": "export-job-test-v1",
            "source": "DB report marts",
        },
        "readiness": {"status": "ready"},
        "options": {},
        "monthly": [],
        "expenses": [],
        "unitRows": [
            {
                "id": "unit-1",
                "product": "Товар",
                "week": "2026-07-06",
                "month": "Июль 2026",
                "organization": "Организация",
                "cabinet": "Кабинет",
                "articleWb": "WB-1",
                "article1c": "1C-1",
                "sales": 1,
                "returns": 0,
                "revenue": 100,
                "cost": 20,
                "profitBeforeTax": 50,
                "profit": 45,
                "status": "ОК",
            }
        ],
        "returns": [],
        "lostSales": [],
        "reconciliationMonthly": [],
        "documentReconciliation": [],
    }
