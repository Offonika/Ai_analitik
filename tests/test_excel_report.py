from __future__ import annotations

import json
from datetime import date, datetime, timedelta
from decimal import Decimal
from types import SimpleNamespace
from zipfile import ZipFile
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from tests.fixtures import (
    CLIENT_ID,
    account_org_mapping,
    cost_snapshots,
    sku_mappings,
    wb_snapshots,
)
from wb_unit_economics.calculation import build_unit_economics_report
from wb_unit_economics.contracts import (
    DataQualityStatus,
    OnecGrossProfitDocumentRow,
    OnecMarketplaceServiceRow,
    OnecReportKind,
    OnecReportProductRow,
    OnecReportReconciliationRow,
    OnecUnfCostSnapshot,
    ReportStatus,
    SalesModel,
    UnitEconomicsReport,
    UnitEconomicsRow,
    WbSalesReportSummaryRow,
)
from wb_unit_economics.excel import (
    REQUIRED_SHEETS,
    _document_reconciliation_row,
    _expected_sales_week_for_onec_document,
    _report_period_status_label,
    build_excel_report,
)
from wb_unit_economics.onec_opiu import OnecOpiuSummary


def _complete_stock_history_csv(*, zero_date: date) -> str:
    start = date(2026, 3, 1)
    end = date(2026, 6, 17)
    dates = []
    current = start
    while current <= end:
        dates.append(current)
        current += timedelta(days=1)
    headers = ",".join(item.strftime("%d.%m.%Y") for item in dates)
    values = ",".join("0" if item == zero_date else "3" for item in dates)
    return f"NmID,VendorCode,Name,{headers}\n101,A-1,Product 1,{values}\n"


def test_excel_report_has_required_sheets_and_reconciled_summary(tmp_path) -> None:
    stock_history_dir = tmp_path / "wb_stock_history"
    stock_history_dir.mkdir()
    with ZipFile(stock_history_dir / "stock_history.zip", "w") as archive:
        archive.writestr(
            "stock.csv",
                (
                    _complete_stock_history_csv(zero_date=date(2026, 3, 2))
                ),
        )
    (stock_history_dir / "manifest.json").write_text(
        json.dumps(
            {
                "period_start": "2026-03-01",
                    "period_end": "2026-06-17",
                    "stock_type": "wb",
                "results": [
                        {
                            "status": "ok",
                            "seller_account_id": "WB_ACCOUNT_1",
                            "output_file": "stock_history.zip",
                        },
                        {
                            "status": "ok",
                            "seller_account_id": "WB_ACCOUNT_2",
                            "output_file": "stock_history.zip",
                        },
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    onec_stock_dir = tmp_path / "onec_stock"
    onec_stock_dir.mkdir()
    (onec_stock_dir / "manifest.json").write_text(
        json.dumps(
            {
                "results": [
                    {
                        "sample_id": "stock_by_warehouse",
                        "ok": True,
                        "row_count": 1,
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (onec_stock_dir / "stock_by_warehouse.raw.json").write_text(
        json.dumps(
            {
                "value": [
                    {
                        "RecordSet": [
                            {
                                "Active": True,
                                "Номенклатура_Key": "ONEC-1",
                                "Организация_Key": "1C_ORG_1",
                                "Характеристика_Key": "CHAR-1",
                                "Склад_Key": "WAREHOUSE-1",
                                "Количество": "7",
                            }
                        ]
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (onec_stock_dir / "Catalog_СтруктурныеЕдиницы.raw.json").write_text(
        json.dumps(
            {
                "value": [
                    {
                        "Ref_Key": "WAREHOUSE-1",
                        "Description": "Собственный склад",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    report = build_unit_economics_report(
        client_id=CLIENT_ID,
        wb_snapshots=wb_snapshots(),
        cost_snapshots=cost_snapshots(),
        sku_mappings=sku_mappings(),
        account_org_mapping=account_org_mapping(),
        generated_at=datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        as_of_date=date(2026, 6, 16),
    )
    output = build_excel_report(
        report,
        tmp_path / "report.xlsx",
        include_ai_summary=True,
        cost_snapshots=cost_snapshots(),
        sku_mappings=sku_mappings(),
        wb_sales_report_summary_rows=[
            WbSalesReportSummaryRow(
                client_id=CLIENT_ID,
                seller_account_id="WB_ACCOUNT_1",
                account_name="WB_ACCOUNT_1",
                report_id="SUMMARY-1",
                date_from=date(2026, 4, 6),
                date_to=date(2026, 4, 12),
                create_date=date(2026, 4, 13),
                report_type=1,
                retail_amount_sum="1000",
                for_pay_sum="850",
                delivery_service_sum="50",
                paid_storage_sum="20",
                deduction_sum="8",
                penalty_sum="5",
                cashback_discount_sum="25",
                bank_payment_sum="700",
                raw_payload_hash="summary-hash-1",
            ),
            WbSalesReportSummaryRow(
                client_id=CLIENT_ID,
                seller_account_id="WB_ACCOUNT_1",
                account_name="WB_ACCOUNT_1",
                report_id="BUYOUT-1",
                date_from=date(2026, 4, 6),
                date_to=date(2026, 4, 12),
                create_date=date(2026, 4, 6),
                report_type=2,
                retail_amount_sum="66.74",
                for_pay_sum="54.94",
                delivery_service_sum="28.63",
                paid_storage_sum="0",
                deduction_sum="0",
                penalty_sum="0",
                cashback_discount_sum="0",
                bank_payment_sum="24.31",
                raw_payload_hash="summary-buyout-hash-1",
            )
        ],
        onec_marketplace_service_rows=[
            OnecMarketplaceServiceRow(
                client_id=CLIENT_ID,
                organization_id="1C_ORG_1",
                counterparty_id="RWB",
                document_id="SERVICE-DOC-1",
                document_number="УПД-1",
                input_number="052400020047",
                document_comment="Отчет услуг WB",
                document_date=date(2026, 4, 12),
                input_date=date(2026, 4, 12),
                week_start=date(2026, 4, 6),
                week_end=date(2026, 4, 12),
                service_category="Комиссия WB",
                service_name="Комиссионное вознаграждение ВБ",
                amount="185",
                vat="33.30",
                total="185",
                source_row_hash="service-hash-1",
            )
        ],
        onec_gross_profit_rows=[
            OnecGrossProfitDocumentRow(
                client_id=CLIENT_ID,
                organization_id="1C_ORG_1",
                counterparty_id="RWB",
                document_id="DOC-1",
                document_type="РасходнаяНакладная",
                document_date=date(2026, 6, 16),
                week_start=date(2026, 6, 9),
                week_end=date(2026, 6, 15),
                quantity="39",
                revenue="24927.88",
                vat="1187.04",
                cogs="11167.58",
                gross_profit="13760.30",
                source_row_count=22,
            ),
            OnecGrossProfitDocumentRow(
                client_id=CLIENT_ID,
                organization_id="1C_ORG_1",
                counterparty_id="RWB",
                document_id="DOC-COMMISSIONER-1",
                document_type="ОтчетКомиссионера",
                document_number="ОК-000001",
                input_number="SUMMARY-1",
                document_date=date(2026, 4, 12),
                week_start=date(2026, 4, 6),
                week_end=date(2026, 4, 12),
                quantity="2",
                revenue="1000",
                vat="47.62",
                cogs="430",
                gross_profit="570",
                external_report_id="SUMMARY-1",
                settlement_total="850",
                source_row_count=10,
            ),
            OnecGrossProfitDocumentRow(
                client_id=CLIENT_ID,
                organization_id="1C_ORG_1",
                counterparty_id="RWB",
                document_id="DOC-COMMISSIONER-1",
                document_type="ОтчетКомиссионера",
                document_date=date(2026, 4, 30),
                week_start=date(2026, 4, 27),
                week_end=date(2026, 5, 3),
                quantity="0",
                revenue="0",
                vat="0",
                cogs="25",
                gross_profit="-25",
                external_report_id="SUMMARY-1",
                settlement_total="0",
                source_row_count=1,
            ),
            OnecGrossProfitDocumentRow(
                client_id=CLIENT_ID,
                organization_id="1C_ORG_1",
                counterparty_id="RWB",
                document_id="DOC-COMMISSIONER-WRONG",
                document_type="ОтчетКомиссионера",
                document_date=date(2026, 4, 12),
                week_start=date(2026, 4, 6),
                week_end=date(2026, 4, 12),
                quantity="99",
                revenue="9999",
                vat="0",
                cogs="0",
                gross_profit="9999",
                external_report_id="SUMMARY-OTHER",
                settlement_total="9999",
                source_row_count=1,
            )
        ],
        onec_opiu_summary=OnecOpiuSummary(
            source_label="test income_expense_register",
            source_row_count=3,
            values={
                "revenue": Decimal("1000"),
                "vat": Decimal("47.62"),
                "revenue_without_vat": Decimal("952.38"),
                "rwb_total": Decimal("250"),
                "rwb_commission": Decimal("185"),
                "rwb_logistics": Decimal("0"),
                "rwb_promotion": Decimal("0"),
                "rwb_fines": Decimal("0"),
                "rwb_acquiring": Decimal("0"),
                "cogs": Decimal("430"),
                "net_profit": Decimal("350"),
            },
            monthly_values={
                "2026-03": {
                    "revenue": Decimal("1000"),
                    "cogs": Decimal("430"),
                    "rwb_total": Decimal("250"),
                },
                "2026-04": {
                    "revenue": Decimal("1000"),
                    "cogs": Decimal("430"),
                    "rwb_total": Decimal("250"),
                },
                "2026-06": {
                    "revenue": Decimal("1000"),
                    "cogs": Decimal("430"),
                    "rwb_total": Decimal("250"),
                },
            },
        ),
        account_labels={
            "WB_ACCOUNT_1": "Организация Минзифа",
            "WB_ACCOUNT_2": "Организация Султан",
        },
        organization_labels={
            "1C_ORG_1": "Организация Минзифа",
            "1C_ORG_2": "Организация Султан",
        },
        stock_history_dir=stock_history_dir,
        onec_stock_dir=onec_stock_dir,
    )
    workbook = load_workbook(output, data_only=True)
    formula_workbook = load_workbook(output, data_only=False)
    assert [
        name for name in REQUIRED_SHEETS if name in workbook.sheetnames
    ] == REQUIRED_SHEETS
    assert "ИИ-резюме" in workbook.sheetnames
    summary = workbook["Сводка"]
    summary_values = {
        row[0].value: row[1].value
        for row in summary.iter_rows(min_row=2, max_col=2)
        if row[0].value
    }
    assert summary_values["Выручка до СПП"] == 1000
    assert summary_values["СПП"] == 0
    assert summary_values["Выручка после СПП"] == 1000
    total_cogs = sum(
        (row.cogs_from_1c_with_extra_costs for row in report.rows),
        Decimal("0"),
    )
    assert summary_values[
        "Себестоимость 1С, включая распределенные допрасходы"
    ] == float(total_cogs)
    assert summary_values["Маржинальный доход WB до налогов"] == float(
        report.total_gross_profit
    )
    assert summary_values["НДС к уплате"] == 47.62
    assert summary_values["Управленческая прибыль WB"] == float(
        report.total_profit_after_taxes
    )
    assert summary.tables
    wb_fields = workbook["WB поля отчета"]
    wb_fields_headers = [cell.value for cell in wb_fields[4]]
    assert "retailAmountSum" in wb_fields_headers
    assert "bankPaymentSum" in wb_fields_headers
    payout_status_col = wb_fields_headers.index("Статус выплаты") + 1
    assert wb_fields.cell(5, payout_status_col).value == "Нужен источник выплаты 1С"
    report_reconciliation = workbook["Сверка по отчетам WB"]
    assert report_reconciliation["B1"].value == "Номер отчета WB"
    assert report_reconciliation["B2"].value == "WB-REPORT-1"
    assert report_reconciliation["C2"].value == "Организация Минзифа"
    assert report_reconciliation["D2"].value == "Организация Минзифа"
    assert report_reconciliation["E2"].value == 2
    report_headers = [cell.value for cell in report_reconciliation[1]]
    report_idx = {header: index for index, header in enumerate(report_headers)}
    assert (
        report_reconciliation.cell(2, report_idx["Выручка после СПП"] + 1).value == 1000
    )
    assert (
        report_reconciliation.cell(2, report_idx["НДС к уплате"] + 1).value == 47.62
    )
    assert (
        report_reconciliation.cell(2, report_idx["Налог с выручки/НДФЛ"] + 1).value
        == 10
    )
    assert (
        report_reconciliation.cell(
            2, report_idx["Маржинальный доход WB до налогов"] + 1
        ).value
        == 570
    )
    assert (
        report_reconciliation.cell(
            2, report_idx["Управленческая прибыль WB"] + 1
        ).value
        == 512.38
    )
    assert report_reconciliation.tables
    onec_reconciliation = workbook["Сверка с 1С"]
    assert onec_reconciliation["A1"].value == "Дата документа 1С"
    assert onec_reconciliation["C2"].value == "Отчет комиссионера"
    assert onec_reconciliation["E2"].value == "Организация Минзифа"
    onec_headers = [cell.value for cell in onec_reconciliation[1]]
    onec_idx = {header: index for index, header in enumerate(onec_headers)}
    assert (
        onec_reconciliation.cell(
            2, onec_idx["Номер отчета WB (сводный)"] + 1
        ).value
        == "SUMMARY-1"
    )
    assert (
        onec_reconciliation.cell(
            2, onec_idx["PDF 1. Стоимость реализовано"] + 1
        ).value
        == 1000
    )
    assert (
        onec_reconciliation.cell(2, onec_idx["PDF 1.1 Товар реализован"] + 1).value
        == 975
    )
    assert (
        onec_reconciliation.cell(2, onec_idx["Дельта детализация - PDF 1"] + 1).value
        == 0
    )
    assert onec_reconciliation.tables
    onec_document_reconciliation = workbook["Сверка документов 1С"]
    document_headers = [cell.value for cell in onec_document_reconciliation[1]]
    document_idx = {header: index for index, header in enumerate(document_headers)}
    assert onec_document_reconciliation["A2"].value == "OK"
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["Номер отчета WB (сводный)"] + 1
        ).value
        == "SUMMARY-1"
    )
    assert (
        onec_document_reconciliation.cell(2, document_idx["WB отчет продаж"] + 1).value
        == "SUMMARY-1"
    )
    assert (
        onec_document_reconciliation.cell(2, document_idx["WB отчет выкупов"] + 1).value
        == "BUYOUT-1"
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["WB сумма документа"] + 1
        ).value
        == 1000
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["1С сумма документа"] + 1
        ).value
        == 1000
    )
    assert (
        onec_document_reconciliation.cell(2, document_idx["Документы 1С"] + 1).value
        == "ОтчетКомиссионера № ОК-000001 / вх. SUMMARY-1"
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["Строк регистра 1С"] + 1
        ).value
        == 10
    )
    assert (
        onec_document_reconciliation.cell(2, document_idx["Дельта сумма"] + 1).value
        == 0
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["WB к перечислению (forPaySum)"] + 1
        ).value
        == 850
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["1С оборот взаиморасчетов"] + 1
        ).value
        == 850
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["Дельта к обороту 1С"] + 1
        ).value
        == 0
    )
    assert (
        onec_document_reconciliation.cell(
            2, document_idx["Статус выплаты"] + 1
        ).value
        == "Нужен источник выплаты 1С"
    )
    assert onec_document_reconciliation.cell(2, document_idx["WB продажи"] + 1).value
    assert onec_document_reconciliation.cell(2, document_idx["1С продажи"] + 1).value
    assert onec_document_reconciliation.tables
    opiu_reconciliation = workbook["Сверка с 1С ОПиУ"]
    assert opiu_reconciliation["A1"].value == "Сверка с 1С ОПиУ"
    opiu_rows = {
        row[0].value: row
        for row in opiu_reconciliation.iter_rows(min_row=5, max_col=5)
        if row[0].value
    }
    assert opiu_rows["Выручка после СПП"][1].value == 1000
    assert opiu_rows["Выручка после СПП"][2].value == 35926.88
    assert opiu_rows["Себестоимость РВБ"][1].value == 5
    assert opiu_rows["Себестоимость РВБ"][2].value == 11622.58
    assert opiu_rows["Валовая прибыль РВБ"][2].value == 24304.3
    assert "Выручка 1С/ОПиУ всего" not in opiu_rows
    assert "Себестоимость 1С в юнит-экономике" not in opiu_rows
    assert "Чистая прибыль 1С/ОПиУ" not in opiu_rows

    monthly_header_row = None
    for row_idx in range(1, opiu_reconciliation.max_row + 1):
        if opiu_reconciliation.cell(row_idx, 1).value == "Месяц":
            monthly_header_row = row_idx
            break
    assert monthly_header_row is not None
    monthly_rows = {
        row[0].value: row
        for row in opiu_reconciliation.iter_rows(
            min_row=monthly_header_row + 1,
            max_col=8,
        )
        if row[0].value
    }
    monthly_headers = [
        opiu_reconciliation.cell(monthly_header_row, col).value
        for col in range(1, 12)
    ]
    assert monthly_headers[1] == "WB количество"
    assert monthly_headers[2] == "1С количество"
    assert monthly_headers[4] == "Себестоимость 1С в WB-расчете"
    assert monthly_headers[5] == "Себестоимость по валовой прибыли 1С"
    assert monthly_rows["Апрель 2026"][5].value == 430
    assert monthly_rows["Апрель 2026"][6].value == (
        monthly_rows["Апрель 2026"][4].value
        - monthly_rows["Апрель 2026"][5].value
    )
    onec_gross_profit = workbook["Валовая прибыль 1С"]
    assert onec_gross_profit["C2"].value == "Организация Минзифа"
    assert onec_gross_profit["E2"].value == "РасходнаяНакладная"
    assert onec_gross_profit["G2"].value == 39
    assert onec_gross_profit["H2"].value == 24927.88
    assert onec_gross_profit["J2"].value == 11167.58
    assert onec_gross_profit.tables
    service_reconciliation = workbook["Сверка услуг WB"]
    assert service_reconciliation["A1"].value == "Контрольный блок сверки услуг"
    check_rows = {
        (row[0].value, row[1].value, row[2].value): row
        for row in service_reconciliation.iter_rows(min_row=3, max_col=9)
        if row[2].value
    }
    combined_check = check_rows[
        (
            "2026-04-06",
            "Организация Минзифа",
            "Комиссия + Логистика + Хранение + Эквайринг",
        )
    ]
    assert combined_check[3].value == 185
    assert combined_check[4].value == 185
    assert service_reconciliation.tables
    service_breakdown = workbook["Расшифровка услуг 1С"]
    assert service_breakdown["E2"].value == "УПД-1"
    assert service_breakdown["F2"].value == "052400020047"
    assert service_breakdown["H2"].value == "Отчет услуг WB"
    assert service_breakdown["I2"].value == "Комиссионное вознаграждение ВБ"
    assert service_breakdown["J2"].value == "Комиссия WB"
    assert service_breakdown.tables
    expense_allocation = workbook["Распределение расходов"]
    assert expense_allocation["G1"].value == "Статья"
    assert expense_allocation["R1"].value == "Метод"
    assert expense_allocation["S1"].value == "Статус"
    allocation_statuses = {
        row[18].value
        for row in expense_allocation.iter_rows(min_row=2, max_col=19)
        if row[18].value
    }
    assert "Нет недельного фин. отчета WB, взята детализация" in allocation_statuses
    assert expense_allocation.tables
    onec_products = workbook["Товары по отчетам 1С"]
    assert onec_products["C2"].value == "Отчет комиссионера"
    assert onec_products["E2"].value == "Организация Минзифа"
    assert onec_products["F2"].value == 101
    assert onec_products.tables
    account_summary = workbook["Сводка по кабинетам WB"]
    account_summary_labels = {
        row[0].value for row in account_summary.iter_rows(min_row=2, max_col=1)
    }
    assert "Организация Минзифа" in account_summary_labels
    assert "Организация Султан" in account_summary_labels
    products = workbook["Товары"]
    product_account_labels = {
        row[1].value for row in products.iter_rows(min_row=2, max_col=2)
    }
    assert "Организация Минзифа" in product_account_labels
    assert "Организация Султан" in product_account_labels
    product_org_labels = {
        row[2].value for row in products.iter_rows(min_row=2, max_col=3)
    }
    assert "Организация Минзифа" in product_org_labels
    unit_economics = workbook["Юнит экономика"]
    unit_headers = [cell.value for cell in unit_economics[1]]
    unit_idx = {header: index for index, header in enumerate(unit_headers)}
    assert unit_headers[unit_idx["Документ-отчет"]] == "Документ-отчет"
    assert unit_headers[unit_idx["Номер отчета WB"]] == "Номер отчета WB"
    assert unit_headers[unit_idx["Дата отчета WB"]] == "Дата отчета WB"
    assert unit_headers[unit_idx["Организация 1С"]] == "Организация 1С"
    assert unit_headers[unit_idx["Товар"]] == "Товар"
    assert unit_headers[unit_idx["nmId WB"]] == "nmId WB"
    assert unit_headers[unit_idx["Артикул WB"]] == "Артикул WB"
    assert unit_headers[unit_idx["Артикул 1С"]] == "Артикул 1С"
    assert unit_headers[unit_idx["Продажи, шт"]] == "Продажи, шт"
    assert unit_headers[unit_idx["Возвраты, шт"]] == "Возвраты, шт"
    assert unit_headers[unit_idx["Чистое кол-во"]] == "Чистое кол-во"
    assert unit_headers[unit_idx["% возвратов"]] == "% возвратов"
    assert unit_headers[unit_idx["Выручка до СПП"]] == "Выручка до СПП"
    assert unit_headers[unit_idx["СПП"]] == "СПП"
    assert unit_headers[unit_idx["% СПП"]] == "% СПП"
    assert unit_headers[unit_idx["Выручка после СПП"]] == "Выручка после СПП"
    assert unit_headers[unit_idx["НДС к уплате"]] == "НДС к уплате"
    assert unit_headers[unit_idx["Исходящий НДС"]] == "Исходящий НДС"
    assert unit_headers[unit_idx["Входящий НДС"]] == "Входящий НДС"
    assert unit_headers[unit_idx["Хранение WB"]] == "Хранение WB"
    assert unit_headers[unit_idx["Продвижение WB"]] == "Продвижение WB"
    assert (
        unit_headers[unit_idx["Налог с выручки/НДФЛ"]]
        == "Налог с выручки/НДФЛ"
    )
    assert (
        unit_headers[unit_idx["Управленческая прибыль WB"]]
        == "Управленческая прибыль WB"
    )
    assert unit_headers[unit_idx["Причина статуса"]] == "Причина статуса"
    assert unit_headers[unit_idx["Статус СПП"]] == "Статус СПП"
    assert unit_headers[unit_idx["Налоговый метод"]] == "Налоговый метод"
    assert (
        unit_headers[unit_idx["Источник налогового профиля"]]
        == "Источник налогового профиля"
    )
    assert (
        unit_headers[unit_idx["Полнота налогового расчета"]]
        == "Полнота налогового расчета"
    )
    assert unit_headers[unit_idx["Режим P&L НДС"]] == "Режим P&L НДС"
    assert unit_headers[unit_idx["НДС входящий WB"]] == "НДС входящий WB"
    assert unit_headers[unit_idx["НДС входящий 1С"]] == "НДС входящий 1С"
    assert unit_headers[unit_idx["Расхождение НДС"]] == "Расхождение НДС"
    assert unit_headers[unit_idx["Полнота НДС"]] == "Полнота НДС"
    assert "Сверка входящего НДС" in workbook.sheetnames
    unit_rows = [
        row
        for row in unit_economics.iter_rows(min_row=2, max_col=len(unit_headers))
        if row[unit_idx["Товар"]].value == "Product 1"
    ]
    assert unit_rows
    assert (
        unit_rows[0][unit_idx["Документ-отчет"]].value
        == "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
    )
    assert unit_rows[0][unit_idx["Номер отчета WB"]].value == "WB-REPORT-1"
    assert unit_rows[0][unit_idx["Дата отчета WB"]].value == "2026-04-13"
    assert unit_rows[0][unit_idx["Организация 1С"]].value == "Организация Минзифа"
    assert unit_rows[0][unit_idx["Кабинет WB"]].value == "Организация Минзифа"
    assert unit_rows[0][unit_idx["Артикул WB"]].value == "A-1"
    assert unit_rows[0][unit_idx["Артикул 1С"]].value == "A-1"
    assert unit_rows[0][unit_idx["Продажи, шт"]].value == 2
    assert unit_rows[0][unit_idx["Возвраты, шт"]].value == 0
    assert unit_rows[0][unit_idx["Чистое кол-во"]].value == 2
    assert unit_rows[0][unit_idx["% возвратов"]].value == 0
    assert unit_rows[0][unit_idx["Выручка до СПП"]].value == 1000
    assert unit_rows[0][unit_idx["СПП"]].value == 0
    assert unit_rows[0][unit_idx["Выручка после СПП"]].value == 1000
    assert unit_rows[0][unit_idx["Хранение WB"]].value == 20
    assert unit_rows[0][unit_idx["Продвижение WB"]].value == 0
    assert unit_rows[0][unit_idx["Налог с выручки/НДФЛ"]].value == 10
    assert unit_rows[0][unit_idx["Управленческая прибыль WB"]].value == 512.38
    assert unit_rows[0][unit_idx["Режим P&L НДС"]].value == "legacy_tax_layer"
    assert (
        unit_rows[0][unit_idx["Налоговый метод"]].value
        == "legacy: НДС внутри цены 5/105; налог с выручки 1%"
    )
    assert (
        unit_rows[0][unit_idx["Источник налогового профиля"]].value
        == "legacy-default"
    )
    assert (
        unit_rows[0][unit_idx["Полнота налогового расчета"]].value
        == "legacy_complete"
    )
    assert unit_rows[0][unit_idx["Статус СПП"]].value == (
        "СПП не передается текущим источником"
    )
    assert all(
        row[unit_idx["Товар"]].value != "Товар не определен"
        for row in unit_economics.iter_rows(min_row=2, max_col=len(unit_headers))
        if row[unit_idx["Товар"]].value
    )
    assert unit_economics.tables
    liquidity = workbook["Ликвидность МД"]
    liquidity_headers = [cell.value for cell in liquidity[1]]
    liquidity_idx = {header: index for index, header in enumerate(liquidity_headers)}
    assert liquidity_headers[liquidity_idx["МД1 Наценка"]] == "МД1 Наценка"
    assert (
        liquidity_headers[liquidity_idx["МД4 после логистики и приемки"]]
        == "МД4 после логистики и приемки"
    )
    assert (
        liquidity_headers[liquidity_idx["Упр. прибыль"]]
        == "Упр. прибыль"
    )
    assert (
        liquidity_headers[liquidity_idx["Статус ликвидности"]]
        == "Статус ликвидности"
    )
    assert liquidity_headers[liquidity_idx["Статус данных"]] == "Статус данных"
    assert liquidity.tables
    liquidity_profit = sum(
        row[liquidity_idx["Упр. прибыль"]].value or 0
        for row in liquidity.iter_rows(min_row=2, max_col=len(liquidity_headers))
        if row[liquidity_idx["Товар"]].value
    )
    unit_profit = sum(
        row[unit_idx["Управленческая прибыль WB"]].value or 0
        for row in unit_economics.iter_rows(min_row=2, max_col=len(unit_headers))
        if row[unit_idx["Товар"]].value
    )
    assert round(liquidity_profit, 2) == round(unit_profit, 2)
    liquidity_formula_sheet = formula_workbook["Ликвидность МД"]
    assert not any(
        cell.data_type == "f"
        for row in liquidity_formula_sheet.iter_rows()
        for cell in row
    )
    assert not any(
        cell.value in {"#DIV/0!", "#VALUE!", "#REF!", "#NAME?"}
        for row in liquidity.iter_rows()
        for cell in row
    )
    visible_sheets = {
        name for name in workbook.sheetnames if workbook[name].sheet_state == "visible"
    }
    assert visible_sheets == {
        "Дашборд",
        "Юнит экономика",
        "Ликвидность МД",
        "Динамика",
        "Расходы WB",
        "Возвраты",
        "Упущенные продажи",
        "Сверка документов 1С",
        "Сверка с 1С ОПиУ",
        "Ошибки данных",
        "Методика",
    }
    assert "Сводные" not in workbook.sheetnames
    assert workbook.active.title == "Дашборд"
    dashboard = workbook["Дашборд"]
    assert dashboard["D1"].value == "Период: 01.03.2026 - 17.06.2026"
    assert dashboard["G1"].value == "Статус: Неполный период"
    assert dashboard["J1"].value == "Дата расчета: 16.06.2026"
    readme = workbook["README"]
    readme_rows = {
        row[0].value: row[1].value
        for row in readme.iter_rows(min_row=2, max_col=2)
        if row[0].value
    }
    assert readme_rows["Статус готовности"] == "source_coverage_gap"
    assert readme_rows["Статус периода"] == "Неполный период"
    assert dashboard["A2"].value == "Ключевые показатели"
    assert dashboard["D2"].value == "Статусы данных"
    dashboard_labels = [cell.value for cell in dashboard["A"] if cell.value]
    assert "Динамика месяц к месяцу" in dashboard_labels
    assert "Динамика по неделям" in dashboard_labels
    assert "Продажи, шт" in dashboard_labels
    assert "Возвраты, шт" in dashboard_labels
    assert "Выручка до СПП" in dashboard_labels
    assert "Выручка после СПП" in dashboard_labels
    assert "Управленческая прибыль WB" in dashboard_labels
    assert "Доля возвратов" in dashboard_labels
    assert "Убыточных SKU" in dashboard_labels
    assert dashboard.tables
    assert dashboard.conditional_formatting
    assert unit_economics.conditional_formatting
    dynamics = workbook["Динамика"]
    assert dynamics["A1"].value == "Динамика месяц к месяцу"
    assert dynamics["A2"].value == "Месяц"
    assert dynamics["F2"].value == "Выручка до СПП"
    assert dynamics["I2"].value == "Выручка после СПП"
    expenses = workbook["Расходы WB"]
    assert expenses["A1"].value == "Структура расходов"
    expense_headers = [cell.value for cell in expenses[2]]
    assert "% от выручки" in expense_headers
    assert "Март" in expense_headers
    assert "Апрель к Март" in expense_headers
    returns = workbook["Возвраты"]
    return_headers = [cell.value for cell in returns[1]]
    return_idx = {header: index for index, header in enumerate(return_headers)}
    assert return_headers[return_idx["Баркод"]] == "Баркод"
    assert return_headers[return_idx["Причина возврата"]] == "Причина возврата"
    assert (
        return_headers[return_idx["Управленческая прибыль WB"]]
        == "Управленческая прибыль WB"
    )
    return_rows = [
        row
        for row in returns.iter_rows(min_row=2, max_col=len(return_headers))
        if row[return_idx["Товар"]].value == "Product 2"
    ]
    assert return_rows
    assert return_rows[0][return_idx["Баркод"]].value == "222"
    assert return_rows[0][return_idx["Возвраты, шт"]].value == 1
    assert return_rows[0][return_idx["% возвратов"]].value is None
    assert return_rows[0][return_idx["Причина возврата"]].value == (
        "Причина возврата не передается текущими источниками"
    )
    lost_sales = workbook["Упущенные продажи"]
    assert lost_sales["A1"].value == (
        "Оценка недополученного маржинального дохода до налогов"
    )
    assert lost_sales.row_dimensions[2].height == 18
    assert lost_sales.row_dimensions[3].height == 18
    assert lost_sales["A14"].value == "Кабинет WB"
    assert lost_sales["H14"].value == "Дней без остатка WB"
    assert lost_sales["J14"].value == "Остаток 1С на складах, шт"
    assert lost_sales["K14"].value == "Склады 1С с остатком"
    assert lost_sales["P14"].value == (
        "Оценка недополученного маржинального дохода до налогов"
    )
    lost_headers = [cell.value for cell in lost_sales[14]]
    lost_idx = {header: index for index, header in enumerate(lost_headers)}
    product_1_lost_rows = [
        row
        for row in lost_sales.iter_rows(min_row=15, max_col=len(lost_headers))
        if row[lost_idx["Товар"]].value == "Product 1"
    ]
    assert product_1_lost_rows
    assert product_1_lost_rows[0][lost_idx["Остаток 1С на складах, шт"]].value == 7
    assert (
        product_1_lost_rows[0][lost_idx["Склады 1С с остатком"]].value
        == "Собственный склад: 7"
    )
    assert product_1_lost_rows[0][lost_idx["Вывод"]].value == (
        "Переместить собственный остаток на WB"
    )
    org_summary = workbook["Сводка по организациям"]
    assert org_summary["A2"].value == "Организация Минзифа"
    errors = workbook["Ошибки данных"]
    statuses = [row[5].value for row in errors.iter_rows(min_row=2, max_col=6)]
    assert "Неоднозначное сопоставление" in statuses
    assert "Кабинет WB не совпадает с организацией 1С" in statuses
    assert errors.tables
    costs = workbook["Себестоимость 1С"]
    assert costs["A2"].value == "Организация Минзифа"
    mappings = workbook["Маппинг"]
    assert mappings["A2"].value == "Организация Минзифа"
    assert mappings["B2"].value == "Организация Минзифа"
    assert mappings["K2"].value == "Сопоставлено"
    assert mappings.tables


def test_period_status_label_prefers_partial_period_over_final_status() -> None:
    report = SimpleNamespace(
        rows=[SimpleNamespace(data_quality_status=DataQualityStatus.RELIABLE)],
        source_coverage_start=date(2026, 3, 1),
        source_coverage_end=date(2026, 6, 17),
        report_period_start=date(2026, 3, 1),
        report_period_end=date(2026, 6, 17),
        status=ReportStatus.FINAL,
    )

    assert _report_period_status_label(report) == "Неполный период"


def test_onec_monthly_reconciliation_uses_actual_document_date(tmp_path) -> None:
    report = UnitEconomicsReport(
        client_id=CLIENT_ID,
        report_period_start=date(2026, 4, 1),
        report_period_end=date(2026, 5, 31),
        generated_at=datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        status=ReportStatus.FINAL,
        methodology_version="test",
        rows=[
            UnitEconomicsRow(
                client_id=CLIENT_ID,
                seller_account_id="WB_ACCOUNT_1",
                organization_id="1C_ORG_1",
                week_start=date(2026, 4, 27),
                week_end=date(2026, 5, 3),
                is_partial_week=False,
                nm_id=101,
                vendor_code="A-1",
                barcode="111",
                sales_model=SalesModel.FBO,
                quantity=Decimal("1"),
                sales_quantity=Decimal("1"),
                return_quantity=Decimal("0"),
                revenue_before_spp=Decimal("300"),
                spp_discount=Decimal("0"),
                revenue_after_spp=Decimal("300"),
                net_revenue=Decimal("300"),
                wb_commission=Decimal("0"),
                logistics=Decimal("0"),
                storage=Decimal("0"),
                acceptance=Decimal("0"),
                acquiring=Decimal("0"),
                cogs_from_1c_with_extra_costs=Decimal("100"),
                gross_profit=Decimal("200"),
                margin=Decimal("0.6667"),
                profit_per_unit=Decimal("200"),
                data_quality_status=DataQualityStatus.RELIABLE,
                methodology_version="test",
                source_snapshot_hashes=("hash-border",),
            )
        ],
        onec_report_reconciliation_rows=[
            OnecReportReconciliationRow(
                client_id=CLIENT_ID,
                seller_account_id="WB_ACCOUNT_1",
                organization_id="1C_ORG_1",
                document_date=date(2026, 5, 3),
                week_start=date(2026, 4, 27),
                week_end=date(2026, 5, 3),
                document_kind=OnecReportKind.COMMISSIONER_REPORT,
                document_label="Отчет комиссионера",
                wb_report_ids=("SUMMARY-BORDER",),
                sales_quantity=Decimal("1"),
                return_quantity=Decimal("0"),
                quantity=Decimal("1"),
                sales_amount=Decimal("300"),
                return_amount=Decimal("0"),
                revenue_after_spp=Decimal("300"),
                net_revenue=Decimal("300"),
                wb_commission=Decimal("0"),
                logistics=Decimal("0"),
                storage=Decimal("0"),
                acceptance=Decimal("0"),
                acquiring=Decimal("0"),
                cogs_from_1c_with_extra_costs=Decimal("100"),
                gross_profit=Decimal("200"),
                margin=Decimal("0.6667"),
                data_quality_status=DataQualityStatus.RELIABLE,
                source_row_count=1,
            )
        ],
        onec_report_product_rows=[
            OnecReportProductRow(
                client_id=CLIENT_ID,
                seller_account_id="WB_ACCOUNT_1",
                organization_id="1C_ORG_1",
                document_date=date(2026, 5, 3),
                week_start=date(2026, 4, 27),
                week_end=date(2026, 5, 3),
                document_kind=OnecReportKind.COMMISSIONER_REPORT,
                document_label="Отчет комиссионера",
                wb_report_ids=("SUMMARY-BORDER",),
                nm_id=101,
                vendor_code="A-1",
                barcode="111",
                onec_item_id="ONEC-1",
                sales_model=SalesModel.FBO,
                sales_quantity=Decimal("1"),
                return_quantity=Decimal("0"),
                quantity=Decimal("1"),
                sales_amount=Decimal("300"),
                return_amount=Decimal("0"),
                revenue_after_spp=Decimal("300"),
                net_revenue=Decimal("300"),
                wb_commission=Decimal("0"),
                logistics=Decimal("0"),
                storage=Decimal("0"),
                acceptance=Decimal("0"),
                acquiring=Decimal("0"),
                cogs_from_1c_with_extra_costs=Decimal("100"),
                gross_profit=Decimal("200"),
                margin=Decimal("0.6667"),
                profit_per_unit=Decimal("200"),
                data_quality_status=DataQualityStatus.RELIABLE,
                source_row_count=1,
                source_snapshot_hashes=("hash-border",),
            )
        ],
    )
    output = build_excel_report(
        report,
        tmp_path / "actual-document-date-report.xlsx",
        onec_gross_profit_rows=[
            OnecGrossProfitDocumentRow(
                client_id=CLIENT_ID,
                organization_id="1C_ORG_1",
                counterparty_id="RWB",
                document_id="DOC-BORDER",
                document_type="ОтчетКомиссионера",
                document_number="ОК-000042",
                input_number="SUMMARY-BORDER",
                document_date=date(2026, 4, 30),
                week_start=date(2026, 4, 27),
                week_end=date(2026, 5, 3),
                quantity="1",
                revenue="300",
                vat="14.29",
                cogs="100",
                gross_profit="200",
                external_report_id="SUMMARY-BORDER",
                settlement_total="250",
                source_row_count=1,
            )
        ],
        wb_sales_report_summary_rows=[
            WbSalesReportSummaryRow(
                client_id=CLIENT_ID,
                seller_account_id="WB_ACCOUNT_1",
                account_name="WB_ACCOUNT_1",
                report_id="SUMMARY-BORDER",
                date_from=date(2026, 4, 27),
                date_to=date(2026, 5, 3),
                create_date=date(2026, 5, 3),
                report_type=1,
                retail_amount_sum="300",
                for_pay_sum="250",
                delivery_service_sum="0",
                paid_storage_sum="0",
                deduction_sum="0",
                penalty_sum="0",
                cashback_discount_sum="0",
                bank_payment_sum="250",
                raw_payload_hash="summary-border-hash",
            )
        ],
    )

    workbook = load_workbook(output, data_only=True)
    products = workbook["Товары по отчетам 1С"]
    assert products["A2"].value == "2026-04-30"

    reconciliation = workbook["Сверка с 1С ОПиУ"]
    header_row = next(
        row_idx
        for row_idx in range(1, reconciliation.max_row + 1)
        if reconciliation.cell(row_idx, 1).value == "Месяц"
    )
    monthly_rows = {
        row[0].value: row
        for row in reconciliation.iter_rows(min_row=header_row + 1, max_col=8)
        if row[0].value
    }
    assert monthly_rows["Апрель 2026"][1].value == 1
    assert monthly_rows["Апрель 2026"][2].value == 1
    assert monthly_rows["Апрель 2026"][4].value == 100
    assert monthly_rows["Апрель 2026"][5].value == 100
    assert monthly_rows["Апрель 2026"][6].value == 0


def test_buyout_return_quantity_is_diagnostic_when_invoice_sales_match() -> None:
    expected = SimpleNamespace(
        seller_account_id="WB_ACCOUNT_1",
        organization_id="1C_ORG_1",
        week_start=date(2026, 4, 6),
        week_end=date(2026, 4, 12),
        document_date=date(2026, 4, 6),
        document_label="Уведомление о выкупе",
        sales_quantity=Decimal("195"),
        return_quantity=Decimal("11"),
        quantity=Decimal("184"),
        revenue_after_spp=Decimal("85079.99"),
        wb_report_ids=("535699202604061",),
    )
    actual = OnecGrossProfitDocumentRow(
        client_id=CLIENT_ID,
        organization_id="1C_ORG_1",
        counterparty_id="RWB",
        document_id="DOC-BUYOUT-1",
        document_type="РасходнаяНакладная",
        document_date=date(2026, 4, 6),
        week_start=date(2026, 4, 6),
        week_end=date(2026, 4, 12),
        quantity="195",
        revenue="51532.81",
        vat="2453.94",
        cogs="20000",
        gross_profit="31532.81",
        external_report_id="685214500",
        source_row_count=22,
    )
    summary = WbSalesReportSummaryRow(
        client_id=CLIENT_ID,
        seller_account_id="WB_ACCOUNT_1",
        account_name="WB_ACCOUNT_1",
        report_id="685214500",
        date_from=date(2026, 4, 6),
        date_to=date(2026, 4, 12),
        create_date=date(2026, 4, 6),
        report_type=2,
        retail_amount_sum="85079.99",
        for_pay_sum="68904.85",
        delivery_service_sum="0",
        paid_storage_sum="0",
        deduction_sum="0",
        penalty_sum="0",
        cashback_discount_sum="0",
        bank_payment_sum="43769.73",
        raw_payload_hash="summary-buyout-hash",
    )

    row = _document_reconciliation_row(
        expected,
        [actual],
        [summary],
        report_period_end=date(2026, 6, 17),
        sales_summaries=[],
        buyout_summaries=[summary],
        account_labels=None,
        organization_labels=None,
    )

    assert row["status"] == "Документ найден"
    assert row["expected_quantity"] == Decimal("195")
    assert row["onec_quantity"] == Decimal("195")
    assert row["quantity_delta"] == Decimal("0")
    assert row["return_quantity_delta"] == Decimal("11")
    assert "Возвраты WB из выкупного отчета показаны справочно" in row["comment"]


def test_buyout_sales_quantity_mismatch_requires_review() -> None:
    expected = SimpleNamespace(
        seller_account_id="WB_ACCOUNT_1",
        organization_id="1C_ORG_1",
        week_start=date(2026, 4, 6),
        week_end=date(2026, 4, 12),
        document_date=date(2026, 4, 6),
        document_label="Уведомление о выкупе",
        sales_quantity=Decimal("195"),
        return_quantity=Decimal("11"),
        quantity=Decimal("184"),
        revenue_after_spp=Decimal("85079.99"),
        wb_report_ids=("535699202604061",),
    )
    actual = OnecGrossProfitDocumentRow(
        client_id=CLIENT_ID,
        organization_id="1C_ORG_1",
        counterparty_id="RWB",
        document_id="DOC-BUYOUT-1",
        document_type="РасходнаяНакладная",
        document_date=date(2026, 4, 6),
        week_start=date(2026, 4, 6),
        week_end=date(2026, 4, 12),
        quantity="62",
        revenue="51532.81",
        vat="2453.94",
        cogs="20000",
        gross_profit="31532.81",
        external_report_id="685214500",
        source_row_count=22,
    )
    summary = WbSalesReportSummaryRow(
        client_id=CLIENT_ID,
        seller_account_id="WB_ACCOUNT_1",
        account_name="WB_ACCOUNT_1",
        report_id="685214500",
        date_from=date(2026, 4, 6),
        date_to=date(2026, 4, 12),
        create_date=date(2026, 4, 6),
        report_type=2,
        retail_amount_sum="85079.99",
        for_pay_sum="68904.85",
        delivery_service_sum="0",
        paid_storage_sum="0",
        deduction_sum="0",
        penalty_sum="0",
        cashback_discount_sum="0",
        bank_payment_sum="43769.73",
        raw_payload_hash="summary-buyout-hash",
    )

    row = _document_reconciliation_row(
        expected,
        [actual],
        [summary],
        report_period_end=date(2026, 6, 17),
        sales_summaries=[],
        buyout_summaries=[summary],
        account_labels=None,
        organization_labels=None,
    )

    assert row["status"] == "Проверить количество"
    assert row["expected_quantity"] == Decimal("195")
    assert row["onec_quantity"] == Decimal("62")
    assert row["quantity_delta"] == Decimal("133")
    assert "количество расходной накладной сверяется с WB продажами" in row["comment"]


def test_commissioner_amount_uses_weekly_retail_control_sum() -> None:
    expected = SimpleNamespace(
        seller_account_id="WB_ACCOUNT_1",
        organization_id="1C_ORG_1",
        week_start=date(2026, 4, 6),
        week_end=date(2026, 4, 12),
        document_date=date(2026, 4, 12),
        document_label="Отчет комиссионера",
        sales_quantity=Decimal("644"),
        return_quantity=Decimal("58"),
        quantity=Decimal("586"),
        revenue_after_spp=Decimal("714787.16"),
        wb_report_ids=("439356720260406",),
    )
    actual = OnecGrossProfitDocumentRow(
        client_id=CLIENT_ID,
        organization_id="1C_ORG_1",
        counterparty_id="RWB",
        document_id="DOC-COMMISSIONER-1",
        document_type="ОтчетКомиссионера",
        document_date=date(2026, 4, 12),
        week_start=date(2026, 4, 6),
        week_end=date(2026, 4, 12),
        sales_quantity="777",
        return_quantity="66",
        quantity="711",
        revenue="861806.16",
        vat="41038.39",
        cogs="300000",
        gross_profit="561806.16",
        external_report_id="685214499",
        source_row_count=258,
    )
    summary = WbSalesReportSummaryRow(
        client_id=CLIENT_ID,
        seller_account_id="WB_ACCOUNT_1",
        account_name="WB_ACCOUNT_1",
        report_id="685214499",
        date_from=date(2026, 4, 6),
        date_to=date(2026, 4, 12),
        create_date=date(2026, 4, 12),
        report_type=1,
        retail_amount_sum="861806.16",
        for_pay_sum="809205.78",
        delivery_service_sum="0",
        paid_storage_sum="0",
        deduction_sum="0",
        penalty_sum="0",
        cashback_discount_sum="0",
        bank_payment_sum="343889.53",
        raw_payload_hash="summary-commissioner-hash",
    )

    row = _document_reconciliation_row(
        expected,
        [actual],
        [summary],
        report_period_end=date(2026, 6, 17),
        sales_summaries=[summary],
        buyout_summaries=[],
        account_labels=None,
        organization_labels=None,
    )

    assert row["status"] == "Проверить товарные строки"
    assert row["expected_amount"] == Decimal("861806.16")
    assert row["amount_delta"] == Decimal("0.00")
    assert "retailAmountSum" in row["comment"]


def test_excel_report_shows_sales_register_cost_method(tmp_path) -> None:
    sales_cost = OnecUnfCostSnapshot(
        client_id=CLIENT_ID,
        organization_id="1C_ORG_1",
        loaded_at=datetime(2026, 6, 16, 10, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        onec_item_id="ONEC-1",
        article="A-1",
        barcode="111",
        name="Product 1",
        cost_value="100",
        extra_costs_value="0",
        cost_method="sales_register_weighted_average_allocated_extra_costs",
        effective_from=date(2026, 1, 1),
        source_document="AccumulationRegister_Продажи",
        raw_payload_hash="sales-register-cost-hash",
    )
    report = build_unit_economics_report(
        client_id=CLIENT_ID,
        wb_snapshots=[wb_snapshots()[0]],
        cost_snapshots=[sales_cost],
        sku_mappings=[sku_mappings()[0]],
        account_org_mapping=account_org_mapping(),
        generated_at=datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        as_of_date=date(2026, 6, 16),
    )
    output = build_excel_report(
        report,
        tmp_path / "sales-register-report.xlsx",
        cost_snapshots=[sales_cost],
    )

    workbook = load_workbook(output, data_only=True)
    costs = workbook["Себестоимость 1С"]
    assert costs["H1"].value == "Допрасходы отдельно"
    assert costs["H2"].value == 0
    assert costs["J2"].value == "sales_register_weighted_average_allocated_extra_costs"
    assert costs["L2"].value == "AccumulationRegister_Продажи"


def test_excel_report_explains_nearest_available_cost(tmp_path) -> None:
    sales_cost = OnecUnfCostSnapshot(
        client_id=CLIENT_ID,
        organization_id="1C_ORG_1",
        loaded_at=datetime(2026, 6, 16, 10, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        onec_item_id="ONEC-1",
        article="A-1",
        barcode="111",
        name="Product 1",
        cost_value="100",
        extra_costs_value="0",
        cost_method="sales_register_weighted_average_allocated_extra_costs",
        effective_from=date(2026, 4, 6),
        effective_to=date(2026, 4, 12),
        source_document="AccumulationRegister_Продажи",
        raw_payload_hash="sales-register-cost-hash",
    )
    snapshot = wb_snapshots()[0].model_copy(
        update={
            "period_start": date(2026, 5, 20),
            "period_end": date(2026, 5, 20),
        }
    )
    report = build_unit_economics_report(
        client_id=CLIENT_ID,
        wb_snapshots=[snapshot],
        cost_snapshots=[sales_cost],
        sku_mappings=[sku_mappings()[0]],
        account_org_mapping=account_org_mapping(),
        generated_at=datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        as_of_date=date(2026, 6, 16),
    )
    output = build_excel_report(
        report,
        tmp_path / "nearest-cost-report.xlsx",
        cost_snapshots=[sales_cost],
        sku_mappings=[sku_mappings()[0]],
    )

    workbook = load_workbook(output, data_only=True)
    unit_economics = workbook["Юнит экономика"]
    headers = [cell.value for cell in unit_economics[1]]
    idx = {header: index for index, header in enumerate(headers)}

    assert unit_economics.cell(2, idx["Себестоимость 1С"] + 1).value == 200
    assert (
        unit_economics.cell(2, idx["Статус данных"] + 1).value
        == "Себестоимость 1С требует сверки"
    )
    assert (
        unit_economics.cell(2, idx["Причина статуса"] + 1).value
        == "Себестоимость взята из ближайшей доступной недели 1С; "
        "нужна сверка после закрытия месяца"
    )


def test_excel_dashboard_shows_source_warnings(tmp_path) -> None:
    report = build_unit_economics_report(
        client_id=CLIENT_ID,
        wb_snapshots=[wb_snapshots()[0]],
        cost_snapshots=cost_snapshots(),
        sku_mappings=sku_mappings(),
        account_org_mapping=account_org_mapping(),
        generated_at=datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        as_of_date=date(2026, 6, 16),
    )
    output = build_excel_report(
        report,
        tmp_path / "source-warning-report.xlsx",
        source_notes=[
            "WB кабинет Организация Минзифа загружен не полностью: ошибка HTTP 429.",
            "Себестоимость 1С берется из регистра Продажи.",
        ],
    )

    workbook = load_workbook(output, data_only=True)
    dashboard = workbook["Дашборд"]

    assert dashboard["G2"].value == "Проблемы источников"
    assert dashboard["G3"].value == "Что проверить"
    assert "HTTP 429" in dashboard["G4"].value


def test_excel_dashboard_top_profit_excludes_invalid_costs(tmp_path) -> None:
    generated_at = datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow"))

    def row(
        *,
        vendor_code: str,
        cogs: str,
        profit_after_taxes: str,
        profit_after_taxes_per_unit: str,
        data_quality_status: DataQualityStatus = DataQualityStatus.RELIABLE,
    ) -> UnitEconomicsRow:
        return UnitEconomicsRow(
            client_id=CLIENT_ID,
            seller_account_id="WB_ACCOUNT_1",
            organization_id="1C_ORG_1",
            week_start=date(2026, 4, 6),
            week_end=date(2026, 4, 12),
            is_partial_week=False,
            nm_id=1000,
            vendor_code=vendor_code,
            barcode=f"barcode-{vendor_code}",
            sales_model=SalesModel.FBO,
            quantity=Decimal("1"),
            sales_quantity=Decimal("1"),
            return_quantity=Decimal("0"),
            revenue_before_spp=Decimal("1000"),
            revenue_after_spp=Decimal("1000"),
            net_revenue=Decimal("1000"),
            wb_commission=Decimal("0"),
            logistics=Decimal("0"),
            storage=Decimal("0"),
            acceptance=Decimal("0"),
            acquiring=Decimal("0"),
            cogs_from_1c_with_extra_costs=Decimal(cogs),
            gross_profit=Decimal("1000") - Decimal(cogs),
            profit_after_taxes=Decimal(profit_after_taxes),
            margin=Decimal("0.9"),
            margin_after_taxes=Decimal("0.9"),
            profit_per_unit=Decimal(profit_after_taxes_per_unit),
            profit_after_taxes_per_unit=Decimal(profit_after_taxes_per_unit),
            data_quality_status=data_quality_status,
            methodology_version="test",
            source_snapshot_hashes=(f"hash-{vendor_code}",),
        )

    report = UnitEconomicsReport(
        client_id=CLIENT_ID,
        report_period_start=date(2026, 4, 1),
        report_period_end=date(2026, 4, 30),
        generated_at=generated_at,
        status=ReportStatus.FINAL,
        methodology_version="test",
        rows=[
            row(
                vendor_code="GOOD",
                cogs="100",
                profit_after_taxes="900",
                profit_after_taxes_per_unit="900",
            ),
            row(
                vendor_code="REPORT-TYPE-FALLBACK",
                cogs="120",
                profit_after_taxes="880",
                profit_after_taxes_per_unit="880",
                data_quality_status=DataQualityStatus.REPORT_TYPE_FALLBACK,
            ),
            row(
                vendor_code="ZERO-COST",
                cogs="0",
                profit_after_taxes="1000",
                profit_after_taxes_per_unit="1000",
            ),
            row(
                vendor_code="NEGATIVE-SALE",
                cogs="-50",
                profit_after_taxes="1050",
                profit_after_taxes_per_unit="1050",
            ),
        ],
    )
    output = build_excel_report(report, tmp_path / "top-profit-report.xlsx")

    workbook = load_workbook(output, data_only=True)
    dashboard = workbook["Дашборд"]
    section_row = next(
        cell.row
        for (cell,) in dashboard.iter_rows(min_col=1, max_col=1)
        if cell.value == "Топ прибыльных товаров (себестоимость > 0)"
    )
    header_values = [
        dashboard.cell(section_row + 1, col_idx).value for col_idx in range(1, 10)
    ]
    assert header_values[-1] == "Статус данных"
    top_names = [
        dashboard.cell(row_idx, 1).value
        for row_idx in range(section_row + 2, section_row + 12)
        if dashboard.cell(row_idx, 1).value
    ]

    assert top_names == ["GOOD", "REPORT-TYPE-FALLBACK"]


def test_expense_invoice_reconciliation_uses_its_own_week() -> None:
    row = OnecGrossProfitDocumentRow(
        client_id=CLIENT_ID,
        organization_id="1C_ORG_1",
        counterparty_id="RWB",
        document_id="DOC-BUYOUT-1",
        document_type="РасходнаяНакладная",
        document_date=date(2026, 4, 13),
        week_start=date(2026, 4, 13),
        week_end=date(2026, 4, 19),
        quantity="77",
        revenue="77863.77",
        vat="0",
        cogs="24892.60",
        gross_profit="52971.17",
        source_row_count=12,
    )

    assert _expected_sales_week_for_onec_document(row) == date(2026, 4, 13)


def test_excel_report_does_not_contain_secret_markers(tmp_path) -> None:
    report = build_unit_economics_report(
        client_id=CLIENT_ID,
        wb_snapshots=wb_snapshots(),
        cost_snapshots=cost_snapshots(),
        sku_mappings=sku_mappings(),
        account_org_mapping=account_org_mapping(),
        generated_at=datetime(2026, 6, 16, 12, 0, tzinfo=ZoneInfo("Europe/Moscow")),
        as_of_date=date(2026, 6, 16),
    )
    output = build_excel_report(report, tmp_path / "report.xlsx")
    with ZipFile(output) as archive:
        content = "\n".join(
            archive.read(name).decode("utf-8", errors="ignore")
            for name in archive.namelist()
            if name.endswith(".xml")
        )
    forbidden = [".env", "API_KEY", "webhook", "WB_ACCOUNT_1_API_KEY"]
    assert not any(marker in content for marker in forbidden)
