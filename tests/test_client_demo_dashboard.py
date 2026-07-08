from pathlib import Path

from openpyxl import Workbook, load_workbook

from scripts.build_client_demo_dashboard import (
    _row_month,
    build_dashboard,
    collect_dashboard_data,
)
from scripts.build_power_query_client_package import build_power_query_client_package
from scripts.export_power_bi_marts import export_power_bi_marts


def _create_demo_workbook(path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    dashboard = workbook.create_sheet("Дашборд")
    dashboard.append(
        [
            "Дашборд юнит-экономики",
            None,
            None,
            "Период: 01.04.2026 - 17.06.2026",
        ]
    )
    dashboard.append(["Ключевые показатели"])
    dashboard.append(["Показатель", "Значение"])
    dashboard.append(["Выручка до СПП", 101000])
    dashboard.append(["СПП", 1000])
    dashboard.append(["Выручка после СПП", 100000])
    dashboard.append(["Продажи, шт", 10])
    dashboard.append(["Возвраты, шт", 2])
    dashboard.append(["Чистое кол-во, шт", 8])
    dashboard.append(["Маржинальный доход WB после налогов", 15000])

    dynamics = workbook.create_sheet("Динамика")
    dynamics.append(["Динамика месяц к месяцу"])
    dynamics.append(
        [
            "Месяц",
            "Статус",
            "Продажи, шт",
            "Возвраты, шт",
            "% возвратов",
            "Выручка",
            "Логистика",
            "Расходы WB",
            "Маржинальный доход WB после налогов",
            "Маржа после налогов",
        ]
    )
    dynamics.append(
        [
            "Апрель 2026",
            "полный месяц",
            10,
            2,
            0.2,
            100000,
            7000,
            12000,
            15000,
            0.15,
        ]
    )
    dynamics.append(["Май 2026", "полный месяц", 0, 0, None, 0, 0, 0, 0, None])
    dynamics.append(
        ["Июнь 2026 (неполный месяц)", "неполный месяц", 0, 0, None, 0, 0, 0, 0, None]
    )
    dynamics.append([])
    dynamics.append(["Изменение м/м"])
    dynamics.append(["Период", "Выручка после СПП, Δ", "Выручка после СПП, %"])
    dynamics.append(["Май 2026 к Апрель 2026", -100000, -1])

    expenses = workbook.create_sheet("Расходы WB")
    expenses.append(["Структура расходов"])
    expenses.append(
        [
            "Статья",
            "Сумма",
            "% от выручки",
            "Апрель",
            "Май",
            "Июнь",
            "Май к апрелю",
            "Июнь к маю",
        ]
    )
    expenses.append(["Себестоимость 1С", 40000, 0.4, 40000, 0, 0, -40000, 0])
    expenses.append(["Логистика WB", 7000, 0.07, 7000, 0, 0, -7000, 0])

    returns = workbook.create_sheet("Возвраты")
    returns.append(
        [
            "Неделя",
            "Документ-отчет",
            "Организация 1С",
            "Кабинет WB",
            "Товар",
            "nmId WB",
            "Артикул WB",
            "Артикул 1С",
            "Баркод",
            "Продажи, шт",
            "Возвраты, шт",
            "% возвратов",
            "Сумма возвратов",
            "Прибыль после налогов",
            "Прибыль после налогов на шт",
            "Статус данных",
            "Главная причина",
        ]
    )
    returns.append(
        [
            "2026-04-01",
            "Отчет комиссионера · 01.04.2026-05.04.2026 · закрытие 05.04.2026",
            "Организация",
            "WB_ACCOUNT_1",
            "Товар A",
            123,
            "ART-A",
            "A-1",
            "460000000001",
            10,
            2,
            0.2,
            20000,
            -3000,
            -375,
            "ОК",
            "Возвраты + логистика",
        ]
    )

    lost = workbook.create_sheet("Упущенные продажи")
    for _ in range(13):
        lost.append([])
    lost.append(
        [
            "Товар",
            "Артикул 1С",
            "Баркод",
            "Кабинет WB",
            "Дней без остатка WB",
            "Продажи, шт",
            "Потенциально упущено, шт",
            "Упущенная выручка",
            "Упущенная прибыль",
            "Ограничение",
        ]
    )
    lost.append(
        [
            "Товар A",
            "A-1",
            "460000000001",
            "WB_ACCOUNT_1",
            5,
            10,
            2.5,
            25000,
            6000,
            "Сверить с 1С",
        ]
    )

    reconciliation = workbook.create_sheet("Сверка с 1С ОПиУ")
    reconciliation.append(["Сверка с 1С ОПиУ"])
    reconciliation.append([])
    reconciliation.append([])
    reconciliation.append(
        ["Показатель", "WB-витрина", "1С/ОПиУ", "Дельта", "Комментарий"]
    )
    reconciliation.append(
        [
            "Выручка после СПП",
            100000,
            None,
            None,
            "Контрольная строка внутри WB-методики.",
        ]
    )
    reconciliation.append(
        ["Главная сверка по месяцам: количество, себестоимость 1С и расходы МП"]
    )
    reconciliation.append(
        [
            "Месяц",
            "WB количество",
            "1С количество",
            "Дельта количества",
            "Себестоимость 1С в WB-расчете",
            "Себестоимость по валовой прибыли 1С",
            "Дельта себестоимости",
            "WB расходы МП",
            "1С расходы МП",
            "Дельта расходов МП",
            "Комментарий",
        ]
    )
    reconciliation.append(
        [
            "Апрель 2026",
            98,
            100,
            -2,
            40000,
            42000,
            -2000,
            12000,
            11000,
            1000,
            "Тестовая сверка.",
        ]
    )
    reconciliation.append(
        [
            "Итого",
            98,
            100,
            -2,
            40000,
            42000,
            -2000,
            12000,
            11000,
            1000,
            "Итог по месяцам.",
        ]
    )
    reconciliation.append([])
    reconciliation.append(["Справочный блок: выручка, СПП, статьи РВБ и прибыль"])
    reconciliation.append(
        ["Показатель", "WB-витрина", "1С/ОПиУ", "Дельта", "Комментарий"]
    )
    reconciliation.append(["% СПП", 0.01, None, None, "Доля СПП."])
    reconciliation.append(
        ["Выручка 1С/ОПиУ всего", None, 57410238.38, None, "Справочно."]
    )
    reconciliation.append(["Выручка без НДС", 50442679.37, None, None, "Справочно."])

    unit = workbook.create_sheet("Юнит экономика")
    unit.append(
        [
            "Неделя",
            "Организация 1С",
            "Кабинет WB",
            "Товар",
            "nmId WB",
            "Артикул WB",
            "Артикул 1С",
            "Баркод",
            "Продажи, шт",
            "Возвраты, шт",
            "Выручка после СПП",
            "Сумма возвратов",
            "Себестоимость 1С",
            "Комиссия WB",
            "Логистика WB",
            "Хранение WB",
            "Приемка WB",
            "Продвижение WB",
            "Штрафы/доплаты WB",
            "Эквайринг WB",
            "НДС",
            "Налог с выручки",
            "Маржинальный доход WB после налогов",
            "Статус данных",
        ]
    )
    unit.append(
        [
            "2026-04-01",
            "Организация",
            "WB_ACCOUNT_1",
            "Товар A",
            123,
            "ART-A",
            "A-1",
            "460000000001",
            10,
            2,
            100000,
            20000,
            40000,
            5000,
            7000,
            1000,
            0,
            0,
            0,
            1200,
            5000,
            1000,
            -3000,
            "ОК",
        ]
    )

    workbook.save(path)


def test_build_client_demo_dashboard_from_visible_sheets(tmp_path: Path) -> None:
    workbook_path = tmp_path / "demo.xlsx"
    output_path = tmp_path / "dashboard.html"
    _create_demo_workbook(workbook_path)

    payload = collect_dashboard_data(workbook_path)
    assert payload["views"]["Все кабинеты"]["kpis"]["sales"] == 10
    assert payload["views"]["Все кабинеты"]["kpis"]["lost_products"] == 1
    assert [row["month"] for row in payload["views"]["Все кабинеты"]["monthly"]] == [
        "Апрель 2026",
        "Май 2026",
        "Июнь 2026 (неполный месяц)",
    ]
    assert (
        payload["views"]["WB_ACCOUNT_1"]["top_losses"][0]["barcode"] == "460000000001"
    )
    assert (
        payload["views"]["Все кабинеты"]["reconciliation_monthly"][0][
            "mp_expenses_delta"
        ]
        == 1000
    )
    reconciliation_months = [
        row["month"]
        for row in payload["views"]["Все кабинеты"]["reconciliation_monthly"]
    ]
    assert reconciliation_months == ["Апрель 2026", "Итого"]
    assert "% СПП" not in reconciliation_months
    assert "Выручка 1С/ОПиУ всего" not in reconciliation_months
    assert "Выручка без НДС" not in reconciliation_months

    build_dashboard(workbook_path, output_path, logo_path=None)
    html = output_path.read_text(encoding="utf-8")

    assert "Демо-дашборд юнит-экономики WB" in html
    assert "Упущенные продажи" in html
    assert "Сверка с 1С" in html
    assert "одинаковой выборке РВБ-документов" in html
    assert "1С себестоимость здесь берется из ОПиУ/регистра ДоходыИРасходы" not in html
    assert "5 200 476,68 ₽" not in html
    assert "Причина возврата не передается текущими источниками" in html
    assert "Сейчас собран скелет аналитической архитектуры" not in html
    assert ".env" not in html


def test_dashboard_month_filter_uses_onec_week_closing_month() -> None:
    assert _row_month("2026-03-30") == "Апрель 2026"
    assert _row_month("2026-04-27") == "Апрель 2026"
    assert _row_month("2026-05-04") == "Май 2026"


def test_export_power_bi_marts_from_client_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "demo.xlsx"
    output_dir = tmp_path / "marts"
    _create_demo_workbook(workbook_path)

    paths = export_power_bi_marts(workbook_path, output_dir)
    names = {path.name for path in paths}

    assert "monthly_dynamics.csv" in names
    assert "unit_economics.csv" in names
    assert "onec_opiu_reconciliation.csv" in names
    assert (output_dir / "README.md").exists()
    assert ".env" not in (output_dir / "README.md").read_text(encoding="utf-8")


def test_build_power_query_client_package_from_client_workbook(tmp_path: Path) -> None:
    workbook_path = tmp_path / "demo.xlsx"
    output_dir = tmp_path / "power_query"
    _create_demo_workbook(workbook_path)

    paths = build_power_query_client_package(
        workbook_path=workbook_path,
        output_dir=output_dir,
        database="test_unit_economics",
        schema="bi",
        host="localhost",
        port=55433,
        user="",
        publish_postgres=False,
    )
    names = {path.name for path in paths}

    assert "unit_economics.pq" in names
    assert "README.md" in names
    template_paths = [path for path in paths if path.suffix == ".xlsx"]
    assert template_paths

    workbook = load_workbook(template_paths[0], read_only=True, data_only=True)
    try:
        assert "Инструкция" in workbook.sheetnames
        assert "Параметры подключения" in workbook.sheetnames
    finally:
        workbook.close()

    unit_query = (output_dir / "power_query_m" / "unit_economics.pq").read_text(
        encoding="utf-8"
    )
    assert 'fxLoadView("unit_economics")' in unit_query
    assert '{"c008", "Баркод"}' in unit_query
    assert "Table.RenameColumns" in unit_query
