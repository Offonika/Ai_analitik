from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from wb_unit_economics.onec_odata import (
    OnecODataMetadataCheckResult,
    OnecODataSettings,
    OnecSampleExportResult,
)
from wb_unit_economics.web import repository
from wb_unit_economics.web.app import create_app
from wb_unit_economics.web.database import init_db, make_engine, make_session_factory
from wb_unit_economics.web.models import (
    AuditEvent,
    OrganizationTaxProfile,
    ReportRun,
    SourceRefreshCollection,
    SourceRefreshRun,
    SourceSnapshotRow,
    User,
)
from wb_unit_economics.web.reports.builders import (
    build_month_close_control_payload,
    build_tax_load_payload,
    canonical_payload_sha256,
    fns_paid_taxes_numerator,
    fns_tax_burden_ratio,
)
from wb_unit_economics.web.reports.evidence import (
    AccountingEvidenceSource,
    _bank_tax_payments,
    _date_text,
    _source_gap_issues,
    materialize_accounting_evidence,
)
from wb_unit_economics.web.reports.excel import (
    MONTH_CLOSE_SHEETS,
    TAX_LOAD_SHEETS,
    write_scenario_excel,
)
from wb_unit_economics.web.settings import WebSettings
from wb_unit_economics.web.source_refresh import SourceCredentials, SourceRefreshService


def _report(kind: str, organization_id: str = "ORG-1") -> SimpleNamespace:
    return SimpleNamespace(
        id=f"report-{kind}",
        tenant_id="tenant-a",
        client_id="tenant-a",
        report_kind=kind,
        organization_id=organization_id,
        period_start=date(2026, 1, 1),
        period_end=date(2026, 6, 30),
        methodology_version=f"{kind}-v1",
        source_snapshot_set_id="snapshot-1",
        generated_at=repository.security.utcnow(),
        publication_status="draft",
    )


def _month_close_evidence() -> dict:
    return {
        "organizationId": "ORG-1",
        "sourceCoverage": [
            {
                "sourceKind": "onec_osv",
                "periodStart": "2026-06-01",
                "periodEnd": "2026-06-30",
                "status": "loaded",
                "snapshotId": "snapshot-1",
            }
        ],
        "controls": [
            {
                "controlCode": "osv",
                "section": "ОСВ",
                "title": "Сверка ОСВ",
                "status": "confirmed",
                "sourceKind": "onec_osv",
                "evidenceStatus": "loaded",
                "issueCode": None,
                "nextAction": "Проверить дельту.",
            }
        ],
        "osvBalanceAndTurnovers": {
            "status": "loaded",
            "rows": [
                {
                    "accountCode": "10",
                    "accountName": "Материалы",
                    "openingDebit": "100",
                    "openingCredit": "0",
                    "debitTurnover": "40",
                    "creditTurnover": "10",
                    "closingDebit": "130",
                    "closingCredit": "0",
                },
                {
                    "accountCode": "10.01",
                    "accountName": "Сырье",
                    "openingDebit": None,
                    "openingCredit": "0",
                    "debitTurnover": "20",
                    "creditTurnover": "5",
                    "closingDebit": "15",
                    "closingCredit": "0",
                },
            ],
        },
        "osvRecordTypeFallback": {
            "status": "loaded",
            "rows": [{"accountCode": "99", "closingDebit": "999"}],
        },
        "osvReferenceRows": [
            {
                "accountCode": "10",
                "openingDebit": "100",
                "openingCredit": "0",
                "debitTurnover": "40",
                "creditTurnover": "9",
                "closingDebit": "130",
                "closingCredit": "0",
            },
            {
                "accountCode": "10.01",
                "openingDebit": "5",
                "openingCredit": "0",
                "debitTurnover": "20",
                "creditTurnover": "5",
                "closingDebit": "15",
                "closingCredit": "0",
            },
        ],
    }


def _tax_evidence() -> dict:
    return {
        "organizationId": "ORG-1",
        "incomeEvidence": {
            "value": "1000",
            "status": "confirmed",
            "sourceKind": "onec_official_financial_results",
        },
        "sourceCoverage": [
            {
                "sourceKind": "onec_tax",
                "periodStart": "2026-01-01",
                "periodEnd": "2026-06-30",
                "status": "loaded",
                "snapshotId": "snapshot-1",
            }
        ],
        "taxRows": [
            {
                "taxCode": "profit_tax",
                "taxName": "Налог на прибыль",
                "paid": "100",
                "balance": "25",
                "dueDate": "2026-07-28",
                "paymentKind": "own_tax",
                "includedInFnsTaxBurden": True,
                "evidenceStatus": "loaded",
            },
            {
                "taxCode": "ndfl_agent",
                "taxName": "НДФЛ налогового агента",
                "paid": "50",
                "paymentKind": "agent_ndfl",
                "includedInFnsTaxBurden": False,
                "exclusionReason": "agent_payment",
                "evidenceStatus": "loaded",
            },
            {
                "taxCode": "insurance",
                "taxName": "Страховые взносы",
                "paid": "25",
                "paymentKind": "insurance_contribution",
                "includedInFnsTaxBurden": False,
                "exclusionReason": "insurance_contribution",
                "evidenceStatus": "loaded",
            },
        ],
    }


def _usn_tax_evidence() -> dict:
    """УСН-разновидность налоговых фактов: собственный налог УСН + взносы."""

    evidence = _tax_evidence()
    evidence["taxRows"] = [
        {
            "taxCode": "usn_income",
            "taxName": "УСН доходы",
            "paid": "100",
            "balance": "0",
            "dueDate": "2026-07-28",
            "paymentKind": "own_tax",
            "includedInFnsTaxBurden": True,
            "evidenceStatus": "loaded",
        },
        {
            "taxCode": "insurance",
            "taxName": "Страховые взносы",
            "paid": "25",
            "paymentKind": "insurance_contribution",
            "includedInFnsTaxBurden": False,
            "exclusionReason": "insurance_contribution",
            "evidenceStatus": "loaded",
        },
    ]
    return evidence


def test_month_close_prefers_balance_and_turnovers_and_warns_on_any_delta() -> None:
    payload = build_month_close_control_payload(
        _report("month_close_control"), _month_close_evidence()
    )

    assert payload["businessRecommendation"] == "review_required"
    assert payload["accountantApproval"] is None
    assert payload["osvSummary"] == {
        "sourceKind": "balance_and_turnovers",
        "sourceStatus": "loaded",
        "reconciliationStatus": "warning",
        "mismatchCount": 1,
    }
    account_10 = next(row for row in payload["osvRows"] if row["accountCode"] == "10")
    assert account_10["creditTurnoverDelta"] == "1"
    assert account_10["reconciliationStatus"] == "warning"
    subaccount = next(
        row for row in payload["osvRows"] if row["accountCode"] == "10.01"
    )
    assert subaccount["openingDebit"] is None
    assert subaccount["openingDebitDelta"] is None
    assert subaccount["reconciliationStatus"] == "missing"
    assert {row["accountCode"] for row in payload["osvRows"]} == {"10", "10.01"}


def test_month_close_uses_fallback_after_source_error_and_does_not_fake_match() -> None:
    evidence = _month_close_evidence()
    evidence["osvBalanceAndTurnovers"] = {
        "status": "source_error",
        "rows": [{"accountCode": "51", "closingDebit": "999"}],
    }
    evidence["osvRecordTypeFallback"] = {
        "status": "loaded",
        "rows": [
            {"accountCode": "51", "debitTurnover": "10"},
            {"accountCode": "51", "debitTurnover": "15"},
        ],
    }
    evidence.pop("osvReferenceRows")

    payload = build_month_close_control_payload(
        _report("month_close_control"), evidence
    )

    assert payload["osvSummary"] == {
        "sourceKind": "record_type_fallback",
        "sourceStatus": "loaded",
        "reconciliationStatus": "not_checked",
        "mismatchCount": None,
    }
    assert payload["osvRows"][0]["debitTurnover"] == "25"
    assert payload["osvRows"][0]["reconciliationStatus"] == "not_checked"


def test_month_close_adds_reference_only_account_as_missing() -> None:
    evidence = _month_close_evidence()
    evidence["osvReferenceRows"].append({"accountCode": "62", "closingDebit": "40"})

    payload = build_month_close_control_payload(
        _report("month_close_control"), evidence
    )

    reference_only = next(
        row for row in payload["osvRows"] if row["accountCode"] == "62"
    )
    assert reference_only["closingDebit"] is None
    assert reference_only["closingDebitDelta"] is None
    assert reference_only["reconciliationStatus"] == "missing"


def test_month_close_materialization_uses_normalized_recordtype_fallback_and_period(
) -> None:
    sources = {
        "onec_accounting_chart": AccountingEvidenceSource(
            source_type="onec_accounting_chart",
            status="loaded",
            snapshot_id="chart-sha",
            rows=(
                {"Ref_Key": "ACC-51", "Code": "51", "Description": "Банк"},
            ),
        ),
        "onec_accounting_balance_and_turnovers": AccountingEvidenceSource(
            source_type="onec_accounting_balance_and_turnovers",
            status="failed",
            snapshot_id="primary-failed",
            rows=(),
        ),
        "onec_accounting_register_balances": AccountingEvidenceSource(
            source_type="onec_accounting_register_balances",
            status="loaded",
            snapshot_id="fallback-sha",
            rows=(
                {
                    "Organization_Key": "ORG-1",
                    "Account_Key": "ACC-51",
                    "OpeningDebit": "100",
                    "OpeningCredit": "0",
                    "DebitTurnover": "50",
                    "CreditTurnover": "20",
                    "ClosingDebit": "130",
                    "ClosingCredit": "0",
                },
            ),
        ),
        "onec_accounting_bank_in": AccountingEvidenceSource(
            source_type="onec_accounting_bank_in",
            status="loaded",
            snapshot_id="bank-sha",
            rows=(
                {
                    "Организация_Key": "ORG-1",
                    "Date": "2026-04-30T12:00:00",
                    "СуммаДокумента": "999",
                },
                {
                    "Организация_Key": "ORG-1",
                    "Date": "2026-05-15T12:00:00",
                    "СуммаДокумента": "50",
                },
                {
                    "Организация_Key": "ORG-2",
                    "Date": "2026-05-15T12:00:00",
                    "СуммаДокумента": "777",
                },
            ),
        ),
    }

    evidence = materialize_accounting_evidence(
        report_kind="month_close_control",
        organization_id="ORG-1",
        period_start=date(2026, 5, 1),
        period_end=date(2026, 5, 31),
        refresh_run_id="generation-1",
        sources=sources,
    )

    assert evidence["osvRecordTypeFallback"]["status"] == "loaded"
    assert evidence["osvRecordTypeFallback"]["rows"][0]["accountCode"] == "51"
    assert evidence["osvRecordTypeFallback"]["rows"][0]["closingDebit"] == "130"
    assert evidence["bankSummary"]["inflow"] == "50"
    assert not any(issue["code"] == "osv_source_gap" for issue in evidence["issues"])


def test_tax_load_formula_excludes_agent_and_insurance_rows() -> None:
    evidence = _tax_evidence()

    assert fns_paid_taxes_numerator(evidence["taxRows"]) == Decimal("100")
    assert fns_tax_burden_ratio("100", "1000") == Decimal("10.0000")
    assert fns_tax_burden_ratio(None, "1000") is None
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "osno",
            "profileStatus": "ready",
            "vatRate": "20",
            "vatMode": "included",
            "vatDeductionMode": "allowed",
            "sourceKind": "1c",
        },
        evidence=evidence,
    )

    summary = payload["taxLoadSummary"]
    assert summary["numeratorValue"] == "100"
    assert summary["fnsTaxBurdenRatio"] == "10.0000"
    assert summary["calculationPeriodKind"] == "preliminary_ytd"
    assert summary["comparisonStatus"] == "pending_methodology_confirmation"
    assert summary["benchmarkValue"] is None
    assert payload["businessStatus"] == "accountant_review_required"
    assert payload["accountantApproval"] is None
    assert payload["contractVersion"] == "tax-load-report-v2"


def test_tax_load_requires_confirmed_classified_numerator_and_denominator() -> None:
    evidence = _tax_evidence()
    evidence["taxRows"][0]["evidenceStatus"] = "partial_source"
    partial = build_tax_load_payload(
        _report("tax_load"), tax_profile={}, evidence=evidence
    )
    assert partial["taxLoadSummary"]["numeratorValue"] is None
    assert partial["taxLoadSummary"]["fnsTaxBurdenRatio"] is None

    evidence = _tax_evidence()
    evidence["taxRows"].append(
        {
            "taxCode": "unknown",
            "taxName": "Не классифицирован",
            "paid": "20",
            "paymentKind": "unclassified",
            "includedInFnsTaxBurden": False,
            "evidenceStatus": "loaded",
        }
    )
    unclassified = build_tax_load_payload(
        _report("tax_load"), tax_profile={}, evidence=evidence
    )
    assert unclassified["taxLoadSummary"]["numeratorValue"] is None

    evidence = _tax_evidence()
    evidence["incomeEvidence"]["status"] = "partial_source"
    denominator = build_tax_load_payload(
        _report("tax_load"), tax_profile={}, evidence=evidence
    )
    assert denominator["taxLoadSummary"]["numeratorValue"] == "100"
    assert denominator["taxLoadSummary"]["denominatorValue"] is None
    assert denominator["taxLoadSummary"]["fnsTaxBurdenRatio"] is None


def test_tax_load_usn_own_tax_counts_in_numerator_and_insurance_excluded() -> None:
    evidence = _usn_tax_evidence()

    # Налог УСН считается собственным налогом наравне с ОСНО; взносы исключены.
    assert fns_paid_taxes_numerator(evidence["taxRows"]) == Decimal("100")
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "usn_income",
            "profileStatus": "ready",
            "revenueTaxRate": "1",
            "vatMode": "none",
            "sourceKind": "1c",
        },
        evidence=evidence,
    )

    summary = payload["taxLoadSummary"]
    assert summary["numeratorValue"] == "100"
    assert summary["denominatorValue"] == "1000"
    assert summary["fnsTaxBurdenRatio"] == "10.0000"
    assert payload["taxProfile"]["taxSystem"] == "usn_income"
    assert payload["taxProfile"]["revenueTaxRate"] == "1"


def test_tax_load_ip_usn_without_financial_results_keeps_ratio_null_without_zero(
) -> None:
    evidence = _usn_tax_evidence()
    # ИП на УСН не составляет отчет о финансовых результатах: доход пришел из
    # поступлений УСН, это неофициальный организационный знаменатель. Коэффициент
    # ФНС не считается, ноль не подставляется, суммы налога УСН показываются.
    evidence["incomeEvidence"] = {
        "value": "5000",
        "status": "confirmed",
        "sourceKind": "usn_income_receipts",
    }

    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "УСН Доходы",
            "profileStatus": "ready",
            "revenueTaxRate": "0.01",
            "sourceKind": "1c",
        },
        evidence=evidence,
    )

    summary = payload["taxLoadSummary"]
    assert summary["numeratorValue"] == "100"
    assert summary["denominatorValue"] is None
    assert summary["fnsTaxBurdenRatio"] is None
    assert payload["businessStatus"] == "preliminary"
    assert any(
        issue["code"] == "fns_ratio_source_gap" for issue in payload["issues"]
    )
    # Суммы налога УСН сохранены в строках отчета.
    assert any(row["taxCode"] == "usn_income" for row in payload["taxRows"])


def test_tax_load_ip_usn_management_ratio_from_receipts_when_no_financial_results(
) -> None:
    evidence = _usn_tax_evidence()
    # ИП на УСН: официального ОФР нет -> fns_tax_burden_ratio остается null; доход
    # по УСН без НДС из поступлений дает управленческий ориентир.
    evidence["incomeEvidence"] = {
        "value": None,
        "status": "source_error",
        "sourceKind": "onec_official_financial_results",
    }
    evidence["usnIncomeEvidence"] = {
        "value": "2000",
        "status": "confirmed",
        "sourceKind": "onec_kudir",
    }

    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "usn_income",
            "profileStatus": "ready",
            "revenueTaxRate": "0.01",
            "sourceKind": "1c",
        },
        evidence=evidence,
    )

    summary = payload["taxLoadSummary"]
    # Официальный коэффициент ФНС для ИП не считается.
    assert summary["fnsTaxBurdenRatio"] is None
    # Управленческий показатель: 100 / 2000 * 100 = 5.
    assert summary["usnIncomeValue"] == "2000"
    assert summary["usnIncomeTaxBurden"] == "5.0000"
    assert summary["usnIncomeStatus"] == "management_reference"
    assert summary["usnIncomeDenominatorKind"] == "usn_income_receipts_excluding_vat"
    assert payload["usnDetail"]["calculatedTaxYtd"] == "20.00"
    assert payload["usnDetail"]["paidTaxYtd"] == "100"
    assert payload["usnDetail"]["taxPayable"] == "-80.00"
    assert payload["usnDetail"]["status"] == "ready"
    assert payload["businessStatus"] == "preliminary"


def test_tax_load_usn_income_minus_expenses_has_no_management_ratio() -> None:
    evidence = _usn_tax_evidence()
    evidence["usnIncomeEvidence"] = {
        "value": "2000",
        "status": "confirmed",
        "sourceKind": "onec_kudir",
    }

    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "УСН Доходы минус расходы",
            "profileStatus": "ready",
        },
        evidence=evidence,
    )

    summary = payload["taxLoadSummary"]
    assert summary["usnIncomeValue"] is None
    assert summary["usnIncomeTaxBurden"] is None
    assert summary["usnIncomeStatus"] is None


def test_tax_load_evidence_reads_usn_income_base_from_kudir() -> None:
    sources = {
        "onec_kudir": AccountingEvidenceSource(
            source_type="onec_kudir",
            status="loaded",
            snapshot_id="kudir-sha",
            rows=(
                {"Организация_Key": "ORG-1", "Period": "2026-05-10T00:00:00",
                 "ДоходБаза": "1200", "ВидЗаписи": "Приход"},
                {"Организация_Key": "ORG-1", "Period": "2026-06-20T00:00:00",
                 "ДоходБаза": "800", "ВидЗаписи": "Приход"},
                # Другая организация — не суммируется.
                {"Организация_Key": "ORG-2", "Period": "2026-05-10T00:00:00",
                 "ДоходБаза": "999", "ВидЗаписи": "Приход"},
                # До начала года — вне YTD-периода.
                {"Организация_Key": "ORG-1", "Period": "2025-12-31T00:00:00",
                 "ДоходБаза": "500", "ВидЗаписи": "Приход"},
            ),
        ),
    }

    evidence = materialize_accounting_evidence(
        report_kind="tax_load",
        organization_id="ORG-1",
        period_start=date(2026, 5, 1),
        period_end=date(2026, 6, 30),
        refresh_run_id="gen-1",
        sources=sources,
    )

    usn = evidence["usnIncomeEvidence"]
    assert usn["sourceKind"] == "onec_kudir"
    assert usn["status"] == "loaded"
    # Только ORG-1 и только период с начала года: 1200 + 800.
    assert usn["value"] == "2000"
    assert usn["monthlyValues"] == [
        {
            "month": "2026-05",
            "value": "1200",
            "status": "loaded",
            "rowCount": 1,
        },
        {
            "month": "2026-06",
            "value": "800",
            "status": "loaded",
            "rowCount": 1,
        },
    ]


def test_tax_load_uses_classified_bank_tax_payments_when_ens_is_empty() -> None:
    sources = {
        "onec_tax_kinds": AccountingEvidenceSource(
            source_type="onec_tax_kinds",
            status="loaded",
            snapshot_id="tax-kinds-sha",
            rows=(
                {"Ref_Key": "TAX-USN", "Description": "Налог при УСН"},
                {"Ref_Key": "TAX-VAT", "Description": "НДС"},
            ),
        ),
        "onec_accounting_taxes": AccountingEvidenceSource(
            source_type="onec_accounting_taxes",
            status="loaded",
            snapshot_id="tax-register-sha",
            rows=(
                {
                    "Организация_Key": "ORG-1",
                    "Period": "2026-05-31T00:00:00",
                    "ВидНалога_Key": "TAX-USN",
                    "Сумма": "100",
                },
                {
                    "Организация_Key": "ORG-1",
                    "Period": "2026-05-31T00:00:00",
                    "ВидНалога_Key": "TAX-VAT",
                    "Сумма": "50",
                },
            ),
        ),
        "onec_accounting_taxes_on_ens": AccountingEvidenceSource(
            source_type="onec_accounting_taxes_on_ens",
            status="empty_expected",
            snapshot_id="ens-empty-sha",
            rows=(),
        ),
        "onec_accounting_bank_out": AccountingEvidenceSource(
            source_type="onec_accounting_bank_out",
            status="loaded",
            snapshot_id="bank-out-sha",
            rows=(
                {
                    "Организация_Key": "ORG-1",
                    "Date": "2026-03-20T00:00:00",
                    "Posted": True,
                    "DeletionMark": False,
                    "ВидОперации": "Налоги",
                    "НазначениеПлатежа": "Налог УСН",
                    "СуммаДокумента": "100",
                },
                {
                    "Организация_Key": "ORG-1",
                    "Date": "2026-05-21T00:00:00",
                    "Posted": True,
                    "DeletionMark": False,
                    "ВидОперации": "Налоги",
                    "НазначениеПлатежа": "НДС",
                    "СуммаДокумента": "50",
                },
            ),
        ),
        "onec_kudir": AccountingEvidenceSource(
            source_type="onec_kudir",
            status="loaded",
            snapshot_id="kudir-sha",
            rows=(
                {
                    "Организация_Key": "ORG-1",
                    "Period": "2026-05-31T00:00:00",
                    "ДоходБаза": "2000",
                },
            ),
        ),
    }

    evidence = materialize_accounting_evidence(
        report_kind="tax_load",
        organization_id="ORG-1",
        period_start=date(2026, 5, 1),
        period_end=date(2026, 5, 31),
        refresh_run_id="gen-bank-payments",
        sources=sources,
    )
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={"taxSystem": "УСН Доходы", "profileStatus": "ready"},
        evidence=evidence,
    )

    assert {row["sourceKind"] for row in evidence["taxRows"]} == {
        "onec_accounting_bank_out"
    }
    assert fns_paid_taxes_numerator(evidence["taxRows"]) == Decimal("150")
    assert not any(
        issue["code"] in {
            "onec_accounting_taxes_on_ens_gap",
            "paid_tax_fact_unconfirmed",
        }
        for issue in evidence["issues"]
    )
    assert evidence["usnTaxPaymentEvidence"]["monthlyValues"] == [
        {
            "month": "2026-03",
            "value": "100",
            "status": "loaded",
            "rowCount": 1,
        }
    ]
    assert payload["taxLoadSummary"]["usnIncomeTaxBurden"] == "7.5000"


def test_bank_tax_payment_fallback_rejects_unclassified_document() -> None:
    payments, classified = _bank_tax_payments(
        [
            {
                "Posted": True,
                "DeletionMark": False,
                "ВидОперации": "Налоги",
                "НазначениеПлатежа": "Налог УСН",
                "СуммаДокумента": "100",
            },
            {
                "Posted": True,
                "DeletionMark": False,
                "ВидОперации": "Налоги",
                "НазначениеПлатежа": "Налоговый платеж без расшифровки",
                "СуммаДокумента": "50",
            },
        ]
    )

    assert payments == {}
    assert classified is False


def test_tax_load_usn_management_ratio_source_gap_without_receipts() -> None:
    # УСН без подтвержденного дохода из поступлений: показатель null, не ноль.
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={"taxSystem": "usn_income", "profileStatus": "ready"},
        evidence=_usn_tax_evidence(),
    )
    summary = payload["taxLoadSummary"]
    assert summary["usnIncomeValue"] is None
    assert summary["usnIncomeTaxBurden"] is None
    assert summary["usnIncomeStatus"] == "source_gap"
    assert payload["usnDetail"]["status"] == "source_gap"


def test_scenario_excel_has_exact_sheets_and_traceable_overview(tmp_path: Path) -> None:
    cases = [
        (
            build_month_close_control_payload(
                _report("month_close_control"), _month_close_evidence()
            ),
            list(MONTH_CLOSE_SHEETS),
            "Сводка закрытия",
        ),
        (
            build_tax_load_payload(
                _report("tax_load"),
                tax_profile={},
                evidence=_tax_evidence(),
            ),
            list(TAX_LOAD_SHEETS),
            "Обзор",
        ),
    ]
    for index, (payload, expected_sheets, summary_sheet) in enumerate(cases):
        payload_hash = canonical_payload_sha256(payload)
        path = tmp_path / f"scenario-{index}.xlsx"
        write_scenario_excel(payload, payload_hash, path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        assert workbook.sheetnames == expected_sheets
        values = {
            row[0]: row[1]
            for row in workbook[summary_sheet].iter_rows(values_only=True)
            if row[0]
        }
        if payload["reportKind"] == "tax_load":
            assert "ID отчёта" not in values
            assert "Версия методики" not in values
            assert "reportId" not in values
            assert "payloadSha256" not in values
            assert workbook.properties.identifier == payload["meta"]["reportId"]
            assert workbook.properties.version == payload["meta"]["methodologyVersion"]
            assert workbook.properties.language == "ru-RU"
        else:
            assert values["reportId"] == payload["meta"]["reportId"]
            assert values["payloadSha256"] == payload_hash


def test_tax_load_excel_localizes_headers_and_enum_values(tmp_path: Path) -> None:
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "УСН Доходы",
            "profileStatus": "ready",
            "revenueTaxRate": "0.06",
        },
        evidence=_tax_evidence(),
    )
    payload["issues"].append(
        {
            "severity": "warning",
            "section": "Доходный знаменатель",
            "message": (
                "Источник onec_official_financial_results не подтвержден "
                "за выбранный период."
            ),
            "nextAction": "Повторить read-only загрузку.",
        }
    )
    path = tmp_path / "tax-load-russian.xlsx"
    write_scenario_excel(
        payload,
        canonical_payload_sha256(payload),
        path,
        export_context={
            "clientName": "Клиент А",
            "organizationName": "ИП Клиент А",
        },
    )
    workbook = load_workbook(path, data_only=True)

    overview = {
        row[0]: row[1]
        for row in workbook["Обзор"].iter_rows(values_only=True)
        if row[0]
    }
    assert overview["Клиент"] == "Клиент А"
    assert overview["Организация 1С"] == "ИП Клиент А"
    assert overview["Вид отчёта"] == "Налоговая нагрузка"
    assert overview["Выбранный месяц"] == "Январь 2026"
    assert overview["Период расчёта"] == "Предварительно, с начала года"
    assert overview["Налоговый режим"] == "УСН «Доходы»"
    assert overview["Статус налогового профиля"] == "Готово"
    assert overview["Статус отчёта"] == "Нужна проверка бухгалтера"
    assert overview["Подтверждение бухгалтера"] == "Не подтверждено"
    assert overview["Начало отчётного периода"].date() == date(2026, 1, 1)
    assert overview["Начало периода с начала года"].date() == date(2026, 1, 1)

    taxes = list(workbook["Налоги"].iter_rows(values_only=True))
    assert taxes[0] == (
        "Налог",
        "Период",
        "Налоговая база",
        "Начислено",
        "Уплачено",
        "Сальдо",
        "Срок уплаты",
        "Включён в нагрузку ФНС",
        "Причина исключения",
        "Статус подтверждения",
        "Источник",
    )
    assert taxes[1][7] == "Да"
    assert taxes[1][9] == "Загружено"
    assert taxes[2][8] == "Агентский платёж"
    assert "Код налога" not in taxes[0]
    assert "Код замечания" not in taxes[0]

    coverage = list(workbook["Источники и статус"].iter_rows(values_only=True))
    assert coverage[0] == (
        "Источник",
        "Начало отчётного периода",
        "Окончание отчётного периода",
        "Статус",
    )
    assert coverage[1][0] == "Налоговый учёт 1С"
    assert coverage[1][3] == "Загружено"
    assert "ID снимка" not in coverage[0]

    assert workbook["Налоги"].freeze_panes == "A2"
    assert workbook["Налоги"].sheet_view.showGridLines is False
    tax_table = next(iter(workbook["Налоги"].tables.values()))
    assert tax_table.tableStyleInfo.name == "TableStyleMedium2"
    assert tax_table.autoFilter is not None

    paid_column = taxes[0].index("Уплачено") + 1
    paid_cell = workbook["Налоги"].cell(row=2, column=paid_column)
    assert paid_cell.data_type == "n"
    assert "₽" in paid_cell.number_format
    ratio_row = next(
        row
        for row in workbook["Обзор"].iter_rows()
        if row[0].value == "Налоговая нагрузка по методике ФНС, %"
    )
    assert ratio_row[1].data_type == "n"
    assert "%" in ratio_row[1].number_format
    due_date_column = taxes[0].index("Срок уплаты") + 1
    due_date_cell = workbook["Налоги"].cell(row=2, column=due_date_column)
    assert due_date_cell.data_type == "d"
    assert due_date_cell.number_format == "DD.MM.YYYY"
    workbook_text = " ".join(
        str(cell.value or "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert "onec_official_financial_results" not in workbook_text
    assert "read-only" not in workbook_text
    assert "Отчёт о финансовых результатах 1С" in workbook_text
    assert "только для чтения" in workbook_text
    assert payload["meta"]["reportId"] not in workbook_text
    assert payload["meta"]["methodologyVersion"] not in workbook_text


def test_tax_load_excel_builds_detailed_usn_monthly_matrix(tmp_path: Path) -> None:
    evidence = _usn_tax_evidence()
    evidence["usnIncomeEvidence"] = {
        "value": "600",
        "status": "loaded",
        "sourceKind": "onec_kudir",
        "monthlyValues": [
            {
                "month": f"2026-{month:02d}",
                "value": "100",
                "status": "loaded",
                "rowCount": 1,
            }
            for month in range(1, 7)
        ],
    }
    evidence["usnTaxPaymentEvidence"] = {
        "status": "loaded",
        "sourceKind": "onec_accounting_bank_out",
        "monthlyValues": [
            {"month": "2026-01", "value": "40", "status": "loaded", "rowCount": 1},
            {"month": "2026-04", "value": "60", "status": "loaded", "rowCount": 1},
        ],
    }
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={
            "taxSystem": "УСН Доходы",
            "profileStatus": "ready",
            "revenueTaxRate": "0.01",
        },
        evidence=evidence,
    )
    path = tmp_path / "tax-load-usn-detail.xlsx"
    write_scenario_excel(payload, canonical_payload_sha256(payload), path)
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Расчёт УСН"]
    rows = {
        row[0].value: row
        for row in sheet.iter_rows(min_row=2)
        if row[0].value is not None
    }
    headers = [cell.value for cell in sheet[1]]

    assert headers == [
        "Показатель",
        "Январь",
        "Февраль",
        "Март",
        "Итого за I квартал",
        "Апрель",
        "Май",
        "Июнь",
        "Итого за полугодие",
    ]
    assert rows["Итого доход без НДС"][4].value == Decimal("300")
    assert rows["Итого доход без НДС"][8].value == Decimal("600")
    assert rows["Ставка УСН"][8].value == pytest.approx(0.01)
    assert rows["Ставка УСН"][8].number_format == "0.00%"
    assert rows["Исчислено УСН с начала года"][8].value == Decimal("6.00")
    assert rows["Уплачено УСН"][8].value == Decimal("100")
    assert rows["К доплате / переплата УСН"][8].value == Decimal("-94.00")
    assert rows["Итого доход без НДС"][0].fill.fgColor.rgb.endswith("FFF200")
    assert len(sheet.tables) == 1
    assert not any(
        cell.data_type == "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


def test_tax_load_excel_hides_1c_placeholder_due_date(tmp_path: Path) -> None:
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={},
        evidence=_tax_evidence(),
    )
    payload["taxRows"][0]["dueDate"] = "0001-01-01T00:00:00"
    payload["paymentSchedule"][0]["dueDate"] = "0001-01-01T00:00:00"
    path = tmp_path / "tax-load-placeholder-date.xlsx"
    write_scenario_excel(payload, canonical_payload_sha256(payload), path)
    workbook = load_workbook(path, data_only=True)

    for title in ("Налоги", "График платежей"):
        sheet = workbook[title]
        headers = [cell.value for cell in sheet[1]]
        due_date_cell = sheet.cell(row=2, column=headers.index("Срок уплаты") + 1)
        assert due_date_cell.value == "Не указано"
        assert due_date_cell.data_type == "s"


def test_tax_load_evidence_rejects_1c_placeholder_date() -> None:
    assert _date_text("0001-01-01T00:00:00") is None
    assert _date_text("2026-05-31T00:00:00") == "2026-05-31"


def test_tax_load_source_gap_issue_has_only_user_facing_russian_text() -> None:
    issue = _source_gap_issues(
        {}, {"onec_official_financial_results": "Доходный знаменатель"}
    )[0]

    assert issue["message"] == (
        "Источник «Доходный знаменатель» не подтверждён за выбранный период."
    )
    assert issue["nextAction"] == (
        "Проверить публикацию 1С и повторить загрузку только для чтения."
    )
    assert "onec_" not in issue["message"]
    assert "read-only" not in issue["nextAction"]


def test_tax_load_excel_ignores_unapproved_internal_field(tmp_path: Path) -> None:
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={},
        evidence=_tax_evidence(),
    )
    payload["taxLoadSummary"]["futureInternalField"] = "internal_value"

    path = tmp_path / "tax-load-untranslated.xlsx"
    write_scenario_excel(payload, canonical_payload_sha256(payload), path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    values = {
        value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    }

    assert "futureInternalField" not in values
    assert "internal_value" not in values


def test_tax_load_excel_hides_unknown_enum_value(tmp_path: Path) -> None:
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={},
        evidence=_tax_evidence(),
    )
    payload["businessStatus"] = "future_internal_status"
    path = tmp_path / "tax-load-unknown-status.xlsx"
    write_scenario_excel(payload, canonical_payload_sha256(payload), path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    overview = {
        row[0]: row[1]
        for row in workbook["Обзор"].iter_rows(values_only=True)
        if row[0]
    }

    assert overview["Статус отчёта"] == "Не определено"
    assert "future_internal_status" not in overview.values()


def test_tax_load_excel_neutralizes_formula_text(tmp_path: Path) -> None:
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={},
        evidence=_tax_evidence(),
    )
    payload["taxRows"][0]["taxName"] = '=HYPERLINK("https://example.test")'
    path = tmp_path / "tax-load-safe-text.xlsx"
    write_scenario_excel(payload, canonical_payload_sha256(payload), path)
    workbook = load_workbook(path, data_only=False)
    cell = workbook["Налоги"].cell(row=2, column=1)

    assert cell.value.startswith("=HYPERLINK")
    assert cell.data_type == "s"


def test_scenario_excel_atomic_save_preserves_previous_file_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    payload = build_tax_load_payload(
        _report("tax_load"),
        tax_profile={},
        evidence=_tax_evidence(),
    )
    path = tmp_path / "tax-load-atomic.xlsx"
    path.write_bytes(b"previous-complete-file")

    def fail_save(_workbook: Workbook, temporary_path: Path) -> None:
        Path(temporary_path).write_bytes(b"partial")
        raise RuntimeError("simulated save failure")

    monkeypatch.setattr(Workbook, "save", fail_save)
    with pytest.raises(RuntimeError, match="simulated save failure"):
        write_scenario_excel(payload, canonical_payload_sha256(payload), path)

    assert path.read_bytes() == b"previous-complete-file"
    assert not list(tmp_path.glob(".tax-load-atomic-*.xlsx"))


def test_fns_2025_reference_is_versioned_and_comparison_disabled() -> None:
    path = (
        Path(__file__).parents[1]
        / "config"
        / "reference"
        / "fns_tax_burden"
        / "2025.json"
    )
    payload = json.loads(path.read_text(encoding="utf-8"))

    assert payload["year"] == 2025
    assert payload["sourceSha256"] == (
        "6fe17c81b8f86ae93f979ec4d67657c4ce27d405e3e7504c898f800cbbee8737"
    )
    assert payload["sourceIncludesAgentNdfl"] is True
    assert payload["comparisonEnabled"] is False
    assert payload["comparisonStatus"] == "pending_methodology_confirmation"


def _make_api_client(
    tmp_path: Path,
    *,
    enabled_report_kinds: str = (
        "marketplace_unit_economics,month_close_control,tax_load"
    ),
) -> tuple[TestClient, object]:
    database_url = f"sqlite:///{tmp_path / 'web.sqlite3'}"
    engine = make_engine(database_url)
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        repository.ensure_tenant(db, "tenant-a", "Клиент А")
        company = repository.ensure_client_company(
            db,
            tenant_id="tenant-a",
            client_id="tenant-a",
            display_name="ООО Клиент А",
        )
        assert company is not None
        company.onec_organization_id = "ORG-1"
        company_two = repository.ensure_client_company(
            db,
            tenant_id="tenant-a",
            client_id="tenant-a",
            display_name="ООО Клиент А-2",
        )
        assert company_two is not None
        company_two.onec_organization_id = "ORG-2"
        admin = repository.upsert_user(
            db,
            email="admin@example.com",
            password="secret",
            tenant_id="tenant-a",
            role="admin",
        )
        repository.upsert_user(
            db,
            email="client@example.com",
            password="secret",
            tenant_id="tenant-a",
            role="client_viewer",
        )
        repository.ensure_tenant(db, "tenant-b", "Клиент Б")
        repository.upsert_user(
            db,
            email="other@example.com",
            password="secret",
            tenant_id="tenant-b",
            role="admin",
        )
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id="tenant-a",
            client_id="tenant-a",
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="snapshot-1",
            period_start=date(2026, 1, 1),
            period_end=date(2026, 6, 30),
            user=admin,
        )
        repository.add_source_refresh_collection(
            db,
            refresh,
            source_type="month_close_control_evidence",
            source_label="Normalized month close evidence",
            required=True,
            status="loaded",
            payload={"normalizedEvidence": _month_close_evidence()},
        )
        repository.add_source_refresh_collection(
            db,
            refresh,
            source_type="tax_load_evidence",
            source_label="Normalized tax evidence",
            required=True,
            status="loaded",
            payload={"normalizedEvidence": _tax_evidence()},
        )
        repository.update_source_refresh_run(
            db,
            refresh,
            status="needs_review",
            finished_at=repository.security.utcnow(),
        )
        db.add(
            OrganizationTaxProfile(
                id="tax-profile-1",
                tenant_id="tenant-a",
                client_id="tenant-a",
                client_company_id=company.id,
                organization_id="ORG-1",
                tax_system="osno",
                vat_rate=Decimal("20"),
                vat_mode="included",
                vat_deduction_mode="allowed",
                revenue_tax_rate=Decimal("20"),
                income_tax_kind="profit_tax",
                valid_from=date(2026, 1, 1),
                valid_to=None,
                source="1c",
                rate_basis_kind="onec_profile",
                basis_document="",
                confirmed_by="",
                source_object_ids="[]",
                source_refresh_run_id=refresh.id,
                source_snapshot_hash="snapshot-hash",
                methodology_version="tax-profile-v1",
                status="active",
                created_at=repository.security.utcnow(),
            )
        )
        db.commit()
    export_root = tmp_path / "reports"
    export_root.mkdir()
    settings = WebSettings(
        database_url=database_url,
        cookie_secure=False,
        allowed_export_root=str(export_root),
        openai_api_key="",
        source_refresh_tenant="tenant-a",
        enabled_report_kinds=enabled_report_kinds,
    )

    class _QueuedOnlySourceRefreshService:
        def run_existing(self, *_args, **_kwargs):
            return None

    class _QueuedOnlyAutoRefreshService:
        source_refresh_service = _QueuedOnlySourceRefreshService()

    app = create_app(
        settings=settings,
        session_factory=session_factory,
        auto_refresh_service=_QueuedOnlyAutoRefreshService(),
    )
    return TestClient(app), session_factory


def _complete_generation(
    session_factory,
    generation_run_id: str,
    evidence: dict,
) -> str:
    with session_factory() as db:
        generation = db.get(SourceRefreshRun, generation_run_id)
        assert generation is not None
        user = db.scalar(select(User).where(User.email == "admin@example.com"))
        base_tax_profile = db.get(OrganizationTaxProfile, "tax-profile-1")
        if (
            base_tax_profile is not None
            and base_tax_profile.organization_id == generation.organization_id
        ):
            db.add(
                OrganizationTaxProfile(
                    id=f"tax-profile-{generation.id}",
                    tenant_id=base_tax_profile.tenant_id,
                    client_id=base_tax_profile.client_id,
                    client_company_id=base_tax_profile.client_company_id,
                    organization_id=base_tax_profile.organization_id,
                    tax_system=base_tax_profile.tax_system,
                    vat_rate=base_tax_profile.vat_rate,
                    vat_mode=base_tax_profile.vat_mode,
                    vat_deduction_mode=base_tax_profile.vat_deduction_mode,
                    revenue_tax_rate=base_tax_profile.revenue_tax_rate,
                    income_tax_kind=base_tax_profile.income_tax_kind,
                    valid_from=base_tax_profile.valid_from,
                    valid_to=base_tax_profile.valid_to,
                    source=base_tax_profile.source,
                    rate_basis_kind=base_tax_profile.rate_basis_kind,
                    basis_document=base_tax_profile.basis_document,
                    confirmed_by=base_tax_profile.confirmed_by,
                    source_object_ids=base_tax_profile.source_object_ids,
                    source_refresh_run_id=generation.id,
                    source_snapshot_hash=base_tax_profile.source_snapshot_hash,
                    methodology_version=base_tax_profile.methodology_version,
                    status=base_tax_profile.status,
                    created_at=repository.security.utcnow(),
                )
            )
        repository.add_source_refresh_collection(
            db,
            generation,
            source_type=f"{generation.target_report_kind}_evidence",
            source_label="Test evidence",
            required=True,
            status="loaded",
            snapshot_hash="test-evidence-sha256",
            row_count=1,
            payload={
                "contractVersion": "test-v2",
                "organizationId": generation.organization_id,
                "payloadSha256": "test-evidence-sha256",
                "normalizedEvidence": {
                    **evidence,
                    "sourceRefreshRunId": generation.id,
                    "evidenceSha256": "test-evidence-sha256",
                },
            },
            organization_id=generation.organization_id,
        )
        report = repository.complete_accounting_report_generation(
            db,
            generation=generation,
            user=user,
        )
        db.commit()
        return report.id


def _login(client: TestClient, email: str) -> None:
    response = client.post(
        "/api/auth/login", json={"email": email, "password": "secret"}
    )
    assert response.status_code == 200


def test_staff_api_generation_idempotency_current_audit_and_excel(
    tmp_path: Path,
) -> None:
    client, session_factory = _make_api_client(tmp_path)
    _login(client, "admin@example.com")

    kinds = client.get("/api/clients/tenant-a/report-kinds")
    assert kinds.status_code == 200
    assert {item["kind"] for item in kinds.json()["reportKinds"]} == {
        "marketplace_unit_economics",
        "month_close_control",
        "tax_load",
    }
    request_payload = {
        "reportKind": "month_close_control",
        "organizationId": "ORG-1",
        "periodMonth": "2026-06",
    }
    first = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "month-close-click-1"},
        json=request_payload,
    )
    duplicate = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "month-close-click-1"},
        json=request_payload,
    )

    assert first.status_code == 202
    assert duplicate.status_code == 202
    assert first.json()["generationRunId"] == duplicate.json()["generationRunId"]
    assert first.json()["deduplicated"] is False
    assert duplicate.json()["deduplicated"] is True
    reused_for_other_request = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "month-close-click-1"},
        json={
            "reportKind": "tax_load",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )
    assert reused_for_other_request.status_code == 400
    assert first.json()["reportId"] is None
    report_id = _complete_generation(
        session_factory,
        first.json()["generationRunId"],
        _month_close_evidence(),
    )
    scenario = client.get(f"/api/reports/{report_id}/scenario")
    assert scenario.status_code == 200
    assert scenario.json()["meta"]["reportId"] == report_id
    assert scenario.json()["payloadSha256"]
    generation = client.get(
        f"/api/report-generations/{first.json()['generationRunId']}"
    )
    assert generation.status_code == 200
    assert generation.json()["status"] == "completed"
    export = client.get(f"/api/reports/{report_id}/export.xlsx")
    assert export.status_code == 200
    workbook = load_workbook(filename=BytesIO(export.content), read_only=True)
    assert workbook.sheetnames == list(MONTH_CLOSE_SHEETS)

    second_revision = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "month-close-click-2"},
        json=request_payload,
    )
    other_kind = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "tax-load-click-1"},
        json={
            "reportKind": "tax_load",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )
    other_organization = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "month-close-org-2"},
        json={
            "reportKind": "month_close_control",
            "organizationId": "ORG-2",
            "periodMonth": "2026-06",
        },
    )
    assert second_revision.status_code == 202
    assert other_kind.status_code == 202
    assert other_organization.status_code == 202
    second_revision_report_id = _complete_generation(
        session_factory,
        second_revision.json()["generationRunId"],
        _month_close_evidence(),
    )
    other_kind_report_id = _complete_generation(
        session_factory,
        other_kind.json()["generationRunId"],
        _tax_evidence(),
    )
    other_organization_report_id = _complete_generation(
        session_factory,
        other_organization.json()["generationRunId"],
        {**_month_close_evidence(), "organizationId": "ORG-2"},
    )
    tax_export = client.get(f"/api/reports/{other_kind_report_id}/export.xlsx")
    assert tax_export.status_code == 200
    assert "xlsx" in tax_export.headers["content-disposition"]
    tax_workbook = load_workbook(BytesIO(tax_export.content), data_only=True)
    tax_overview = {
        row[0]: row[1]
        for row in tax_workbook["Обзор"].iter_rows(values_only=True)
        if row[0]
    }
    assert tax_overview["Клиент"] == "Клиент А"
    assert tax_overview["Организация 1С"] == "ООО Клиент А"
    assert tax_overview["Налоговый режим"] == "ОСНО"
    assert tax_overview["Начало периода с начала года"].date() == date(2026, 1, 1)
    assert "Код налога" not in {cell.value for cell in tax_workbook["Налоги"][1]}

    with session_factory() as db:
        reports = list(
            db.scalars(select(ReportRun).where(ReportRun.tenant_id == "tenant-a"))
        )
        current_scopes = {
            (report.report_kind, report.organization_id, report.id)
            for report in reports
            if report.is_current and report.report_kind != "marketplace_unit_economics"
        }
        assert current_scopes == {
            ("month_close_control", "ORG-1", second_revision_report_id),
            ("month_close_control", "ORG-2", other_organization_report_id),
            ("tax_load", "ORG-1", other_kind_report_id),
        }
        current_reports = [
            report
            for report in reports
            if report.is_current and report.report_kind != "marketplace_unit_economics"
        ]
        assert all(report.publication_status == "draft" for report in current_reports)
        actions = set(db.scalars(select(AuditEvent.action)))
        assert {
            "report_generation_requested",
            "report_generation_deduplicated",
            "report_generation_completed",
            "report_viewed",
            "report_exported",
        }.issubset(actions)


def test_accounting_scenarios_are_staff_only_and_tenant_isolated(
    tmp_path: Path,
) -> None:
    client, _session_factory = _make_api_client(tmp_path)
    _login(client, "client@example.com")

    kinds = client.get("/api/clients/tenant-a/report-kinds")
    assert [item["kind"] for item in kinds.json()["reportKinds"]] == [
        "marketplace_unit_economics"
    ]
    forbidden = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "client-attempt"},
        json={
            "reportKind": "month_close_control",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )
    assert forbidden.status_code == 404

    client.post("/api/auth/logout")
    _login(client, "other@example.com")
    isolated = client.get("/api/clients/tenant-a/report-kinds")
    assert isolated.status_code == 404


def test_disabled_accounting_kind_is_hidden_and_rejected(tmp_path: Path) -> None:
    client, _session_factory = _make_api_client(
        tmp_path, enabled_report_kinds="marketplace_unit_economics"
    )
    _login(client, "admin@example.com")

    kinds = client.get("/api/clients/tenant-a/report-kinds")
    assert [item["kind"] for item in kinds.json()["reportKinds"]] == [
        "marketplace_unit_economics"
    ]
    response = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "disabled-kind"},
        json={
            "reportKind": "tax_load",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )
    assert response.status_code == 404


def test_disabled_kind_is_hidden_by_all_read_and_export_routes(tmp_path: Path) -> None:
    client, session_factory = _make_api_client(tmp_path)
    _login(client, "admin@example.com")
    queued = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "disable-after-build"},
        json={
            "reportKind": "month_close_control",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )
    report_id = _complete_generation(
        session_factory,
        queued.json()["generationRunId"],
        _month_close_evidence(),
    )
    disabled_settings = WebSettings(
        database_url=f"sqlite:///{tmp_path / 'web.sqlite3'}",
        cookie_secure=False,
        allowed_export_root=str(tmp_path / "reports"),
        openai_api_key="",
        source_refresh_tenant="tenant-a",
        enabled_report_kinds="marketplace_unit_economics",
    )
    disabled_client = TestClient(
        create_app(settings=disabled_settings, session_factory=session_factory)
    )
    _login(disabled_client, "admin@example.com")

    requests = (
        (
            "/api/reports",
            {"client_id": "tenant-a", "report_kind": "month_close_control"},
        ),
        (
            "/api/clients/tenant-a/reports",
            {"report_kind": "month_close_control", "organization_id": "ORG-1"},
        ),
        (
            "/api/reports/latest/summary",
            {
                "client_id": "tenant-a",
                "report_kind": "month_close_control",
                "organization_id": "ORG-1",
            },
        ),
        (f"/api/reports/{report_id}/summary", None),
        (f"/api/reports/{report_id}/scenario", None),
        (f"/api/reports/{report_id}/export.xlsx", None),
    )
    for path, params in requests:
        response = disabled_client.get(path, params=params)
        assert response.status_code == 404, path


def test_parallel_scope_returns_existing_active_generation(tmp_path: Path) -> None:
    client, session_factory = _make_api_client(tmp_path)
    with session_factory() as db:
        active = repository.create_source_refresh_run(
            db,
            tenant_id="tenant-a",
            client_id="tenant-a",
            mode="report-generation",
            credential_source="stored_snapshots",
            dry_run=False,
            snapshot_set_id="active-generation-snapshot",
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            reason="parallel request test",
            enforce_active_check=False,
        )
        active.target_report_kind = "tax_load"
        active.organization_id = "ORG-1"
        active.idempotency_key = "active-original-key"
        active.status = "running"
        active.started_at = repository.security.utcnow()
        active.finished_at = None
        active_id = active.id
        db.commit()
    _login(client, "admin@example.com")

    response = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "parallel-new-key"},
        json={
            "reportKind": "tax_load",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )

    assert response.status_code == 202
    assert response.json()["generationRunId"] == active_id
    assert response.json()["reportId"] is None
    assert response.json()["deduplicated"] is True


def test_month_close_uses_month_window_and_tax_load_uses_ytd(tmp_path: Path) -> None:
    client, session_factory = _make_api_client(tmp_path)
    _login(client, "admin@example.com")

    for report_kind, expected_start in (
        ("month_close_control", date(2026, 6, 1)),
        ("tax_load", date(2026, 1, 1)),
    ):
        response = client.post(
            "/api/clients/tenant-a/reports/generate",
            headers={"Idempotency-Key": f"source-window-{report_kind}"},
            json={
                "reportKind": report_kind,
                "organizationId": "ORG-1",
                "periodMonth": "2026-06",
            },
        )
        assert response.status_code == 202
        with session_factory() as db:
            run = db.get(SourceRefreshRun, response.json()["generationRunId"])
            assert run is not None
            assert run.source_window_start == expected_start


def test_report_generation_does_not_replace_source_freshness_health(
    tmp_path: Path,
) -> None:
    client, _session_factory = _make_api_client(tmp_path)
    before = client.get("/api/health").json()
    _login(client, "admin@example.com")

    queued = client.post(
        "/api/clients/tenant-a/reports/generate",
        headers={"Idempotency-Key": "health-isolation"},
        json={
            "reportKind": "month_close_control",
            "organizationId": "ORG-1",
            "periodMonth": "2026-06",
        },
    )

    after = client.get("/api/health").json()
    assert queued.status_code == 202
    assert queued.json()["status"] == "queued"
    assert after["latestSourceRefreshRunId"] == before["latestSourceRefreshRunId"]
    assert after["latestSourceRefreshMode"] != "report-generation"
    assert after["latestSourceRefreshStatus"] == before["latestSourceRefreshStatus"]


def test_raw_onec_rows_flow_to_evidence_report_web_and_excel(tmp_path: Path) -> None:
    _client, session_factory = _make_api_client(tmp_path)
    settings = WebSettings(
        database_url=f"sqlite:///{tmp_path / 'web.sqlite3'}",
        cookie_secure=False,
        allowed_export_root=str(tmp_path / "reports"),
        source_refresh_enabled=True,
        source_refresh_root=str(tmp_path / "source-refresh"),
        source_refresh_min_free_gb=0,
        enabled_report_kinds=(
            "marketplace_unit_economics,month_close_control,tax_load"
        ),
        accounting_recordtype_page_size=50000,
    )

    source_rows = {
        "organizations": [{"Ref_Key": "ORG-1", "Description": "ООО Клиент А"}],
        "accounting_chart": [
            {"Ref_Key": "ACC-51", "Code": "51", "Description": "Счета"}
        ],
        "accounting_register_records": [
            {
                "Организация_Key": "ORG-1",
                "Period": "2026-06-10T00:00:00",
                "AccountDr_Key": "ACC-51",
                "AccountCr_Key": "ACC-51",
                "Сумма": "25",
            }
        ],
        "accounting_taxes": [{"Организация_Key": "ORG-1", "Сумма": "100"}],
        "accounting_ens": [{"Организация_Key": "ORG-1", "Сумма": "80"}],
        "accounting_bank_in": [{"Организация_Key": "ORG-1", "СуммаДокумента": "500"}],
        "accounting_bank_out": [{"Организация_Key": "ORG-1", "СуммаДокумента": "300"}],
    }

    captured_export_kwargs: dict[str, object] = {}

    def fake_onec_exporter(_settings, collections, output_dir, **kwargs):
        captured_export_kwargs.update(kwargs)
        output_dir.mkdir(parents=True, exist_ok=True)
        results = []
        for item in collections:
            rows = source_rows.get(item.sample_id, [])
            output_path = output_dir / f"{item.sample_id}.raw.json"
            output_path.write_text(
                json.dumps({"value": rows}, ensure_ascii=False),
                encoding="utf-8",
            )
            results.append(
                OnecSampleExportResult(
                    sample_id=item.sample_id,
                    collection_name=item.collection_name,
                    ok=True,
                    row_count=len(rows),
                    page_count=1,
                    raw_payload_hash=f"sha-{item.sample_id}",
                    output_path=output_path,
                    status_code=200,
                )
            )
        return results

    def fake_balance_exporter(_settings, output_dir, **_kwargs):
        return OnecSampleExportResult(
            sample_id="accounting_balance_and_turnovers",
            collection_name="AccountingRegister/BalanceAndTurnovers",
            ok=False,
            row_count=0,
            status_code=500,
            error="HTTP 500",
            status="failed",
        )

    captured_recordtype_kwargs: dict[str, object] = {}

    def fake_recordtype_exporter(_settings, output_dir, **kwargs):
        captured_recordtype_kwargs.update(kwargs)
        rows = [
            {
                "Organization_Key": "ORG-1",
                "Account_Key": "ACC-51",
                "OpeningDebit": "100",
                "OpeningCredit": "0",
                "DebitTurnover": "500",
                "CreditTurnover": "300",
                "ClosingDebit": "300",
                "ClosingCredit": "0",
            }
        ]
        output_path = output_dir / "accounting_register_balances.raw.json"
        output_path.write_text(
            json.dumps({"value": rows}, ensure_ascii=False),
            encoding="utf-8",
        )
        return OnecSampleExportResult(
            sample_id="accounting_register_balances",
            collection_name="AccountingRegister_Управленческий_RecordType",
            ok=True,
            row_count=1,
            page_count=1,
            raw_payload_hash="sha-recordtype-balances",
            output_path=output_path,
            status_code=200,
        )

    service = SourceRefreshService(
        settings,
        onec_exporter=fake_onec_exporter,
        onec_accounting_balance_exporter=fake_balance_exporter,
        onec_accounting_recordtype_exporter=fake_recordtype_exporter,
        onec_metadata_checker=lambda _settings: OnecODataMetadataCheckResult(
            ok=True,
            status_code=200,
            content_type="application/xml",
        ),
    )
    credentials = SourceCredentials(
        wb_settings=None,
        onec_settings=OnecODataSettings(
            base_url="https://onec.invalid/odata/standard.odata",
            username="read-only",
            password="test-only",
        ),
        ozon_settings=None,
        wb_cabinet_ids={},
        ozon_cabinet_ids={},
        issues=(),
    )
    service._credentials = lambda *_args, **_kwargs: credentials

    with session_factory() as db:
        user = db.scalar(select(User).where(User.email == "admin@example.com"))
        assert user is not None
        generation, deduplicated = repository.generate_accounting_report(
            db,
            user=user,
            client_id="tenant-a",
            report_kind="month_close_control",
            organization_id="ORG-1",
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            idempotency_key="raw-onec-e2e",
        )
        assert deduplicated is False
        generation_id = generation.id
        db.commit()

    with session_factory() as db:
        result = service.run_existing(db, generation_id, worker_id="test-worker")
        db.commit()
        assert int(captured_export_kwargs["max_pages"]) >= 1000
        assert int(captured_export_kwargs["top"]) >= 5000
        assert int(captured_recordtype_kwargs["page_size"]) == 50000
        assert result["status"] == "completed"
        assert result["reportId"]
        report = db.get(ReportRun, result["reportId"])
        assert report is not None
        source_row_count = len(
            list(
                db.scalars(
                    select(SourceSnapshotRow)
                    .join(SourceRefreshCollection)
                    .where(
                        SourceRefreshCollection.refresh_run_id == generation_id,
                        SourceRefreshCollection.source_type
                        == "onec_accounting_register_balances",
                    )
                )
            )
        )
        assert source_row_count == 1
        scenario = repository.scenario_payload_for_report(db, report)
        assert scenario["meta"]["sourceRefreshRunId"] == generation_id
        assert scenario["osvRows"][0]["accountCode"] == "51"
        assert scenario["osvRows"][0]["closingDebit"] == "300"

        path = tmp_path / "raw-onec-e2e.xlsx"
        write_scenario_excel(scenario, scenario["payloadSha256"], path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        values = {
            row[0]: row[1]
            for row in workbook["Сводка закрытия"].iter_rows(values_only=True)
            if row[0]
        }
        assert values["reportId"] == report.id
        assert values["payloadSha256"] == scenario["payloadSha256"]
