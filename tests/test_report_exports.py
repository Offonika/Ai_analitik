from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from wb_unit_economics.excel import REQUIRED_SHEETS
from wb_unit_economics.report_exports import (
    write_excel_from_marts,
    write_ozon_diagnostics_excel,
)


def test_db_first_excel_export_uses_client_facing_russian_headers(
    tmp_path: Path,
) -> None:
    output = tmp_path / "report.xlsx"
    write_excel_from_marts(
        {
            "meta": {
                "client": "Тестовый клиент",
                "period": "01.03.2026 - 31.03.2026",
                "methodologyVersion": "test",
                "source": "DB report marts",
                "lineageType": "db_first_report_marts",
            },
            "readiness": {"status": "source_coverage_gap"},
            "unitRows": [
                {
                    "product": "Товар",
                    "week": "2026-03-02",
                    "month": "Март 2026",
                    "organization": "Организация A",
                    "cabinet": "Кабинет A",
                    "articleWb": "WB-1",
                    "article1c": "A-1",
                    "barcode": "BAR-1",
                    "sales": 1,
                    "returns": 0,
                    "revenue": 1000.0,
                    "profit": 120.0,
                    "status": "ОК",
                }
            ],
            "liquidityRows": [
                {
                    "product": "Товар",
                    "month": "Март 2026",
                    "organization": "Организация A",
                    "cabinet": "Кабинет A",
                    "articleWb": "WB-1",
                    "article1c": "A-1",
                    "barcode": "BAR-1",
                    "sales": 1,
                    "returns": 0,
                    "revenue": 1000.0,
                    "profit": 120.0,
                    "liquidityStatus": "Прибыльный до 500 руб. в месяц",
                    "status": "ОК",
                }
            ],
            "monthly": [{"month": "Март 2026", "status": "ok", "profit": 120.0}],
            "expenses": [{"expense": "Комиссия WB", "amount": 100.0, "share": 0.1}],
            "returns": [],
            "lostSales": [{"product": "Товар", "sourceStatus": "ok"}],
            "documentReconciliation": [
                {"documentReport": "Отчет комиссионера", "status": "OK"}
            ],
            "reconciliationMonthly": [{"month": "Март 2026", "wb_cogs": 300.0}],
            "taxInputReconciliation": [
                {
                    "week": "2026-03-02",
                    "cabinet": "Кабинет A",
                    "organization": "Организация A",
                    "vatInputFromWb": 10.0,
                    "vatInputFrom1c": 12.0,
                    "vatInputDifference": 2.0,
                    "vatInputCompleteness": "partial",
                }
            ],
        },
        output,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert [
            name for name in workbook.sheetnames if name in REQUIRED_SHEETS
        ] == REQUIRED_SHEETS
        assert workbook.active.title == "Дашборд"
        assert "Ликвидность МД" in workbook.sheetnames
        readme_rows = {
            row[0].value: row[1].value
            for row in workbook["README"].iter_rows(max_col=2)
            if row[0].value
        }
        unit_headers = [cell.value for cell in workbook["Юнит экономика"][1]]
        liquidity_sheet = workbook["Ликвидность МД"]
        liquidity_headers = [cell.value for cell in liquidity_sheet[1]]
        liquidity_values = [cell.value for cell in liquidity_sheet[2]]
        returns_headers = [cell.value for cell in workbook["Возвраты"][1]]
        lost_sales_headers = [cell.value for cell in workbook["Упущенные продажи"][1]]
        lost_sales_values = [cell.value for cell in workbook["Упущенные продажи"][2]]
        document_values = [cell.value for cell in workbook["Сверка документов 1С"][2]]
        tax_input_headers = [cell.value for cell in workbook["Сверка входящего НДС"][1]]
    finally:
        workbook.close()

    assert readme_rows["Источник"] == "Расчетные витрины отчета"
    assert (
        readme_rows["Статус готовности"] == "Недостаточное покрытие источников"
    )
    assert readme_rows["Происхождение данных"] == "DB-first витрины отчета"
    assert "Товар" in unit_headers
    assert "НДС входящий WB" in unit_headers
    assert "product" not in unit_headers
    assert "НДС входящий 1С" in tax_input_headers
    assert "Статус ликвидности" in liquidity_headers
    assert "Драйвер ликвидности" in liquidity_headers
    assert "Упр. прибыль" in liquidity_headers
    assert "liquidityStatus" not in liquidity_headers
    assert "product" not in liquidity_headers
    assert liquidity_values[liquidity_headers.index("Товар")] == "Товар"
    assert (
        liquidity_values[liquidity_headers.index("Статус ликвидности")]
        == "Прибыльный до 500 руб. в месяц"
    )
    assert returns_headers[:3] == ["Неделя", "Месяц", "Организация 1С"]
    assert "sourceStatus" not in lost_sales_headers
    assert (
        lost_sales_values[lost_sales_headers.index("Статус источника")] == "ОК"
    )
    assert document_values[0] == "ОК"


def test_db_first_excel_explains_incomplete_stock_history(tmp_path: Path) -> None:
    output = tmp_path / "report.xlsx"
    write_excel_from_marts(
        {
            "meta": {"client": "Клиент", "period": "01.03.2026 - 10.07.2026"},
            "readiness": {"status": "needs_review"},
            "unitRows": [],
            "lostSales": [],
            "lostSalesCoverage": {
                "calculated": False,
                "coveredDays": 92,
                "totalDays": 132,
                "message": "Не рассчитано: история остатков покрывает 92 из 132 дней",
            },
        },
        output,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        values = [
            cell.value
            for row in workbook["Упущенные продажи"].iter_rows()
            for cell in row
        ]
    finally:
        workbook.close()

    assert "Не рассчитано: история остатков покрывает 92 из 132 дней" in values


def test_db_first_excel_publishes_partial_stock_calculation_window(
    tmp_path: Path,
) -> None:
    output = tmp_path / "partial-window-report.xlsx"
    message = (
        "Рассчитано за доступный период: история остатков покрывает "
        "92 из 132 дней, без экстраполяции."
    )
    write_excel_from_marts(
        {
            "meta": {"client": "Клиент", "period": "01.03.2026 - 10.07.2026"},
            "readiness": {"status": "needs_review"},
            "unitRows": [],
            "lostSales": [],
            "lostSalesCoverage": {
                "calculated": True,
                "providerWindowCalculated": True,
                "fullCoverage": False,
                "coveredDays": 92,
                "totalDays": 132,
                "calculationPeriodStart": "2026-04-10",
                "calculationPeriodEnd": "2026-07-10",
                "extrapolated": False,
                "message": message,
            },
        },
        output,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        values = [
            cell.value
            for row in workbook["Упущенные продажи"].iter_rows()
            for cell in row
        ]
    finally:
        workbook.close()

    assert message in values


def test_ozon_diagnostics_excel_keeps_unmatched_1c_only_in_reconciliation(
    tmp_path: Path,
) -> None:
    output = tmp_path / "ozon.xlsx"
    write_ozon_diagnostics_excel(
        {
            "status": "ready",
            "message": "Ozon diagnostics ready",
            "latestRun": {
                "snapshotSetId": "snapshot-1",
                "periodStart": "2026-04-01",
                "periodEnd": "2026-04-30",
            },
            "expenseReconciliation": {
                "status": "review",
                "articleRows": [
                    {
                        "kind": "onec_unmatched",
                        "label": "1C без пары в Ozon: приходная 175",
                        "ozonAmount": 0.0,
                        "onecAmount": 550.0,
                        "deltaAmount": 550.0,
                        "includedInExpense": True,
                        "note": "Проверить соседний месяц mutual settlement.",
                    }
                ],
            },
            "ozonMart": {
                "totals": {
                    "onecRevenue": 1000.0,
                    "profitBeforeTax": 900.0,
                    "profitAfterTax": 846.0,
                    "vatOutput": 120.0,
                    "vatInput": 20.0,
                    "vatPayable": 100.0,
                    "revenueTax": 54.0,
                    "taxSystem": "УСН Доходы",
                    "taxProfileSource": "Catalog_Организации",
                    "taxCompleteness": "profile_complete",
                },
                "expenseAttribution": {
                    "status": "sku_direct",
                    "basis": "ozon_mutual_settlement_expense_documents",
                    "periodExpenseAmount": 100.0,
                    "skuAttributedExpenseAmount": 100.0,
                    "unattributedExpenseAmount": 0.0,
                    "allocatedUnattributedExpenseAmount": 0.0,
                    "overAttributedExpenseAmount": 0.0,
                    "periodExpenseDeltaAmount": 0.0,
                    "roundingDeltaAmount": 0.0,
                },
                "articleRows": [
                    {
                        "articleId": "revenue",
                        "label": "Выручка 1C Ozon SKU",
                        "group": "revenue",
                        "amount": 1000.0,
                        "effectAmount": 1000.0,
                    },
                    {
                        "articleId": "services",
                        "label": "Услуги Ozon",
                        "group": "services",
                        "amount": 100.0,
                        "effectAmount": -100.0,
                    },
                ],
                "rows": [
                    {
                        "offerId": "OZ-1",
                        "sku": "123",
                        "productName": "Товар",
                        "onecRevenue": 1000.0,
                        "ozonServices": 100.0,
                        "ozonPartnerServices": 0.0,
                        "profit": 900.0,
                        "profitBeforeTax": 900.0,
                        "marginBeforeTax": 0.9,
                        "revenueTax": 54.0,
                        "profitAfterTax": 846.0,
                        "marginAfterTax": 0.846,
                        "taxSystem": "УСН Доходы",
                        "taxProfileSource": "Catalog_Организации",
                        "taxCompleteness": "profile_complete",
                        "qualityStatus": "ready",
                    }
                ],
                "articleDrilldown": [
                    {
                        "kind": "sku_direct",
                        "articleId": "services",
                        "label": "Услуги Ozon",
                        "offerId": "OZ-1",
                        "sku": "123",
                        "productName": "Товар",
                        "amount": 100.0,
                        "effectAmount": -100.0,
                        "includedInSkuProfit": True,
                        "basis": "ozon_mutual_settlement_expense_documents",
                        "expenseBasis": "ozon_realization_sku_fields",
                        "attributionType": "sku_direct",
                        "status": "ready",
                    }
                ],
            },
        },
        output,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "Сводная Ozon",
            "Юнит экономика Ozon",
            "Начисления услуг Ozon",
            "Статьи по SKU",
            "Сверка Ozon 1C",
            "Методика",
        ]
        sku_values = [
            cell.value for row in workbook["Статьи по SKU"].iter_rows() for cell in row
        ]
        sku_headers = [cell.value for cell in workbook["Статьи по SKU"][1]]
        reconciliation_values = [
            cell.value
            for row in workbook["Сверка Ozon 1C"].iter_rows()
            for cell in row
        ]
        unit_headers = [cell.value for cell in workbook["Юнит экономика Ozon"][1]]
        summary_headers = [cell.value for cell in workbook["Сводная Ozon"][1]]
        summary_data = [
            [cell.value for cell in row]
            for row in workbook["Сводная Ozon"].iter_rows(min_row=2)
        ]
    finally:
        workbook.close()

    assert "Услуги партнеров / перевыставление" in unit_headers
    assert "База расхода" in unit_headers
    assert "Тип атрибуции" in unit_headers
    assert "Остаток периода" in unit_headers
    assert "Управленческая прибыль WB" in unit_headers
    assert "Прибыль до НДФЛ" in unit_headers
    assert "Налоговый режим" in unit_headers
    assert "Источник налогового профиля" in unit_headers
    assert "База расхода" in sku_headers
    assert "Тип атрибуции" in sku_headers
    assert "Остаток периода" in sku_headers
    assert "Расходы из Ozon detail" in summary_headers
    assert "Строк выручки без сопоставления" in summary_headers
    assert "Строк выручки с неоднозначным сопоставлением" in summary_headers
    assert "Режим порога существенности" in summary_headers
    assert "Минимальный месячный порог" in summary_headers
    assert "Максимальный месячный порог" in summary_headers
    article_index = summary_headers.index("Код статьи")
    effect_index = summary_headers.index("Влияние на прибыль")
    tax_effects = {
        row[article_index]: row[effect_index]
        for row in summary_data
        if row[article_index]
    }
    assert tax_effects["revenue_tax"] == -54.0
    assert tax_effects["vat_output"] is None
    assert tax_effects["vat_input"] is None
    assert tax_effects["vat_payable"] == -100.0
    assert tax_effects["profit_after_tax"] == 846.0
    assert "550" not in {str(value) for value in sku_values}
    assert "1C без пары в Ozon: приходная 175" in reconciliation_values
