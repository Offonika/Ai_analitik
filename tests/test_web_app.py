from __future__ import annotations

from copy import deepcopy
from dataclasses import replace
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import quote

import pytest
from cryptography.fernet import Fernet
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import event, select, text

from wb_unit_economics.logistics_analysis import (
    CHAIN_KEY_VERSION,
    LOGISTICS_CLASSIFIER_VERSION,
    LOGISTICS_FACTORS_METHODOLOGY_VERSION,
    LOGISTICS_MEASUREMENTS_METHODOLOGY_VERSION,
    LOGISTICS_METHODOLOGY_VERSION,
    LOGISTICS_ROUTES_METHODOLOGY_VERSION,
    LOGISTICS_TARIFFS_METHODOLOGY_VERSION,
    LogisticsSourceRow,
    UnitEconomicsSlice,
    build_dimension_rows,
    build_logistics_analysis,
    build_measurement_rows,
    build_route_rows,
    build_tariff_rows,
)
from wb_unit_economics.return_reason_analysis import build_return_reason_analysis
from wb_unit_economics.wb_goods_return import normalize_goods_return_source_row
from wb_unit_economics.wb_return_claims import normalize_claim_source_row
from wb_unit_economics.web import dashboard_payload, integrations, repository
from wb_unit_economics.web.ai import AiAnalyst
from wb_unit_economics.web.app import create_app
from wb_unit_economics.web.dashboard_payload import (
    analysis_period_text,
    document_reconciliation_rows,
    period_boundaries_from_label,
    period_label_from_value,
)
from wb_unit_economics.web.database import init_db, make_engine, make_session_factory
from wb_unit_economics.web.models import (
    OrganizationTaxProfile,
    OrganizationTaxProfileOverride,
    ReportLogisticsAnalysisContext,
    ReportLogisticsOrderRow,
    ReportLogisticsSkuRow,
    ReportLostSalesRow,
    SourceLoad,
    SourceRefreshRun,
    TenantIntegration,
    WbCabinet,
)
from wb_unit_economics.web.prompt_loader import load_prompt, render_prompt
from wb_unit_economics.web.refresh import (
    AutoRefreshBusyError,
    AutoRefreshUnavailableError,
    OnecAutoRefreshService,
)
from wb_unit_economics.web.repository import import_dashboard_payload, upsert_user
from wb_unit_economics.web.settings import WebSettings
from wb_unit_economics.web.source_refresh_worker import SourceRefreshWorkerLaunchError


@pytest.mark.parametrize(
    ("runtime_environment", "external_integrations_enabled", "expected_status"),
    (
        ("test", False, "ok"),
        ("test", True, "degraded"),
        ("production", False, "degraded"),
    ),
)
def test_health_accepts_missing_refresh_configuration_only_for_disabled_test(
    tmp_path: Path,
    runtime_environment: str,
    external_integrations_enabled: bool,
    expected_status: str,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "runtime_environment": runtime_environment,
            "external_integrations_enabled": external_integrations_enabled,
        },
    )
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="full-needs-configuration",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 10),
            reason="health configuration test",
        )
        repository.update_source_refresh_run(
            db,
            refresh,
            status="needs_configuration",
            finished_at=datetime(2026, 7, 10, 1, 0),
        )
        db.commit()

    payload = client.get("/api/health").json()

    assert payload["status"] == expected_status
    assert payload["sourceRefreshHealthStatus"] == "needs_configuration"


def test_health_is_degraded_while_new_run_follows_failed_completed_run(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        failed = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            mode="daily",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="daily-failed",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 9),
            reason="health test failure",
        )
        repository.update_source_refresh_run(
            db,
            failed,
            status="failed",
            finished_at=datetime(2026, 7, 10, 1, 0),
        )
        running = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            mode="daily",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="daily-running",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 9),
            reason="health test active",
        )
        repository.update_source_refresh_run(
            db,
            running,
            status="running",
            started_at=datetime(2026, 7, 10, 1, 1),
        )
        db.commit()

    response = client.get("/api/health")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "degraded"
    assert payload["latestSourceRefreshStatus"] == "running"
    assert payload["latestSourceRefreshActive"] is True
    assert payload["latestCompletedSourceRefreshStatus"] == "failed"
    assert payload["sourceRefreshHealthStatus"] == "failed"


def test_health_uses_most_recent_finish_not_later_blocked_attempt(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        completed = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="full-completed",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 7, 10),
            reason="completed after blocked attempt",
        )
        completed.created_at = datetime(2026, 7, 10, 16, 0)
        repository.update_source_refresh_run(
            db,
            completed,
            status="needs_review",
            finished_at=datetime(2026, 7, 10, 16, 17),
        )
        blocked = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            mode="daily",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="daily-blocked",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 10),
            reason="blocked while full was active",
        )
        blocked.created_at = datetime(2026, 7, 10, 16, 15)
        repository.update_source_refresh_run(
            db,
            blocked,
            status="blocked_active_refresh",
            finished_at=datetime(2026, 7, 10, 16, 15),
        )
        db.commit()

    payload = client.get("/api/health").json()

    assert payload["status"] == "ok"
    assert payload["latestSourceRefreshStatus"] == "blocked_active_refresh"
    assert payload["latestCompletedSourceRefreshStatus"] == "needs_review"
    assert payload["sourceRefreshHealthStatus"] == "needs_review"


def test_health_ignores_later_refresh_from_another_tenant(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        shumeyko_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            mode="daily",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="shumeyko-completed",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 10),
            reason="tenant-scoped health success",
        )
        repository.update_source_refresh_run(
            db,
            shumeyko_run,
            status="needs_review",
            finished_at=datetime(2026, 7, 10, 16, 0),
        )
        other_run = repository.create_source_refresh_run(
            db,
            tenant_id="other",
            mode="onec-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="other-failed",
            period_start=date(2026, 4, 1),
            period_end=date(2026, 4, 30),
            reason="other tenant canary",
        )
        repository.update_source_refresh_run(
            db,
            other_run,
            status="failed",
            finished_at=datetime(2026, 7, 10, 16, 1),
        )
        db.commit()

    payload = client.get("/api/health").json()

    assert payload["status"] == "ok"
    assert payload["sourceRefreshTenantId"] == "shumeyko"
    assert payload["latestSourceRefreshRunId"] == shumeyko_run.id
    assert payload["latestCompletedSourceRefreshStatus"] == "needs_review"


def test_health_exposes_safe_runtime_contour_and_maintenance_message(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "runtime_environment": "test",
            "maintenance_message": "Проверяем новую версию до 18:00.",
            "chatkit_enabled": True,
        },
    )

    payload = client.get("/api/health").json()

    assert payload["runtimeEnvironment"] == "test"
    assert payload["maintenanceMessage"] == "Проверяем новую версию до 18:00."
    assert payload["chatkitEnabled"] is True
    assert "chatkitDomainKey" not in payload


def test_test_contour_blocks_client_login_but_keeps_staff_access(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "runtime_environment": "test",
            "client_login_enabled": False,
        },
    )
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="client-only@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()

    rejected = client.post(
        "/api/auth/login",
        json={"email": "client-only@example.com", "password": "secret"},
    )
    staff = client.post(
        "/api/auth/login",
        json={"email": "admin@example.com", "password": "secret"},
    )

    assert rejected.status_code == 401
    assert staff.status_code == 200


def test_external_integration_master_switch_blocks_live_check(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"external_integrations_enabled": False},
    )
    login(client)

    response = client.post(
        "/api/integrations/wb_api/check",
        json={"tenant_id": "shumeyko"},
    )

    assert response.status_code == 409
    assert "Внешние проверки отключены" in response.json()["detail"]


def test_ozon_mapping_candidate_uses_later_precise_match_after_ambiguous() -> None:
    result = repository._check_ozon_mapping_candidate(
        {
            "rowNumber": 1,
            "sourceRowId": "ozon-row-1",
            "productName": "Ozon item",
            "offerId": "OZ-1",
            "barcode": "460000000001",
        },
        {
            "byOzonMarketplaceOffer": {
                "oz-1": [
                    {"id": "ITEM-1", "name": "Wrong candidate", "article": "OZ-1"},
                    {"id": "ITEM-2", "name": "Right candidate", "article": "OZ-1"},
                ]
            },
            "byOzonMarketplaceBarcode": {
                "460000000001": [
                    {
                        "id": "ITEM-2",
                        "name": "Right candidate",
                        "article": "OZ-1",
                    }
                ]
            },
        },
    )

    assert result["statusCounter"] == "matched"
    assert result["row"]["status"] == "matched"
    assert result["row"]["matchMethod"] == "onec_marketplace_ozon_barcode"
    assert result["row"]["onecItemId"] == "ITEM-2"


def test_ozon_mapping_candidate_uses_period_financials_to_narrow_ambiguous() -> None:
    result = repository._check_ozon_mapping_candidate(
        {
            "rowNumber": 1,
            "sourceRowId": "ozon-row-1",
            "productName": "Ozon item",
            "offerId": "OZ-1",
        },
        {
            "byArticle": {
                "oz-1": [
                    {"id": "ITEM-1", "name": "Old duplicate", "article": "OZ-1"},
                    {"id": "ITEM-2", "name": "Current item", "article": "OZ-1"},
                ]
            }
        },
        preferred_onec_item_ids={"ITEM-2"},
    )

    assert result["statusCounter"] == "matched"
    assert result["row"]["status"] == "matched"
    assert result["row"]["matchMethod"] == "offer_id_period_financials"
    assert result["row"]["onecItemId"] == "ITEM-2"


def test_ozon_expense_reconciliation_uses_api_expenses_and_onec_control() -> None:
    ozon = repository._ozon_cash_flow_expenses_payload(
        [
            SimpleNamespace(
                row_payload={
                    "details": [
                        {
                            "period": {
                                "begin": "2026-04-01T00:00:00Z",
                                "end": "2026-04-30T00:00:00Z",
                            },
                            "delivery": {"total": "1000"},
                            "services": {"total": "-100"},
                            "others": {"total": "10"},
                        }
                    ]
                }
            )
        ],
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )
    onec = repository._onec_incoming_invoice_expense_payload(
        [
            SimpleNamespace(
                row_payload={
                    "Date": "2026-04-30T00:00:00",
                    "Контрагент_Key": "OZON-CP",
                    "ВидОперации": "ПоступлениеОтПоставщика",
                    "СуммаДокумента": "95",
                }
            )
        ],
        counterparty_ids=["OZON-CP"],
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )
    reconciliation = repository._ozon_expense_reconciliation_payload(ozon, onec)

    assert ozon["summary"]["expenseAmount"] == 90.0
    assert ozon["summary"]["deliveryAmount"] == 1000.0
    assert ozon["summary"]["positiveAdjustmentAmount"] == 10.0
    assert {
        "category": "delivery",
        "label": "Ozon доставка / денежный блок",
        "signedAmount": 1000.0,
        "expenseEffectAmount": None,
        "includedInExpense": False,
        "note": "Не входит в расходы V1: это отдельный денежный блок.",
    } in ozon["categoryRows"]
    assert onec["amount"] == 95.0
    assert onec["operationRows"] == [
        {
            "operation": "ПоступлениеОтПоставщика",
            "amount": 95.0,
            "rowCount": 1,
            "includedInControl": True,
            "note": "Входит в 1C контроль расходов.",
        }
    ]
    assert reconciliation["status"] == "review"
    assert reconciliation["deltaAmount"] == 5.0
    assert reconciliation["detailRows"][-1] == {
        "kind": "total",
        "label": "Итого к расчету",
        "ozonAmount": 90.0,
        "ozonSignedAmount": None,
        "onecAmount": 95.0,
        "deltaAmount": 5.0,
        "includedInExpense": True,
        "note": "Дельта = 1C контроль минус Ozon API.",
    }


def test_ozon_expense_reconciliation_keeps_onec_control_without_ozon_expenses() -> None:
    ozon = {
        "status": "missing",
        "summary": {},
        "categoryRows": [],
        "topItems": [],
    }
    onec = repository._onec_incoming_invoice_expense_payload(
        [
            SimpleNamespace(
                row_payload={
                    "Date": "2026-04-30T00:00:00",
                    "Контрагент_Key": "OZON-CP",
                    "ВидОперации": "ПоступлениеОтПоставщика",
                    "СуммаДокумента": "550",
                }
            )
        ],
        counterparty_ids=["OZON-CP"],
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )

    reconciliation = repository._ozon_expense_reconciliation_payload(ozon, onec)

    assert reconciliation["status"] == "missing"
    assert reconciliation["ozonExpenseAmount"] is None
    assert reconciliation["onecExpenseAmount"] == 550.0
    assert any(
        item["kind"] == "onec_operation" for item in reconciliation["detailRows"]
    )


def test_ozon_mutual_settlement_expenses_use_document_rows() -> None:
    mutual = repository._ozon_mutual_settlement_expenses_payload(
        [
            SimpleNamespace(
                row_payload={
                    "Наименование": "Акт выполненных работ",
                    "Сумма дебиторской задолженности, RUR": "4898080.79",
                    "Сумма кредиторской задолженности, RUR": "",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Наименование": "Отчет о перевыставлении услуг",
                    "Сумма дебиторской задолженности, RUR": "535869.81",
                    "Сумма кредиторской задолженности, RUR": "",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Наименование": "Отчет о реализации",
                    "Сумма дебиторской задолженности, RUR": "151715.49",
                    "Сумма кредиторской задолженности, RUR": "26149512.63",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Наименование": "Фактическая оплата селлеров",
                    "Сумма дебиторской задолженности, RUR": "17286376.74",
                    "Сумма кредиторской задолженности, RUR": "",
                }
            ),
        ]
    )

    assert mutual["status"] == "loaded"
    assert mutual["basis"] == "ozon_mutual_settlement_expense_documents"
    assert mutual["summary"]["expenseAmount"] == 5433950.6
    assert any(
        item["label"] == "Отчет о реализации"
        and item["includedInExpense"] is False
        and item["expenseEffectAmount"] == 0.0
        for item in mutual["categoryRows"]
    )
    assert any(
        item["label"] == "Фактическая оплата селлеров"
        and item["includedInExpense"] is False
        for item in mutual["categoryRows"]
    )


def test_ozon_expense_reconciliation_shows_unmatched_onec_article() -> None:
    ozon = repository._ozon_mutual_settlement_expenses_payload(
        [
            SimpleNamespace(
                row_payload={
                    "Наименование": "Акт выполненных работ",
                    "Сумма дебиторской задолженности, RUR": "4898080.79",
                    "Сумма кредиторской задолженности, RUR": "",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Наименование": "Отчет о перевыставлении услуг",
                    "Сумма дебиторской задолженности, RUR": "535869.81",
                    "Сумма кредиторской задолженности, RUR": "",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Наименование": "Отчет о реализации",
                    "Сумма дебиторской задолженности, RUR": "151715.49",
                    "Сумма кредиторской задолженности, RUR": "",
                }
            ),
        ]
    )
    onec = repository._onec_incoming_invoice_expense_payload(
        [
            SimpleNamespace(
                row_payload={
                    "Date": "2026-04-30T00:00:00",
                    "Номер": "DOC-1",
                    "Контрагент_Key": "OZON-CP",
                    "ВидОперации": "ПоступлениеОтПоставщика",
                    "СуммаДокумента": "4898080.79",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Date": "2026-04-30T00:00:00",
                    "Номер": "DOC-2",
                    "Контрагент_Key": "OZON-CP",
                    "ВидОперации": "ПоступлениеОтПоставщика",
                    "СуммаДокумента": "535869.81",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Date": "2026-04-30T00:00:00",
                    "Номер": "DOC-3",
                    "Контрагент_Key": "OZON-CP",
                    "ВидОперации": "ПоступлениеОтПоставщика",
                    "СуммаДокумента": "151715.49",
                }
            ),
            SimpleNamespace(
                row_payload={
                    "Date": "2026-04-30T00:00:00",
                    "Номер": "DOC-4",
                    "Контрагент_Key": "OZON-CP",
                    "ВидОперации": "ПоступлениеОтПоставщика",
                    "СуммаДокумента": "550",
                }
            ),
        ],
        counterparty_ids=["OZON-CP"],
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )

    reconciliation = repository._ozon_expense_reconciliation_payload(ozon, onec)
    unmatched = [
        item
        for item in reconciliation["articleRows"]
        if item["kind"] == "onec_unmatched"
    ]

    assert reconciliation["deltaAmount"] == 550.0
    assert reconciliation["controlOnlyMatchedAmount"] == 151715.49
    assert reconciliation["onecComparableExpenseAmount"] == 5434500.6
    assert reconciliation["status"] == "review"
    assert "статьи без пары" in reconciliation["message"]
    assert len(unmatched) == 1
    assert unmatched[0]["ozonAmount"] == 0.0
    assert unmatched[0]["onecAmount"] == 550.0
    assert unmatched[0]["deltaAmount"] == 550.0
    assert "1C без пары в Ozon" in unmatched[0]["label"]
    assert "отчёт о взаиморасчётах за соседний месяц" in unmatched[0]["note"]
    assert any(
        item["kind"] == "control_matched"
        and item["onecAmount"] == 151715.49
        and item["includedInExpense"] is False
        for item in reconciliation["articleRows"]
    )


def test_ozon_period_from_output_file_accepts_mutual_settlement_xlsx() -> None:
    assert repository._ozon_period_from_output_file(
        "OZON_API_ozon_mutual_settlement_2026-04_file.raw.xlsx"
    ) == (date(2026, 4, 1), date(2026, 4, 30))


def test_ozon_cost_index_is_scoped_by_month_and_organization() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-30",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "2",
                        "Себестоимость": "200",
                        "ВходящийНДСИтого": "20",
                    },
                    {
                        "Period": "2026-05-31",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "2",
                        "Себестоимость": "600",
                        "ВходящийНДСИтого": "60",
                    },
                    {
                        "Period": "2026-04-30",
                        "Организация_Key": "ORG-2",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "2",
                        "Себестоимость": "1800",
                        "ВходящийНДСИтого": "180",
                    },
                ]
            }
        )
    ]

    april_org_1 = repository._onec_sales_cost_index(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
    )
    may_org_1 = repository._onec_sales_cost_index(
        rows,
        period_start=date(2026, 5, 1),
        period_end=date(2026, 5, 31),
        organization_id="ORG-1",
    )
    april_org_2 = repository._onec_sales_cost_index(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-2",
    )
    april_input_vat_org_1 = repository._onec_sales_input_vat_index(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
    )

    assert april_org_1 == {"ITEM-1": 100}
    assert may_org_1 == {"ITEM-1": 300}
    assert april_org_2 == {"ITEM-1": 900}
    assert april_input_vat_org_1 == {"ITEM-1": 20}


def sample_payload() -> dict:
    return {
        "meta": {
            "title": "Кабинет юнит-экономики WB",
            "client": "Шумейко и Партнеры",
            "period": "01.03.2026 - 17.06.2026",
            "reportPeriod": "01.03.2026 - 17.06.2026",
            "periodText": "март, апрель, май, июнь; июнь неполный, по 17.06.2026",
            "periodStatus": "предварительный: июнь неполный",
            "sourceCoverage": "01.03.2026 - 17.06.2026",
            "sourceCoverageStart": "2026-03-01",
            "sourceCoverageEnd": "2026-06-17",
            "methodologyVersion": "Excel MVP / test",
            "generatedAt": "20.06.2026 12:00",
            "sourceWorkbook": "shumeyko_wb_excel_mvp.xlsx",
            "returnReasonLimitation": (
                "Причина возврата не передается текущими источниками"
            ),
        },
        "options": {},
        "monthly": [],
        "expenses": [],
        "unitRows": [
            {
                "id": "unit-1",
                "week": "2026-04-06",
                "month": "Апрель 2026",
                "documentReport": (
                    "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
                ),
                "wbReportId": "726807272",
                "wbReportDate": "2026-04-13",
                "organization": "Организация A",
                "cabinet": "Кабинет A",
                "product": "Убыточный товар",
                "nmId": "1001",
                "articleWb": "WB-LOSS",
                "article1c": "A-LOSS",
                "barcode": "BAR-LOSS",
                "scheme": "FBO",
                "sales": 20,
                "returns": 8,
                "netQty": 12,
                "returnRate": 0.4,
                "revenueBeforeSpp": 100000,
                "spp": 1000,
                "revenue": 99000,
                "vat": 4714,
                "revenueWithoutVat": 94286,
                "cost": 65000,
                "commission": 10000,
                "logistics": 27000,
                "storage": 3000,
                "acceptance": 0,
                "promotion": 4000,
                "penalties": 0,
                "acquiring": 1200,
                "usn": 990,
                "profitBeforeTax": -9000,
                "profit": -14704,
                "margin": -0.1485,
                "unitProfit": -1225.3,
                "status": "ОК",
                "statusReason": "Данные достаточны для расчета",
                "sppStatus": "ОК",
                "lossClass": "Возвраты + логистика",
                "lossDriver": "Возвраты + логистика",
            },
            {
                "id": "unit-2",
                "week": "2026-06-02",
                "month": "Июнь 2026 (неполный месяц)",
                "documentReport": (
                    "Отчет комиссионера · 01.06.2026-07.06.2026 · закрытие 07.06.2026"
                ),
                "organization": "Организация B",
                "cabinet": "Кабинет B",
                "product": "Товар без себестоимости",
                "nmId": "1003",
                "articleWb": "WB-NOCOST",
                "article1c": "A-NOCOST",
                "barcode": "BAR-NOCOST",
                "scheme": "FBO",
                "sales": 5,
                "returns": 0,
                "netQty": 5,
                "returnRate": 0,
                "revenueBeforeSpp": 20000,
                "spp": 0,
                "revenue": 20000,
                "vat": 952,
                "revenueWithoutVat": 19048,
                "cost": 0,
                "commission": 2000,
                "logistics": 1500,
                "storage": 500,
                "acceptance": 0,
                "promotion": 0,
                "penalties": 0,
                "acquiring": 200,
                "usn": 200,
                "profitBeforeTax": 16000,
                "profit": 14648,
                "margin": 0.7324,
                "unitProfit": 2929.6,
                "status": "Нет себестоимости 1С",
                "statusReason": "Нет действующей себестоимости 1С",
                "sppStatus": "ОК",
                "lossClass": "Нужна проверка данных",
                "lossDriver": "Нет себестоимости 1С",
            },
        ],
        "returns": [],
        "lostSales": [
            {
                "id": "lost-1",
                "cabinet": "Кабинет A",
                "product": "Убыточный товар",
                "article1c": "A-LOSS",
                "barcode": "BAR-LOSS",
                "zeroStockDays": 10,
                "onecStock": 12,
                "onecWarehouses": "Собственный склад: 12",
                "sales": 20,
                "lostUnits": 5,
                "lostRevenue": 25000,
                "lostProfit": 3000,
                "note": "Сверить с 1С",
            }
        ],
        "reconciliation": [],
        "reconciliationMonthly": [
            {
                "month": "Апрель 2026",
                "wb_quantity": 90,
                "onec_quantity": 91,
                "quantity_delta": -1,
                "wb_cogs": 90000,
                "onec_cogs": 91000,
                "cogs_delta": -1000,
                "wb_mp_expenses": 55000,
                "onec_mp_expenses": 53000,
                "mp_expenses_delta": 2000,
                "comment": "Тестовая сверка",
            }
        ],
        "documentReconciliation": [
            {
                "id": "doc-recon-1",
                "status": "OK",
                "payoutStatus": "Нужен источник выплаты 1С",
                "periodStatus": "полный период",
                "documentReport": (
                    "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
                ),
                "salesPeriod": "2026-04-06 - 2026-04-12",
                "salesPeriodStart": "2026-04-06",
                "salesPeriodEnd": "2026-04-12",
                "expectedDocumentDate": "2026-04-12",
                "documentType": "Отчет комиссионера",
                "cabinet": "Кабинет A",
                "organization": "Организация A",
                "summaryReportId": "SUMMARY-1",
                "weeklySalesReportId": "SUMMARY-1",
                "weeklyBuyoutReportId": "BUYOUT-1",
                "wbReportIds": "726807272",
                "onecDocuments": "DOC-COMMISSIONER-1",
                "onecDocumentTypes": "ОтчетКомиссионера",
                "onecDocumentDates": "2026-04-12",
                "wbSalesQuantity": 22,
                "wbReturnQuantity": 2,
                "wbNetQuantity": 20,
                "onecSalesQuantity": 22,
                "onecReturnQuantity": 2,
                "onecNetQuantity": 20,
                "salesQuantityDelta": 0,
                "returnQuantityDelta": 0,
                "netQuantityDelta": 0,
                "wbQuantity": 20,
                "onecQuantity": 20,
                "quantityDelta": 0,
                "wbAmount": 99000,
                "onecAmount": 99000,
                "amountDelta": 0,
                "buyoutRetailAmountSum": None,
                "buyoutForPaySum": None,
                "buyoutBankPaymentSum": None,
                "onecExpenseInvoiceAmount": None,
                "buyoutRetailDelta": None,
                "buyoutForPayDelta": None,
                "buyoutBankDelta": None,
                "pdfBankPayment": 85000,
                "wbForPaySum": 85000,
                "onecSettlementTotal": 85000,
                "settlementDelta": 0,
                "onecVat": 0,
                "onecCogs": 600,
                "onecCogsWithoutVat": 550,
                "onecGrossProfit": 98400,
                "onecSourceRows": 10,
                "comment": "Документ совпал",
            }
        ],
    }


def test_tax_input_reconciliation_keeps_signed_charges_reversals_and_net() -> None:
    rows = [
        SimpleNamespace(
            week=date(2026, 3, 9),
            cabinet="Кабинет A",
            organization="Организация A",
            vat_input_from_wb=Decimal("616959.76"),
            vat_input_from_1c=Decimal("0"),
            vat_input_completeness="partial",
        ),
        SimpleNamespace(
            week=date(2026, 3, 2),
            cabinet="Кабинет A",
            organization="Организация A",
            vat_input_from_wb=Decimal("-501265.06"),
            vat_input_from_1c=Decimal("0"),
            vat_input_completeness="partial",
        ),
    ]

    result = repository._tax_input_reconciliation_payload_from_unit_rows(
        rows,
        tax_context={"vatDeductionMode": "unknown"},
    )

    assert len(result) == 2
    assert sum(item["vatInputFromWbCharges"] for item in result) == 616959.76
    assert sum(item["vatInputFromWbReversals"] for item in result) == -501265.06
    assert sum(item["vatInputFromWb"] for item in result) == pytest.approx(115694.70)
    assert all(item["onecEvidenceStatus"] == "missing" for item in result)
    assert all(item["vatInputCompleteness"] == "missing" for item in result)
    assert all(item["vatDeductionMode"] == "unknown" for item in result)


def test_tax_input_reconciliation_uses_organization_deduction_modes() -> None:
    rows = [
        SimpleNamespace(
            week=date(2026, 3, 9),
            cabinet="Кабинет A",
            organization="Организация ОСНО",
            vat_input_from_wb=Decimal("22"),
            vat_input_from_1c=Decimal("22"),
            vat_input_completeness="confirmed",
        ),
        SimpleNamespace(
            week=date(2026, 3, 9),
            cabinet="Кабинет B",
            organization="Организация УСН",
            vat_input_from_wb=Decimal("5"),
            vat_input_from_1c=Decimal("0"),
            vat_input_completeness="partial",
        ),
    ]
    tax_context = {
        "vatDeductionMode": "mixed",
        "profiles": [
            {
                "organization": "Организация ОСНО",
                "checks": [{"vatDeductionMode": "allowed"}],
            },
            {
                "organization": "Организация УСН",
                "checks": [{"vatDeductionMode": "not_allowed"}],
            },
        ],
    }

    result = repository._tax_input_reconciliation_payload_from_unit_rows(
        rows,
        tax_context=tax_context,
    )

    assert {
        item["organization"]: item["vatDeductionMode"] for item in result
    } == {
        "Организация ОСНО": "allowed",
        "Организация УСН": "not_allowed",
    }


def test_filtered_analytics_scopes_tax_input_reconciliation(tmp_path: Path) -> None:
    payload = deepcopy(sample_payload())
    payload["unitRows"][0].update(
        {
            "vatInputFromWb": 22,
            "vatInputFrom1c": 20,
            "vatInputCompleteness": "mismatch",
        }
    )
    payload["unitRows"][1].update(
        {
            "vatInputFromWb": 11,
            "vatInputFrom1c": 11,
            "vatInputCompleteness": "confirmed",
        }
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    april = client.get(
        "/api/reports/report-1/rows",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    ).json()
    cabinet_b = client.get(
        "/api/reports/report-1/rows",
        params={"wb_cabinet_id": "Кабинет B"},
    ).json()

    assert [
        item["cabinet"] for item in april["analytics"]["taxInputReconciliation"]
    ] == ["Кабинет A"]
    assert [
        item["cabinet"]
        for item in cabinet_b["analytics"]["taxInputReconciliation"]
    ] == ["Кабинет B"]


def ready_payload() -> dict:
    payload = deepcopy(sample_payload())
    payload["meta"] = {
        **payload["meta"],
        "period": "01.04.2026 - 30.04.2026",
        "periodText": "апрель 2026",
        "periodStatus": "",
        "sourceWorkbook": "ready-report.xlsx",
    }
    clean_rows = []
    for row in payload["unitRows"]:
        clean_rows.append(
            {
                **row,
                "month": "Апрель 2026",
                "status": "ОК",
                "statusReason": "Данные достаточны для расчета",
                "lossClass": "Без критичных проблем",
                "lossDriver": "Без критичных проблем",
            }
        )
    clean_rows[1] = {
        **clean_rows[1],
        "cost": 9000,
        "profit": 5648,
        "unitProfit": 1129.6,
    }
    payload["unitRows"] = clean_rows
    payload["documentReconciliation"] = [
        {
            **row,
            "payoutStatus": "",
            "periodStatus": "полный период",
            "comment": "Документ совпал",
        }
        for row in payload["documentReconciliation"]
    ]
    return payload


def client_ready_draft_text() -> str:
    return (
        "Ключевой вывод\n"
        "Отчет можно отправлять клиенту после стандартной проверки консультанта.\n\n"
        "Факты\n"
        "- Строки отчета имеют статус ОК.\n\n"
        "Что требует проверки\n"
        "- Дополнительных блокеров по данным нет.\n\n"
        "Ограничения\n"
        "- AI не меняет данные WB/1C и не выполняет отправку клиенту.\n\n"
        "Следующий шаг\n"
        "Передать отчет клиенту."
    )


class FakeAutoRefreshService:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.last_reason = ""

    def run(self, db, *, user, report, reason, thread_id=None):
        self.last_reason = reason
        try:
            refresh_run = repository.create_source_refresh_run(
                db,
                tenant_id=report.tenant_id,
                mode="onec-only",
                credential_source="tenant",
                dry_run=False,
                snapshot_set_id="onec-only-test",
                period_start=date(2026, 3, 1),
                period_end=date(2026, 6, 17),
                user=user,
                source_report=report,
                reason=reason,
            )
        except ValueError as exc:
            raise AutoRefreshBusyError(str(exc)) from exc
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="running",
            started_at=repository.security.utcnow(),
            root_dir="data/source_refresh/onec-only-test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_sales_register",
            source_label="AccumulationRegister_Продажи",
            required=True,
            status="loaded",
            snapshot_hash="hash",
            row_count=7,
            raw_path="data/source_refresh/onec-only-test/onec/sales_register.json",
        )
        repository.update_source_refresh_run(db, refresh_run, status="rebuilding")
        payload = sample_payload()
        payload["meta"] = {
            **payload["meta"],
            "sourceWorkbook": "auto-refresh.xlsx",
            "generatedAt": "20.06.2026 13:00",
        }
        payload["unitRows"][1] = {
            **payload["unitRows"][1],
            "status": "ОК",
            "statusReason": "Себестоимость обновлена из read-only 1С job",
            "cost": 9000,
            "profit": 5648,
        }
        report_id = f"{report.id}-refresh"
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook_path.write_bytes(b"auto-xlsx")
        new_report = repository.import_dashboard_payload(
            db,
            payload,
            tenant_id=report.tenant_id,
            tenant_name=report.client_name,
            report_id=report_id,
            source_workbook_path=str(self.workbook_path),
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="report_created",
            workbook_path=str(self.workbook_path),
            new_report_run_id=new_report.id,
            finished_at=repository.security.utcnow(),
        )
        repository.audit(
            db,
            action="source_refresh_report_created",
            user=user,
            tenant_id=report.tenant_id,
            entity_type="source_refresh_run",
            entity_id=refresh_run.id,
            payload={
                "source_report_run_id": report.id,
                "new_report_run_id": new_report.id,
                "status": "report_created",
            },
        )
        result = repository.source_refresh_run_payload(refresh_run)
        result["jobType"] = "source_refresh"
        result["sourceRefreshRunId"] = result["id"]
        if thread_id:
            result["threadId"] = thread_id
        return result


class FakeSourceRefreshService:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.calls = []

    def run(
        self,
        db,
        *,
        tenant_id,
        client_id=None,
        mode,
        credential_source,
        dry_run,
        user,
        reason,
        source_report=None,
        period_start=None,
        period_end=None,
        resume_mode="auto",
        resume_from_run_id=None,
    ):
        self.calls.append(
            {
                "tenant_id": tenant_id,
                "client_id": client_id,
                "mode": mode,
                "credential_source": credential_source,
                "dry_run": dry_run,
                "reason": reason,
                "period_start": period_start,
                "period_end": period_end,
                "resume_mode": resume_mode,
                "resume_from_run_id": resume_from_run_id,
            }
        )
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id=tenant_id,
            mode=mode,
            credential_source=credential_source,
            dry_run=dry_run,
            snapshot_set_id=("dry-run-test" if dry_run else "full-test"),
            period_start=period_start or date(2026, 3, 1),
            period_end=period_end or date(2026, 6, 17),
            client_id=client_id,
            user=user,
            source_report=source_report,
            reason=reason,
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="running",
            started_at=repository.security.utcnow(),
            root_dir="data/source_refresh/full-test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="sku_mapping",
            source_label="WB ↔ 1C mapping",
            required=True,
            status="loaded",
            snapshot_hash="hash",
            row_count=3,
            raw_path="data/source_refresh/full-test/mapping/manifest.json",
        )
        if dry_run:
            repository.update_source_refresh_run(
                db,
                refresh_run,
                status="dry_run_ready",
                finished_at=repository.security.utcnow(),
            )
            return repository.source_refresh_run_payload(refresh_run)

        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook_path.write_bytes(b"full-refresh-xlsx")
        new_report = repository.import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id=tenant_id,
            tenant_name="Refresh tenant",
            report_id="client-full-refresh-report",
            source_workbook_path=str(self.workbook_path),
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="report_created",
            workbook_path=str(self.workbook_path),
            new_report_run_id=new_report.id,
            finished_at=repository.security.utcnow(),
        )
        return repository.source_refresh_run_payload(refresh_run)

    def enqueue(
        self,
        db,
        *,
        tenant_id,
        client_id=None,
        mode,
        credential_source,
        user,
        reason,
        source_report=None,
        period_start=None,
        period_end=None,
        resume_mode="auto",
        resume_from_run_id=None,
    ):
        self.calls.append(
            {
                "tenant_id": tenant_id,
                "client_id": client_id,
                "mode": mode,
                "credential_source": credential_source,
                "dry_run": False,
                "reason": reason,
                "period_start": period_start,
                "period_end": period_end,
                "resume_mode": resume_mode,
                "resume_from_run_id": resume_from_run_id,
            }
        )
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id=tenant_id,
            mode=mode,
            credential_source=credential_source,
            dry_run=False,
            snapshot_set_id="full-test",
            period_start=period_start or date(2026, 3, 1),
            period_end=period_end or date(2026, 6, 17),
            client_id=client_id,
            user=user,
            source_report=source_report,
            reason=reason,
        )
        return repository.source_refresh_run_payload(refresh_run)

    def run_existing(self, db, refresh_run_id):
        refresh_run = db.get(SourceRefreshRun, refresh_run_id)
        assert refresh_run is not None
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="running",
            started_at=repository.security.utcnow(),
            root_dir="data/source_refresh/full-test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="sku_mapping",
            source_label="WB ↔ 1C mapping",
            required=True,
            status="loaded",
            snapshot_hash="hash",
            row_count=3,
            raw_path="data/source_refresh/full-test/mapping/manifest.json",
        )
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook_path.write_bytes(b"full-refresh-xlsx")
        new_report = repository.import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id=refresh_run.tenant_id,
            tenant_name="Refresh tenant",
            report_id="client-full-refresh-report",
            source_workbook_path=str(self.workbook_path),
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="report_created",
            workbook_path=str(self.workbook_path),
            new_report_run_id=new_report.id,
            finished_at=repository.security.utcnow(),
        )
        return repository.source_refresh_run_payload(refresh_run)


def make_client(
    tmp_path: Path,
    *,
    payload: dict | None = None,
    settings_overrides: dict | None = None,
    auto_refresh_service=None,
    publish_report: bool = True,
) -> TestClient:
    export = tmp_path / "reports" / "shumeyko_wb_excel_mvp.xlsx"
    export.parent.mkdir()
    export.write_bytes(b"xlsx")
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        import_dashboard_payload(
            db,
            payload or sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
            source_workbook_path=str(export),
            publication_status="published" if publish_report else "draft",
            publish=publish_report,
        )
        upsert_user(
            db,
            email="admin@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="admin",
        )
        import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="other",
            tenant_name="Другой клиент",
            report_id="other-report",
            source_workbook_path=str(export),
        )
        db.commit()
    settings_values = {
        "database_url": f"sqlite:///{tmp_path / 'web.sqlite3'}",
        "cookie_secure": False,
        "allowed_export_root": str(export.parent),
        "openai_api_key": "",
    }
    settings_values.update(settings_overrides or {})
    settings = WebSettings(**settings_values)
    app = create_app(
        settings=settings,
        session_factory=session_factory,
        auto_refresh_service=auto_refresh_service,
    )
    return TestClient(app)


def login(client: TestClient) -> None:
    login_as(client, "admin@example.com", "secret")


def login_as(client: TestClient, email: str, password: str) -> None:
    response = client.post(
        "/api/auth/login",
        json={"email": email, "password": password},
    )
    assert response.status_code == 200


def _ensure_logistics_dimensions(
    db,
    report,
    *,
    cabinet_id: str = "cabinet-logistics",
    company_id: str = "company-logistics",
    tenant_id: str | None = None,
    client_id: str | None = None,
) -> None:
    scope_tenant_id = tenant_id or report.tenant_id
    scope_client_id = client_id or report.client_id
    now = repository.security.utcnow()
    if db.get(repository.ClientCompany, company_id) is None:
        db.add(
            repository.ClientCompany(
                id=company_id,
                tenant_id=scope_tenant_id,
                client_id=scope_client_id,
                display_name=f"Организация {company_id}",
                source_key=f"test-{company_id}",
                status="active",
                created_at=now,
                updated_at=now,
            )
        )
        db.flush()
    if db.get(repository.WbCabinet, cabinet_id) is None:
        db.add(
            repository.WbCabinet(
                id=cabinet_id,
                tenant_id=scope_tenant_id,
                client_id=scope_client_id,
                client_company_id=company_id,
                display_name=f"Кабинет {cabinet_id}",
                cabinet_key=f"test-{cabinet_id}",
                provider="wb_api",
                status="active",
                created_at=now,
                updated_at=now,
            )
        )
        db.flush()


def _logistics_fixture_result(report, *, product_count: int = 1):
    source_rows = [
        LogisticsSourceRow(
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            wb_cabinet_id="cabinet-logistics",
            client_company_id="company-logistics",
            source_row_id="logistics-1",
            source_hash="safe-source-hash-1",
            financial_date=date(2026, 4, 6),
            order_date=date(2026, 4, 5),
            order_uid="external-order-must-not-leak",
            nm_id="101",
            sku="sku-101",
            vendor_code="A-101",
            product="Товар для проверки логистики",
            scheme="fbo",
            warehouse="Коледино",
            destination="Россия",
            document_type="Логистика",
            operation_name="Логистика",
            quantity=Decimal("0"),
            retail_amount=Decimal("0"),
            delivery_service=Decimal("10"),
            delivery_amount=Decimal("1"),
            return_amount=Decimal("0"),
            rebill_logistic_cost=Decimal("0"),
        )
    ]
    unit_rows = [
        UnitEconomicsSlice(
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            financial_week_start=date(2026, 4, 6),
            wb_cabinet_id="cabinet-logistics",
            client_company_id="company-logistics",
            scheme="fbo",
            nm_id="101",
            sku="sku-101",
            vendor_code="A-101",
            product="Товар для проверки логистики",
            revenue=Decimal("100"),
            profit_before_tax=Decimal("20"),
            logistics=Decimal("10"),
        )
    ]
    for index in range(2, product_count + 1):
        nm_id = str(100 + index)
        source_rows.append(
            replace(
                source_rows[0],
                source_row_id=f"logistics-{index}",
                source_hash=f"safe-source-hash-{index}",
                order_uid=f"external-order-{index}",
                nm_id=nm_id,
                sku=f"sku-{nm_id}",
                vendor_code=f"A-{nm_id}",
                product=f"Товар {nm_id}",
            )
        )
        unit_rows.append(
            replace(
                unit_rows[0],
                nm_id=nm_id,
                sku=f"sku-{nm_id}",
                vendor_code=f"A-{nm_id}",
                product=f"Товар {nm_id}",
            )
        )
    return build_logistics_analysis(source_rows, unit_rows)


def _return_reason_fixture_result(
    report,
    *,
    claims_state: str = "access_denied",
):
    finance_row = LogisticsSourceRow(
        tenant_id=report.tenant_id,
        client_id=report.client_id,
        wb_cabinet_id="cabinet-logistics",
        client_company_id="company-logistics",
        source_row_id="return-finance-row",
        source_hash="safe-return-finance-hash",
        financial_date=date(2026, 4, 7),
        order_date=date(2026, 4, 5),
        order_uid="return-order-must-not-leak",
        nm_id="101",
        sku="sku-101",
        vendor_code="A-101",
        product="Товар для возврата",
        scheme="fbo",
        warehouse="Коледино",
        destination="Россия",
        document_type="Возврат",
        operation_name="Возврат",
        quantity=Decimal("-1"),
        retail_amount=Decimal("-100"),
        delivery_service=Decimal("10"),
        delivery_amount=Decimal("0"),
        return_amount=Decimal("0"),
        rebill_logistic_cost=Decimal("0"),
        finance_srid="finance-srid-safe-internal",
    )
    base = build_logistics_analysis(
        [finance_row],
        [
            UnitEconomicsSlice(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                financial_week_start=date(2026, 4, 6),
                wb_cabinet_id="cabinet-logistics",
                client_company_id="company-logistics",
                scheme="fbo",
                nm_id="101",
                sku="sku-101",
                vendor_code="A-101",
                product="Товар для возврата",
                revenue=Decimal("-100"),
                profit_before_tax=Decimal("-20"),
                logistics=Decimal("10"),
            )
        ],
    )
    logistics_result = replace(
        base,
        order_rows=tuple(
            replace(
                row,
                logistics_reverse=Decimal("10"),
                return_quantity=Decimal("1"),
            )
            for row in base.order_rows
        ),
    )
    goods_rows = [
        normalize_goods_return_source_row(
            {
                "srid": "finance-srid-safe-internal",
                "order_id": "provider-order-must-not-leak",
                "nm_id": "101",
                "barcode": "provider-barcode-must-not-leak",
                "reason": "Не подошёл размер",
                "status": "returned",
                "return_type": "buyer",
            },
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            wb_cabinet_id="cabinet-logistics",
        )
    ]
    claim_rows = (
        [
            normalize_claim_source_row(
                {
                    "srid": "finance-srid-safe-internal",
                    "nm_id": "101",
                    "is_archive": False,
                    "has_user_comment": True,
                },
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="cabinet-logistics",
            )
        ]
        if claims_state == "confirmed_nonempty"
        else []
    )
    return_reason_result = build_return_reason_analysis(
        [finance_row],
        logistics_result.order_rows,
        goods_rows,
        claim_rows,
        goods_return_snapshot_hash="safe-goods-return-snapshot",
        claims_snapshot_hash="safe-claims-snapshot",
        goods_return_coverage_start=date(2026, 4, 1),
        goods_return_coverage_end=date(2026, 4, 30),
        claims_coverage_start=date(2026, 4, 1),
        claims_coverage_end=date(2026, 4, 14),
        claims_source_status=claims_state,
        claims_review_reasons=(
            ("return_claims_source_access_denied",)
            if claims_state == "access_denied"
            else ()
        ),
    )
    return logistics_result, return_reason_result


def _dimension_context(report, rows, *, data_status: str = "ready"):
    return {
        "tenant_id": report.tenant_id,
        "client_id": report.client_id,
        "factor_methodology_version": LOGISTICS_FACTORS_METHODOLOGY_VERSION,
        "data_status": data_status,
        "input_hash": "safe-dimension-input-hash",
        "source_snapshot_hash": "safe-card-snapshot-hash",
        "source_loaded_at": datetime(2026, 7, 20, 12, 0),
        "source_row_count": len(rows),
        "dimension_row_count": len(rows),
        "matched_product_count": sum(
            row["coverage_status"] == "ready" for row in rows
        ),
        "missing_product_count": sum(
            row["coverage_status"] == "missing_dimensions" for row in rows
        ),
        "invalid_product_count": sum(
            row["coverage_status"] in {"invalid_dimensions", "identity_conflict"}
            for row in rows
        ),
        "conflicting_product_count": sum(
            row["coverage_status"] == "conflicting_dimensions" for row in rows
        ),
        "signal_product_count": sum(
            row["dimensions_valid"] is False for row in rows
        ),
        "blocking_reasons": [],
        "review_reasons": [] if data_status == "ready" else ["test_partial"],
        "created_at": datetime(2026, 7, 20, 12, 1),
    }


def _tariff_context(report, rows, *, data_status: str = "ready"):
    points: dict[tuple, set[str]] = {}
    for row in rows:
        point = (
            row["wb_cabinet_id"],
            row["client_company_id"],
            row["scheme"],
            row["financial_week_start"],
            row["tariff_type"],
        )
        points.setdefault(point, set()).add(row["evidence_type"])
    factual = sum("fact" in values for values in points.values())
    estimated = sum(
        "fact" not in values and "estimate" in values
        for values in points.values()
    )
    expected = len(points)
    return {
        "tenant_id": report.tenant_id,
        "client_id": report.client_id,
        "factor_methodology_version": LOGISTICS_TARIFFS_METHODOLOGY_VERSION,
        "data_status": data_status,
        "input_hash": "safe-tariff-input-hash",
        "source_snapshot_hash": "safe-tariff-snapshot-hash",
        "source_loaded_at": datetime(2026, 7, 21, 12, 0),
        "factor_snapshot_date": date(2026, 7, 21),
        "source_row_count": len(rows),
        "tariff_row_count": len(rows),
        "expected_point_count": expected,
        "factual_point_count": factual,
        "estimated_point_count": estimated,
        "unavailable_point_count": expected - factual - estimated,
        "invalid_row_count": sum(
            row["coverage_status"] == "invalid_tariff" for row in rows
        ),
        "conflicting_row_count": sum(
            row["coverage_status"] == "conflicting_tariff" for row in rows
        ),
        "warehouse_count": len(
            {row["warehouse"] for row in rows if row["warehouse"]}
        ),
        "blocking_reasons": [],
        "review_reasons": [] if data_status == "ready" else ["test_partial"],
        "created_at": datetime(2026, 7, 21, 12, 1),
    }


def _route_context(report, rows, *, data_status: str = "ready"):
    chains = {row["chain_key"] for row in rows}
    matched = {
        row["chain_key"] for row in rows if row["coverage_status"] == "ready"
    }
    conflicting = {
        row["chain_key"]
        for row in rows
        if row["coverage_status"] == "conflicting_route"
    }
    total = sum((row["logistics_total"] for row in rows), Decimal("0"))
    linked = sum(
        (
            row["logistics_total"]
            for row in rows
            if row["coverage_status"] == "ready"
        ),
        Decimal("0"),
    )
    return {
        "tenant_id": report.tenant_id,
        "client_id": report.client_id,
        "factor_methodology_version": LOGISTICS_ROUTES_METHODOLOGY_VERSION,
        "data_status": data_status,
        "input_hash": "safe-route-input-hash",
        "source_snapshot_hash": "safe-route-snapshot-hash",
        "source_loaded_at": datetime(2026, 7, 21, 12, 0),
        "source_coverage_start": date(2026, 4, 1),
        "source_coverage_end": date(2026, 4, 30),
        "source_row_count": len(rows),
        "route_row_count": len(rows),
        "total_chain_count": len(chains),
        "matched_chain_count": len(matched),
        "missing_chain_count": len(chains - matched - conflicting),
        "conflicting_chain_count": len(conflicting),
        "warehouse_count": len(
            {row["warehouse"] for row in rows if row["warehouse_status"] == "ready"}
        ),
        "destination_count": len(
            {
                row["destination"]
                for row in rows
                if row["destination_status"] == "ready"
            }
        ),
        "total_logistics": total,
        "linked_logistics": linked,
        "reconciliation_delta": Decimal("0"),
        "blocking_reasons": [],
        "review_reasons": [] if data_status == "ready" else ["test_partial"],
        "created_at": datetime(2026, 7, 21, 12, 1),
    }


def _measurement_context(report, rows, *, data_status: str = "ready"):
    return {
        "tenant_id": report.tenant_id,
        "client_id": report.client_id,
        "factor_methodology_version": LOGISTICS_MEASUREMENTS_METHODOLOGY_VERSION,
        "data_status": data_status,
        "input_hash": "safe-measurement-input-hash",
        "penalty_source_snapshot_hash": "safe-penalty-snapshot-hash",
        "warehouse_source_snapshot_hash": "safe-warehouse-snapshot-hash",
        "penalty_source_loaded_at": datetime(2026, 7, 21, 12, 0),
        "warehouse_source_loaded_at": datetime(2026, 7, 21, 12, 0),
        "factor_snapshot_at": datetime(2026, 7, 21, 12, 0),
        "source_coverage_start": report.period_start,
        "source_coverage_end": report.period_end,
        "expected_endpoint_count": 2,
        "complete_endpoint_count": 2,
        "unavailable_endpoint_count": 0,
        "source_event_count": len(rows),
        "provider_event_count": len(rows),
        "measurement_row_count": len(rows),
        "scoped_product_count": 2,
        "product_with_event_count": len(
            {row["product_ref"] for row in rows if row.get("product_ref")}
        ),
        "matched_event_count": sum(bool(row.get("product_ref")) for row in rows),
        "unmatched_event_count": sum(
            row["coverage_status"] == "unmatched_product" for row in rows
        ),
        "ambiguous_event_count": sum(
            row["coverage_status"] == "ambiguous_product_scope" for row in rows
        ),
        "invalid_event_count": sum(
            row["coverage_status"] == "invalid_measurement" for row in rows
        ),
        "conflicting_event_count": sum(
            row["coverage_status"] == "conflicting_measurement" for row in rows
        ),
        "penalty_event_count": sum(
            (row.get("penalty_amount") or 0) > 0 for row in rows
        ),
        "reversal_event_count": sum(
            (row.get("reversal_amount") or 0) > 0 for row in rows
        ),
        "warehouse_only_event_count": sum(
            row.get("event_kind") == "warehouse_measurement" for row in rows
        ),
        "blocking_reasons": [],
        "review_reasons": [] if data_status == "ready" else ["test_partial"],
        "created_at": datetime(2026, 7, 21, 12, 1),
    }


def persist_logistics_fixture(client: TestClient) -> None:
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        repository.replace_report_logistics_analysis(
            db,
            report,
            _logistics_fixture_result(report),
        )
        db.commit()


def test_dimension_mart_persist_round_trip(tmp_path: Path) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        result = _logistics_fixture_result(report)
        card_rows = [
            {
                "wb_cabinet_id": "cabinet-logistics",
                "nm_id": "101",
                "length_cm": 30,
                "width_cm": 20,
                "height_cm": 10,
                "weight_brutto_kg": 2,
                "dimensions_valid": False,
            }
        ]
        rows = build_dimension_rows(result.sku_rows, card_rows)
        _ensure_logistics_dimensions(db, report)

        context = repository.replace_report_logistics_dimension_analysis(
            db,
            report,
            context=_dimension_context(report, rows),
            rows=rows,
        )
        db.commit()
        assert context.dimension_row_count == len(rows)

        stored = repository.report_logistics_dimension_rows(db, report.id)
        by_nm = {row.nm_id: row for row in stored}
        assert "101" in by_nm
        assert by_nm["101"].length_cm == Decimal("30")
        assert by_nm["101"].volume_l == Decimal("6")
        assert by_nm["101"].dimensions_valid is False
        assert by_nm["101"].evidence_type == "fact"
        assert by_nm["101"].measured_penalty_amount is None

        # повторная запись перезаписывает срез (delete + insert)
        repository.replace_report_logistics_dimension_analysis(
            db,
            report,
            context=_dimension_context(report, rows),
            rows=rows,
        )
        db.commit()
        assert len(repository.report_logistics_dimension_rows(db, report.id)) == len(
            rows
        )


def test_logistics_dimensions_api_returns_persisted_mart(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        result = _logistics_fixture_result(report)
        rows = build_dimension_rows(
            result.sku_rows,
            [
                {
                    "wb_cabinet_id": "cabinet-logistics",
                    "nm_id": "101",
                    "length_cm": 30,
                    "width_cm": 20,
                    "height_cm": 10,
                    "weight_brutto_kg": 2,
                    "dimensions_valid": False,
                }
            ],
        )
        _ensure_logistics_dimensions(db, report)
        repository.replace_report_logistics_analysis(db, report, result)
        repository.replace_report_logistics_dimension_analysis(
            db,
            report,
            context=_dimension_context(report, rows),
            rows=rows,
        )
        db.commit()

    response = client.get("/api/reports/report-1/logistics/dimensions")

    assert response.status_code == 200
    body = response.json()
    assert body["reportId"] == "report-1"
    assert body["coverage"]["total"] >= 1
    row = next(item for item in body["rows"] if item["nmId"] == "101")
    assert Decimal(row["lengthCm"]) == Decimal("30")
    assert Decimal(row["volumeL"]) == Decimal("6")
    assert row["dimensionsValid"] is False
    assert row["evidenceType"] == "fact"
    assert row["measuredPenaltyAmount"] is None


def test_logistics_dimensions_api_is_feature_gated(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)
    response = client.get("/api/reports/report-1/logistics/dimensions")
    assert response.status_code == 404


def test_logistics_dimensions_api_partial_coverage_uses_full_filtered_slice(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report, product_count=2)
        repository.replace_report_logistics_analysis(db, report, result)
        rows = build_dimension_rows(
            result.sku_rows,
            [
                {
                    "wb_cabinet_id": "cabinet-logistics",
                    "nm_id": "101",
                    "length_cm": 30,
                    "width_cm": 20,
                    "height_cm": 10,
                    "weight_brutto_kg": 2,
                    "dimensions_valid": False,
                }
            ],
        )
        repository.replace_report_logistics_dimension_analysis(
            db,
            report,
            context=_dimension_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.commit()

    response = client.get(
        "/api/reports/report-1/logistics/dimensions",
        params={"limit": 1, "sortBy": "product", "sortOrder": "asc"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["dataStatus"] == "partial"
    assert body["sliceStatus"] == "partial"
    assert body["total"] == 2
    assert len(body["rows"]) == 1
    assert body["coverage"] == {
        "total": 2,
        "withDimensions": 1,
        "missingDimensions": 1,
        "invalidDimensions": 0,
        "conflictingDimensions": 0,
        "signalCount": 1,
        "coveragePct": 50,
    }
    assert {item["evidenceType"] for item in body["recommendations"]} == {
        "limitation",
        "data_unavailable",
    }
    assert "sourceHashDigest" not in body["rows"][0]
    assert "sourceSnapshotHash" not in body

    empty = client.get(
        "/api/reports/report-1/logistics/dimensions",
        params={"periodStart": "2026-06-01", "periodEnd": "2026-06-07"},
    ).json()
    assert empty["sliceStatus"] == "empty"
    assert empty["coverage"]["total"] == 0

    filtered = client.get(
        "/api/reports/report-1/logistics/dimensions",
        params={"product": "102"},
    ).json()
    assert filtered["total"] == 1
    assert filtered["rows"][0]["nmId"] == "102"
    assert filtered["coverage"]["missingDimensions"] == 1

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsDimensionContext, "report-1")
        assert context is not None
        context.tenant_id = "other"
        db.commit()
    blocked = client.get("/api/reports/report-1/logistics/dimensions").json()
    assert blocked["dataStatus"] == "blocked"
    assert blocked["sliceStatus"] == "blocked"
    assert blocked["rows"] == []


def test_logistics_dimensions_role_and_flag_matrix(tmp_path: Path) -> None:
    staff_path = tmp_path / "staff"
    client_path = tmp_path / "client"
    staff_path.mkdir()
    client_path.mkdir()
    staff = make_client(
        staff_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
        },
    )
    login(staff)
    assert staff.get("/api/reports/report-1/logistics/dimensions").status_code == 200
    assert staff.get("/api/me").json()["logisticsFactorsEnabled"] is True

    client = make_client(
        client_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_analysis_client_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_factors_client_enabled": False,
        },
    )
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="factor-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "factor-client@example.com", "secret")
    assert client.get("/api/reports/report-1/logistics/dimensions").status_code == 404
    assert client.get("/api/me").json()["logisticsFactorsEnabled"] is False

    client.app.state.settings.logistics_factors_client_enabled = True
    assert client.get("/api/reports/report-1/logistics/dimensions").status_code == 200
    assert client.get("/api/me").json()["logisticsFactorsEnabled"] is True


def test_logistics_return_reasons_api_states_filters_and_safe_payload(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_return_reasons_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    path = "/api/reports/report-1/logistics/return-reasons"
    legacy = client.get(path).json()
    assert legacy["dataStatus"] == "needs_rebuild"
    assert legacy["sliceStatus"] == "needs_rebuild"
    assert legacy["rows"] == []

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        logistics_result, return_reason_result = _return_reason_fixture_result(report)
        repository.replace_report_logistics_analysis(
            db,
            report,
            logistics_result,
        )
        context = repository.replace_report_logistics_return_reason_analysis(
            db,
            report,
            return_reason_result,
            goods_return_source_loaded_at=datetime(2026, 7, 23, 10, 0),
            claims_source_loaded_at=datetime(2026, 7, 23, 10, 1),
        )
        db.commit()
        assert context.return_reason_row_count == 1

    response = client.get(path)

    assert response.status_code == 200
    body = response.json()
    assert body["dataStatus"] == "partial"
    assert body["sliceStatus"] == "partial"
    assert body["total"] == 1
    assert body["coverage"] == {
        "totalReturnChains": 1,
        "reasonAvailable": 1,
        "reasonUnavailable": 0,
        "claimAvailable": 0,
        "hasUserComment": 0,
        "claimCoverageUnknown": 1,
        "reasonCoveragePct": 100,
        "claimCoveragePct": 0,
    }
    assert body["sourceCoverage"]["claims"]["status"] == "access_denied"
    assert (
        body["sourceCoverage"]["claims"]["message"]
        == "Источник заявок недоступен"
    )
    row = body["rows"][0]
    assert row["reasonCategory"] == "Не подошёл размер"
    assert row["reasonSource"] == "goods_return"
    assert row["evidenceType"] == "fact"
    assert row["claimAvailable"] is None
    assert row["hasUserComment"] is None
    assert len(row["chainRef"]) == 12
    rendered = response.text
    for forbidden in (
        "finance-srid-safe-internal",
        "provider-order-must-not-leak",
        "provider-barcode-must-not-leak",
        "sourceHashDigest",
        "snapshotHash",
        "rowHash",
    ):
        assert forbidden not in rendered

    filtered = client.get(
        path,
        params={
            "product": "возврата",
            "reasonSource": "goods_return",
            "evidenceType": "fact",
            "matchStatus": "ready",
        },
    ).json()
    assert filtered["total"] == 1
    assert filtered["coverage"]["reasonAvailable"] == 1

    empty = client.get(
        path,
        params={"periodStart": "2026-04-20", "periodEnd": "2026-04-25"},
    ).json()
    assert empty["sliceStatus"] == "empty"
    assert empty["coverage"]["totalReturnChains"] == 0

    with client.app.state.session_factory() as db:
        context = db.get(
            repository.ReportLogisticsReturnReasonContext,
            "report-1",
        )
        assert context is not None
        context.return_reason_row_count = 2
        db.commit()
    blocked = client.get(path).json()
    assert blocked["dataStatus"] == "blocked"
    assert blocked["sliceStatus"] == "blocked"
    assert blocked["rows"] == []


def test_logistics_return_reasons_exact_claim_activates_safe_booleans(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_return_reasons_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        logistics_result, return_reason_result = _return_reason_fixture_result(
            report,
            claims_state="confirmed_nonempty",
        )
        repository.replace_report_logistics_analysis(db, report, logistics_result)
        repository.replace_report_logistics_return_reason_analysis(
            db,
            report,
            return_reason_result,
        )
        db.commit()

    body = client.get(
        "/api/reports/report-1/logistics/return-reasons"
    ).json()
    assert body["dataStatus"] == "ready"
    assert body["sliceStatus"] == "ready"
    assert body["coverage"]["claimAvailable"] == 1
    assert body["coverage"]["hasUserComment"] == 1
    assert body["rows"][0]["claimAvailable"] is True
    assert body["rows"][0]["hasUserComment"] is True


def test_logistics_return_reason_analysis_is_atomic_and_published_immutable(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        _logistics_result, result = _return_reason_fixture_result(report)
        repository.replace_report_logistics_return_reason_analysis(
            db,
            report,
            result,
        )
        db.flush()
        persisted_uids = list(
            db.scalars(
                select(repository.ReportLogisticsReturnReasonRow.row_uid).where(
                    repository.ReportLogisticsReturnReasonRow.report_run_id
                    == report.id
                )
            )
        )

        invalid_row = replace(result.rows[0], tenant_id="other-tenant")
        invalid_result = replace(result, rows=(invalid_row,))
        with pytest.raises(ValueError, match="tenant does not match report"):
            repository.replace_report_logistics_return_reason_analysis(
                db,
                report,
                invalid_result,
            )
        assert list(
            db.scalars(
                select(repository.ReportLogisticsReturnReasonRow.row_uid).where(
                    repository.ReportLogisticsReturnReasonRow.report_run_id
                    == report.id
                )
            )
        ) == persisted_uids
        assert report.logistics_return_reasons_required is True

        report.publication_status = "published"
        with pytest.raises(ValueError, match="published logistics return-reason"):
            repository.replace_report_logistics_return_reason_analysis(
                db,
                report,
                result,
            )


def test_logistics_return_reasons_role_and_flag_matrix(tmp_path: Path) -> None:
    (tmp_path / "staff").mkdir()
    (tmp_path / "client").mkdir()
    path = "/api/reports/report-1/logistics/return-reasons"
    staff = make_client(
        tmp_path / "staff",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_return_reasons_enabled": False,
        },
    )
    login(staff)
    assert staff.get(path).status_code == 404
    assert staff.get("/api/me").json()["logisticsReturnReasonsEnabled"] is False
    staff.app.state.settings.logistics_return_reasons_enabled = True
    assert staff.get(path).status_code == 200
    assert staff.get("/api/me").json()["logisticsReturnReasonsEnabled"] is True

    client = make_client(
        tmp_path / "client",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_analysis_client_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_factors_client_enabled": True,
            "logistics_return_reasons_enabled": True,
            "logistics_return_reasons_client_enabled": False,
        },
    )
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="return-reason-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "return-reason-client@example.com", "secret")
    assert client.get(path).status_code == 404
    me = client.get("/api/me").json()
    assert me["logisticsReturnReasonsEnabled"] is False
    assert me["logisticsReturnReasonsClientEnabled"] is False
    client.app.state.settings.logistics_return_reasons_client_enabled = True
    assert client.get(path).status_code == 200
    me = client.get("/api/me").json()
    assert me["logisticsReturnReasonsEnabled"] is True
    assert me["logisticsReturnReasonsClientEnabled"] is True


def test_logistics_measurements_api_states_filters_and_full_slice_coverage(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_measurements_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    legacy = client.get("/api/reports/report-1/logistics/measurements").json()
    assert legacy["dataStatus"] == "needs_rebuild"
    assert legacy["sliceStatus"] == "needs_rebuild"
    assert legacy["rows"] == []

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report, product_count=2)
        repository.replace_report_logistics_analysis(db, report, result)
        penalty_rows = [
            {
                "tenant_id": report.tenant_id,
                "client_id": report.client_id,
                "wb_cabinet_id": "cabinet-logistics",
                "dim_id": "dimension-event-one-must-not-leak",
                "nm_id": "101",
                "volume": "2.5",
                "width": "10",
                "length": "25",
                "height": "10",
                "volume_sup": "2",
                "width_sup": "10",
                "length_sup": "20",
                "height_sup": "10",
                "prc_over": "125",
                "dt_bonus": "2026-04-07T10:00:00Z",
                "is_valid": False,
                "is_valid_dt": "2026-04-07T11:00:00Z",
                "penalty_amount": "10",
                "reversal_amount": "5",
                "source_hash": "penalty-source-hash-must-not-leak",
            },
            {
                "tenant_id": report.tenant_id,
                "client_id": report.client_id,
                "wb_cabinet_id": "cabinet-logistics",
                "dim_id": "dimension-event-unmatched-must-not-leak",
                "nm_id": "999",
                "volume": "1",
                "width": "10",
                "length": "10",
                "height": "10",
                "prc_over": "110",
                "dt_bonus": "2026-04-08T10:00:00Z",
                "penalty_amount": "3",
                "reversal_amount": "0",
                "source_hash": "unmatched-source-hash-must-not-leak",
            },
            {
                "tenant_id": report.tenant_id,
                "client_id": report.client_id,
                "wb_cabinet_id": "cabinet-logistics",
                "dim_id": "dimension-event-invalid-date-must-not-leak",
                "nm_id": "101",
                "volume": "1",
                "width": "10",
                "length": "10",
                "height": "10",
                "dt_bonus": "not-a-timestamp",
                "source_hash": "invalid-source-hash-must-not-leak",
            },
        ]
        warehouse_rows = [
            {
                "tenant_id": report.tenant_id,
                "client_id": report.client_id,
                "wb_cabinet_id": "cabinet-logistics",
                "dim_id": "dimension-event-one-must-not-leak",
                "nm_id": "101",
                "volume": "2.5",
                "width": "10",
                "length": "25",
                "height": "10",
                "dt": "2026-04-07T09:00:00Z",
                "source_hash": "warehouse-source-hash-must-not-leak",
            },
            {
                "tenant_id": report.tenant_id,
                "client_id": report.client_id,
                "wb_cabinet_id": "cabinet-logistics",
                "dim_id": "dimension-event-two-must-not-leak",
                "nm_id": "102",
                "volume": "1.5",
                "width": "10",
                "length": "15",
                "height": "10",
                "dt": "2026-04-09T09:00:00Z",
                "source_hash": "warehouse-two-hash-must-not-leak",
            },
        ]
        rows = build_measurement_rows(
            result.sku_rows,
            penalty_rows,
            warehouse_rows,
        )
        repository.replace_report_logistics_measurement_analysis(
            db,
            report,
            context=_measurement_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.commit()

    response = client.get(
        "/api/reports/report-1/logistics/measurements",
        params={"limit": 1, "sortBy": "penaltyAmount", "sortOrder": "desc"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["dataStatus"] == "partial"
    assert body["sliceStatus"] == "partial"
    assert body["total"] == 4
    assert len(body["rows"]) == 1
    assert body["coverage"] == {
        "expectedEndpoints": 2,
        "completeEndpoints": 2,
        "unavailableEndpoints": 0,
        "scopedProducts": 2,
        "productsWithEvents": 2,
        "totalEvents": 4,
        "penaltyEvents": 2,
        "reversalEvents": 1,
        "warehouseOnlyEvents": 1,
        "matchedEvents": 3,
        "unmatchedEvents": 1,
        "ambiguousEvents": 0,
        "invalidEvents": 1,
        "conflictingEvents": 0,
        "measurementIncidencePercent": 100,
    }
    assert body["accountingTreatment"]["includedInFinancialKpi"] is False
    assert body["rows"][0]["includedInFinancialKpi"] is False
    assert body["rows"][0]["accountingReconciliationStatus"] == "unreconciled"
    assert {item["evidenceType"] for item in body["recommendations"]} == {
        "fact",
        "data_unavailable",
    }
    serialized = response.text
    for forbidden in (
        "dimension-event-one-must-not-leak",
        "penalty-source-hash-must-not-leak",
        "warehouse-source-hash-must-not-leak",
        "invalid-source-hash-must-not-leak",
        "sourceHashDigest",
        "sourceSnapshotHash",
        "dimId",
        "nmId",
            "photoUrls",
            "sellerAccountId",
        ):
        assert forbidden not in serialized

    penalty_only = client.get(
        "/api/reports/report-1/logistics/measurements",
        params={"hasPenalty": "true"},
    ).json()
    assert penalty_only["total"] == 2
    assert penalty_only["coverage"]["penaltyEvents"] == 2

    warehouse_only = client.get(
        "/api/reports/report-1/logistics/measurements",
        params={"eventKind": "warehouse_measurement"},
    ).json()
    assert warehouse_only["total"] == 1
    assert warehouse_only["rows"][0]["eventKind"] == "warehouse_measurement"

    product_filtered = client.get(
        "/api/reports/report-1/logistics/measurements",
        params={"product": "Товар 102"},
    ).json()
    assert product_filtered["total"] == 1
    assert product_filtered["coverage"]["scopedProducts"] == 1
    assert product_filtered["coverage"]["measurementIncidencePercent"] == 100

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsMeasurementContext, "report-1")
        assert context is not None
        unmatched = (
            db.query(repository.ReportLogisticsMeasurementRow)
            .filter_by(
                report_run_id="report-1",
                coverage_status="unmatched_product",
            )
            .one()
        )
        invalid = (
            db.query(repository.ReportLogisticsMeasurementRow)
            .filter_by(
                report_run_id="report-1",
                coverage_status="invalid_measurement",
            )
            .one()
        )
        db.delete(unmatched)
        db.delete(invalid)
        context.measurement_row_count = 2
        context.data_status = "ready"
        context.review_reasons = []
        db.commit()
    ready = client.get("/api/reports/report-1/logistics/measurements").json()
    assert ready["dataStatus"] == "ready"
    assert ready["sliceStatus"] == "ready"
    assert ready["total"] == 2

    empty = client.get(
        "/api/reports/report-1/logistics/measurements",
        params={"periodStart": "2026-06-01", "periodEnd": "2026-06-07"},
    ).json()
    assert empty["sliceStatus"] == "empty"
    assert empty["coverage"]["totalEvents"] == 0

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsMeasurementContext, "report-1")
        assert context is not None
        context.measurement_row_count = 3
        db.commit()
    blocked = client.get("/api/reports/report-1/logistics/measurements").json()
    assert blocked["dataStatus"] == "blocked"
    assert blocked["sliceStatus"] == "blocked"
    assert blocked["rows"] == []


def test_logistics_measurements_role_and_flag_matrix(tmp_path: Path) -> None:
    (tmp_path / "staff").mkdir()
    (tmp_path / "client").mkdir()
    staff = make_client(
        tmp_path / "staff",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_measurements_enabled": False,
        },
    )
    login(staff)
    path = "/api/reports/report-1/logistics/measurements"
    assert staff.get(path).status_code == 404
    assert staff.get("/api/me").json()["logisticsMeasurementsEnabled"] is False
    staff.app.state.settings.logistics_measurements_enabled = True
    assert staff.get(path).status_code == 200
    assert staff.get("/api/me").json()["logisticsMeasurementsEnabled"] is True

    client = make_client(
        tmp_path / "client",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_analysis_client_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_factors_client_enabled": True,
            "logistics_measurements_enabled": True,
            "logistics_measurements_client_enabled": False,
        },
    )
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="measurement-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "measurement-client@example.com", "secret")
    assert client.get(path).status_code == 404
    assert client.get("/api/me").json()["logisticsMeasurementsEnabled"] is False
    client.app.state.settings.logistics_measurements_client_enabled = True
    assert client.get(path).status_code == 200
    assert client.get("/api/me").json()["logisticsMeasurementsEnabled"] is True


def test_published_dimension_mart_is_immutable(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        with pytest.raises(ValueError, match="published logistics dimension"):
            repository.replace_report_logistics_dimension_analysis(
                db,
                report,
                context=_dimension_context(report, []),
                rows=[],
            )


def test_logistics_tariffs_api_partial_coverage_uses_full_filtered_slice(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_tariffs_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report, product_count=2)
        repository.replace_report_logistics_analysis(db, report, result)
        source_rows = [
            {
                "wb_cabinet_id": "cabinet-logistics",
                "requested_date": "2026-04-06",
                "tariff_type": "box",
                "warehouse_name": warehouse,
                "box_delivery_base": "48",
                "box_delivery_liter": "11,2",
                "box_delivery_coef_expr": "125",
                "box_storage_base": "0,14",
                "box_storage_liter": "0,07",
                "box_storage_coef_expr": "115",
                "source_hash": f"box-{index}",
            }
            for index, warehouse in enumerate(("Склад A", "Склад B"), 1)
        ]
        source_rows.append(
            {
                "wb_cabinet_id": "cabinet-logistics",
                "requested_date": "2026-07-21",
                "tariff_type": "pallet",
                "warehouse_name": "Склад A",
                "pallet_delivery_expr": "170",
                "pallet_delivery_value_base": "51",
                "pallet_storage_expr": "155",
                "pallet_storage_value_expr": "35.65",
                "source_hash": "pallet-current",
            }
        )
        rows = build_tariff_rows(
            result.sku_rows,
            source_rows,
            factor_snapshot_date=date(2026, 7, 21),
        )
        repository.replace_report_logistics_tariff_analysis(
            db,
            report,
            context=_tariff_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.commit()

    response = client.get(
        "/api/reports/report-1/logistics/tariffs",
        params={"limit": 1, "sortBy": "warehouse", "sortOrder": "asc"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["dataStatus"] == "partial"
    assert body["sliceStatus"] == "partial"
    assert body["total"] == 3
    assert len(body["rows"]) == 1
    assert body["coverage"] == {
        "expectedPoints": 2,
        "factualPoints": 1,
        "estimatedPoints": 1,
        "unavailablePoints": 0,
        "invalidRows": 0,
        "conflictingRows": 0,
        "warehouses": 2,
        "factualCoveragePct": 50,
    }
    assert body["financialEffect"] is None
    assert body["recommendations"][0]["evidenceType"] == "limitation"
    assert "sourceHashDigest" not in body["rows"][0]
    assert "sourceSnapshotHash" not in body
    assert "wbCabinetId" not in body["rows"][0]

    box = client.get(
        "/api/reports/report-1/logistics/tariffs",
        params={"tariffType": "box", "warehouse": "Склад B"},
    ).json()
    assert box["total"] == 1
    assert box["rows"][0]["warehouse"] == "Склад B"
    assert box["rows"][0]["evidenceType"] == "fact"

    empty = client.get(
        "/api/reports/report-1/logistics/tariffs",
        params={"periodStart": "2026-06-01", "periodEnd": "2026-06-07"},
    ).json()
    assert empty["sliceStatus"] == "empty"
    assert empty["coverage"]["expectedPoints"] == 0

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsTariffContext, "report-1")
        assert context is not None
        pallet = (
            db.query(repository.ReportLogisticsTariffRow)
            .filter_by(report_run_id="report-1", tariff_type="pallet")
            .one()
        )
        pallet.evidence_type = "fact"
        pallet.tariff_date = pallet.requested_date
        context.data_status = "ready"
        db.commit()
    ready = client.get("/api/reports/report-1/logistics/tariffs").json()
    assert ready["dataStatus"] == "ready"
    assert ready["sliceStatus"] == "ready"
    assert ready["coverage"]["factualCoveragePct"] == 100

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsTariffContext, "report-1")
        assert context is not None
        context.tenant_id = "other"
        db.commit()
    blocked = client.get("/api/reports/report-1/logistics/tariffs").json()
    assert blocked["dataStatus"] == "blocked"
    assert blocked["rows"] == []


def test_logistics_tariffs_role_and_flag_matrix(tmp_path: Path) -> None:
    (tmp_path / "staff").mkdir()
    (tmp_path / "client").mkdir()
    staff = make_client(
        tmp_path / "staff",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_tariffs_enabled": False,
        },
    )
    login(staff)
    assert staff.get("/api/reports/report-1/logistics/tariffs").status_code == 404
    assert staff.get("/api/me").json()["logisticsTariffsEnabled"] is False
    staff.app.state.settings.logistics_tariffs_enabled = True
    assert staff.get("/api/reports/report-1/logistics/tariffs").status_code == 200
    assert staff.get("/api/me").json()["logisticsTariffsEnabled"] is True

    client = make_client(
        tmp_path / "client",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_analysis_client_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_factors_client_enabled": True,
            "logistics_tariffs_enabled": True,
            "logistics_tariffs_client_enabled": False,
        },
    )
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="tariff-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "tariff-client@example.com", "secret")
    assert client.get("/api/reports/report-1/logistics/tariffs").status_code == 404
    assert client.get("/api/me").json()["logisticsTariffsEnabled"] is False
    client.app.state.settings.logistics_tariffs_client_enabled = True
    assert client.get("/api/reports/report-1/logistics/tariffs").status_code == 200
    assert client.get("/api/me").json()["logisticsTariffsEnabled"] is True


def test_logistics_routes_api_partial_coverage_uses_full_filtered_slice(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_routes_enabled": True,
        },
        publish_report=False,
    )
    login(client)
    legacy = client.get("/api/reports/report-1/logistics/routes").json()
    assert legacy["dataStatus"] == "needs_rebuild"
    assert legacy["sliceStatus"] == "needs_rebuild"
    assert legacy["rows"] == []
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report, product_count=2)
        repository.replace_report_logistics_analysis(db, report, result)
        rows = build_route_rows(
            result.order_rows,
            [
                {
                    "tenant_id": report.tenant_id,
                    "client_id": report.client_id,
                    "wb_cabinet_id": "cabinet-logistics",
                    "srid": "external-order-must-not-leak",
                    "nm_id": "101",
                    "warehouse_name": "Склад A",
                    "country_name": "Страна",
                    "region_name": "Регион A",
                    "source_hash": "route-source-must-not-leak",
                }
            ],
        )
        repository.replace_report_logistics_route_analysis(
            db,
            report,
            context=_route_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.commit()

    response = client.get(
        "/api/reports/report-1/logistics/routes",
        params={"limit": 1, "sortBy": "logisticsTotal", "sortOrder": "desc"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["dataStatus"] == "partial"
    assert body["sliceStatus"] == "partial"
    assert body["total"] == 2
    assert len(body["rows"]) == 1
    assert body["coverage"] == {
        "totalChains": 2,
        "matchedChains": 1,
        "missingChains": 1,
        "conflictingChains": 0,
        "linkedLogistics": 10,
        "unlinkedLogistics": 10,
        "warehouses": 1,
        "destinations": 1,
        "coveragePct": 50,
    }
    assert body["financialEffect"] is None
    assert body["recommendations"][0]["evidenceType"] == "fact"
    assert body["recommendations"][1]["evidenceType"] == "data_unavailable"
    serialized = response.text
    for forbidden in (
        "external-order-must-not-leak",
        "route-source-must-not-leak",
        "sourceHashDigest",
        "sourceSnapshotHash",
        "chainKey",
        "nmId",
    ):
        assert forbidden not in serialized

    filtered = client.get(
        "/api/reports/report-1/logistics/routes",
        params={"product": "Товар 102"},
    ).json()
    assert filtered["coverage"]["totalChains"] == 1
    assert filtered["coverage"]["missingChains"] == 1
    assert filtered["rows"][0]["evidenceType"] == "data_unavailable"

    fully_filtered = client.get(
        "/api/reports/report-1/logistics/routes",
        params={
            "periodStart": "2026-04-01",
            "periodEnd": "2026-04-30",
            "wbCabinetId": "cabinet-logistics",
            "clientCompanyId": "company-logistics",
            "scheme": "fbo",
            "product": "A-101",
            "warehouse": "Склад A",
            "destination": "Регион A",
            "sortBy": "warehouse",
            "sortOrder": "asc",
        },
    ).json()
    assert fully_filtered["total"] == 1
    assert fully_filtered["coverage"]["totalChains"] == 1
    assert fully_filtered["coverage"]["matchedChains"] == 1
    assert fully_filtered["rows"][0]["warehouse"] == "Склад A"

    empty = client.get(
        "/api/reports/report-1/logistics/routes",
        params={"periodStart": "2026-06-01", "periodEnd": "2026-06-07"},
    ).json()
    assert empty["sliceStatus"] == "empty"
    assert empty["coverage"]["totalChains"] == 0

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsRouteContext, "report-1")
        assert context is not None
        missing_row = (
            db.query(repository.ReportLogisticsRouteRow)
            .filter_by(report_run_id="report-1", coverage_status="data_unavailable")
            .one()
        )
        missing_row.warehouse = "Склад B"
        missing_row.warehouse_status = "ready"
        missing_row.destination = "Страна · Регион B"
        missing_row.destination_status = "ready"
        missing_row.evidence_type = "fact"
        missing_row.coverage_status = "ready"
        missing_row.data_quality_status = "ready"
        context.data_status = "ready"
        context.matched_chain_count = 2
        context.missing_chain_count = 0
        context.review_reasons = []
        db.commit()
    ready = client.get("/api/reports/report-1/logistics/routes").json()
    assert ready["dataStatus"] == "ready"
    assert ready["sliceStatus"] == "ready"
    assert ready["coverage"]["coveragePct"] == 100

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsRouteContext, "report-1")
        assert context is not None
        context.factor_methodology_version = "wb-logistics-routes-legacy"
        db.commit()
    outdated = client.get("/api/reports/report-1/logistics/routes").json()
    assert outdated["dataStatus"] == "needs_rebuild"
    assert outdated["rows"] == []

    with client.app.state.session_factory() as db:
        context = db.get(repository.ReportLogisticsRouteContext, "report-1")
        assert context is not None
        context.factor_methodology_version = LOGISTICS_ROUTES_METHODOLOGY_VERSION
        context.tenant_id = "other"
        db.commit()
    blocked = client.get("/api/reports/report-1/logistics/routes").json()
    assert blocked["dataStatus"] == "blocked"
    assert blocked["rows"] == []


def test_logistics_routes_role_and_flag_matrix(tmp_path: Path) -> None:
    (tmp_path / "staff").mkdir()
    (tmp_path / "client").mkdir()
    staff = make_client(
        tmp_path / "staff",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_routes_enabled": False,
        },
    )
    login(staff)
    assert staff.get("/api/reports/report-1/logistics/routes").status_code == 404
    assert staff.get("/api/me").json()["logisticsRoutesEnabled"] is False
    staff.app.state.settings.logistics_routes_enabled = True
    assert staff.get("/api/reports/report-1/logistics/routes").status_code == 200
    assert staff.get("/api/me").json()["logisticsRoutesEnabled"] is True

    client = make_client(
        tmp_path / "client",
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_analysis_client_enabled": True,
            "logistics_factors_enabled": True,
            "logistics_factors_client_enabled": True,
            "logistics_routes_enabled": True,
            "logistics_routes_client_enabled": False,
        },
    )
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="route-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "route-client@example.com", "secret")
    assert client.get("/api/reports/report-1/logistics/routes").status_code == 404
    assert client.get("/api/me").json()["logisticsRoutesEnabled"] is False
    client.app.state.settings.logistics_routes_client_enabled = True
    assert client.get("/api/reports/report-1/logistics/routes").status_code == 200
    assert client.get("/api/me").json()["logisticsRoutesEnabled"] is True


def test_logistics_api_is_feature_gated_and_old_report_needs_rebuild(
    tmp_path: Path,
) -> None:
    disabled_path = tmp_path / "disabled"
    enabled_path = tmp_path / "enabled"
    disabled_path.mkdir()
    enabled_path.mkdir()
    disabled = make_client(disabled_path)
    login(disabled)
    assert disabled.get("/api/reports/report-1/logistics/summary").status_code == 404
    assert (
        "logisticsAnalysis" not in disabled.get("/api/reports/report-1/summary").json()
    )
    assert disabled.get("/api/me").json()["logisticsAnalysisEnabled"] is False

    enabled = make_client(
        enabled_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    login(enabled)
    response = enabled.get("/api/reports/report-1/logistics/summary")

    assert response.status_code == 200
    payload = response.json()
    assert payload["dataStatus"] == "needs_rebuild"
    assert payload["sliceStatus"] == "needs_rebuild"
    assert payload["reportCoverage"] is None
    assert payload["kpis"]["logisticsTotal"] is None
    assert payload["dynamics"] == []
    assert enabled.get("/api/me").json()["logisticsAnalysisEnabled"] is True


def test_logistics_api_returns_reconciled_safe_staff_payload(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    persist_logistics_fixture(client)
    login(client)

    cabinet = client.get("/cabinet")
    script = client.get("/static/app.js")
    styles = client.get("/static/styles.css")
    assert 'data-workspace-nav="logistics"' not in cabinet.text
    assert 'data-workspace-nav="tables"' in cabinet.text
    assert 'id="logistics-scenario-nav"' in cabinet.text
    assert 'data-table-scenario-nav="logistics"' in cabinet.text
    assert 'data-table-scenario-panel="logistics"' in cabinet.text
    assert "Аналитика и таблицы" in cabinet.text
    assert "Логистика: расходы и зоны проверки" in cabinet.text
    assert 'id="table-scenario-kpi-grid"' in cabinet.text
    assert 'id="table-scenario-summary-status"' in cabinet.text
    assert 'id="logistics-state-message"' in cabinet.text
    assert 'id="logistics-data-status"' in cabinet.text
    assert 'id="logistics-trust-freshness"' in cabinet.text
    assert 'id="logistics-trust-low-sample"' in cabinet.text
    assert 'id="logistics-state-action"' in cabinet.text
    assert 'id="logistics-products-rows"' in cabinet.text
    assert 'id="logistics-products-pagination"' in cabinet.text
    assert 'id="logistics-dimensions"' in cabinet.text
    assert 'id="logistics-dimensions-coverage"' in cabinet.text
    assert 'id="logistics-dimensions-rows"' in cabinet.text
    assert 'id="logistics-dimensions-pagination"' in cabinet.text
    assert "Габариты в карточке WB" in cabinet.text
    assert "не исторический" in cabinet.text
    assert 'id="logistics-measurements"' in cabinet.text
    assert 'id="logistics-measurements-coverage"' in cabinet.text
    assert 'id="logistics-measurements-rows"' in cabinet.text
    assert 'id="logistics-measurements-pagination"' in cabinet.text
    assert "Контрольные замеры и удержания WB" in cabinet.text
    assert "не прибавляются к расходам" in cabinet.text
    assert ".logistics-dimensions-table {" in styles.text
    assert "table-layout: fixed;" in styles.text
    assert ".logistics-dimensions-table th:nth-child(5)" in styles.text
    assert cabinet.text.index('id="logistics-dimensions"') < cabinet.text.index(
        'id="logistics-measurements"'
    )
    assert cabinet.text.index('id="logistics-measurements"') < cabinet.text.index(
        'aria-labelledby="logistics-products-title"'
    )
    assert 'id="logistics-orders-rows"' in cabinet.text
    assert 'id="logistics-orders-pagination"' in cabinet.text
    assert cabinet.text.index('id="logistics-kpi-grid"') < cabinet.text.index(
        'id="logistics-recommendations"'
    )
    assert cabinet.text.index('id="logistics-recommendations"') < cabinet.text.index(
        'id="logistics-trust-strip"'
    )
    assert cabinet.text.index('id="logistics-trust-strip"') < cabinet.text.index(
        'id="logistics-filter-form"'
    )
    assert 'id="logistics-scheme-filter"' in cabinet.text
    assert 'id="logistics-product-filter"' in cabinet.text
    # Разрез по организации совпадает с выбором кабинета WB наверху
    # (1 организация = 1 кабинет), поэтому отдельный контрол не выводится.
    assert 'id="logistics-organization-filter"' not in cabinet.text
    assert "logisticsOrganizationFilter" not in script.text
    assert "function loadLogisticsDimensions" in script.text
    assert "function resetLogisticsDimensions" in script.text
    assert "function loadLogisticsMeasurements" in script.text
    assert "function resetLogisticsMeasurements" in script.text
    assert "Замеры временно недоступны. Основная логистика" in script.text
    assert "Сигнал записи WB" in script.text
    assert "Справочные суммы" in script.text
    assert "state.logisticsMeasurementsOffset = 0;" in script.text
    assert "state.logisticsMeasurementsRequestId += 1;" in script.text
    assert "Основная логистика продолжает работать" in script.text
    assert "measuredPenaltyAmount" not in script.text
    assert "loadLogisticsAnalysis" in script.text
    assert "function renderTableScenarioSummary" in script.text
    assert "Текущий отчёт собран до появления витрины логистики v5" in script.text
    assert 'reportId: params.get("report_id") || ""' in script.text
    assert "item.id === requestedReportId" in script.text
    assert "state.logisticsProductsTotal = Number(products.total || 0)" in script.text
    assert "state.logisticsOrdersTotal = Number(payload.total || 0)" in script.text
    assert "logisticsProfitEffectText(item.profitEffectAmount)" in script.text
    assert "Финансовая связь с отчётом отсутствует" in script.text
    assert "Финансовая связь отсутствует" in script.text
    assert "Корректировка — схема не применяется" in script.text
    assert 'dataStatus === "partial" && sliceStatus === "ready"' in script.text
    assert 'normalize(item.dataQualityStatus) === "missing_profit_link"' in script.text
    assert '? "Проверить данные"' in script.text
    assert '"Основание / ограничение"' in script.text
    assert '"Что сделать"' in script.text
    assert 'status === "empty"' in script.text

    full_week = {"periodStart": "2026-04-06", "periodEnd": "2026-04-12"}
    summary = client.get("/api/reports/report-1/logistics/summary", params=full_week)
    products = client.get(
        "/api/reports/report-1/logistics/products",
        params={**full_week, "product": "проверки", "limit": 10000},
    )

    assert summary.status_code == 200
    assert summary.json()["dataStatus"] == "ready"
    assert summary.json()["sliceStatus"] == "ready"
    assert summary.json()["financialMetricStatus"] == "ready"
    assert summary.json()["methodologyVersion"] == LOGISTICS_METHODOLOGY_VERSION
    assert summary.json()["classifierVersion"] == LOGISTICS_CLASSIFIER_VERSION
    assert summary.json()["chainKeyVersion"] == CHAIN_KEY_VERSION
    assert datetime.fromisoformat(summary.json()["generatedAt"])
    assert summary.json()["coverage"]["lowSampleProductCount"] == 1
    assert summary.json()["filterContext"]["dateGrain"] == "calendar_day"
    assert summary.json()["reportCoverage"]["maxDimensionDelta"] == 0
    assert summary.json()["reportCoverage"]["invalidReportRows"] == 0
    assert summary.json()["reportCoverage"]["reportRequiredFieldErrors"] == 0
    assert summary.json()["reportCoverage"]["chainDimensionConflicts"] == 0
    assert summary.json()["reportCoverage"]["invalidSourcePayloadShapes"] == 0
    assert summary.json()["reportCoverage"]["sourceIdentityErrors"] == 0
    assert summary.json()["reportCoverage"]["sourceRevisionConflicts"] == 0
    assert summary.json()["reportCoverage"]["scopeMismatches"] == 0
    assert summary.json()["kpis"]["logisticsTotal"] == 10
    assert summary.json()["kpis"]["logisticsSharePct"] == 10
    assert summary.json()["components"] == {
        "forward": 10,
        "reverse": 0,
        "adjustment": 0,
        "unclassified": 0,
    }
    recommendation = summary.json()["recommendations"][0]
    assert {
        "impactAmount",
        "evidenceType",
        "actionTarget",
        "actionLabel",
    } <= recommendation.keys()
    assert products.status_code == 200
    assert products.json()["limit"] == 1000
    assert products.json()["items"][0]["lowSample"] is True
    product_ref = products.json()["items"][0]["productRef"]
    orders = client.get(
        "/api/reports/report-1/logistics/orders",
        params={**full_week, "productRef": product_ref},
    )
    assert orders.status_code == 200
    assert orders.json()["items"][0]["chainRef"]
    serialized = f"{summary.text}{products.text}{orders.text}"
    assert "external-order-must-not-leak" not in serialized
    assert "safe-source-hash-1" not in serialized

    partial = client.get(
        "/api/reports/report-1/logistics/summary",
        params={"periodStart": "2026-04-06", "periodEnd": "2026-04-06"},
    ).json()
    assert partial["kpis"]["logisticsTotal"] == 10
    assert partial["financialMetricStatus"] == "not_available_partial_week"
    assert partial["kpis"]["revenue"] is None
    assert partial["kpis"]["logisticsSharePct"] is None
    assert partial["kpis"]["profitBeforeTax"] is None
    assert partial["kpis"]["profitEffectAmount"] is None
    assert partial["rankings"]["byProfitEffect"] == []
    partial_products = client.get(
        "/api/reports/report-1/logistics/products",
        params={"periodStart": "2026-04-06", "periodEnd": "2026-04-06"},
    ).json()
    assert partial_products["financialMetricStatus"] == "not_available_partial_week"
    assert partial_products["items"][0]["profitEffectAmount"] is None

    empty = client.get(
        "/api/reports/report-1/logistics/summary",
        params={**full_week, "product": "товар-которого-нет"},
    ).json()
    assert empty["sliceStatus"] == "empty"
    assert empty["financialMetricStatus"] == "not_available_empty_slice"
    assert empty["kpis"]["logisticsTotal"] is None
    assert empty["components"]["forward"] is None
    assert empty["recommendations"] == []


def test_logistics_correction_segment_does_not_count_as_order(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        order_row = db.query(ReportLogisticsOrderRow).one()
        order_row.countable_order = False
        order_row.scheme = "not_applicable"
        sku_row = db.query(ReportLogisticsSkuRow).one()
        sku_row.scheme = "not_applicable"
        db.commit()
    login(client)

    filters = {
        "periodStart": "2026-04-06",
        "periodEnd": "2026-04-12",
        "scheme": "not_applicable",
    }
    summary = client.get(
        "/api/reports/report-1/logistics/summary", params=filters
    ).json()
    products = client.get(
        "/api/reports/report-1/logistics/products", params=filters
    ).json()

    assert summary["kpis"]["logisticsTotal"] == 10
    assert summary["kpis"]["orderCount"] == 0
    assert summary["kpis"]["logisticsPerOrder"] is None
    assert products["items"][0]["orderCount"] == 0
    assert products["items"][0]["logisticsPerOrder"] is None


def test_logistics_product_filter_uses_one_canonical_product_reference(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        order_row = db.query(ReportLogisticsOrderRow).one()
        sku_row = db.query(ReportLogisticsSkuRow).one()
        order_row.product = "Старое название только в WB"
        sku_row.product = "Другое название в отчёте"
        values = {
            column.name: getattr(order_row, column.name)
            for column in ReportLogisticsOrderRow.__table__.columns
            if column.name != "id"
        }
        values.update(
            {
                "chain_key": "f" * 64,
                "chain_segment_key": "e" * 64,
                "product": "Новое название того же товара",
                "source_hash_digest": "d" * 64,
            }
        )
        db.add(ReportLogisticsOrderRow(**values))
        db.commit()
    login(client)

    filters = {
        "periodStart": "2026-04-06",
        "periodEnd": "2026-04-12",
        "product": "Старое название",
    }
    summary = client.get(
        "/api/reports/report-1/logistics/summary", params=filters
    ).json()
    products = client.get(
        "/api/reports/report-1/logistics/products", params=filters
    ).json()

    assert summary["kpis"]["logisticsTotal"] == 20
    assert summary["kpis"]["revenue"] == 100
    assert summary["kpis"]["profitBeforeTax"] == 20
    assert summary["dynamics"][0]["revenue"] == 100
    assert products["total"] == 1
    assert products["items"][0]["logisticsTotal"] == 20
    assert products["items"][0]["revenue"] == 100
    assert products["items"][0]["profitBeforeTax"] == 20
    orders = client.get(
        "/api/reports/report-1/logistics/orders",
        params={**filters, "productRef": products["items"][0]["productRef"]},
    ).json()
    assert orders["total"] == 2


def test_logistics_quality_status_is_visible_in_slice_and_product(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        context = db.get(ReportLogisticsAnalysisContext, "report-1")
        order_row = db.query(ReportLogisticsOrderRow).one()
        assert context is not None
        context.data_status = "partial"
        order_row.data_quality_status = "partial"
        db.commit()
    login(client)

    filters = {"periodStart": "2026-04-06", "periodEnd": "2026-04-12"}
    summary = client.get(
        "/api/reports/report-1/logistics/summary", params=filters
    ).json()
    products = client.get(
        "/api/reports/report-1/logistics/products", params=filters
    ).json()

    assert summary["dataStatus"] == "partial"
    assert summary["sliceStatus"] == "partial"
    assert summary["coverage"]["dataQualityIssues"] == 1
    recommendations = summary["recommendations"]
    assert "review_data_quality" in {item["code"] for item in recommendations}
    assert [item["priority"] for item in recommendations] == sorted(
        item["priority"] for item in recommendations
    )
    assert [item["code"] for item in recommendations].index(
        "review_data_quality"
    ) < [item["code"] for item in recommendations].index("check_margin")
    assert products["items"][0]["dataQualityStatus"] == "partial"


def test_logistics_missing_profit_link_fails_financial_slice_closed(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        order_row = db.query(ReportLogisticsOrderRow).one()
        sku_row = db.query(ReportLogisticsSkuRow).one()
        order_values = {
            column.name: getattr(order_row, column.name)
            for column in ReportLogisticsOrderRow.__table__.columns
            if column.name != "id"
        }
        order_values.update(
            {
                "chain_key": "a" * 64,
                "chain_segment_key": "b" * 64,
                "product_ref": "product:missing-profit-link",
                "product_key": "nm:202",
                "nm_id": "202",
                "sku": "sku-202",
                "vendor_code": "A-202",
                "product": "Товар без финансовой связи",
                "source_revenue": Decimal("999"),
                "source_hash_digest": "c" * 64,
            }
        )
        db.add(ReportLogisticsOrderRow(**order_values))
        sku_values = {
            column.name: getattr(sku_row, column.name)
            for column in ReportLogisticsSkuRow.__table__.columns
            if column.name != "id"
        }
        sku_values.update(
            {
                "row_uid": "d" * 64,
                "product_ref": "product:missing-profit-link",
                "product_key": "nm:202",
                "nm_id": "202",
                "sku": "sku-202",
                "vendor_code": "A-202",
                "product": "Товар без финансовой связи",
                "revenue": Decimal("999"),
                "financial_revenue": None,
                "profit_before_tax": None,
                "profit_without_logistics": None,
                "profit_effect_amount": Decimal("-10"),
                "logistics_share_pct": None,
                "data_quality_status": "missing_profit_link",
                "recommendation_flags": ["restore_profit_link"],
                "source_hash_digest": "e" * 64,
            }
        )
        db.add(ReportLogisticsSkuRow(**sku_values))
        db.commit()
    login(client)

    filters = {"periodStart": "2026-04-06", "periodEnd": "2026-04-12"}
    summary = client.get(
        "/api/reports/report-1/logistics/summary", params=filters
    ).json()
    products = client.get(
        "/api/reports/report-1/logistics/products", params=filters
    ).json()

    assert summary["dataStatus"] == "ready"
    assert summary["sliceStatus"] == "partial"
    assert (
        summary["financialMetricStatus"]
        == "not_available_missing_profit_link"
    )
    assert summary["coverage"]["missingProfitLinks"] == 1
    assert summary["coverage"]["missingProfitLinkAmount"] == 10
    assert summary["coverage"]["dataQualityIssues"] == 0
    assert summary["kpis"]["logisticsTotal"] == 20
    for key in (
        "revenue",
        "logisticsSharePct",
        "profitBeforeTax",
        "profitWithoutLogistics",
        "profitEffectAmount",
    ):
        assert summary["kpis"][key] is None
    assert summary["dynamics"][0]["revenue"] is None
    assert summary["dynamics"][0]["logisticsSharePct"] is None
    assert len(summary["rankings"]["byTotal"]) == 2
    assert summary["rankings"]["byRevenueShare"] == []
    assert summary["rankings"]["byProfitEffect"] == []
    restore_link = next(
        item
        for item in summary["recommendations"]
        if item["code"] == "restore_profit_link"
    )
    assert restore_link["impactAmount"] == 10
    assert restore_link["evidenceType"] == "data_quality"
    assert restore_link["actionTarget"] == "source"
    assert "review_data_quality" not in {
        item["code"] for item in summary["recommendations"]
    }

    assert products["financialMetricStatus"] == (
        "not_available_missing_profit_link"
    )
    assert all(item["revenue"] is None for item in products["items"])
    assert all(item["profitEffectAmount"] is None for item in products["items"])
    missing = next(
        item for item in products["items"] if item["productKey"] == "nm:202"
    )
    assert missing["logisticsTotal"] == 10
    assert missing["dataQualityStatus"] == "missing_profit_link"
    assert missing["recommendationFlags"] == ["restore_profit_link"]


def test_logistics_api_rejects_inverted_and_outside_periods(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_factors_enabled": True,
        },
    )
    persist_logistics_fixture(client)
    login(client)

    for endpoint in ("summary", "products", "dimensions", "orders"):
        inverted = client.get(
            f"/api/reports/report-1/logistics/{endpoint}",
            params={"periodStart": "2026-04-12", "periodEnd": "2026-04-06"},
        )
        outside = client.get(
            f"/api/reports/report-1/logistics/{endpoint}",
            params={"periodStart": "2020-04-05", "periodEnd": "2026-04-12"},
        )
        assert inverted.status_code == 400
        assert outside.status_code == 400
        assert inverted.json()["detail"]["code"] == "invalid_logistics_period"
        assert outside.json()["detail"]["code"] == "invalid_logistics_period"

    openapi = client.get("/openapi.json").json()
    for endpoint in ("summary", "products", "dimensions", "orders"):
        operation = openapi["paths"][
            f"/api/reports/{{report_id}}/logistics/{endpoint}"
        ]["get"]
        assert "400" in operation["responses"]


def test_invalid_logistics_context_status_blocks_publication_and_api(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        context = db.get(repository.ReportLogisticsAnalysisContext, "report-1")
        assert report is not None
        assert context is not None
        context.data_status = "unexpected"
        db.commit()
        blocker = next(
            item
            for item in repository.report_publication_blockers(db, report)
            if item["code"] == "logistics_analysis_invalid_status"
        )

    login(client)
    response = client.get("/api/reports/report-1/logistics/summary")

    assert blocker["nonOverridable"] is True
    assert response.status_code == 200
    assert response.json()["dataStatus"] == "needs_rebuild"
    assert response.json()["kpis"]["logisticsTotal"] is None
    assert response.json()["rankings"]["byTotal"] == []


def test_logistics_persistence_rejects_result_from_another_scope(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        source = LogisticsSourceRow(
            tenant_id="other-tenant",
            client_id="other-client",
            wb_cabinet_id="other-cabinet",
            client_company_id="other-company",
            source_row_id="other-row",
            source_hash="other-hash",
            financial_date=date(2026, 4, 6),
            order_date=date(2026, 4, 6),
            order_uid="other-order",
            nm_id="101",
            sku="sku-101",
            vendor_code="A-101",
            product="Чужой товар",
            scheme="fbo",
            warehouse="Коледино",
            destination="Россия",
            document_type="Логистика",
            operation_name="Логистика",
            quantity=Decimal("0"),
            retail_amount=Decimal("0"),
            delivery_service=Decimal("10"),
            delivery_amount=Decimal("1"),
            return_amount=Decimal("0"),
            rebill_logistic_cost=Decimal("0"),
        )
        unit = UnitEconomicsSlice(
            tenant_id="other-tenant",
            client_id="other-client",
            financial_week_start=date(2026, 4, 6),
            wb_cabinet_id="other-cabinet",
            client_company_id="other-company",
            scheme="fbo",
            nm_id="101",
            sku="sku-101",
            vendor_code="A-101",
            product="Чужой товар",
            revenue=Decimal("100"),
            profit_before_tax=Decimal("20"),
            logistics=Decimal("10"),
        )
        result = build_logistics_analysis([source], [unit])
        assert result.context.data_status == "ready"

        with pytest.raises(ValueError, match="scope does not match report"):
            repository.replace_report_logistics_analysis(db, report, result)


def test_logistics_persistence_rejects_foreign_cabinet_and_company(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(
            db,
            report,
            cabinet_id="foreign-logistics-cabinet",
            company_id="foreign-logistics-company",
            tenant_id="other",
            client_id="other",
        )
        source = LogisticsSourceRow(
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            wb_cabinet_id="foreign-logistics-cabinet",
            client_company_id="foreign-logistics-company",
            source_row_id="foreign-row",
            source_hash="foreign-hash",
            financial_date=date(2026, 4, 6),
            order_date=date(2026, 4, 6),
            order_uid="foreign-order",
            nm_id="101",
            sku="sku-101",
            vendor_code="A-101",
            product="Чужой кабинет",
            scheme="fbo",
            warehouse="Коледино",
            destination="Россия",
            document_type="Логистика",
            operation_name="Логистика",
            quantity=Decimal("0"),
            retail_amount=Decimal("0"),
            delivery_service=Decimal("10"),
            delivery_amount=Decimal("1"),
            return_amount=Decimal("0"),
            rebill_logistic_cost=Decimal("0"),
        )
        unit = UnitEconomicsSlice(
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            financial_week_start=date(2026, 4, 6),
            wb_cabinet_id="foreign-logistics-cabinet",
            client_company_id="foreign-logistics-company",
            scheme="fbo",
            nm_id="101",
            sku="sku-101",
            vendor_code="A-101",
            product="Чужой кабинет",
            revenue=Decimal("100"),
            profit_before_tax=Decimal("20"),
            logistics=Decimal("10"),
        )
        result = build_logistics_analysis([source], [unit])
        assert result.context.data_status == "ready"

        with pytest.raises(ValueError, match="cabinet or company scope"):
            repository.replace_report_logistics_analysis(db, report, result)


def test_logistics_persistence_rejects_repeat_for_same_report(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        repository.replace_report_logistics_analysis(db, report, result)
        db.flush()

        with pytest.raises(ValueError, match="immutable for a report"):
            repository.replace_report_logistics_analysis(db, report, result)


def test_logistics_persistence_keeps_missing_financial_link_without_zero_sentinel(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        source_only_sku = replace(
            result.sku_rows[0],
            source_revenue=Decimal("125"),
            revenue=None,
            profit_before_tax=None,
            profit_without_logistics=None,
            profit_effect_amount=None,
            logistics_share_pct=None,
            data_quality_status="missing_profit_link",
            recommendation_flags=("restore_profit_link",),
        )
        repository.replace_report_logistics_analysis(
            db,
            report,
            replace(result, sku_rows=(source_only_sku,)),
        )
        db.flush()

        persisted = db.query(ReportLogisticsSkuRow).one()
        assert persisted.revenue == Decimal("125")
        assert persisted.financial_revenue is None
        assert persisted.profit_before_tax is None
        assert persisted.profit_effect_amount == Decimal("-10")


def test_logistics_recommendation_uses_full_slice_not_by_total_top_ten(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        source_rows = []
        unit_rows = []
        for index in range(1, 12):
            reverse = index == 11
            logistics = Decimal("1") if reverse else Decimal(110 - index)
            source_rows.append(
                LogisticsSourceRow(
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                    wb_cabinet_id="cabinet-logistics",
                    client_company_id="company-logistics",
                    source_row_id=f"row-{index}",
                    source_hash=f"hash-{index}",
                    financial_date=date(2026, 4, 6),
                    order_date=date(2026, 4, 6),
                    order_uid=f"order-{index}",
                    nm_id=str(index),
                    sku=f"sku-{index}",
                    vendor_code=f"A-{index}",
                    product=f"Product {index}",
                    scheme="fbo",
                    warehouse="Коледино",
                    destination="Россия",
                    document_type="Логистика",
                    operation_name="Логистика",
                    quantity=Decimal("0"),
                    retail_amount=Decimal("0"),
                    delivery_service=logistics,
                    delivery_amount=Decimal("0" if reverse else "1"),
                    return_amount=Decimal("1" if reverse else "0"),
                    rebill_logistic_cost=Decimal("0"),
                )
            )
            unit_rows.append(
                UnitEconomicsSlice(
                    tenant_id=report.tenant_id,
                    client_id=report.client_id,
                    financial_week_start=date(2026, 4, 6),
                    wb_cabinet_id="cabinet-logistics",
                    client_company_id="company-logistics",
                    scheme="fbo",
                    nm_id=str(index),
                    sku=f"sku-{index}",
                    vendor_code=f"A-{index}",
                    product=f"Product {index}",
                    revenue=Decimal("100"),
                    profit_before_tax=Decimal("20"),
                    logistics=logistics,
                )
            )
        repository.replace_report_logistics_analysis(
            db,
            report,
            build_logistics_analysis(
                source_rows,
                unit_rows,
                expected_tenant_id=report.tenant_id,
                expected_client_id=report.client_id,
            ),
        )
        db.commit()

    login(client)
    payload = client.get(
        "/api/reports/report-1/logistics/summary",
        params={"periodStart": "2026-04-06", "periodEnd": "2026-04-12"},
    ).json()
    recommendation = next(
        item for item in payload["recommendations"] if item["code"] == "check_returns"
    )

    assert len(payload["rankings"]["byTotal"]) == 10
    assert all(
        item["product"] != "Product 11" for item in payload["rankings"]["byTotal"]
    )
    assert recommendation["evidence"]["product"] == "Product 11"
    assert recommendation["evidenceType"] == "limitation"
    assert recommendation["actionTarget"] == "products"
    assert "Причина недоступна в Finance" in recommendation["message"]


def test_logistics_api_scopes_sku_fallback_and_recomputes_slice_coverage(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(
            db,
            report,
            cabinet_id="cabinet-1",
            company_id="company-1",
        )
        _ensure_logistics_dimensions(
            db,
            report,
            cabinet_id="cabinet-2",
            company_id="company-2",
        )
        rows = [
            LogisticsSourceRow(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="cabinet-1",
                client_company_id="company-1",
                source_row_id="row-1",
                source_hash="hash-1",
                financial_date=date(2026, 4, 6),
                order_date=date(2026, 4, 6),
                order_uid="order-1",
                nm_id="",
                sku="same-sku",
                vendor_code="A",
                product="Первый",
                scheme="fbo",
                warehouse="Коледино",
                destination="Москва",
                document_type="Логистика",
                operation_name="Логистика",
                quantity=Decimal("0"),
                retail_amount=Decimal("0"),
                delivery_service=Decimal("5"),
                delivery_amount=Decimal("1"),
                return_amount=Decimal("0"),
                rebill_logistic_cost=Decimal("0"),
            ),
            LogisticsSourceRow(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="cabinet-2",
                client_company_id="company-2",
                source_row_id="row-2",
                source_hash="hash-2",
                financial_date=date(2026, 4, 6),
                order_date=date(2026, 4, 6),
                order_uid="order-2",
                nm_id="",
                sku="same-sku",
                vendor_code="B",
                product="Второй",
                scheme="fbo",
                warehouse="Коледино",
                destination="Москва",
                document_type="Логистика",
                operation_name="Логистика",
                quantity=Decimal("0"),
                retail_amount=Decimal("0"),
                delivery_service=Decimal("5"),
                delivery_amount=Decimal("1"),
                return_amount=Decimal("0"),
                rebill_logistic_cost=Decimal("0"),
            ),
            LogisticsSourceRow(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="cabinet-1",
                client_company_id="company-1",
                source_row_id="row-3",
                source_hash="hash-3",
                financial_date=date(2026, 4, 7),
                order_date=date(2026, 4, 6),
                order_uid="order-3",
                nm_id="",
                sku="same-sku",
                vendor_code="A",
                product="Первый",
                scheme="fbo",
                warehouse="Коледино",
                destination="Москва",
                document_type="Логистика",
                operation_name="Неизвестно",
                quantity=Decimal("0"),
                retail_amount=Decimal("0"),
                delivery_service=Decimal("3"),
                delivery_amount=Decimal("0"),
                return_amount=Decimal("0"),
                rebill_logistic_cost=Decimal("0"),
            ),
        ]
        units = [
            UnitEconomicsSlice(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                financial_week_start=date(2026, 4, 6),
                wb_cabinet_id="cabinet-1",
                client_company_id="company-1",
                scheme="fbo",
                nm_id="",
                sku="same-sku",
                vendor_code="A",
                product="Первый",
                revenue=Decimal("100"),
                profit_before_tax=Decimal("20"),
                logistics=Decimal("8"),
            ),
            UnitEconomicsSlice(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                financial_week_start=date(2026, 4, 6),
                wb_cabinet_id="cabinet-2",
                client_company_id="company-2",
                scheme="fbo",
                nm_id="",
                sku="same-sku",
                vendor_code="B",
                product="Второй",
                revenue=Decimal("100"),
                profit_before_tax=Decimal("20"),
                logistics=Decimal("5"),
            ),
        ]
        repository.replace_report_logistics_analysis(
            db, report, build_logistics_analysis(rows, units)
        )
        db.commit()

    login(client)
    for sort_by in sorted(repository.LOGISTICS_PRODUCT_SORT_KEYS):
        response = client.get(
            "/api/reports/report-1/logistics/products",
            params={
                "periodStart": "2026-04-06",
                "periodEnd": "2026-04-12",
                "sortBy": sort_by,
                "sortOrder": "asc",
                "limit": 1,
            },
        )
        assert response.status_code == 200, (sort_by, response.text)
    products = client.get(
        "/api/reports/report-1/logistics/products",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "limit": 1,
        },
    ).json()
    assert products["total"] == 2
    assert len(products["items"]) == 1
    products_ascending = client.get(
        "/api/reports/report-1/logistics/products",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "sortBy": "logisticsTotal",
            "sortOrder": "asc",
            "limit": 1,
        },
    ).json()
    products_descending = client.get(
        "/api/reports/report-1/logistics/products",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "sortBy": "logisticsTotal",
            "sortOrder": "desc",
            "limit": 1,
        },
    ).json()
    assert (
        products_ascending["items"][0]["logisticsTotal"]
        < products_descending["items"][0]["logisticsTotal"]
    )
    product_ref = products["items"][0]["productRef"]
    for sort_by in sorted(repository.LOGISTICS_ORDER_SORT_KEYS):
        response = client.get(
            "/api/reports/report-1/logistics/orders",
            params={
                "periodStart": "2026-04-06",
                "periodEnd": "2026-04-12",
                "productRef": product_ref,
                "sortBy": sort_by,
                "sortOrder": "asc",
                "limit": 1,
            },
        )
        assert response.status_code == 200, (sort_by, response.text)
    first_order_page = client.get(
        "/api/reports/report-1/logistics/orders",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "productRef": product_ref,
            "limit": 1,
            "offset": 0,
        },
    ).json()
    second_order_page = client.get(
        "/api/reports/report-1/logistics/orders",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "productRef": product_ref,
            "limit": 1,
            "offset": 1,
        },
    ).json()
    assert first_order_page["total"] == 2
    assert second_order_page["total"] == 2
    assert (
        first_order_page["items"][0]["chainRef"]
        != second_order_page["items"][0]["chainRef"]
    )
    ascending_order = client.get(
        "/api/reports/report-1/logistics/orders",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "productRef": product_ref,
            "sortBy": "logisticsTotal",
            "sortOrder": "asc",
            "limit": 1,
        },
    ).json()
    descending_order = client.get(
        "/api/reports/report-1/logistics/orders",
        params={
            "periodStart": "2026-04-06",
            "periodEnd": "2026-04-12",
            "productRef": product_ref,
            "sortBy": "logisticsTotal",
            "sortOrder": "desc",
            "limit": 1,
        },
    ).json()
    assert (
        ascending_order["items"][0]["logisticsTotal"]
        < descending_order["items"][0]["logisticsTotal"]
    )

    classified = client.get(
        "/api/reports/report-1/logistics/summary",
        params={"periodStart": "2026-04-06", "periodEnd": "2026-04-06"},
    ).json()
    unclassified = client.get(
        "/api/reports/report-1/logistics/summary",
        params={"periodStart": "2026-04-07", "periodEnd": "2026-04-07"},
    ).json()
    assert classified["coverage"]["classificationPct"] == 100
    assert classified["sliceStatus"] == "ready"
    assert unclassified["coverage"]["classificationPct"] == 0
    assert unclassified["sliceStatus"] == "partial"
    assert "restore_classification" not in {
        item["code"] for item in classified["recommendations"]
    }
    assert "restore_classification" in {
        item["code"] for item in unclassified["recommendations"]
    }


def test_blocked_logistics_gate_is_non_overridable_publication_blocker(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        source_rows = [
            LogisticsSourceRow(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="cabinet-logistics",
                client_company_id="company-logistics",
                source_row_id="missing-order-key",
                source_hash="missing-order-key-hash",
                financial_date=date(2026, 4, 6),
                order_date=date(2026, 4, 5),
                order_uid="",
                nm_id="101",
                sku="sku-101",
                vendor_code="A-101",
                product="Товар без ключа заказа",
                scheme="fbo",
                warehouse="Коледино",
                destination="Россия",
                document_type="Логистика",
                operation_name="Логистика",
                quantity=Decimal("0"),
                retail_amount=Decimal("0"),
                delivery_service=Decimal("10"),
                delivery_amount=Decimal("1"),
                return_amount=Decimal("0"),
                rebill_logistic_cost=Decimal("0"),
            )
        ]
        unit_rows = [
            UnitEconomicsSlice(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                financial_week_start=date(2026, 4, 6),
                wb_cabinet_id="cabinet-logistics",
                client_company_id="company-logistics",
                scheme="fbo",
                nm_id="101",
                sku="sku-101",
                vendor_code="A-101",
                product="Товар без ключа заказа",
                revenue=Decimal("100"),
                profit_before_tax=Decimal("20"),
                logistics=Decimal("10"),
            )
        ]
        result = build_logistics_analysis(source_rows, unit_rows)
        assert result.context.data_status == "blocked"
        repository.replace_report_logistics_analysis(db, report, result)
        db.flush()

        blocker = next(
            item
            for item in repository.report_publication_blockers(db, report)
            if item["code"] == "logistics_analysis_blocked"
        )

    assert blocker["nonOverridable"] is True


def test_required_logistics_context_missing_or_outdated_blocks_publication(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"logistics_analysis_enabled": True},
    )
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        report.logistics_analysis_required = True
        db.flush()
        missing = repository.report_publication_blockers(db, report)
        blocker = next(
            item for item in missing if item["code"] == "logistics_analysis_missing"
        )
        assert blocker["nonOverridable"] is True
        report.logistics_analysis_required = False
        assert not any(
            item["code"].startswith("logistics_analysis_")
            for item in repository.report_publication_blockers(db, report)
        )
        db.rollback()

    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        context = db.get(ReportLogisticsAnalysisContext, report.id)
        assert context is not None
        context.methodology_version = "wb-logistics-v4"
        db.commit()
        outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item for item in outdated if item["code"] == "logistics_analysis_outdated"
        )
        assert blocker["nonOverridable"] is True

    login(client)
    payload = client.get("/api/reports/report-1/logistics/summary").json()
    assert payload["dataStatus"] == "needs_rebuild"
    assert payload["kpis"]["logisticsTotal"] is None

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        context = db.get(ReportLogisticsAnalysisContext, report.id)
        assert context is not None
        context.methodology_version = LOGISTICS_METHODOLOGY_VERSION
        context.chain_key_version = "wb-order-product-legacy"
        db.commit()
        key_outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in key_outdated
            if item["code"] == "logistics_analysis_key_outdated"
        )
        assert blocker["nonOverridable"] is True

    payload = client.get("/api/reports/report-1/logistics/summary").json()
    assert payload["dataStatus"] == "needs_rebuild"
    assert payload["chainKeyVersion"] == CHAIN_KEY_VERSION
    assert payload["kpis"]["logisticsTotal"] is None


def test_required_dimension_context_controls_publication_readiness(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        report.logistics_dimensions_required = True
        missing = repository.report_publication_blockers(db, report)
        blocker = next(
            item for item in missing if item["code"] == "logistics_dimensions_missing"
        )
        assert blocker["nonOverridable"] is True
        db.rollback()

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        rows = build_dimension_rows(result.sku_rows, [])
        repository.replace_report_logistics_dimension_analysis(
            db,
            report,
            context=_dimension_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        assert not any(
            item["code"].startswith("logistics_dimensions_")
            for item in readiness["blockingReasons"]
        )
        assert any(
            item["code"] == "logistics_dimensions_partial"
            for item in readiness["reviewReasons"]
        )

        context = db.get(repository.ReportLogisticsDimensionContext, report.id)
        assert context is not None
        context.dimension_row_count += 1
        mismatch = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in mismatch
            if item["code"] == "logistics_dimensions_row_count_mismatch"
        )
        assert blocker["nonOverridable"] is True

        context.dimension_row_count -= 1
        context.factor_methodology_version = "wb-logistics-factors-legacy"
        outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in outdated
            if item["code"] == "logistics_dimensions_outdated"
        )
        assert blocker["nonOverridable"] is True


def test_required_tariff_context_controls_publication_readiness(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        report.logistics_tariffs_required = True
        missing = repository.report_publication_blockers(db, report)
        blocker = next(
            item for item in missing if item["code"] == "logistics_tariffs_missing"
        )
        assert blocker["nonOverridable"] is True
        db.rollback()

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        rows = build_tariff_rows(
            result.sku_rows,
            [
                {
                    "wb_cabinet_id": "cabinet-logistics",
                    "requested_date": "2026-04-06",
                    "tariff_type": "box",
                    "warehouse_name": "Склад A",
                    "box_delivery_coef_expr": "125",
                    "box_storage_coef_expr": "115",
                    "source_hash": "tariff-readiness",
                }
            ],
            factor_snapshot_date=date(2026, 7, 21),
        )
        repository.replace_report_logistics_tariff_analysis(
            db,
            report,
            context=_tariff_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        assert not any(
            item["code"].startswith("logistics_tariffs_")
            for item in readiness["blockingReasons"]
        )
        assert any(
            item["code"] == "logistics_tariffs_partial"
            for item in readiness["reviewReasons"]
        )

        context = db.get(repository.ReportLogisticsTariffContext, report.id)
        assert context is not None
        context.tariff_row_count += 1
        mismatch = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in mismatch
            if item["code"] == "logistics_tariffs_row_count_mismatch"
        )
        assert blocker["nonOverridable"] is True

        context.tariff_row_count -= 1
        context.factor_methodology_version = "wb-logistics-tariffs-legacy"
        outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in outdated
            if item["code"] == "logistics_tariffs_outdated"
        )
        assert blocker["nonOverridable"] is True


def test_required_route_context_controls_publication_readiness(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        report.logistics_routes_required = True
        missing = repository.report_publication_blockers(db, report)
        blocker = next(
            item for item in missing if item["code"] == "logistics_routes_missing"
        )
        assert blocker["nonOverridable"] is True
        db.rollback()

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        rows = build_route_rows(result.order_rows, [])
        repository.replace_report_logistics_route_analysis(
            db,
            report,
            context=_route_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        assert not any(
            item["code"].startswith("logistics_routes_")
            for item in readiness["blockingReasons"]
        )
        assert any(
            item["code"] == "logistics_routes_partial"
            for item in readiness["reviewReasons"]
        )

        context = db.get(repository.ReportLogisticsRouteContext, report.id)
        assert context is not None
        context.route_row_count += 1
        mismatch = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in mismatch
            if item["code"] == "logistics_routes_row_count_mismatch"
        )
        assert blocker["nonOverridable"] is True

        context.route_row_count -= 1
        context.factor_methodology_version = "wb-logistics-routes-legacy"
        outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in outdated
            if item["code"] == "logistics_routes_outdated"
        )
        assert blocker["nonOverridable"] is True


def test_required_measurement_context_controls_publication_readiness(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        report.logistics_measurements_required = True
        missing = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in missing
            if item["code"] == "logistics_measurements_missing"
        )
        assert blocker["nonOverridable"] is True
        db.rollback()

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        rows = build_measurement_rows(
            result.sku_rows,
            [
                {
                    "tenant_id": report.tenant_id,
                    "client_id": report.client_id,
                    "wb_cabinet_id": "cabinet-logistics",
                    "dim_id": "safe-readiness-event",
                    "nm_id": "101",
                    "volume": "2.5",
                    "width": "10",
                    "length": "25",
                    "height": "10",
                    "dt_bonus": "2026-04-07T10:00:00Z",
                    "penalty_amount": "0",
                    "reversal_amount": "0",
                }
            ],
            [],
        )
        repository.replace_report_logistics_measurement_analysis(
            db,
            report,
            context=_measurement_context(report, rows, data_status="partial"),
            rows=rows,
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        assert not any(
            item["code"].startswith("logistics_measurements_")
            for item in readiness["blockingReasons"]
        )
        assert any(
            item["code"] == "logistics_measurements_partial"
            for item in readiness["reviewReasons"]
        )

        context = db.get(repository.ReportLogisticsMeasurementContext, report.id)
        assert context is not None
        context.measurement_row_count += 1
        mismatch = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in mismatch
            if item["code"] == "logistics_measurements_row_count_mismatch"
        )
        assert blocker["nonOverridable"] is True

        context.measurement_row_count -= 1
        context.factor_methodology_version = "wb-logistics-measurements-legacy"
        outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in outdated
            if item["code"] == "logistics_measurements_outdated"
        )
        assert blocker["nonOverridable"] is True


def test_required_return_reason_context_controls_publication_readiness(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        report.logistics_return_reasons_required = True
        missing = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in missing
            if item["code"] == "logistics_return_reasons_missing"
        )
        assert blocker["nonOverridable"] is True
        db.rollback()

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        logistics_result, return_reason_result = _return_reason_fixture_result(report)
        repository.replace_report_logistics_analysis(db, report, logistics_result)
        repository.replace_report_logistics_return_reason_analysis(
            db,
            report,
            return_reason_result,
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        assert not any(
            item["code"].startswith("logistics_return_reasons_")
            for item in readiness["blockingReasons"]
        )
        assert any(
            item["code"] == "logistics_return_reasons_partial"
            for item in readiness["reviewReasons"]
        )

        context = db.get(
            repository.ReportLogisticsReturnReasonContext,
            report.id,
        )
        assert context is not None
        context.return_reason_row_count += 1
        mismatch = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in mismatch
            if item["code"] == "logistics_return_reasons_row_count_mismatch"
        )
        assert blocker["nonOverridable"] is True

        context.return_reason_row_count -= 1
        context.methodology_version = "wb-logistics-return-reasons-legacy"
        outdated = repository.report_publication_blockers(db, report)
        blocker = next(
            item
            for item in outdated
            if item["code"] == "logistics_return_reasons_outdated"
        )
        assert blocker["nonOverridable"] is True


def test_dimension_context_and_rows_validation_is_atomic(tmp_path: Path) -> None:
    client = make_client(tmp_path, publish_report=False)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        _ensure_logistics_dimensions(db, report)
        result = _logistics_fixture_result(report)
        rows = build_dimension_rows(
            result.sku_rows,
            [
                {
                    "wb_cabinet_id": "cabinet-logistics",
                    "nm_id": "101",
                    "length_cm": 30,
                    "width_cm": 20,
                    "height_cm": 10,
                    "weight_brutto_kg": 2,
                    "dimensions_valid": True,
                }
            ],
        )
        repository.replace_report_logistics_dimension_analysis(
            db,
            report,
            context=_dimension_context(report, rows),
            rows=rows,
        )
        db.flush()
        original_uid = repository.report_logistics_dimension_rows(db, report.id)[
            0
        ].row_uid

        foreign = [{**rows[0], "tenant_id": "other"}]
        with pytest.raises(ValueError, match="tenant does not match report"):
            repository.replace_report_logistics_dimension_analysis(
                db,
                report,
                context=_dimension_context(report, foreign),
                rows=foreign,
            )

        persisted = repository.report_logistics_dimension_rows(db, report.id)
        context = db.get(repository.ReportLogisticsDimensionContext, report.id)
        assert [row.row_uid for row in persisted] == [original_uid]
        assert context is not None
        assert context.input_hash == "safe-dimension-input-hash"

def test_report_with_logistics_context_cannot_be_reimported_in_place(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    persist_logistics_fixture(client)
    with (
        client.app.state.session_factory() as db,
        pytest.raises(ValueError, match="create a new report run"),
    ):
        repository.import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
            source_workbook_path="",
        )


def test_logistics_client_flag_does_not_expose_order_chains(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "logistics_analysis_enabled": True,
            "logistics_analysis_client_enabled": True,
        },
    )
    persist_logistics_fixture(client)
    with client.app.state.session_factory() as db:
        repository.upsert_user(
            db,
            email="logistics-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "logistics-client@example.com", "secret")

    assert client.get("/api/reports/report-1/logistics/summary").status_code == 200
    assert client.get("/api/reports/report-1/logistics/products").status_code == 200
    assert client.get("/api/reports/report-1/logistics/orders").status_code == 404
    assert client.get("/api/reports/other-report/logistics/summary").status_code == 404


def test_import_dashboard_payload_replaces_existing_report_rows(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
        )
        db.commit()
        replacement = deepcopy(sample_payload())
        replacement["unitRows"] = [replacement["unitRows"][0]]
        replacement["lostSales"] = []
        replacement["documentReconciliation"] = []
        import_dashboard_payload(
            db,
            replacement,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
        )
        db.commit()
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        summary = repository.report_full_payload(db, report)

    assert len(summary["unitRows"]) == 1
    assert (
        summary["unitRows"][0]["documentReport"]
        == "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
    )
    assert summary["unitRows"][0]["wbReportId"] == "726807272"
    assert summary["unitRows"][0]["wbReportDate"] == "2026-04-13"
    assert summary["meta"]["sourceCoverage"] == "01.03.2026 - 17.06.2026"
    assert summary["meta"]["sourceCoverageStart"] == "2026-03-01"
    assert summary["meta"]["sourceCoverageEnd"] == "2026-06-17"
    assert len(summary["liquidityRows"]) == 1
    assert summary["lostSales"] == []
    assert summary["reconciliationMonthly"][0]["wb_quantity"] == 90.0
    assert summary["reconciliationMonthly"][0]["onec_quantity"] == 91.0
    assert summary["reconciliationMonthly"][0]["quantity_delta"] == -1.0
    assert summary["documentReconciliation"] == []
    expense_labels = {item["expense"] for item in summary["expenses"]}
    assert "НДС к уплате" not in expense_labels
    assert "Налог с выручки/НДФЛ" not in expense_labels


def test_multi_client_backfill_is_idempotent(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
        )
        client = db.query(repository.Client).filter_by(tenant_id="shumeyko").one()
        client.name = "Реальный клиент"
        db.commit()

    init_db(engine)
    init_db(engine)

    with session_factory() as db:
        client = db.query(repository.Client).filter_by(tenant_id="shumeyko").one()
        report = db.get(repository.ReportRun, "report-1")
        unit_rows = db.query(repository.ReportUnitRow).all()
        lost_rows = db.query(repository.ReportLostSalesRow).all()
        document_rows = db.query(repository.ReportDocumentReconciliationRow).all()

        assert client.name == "Реальный клиент"
        assert db.query(repository.Client).filter_by(tenant_id="shumeyko").count() == 1
        company_count = (
            db.query(repository.ClientCompany).filter_by(client_id="shumeyko").count()
        )
        cabinet_count = (
            db.query(repository.WbCabinet).filter_by(client_id="shumeyko").count()
        )
        assert company_count == 2
        assert cabinet_count == 2

    assert report is not None
    assert report.client_id == "shumeyko"
    assert {row.client_id for row in unit_rows} == {"shumeyko"}
    assert all(row.client_company_id for row in unit_rows)
    assert all(row.wb_cabinet_id for row in unit_rows)
    assert {row.client_id for row in lost_rows} == {"shumeyko"}
    assert all(row.wb_cabinet_id for row in lost_rows)
    assert {row.client_id for row in document_rows} == {"shumeyko"}
    assert all(row.client_company_id for row in document_rows)
    assert all(row.wb_cabinet_id for row in document_rows)


def test_import_uses_existing_client_name_over_legacy_meta(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
        )
        client = db.query(repository.Client).filter_by(tenant_id="shumeyko").one()
        client.name = "Реальный клиент"
        db.commit()

        import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-2",
        )
        db.commit()
        report = db.get(repository.ReportRun, "report-2")
        client = db.query(repository.Client).filter_by(tenant_id="shumeyko").one()

    assert report is not None
    assert report.client_name == "Реальный клиент"
    assert client.name == "Реальный клиент"


def test_import_prefers_single_active_wb_provider_cabinet(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    now = repository.security.utcnow()
    with session_factory() as db:
        repository.ensure_tenant(db, "galustov", "Галустов")
        repository.ensure_client_for_tenant(
            db,
            tenant_id="galustov",
            name="Галустов",
        )
        active_company = repository.ensure_client_company(
            db,
            tenant_id="galustov",
            client_id="galustov",
            display_name="Галустов",
        )
        assert active_company is not None
        db.add(
            WbCabinet(
                id="wb-active",
                tenant_id="galustov",
                client_id="galustov",
                client_company_id=active_company.id,
                display_name="ИП Галустов",
                cabinet_key="ip-galustov",
                provider="wb_api",
                status="active",
                created_at=now,
                updated_at=now,
            )
        )
        db.add(
            WbCabinet(
                id="wb-disabled",
                tenant_id="galustov",
                client_id="galustov",
                client_company_id=None,
                display_name="Галустов Рафаэль Рудольфович",
                cabinet_key="galustov-rafael-rudolfovich",
                provider="",
                status="disabled",
                created_at=now,
                updated_at=now,
            )
        )
        db.commit()

        payload = deepcopy(sample_payload())
        payload["meta"]["client"] = "Галустов"
        for row in payload["unitRows"]:
            row["organization"] = "Галустов Рафаэль Рудольфович"
            row["cabinet"] = "Галустов Рафаэль Рудольфович"
        for row in payload["lostSales"]:
            row["cabinet"] = "Галустов Рафаэль Рудольфович"
        for row in payload["documentReconciliation"]:
            row["organization"] = "Галустов Рафаэль Рудольфович"
            row["cabinet"] = "Галустов Рафаэль Рудольфович"

        import_dashboard_payload(
            db,
            payload,
            tenant_id="galustov",
            tenant_name="Галустов",
            report_id="report-galustov",
        )

        unit_rows = db.query(repository.ReportUnitRow).all()
        document_rows = db.query(repository.ReportDocumentReconciliationRow).all()

    assert {row.wb_cabinet_id for row in unit_rows} == {"wb-active"}
    assert {row.client_company_id for row in unit_rows} == {active_company.id}
    assert {row.cabinet for row in unit_rows} == {"ИП Галустов"}
    assert {row.wb_cabinet_id for row in document_rows} == {"wb-active"}
    assert all(row.wb_cabinet_id != "wb-disabled" for row in unit_rows)


def test_report_summary_preserves_lost_sales_onec_stock(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        import_dashboard_payload(
            db,
            sample_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-1",
        )
        db.commit()
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        lost_sales = db.query(repository.ReportLostSalesRow).filter_by(
            report_run_id=report.id
        ).all()

    assert float(lost_sales[0].onec_stock_quantity) == 12.0
    assert lost_sales[0].onec_warehouses == "Собственный склад: 12"


def test_report_summary_includes_document_reconciliation(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    response = client.get("/api/reports/report-1/summary")

    assert response.status_code == 200
    summary = response.json()
    assert datetime.fromisoformat(summary["meta"]["generatedAtIso"])
    assert summary["documentReconciliation"][0]["status"] == "OK"
    assert (
        summary["documentReconciliation"][0]["payoutStatus"]
        == "Нужен источник выплаты 1С"
    )
    assert (
        summary["documentReconciliation"][0]["documentReport"]
        == "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
    )
    assert summary["documentReconciliation"][0]["wbSalesQuantity"] == 22
    assert summary["documentReconciliation"][0]["onecReturnQuantity"] == 2
    assert summary["documentReconciliation"][0]["netQuantityDelta"] == 0
    assert summary["documentReconciliation"][0]["wbForPaySum"] == 85000
    assert summary["kpis"]["wbForPaySum"] == 85000
    assert summary["kpis"]["wbForPayRowCount"] == 1
    assert summary["documentReconciliation"][0]["weeklySalesReportId"] == "SUMMARY-1"
    assert summary["documentReconciliation"][0]["weeklyBuyoutReportId"] == "BUYOUT-1"
    assert summary["documentReconciliation"][0]["onecSettlementTotal"] == 85000
    assert summary["documentReconciliation"][0]["settlementDelta"] == 0
    assert (
        "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
        in summary["options"]["documentReports"]
    )
    assert summary["quality"]["documentReconciliationRows"] == 1
    assert summary["quality"]["documentReconciliationIssues"] == 0
    assert summary["quality"]["documentReconciliationMissingOnec"] == 0
    assert "Отчет комиссионера" in summary["options"]["documentTypes"]
    assert "OK" in summary["options"]["documentReconciliationStatuses"]


def test_document_reconciliation_endpoint_filters_and_kpis(tmp_path: Path) -> None:
    payload = deepcopy(sample_payload())
    payload["documentReconciliation"][0] = {
        **payload["documentReconciliation"][0],
        "payoutStatus": "",
        "comment": "Чистая сверка",
    }
    payload["documentReconciliation"].append(
        {
            **payload["documentReconciliation"][0],
            "id": "doc-recon-2",
            "status": "Нужна проверка",
            "documentReport": (
                "Отчет комиссионера · 01.06.2026-07.06.2026 · закрытие 07.06.2026"
            ),
            "salesPeriod": "2026-06-01 - 2026-06-07",
            "salesPeriodStart": "2026-06-01",
            "salesPeriodEnd": "2026-06-07",
            "expectedDocumentDate": "2026-06-07",
            "cabinet": "Кабинет B",
            "wbCabinetId": "",
            "organization": "Организация B",
            "clientCompanyId": "",
            "onecDocuments": "",
            "quantityDelta": 3,
            "amountDelta": 1200,
            "comment": "Нет документа 1С и есть дельта суммы",
        }
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    response = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"limit": 1},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 2
    assert len(body["items"]) == 1
    assert body["kpis"]["documentCount"] == 2
    assert body["kpis"]["okRows"] == 1
    assert body["kpis"]["issueRows"] == 1
    assert body["kpis"]["missingOnecRows"] == 1
    assert body["kpis"]["amountDelta"] == 1200

    delta_rows = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"delta_only": "true"},
    ).json()
    assert delta_rows["total"] == 1
    assert delta_rows["items"][0]["id"] == "doc-recon-2"

    status_rows = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"status": "Нужна проверка", "period_start": "2026-06-01"},
    ).json()
    assert status_rows["total"] == 1
    assert status_rows["items"][0]["cabinet"] == "Кабинет B"


def test_document_reconciliation_period_uses_sales_week_closing_date(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    payload["documentReconciliation"].append(
        {
            **payload["documentReconciliation"][0],
            "id": "doc-recon-may-closing",
            "salesPeriod": "2026-04-27 - 2026-05-03",
            "salesPeriodStart": "2026-04-27",
            "salesPeriodEnd": "2026-05-03",
            "expectedDocumentDate": "2026-04-30",
        }
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    april = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    ).json()
    may = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"period_start": "2026-05-01", "period_end": "2026-05-31"},
    ).json()

    assert [row["id"] for row in april["items"]] == ["doc-recon-1"]
    assert [row["id"] for row in may["items"]] == ["doc-recon-may-closing"]


def test_financial_document_reconciliation_uses_sales_weeks_for_revenue(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    payload["documentReconciliation"].append(
        {
            **payload["documentReconciliation"][0],
            "id": "doc-recon-buyout",
            "documentType": "Уведомление о выкупе",
            "wbAmount": 66003.74,
            "buyoutRetailAmountSum": 66003.74,
            "onecAmount": 60000,
            "onecExpenseInvoiceAmount": 60000,
            "amountDelta": 6003.74,
            "buyoutRetailDelta": 6003.74,
            "status": "Сверено по количеству",
        }
    )
    payload["unitRows"][0]["penalties"] = 100
    payload["unitRows"].append(
        {
            **payload["unitRows"][0],
            "id": "unit-period-boundary",
            "week": "2026-04-27",
            "documentReport": (
                "Отчет комиссионера · 27.04.2026-03.05.2026 · закрытие 03.05.2026"
            ),
            "wbReportId": "BOUNDARY-REPORT",
            "revenue": 0,
            "revenueWithoutVat": 0,
            "penalties": 40,
        }
    )
    client = make_client(tmp_path, payload=payload)
    login(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="onec-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="financial-reconciliation-test",
            period_start=date(2026, 4, 1),
            period_end=date(2026, 5, 3),
            user=user,
            source_report=report,
            reason="financial reconciliation test",
        )
        sales_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_sales_register",
            source_label="AccumulationRegister_Продажи",
            required=True,
            status="loaded",
            row_count=3,
        )
        repository.add_source_snapshot_row(
            db,
            sales_collection,
            row_number=1,
            raw_payload_hash="financial-sales-1",
            source_row_id="financial-sales-1",
            row_payload={
                "RecordSet": [
                    {
                        "Active": True,
                        "Period": "2026-04-12T23:59:59",
                        "Документ": "COMMISSIONER-1",
                        "Документ_Type": ("StandardODATA.Document_ОтчетКомиссионера"),
                        "Сумма": 100000,
                    },
                    {
                        "Active": True,
                        "Period": "2026-04-13T00:00:00",
                        "Документ": "BUYOUT-1",
                        "Документ_Type": "StandardODATA.Document_РасходнаяНакладная",
                        "Сумма": 60000,
                    },
                    {
                        "Active": True,
                        "Period": "2026-04-30T23:59:59",
                        "Документ": "COMMISSIONER-MAY",
                        "Документ_Type": "StandardODATA.Document_ОтчетКомиссионера",
                        "Сумма": 200000,
                    },
                ]
            },
        )
        invoice_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_incoming_invoices",
            source_label="Document_ПриходнаяНакладная",
            required=True,
            status="loaded",
            row_count=2,
        )
        for row_number, invoice_date, number, amount in (
            (1, "2026-04-12T23:59:59", "НФНФ-TEST-1", 95),
            (2, "2026-05-03T23:59:59", "НФНФ-TEST-2", 40),
        ):
            repository.add_source_snapshot_row(
                db,
                invoice_collection,
                row_number=row_number,
                raw_payload_hash=f"financial-invoice-{row_number}",
                source_row_id=f"financial-invoice-{row_number}",
                row_payload={
                    "Ref_Key": f"invoice-{row_number}",
                    "Number": number,
                    "Date": invoice_date,
                    "Posted": True,
                    "DeletionMark": False,
                    "НомерВходящегоДокумента": f"IN-{row_number}",
                    "Расходы": [
                        {
                            "Содержание": "Штрафы",
                            "Сумма": amount,
                        }
                    ],
                },
            )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="needs_review",
            finished_at=repository.security.utcnow(),
        )
        db.commit()

    response = client.get(
        "/api/reports/report-1/financial-document-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["kpis"] == {
        "revenueWb": 99000.0,
        "revenueOnec": 99000.0,
        "revenueDelta": 0.0,
        "buyoutRetailWb": 66003.74,
        "buyoutNetOnec": 60000.0,
        "buyoutAmountsComparable": False,
        "onecCalendarRevenue": 159000.0,
        "onecCalendarCommissionerRevenue": 99000.0,
        "onecCalendarBuyoutRevenue": 60000.0,
        "onecCalendarDocumentCount": 2,
        "penaltiesWb": 100.0,
        "penaltiesOnec": 95.0,
        "penaltiesDelta": -5.0,
        "issueRows": 1,
    }
    assert body["source"]["snapshotSetId"] == "financial-reconciliation-test"
    buyout_row = next(
        row for row in body["items"] if row["documentType"] == "Уведомление о выкупе"
    )
    assert buyout_row["status"] == "Справочно"
    assert buyout_row["delta"] is None
    assert buyout_row["amountsComparable"] is False
    penalty_rows = [row for row in body["items"] if row["controlType"] == "penalties"]
    assert penalty_rows[0]["onecDocuments"].startswith(
        "Приходная накладная НФНФ-TEST-1"
    )
    assert len(penalty_rows) == 1

    filtered = client.get(
        "/api/reports/report-1/financial-document-reconciliation",
        params={
            "period_start": "2026-04-01",
            "period_end": "2026-04-30",
            "control_type": "penalties",
            "delta_only": "true",
        },
    ).json()
    assert filtered["total"] == 1
    assert {row["controlType"] for row in filtered["items"]} == {"penalties"}

    may = client.get(
        "/api/reports/report-1/financial-document-reconciliation",
        params={
            "period_start": "2026-05-01",
            "period_end": "2026-05-31",
            "control_type": "penalties",
        },
    ).json()
    assert may["kpis"]["penaltiesWb"] == 40.0
    assert may["kpis"]["penaltiesOnec"] == 40.0
    assert may["kpis"]["penaltiesDelta"] == 0.0
    assert may["items"][0]["status"] == "Сходится"
    assert "03.05.2026" in may["items"][0]["onecDocuments"]


def test_financial_reconciliation_calendar_onec_total_includes_buyouts(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    payload["documentReconciliation"].extend(
        [
            {
                **payload["documentReconciliation"][0],
                "id": "doc-recon-commissioner-april-30",
                "salesPeriod": "2026-04-27 - 2026-05-03",
                "salesPeriodStart": "2026-04-27",
                "salesPeriodEnd": "2026-05-03",
                "expectedDocumentDate": "2026-05-03",
                "onecDocumentDates": "2026-04-30",
                "wbAmount": 2018535,
                "onecAmount": 2018535,
                "onecQuantity": 30,
                "onecNetQuantity": 30,
                "onecCogs": 400,
                "onecCogsWithoutVat": 400,
                "onecGrossProfit": 2018135,
            },
            {
                **payload["documentReconciliation"][0],
                "id": "doc-recon-buyout-april-27",
                "documentType": "Уведомление о выкупе",
                "salesPeriod": "2026-04-27 - 2026-05-03",
                "salesPeriodStart": "2026-04-27",
                "salesPeriodEnd": "2026-05-03",
                "expectedDocumentDate": "2026-04-27",
                "onecDocumentDates": "2026-04-27",
                "wbAmount": 63705.23,
                "onecAmount": 45889.76,
                "onecQuantity": 5,
                "onecSalesQuantity": 5,
                "onecCogs": 100,
                "onecCogsWithoutVat": 100,
                "onecGrossProfit": 45789.76,
                "buyoutRetailAmountSum": 63705.23,
                "onecExpenseInvoiceAmount": 45889.76,
            },
            {
                **payload["documentReconciliation"][0],
                "id": "doc-recon-cost-adjustment-april-30",
                "status": "Корректировка себестоимости 1С",
                "documentType": "Корректировка себестоимости 1С",
                "documentReport": "Корректировка себестоимости 1С",
                "cabinet": "",
                "salesPeriod": "2026-04-27 - 2026-05-03",
                "salesPeriodStart": "2026-04-27",
                "salesPeriodEnd": "2026-05-03",
                "expectedDocumentDate": "",
                "onecDocumentDates": "2026-04-30",
                "onecDocuments": "Закрытие месяца 26 от 30.04.2026",
                "wbAmount": None,
                "onecAmount": 0,
                "wbQuantity": None,
                "onecQuantity": 0,
                "onecSalesQuantity": 0,
                "onecReturnQuantity": 0,
                "onecNetQuantity": 0,
                "onecCogs": -20,
                "onecCogsWithoutVat": -20,
                "onecGrossProfit": 20,
            },
        ]
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    body = client.get(
        "/api/reports/report-1/financial-document-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    ).json()

    assert body["kpis"]["onecCalendarCommissionerRevenue"] == 2117535.0
    assert body["kpis"]["onecCalendarBuyoutRevenue"] == 45889.76
    assert body["kpis"]["onecCalendarRevenue"] == 2163424.76
    assert body["kpis"]["onecCalendarDocumentCount"] == 3

    summary = client.get("/api/reports/report-1/summary").json()
    assert summary["kpis"]["onecRevenueWithVat"] == 2163424.76
    assert summary["kpis"]["onecRevenueDocumentCount"] == 3
    assert summary["kpis"]["onecCommissionerRevenueWithVat"] == 2117535.0
    assert summary["kpis"]["onecBuyoutRevenueWithVat"] == 45889.76
    assert summary["kpis"]["onecSalesQuantity"] == 55.0
    assert summary["kpis"]["onecCommissionerQuantity"] == 50.0
    assert summary["kpis"]["onecBuyoutQuantity"] == 5.0
    assert summary["kpis"]["onecCogs"] == 1080.0
    assert summary["kpis"]["onecCommissionerCogs"] == 1000.0
    assert summary["kpis"]["onecBuyoutCogs"] == 100.0
    assert summary["kpis"]["onecOtherCogs"] == -20.0
    assert summary["kpis"]["onecCostAdjustmentRows"] == 1
    assert summary["kpis"]["wbCommissionerRevenueWithVat"] == 2117535.0
    assert summary["kpis"]["wbBuyoutRetailRevenueWithVat"] == 63705.23
    assert summary["kpis"]["wbDocumentRevenueWithVat"] == 2181240.23
    assert summary["kpis"]["commissionerRevenueDelta"] == 0.0
    assert summary["kpis"]["buyoutRevenueDelta"] == -17815.47
    assert summary["kpis"]["wbDocumentRevenueDeltaVsOnec"] == -17815.47
    assert summary["kpis"]["accountingReconciliationWbAmount"] is None
    assert summary["kpis"]["accountingReconciliationOnecAmount"] == 2163424.76
    assert summary["kpis"]["accountingReconciliationDelta"] is None
    assert (
        summary["kpis"]["accountingReconciliationStatus"]
        == "Не проверена первичка выкупов WB"
    )
    assert summary["kpis"]["buyoutPrimaryDocumentAmount"] is None
    assert summary["kpis"]["buyoutPrimaryDocumentDelta"] is None
    assert summary["kpis"]["buyoutPrimaryDocumentStatus"] == "not_loaded"
    assert summary["kpis"]["buyoutUnverifiedPrimaryRows"] == 1

    filtered = client.get(
        "/api/reports/report-1/rows",
        params={"month": "Апрель 2026", "limit": 50},
    ).json()
    assert filtered["kpis"]["onecRevenueWithVat"] == 2163424.76
    assert filtered["analytics"]["kpis"]["onecRevenueWithVat"] == 2163424.76
    assert filtered["kpis"]["wbDocumentRevenueWithVat"] == 2181240.23
    assert filtered["kpis"]["accountingReconciliationDelta"] is None


def test_cogs_reconciliation_explains_boundary_week_and_adjustment(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    first = payload["unitRows"][0]
    first.update(
        {
            "cost": 600,
            "unitCost": 30,
            "costMethod": "sales_register_weighted_average",
            "costMatchStatus": "exact_week_exact_kind",
            "costSourceKind": "commissioner_report",
            "costSourcePeriodStart": "2026-04-06",
            "costSourcePeriodEnd": "2026-04-12",
            "costSourceDocument": "DOC-COMMISSIONER-1",
        }
    )
    second = payload["unitRows"][1]
    second.update(
        {
            "week": "2026-04-27",
            "accountingPeriodDate": "2026-04-30",
            "accountingPeriodSource": "onec_document_date",
            "month": "Май 2026",
            "documentReport": (
                "Отчет комиссионера · 27.04.2026-03.05.2026 · закрытие 03.05.2026"
            ),
            "wbReportId": "WB-MAY-BOUNDARY",
            "status": "ОК",
            "statusReason": "Данные достаточны для расчета",
            "netQty": 10,
            "cost": 400,
            "unitCost": 40,
            "costMethod": "sales_register_weighted_average",
            "costMatchStatus": "exact_week_exact_kind",
            "costSourceKind": "commissioner_report",
            "costSourcePeriodStart": "2026-04-27",
            "costSourcePeriodEnd": "2026-05-03",
            "costSourceDocument": "DOC-BOUNDARY",
        }
    )
    base = payload["documentReconciliation"][0]
    payload["documentReconciliation"].extend(
        [
            {
                **base,
                "id": "doc-recon-boundary",
                "salesPeriod": "2026-04-27 - 2026-05-03",
                "salesPeriodStart": "2026-04-27",
                "salesPeriodEnd": "2026-05-03",
                "expectedDocumentDate": "2026-05-03",
                "onecDocumentDates": "2026-04-30",
                "wbReportIds": "WB-MAY-BOUNDARY",
                "onecDocuments": "DOC-BOUNDARY",
                "onecQuantity": 10,
                "onecCogs": 400,
            },
            {
                **base,
                "id": "doc-recon-adjustment",
                "status": "Корректировка себестоимости 1С",
                "documentType": "Корректировка себестоимости 1С",
                "documentReport": "Корректировка себестоимости 1С",
                "salesPeriod": "2026-04-27 - 2026-05-03",
                "salesPeriodStart": "2026-04-27",
                "salesPeriodEnd": "2026-05-03",
                "onecDocumentDates": "2026-04-30",
                "onecDocuments": "Закрытие месяца 26 от 30.04.2026",
                "onecQuantity": 0,
                "onecCogs": -20,
            },
        ]
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    response = client.get(
        "/api/reports/report-1/cogs-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["supported"] is True
    assert body["summary"]["status"] == "explained"
    assert body["summary"]["pnlCogs"] == 1000.0
    assert body["summary"]["onecCogs"] == 980.0
    assert body["summary"]["delta"] == -20.0
    assert body["summary"]["commissionerBoundaryDelta"] == 0.0
    assert body["summary"]["commissionerSameScopeDelta"] == 0.0
    assert body["summary"]["adjustmentDelta"] == -20.0
    assert body["summary"]["unexplainedDelta"] == 0.0
    assert not any(
        item["status"] == "Переходящая неделя" for item in body["items"]
    )
    assert body["costItems"] == []


def test_cogs_reconciliation_keeps_legacy_report_readable(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    body = client.get(
        "/api/reports/report-1/cogs-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    ).json()

    assert body["supported"] is False
    assert body["summary"]["pnlCogs"] == 65000.0
    assert "Пересоберите immutable report" in body["supportMessage"]


def test_marketplace_expense_reconciliation_is_filterable_and_matches_groups(
    tmp_path: Path,
) -> None:
    payload = ready_payload()
    payload["meta"]["marketplaceExpenseContextVersion"] = (
        "marketplace-expense-reconciliation-v1"
    )
    common = {
        "clientId": "shumeyko",
        "sellerAccountId": "cabinet-a",
        "cabinet": "Кабинет A",
        "organizationId": "ORG-A",
        "organization": "Организация A",
        "counterpartyId": "WB",
        "periodStart": "2026-04-06",
        "periodEnd": "2026-04-12",
        "recognitionDate": "2026-04-12",
        "documentDate": "2026-04-13",
        "sourceKind": "incoming_invoice_expenses",
        "matchStatus": "matched_marketplace_pair",
    }
    payload["marketplaceServiceRows"] = [
        {
            **common,
            "id": "service-core-1",
            "documentNumber": "УПД-1",
            "serviceCategory": "Комиссия WB",
            "controlGroup": "core_services",
            "serviceName": "Комиссия и логистика",
            "amountWithoutVat": 40200,
            "vat": 1000,
            "amountWithVat": 41200,
            "sourceRowHash": "service-core-1-hash",
        },
        {
            **common,
            "id": "service-promotion-1",
            "documentNumber": "УПД-2",
            "serviceCategory": "WB Продвижение",
            "controlGroup": "promotion",
            "serviceName": "WB Продвижение",
            "amountWithoutVat": 4000,
            "vat": 0,
            "amountWithVat": 4000,
            "sourceRowHash": "service-promotion-1-hash",
        },
    ]
    client = make_client(tmp_path, payload=payload)
    login(client)

    response = client.get(
        "/api/reports/report-1/marketplace-expense-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    )

    assert response.status_code == 200
    result = response.json()
    assert result["kpis"]["wbMarketplaceDocumentExpensesWithVat"] == 45200
    assert result["kpis"]["onecMarketplaceExpensesWithVat"] == 45200
    assert result["kpis"]["marketplaceExpenseDeltaWithVat"] == 0
    assert result["kpis"]["marketplaceExpenseReconciliationStatus"] == "matched"
    assert {item["controlGroup"] for item in result["groups"]} == {
        "core_services",
        "promotion",
    }
    assert all(item["status"] == "matched" for item in result["groups"])
    summary = client.get("/api/reports/report-1/summary").json()
    assert summary["kpis"]["onecMarketplaceExpensesWithVat"] == 45200
    filtered = client.get("/api/reports/report-1/rows").json()
    assert filtered["kpis"]["wbMarketplacePnlExpenses"] is not None
    assert (
        filtered["kpis"]["wbMarketplacePnlExpenses"]
        == filtered["analytics"]["kpis"]["wbMarketplacePnlExpenses"]
    )


def test_rows_filters_recalculate_after_tax_profit_and_margin(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    payload = ready_payload()
    for row in payload["unitRows"]:
        row["vatPayable"] = row["vat"]
        row["profitBeforeTax"] = round(
            row["profit"] + row["vatPayable"] + row["usn"],
            2,
        )
        row["pnlVatMode"] = "legacy"
        row["taxProfileSource"] = "test-profile"

    client = make_client(tmp_path, payload=payload)
    login_as(client, "admin@example.com", "secret")
    monkeypatch.setattr(
        repository,
        "_tax_context_payload",
        lambda *_args, **_kwargs: {
            "calculated": True,
            "taxSystem": "УСН",
            "revenueTaxRate": 0.01,
        },
    )

    full = client.get("/api/reports/report-1/rows").json()
    assert full["kpis"]["taxBridgeCalculated"] is True
    assert full["kpis"]["marginAfterTax"] == pytest.approx(
        full["kpis"]["profitAfterTax"] / full["kpis"]["revenue"]
    )
    assert full["kpis"]["marginAfterTax"] == full["analytics"]["kpis"][
        "marginAfterTax"
    ]

    slices = [
        {"cabinet": "Кабинет A"},
        {"organization": "Организация A"},
        {"period_start": "2026-04-01", "period_end": "2026-04-30"},
    ]
    for params in slices:
        filtered = client.get(
            "/api/reports/report-1/rows",
            params=params,
        ).json()
        kpis = filtered["kpis"]
        assert filtered["total"] == 1
        assert kpis["taxBridgeCalculated"] is True
        assert kpis["profitAfterTax"] != full["kpis"]["profitAfterTax"]
        assert kpis["marginAfterTax"] == pytest.approx(
            kpis["profitAfterTax"] / kpis["revenue"]
        )
        assert kpis["profitAfterTax"] == filtered["analytics"]["kpis"][
            "profitAfterTax"
        ]
        assert kpis["marginAfterTax"] == filtered["analytics"]["kpis"][
            "marginAfterTax"
        ]


def test_marketplace_expense_reconciliation_requires_rebuild_for_legacy_report(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, payload=ready_payload())
    login(client)

    payload = client.get(
        "/api/reports/report-1/marketplace-expense-reconciliation"
    ).json()

    assert payload["kpis"]["onecMarketplaceExpensesWithVat"] is None
    assert (
        payload["kpis"]["marketplaceExpenseReconciliationStatus"]
        == "legacy_rebuild_required"
    )
    assert payload["source"]["status"] == "legacy_rebuild_required"


def test_document_reconciliation_endpoint_caps_limit(tmp_path: Path) -> None:
    payload = deepcopy(sample_payload())
    base_row = {
        **payload["documentReconciliation"][0],
        "payoutStatus": "",
        "comment": "Чистая сверка",
    }
    payload["documentReconciliation"] = [
        {
            **base_row,
            "id": f"doc-recon-{index}",
            "documentReport": f"Документ сверки {index}",
        }
        for index in range(1005)
    ]
    client = make_client(tmp_path, payload=payload)
    login(client)

    response = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"limit": 5000},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1005
    assert body["kpis"]["documentCount"] == 1005
    assert len(body["items"]) == 1000


def test_dashboard_payload_period_helpers_separate_report_period_and_coverage() -> None:
    report_period = period_label_from_value(
        "2026-03-01 - 2026-06-17",
        "01.04.2026 - 17.06.2026",
    )
    coverage = period_label_from_value(
        "2026-04-01 - 2026-06-17",
        "",
    )

    assert report_period == "01.03.2026 - 17.06.2026"
    assert coverage == "01.04.2026 - 17.06.2026"
    assert period_boundaries_from_label(coverage) == ("2026-04-01", "2026-06-17")
    assert (
        analysis_period_text(
            "Период анализа: март, апрель, май, июнь; июнь неполный",
            "fallback",
        )
        == "март, апрель, май, июнь; июнь неполный"
    )


def test_report_option_bounds_use_week_closing_dates() -> None:
    rows = [
        {
            "week": "2026-06-29",
            "month": "Июль 2026",
        },
        {
            "week": "2026-07-06",
            "accountingPeriodDate": "2026-07-07",
            "month": "Июль 2026",
        },
    ]

    assert repository.options_payload(rows)["periodStart"] == "2026-07-05"
    assert repository.options_payload(rows)["periodEnd"] == "2026-07-12"
    assert dashboard_payload.options(rows)["periodStart"] == "2026-07-05"
    assert dashboard_payload.options(rows)["periodEnd"] == "2026-07-12"


def test_document_reconciliation_parser_reads_excel_control_columns() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сверка документов 1С"
    sheet.append(
        [
            "Статус сверки",
            "Статус выплаты",
            "Статус периода",
            "Период продаж",
            "Ожидаемая дата документа",
            "Тип документа WB/1С",
            "Кабинет WB",
            "Организация 1С",
            "Номер отчета WB (сводный)",
            "WB отчет продаж",
            "WB отчет выкупов",
            "WB reportId в пакете",
            "Документы 1С",
            "Типы документов 1С",
            "Даты документов 1С",
            "WB продажи",
            "WB возвраты",
            "WB чистое",
            "1С продажи",
            "1С возвраты",
            "1С чистое",
            "Дельта продажи",
            "Дельта возвраты",
            "Дельта чистое",
            "WB количество для 1С",
            "1С количество",
            "Дельта количество",
            "WB сумма документа",
            "1С сумма документа",
            "Дельта сумма",
            "WB выкуп: retailAmountSum",
            "WB выкуп: forPaySum",
            "WB выкуп: bankPaymentSum",
            "1С расходная накладная",
            "Δ выкуп retail",
            "Δ выкуп к перечислению",
            "Δ выкуп банк",
            "PDF 8. К перечислению",
            "WB к перечислению (forPaySum)",
            "1С оборот взаиморасчетов",
            "Дельта к обороту 1С",
            "Строк регистра 1С",
            "Комментарий",
        ]
    )
    sheet.append(
        [
            "OK",
            "Нужен источник выплаты 1С",
            "неполный период",
            "2026-04-06 - 2026-04-12",
            "2026-04-12",
            "Уведомление о выкупе",
            "Кабинет A",
            "Организация A",
            "SUMMARY-1",
            "SUMMARY-SALES-1",
            "SUMMARY-BUYOUT-1",
            "7268072721",
            "DOC-BUYOUT-1",
            "РасходнаяНакладная",
            "2026-04-12",
            42,
            0,
            42,
            42,
            0,
            42,
            0,
            0,
            0,
            42,
            42,
            0,
            66003.74,
            66003.74,
            0,
            66003.74,
            53420.94,
            24541.31,
            39464.41,
            26539.33,
            13956.53,
            -14923.1,
            64000,
            64000,
            None,
            None,
            5,
            "Нужен источник выплаты 1С",
        ]
    )

    rows = document_reconciliation_rows(workbook)

    assert rows == [
        {
            "id": "document-reconciliation-1",
            "status": "OK",
            "payoutStatus": "Нужен источник выплаты 1С",
            "periodStatus": "неполный период",
            "documentReport": (
                "Уведомление о выкупе · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
            ),
            "salesPeriod": "2026-04-06 - 2026-04-12",
            "salesPeriodStart": "2026-04-06",
            "salesPeriodEnd": "2026-04-12",
            "expectedDocumentDate": "2026-04-12",
            "documentType": "Уведомление о выкупе",
            "cabinet": "Кабинет A",
            "organization": "Организация A",
            "summaryReportId": "SUMMARY-1",
            "weeklySalesReportId": "SUMMARY-SALES-1",
            "weeklyBuyoutReportId": "SUMMARY-BUYOUT-1",
            "wbReportIds": "7268072721",
            "onecDocuments": "DOC-BUYOUT-1",
            "onecDocumentTypes": "РасходнаяНакладная",
            "onecDocumentDates": "2026-04-12",
            "wbSalesQuantity": 42,
            "wbReturnQuantity": 0,
            "wbNetQuantity": 42,
            "onecSalesQuantity": 42,
            "onecReturnQuantity": 0,
            "onecNetQuantity": 42,
            "salesQuantityDelta": 0,
            "returnQuantityDelta": 0,
            "netQuantityDelta": 0,
            "wbQuantity": 42,
            "onecQuantity": 42,
            "quantityDelta": 0,
            "wbAmount": 66003.74,
            "onecAmount": 66003.74,
            "amountDelta": 0,
            "buyoutRetailAmountSum": 66003.74,
            "buyoutForPaySum": 53420.94,
            "buyoutBankPaymentSum": 24541.31,
            "onecExpenseInvoiceAmount": 39464.41,
            "buyoutRetailDelta": 26539.33,
            "buyoutForPayDelta": 13956.53,
            "buyoutBankDelta": -14923.1,
            "pdfBankPayment": 64000,
            "wbForPaySum": 64000,
            "onecSettlementTotal": None,
            "settlementDelta": None,
            "onecSourceRows": 5,
            "comment": "Нужен источник выплаты 1С",
        }
    ]


def test_report_requires_auth(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    assert client.get("/api/reports/report-1/summary").status_code == 401
    assert (
        client.get("/api/reports/report-1/document-reconciliation").status_code == 401
    )


def test_login_remember_me_extends_session_cookie(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={
            "session_ttl_hours": 1,
            "remember_me_session_ttl_hours": 48,
        },
    )

    response = client.post(
        "/api/auth/login",
        json={
            "email": "admin@example.com",
            "password": "secret",
            "remember_me": True,
        },
    )

    assert response.status_code == 200
    assert "Max-Age=172800" in response.headers["set-cookie"]


def test_overview_orders_kpis_analytics_and_readiness(tmp_path: Path) -> None:
    cabinet = make_client(tmp_path).get("/cabinet")

    assert cabinet.status_code == 200
    assert cabinet.text.index('id="kpi-grid"') < cabinet.text.index(
        'id="analytics-panel"'
    )
    assert cabinet.text.index('id="ozon-diagnostics-panel"') < cabinet.text.index(
        'id="analytics-panel"'
    )
    assert cabinet.text.index('id="analytics-panel"') < cabinet.text.index(
        'id="readiness-card"'
    )
    assert cabinet.text.index('id="preflight-title"') < cabinet.text.index(
        'id="data-trust-strip"'
    )
    assert cabinet.text.index('id="data-trust-strip"') < cabinet.text.index(
        'id="secondary-kpi-section"'
    )
    assert cabinet.text.index('id="secondary-kpi-section"') < cabinet.text.index(
        'id="tax-input-check-card"'
    )
    assert cabinet.text.index('id="tax-input-check-card"') < cabinet.text.index(
        'id="onec-kpi-section"'
    )
    assert cabinet.text.index('id="tax-input-check-card"') < cabinet.text.index(
        'id="analytics-panel"'
    )


def test_cabinet_shell_serves_login_without_report_data(tmp_path: Path) -> None:
    client = make_client(tmp_path)

    health = client.get("/api/health")
    assert health.status_code == 200
    assert health.json()["backendBuildId"] == (
        "20260723-logistics-r4-return-reasons-v1"
    )
    assert health.json()["staticBuildId"] == (
        "20260723-logistics-r4-return-reasons-v1"
    )

    page = client.get("/")
    assert page.status_code == 200
    assert "Кабинет отчета" in page.text
    assert 'id="runtime-banner"' in page.text
    assert "Убыточный товар" not in page.text
    assert "Нет себестоимости 1С" not in page.text

    cabinet = client.get("/cabinet")
    assert cabinet.status_code == 200
    assert "/static/app.js" in cabinet.text
    integrations = client.get("/integrations")
    assert integrations.status_code == 200
    assert "/static/app.js" in integrations.text
    ai_page = client.get("/ai")
    assert ai_page.status_code == 200
    assert "/static/app.js" in ai_page.text
    assert 'id="ai-widget-overlay"' in ai_page.text
    assert 'id="ai-widget-close"' in ai_page.text
    assert 'id="client-output-widget-overlay"' in cabinet.text
    assert 'id="client-output-widget-close"' in cabinet.text
    assert 'id="accounting-scenario-checks"' in cabinet.text
    assert 'data-check-panel="summary"' in cabinet.text
    assert 'id="integrations-widget-overlay"' in integrations.text
    assert 'id="integrations-widget-close"' in integrations.text
    assert 'id="mapping-widget-overlay"' in cabinet.text
    assert 'id="mapping-widget-close"' in cabinet.text
    assert 'aria-label="Закрыть сервис сопоставления"' in cabinet.text
    assert 'id="drilldown-widget-overlay"' in cabinet.text
    assert 'id="drilldown-widget-close"' in cabinet.text
    assert 'id="drilldown-sources"' in cabinet.text
    assert 'id="drilldown-guidance"' in cabinet.text
    assert 'id="drilldown-table-wrap"' in cabinet.text
    assert 'id="drilldown-rows-head"' in cabinet.text
    assert 'data-drilldown-preset="sources"' in cabinet.text
    assert 'data-drilldown-preset="missingCost"' in cabinet.text
    assert 'data-drilldown-preset="missingMapping"' in cabinet.text
    assert "Расшифровки проблем" in cabinet.text
    assert 'id="integrations-back-link"' not in integrations.text
    assert 'href="/ai"' not in cabinet.text
    assert 'href="/integrations"' not in cabinet.text
    assert 'aria-controls="client-output-widget-overlay"' in cabinet.text
    assert 'aria-controls="ai-widget-overlay"' in cabinet.text
    assert 'aria-controls="integrations-widget-overlay"' in cabinet.text
    assert 'id="rows-filter-form"' in cabinet.text
    assert "filter-document-report" not in cabinet.text
    assert 'id="apply-filters-button"' not in cabinet.text
    assert 'id="report-select"' not in cabinet.text
    assert 'class="report-switcher"' not in cabinet.text
    assert "Применить" not in cabinet.text
    assert 'id="topbar-cabinet-select"' in cabinet.text
    assert 'id="topbar-period-start"' in cabinet.text
    assert 'id="topbar-period-end"' in cabinet.text
    assert 'id="topbar-period-select"' not in cabinet.text
    assert 'id="new-client-button"' in cabinet.text
    assert 'id="new-client-widget-overlay"' in cabinet.text
    assert 'id="new-client-form"' in cabinet.text
    assert "Добавить клиента" in cabinet.text
    assert "Новый клиент" in cabinet.text
    assert 'id="report-build-button"' in cabinet.text
    assert 'aria-controls="report-wizard-overlay"' in cabinet.text
    assert 'id="report-wizard-overlay"' in cabinet.text
    assert 'id="report-wizard-form"' in cabinet.text
    assert 'id="report-wizard-period-start"' in cabinet.text
    assert 'id="report-wizard-period-end"' in cabinet.text
    assert 'id="report-wizard-period-hint"' in cabinet.text
    assert 'id="report-wizard-dry-run"' not in cabinet.text
    assert '<ol class="report-wizard-steps"' in cabinet.text
    assert 'aria-current="step"' in cabinet.text
    assert 'id="report-wizard-current"' in cabinet.text
    assert 'id="report-wizard-current-period"' in cabinet.text
    assert 'id="report-wizard-current-download"' in cabinet.text
    assert 'id="report-wizard-check"' in cabinet.text
    assert 'id="report-wizard-reset"' in cabinet.text
    assert "Не относится к настройкам нового отчёта ниже" in cabinet.text
    assert "Проверить источники без создания" in cabinet.text
    assert "Сформировать другой период" in cabinet.text
    assert 'id="report-wizard-result"' in cabinet.text
    assert 'id="report-wizard-excel-download"' in cabinet.text
    assert 'id="report-wizard-client-report-generate"' in cabinet.text
    assert 'id="report-wizard-docx-download"' in cabinet.text
    assert 'id="report-wizard-pdf-download"' in cabinet.text
    assert "Отчёт готов — выберите файл" in cabinet.text
    assert "Сформировать отчёт" in cabinet.text
    assert 'id="report-download-button"' in cabinet.text
    assert "Скачать Excel" in cabinet.text
    assert "Кабинет МП" in cabinet.text
    assert 'aria-label="Фильтр по кабинету маркетплейса"' in cabinet.text
    assert "Дата начала" in cabinet.text
    assert "Дата конца" in cabinet.text
    assert "products-table" in cabinet.text
    assert 'class="products-table data-table report-rows-table"' in cabinet.text
    assert 'id="rows-pagination"' in cabinet.text
    assert 'id="rows-page-prev"' in cabinet.text
    assert 'id="rows-page-next"' in cabinet.text
    assert "Себестоимость" in cabinet.text
    assert "Остаток после" not in cabinet.text
    assert "НДС-корректировка P&amp;L" in cabinet.text
    assert "Итог после включ. налогов" in cabinet.text
    assert 'class="products-table data-table liquidity-table"' in cabinet.text
    assert 'class="products-table data-table lost-sales-table"' in cabinet.text
    assert 'id="analytics-panel"' in cabinet.text
    assert 'id="action-insights-panel"' in cabinet.text
    assert 'id="action-insights-list"' in cabinet.text
    assert "Аналитика" in cabinet.text
    assert 'id="money-trend-chart"' in cabinet.text
    assert 'id="unit-pl-table"' in cabinet.text
    assert "Прибыли и убытки юнит-экономики" in cabinet.text
    assert "Для ОСНО выручка и расходы показываются без НДС." in cabinet.text
    assert 'id="loss-drivers-chart"' in cabinet.text
    assert 'id="returns-chart"' in cabinet.text
    assert 'id="data-trust-strip"' in cabinet.text
    assert 'id="lost-margin-chart"' in cabinet.text
    assert 'id="tax-input-check-card"' in cabinet.text
    assert 'class="analytics-chart tax-input-chart-card"' in cabinet.text
    assert 'id="ozon-article-economics-card"' in cabinet.text
    assert 'id="ozon-article-economics-chart"' in cabinet.text
    assert (
        'class="panel full-width detail-workspace report-page-section"' in cabinet.text
    )
    assert 'data-detail-tab="liquidity"' in cabinet.text
    assert 'data-detail-tab="lostSales"' in cabinet.text
    assert 'id="reconciliation-hub-panel"' in cabinet.text
    assert 'id="reconciliation-hub-overlay"' not in cabinet.text
    assert 'data-reconciliation-tab="documents"' in cabinet.text
    assert 'data-reconciliation-tab="cogs"' in cabinet.text
    assert 'data-reconciliation-tab="expenses"' in cabinet.text
    assert 'data-reconciliation-tab="buyouts"' in cabinet.text
    assert 'data-detail-tab="products"' in cabinet.text
    assert 'id="integration-provider-tabs"' in cabinet.text
    assert 'data-integration-provider-filter="ozon_api"' in cabinet.text
    assert 'id="ozon-diagnostics-panel"' in cabinet.text
    assert 'id="ozon-preview-rows"' in cabinet.text
    assert 'id="ozon-issue-list"' in cabinet.text
    assert 'id="ozon-vitrine-status"' in cabinet.text
    assert 'id="ozon-pnl-grid"' not in cabinet.text
    assert 'id="ozon-buyout-rows"' in cabinet.text
    assert 'id="ozon-diagnostic-message"' in cabinet.text
    assert 'id="ozon-mapping-rows"' in cabinet.text
    assert "Служебная витрина" in cabinet.text
    assert "Ozon: расчет экономики" in cabinet.text
    assert "Ozon: расчет и сверка" not in cabinet.text
    assert "Ошибки Ozon" in cabinet.text
    assert "Финансовые сигналы" in cabinet.text
    assert "Итоги P&amp;L" not in cabinet.text
    assert "Ozon + 1C" in cabinet.text
    assert "Выкупы Ozon" in cabinet.text
    assert "Ozon + 1C" in cabinet.text
    assert (
            "styles.css?v=20260723-logistics-r4-return-reasons-v1"
        in cabinet.text
    )
    assert (
            "app.js?v=20260723-logistics-r4-return-reasons-v1"
        in cabinet.text
    )
    assert "Очередь аналитика" in cabinet.text
    assert "не выбирает номенклатуру 1C автоматически" in cabinet.text
    assert "Источники и сопоставление" in cabinet.text
    assert "Техническая сверка загрузки Ozon" in cabinet.text
    assert "Источники Ozon + 1C" in cabinet.text
    assert "Сопоставление Ozon → 1C" in cabinet.text
    assert "Ozon finance" not in cabinet.text
    assert "cash-flow" not in cabinet.text
    assert "Ozon: детализация по товарам" in cabinet.text
    assert "Offer / SKU" in cabinet.text
    assert "Комиссии / услуги" in cabinet.text
    assert "Партнерские услуги" in cabinet.text
    assert "Прибыль до налогов / маржа" in cabinet.text
    assert "Причина / действие" in cabinet.text
    assert "Номенклатура 1С" in cabinet.text
    assert cabinet.text.index('id="kpi-grid"') < cabinet.text.index(
        'id="ozon-diagnostics-panel"'
    )
    assert cabinet.text.index('id="ozon-diagnostics-panel"') < cabinet.text.index(
        'id="preflight-title"'
    )
    assert cabinet.text.index('id="preflight-title"') < cabinet.text.index(
        'id="analytics-panel"'
    )
    assert 'id="liquidity-summary"' in cabinet.text
    assert 'class="metric-grid liquidity-insight-grid"' in cabinet.text
    assert "МД1 наценка" in cabinet.text
    assert "МД6 до налогов" in cabinet.text
    assert "Маржа" in cabinet.text
    assert "Юнит-экономика" in cabinet.text
    assert "Расчетная таблица" in cabinet.text
    assert cabinet.text.index('data-detail-tab="products"') < cabinet.text.index(
        'data-detail-tab="liquidity"'
    )
    assert 'data-detail-panel="liquidity"' in cabinet.text
    assert 'data-detail-panel="lostSales"' in cabinet.text
    assert 'data-reconciliation-panel="documents"' in cabinet.text
    assert 'data-reconciliation-panel="cogs"' in cabinet.text
    assert 'data-reconciliation-panel="expenses"' in cabinet.text
    assert 'data-reconciliation-panel="buyouts"' in cabinet.text
    assert 'data-detail-panel="products"' in cabinet.text
    assert 'data-row-preset="returns"' in cabinet.text
    assert 'id="onec-reconciliation-filter-form"' in cabinet.text
    assert 'id="onec-filter-delta-only"' in cabinet.text
    assert 'id="ai-panel"' in cabinet.text
    assert 'id="integrations-panel"' in cabinet.text
    assert 'id="next-action-upload-form"' in cabinet.text
    assert 'id="next-action-upload-file"' in cabinet.text
    assert 'class="file-picker upload-file-button"' in cabinet.text
    assert "Вставить файл из 1С" in cabinet.text
    assert "Обновить сопоставление" in cabinet.text
    assert "data-tooltip" in cabinet.text
    assert 'id="source-refresh-panel"' in cabinet.text
    assert 'id="source-refresh-mapping-form"' in cabinet.text
    assert 'id="source-refresh-steps"' in cabinet.text
    assert 'id="source-refresh-incremental-run"' in cabinet.text
    assert 'id="source-refresh-full-run"' in cabinet.text
    assert 'id="source-refresh-ozon-run"' in cabinet.text
    assert cabinet.text.index('id="source-refresh-panel"') < cabinet.text.index(
        'id="kpi-grid"'
    )
    assert cabinet.text.index('id="source-refresh-panel"') < cabinet.text.index(
        'id="integrations-widget-overlay"'
    )
    assert "Данные и расчёт" in cabinet.text
    assert "Сопоставление клиента" in cabinet.text
    assert "Проверить готовность" in cabinet.text
    assert "Обновить последние данные" in cabinet.text
    assert "Полная пересборка истории" in cabinet.text
    assert "Обновить статус" in cabinet.text
    assert "Обновление данных маркетплейса" not in cabinet.text
    assert 'id="mapping-upload-form"' not in cabinet.text
    assert 'id="mapping-upload-file"' not in cabinet.text
    assert 'id="client-structure-panel"' not in cabinet.text
    assert "Источники и свежесть" not in cabinet.text
    assert "Организации и WB-кабинеты" not in cabinet.text
    assert 'class="control-room report-page-section"' in cabinet.text
    assert 'class="decision-strip readiness-neutral"' in cabinet.text
    assert 'class="panel money-strip report-page-section"' in cabinet.text
    assert 'id="secondary-kpi-section"' in cabinet.text
    assert 'id="secondary-kpi-grid"' in cabinet.text
    assert "Дополнительные показатели" in cabinet.text
    assert "Ключевые показатели" in cabinet.text
    assert "12 месяцев показываются как год" in cabinet.text
    assert cabinet.text.index('id="preflight-title"') < cabinet.text.index(
        'id="secondary-kpi-section"'
    )
    assert cabinet.text.index('id="secondary-kpi-section"') < cabinet.text.index(
        'id="onec-kpi-section"'
    )
    assert "analytics-chart-wide sales-trend-card" in cabinet.text
    assert "sales-trend-chart" in cabinet.text
    assert 'class="decision-support-grid report-page-section"' in cabinet.text
    assert 'class="panel preflight-panel report-page-section"' in cabinet.text
    assert "Готовность, деньги и следующий шаг по клиенту" not in cabinet.text
    assert "Финансовая картина" not in cabinet.text
    assert "Что важно по деньгам" not in cabinet.text
    assert '<h2 id="preflight-title">Перед отправкой' not in cabinet.text
    assert "Смарт-процесс подготовки" not in cabinet.text
    assert cabinet.text.index('id="kpi-title"') < cabinet.text.index(
        'id="readiness-card"'
    )
    assert cabinet.text.index('id="quality-title"') < cabinet.text.index(
        'id="blocking-title"'
    )
    assert 'id="quality-summary-text"' in cabinet.text
    assert 'id="quality-progress-fill"' in cabinet.text
    assert "Что проверить в отчете" in cabinet.text
    assert "Как выгрузить файл из 1С" in cabinet.text
    assert "Вывести список" in cabinet.text
    assert "MXL сюда не загружаем" in cabinet.text
    assert "Открыть проблемные строки" in cabinet.text
    assert "Исправить сейчас" in cabinet.text
    assert "В работе у аналитика" in cabinet.text
    assert "Готово к отправке" in cabinet.text
    assert 'id="command-checklist"' not in cabinet.text
    assert 'id="done-reasons"' in cabinet.text
    assert 'id="next-action-button"' in cabinet.text
    assert 'id="client-output-button"' in cabinet.text
    assert 'id="client-report-generate-button"' in cabinet.text
    assert 'id="client-report-excel-download"' in cabinet.text
    assert 'id="client-report-docx-download"' in cabinet.text
    assert 'id="client-report-pdf-download"' in cabinet.text
    assert "Сформируйте отчёт клиенту" in cabinet.text
    assert 'class="brand-lockup is-info"' in cabinet.text
    assert 'class="workspace-sidebar"' in cabinet.text
    assert 'data-workspace-nav="overview"' in cabinet.text
    assert 'data-workspace-nav="checks"' in cabinet.text
    assert 'data-workspace-nav="tables"' in cabinet.text
    assert 'data-workspace-nav="guide"' in cabinet.text
    assert 'id="user-guide-page"' in cabinet.text
    assert 'data-workspace-panel="guide"' in cabinet.text
    assert "Как пользоваться сервисом" in cabinet.text
    assert 'id="workspace-actions-menu"' in cabinet.text
    assert 'class="secondary-button session-button"' in cabinet.text
    assert "Отчёт клиенту" in cabinet.text
    assert "Текст для клиента" not in cabinet.text
    assert "Клиентский вывод" not in cabinet.text
    assert 'id="cost-review-workflow"' in cabinet.text
    assert 'data-check-panel="cost"' in cabinet.text
    assert "Найти строки" in cabinet.text
    assert "Проверить себестоимость" in cabinet.text
    assert "Подтвердить" in cabinet.text
    assert 'id="ai-open-button"' in cabinet.text
    assert 'class="ai-assistant-icon"' in cabinet.text
    assert 'id="ai-context-strip"' in cabinet.text
    assert "AI-аналитик" in cabinet.text
    assert "Помощник без изменения данных" in cabinet.text
    assert "или готовности" in cabinet.text
    assert "mapping и обязательных" not in cabinet.text
    assert "Read-only помощник" not in cabinet.text
    assert "P&amp;L юнит-экономики" not in cabinet.text
    assert "report-only-control" in cabinet.text
    assert 'id="report-load-retry-button"' in cabinet.text
    assert client.get("/api/reports").status_code == 401


def test_all_web_tables_use_shared_accessible_column_sorting(tmp_path: Path) -> None:
    client = make_client(tmp_path)

    cabinet = client.get("/cabinet")
    workflow_page = client.get("/static/accounting-workflows.html")
    script = client.get("/static/sortable-tables.js")
    app_script = client.get("/static/app.js")
    styles = client.get("/static/sortable-tables.css")

    assert cabinet.status_code == 200
    assert workflow_page.status_code == 200
    assert script.status_code == 200
    assert app_script.status_code == 200
    assert styles.status_code == 200
    asset_path = "/static/sortable-tables.js?v=20260718-column-sorting-v3"
    stylesheet_path = "/static/sortable-tables.css?v=20260718-column-sorting-v1"
    assert asset_path in cabinet.text
    assert asset_path in workflow_page.text
    assert stylesheet_path in cabinet.text
    assert stylesheet_path in workflow_page.text
    assert 'const HEADER_SELECTOR = "table thead th"' in script.text
    assert 'new Intl.Collator("ru"' in script.text
    assert 'header.setAttribute("aria-sort", state.direction)' in script.text
    assert 'event.key !== "Enter" && event.key !== " "' in script.text
    assert "new MutationObserver(handleMutations)" in script.text
    assert "if (indicator.textContent !== value)" in script.text
    assert 'setIndicatorText(indicator, "↕")' in script.text
    assert 'left.kind === "empty" ? 1 : -1' in script.text
    assert 'new CustomEvent("sortable-table-sort"' in script.text
    assert 'table.dataset.sortMode === "remote"' in script.text
    assert 'id="report-rows-table"' in cabinet.text
    assert 'data-sort-key="logisticsTotal"' in cabinet.text
    assert 'data-sort-disabled="true"' in cabinet.text
    assert "sort_by: state.rowsSortBy" in app_script.text
    assert "sortBy: state.logisticsProductsSortBy" in app_script.text
    assert "sortBy: state.logisticsDimensionsSortBy" in app_script.text
    assert "sortBy: state.logisticsOrdersSortBy" in app_script.text
    assert 'table.dataset.sortScope = "tax-input"' in app_script.text
    assert "function sortTaxInputRows(rows)" in app_script.text
    assert "state.rowsOffset = 0;" in app_script.text
    assert ".sortable-table-header[aria-sort]" in styles.text


def test_user_guide_is_generated_from_current_interface_metadata(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    cabinet = client.get("/cabinet")
    app_js = client.get("/static/app.js")
    styles = client.get("/static/styles.css")

    assert cabinet.status_code == 200
    assert 'id="guide-start-list"' in cabinet.text
    assert 'id="guide-sections-list"' in cabinet.text
    assert 'id="guide-actions-list"' in cabinet.text
    assert 'id="guide-checks-list"' in cabinet.text
    assert 'data-guide-entry="start"' in cabinet.text
    assert 'data-guide-entry="sections"' in cabinet.text
    assert 'data-guide-entry="actions"' in cabinet.text
    assert 'data-guide-entry="checks"' in cabinet.text
    assert 'data-guide-roles="consultant,admin"' in cabinet.text
    assert "Как работать с вкладкой «Проверки»" in cabinet.text
    assert "Прочитайте сводку запуска" in cabinet.text
    assert "Обычный рабочий сценарий" in cabinet.text
    assert "Используйте только для первой загрузки" in cabinet.text
    assert 'id="preflight-panel"' in cabinet.text

    assert app_js.status_code == 200
    assert 'if (value === "guide")' in app_js.text
    assert '["overview", "checks", "tables", "guide"]' in app_js.text
    assert "#tables/${tableScenario}" in app_js.text
    assert 'tableScenario: "logistics"' in app_js.text
    assert "function renderUserGuide()" in app_js.text
    assert "document.querySelectorAll(`[data-guide-entry=" in app_js.text
    assert "guideEntryVisibleForRole" in app_js.text
    assert "document.createElement(\"li\")" in app_js.text
    assert "list.replaceChildren(...cards)" in app_js.text
    assert 'renderGuideGroup("checks", els.guideChecksList, role)' in app_js.text

    assert styles.status_code == 200
    assert 'data-active-workspace="guide"' in styles.text
    assert "grid-template-columns: repeat(4, minmax(0, 1fr));" in styles.text


def test_ai_sse_ui_restores_history_and_contains_modal_overflow(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)

    app_js = client.get("/static/app.js")
    styles = client.get("/static/styles.css")

    assert app_js.status_code == 200
    assert styles.status_code == 200
    assert (
        "/api/ai/threads?report_id=${encodeURIComponent(reportId)}&limit=1"
        in app_js.text
    )
    assert "function renderAiThread(thread)" in app_js.text
    assert 'els.aiSourceStatus.textContent = "Анализирую…"' in app_js.text
    assert 'throw new Error("AI stream ended without a final answer")' in app_js.text
    assert "grid-template-rows: auto auto auto minmax(0, 1fr);" in styles.text
    assert ".ai-widget {\n    display: block;\n    overflow-y: auto;" in styles.text
    assert "@media (max-height: 700px)" in styles.text
    assert "overscroll-behavior: contain;" in styles.text
    assert "word-break: break-word;" in styles.text


def test_tax_load_v2_renderer_localizes_contract_and_keeps_tables_accessible(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)

    app_js = client.get("/static/app.js")
    scenarios = client.get("/static/report-scenarios.js")
    tax_load = client.get("/static/tax-load-report.js")
    styles = client.get("/static/styles.css")

    assert app_js.status_code == 200
    assert scenarios.status_code == 200
    assert tax_load.status_code == 200
    assert styles.status_code == 200
    assert 'header.scope = "col"' in scenarios.text
    assert "wrapper.tabIndex = 0" in scenarios.text
    assert 'wrapper.setAttribute("aria-label", options.label)' in scenarios.text
    assert "Нужна проверка бухгалтера" in tax_load.text
    assert "Открытых дозапросов нет" in tax_load.text
    assert "В нагрузке ФНС" in tax_load.text
    assert "formatMoney" in tax_load.text
    assert "formatDate" in tax_load.text
    assert "accountingScenarioStatusLabel" in app_js.text
    assert "Внутренний предварительный отчёт" in app_js.text
    assert "status ${status}" not in app_js.text
    assert ".scenario-table-wrap:focus-visible" in styles.text
    assert ".scenario-table-hint" in styles.text
    assert ".scenario-heading .status-pill" in styles.text


def test_cabinet_static_assets_use_readiness_api_and_safe_rendering(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)

    app_js = client.get("/static/app.js")
    styles = client.get("/static/styles.css")
    cabinet = client.get("/cabinet")
    assert app_js.status_code == 200
    assert styles.status_code == 200
    assert cabinet.status_code == 200
    assert "/api/reports" in app_js.text
    assert "/summary" in app_js.text
    assert "/logistics/tariffs" in app_js.text
    assert "logisticsTariffsAvailable" in app_js.text
    assert 'id="logistics-tariffs"' in cabinet.text
    assert "Тарифы и коэффициенты WB" in cabinet.text
    assert "/logistics/measurements" in app_js.text
    assert "logisticsMeasurementsAvailable" in app_js.text
    assert "if (logisticsMeasurementsAvailable())" in app_js.text
    assert "resetLogisticsMeasurements({ hide: true })" in app_js.text
    assert 'id="logistics-measurements"' in cabinet.text
    assert "Контрольные замеры и удержания WB" in cabinet.text
    assert "/logistics/routes" in app_js.text
    assert "logisticsRoutesAvailable" in app_js.text
    assert 'id="logistics-routes"' in cabinet.text
    assert "Склады и направления" in cabinet.text
    assert cabinet.text.index('id="logistics-measurements"') < cabinet.text.index(
        'id="logistics-tariffs"'
    )
    assert cabinet.text.index('id="logistics-tariffs"') < cabinet.text.index(
        'id="logistics-routes"'
    )
    assert cabinet.text.index('id="logistics-routes"') < cabinet.text.index(
        'id="logistics-products-title"'
    )
    assert "/logistics/return-reasons" in app_js.text
    assert "logisticsReturnReasonsAvailable" in app_js.text
    assert "if (logisticsReturnReasonsAvailable())" in app_js.text
    assert "resetLogisticsReturnReasons({ hide: true })" in app_js.text
    assert 'id="logistics-return-reasons"' in cabinet.text
    assert "Причины возвратов" in cabinet.text
    assert "Покрытие неизвестно" in app_js.text
    assert cabinet.text.index('id="logistics-routes"') < cabinet.text.index(
        'id="logistics-return-reasons"'
    )
    assert cabinet.text.index(
        'id="logistics-return-reasons"'
    ) < cabinet.text.index('id="logistics-products-title"')
    assert "20260723-logistics-r4-return-reasons-v1" in cabinet.text
    assert ".logistics-tariffs-table" in styles.text
    assert ".logistics-return-reasons-coverage" in styles.text
    measurement_cell_rule = styles.text.split(
        ".logistics-table.logistics-measurements-table th,", 1
    )[1].split("}", 1)[0]
    assert "min-width: 0" in measurement_cell_rule
    assert "overflow-wrap: anywhere" in measurement_cell_rule
    assert "white-space: normal" in measurement_cell_rule
    assert "/freshness" in app_js.text
    assert "/client-draft" in app_js.text
    assert "setTopbarNotice" in app_js.text
    assert "reportTopbarTone" in app_js.text
    assert "Открыть отчёт для клиента" in app_js.text
    assert "текст для клиента" not in app_js.text.lower()
    assert "клиентский вывод" not in app_js.text.lower()
    assert "/messages/stream" in app_js.text
    assert 'apiURL: "/api/chatkit"' in app_js.text
    assert "domainKey" not in app_js.text
    assert "answerSource" in app_js.text
    assert "latestSourceRefresh" in app_js.text
    assert "integrationEffectiveStatus" in app_js.text
    assert "runtimeCheckedAt" in app_js.text
    assert "1С недоступна" in app_js.text
    assert "updateReportBuildButton" in app_js.text
    assert "generatedAtIso" in app_js.text
    assert "openReportWizard" in app_js.text
    assert "onReportWizardSubmit" in app_js.text
    assert "onReportWizardCheck" in app_js.text
    assert "renderReportWizardResult" in app_js.text
    assert "reportWizardPublishedReport" in app_js.text
    assert "reportWizardGeneratedReportId" in app_js.text
    assert "syncReportWizardRefresh" in app_js.text
    assert "generateClientAnalyticalReport" in app_js.text
    assert "/analytical-report" in app_js.text
    assert "Отчёт клиенту ещё не сформирован" in app_js.text
    assert "Черновик еще не подготовлен" not in app_js.text
    assert ".report-wizard-result" in styles.text
    assert ".client-report-actions" in styles.text
    assert "period_start: periodStart || null" in app_js.text
    assert "period_end: periodEnd || null" in app_js.text
    assert (
        'const scope = els.clientReportScope.value || "last_closed_week"'
        in app_js.text
    )
    assert "Последняя закрытая неделя" in cabinet.text
    assert "Проверить источники без создания" in cabinet.text
    assert "Создать Excel за ${periodLabel}" in app_js.text
    assert "els.reportWizardPeriodHint.hidden = customPeriod" in app_js.text
    assert "dry_run: Boolean(dryRun)" in app_js.text
    assert "Только проверить готовность" not in cabinet.text
    assert "Только проверить готовность" not in app_js.text
    assert "Не удалось подготовить DOCX и PDF." in app_js.text
    assert "Сформированный Excel остаётся доступен." in app_js.text
    assert "Обновить DOCX и PDF" in app_js.text
    assert "Сформировать заново" not in app_js.text
    assert "Отчёт формируется" in app_js.text
    assert "Данные обновляются" in app_js.text
    assert "Нет подтверждённой активности" in app_js.text
    assert "Нет отклика worker" not in app_js.text
    assert "worker-контролем" not in app_js.text
    assert "За доступный период" in app_js.text
    assert "без экстраполяции" in app_js.text
    assert "calculationPeriodStart" in app_js.text
    assert "calculationPeriodEnd" in app_js.text
    assert "reportFreshnessSubtitle" in app_js.text
    assert "taxProfileNeedsRebuild" in app_js.text
    assert "профиль требует пересборки" in app_js.text
    assert "reportCoversRefresh" in app_js.text
    assert "Данные в отчете" in app_js.text
    assert "Последнее обновление данных" in app_js.text
    assert "sourceRefreshModeText" in app_js.text
    assert 'daily: "ежедневный"' in app_js.text
    assert 'incremental: "последние 28 дней"' in app_js.text
    assert "sourceRefreshAutoOpenRunId" in app_js.text
    assert 'mode: "incremental"' in app_js.text
    assert "sourceRefreshNewReport" not in app_js.text
    assert "sourceRefreshIssueSummary" not in app_js.text
    assert "/mapping-file" in app_js.text
    assert "mappingUpload" in app_js.text
    assert "mappingUploadControls" in app_js.text
    assert "onMappingUpload" in app_js.text
    assert "mappingUploadRefreshStatus" in app_js.text
    assert "/source-refresh/latest" in app_js.text
    assert "/source-refresh" in app_js.text
    assert "/ozon-diagnostics" in app_js.text
    assert "?mode=${encodeURIComponent(mode)}" in app_js.text
    assert "Черновик Ozon — требуется проверка" in app_js.text
    assert "Доступна служебная витрина Ozon + 1C" in app_js.text
    assert "/ozon-diagnostics/export.xlsx" in app_js.text
    assert "reportDownloadContext" in app_js.text
    assert "updateReportDownloadControl" in app_js.text
    assert "Статьи экономики Ozon" in app_js.text
    assert "Расходы по SKU из детализации Ozon" in app_js.text
    assert "Часть расходов распределена" in app_js.text
    assert "нераспределенный остаток" in app_js.text
    assert "сверка 1C/Ozon" in app_js.text
    assert "ozonPartnerServices" in app_js.text
    assert "renderOzonMoneyTrendChart" in app_js.text
    assert "ozonMoneyTrendRows" in app_js.text
    assert "ozonMoneyTrendTotalsRow" in app_js.text
    assert "ozonMoneyTrendRowsFromItems" in app_js.text
    assert "ozonPeriodMonthKey" in app_js.text
    assert "renderOzonArticleEconomics" in app_js.text
    assert "ozonArticleEconomicsCards" in app_js.text
    assert "ozonArticleEconomicsCard" in app_js.text
    assert "ozonArticleDrilldownRows" in app_js.text
    assert "reconciliationTotals" in app_js.text
    assert "регистр продаж 1C · включая выкупы" in app_js.text
    assert "P&L Ozon (включая выкупы)" in app_js.text
    assert "onec_sales_register_including_additional_documents" in app_js.text
    assert "SKU-экономика без выкупов" in app_js.text
    assert 'grid.className = "metric-grid ozon-economics-grid";' in app_js.text
    assert "renderAnalyticsMetricsGrid" in app_js.text
    assert "clearAnalyticsMetricsGrid" in app_js.text
    assert "resetOzonAnalyticsCardGrids" in app_js.text
    assert (
        'document.body.classList.toggle("ozon-analytics-mode", ozonMode)' in app_js.text
    )
    assert "Документный контроль Ozon + 1C" in app_js.text
    assert "Комиссионер, выкупы и расходы: что сходится и что проверить." in app_js.text
    assert "Сверка Ozon ↔ 1C" not in app_js.text
    assert "Комиссионер, выкупы и расходы по статьям." not in app_js.text
    assert 'els.moneyTrendTitle.textContent = "Динамика продаж";' in app_js.text
    assert (
        '"По месяцам текущего загруженного отчёта; 12 месяцев показываются как год."'
        in app_js.text
    )
    assert "renderOzonMartKpis" not in app_js.text
    assert "Расчетная витрина Ozon" not in app_js.text
    assert "Итоги экономики по товарам за выбранный период." not in app_js.text
    assert 'params.set("limit", "50")' in app_js.text
    assert 'params.set("period_start"' in app_js.text
    assert 'params.set("period_end"' in app_js.text
    assert 'params.set("wb_cabinet_id"' in app_js.text
    assert "params !== state.ozonDiagnosticsParams" in app_js.text
    assert 'applyTopbarFilter("cabinet")' in app_js.text
    assert "renderOzonIssues" in app_js.text
    assert "els.ozonIssuesPanel.hidden = false;" in app_js.text
    assert "renderOzonPnl" not in app_js.text
    assert "diagnostics.expenseReconciliation || {}" in app_js.text
    assert "expenseReconciliation.articleRows" in app_js.text
    assert "Из чего состоит дельта расходов" in app_js.text
    assert "Строки без пары" in app_js.text
    assert "setOzonDiagnosticCalculationSectionsVisible" in app_js.text
    assert (
        "const showDiagnosticCalculation = !shouldUseOzonWorkingView();" in app_js.text
    )
    assert "if (showDiagnosticCalculation) {" in app_js.text
    assert "renderOzonBuyouts" in app_js.text
    assert "ozonVitrineStatus" in app_js.text
    assert "ROW_PRESET_LABELS" in app_js.text
    assert "applyRowsFilterMode" in app_js.text
    assert "filteredOzonMartRows" in app_js.text
    assert "ozonMartRowMatchesQuery" in app_js.text
    assert "ozonMartRowMatchesPreset" in app_js.text
    assert "presetBar.hidden = false" in app_js.text
    assert "els.rowsFilterForm.hidden = false" in app_js.text
    assert "setRowsFilterHidden(els.filterMonth, ozonMode)" in app_js.text
    assert "Все статусы Ozon" in app_js.text
    assert "Без связи" in app_js.text
    assert "Выкупы" in app_js.text
    assert "Товар, артикул продавца, SKU, штрихкод, 1C" in app_js.text
    assert "Служебная витрина Ozon" in app_js.text
    assert "sourceRefreshPanel" in app_js.text
    assert "sourceRefreshSteps" in app_js.text
    assert "Сопоставление" in app_js.text
    assert "Проверка" in app_js.text
    assert "Обновление" in app_js.text
    assert "Отчет" in app_js.text
    assert "Диагностика" in app_js.text
    assert "runClientSourceRefresh" in app_js.text
    assert "sourceRefreshOzonRun" in app_js.text
    assert "payload.activeRun" in app_js.text
    assert "payload.latestAttempt" in app_js.text
    assert "sourceRefreshBlockedAttemptMessage" in app_js.text
    assert 'mode: "ozon-only"' in app_js.text
    assert "Загружаем служебную витрину Ozon + 1C без обязательного WB" in app_js.text
    assert "Проверить источники без создания" in cabinet.text
    assert "Запустите refresh" not in app_js.text
    assert "Отправляем файл и запускаем пересборку" in app_js.text
    assert "FormData" in app_js.text
    assert "Открыть детали источников" not in app_js.text
    assert "renderClientStructure" not in app_js.text
    assert "renderCommandChecklist" in app_js.text
    assert "renderNextAction" in app_js.text
    assert "renderAnalytics" in app_js.text
    assert "renderActionInsights" in app_js.text
    assert "renderOzonPreview" in app_js.text
    assert "loadOzonDiagnostics" in app_js.text
    assert "renderOzonDiagnosticsPayload" in app_js.text
    assert "shouldKeepOzonDiagnosticsVisible" in app_js.text
    assert "Ниже показана служебная витрина Ozon + 1C" in app_js.text
    assert "renderOzonPreflightWithoutReport" in app_js.text
    assert "Контроль перед отправкой пока недоступен" in app_js.text
    assert "SKU-P&L Ozon (без выкупов)" in app_js.text
    assert "НДС не выделен: поле «Себестоимость» из 1C" in app_js.text
    assert "Ozon-данные загружены" in app_js.text
    assert "const useOzonWorkingView = shouldUseOzonWorkingView();" in app_js.text
    assert "diagnostics?.ozonMart" in app_js.text
    assert "ozonUnitProfitMarginText" in app_js.text
    assert "partial_source" in app_js.text
    assert "Все кабинеты МП" in app_js.text
    assert "activeMarketplaceCabinets" in app_js.text
    assert "marketplaceCabinetLabel" in app_js.text
    assert "shouldShowOzonPreview" in app_js.text
    assert "ozonMappingRowNode" in app_js.text
    assert "openMappingAnalystQueue" in app_js.text
    assert "Открыть в очереди" in app_js.text
    assert "Выбрать эту номенклатуру" in app_js.text
    assert "Ozon → 1C" in app_js.text
    assert "артикул продавца → артикул 1C" in app_js.text
    assert (
        "loadIntegrations(context),\n      loadSourceRefreshStatus(context),"
        in app_js.text
    )
    assert "ozonFinanceRowNode" not in app_js.text
    assert "движение денежных средств" in app_js.text
    assert "ozonCollections" in app_js.text
    assert "integrationRowsForActiveProvider" in app_js.text
    assert "Ozon еще не подключен" in app_js.text
    assert "isWbClientCabinet(item)" in app_js.text
    assert 'label.includes("ozon seller")' in app_js.text
    assert "syncIntegrationsEntryPoint" in app_js.text
    assert "clientLoadToken" in app_js.text
    assert "currentClientLoadContext" in app_js.text
    assert "isCurrentClientLoad(context)" in app_js.text
    assert "loadReports(currentClientLoadContext())" in app_js.text
    assert "await Promise.allSettled([" in app_js.text
    assert "loadReportFreshness(reportId, context)" in app_js.text
    assert "Не удалось загрузить отчёт" in app_js.text
    assert "Статус свежести источников временно недоступен" in app_js.text
    assert "retryCurrentReportLoad" in app_js.text
    assert "const mappingAutoSync = refresh?.mappingAutoSync;" in app_js.text
    assert "await loadSourceRefreshStatus(context).catch" not in app_js.text
    assert "Загружаем клиента" in app_js.text
    assert "renderMetrics(els.kpiGrid, []);" in app_js.text
    assert "renderMetrics(els.qualityGrid, []);" in app_js.text
    assert "!els.ozonPreviewSummary" in app_js.text
    assert "renderIntegrationRowSafe" in app_js.text
    assert "renderCabinetManagerSafe" in app_js.text
    assert "syncSelectedClientFromControl" in app_js.text
    assert "Получаем подключения выбранного клиента только для чтения." in app_js.text
    assert "renderIntegrationsWithFallback(state.integrationItems)" in app_js.text
    assert "renderIntegrationsRecovery" in app_js.text
    assert "const client = selectedClient();" in app_js.text
    assert "asArray(client?.companies)" in app_js.text
    assert "runAnalyticsAction" in app_js.text
    assert "selectRowsPreset" in app_js.text
    assert "openProductsPreset" in app_js.text
    assert "openProductsMonth" in app_js.text
    assert "els.onecFilterDeltaOnly.checked = true" in app_js.text
    assert 'selectDetailTab("lostSales")' in app_js.text
    assert "renderMoneyTrendChart" in app_js.text
    assert "renderUnitProfitAndLossTable" in app_js.text
    assert "isTaxBridgeExpense" in app_js.text
    assert ".filter((row) => !isTaxBridgeExpense(row))" in app_js.text
    assert (
        "profitManagement ?? kpis.managementProfit ?? kpis.profitBeforeTax"
        in app_js.text
    )
    assert "renderLossDriversChart" in app_js.text
    assert "renderReturnsChart" in app_js.text
    assert "renderColumnChart" in app_js.text
    assert "sales-trend-svg" in app_js.text
    assert "sales-trend-crosshair" in app_js.text
    assert "compactMonthLabel" in app_js.text
    assert "profitAndLossTable" in app_js.text
    assert "analytics-column-chart" in app_js.text
    assert "analytics-pl-table" in app_js.text
    assert "dataset.analyticsAction" in app_js.text
    assert "onecReconciliationDelta" in app_js.text
    assert "renderWaterfallColumns" not in app_js.text
    assert "Chart." not in app_js.text
    assert "chart.js" not in app_js.text.lower()
    assert "echarts" not in app_js.text.lower()
    assert "d3." not in app_js.text
    assert "decisionHeadline" in app_js.text
    assert "preliminaryPeriodNotice" in app_js.text
    assert "Период предварительный: укажите это клиенту" in app_js.text
    assert "revenueWithVat" in app_js.text
    assert (
        'label: showRevenueWithVat ? "Выручка WB без НДС" : "Выручка WB"'
        in app_js.text
    )
    assert '"Выручка WB с НДС"' in app_js.text
    assert "onecRevenueSupportingCaption" in app_js.text
    assert "с НДС · календарный учёт" in app_js.text
    assert '"Выручка 1С с НДС"' not in app_js.text
    assert "onecRevenueWithVat" in app_js.text
    assert "wbDocumentRevenueWithVat" in app_js.text
    assert "accountingReconciliationDelta" in app_js.text
    assert '"Единый стандарт WB ↔ 1С"' in app_js.text
    assert '"Сверка комиссионера WB ↔ 1С"' in app_js.text
    assert '"Выкупы: первичка WB ↔ 1С"' in app_js.text
    assert '"Себестоимость продаж 1С"' in app_js.text
    assert '"Себестоимость 1С"' in app_js.text
    assert '"Расходы WB"' in app_js.text
    assert '"Итого к перечислению"' in app_js.text
    assert "wbForPaySum" in app_js.text
    assert "kpis: analytics.kpis || payload.kpis || {}" in app_js.text
    assert '"Услуги WB по документам 1С"' in app_js.text
    assert '"Сверка расходов WB ↔ 1С"' in app_js.text
    assert "openMarketplaceExpenseReconciliationWidget" in app_js.text
    assert "/marketplace-expense-reconciliation" in app_js.text
    assert "openCogsReconciliationWidget" in app_js.text
    assert "/cogs-reconciliation" in app_js.text
    assert "Юнит-экономика рассчитана. Документальная сверка" in app_js.text
    assert "profitUsesRevenueWithoutVat" in app_js.text
    assert "Сумма выкупа" in app_js.text
    assert "item.dataset.tooltip = String(formula)" in app_js.text
    assert "salesTrendPeriodLabel" in app_js.text
    assert "Выручка 1C Ozon · факт" in app_js.text
    assert "Источник: 1C OData · регистр продаж · включая выкупы" in app_js.text
    assert "Ozon API · ожидается в 1C" in app_js.text
    assert "Первичные документы 1C" in app_js.text
    assert "ozonRevenueDocumentControlNode" in app_js.text
    assert "openOzonDocumentControlDetails" in app_js.text
    assert 'target.querySelector("tr.is-review")' in app_js.text
    assert "Перепроверить после исправления" in app_js.text
    assert 'shareDisplay: "справочно"' in app_js.text
    assert "Недополученный маржинальный доход" in app_js.text
    assert "lostContributionMargin" in app_js.text
    assert "Нет подтверждающих документов" in app_js.text
    assert "tax-input-semantic-table" in app_js.text
    assert "taxInputCabinet" not in app_js.text
    assert "Фильтр сверки НДС по кабинету" not in app_js.text
    assert 'includes(deductionMode)' in app_js.text
    assert 'els.taxInputCard.hidden = true' in app_js.text
    assert "Право на вычет входящего НДС не подтверждено" in app_js.text
    assert "sourceRows.slice(0, 8)" not in app_js.text
    assert "taxInputPage" in app_js.text
    assert "monthStart" in app_js.text
    assert "isPartial" in app_js.text
    assert "Чистые продажи WB" in app_js.text
    assert "Продажи WB" in app_js.text
    assert "Возвратность" in app_js.text
    assert "Выручка / продажа" in app_js.text
    assert "item.unitProfit" in app_js.text
    assert "Убыточные строки" in app_js.text
    assert "Штрафы без продаж" in app_js.text
    assert "Финансовая проверка не пройдена" in app_js.text
    assert 'profitDisplay: profit === null ? "не рассчитано" : ""' in app_js.text
    assert "Прибыли и убытки не рассчитаны: финансовая проверка" not in app_js.text
    assert "Предварительный расчёт: есть замечания к качеству данных" in app_js.text
    assert "content.push(profitAndLossTable(rows, revenue))" in app_js.text
    assert "nonOkSourceCount" in app_js.text
    assert "refreshHasCollectionStatus" in app_js.text
    assert "applyTopbarFilter" in app_js.text
    assert "syncTopbarFiltersFromRows" in app_js.text
    assert "topbarCabinetSelect" in app_js.text
    assert "seenLabels" in app_js.text
    assert "topbarPeriodStart" in app_js.text
    assert "topbarPeriodEnd" in app_js.text
    assert "newClientButton" in app_js.text
    assert "onNewClientSubmit" in app_js.text
    assert "clientCreateErrorMessage" in app_js.text
    assert "Сервер еще не подхватил обновление" in app_js.text
    assert "Такой код контура уже используется" in app_js.text
    assert "reportSelect" not in app_js.text
    assert "renderReportSelect" not in app_js.text
    assert "payload.kpis" in app_js.text
    assert "payload.analytics" in app_js.text
    assert "filteredAnalyticsSummary" in app_js.text
    assert "bindAutoApplyingFilters" in app_js.text
    assert "applyRowsFilters" in app_js.text
    assert "REPORT_ROWS_PAGE_SIZE = 100" in app_js.text
    assert 'params.set("offset", String(state.rowsOffset))' in app_js.text
    assert "renderRowsPagination" in app_js.text
    assert "unitProfitBridge" in app_js.text
    assert 'return mode === "ozon" ? 12 : 40' in app_js.text
    assert "saved.rowPreset" not in app_js.text
    assert "debounce(applyRowsFilters" in app_js.text
    assert "/api/integrations" in app_js.text
    assert "isIntegrationsPage" in app_js.text
    assert "isAiPage" in app_js.text
    assert "renderAiPageHeader" in app_js.text
    assert "openAiWidget" in app_js.text
    assert "closeAiWidget" in app_js.text
    assert "integration-feedback" in app_js.text
    assert "editingIntegrationKey" in app_js.text
    assert "draftIntegration" in app_js.text
    assert "integration-compact-row" in app_js.text
    assert "integration-edit-form" in app_js.text
    assert "integration-more" in app_js.text
    assert "Новая карточка подключения" in app_js.text
    assert "Тип подключения" in app_js.text
    assert "Создать карточку" in app_js.text
    assert "createDraftIntegrationCard" in app_js.text
    assert "provider_base" in app_js.text
    assert "Настроить" in app_js.text
    assert "Изменить" in app_js.text
    assert "Отмена" in app_js.text
    assert "renderCabinetManager" in app_js.text
    assert "onCabinetManagerSubmit" in app_js.text
    assert "clientScopedFilterOptions" in app_js.text
    assert "Сохранить кабинет" in app_js.text
    assert "buildIntegrationRows" in app_js.text
    assert "findIntegrationForCabinet" in app_js.text
    assert "cabinet_name: form.dataset.cabinetName" in app_js.text
    assert "renderOnecSecretControls" in app_js.text
    assert "onec_base_url" in app_js.text
    assert "URL 1С/OData" in app_js.text
    assert "integration-card--onec-provider" in app_js.text
    assert "integration-optional-pill" in app_js.text
    assert "integration-subtle-action" in app_js.text
    assert "Кабинет / организация" not in app_js.text
    assert "Хранение encrypted storage" not in app_js.text
    assert "Опционально" in app_js.text
    assert "Кабинет продавца Ozon — только чтение" in app_js.text
    assert "Ozon кабинет" in app_js.text
    assert "clientId=...; apiKey=..." in app_js.text
    assert "Поля 1С заполнены" in app_js.text
    assert "Новый ключ введен в строке этого кабинета" in app_js.text
    assert "Сохранено. Секрет скрыт" in app_js.text
    assert "Сохраните ключ, затем проверьте подключение" in app_js.text
    assert "Готово к полному обновлению. Отчет еще не создан." in app_js.text
    assert "Отчет создан" in app_js.text
    assert "openClientOutputWidget" in app_js.text
    assert "openIntegrationsWidget" in app_js.text
    assert "openMappingWidget" in app_js.text
    assert 'return "#checks/cost"' in app_js.text
    assert 'dataWorkspace' not in app_js.text
    assert "configureWorkspaceFromLocation" in app_js.text
    assert "renderCostReview" in app_js.text
    assert "toggleCostReviewAcknowledgement" in app_js.text
    assert "renderAiContext" in app_js.text
    assert (
        'openMappingWidget({ marketplace: "wb", status: "review", search: "" })'
        in app_js.text
    )
    assert "closeAllWidgets" in app_js.text
    assert "clientOutputWidgetOverlay" in app_js.text
    assert "integrationsWidgetOverlay" in app_js.text
    assert "mappingWidgetOverlay" in app_js.text
    assert 'body.classList.add("widget-open")' in app_js.text
    assert "renderIntegrationsEmpty" in app_js.text
    assert "integrationsBackLink" not in app_js.text
    assert "aiWidgetOverlay" in app_js.text
    assert "integrationsPanel.scrollIntoView" not in app_js.text
    assert "aiPanel.scrollIntoView" not in app_js.text
    assert "storageMode" in app_js.text
    assert "lastCheck.message" in app_js.text
    assert "status_filter" in app_js.text
    assert "loss_class" in app_js.text
    assert "document_report" not in app_js.text
    assert "liquidityRows" in app_js.text
    assert "liquidity-rows" in app_js.text
    assert "function asArray" in app_js.text
    assert "sourceLoads = asArray" in app_js.text
    assert "summary.unitRows" not in app_js.text
    assert "URLSearchParams" in app_js.text
    assert "readiness" in app_js.text
    assert "qualitySummaryText" in app_js.text
    assert "qualityProgressFill" in app_js.text
    assert "renderDoneTasks" in app_js.text
    assert "doneReasons" in app_js.text
    assert "taskStatusStorageKey" in app_js.text
    assert "setTaskReviewed" in app_js.text
    assert "detailTabs" in app_js.text
    assert "detailPanels" in app_js.text
    assert "selectDetailTab" in app_js.text
    assert "liquiditySummary" in app_js.text
    assert "liquidityMainDriver" in app_js.text
    assert "renderLiquiditySummary" in app_js.text
    assert "Красная зона" in app_js.text
    assert "Потери в выборке" in app_js.text
    assert "statusLabel" in app_js.text
    assert "Проверить тип документа WB" in app_js.text
    assert "md1Markup" in app_js.text
    assert "md6BeforeTax" in app_js.text
    assert 'selectDetailTab(tab = "products")' in app_js.text
    assert "localStorage" in app_js.text
    assert "/document-reconciliation" in app_js.text
    assert "/financial-document-reconciliation" in app_js.text
    assert "/buyout-reconciliation" in app_js.text
    assert "renderFinancialReconciliation" in app_js.text
    assert "openBuyoutReconciliationWidget" in app_js.text
    assert "metric-action" in app_js.text
    assert "Дельта выручки комиссионера · 1С − WB" in app_js.text
    assert "Выкупы · накладные 1С" in app_js.text
    assert "Выкупы · первичка WB" in app_js.text
    assert "loadOnecReconciliation" in app_js.text
    assert "renderOnecReconciliation" in app_js.text
    assert "onecReconciliationFilterParams" in app_js.text
    assert "onec_reconciliation_review" in app_js.text
    assert "Отметить просмотренным" in app_js.text
    assert "Вернуть в работу" in app_js.text
    assert "reasonGuide" in app_js.text
    assert "cogs_reconciliation_failed" in app_js.text
    assert "costIssueBreakdown" in app_js.text
    assert "runReasonAction" in app_js.text
    assert "showRowsPreset" not in app_js.text
    assert "openDrilldownWidget" in app_js.text
    assert "selectDrilldownPreset" in app_js.text
    assert "renderDrilldownRows" in app_js.text
    assert "renderSourceDrilldown" in app_js.text
    assert "sourceStatusText" in app_js.text
    assert "appendTableCells" in app_js.text
    assert "tableRowClass" in app_js.text
    assert "statusTone" in app_js.text
    assert "is-missing-cost" in app_js.text
    assert "has-delta" in app_js.text
    assert 'products-panel").scrollIntoView' not in app_js.text
    assert "missingMapping" in app_js.text
    assert "Показать строки сопоставления" in app_js.text
    assert "Показать источники" in app_js.text
    assert "blocked_low_disk" in app_js.text
    assert "innerHTML" not in app_js.text

    css = client.get("/static/styles.css")
    assert css.status_code == 200
    assert ".analytics-calculation-note" in css.text
    assert "@media (max-width: 560px)" in css.text
    assert ".filters-bar" in css.text
    assert ".control-room" in css.text
    assert ".money-strip" in css.text
    assert ".decision-strip" in css.text
    assert ".decision-support-grid" in css.text
    assert ".preflight-panel" in css.text
    assert ".preflight-layout" in css.text
    assert ".task-board" in css.text
    assert ".task-column" in css.text
    assert ".task-card" in css.text
    assert ".task-card-actions" in css.text
    assert ".task-done-link" in css.text
    assert ".task-reopen-link" in css.text
    assert ".is-done" in css.text
    assert ".detail-workspace" in css.text
    assert ".detail-tabs" in css.text
    assert ".detail-tab-panel" in css.text
    assert ".new-client-widget" in css.text
    assert ".new-client-form" in css.text
    assert ".liquidity-summary" in css.text
    assert ".liquidity-insight-grid" in css.text
    assert ".metric-bad" in css.text
    assert ".liquidity-table" in css.text
    assert "min-width: 2420px" in css.text
    assert ".filter-checkbox" in css.text
    assert ".quality-progress" in css.text
    assert ".reason-item" in css.text
    assert ".reason-action-link" in css.text
    assert ".reason-hint" in css.text
    assert ".quality-diagnostics" in css.text
    assert ".drilldown-widget" in css.text
    assert ".drilldown-tabs" in css.text
    assert ".drilldown-table-wrap" in css.text
    assert ".drilldown-sources" in css.text
    assert ".source-load-card" in css.text
    assert ".source-load-status" in css.text
    assert ".next-action-panel" in css.text
    assert ".command-checklist" in css.text
    assert ".command-metrics" in css.text
    assert ".action-insights-panel" in css.text
    assert ".action-insight-card" in css.text
    assert ".row-preset-bar" in css.text
    assert ".report-rows-table" in css.text
    assert ".table-badge" in css.text
    assert ".products-table tbody tr.is-loss" in css.text
    assert ".products-table tbody tr.has-delta" in css.text
    assert "position: sticky" in css.text
    assert ".next-action-upload-form" in css.text
    assert ".next-action-controls" in css.text
    assert "[data-tooltip]" in css.text
    assert "top: calc(100% + 14px)" in css.text
    assert "border-bottom-color: var(--text)" in css.text
    assert "bottom: calc(100% + 14px)" not in css.text
    assert ".upload-file-button" in css.text
    assert ".upload-guidance" in css.text
    assert ".upload-help" in css.text
    assert ".upload-submit-button" in css.text
    assert "-webkit-overflow-scrolling: touch" in css.text
    assert "width: max-content" in css.text
    assert ".ai-workspace" in css.text
    assert ".widget-overlay" in css.text
    assert ".widget-shell" in css.text
    assert ".widget-actions" in css.text
    assert "body.widget-open" in css.text
    assert ".client-output-widget" in css.text
    assert ".integrations-widget" in css.text
    assert ".ai-widget" in css.text
    assert ".integration-card" in css.text
    assert ".integration-empty" in css.text
    assert ".integration-details" in css.text
    assert ".integration-feedback" in css.text
    assert ".integration-compact-row" in css.text
    assert ".integration-edit-form" in css.text
    assert ".integration-read-badge" in css.text
    assert ".integration-more" in css.text
    assert ".integration-cabinet-manager" in css.text
    assert ".integration-card-creator-form" in css.text
    assert ".integration-type-field" in css.text
    assert ".integration-cabinet-field" in css.text
    assert ".cabinet-manager-form" in css.text
    assert ".integration-list-header" in css.text
    assert ".integration-target" not in css.text
    assert ".integration-subtle-action" in css.text
    assert ".integration-compact-field" in css.text
    assert ".integration-card--onec-provider" in css.text
    assert ".integration-form--onec" in css.text
    assert ".onec-secret-fields" in css.text
    assert ".integration-toggle-field" in css.text
    assert "grid-template-areas" in css.text
    assert '"role status"' in css.text
    assert '"secret status"' in css.text
    assert ".integration-card--ok" in css.text
    assert ".integration-status-pill" in css.text
    assert ".client-structure-grid" not in css.text
    assert ".source-refresh-panel" in css.text
    assert ".source-refresh-collections" in css.text
    assert "#report-build-button.is-warning" in css.text
    assert ".report-wizard-form" in css.text
    assert ".report-wizard-steps" in css.text
    assert ".ozon-preview-grid" in css.text
    assert (
        '.ozon-analytics-mode .analytics-chart[aria-labelledby="loss-drivers-title"],'
        in css.text
    )
    assert (
        '.ozon-analytics-mode .analytics-chart[aria-labelledby="returns-chart-title"],'
        in css.text
    )
    assert (
        ".ozon-analytics-mode "
        '.analytics-chart[aria-labelledby="loss-drivers-title"] '
        ".analytics-chart-body" in css.text
    )
    assert (
        ".ozon-analytics-mode "
        '.analytics-chart[aria-labelledby="returns-chart-title"] '
        ".analytics-chart-body" in css.text
    )
    assert (
        ".ozon-analytics-mode "
        '.analytics-chart[aria-labelledby="ozon-article-economics-title"]'
        in css.text
    )
    assert (
        ".ozon-analytics-mode "
        '.analytics-chart[aria-labelledby="ozon-article-economics-title"] '
        ".analytics-chart-body" in css.text
    )
    assert ".ozon-economics-grid" in css.text
    assert ".ozon-economics-grid .metric" in css.text
    assert ".ozon-economics-grid .metric strong" in css.text
    assert ".ozon-analytics-card-grid" in css.text
    assert ".ozon-analytics-card-grid .metric" in css.text
    assert ".ozon-analytics-card-grid .metric strong" in css.text
    assert ".ozon-issue-panel" in css.text
    assert ".ozon-issue-list" in css.text
    assert ".ozon-technical-details" in css.text
    assert ".ozon-pnl-grid" in css.text
    assert ".ozon-pnl-table" in css.text
    assert ".ozon-buyout-table" in css.text
    assert ".ozon-preview-table" in css.text
    assert ".mapping-upload-form" not in css.text
    assert "reason-columns" not in css.text
    assert ".file-picker" in css.text
    assert "overflow-wrap: anywhere" in css.text


def test_client_logistics_deep_link_waits_for_reports_and_skips_staff_draft_api(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)

    app_js = client.get("/static/app.js")

    assert app_js.status_code == 200
    assert "reportsLoaded: false" in app_js.text
    assert "state.reportsLoaded = true;\n  syncLogisticsEntryPoint();" in app_js.text
    assert "state.reportsLoaded\n        && state.reports.length === 0" in app_js.text
    draft_loader = app_js.text.split(
        "async function loadClientDraft", 1
    )[1].split("async function loadIntegrations", 1)[0]
    assert draft_loader.index("if (!isStaffUser())") < draft_loader.index(
        "/client-draft"
    )


def test_report_wizard_keeps_published_report_and_new_run_separate(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    cabinet = client.get("/cabinet")
    app_js = client.get("/static/app.js")
    styles = client.get("/static/styles.css")

    assert cabinet.status_code == 200
    assert app_js.status_code == 200
    assert styles.status_code == 200
    assert cabinet.text.index('id="report-wizard-current"') < cabinet.text.index(
        'id="report-wizard-mode"'
    )
    assert cabinet.text.index('id="report-wizard-submit"') < cabinet.text.index(
        'id="report-wizard-check"'
    )
    assert (
        'Boolean(item.isCurrent) && normalize(item.publicationStatus) === "published"'
        in app_js.text
    )
    assert "state.reportWizardRefresh?.newReportRunId" in app_js.text
    assert "refresh.id !== state.reportWizardRefresh.id" in app_js.text
    assert "state.reportWizardRefresh = state.latestSourceRefresh" not in app_js.text
    assert "renderReportWizardStatus(state.latestSourceRefresh" not in app_js.text
    assert (
        "/api/reports/${encodeURIComponent(report.id)}/export.xlsx"
        in app_js.text
    )
    assert (
        "/api/reports/${encodeURIComponent(generatedReportId)}/export.xlsx"
        in app_js.text
    )
    assert 'normalize(refresh?.status) === "needs_review"' in app_js.text
    assert "и пока не опубликован как текущий" in app_js.text
    assert "Служебная диагностика готова для скачивания" in app_js.text
    mobile_actions = styles.text.rsplit(".report-wizard-actions {", 1)[1].split(
        "}", 1
    )[0]
    assert "flex-direction: column;" in mobile_actions
    assert "column-reverse" not in mobile_actions


def test_primary_kpi_contract_contains_ten_ordered_after_tax_cards(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    app_js = client.get("/static/app.js")
    css = client.get("/static/styles.css")

    assert app_js.status_code == 200
    render_kpis = app_js.text.split("function renderKpis", 1)[1].split(
        "function lostSalesCoveragePeriodText",
        1,
    )[0]
    ordered_labels = [
        "Выручка WB без НДС",
        "Себестоимость 1С",
        "Расходы WB",
        "Управленческая прибыль WB",
        "Маржинальность WB",
        "Прибыль до налогов",
        "Маржинальность до налогов",
        "Итого к перечислению",
        "Продажи WB",
        "Возвратность",
    ]
    positions = [render_kpis.index(f'"{label}"') for label in ordered_labels]

    assert positions == sorted(positions)
    assert render_kpis.count('"Прибыль до налогов"') == 1
    assert render_kpis.count('"Маржинальность до налогов"') == 1
    assert '"Прибыль после налогов"' not in render_kpis
    assert '"Рентабельность после налогов"' not in render_kpis
    assert "marginAfterTax" in render_kpis
    assert "Налоговый профиль не применён" in render_kpis
    assert "Налоговый мост требует сверки" in render_kpis
    assert "По юнит-экономике · НДФЛ ИП не включён" in render_kpis
    assert '"Нулевая выручка"' in render_kpis

    assert css.status_code == 200
    assert (
        ".money-strip .primary-kpi-grid {\n"
        "  grid-template-columns: repeat(5, minmax(0, 1fr));"
    ) in css.text
    primary_card_rule = css.text.split(
        ".money-strip .primary-kpi-grid .metric {",
        1,
    )[1].split("}", 1)[0]
    primary_label_rule = css.text.split(
        ".money-strip .primary-kpi-grid .metric > span {",
        1,
    )[1].split("}", 1)[0]
    primary_value_rule = css.text.split(
        ".money-strip .primary-kpi-grid .metric > strong {",
        1,
    )[1].split("}", 1)[0]
    tablet_rules = css.text.split(
        "@media (max-width: 1179px) and (min-width: 761px)",
        1,
    )[1].split("@media (max-width: 920px)", 1)[0]
    mobile_rules = css.text.rsplit("@media (max-width: 760px)", 1)[1]

    assert "min-height: 142px" in primary_card_rule
    assert "min-height: 36px" in primary_label_rule
    assert "font-size: 14px" in primary_label_rule
    assert "-webkit-line-clamp: 2" in primary_label_rule
    assert "font-size: clamp(22px, 1.55vw, 26px)" in primary_value_rule
    assert "min-width: 0" in primary_value_rule
    assert "max-width: 100%" in primary_value_rule
    assert "white-space: nowrap" in primary_value_rule
    assert "overflow-wrap: normal" in primary_value_rule
    assert "word-break: keep-all" in primary_value_rule
    assert "font-variant-numeric: tabular-nums" in primary_value_rule
    assert "grid-template-columns: repeat(3, minmax(0, 1fr))" in tablet_rules
    assert "grid-template-columns: repeat(2, minmax(0, 1fr))" in mobile_rules
    assert "gap: 8px" in mobile_rules
    assert ".money-strip {\n    padding-inline: 8px" in mobile_rules
    assert "font-size: 19px" in mobile_rules
    assert 'font-family: "Arial Narrow", Arial, sans-serif' in mobile_rules
    assert "max-width: min(320px, calc(100vw - 24px))" in css.text
    assert ".metric:nth-child(5n + 1)::after" in css.text
    assert ".metric:nth-child(3n + 1)::after" in tablet_rules
    assert ".metric:nth-child(odd)::after" in mobile_rules


def test_frontend_guards_stale_filter_requests(tmp_path: Path) -> None:
    client = make_client(tmp_path)

    app_js = client.get("/static/app.js")
    assert app_js.status_code == 200
    text = app_js.text

    assert "rowsRequestKey(reportId, params)" in text
    assert "isCurrentRowsRequest(context, reportId, requestKey)" in text
    assert "renderRowsLoadingState()" in text
    assert "renderRowsErrorState()" in text
    assert "onecReconciliationRequestKey(reportId, params)" in text
    assert "isCurrentOnecReconciliationRequest(context, reportId, requestKey)" in text
    assert "renderOnecReconciliationStatus(" in text
    assert "mappingItemsRequestKey(clientId, paramsKey)" in text
    assert "isCurrentMappingItemsRequest(context, requestKey)" in text
    assert "params !== state.ozonDiagnosticsParams" in text
    assert 'els.filterQuery.value = "";' not in text
    assert 'els.filterStatus.value = "";' not in text


def test_frontend_login_and_widget_accessibility_regressions(tmp_path: Path) -> None:
    client = make_client(tmp_path)

    app_js = client.get("/static/app.js")
    assert app_js.status_code == 200
    text = app_js.text
    assert "error?.status === 401" in text
    assert "Сессия временно занята" not in text
    assert "openWidgetOverlay" in text
    assert "closeWidgetOverlay" in text
    assert "trapWidgetFocus" in text
    assert "!els.newClientWidgetOverlay.hidden" in text

    cabinet = client.get("/cabinet")
    assert cabinet.status_code == 200
    assert 'aria-label="Вопрос AI-аналитику"' in cabinet.text
    assert (
        'data-drilldown-preset="review" role="tab" aria-selected="true"' in cabinet.text
    )
    assert (
        'data-drilldown-preset="sources" role="tab" aria-selected="false"'
        in cabinet.text
    )


def test_mapping_file_upload_saves_local_source_and_audits(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "mapping_uploads"
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_mapping_dir": str(mapping_dir)},
    )
    login(client)

    content = "Товар WB\tАртикул WB\tnmId\nПлатье\tART-1\t1001\n".encode()
    response = client.post(
        "/api/reports/report-1/mapping-file",
        files={"file": ("Организация A.csv", content, "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "uploaded"
    assert payload["fileName"] == "Организация_A.txt"
    assert payload["autoRefresh"]["status"] == "disabled"
    saved = mapping_dir / "Организация_A.txt"
    assert saved.read_bytes() == content

    audit = client.get("/api/admin/audit").json()["items"]
    event = next(item for item in audit if item["action"] == "mapping_file_uploaded")
    assert event["entityId"] == "report-1"
    assert event["payload"]["fileName"] == "Организация_A.txt"
    assert "Платье" not in str(event)


def test_mapping_file_upload_auto_refreshes_and_returns_new_report(
    tmp_path: Path,
) -> None:
    mapping_dir = tmp_path / "mapping_uploads"
    fake_service = FakeAutoRefreshService(tmp_path / "reports" / "auto-refresh.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={
            "source_refresh_enabled": True,
            "source_refresh_mapping_dir": str(mapping_dir),
        },
        auto_refresh_service=fake_service,
    )
    login(client)

    response = client.post(
        "/api/reports/report-1/mapping-file",
        files={"file": ("СопоставлениеНоменклатуры.txt", b"a\tb\n", "text/plain")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "uploaded"
    assert payload["autoRefresh"]["status"] == "report_created"
    assert payload["autoRefresh"]["jobType"] == "source_refresh"
    assert payload["autoRefresh"]["sourceReportRunId"] == "report-1"
    assert payload["autoRefresh"]["newReportRunId"] == "report-1-refresh"
    assert (
        "Автоматическая пересборка после загрузки сопоставления WB ↔ 1C"
        in fake_service.last_reason
    )
    assert "a\tb" not in str(payload)
    assert (mapping_dir / "СопоставлениеНоменклатуры.txt").read_bytes() == b"a\tb\n"
    new_summary = client.get("/api/reports/report-1-refresh/summary").json()
    assert new_summary["quality"]["missingCostRows"] == 0


def test_mapping_upload_keeps_file_when_worker_is_unavailable(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "mapping_uploads"

    class _UnavailableAutoRefresh:
        def run(self, *_args, **_kwargs):
            raise AutoRefreshUnavailableError("Не удалось запустить обновление.")

    client = make_client(
        tmp_path,
        settings_overrides={
            "source_refresh_enabled": True,
            "source_refresh_mapping_dir": str(mapping_dir),
        },
        auto_refresh_service=_UnavailableAutoRefresh(),
    )
    login(client)

    response = client.post(
        "/api/reports/report-1/mapping-file",
        files={"file": ("mapping.txt", b"a\tb\n", "text/plain")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "uploaded"
    assert payload["autoRefresh"]["status"] == "failed"
    assert payload["autoRefresh"]["reviewStatus"] == "needs_review"
    assert (mapping_dir / "mapping.txt").read_bytes() == b"a\tb\n"


def test_mapping_file_upload_is_staff_only(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "mapping_uploads"
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_mapping_dir": str(mapping_dir)},
    )
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "mapping-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "mapping-client@example.com", created["temporaryPassword"])

    forbidden = client.post(
        "/api/reports/report-1/mapping-file",
        files={"file": ("mapping.txt", b"a\tb\n", "text/plain")},
    )

    assert forbidden.status_code == 403
    assert not mapping_dir.exists()


def test_client_mapping_file_upload_and_source_refresh_controls(
    tmp_path: Path,
) -> None:
    mapping_dir = tmp_path / "client_mapping_uploads"
    fake_refresh = FakeSourceRefreshService(
        tmp_path / "reports" / "client-full-refresh.xlsx"
    )
    client = make_client(
        tmp_path,
        settings_overrides={
            "source_refresh_mapping_dir": str(mapping_dir),
            "source_refresh_enabled": True,
            "source_refresh_incremental_enabled": True,
            "marketplace_daily_facts_enabled": True,
            "db_first_reports_enabled": True,
        },
    )
    client.app.state.source_refresh_service = fake_refresh
    login(client)

    upload = client.post(
        "/api/clients/shumeyko/mapping-file",
        files={"file": ("Галустов mapping.csv", b"wb\tonec\n", "text/csv")},
    )

    assert upload.status_code == 200
    upload_payload = upload.json()
    assert upload_payload["status"] == "uploaded"
    assert upload_payload["fileName"] == "Галустов_mapping.txt"
    assert (mapping_dir / "Галустов_mapping.txt").read_bytes() == b"wb\tonec\n"
    assert "wb\tonec" not in str(upload_payload)

    dry_run = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "full", "dry_run": True},
    )

    assert dry_run.status_code == 200
    assert dry_run.json()["latest"]["status"] == "dry_run_ready"
    assert fake_refresh.calls[-1]["dry_run"] is True
    assert fake_refresh.calls[-1]["tenant_id"] == "shumeyko"
    assert fake_refresh.calls[-1]["client_id"] == "shumeyko"

    incremental = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "incremental", "dry_run": True},
    )

    assert incremental.status_code == 200
    assert incremental.json()["incrementalEnabled"] is True
    assert fake_refresh.calls[-1]["mode"] == "incremental"
    assert fake_refresh.calls[-1]["dry_run"] is True

    incremental_latest = client.get(
        "/api/clients/shumeyko/source-refresh/latest?mode=incremental"
    )
    assert incremental_latest.status_code == 200
    assert incremental_latest.json()["incrementalEnabled"] is True
    assert incremental_latest.json()["incrementalWindowDays"] == 28

    ozon_only = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "ozon-only", "dry_run": True},
    )

    assert ozon_only.status_code == 200
    assert fake_refresh.calls[-1]["mode"] == "ozon-only"
    assert fake_refresh.calls[-1]["dry_run"] is True

    ozon_latest = client.get(
        "/api/clients/shumeyko/source-refresh/latest?mode=ozon-only"
    )
    assert ozon_latest.status_code == 200
    assert ozon_latest.json()["latestAttempt"]["mode"] == "ozon-only"

    full = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "full", "dry_run": False},
    )

    assert full.status_code == 200
    payload = full.json()["latest"]
    assert payload["status"] == "queued"
    assert payload["newReportRunId"] is None
    assert payload["collections"] == []
    assert fake_refresh.calls[-1]["dry_run"] is False
    assert fake_refresh.calls[-1]["client_id"] == "shumeyko"

    latest = client.get("/api/clients/shumeyko/source-refresh/latest")
    assert latest.status_code == 200
    latest_response = latest.json()
    latest_payload = latest_response["latest"]
    assert latest_payload["id"] == payload["id"]
    assert latest_payload["status"] == "queued"
    assert latest_payload["newReportRunId"] is None
    assert latest_response["activeRun"]["id"] == payload["id"]
    assert latest_response["latestAttempt"]["id"] == payload["id"]
    assert latest_response["latestCompleted"]["status"] == "dry_run_ready"


def test_client_source_refresh_hands_production_run_to_worker(tmp_path: Path) -> None:
    fake_refresh = FakeSourceRefreshService(
        tmp_path / "reports" / "client-worker-refresh.xlsx"
    )
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
    )
    client.app.state.source_refresh_service = fake_refresh
    launched: list[str] = []

    class _Launcher:
        def launch(self, refresh_run_id: str) -> str:
            launched.append(refresh_run_id)
            return f"shumeiko-source-refresh-worker@{refresh_run_id}.service"

    client.app.state.source_refresh_worker_launcher = _Launcher()
    login(client)

    response = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "full", "dry_run": False},
    )

    assert response.status_code == 200
    payload = response.json()["latest"]
    assert payload["status"] == "queued"
    assert payload["workerAssigned"] is True
    assert payload["progress"]["stage"] == "queued"
    assert launched == [payload["id"]]


def test_client_source_refresh_marks_run_failed_when_worker_launch_fails(
    tmp_path: Path,
) -> None:
    fake_refresh = FakeSourceRefreshService(
        tmp_path / "reports" / "client-worker-failure.xlsx"
    )
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
    )
    client.app.state.source_refresh_service = fake_refresh

    class _FailingLauncher:
        def launch(self, _refresh_run_id: str) -> str:
            raise SourceRefreshWorkerLaunchError("worker start failed")

    client.app.state.source_refresh_worker_launcher = _FailingLauncher()
    login(client)

    response = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "full", "dry_run": False},
    )

    assert response.status_code == 503
    with client.app.state.session_factory() as db:
        refresh_run = db.query(SourceRefreshRun).order_by(
            SourceRefreshRun.created_at.desc()
        ).first()
        assert refresh_run is not None
        assert refresh_run.status == "failed"
        assert refresh_run.failure_code == "worker_launch_failed"
        assert refresh_run.finished_at is not None


def test_client_ozon_diagnostics_returns_safe_latest_ozon_only_snapshot(
    tmp_path: Path,
) -> None:
    ozon_product_name = "Ozon, product, name, with, commas, in, title, example"
    mapping_dir = tmp_path / "ozon_mapping"
    mapping_dir.mkdir()
    (mapping_dir / "sopostavlenie_ozon.txt").write_text(
        (
            "Номенклатура Ozon\tНоменклатура\tХарактеристика\tУпаковка\n"
            f"{ozon_product_name}\tТовар Ozon 1C\t\t\n"
        ),
        encoding="utf-8",
    )
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
    )
    login(client)
    with client.app.state.session_factory() as db:
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        company = (
            db.query(repository.ClientCompany)
            .filter_by(client_id="shumeyko", status="active")
            .order_by(repository.ClientCompany.id)
            .first()
        )
        assert company is not None
        company.onec_organization_id = "ORG-1"
        repository.ensure_wb_cabinet(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            display_name="Ozon test",
            cabinet_key="ozon-test",
            provider="ozon_api",
            client_company_id=company.id,
        )
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="ozon-only-test",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 6, 17),
            user=user,
            reason="Ozon diagnostic test",
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="running",
            started_at=repository.security.utcnow(),
            root_dir="data/source_refresh/ozon-only-test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="sku_mapping",
            source_label="WB ↔ 1C mapping",
            required=True,
            status="loaded",
            row_count=5,
            raw_path=str(mapping_dir),
        )
        sales_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_sales_register",
            source_label="AccumulationRegister_Продажи",
            required=True,
            status="loaded",
            row_count=48,
        )
        repository.add_source_snapshot_row(
            db,
            sales_collection,
            row_number=1,
            raw_payload_hash="onec-sales-hash-1",
            source_row_id="sales-1",
            row_payload={
                "Организация_Key": "ORG-1",
                "RecordSet": [
                    {
                        "Period": "2026-05-31T01:00:00",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "2",
                        "Себестоимость": "600",
                    },
                    {
                        "Period": "2026-05-31T01:00:00",
                        "Контрагент_Key": "OZON-CP",
                        "Документ": "OZON-DOC-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "3",
                        "Сумма": "900",
                        "Себестоимость": "900",
                    },
                ]
            },
        )
        commissioner_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_commissioner_reports",
            source_label="Document_ОтчетКомиссионера",
            required=True,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            commissioner_collection,
            row_number=1,
            raw_payload_hash="onec-commissioner-hash-1",
            source_row_id="ozon-commissioner-1",
            row_payload={
                "Организация_Key": "ORG-1",
                "Date": "2026-05-31T01:00:00",
                "Number": "НФНФ-000033",
                "Posted": True,
                "Комментарий": (
                    "ОЗОН Отчет комиссионера № 16 567 305 "
                    "от 01.05.2026 0:00:00 по 31.05.2026 0:00:00"
                ),
                "Контрагент_Key": "OZON-CP",
                "Запасы": [
                    {
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "2",
                        "Всего": "600",
                        "СуммаНДС": "100",
                    },
                    {
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "1",
                        "Всего": "400",
                        "СуммаНДС": "80",
                    },
                ],
                "ЗапасыВозвраты": [
                    {
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "1",
                        "Всего": "100",
                        "СуммаНДС": "18",
                    },
                ],
            },
        )
        expense_invoice_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_expense_invoices",
            source_label="Document_РасходнаяНакладная",
            required=False,
            status="loaded",
            row_count=3,
        )
        repository.add_source_snapshot_row(
            db,
            expense_invoice_collection,
            row_number=1,
            raw_payload_hash="onec-buyout-hash-duplicate",
            source_row_id="НФНФ-000041",
            row_payload={
                "Date": "2026-05-10T00:00:00",
                "Number": "НФНФ-000041",
                "ОснованиеПечати": "Выкуп",
                "Комментарий": "Отчет о выкупе №4767782 от 15.05.2026",
                "Запасы": [
                    {"Количество": "239", "Всего": "485503.40"},
                ],
            },
        )
        repository.add_source_snapshot_row(
            db,
            expense_invoice_collection,
            row_number=2,
            raw_payload_hash="onec-buyout-hash-1",
            source_row_id="НФНФ-000040",
            row_payload={
                "Date": "2026-05-15T00:00:00",
                "Number": "НФНФ-000040",
                "ОснованиеПечати": "Выкуп",
                "Комментарий": (
                    "ОЗОН Создан на основании отчета о выкупленных товарах "
                    "№ 4767782 от 01.05.2026 0:00:00 по 15.05.2026 0:00:00"
                ),
                "Запасы": [
                    {"Количество": "239", "Всего": "485503.40"},
                ],
            },
        )
        repository.add_source_snapshot_row(
            db,
            expense_invoice_collection,
            row_number=3,
            raw_payload_hash="onec-buyout-hash-2",
            source_row_id="НФНФ-000107",
            row_payload={
                "Date": "2026-05-31T00:00:00",
                "Number": "НФНФ-000107",
                "Комментарий": (
                    "Создан на основании отчета о выкупленных товарах "
                    "№ 4901196 от 16.05.2026 0:00:00 по 31.05.2026 0:00:00"
                ),
                "Запасы": [
                    {"Количество": "217", "Всего": "446196.64"},
                ],
            },
        )
        onec_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_nomenclature",
            source_label="Catalog_Номенклатура",
            required=True,
            status="loaded",
            row_count=2,
        )
        repository.add_source_snapshot_row(
            db,
            onec_collection,
            row_number=1,
            raw_payload_hash="onec-hash-1",
            source_row_id="ITEM-1",
            row_payload={
                "Ref_Key": "ITEM-1",
                "Description": "Товар Ozon 1C",
                "Артикул": "OZ-1",
            },
        )
        repository.add_source_snapshot_row(
            db,
            onec_collection,
            row_number=2,
            raw_payload_hash="onec-hash-2",
            source_row_id="ITEM-2",
            row_payload={
                "Ref_Key": "ITEM-2",
                "Description": "Товар Ozon дубль",
                "Артикул": "OZ-1",
            },
        )
        barcode_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="onec_barcodes",
            source_label="InformationRegister_ШтрихкодыНоменклатуры",
            required=False,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            barcode_collection,
            row_number=1,
            raw_payload_hash="barcode-hash-1",
            source_row_id="barcode-1",
            row_payload={
                "Штрихкод": "12345",
                "Номенклатура_Key": "ITEM-1",
            },
        )
        ozon_products = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="ozon_products_report",
            source_label="Ozon products report",
            required=False,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            ozon_products,
            row_number=1,
            raw_payload_hash="ozon-product-hash-1",
            source_row_id="product-1",
            row_payload={
                "Название товара": ozon_product_name,
                "Артикул продавца": "OZ-1",
                "ID товара": "product-1",
                "SKU": "12345",
                "Штрихкод": "12345",
                "apiKey": "must-not-leak-product",
            },
        )
        ozon_collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="ozon_finance_cash_flow",
            source_label="Ozon financial cash-flow statement",
            required=True,
            status="loaded",
            row_count=2,
        )
        repository.add_source_snapshot_row(
            db,
            ozon_collection,
            row_number=1,
            raw_payload_hash="hash-1",
            source_row_id="op-1",
            row_payload={
                "marketplace": "ozon",
                "operation_id": "op-1",
                "operation_date": "2026-06-01",
                "operation_type": "cash_flow",
                "offer_id": "OZ-1",
                "sku": "12345",
                "price": "1000",
                "details": [
                    {
                        "period": {
                            "begin": "2026-05-01T00:00:00Z",
                            "end": "2026-05-31T00:00:00Z",
                        },
                        "services": {"total": "-100"},
                        "return": {"total": "0"},
                        "others": {"total": "0"},
                    }
                ],
                "apiKey": "must-not-leak",
                "raw": {"clientId": "must-not-leak-too"},
            },
        )
        repository.add_source_snapshot_row(
            db,
            ozon_collection,
            row_number=2,
            raw_payload_hash="hash-2",
            source_row_id="op-2",
            row_payload={
                "operation_id": "op-2",
                "offer_id": "OZ-2",
                "price": "2500",
            },
        )
        ozon_realization = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="ozon_realization",
            source_label="Ozon realization report",
            required=False,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            ozon_realization,
            row_number=1,
            raw_payload_hash="ozon-realization-hash-1",
            source_row_id="realization-1",
            row_payload={
                "offer_id": "OZ-1",
                "sku": "12345",
                "sale_qty": "2",
                "sale_amount": "1000",
                "commission_amount": "50",
                "services_amount": "10",
                "logistics_amount": "20",
                "storage_amount": "5",
                "other_amount": "15",
            },
        )
        ozon_buyout = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="ozon_products_buyout",
            source_label="Ozon products buyout report",
            required=False,
            status="loaded",
            row_count=1,
            payload={
                "marketplace": "ozon",
                "results": [
                    {
                        "sellerAccountId": "OZON_API",
                        "pageIndex": 1,
                        "outputFile": (
                            "OZON_API_ozon_products_buyout_"
                            "2026-05-01_2026-05-31.raw.json"
                        ),
                    }
                ],
            },
        )
        repository.add_source_snapshot_row(
            db,
            ozon_buyout,
            row_number=1,
            raw_payload_hash="ozon-buyout-hash-1",
            source_row_id="ozon_products_buyout:1:1",
            row_payload={
                "seller_account_id": "OZON_API",
                "products": [
                    {"quantity": "239", "amount": "485503.40"},
                    {"quantity": "217", "amount": "446196.64"},
                ],
            },
        )
        ozon_b2b = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="ozon_b2b_sales_json",
            source_label="Ozon B2B sales JSON",
            required=False,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            ozon_b2b,
            row_number=1,
            raw_payload_hash="ozon-b2b-hash-1",
            source_row_id="ozon_b2b_sales_json:1:1",
            row_payload={
                "invoices": [
                    {
                        "number": "B2B-1",
                        "amount": "777777.77",
                        "items": [{"offer_id": "OZ-1", "amount": "777777.77"}],
                    }
                ],
            },
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="source_loaded",
            finished_at=repository.security.utcnow(),
        )
        ozon_draft = repository.materialize_ozon_draft_report(
            db,
            refresh_run,
            user=user,
        )
        ozon_draft_id = ozon_draft.id
        dry_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=True,
            snapshot_set_id="ozon-only-dry-run-after-real-snapshot",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 6, 17),
            user=user,
            reason="test dry-run must not mask real diagnostics",
        )
        repository.update_source_refresh_run(
            db,
            dry_run,
            status="needs_review",
            started_at=repository.security.utcnow(),
            finished_at=repository.security.utcnow(),
        )
        db.commit()

    response = client.get("/api/clients/shumeyko/ozon-diagnostics?limit=1")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "needs_review"
    assert payload["latestRun"]["mode"] == "ozon-only"
    assert payload["latestAttempt"]["dryRun"] is True
    assert payload["readiness"] == {
        "ozonFinanceLoaded": True,
        "ozonRealizationLoaded": True,
        "mappingLoaded": True,
        "onecRequiredLoaded": True,
        "reportExpected": False,
    }
    assert payload["sourceSummary"]["ozonFinance"]["rowCount"] == 2
    assert payload["sourceSummary"]["ozonRealization"]["rowCount"] == 1
    assert payload["sourceSummary"]["ozonBuyouts"]["rowCount"] == 2
    assert payload["sourceSummary"]["ozonBuyouts"]["snapshotRows"] == 1
    assert payload["sourceSummary"]["ozonBuyouts"]["productRows"] == 2
    assert payload["sourceSummary"]["ozonProducts"]["rowCount"] == 1
    assert payload["sourceSummary"]["mapping"]["rowCount"] == 5
    assert payload["sourceSummary"]["onec"]["rowCount"] == 55
    assert payload["ozonBuyouts"]["summary"] == {
        "foundInOzonApi": 2,
        "missingInOzonApi": 0,
        "matchedByReportNumber": 0,
        "matchedByPeriodTotal": 2,
        "ozonApiRows": 1,
        "ozonApiProductRows": 2,
        "ozonApiLoaded": True,
        "ozonApiAmount": 931700.04,
        "ozonApiQuantity": 456.0,
        "ozonApiLoadedAmount": 931700.04,
        "ozonApiLoadedQuantity": 456.0,
        "ozonApiLoadedProductRows": 2,
        "amount": 931700.04,
        "quantity": 456.0,
    }
    buyout_response = client.get("/api/clients/shumeyko/ozon-diagnostics?limit=5")
    assert buyout_response.status_code == 200
    assert buyout_response.json()["ozonBuyouts"]["rows"] == [
        {
            "rowNumber": 2,
            "sourceRowId": "НФНФ-000040",
            "documentNumber": "НФНФ-000040",
            "documentDate": "2026-05-15",
            "basis": "Выкуп",
            "reportNumber": "4767782",
            "periodFrom": "2026-05-01",
            "periodTo": "2026-05-15",
            "quantity": 239.0,
            "amount": 485503.4,
            "foundInOzonApi": True,
            "ozonMatchStatus": "matched_by_period_total",
            "ozonMatchedPeriodFrom": "2026-05-01",
            "ozonMatchedPeriodTo": "2026-05-31",
            "ozonMatchedQuantity": 456.0,
            "ozonMatchedAmount": 931700.04,
        },
        {
            "rowNumber": 3,
            "sourceRowId": "НФНФ-000107",
            "documentNumber": "НФНФ-000107",
            "documentDate": "2026-05-31",
            "basis": "",
            "reportNumber": "4901196",
            "periodFrom": "2026-05-16",
            "periodTo": "2026-05-31",
            "quantity": 217.0,
            "amount": 446196.64,
            "foundInOzonApi": True,
            "ozonMatchStatus": "matched_by_period_total",
            "ozonMatchedPeriodFrom": "2026-05-01",
            "ozonMatchedPeriodTo": "2026-05-31",
            "ozonMatchedQuantity": 456.0,
            "ozonMatchedAmount": 931700.04,
        },
    ]
    assert payload["ozonMapping"]["status"] == "ready"
    assert payload["ozonMapping"]["checkedRows"] == 1
    assert payload["ozonMapping"]["summary"] == {
        "matched": 1,
        "missing": 0,
        "ambiguous": 0,
        "noKey": 0,
        "notChecked": 0,
    }
    assert payload["ozonMapping"]["rows"] == [
        {
            "rowNumber": 1,
            "sourceRowId": "product-1",
            "productName": ozon_product_name,
            "offerId": "OZ-1",
            "productId": "product-1",
            "sku": "12345",
            "barcode": "12345",
            "status": "matched",
            "matchMethod": "uploaded_mapping_name",
            "matchKey": ozon_product_name,
            "onecItemId": "ITEM-1",
            "onecName": "Товар Ozon 1C",
            "onecArticle": "OZ-1",
        }
    ]
    assert payload["finance"]["rowCount"] == 2
    assert payload["finance"]["previewRowCount"] == 1
    assert payload["finance"]["previewLimited"] is True
    assert payload["finance"]["totals"]["price"] == 1000
    assert payload["financeRows"] == [
        {
            "rowNumber": 1,
            "sourceRowId": "op-1",
            "loadedAt": payload["financeRows"][0]["loadedAt"],
            "operationId": "op-1",
            "operationDate": "2026-06-01",
            "operationType": "cash_flow",
            "offerId": "OZ-1",
            "productId": "",
            "sku": "12345",
            "amount": None,
            "price": 1000.0,
            "income": None,
            "expense": None,
            "sourceEndpoint": "",
            "hasMappingKey": True,
        }
    ]
    assert payload["pnl"]["status"] == "provisional"
    assert payload["pnl"]["cashFlowRows"] == 0
    assert payload["pnl"]["realizationRows"] == 1
    assert payload["pnl"]["itemLevelRows"] == 1
    assert payload["pnl"]["costedItemRows"] == 1
    assert payload["pnl"]["totals"]["cashFlowRevenue"] == 0.0
    assert payload["pnl"]["totals"]["revenue"] == 900.0
    assert payload["pnl"]["totals"]["revenueBasis"] == "onec_sales_register"
    assert payload["pnl"]["totals"]["ozonExpenses"] == 100.0
    assert payload["pnl"]["totals"]["expenseBasis"] == "ozon_cash_flow_statement"
    assert payload["pnl"]["totals"]["profitBeforeCogs"] == 800.0
    assert payload["pnl"]["totals"]["onecCogs"] == 600.0
    assert payload["pnl"]["totals"]["profitAfterCogs"] == 200.0
    assert payload["unitRows"]["rowCount"] == 1
    assert payload["unitRows"]["previewRowCount"] == 1
    assert payload["unitRows"]["previewLimited"] is False
    assert payload["unitRows"]["summary"] == {
        "ready": 1,
        "partialSource": 0,
        "missingMapping": 0,
        "ambiguousMapping": 0,
        "missingCost": 0,
        "missing1cCommissioner": 0,
        "missing1cOrganization": 0,
        "buyoutPeriodOnly": 0,
        "partialExpenses": 0,
    }
    unit_row = payload["unitRows"]["rows"][0]
    assert unit_row["rowType"] == "realization_item"
    assert unit_row["periodStart"] == "2026-05-01"
    assert unit_row["periodEnd"] == "2026-05-31"
    assert unit_row["onecRevenue"] == 900.0
    assert unit_row["cogs"] == 600.0
    assert unit_row["ozonExpenses"] == 100.0
    assert unit_row["profit"] == 200.0
    assert unit_row["profitBeforeTax"] == 200.0
    assert unit_row["profitAfterTax"] is None
    assert unit_row["taxProfileSource"] == "missing"
    assert unit_row["taxCompleteness"] == "missing_tax_profile"
    assert unit_row["profitAliasDeprecated"] is True
    assert unit_row["qualityStatus"] == "ready"
    assert unit_row["costQualityStatus"] == "warning"
    assert unit_row["costQualityReason"] == "insufficient_history"
    assert payload["ozonMart"]["basis"] == (
        "staff_only_ozon_unit_economics_mart_v2_monthly"
    )
    assert payload["ozonMart"]["summary"]["ready"] == 1
    assert payload["ozonMart"]["summary"]["buyoutPeriodOnly"] == 0
    assert payload["ozonMart"]["expenseAttribution"]["status"] == "not_applicable"
    assert payload["ozonMart"]["articleDrilldown"][0]["kind"] == "sku_direct"
    assert payload["ozonMart"]["articleDrilldown"][0]["includedInSkuProfit"] is True
    assert payload["ozonMart"]["totals"]["quantity"] == 2.0
    assert payload["ozonMart"]["totals"]["onecRevenue"] == 900.0
    assert payload["ozonMart"]["totals"]["cogs"] == 1500.0
    assert payload["ozonMart"]["totals"]["ozonExpenses"] == 100.0
    assert payload["ozonMart"]["totals"]["profitBeforeTax"] is None
    assert payload["ozonMart"]["totals"]["profitAfterTax"] is None
    assert payload["ozonMart"]["pnlScope"] == (
        "onec_sales_register_including_additional_documents"
    )
    assert payload["ozonMart"]["costQuality"]["status"] == "warning"
    assert payload["ozonMart"]["costQuality"]["quantityCoveragePct"] == 1.0
    assert payload["ozonMart"]["excludedIncompletePeriods"] == [
        {
            "periodStart": "2026-03-01",
            "periodEnd": "2026-03-31",
            "reason": "missing_ozon_realization",
            "reasons": ["missing_ozon_realization"],
        },
        {
            "periodStart": "2026-04-01",
            "periodEnd": "2026-04-30",
            "reason": "missing_ozon_realization",
            "reasons": ["missing_ozon_realization"],
        },
        {
            "periodStart": "2026-06-01",
            "periodEnd": "2026-06-17",
            "reason": "missing_ozon_realization",
            "reasons": ["missing_ozon_realization"],
        },
    ]
    assert payload["ozonMart"]["profitAliasDeprecated"] is True
    assert payload["pnl"]["deprecated"] is True
    assert payload["pnl"]["replacement"] == "ozonMart"
    assert payload["expenseReconciliation"]["status"] == "review"
    assert payload["expenseReconciliation"]["ozonExpenseAmount"] == 100.0
    assert payload["expenseReconciliation"]["onecExpenseAmount"] is None
    reconciliation = payload["reconciliation"]
    assert reconciliation["status"] == "review"
    assert reconciliation["message"] == (
        "Найдено проблем по первичным документам 1C: 1."
    )
    assert reconciliation["sourceType"] == "onec_sales_register"
    assert reconciliation["sourceLabel"] == "1C · регистр продаж"
    assert reconciliation["ozonSourceLabel"] == "Ozon API · реализация и выкупы"
    assert reconciliation["ozonCommissionerAmount"] == 1000.0
    assert reconciliation["commissionerAmount"] == 900.0
    assert reconciliation["commissionerDeltaAmount"] == -100.0
    assert reconciliation["buyoutAmount"] == 931700.04
    assert reconciliation["onecBuyoutAmount"] == 931700.04
    assert reconciliation["buyoutDeltaAmount"] == 0.0
    assert reconciliation["ozonTotalAmount"] == 932700.04
    assert reconciliation["onecSalesRegisterAmount"] == 900.0
    assert reconciliation["deltaAmount"] == -931800.04
    assert reconciliation["buyoutQuantity"] == 456.0
    assert reconciliation["matchedBuyouts"] == 2
    assert reconciliation["missingBuyouts"] == 0
    assert reconciliation["matchedWithoutReportNumber"] == 2
    document_control = reconciliation["documentControl"]
    assert document_control["status"] == "review"
    assert document_control["issueCount"] == 1
    assert document_control["missingPrimaryCount"] == 0
    assert document_control["wrongDateCount"] == 0
    assert document_control["notPostedCount"] == 0
    assert document_control["amountMismatchCount"] == 1
    assert document_control["rows"][0]["status"] == "amount_mismatch"
    assert document_control["rows"][0]["documents"] == [
        "НФНФ-000033 · 2026-05-31"
    ]
    assert document_control["rows"][1]["status"] == "matched"
    assert payload["pnl"]["onecOzon"] == {
        "status": "loaded",
        "counterpartyLabel": "ООО Интернет Решения",
        "counterpartyIds": ["OZON-CP"],
        "reportCount": 1,
        "salesLines": 2,
        "returnLines": 1,
        "salesQuantity": 3.0,
        "returnQuantity": 1.0,
        "salesAmount": 1000.0,
        "returnsAmount": 100.0,
        "netSalesAmount": 900.0,
        "vatAmount": 180.0,
        "returnVatAmount": 18.0,
        "documentRows": [
            {
                "documentNumber": "НФНФ-000033",
                "documentDate": "2026-05-31",
                "reportNumber": "16567305",
                "periodFrom": "2026-05-01",
                "periodTo": "2026-05-31",
                "amount": 900.0,
                "posted": True,
                "status": "matched",
            }
        ],
        "salesRegister": {
            "rowCount": 1,
                "documentCount": 1,
                "quantity": 3.0,
                "amount": 900.0,
                "cost": 900.0,
            "deltaVsCommissionerNet": 0.0,
        },
    }
    assert payload["pnl"]["ozonRealizationAmount"] == 1000.0
    assert payload["pnl"]["periods"] == []

    draft_summary = client.get(f"/api/reports/{ozon_draft_id}/summary")
    assert draft_summary.status_code == 200
    assert draft_summary.json()["marketplace"] == "ozon"
    assert draft_summary.json()["meta"]["lineageType"] == "ozon_mart_snapshot"
    assert (
        draft_summary.json()["ozonDiagnostics"]["latestRun"]["id"]
        == refresh_run.id
    )
    pinned_diagnostics = client.get(
        f"/api/reports/{ozon_draft_id}/ozon-diagnostics"
        "?period_start=2026-05-01&period_end=2026-05-31"
    )
    assert pinned_diagnostics.status_code == 200
    assert pinned_diagnostics.json()["latestRun"]["id"] == refresh_run.id
    draft_export = client.get(f"/api/reports/{ozon_draft_id}/export.xlsx")
    assert draft_export.status_code == 200
    assert "ozon_unit_economics" in draft_export.headers["content-disposition"]

    export = client.get("/api/clients/shumeyko/ozon-diagnostics/export.xlsx")
    assert export.status_code == 200
    assert "ozon_unit_economics" in export.headers["content-disposition"]
    workbook = load_workbook(BytesIO(export.content), read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == [
            "Сводная Ozon",
            "Юнит экономика Ozon",
            "Начисления услуг Ozon",
            "Статьи по SKU",
            "Сверка Ozon 1C",
            "Методика",
        ]
        unit_headers = [cell.value for cell in workbook["Юнит экономика Ozon"][1]]
        reconciliation_headers = [cell.value for cell in workbook["Сверка Ozon 1C"][1]]
        reconciliation_values = [cell.value for cell in workbook["Сверка Ozon 1C"][2]]
    finally:
        workbook.close()
    assert "Услуги партнеров / перевыставление" in unit_headers
    assert "Качество себестоимости" in unit_headers
    assert "Референсная стоимость единицы" in unit_headers
    assert "Ozon API" in reconciliation_headers
    assert reconciliation_values

    filtered_response = client.get(
        "/api/clients/shumeyko/ozon-diagnostics"
        "?period_start=2026-06-01&period_end=2026-06-30&limit=10"
    )
    assert filtered_response.status_code == 200
    filtered_payload = filtered_response.json()
    assert filtered_payload["pnl"]["cashFlowRows"] == 0
    assert filtered_payload["pnl"]["periodFilter"] == {
        "periodStart": "2026-06-01",
        "periodEnd": "2026-06-30",
    }
    assert filtered_payload["pnl"]["totals"]["revenue"] == 0.0
    assert filtered_payload["pnl"]["totals"]["profitBeforeCogs"] == 0.0
    assert filtered_payload["pnl"]["onecOzon"]["status"] == "missing"
    assert filtered_payload["pnl"]["status"] == "partial_source"
    assert filtered_payload["unitRows"]["rows"] == []
    assert filtered_payload["ozonMart"]["status"] == "partial_source"
    assert filtered_payload["ozonMart"]["totals"]["profitBeforeTax"] is None
    assert filtered_payload["pnl"]["periods"] == []
    assert payload["issues"]["blockingCount"] == 0
    # Профиль 1С не загружен -> витрина экономически готова, но налоговый KPI
    # требует проверки: вместо all-clear карточки выходит tax-review.
    assert payload["issues"]["reviewCount"] == 3
    issue_codes = {item["code"] for item in payload["issues"]["items"]}
    assert issue_codes == {
        "ozon_buyout_matched_without_report_number",
        "ozon_mart_cost_quality_warning",
        "ozon_mart_tax_profile_missing",
    }
    assert "must-not-leak" not in str(payload)
    assert "raw_payload_hash" not in str(payload)

    created = client.post(
        "/api/admin/users",
        json={"email": "ozon-draft-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(
        client,
        "ozon-draft-client@example.com",
        created["temporaryPassword"],
    )
    assert client.get(f"/api/reports/{ozon_draft_id}/summary").status_code == 404
    client_reports = client.get("/api/clients/shumeyko/reports").json()["items"]
    assert ozon_draft_id not in {item["id"] for item in client_reports}


def test_ozon_mapping_prefers_onec_marketplace_ozon_mapping_before_fallback() -> None:
    indexes = {
        "byName": {},
        "byArticle": {
            "oz-1": [{"id": "ITEM-2", "name": "Fallback item", "article": "OZ-1"}]
        },
        "byCode": {},
        "byBarcode": {},
        "byOzonNameMapping": {},
        "nomenclatureRows": 1,
        "barcodeRows": 0,
    }
    indexes.update(
        repository._ozon_onec_marketplace_mapping_indexes_from_rows(
            [
                SimpleNamespace(
                    source_type="onec_marketplace_ozon_mapping",
                    row_payload={
                        "marketplace": "ozon",
                        "offer_id": "OZ-1",
                        "product_id": "product-1",
                        "sku": "12345",
                        "barcode": "4600000000000",
                        "ozon_name": "Ozon product",
                        "onec_item_id": "ITEM-1",
                        "onec_name": "Товар из 1C Ozon mapping",
                        "onec_article": "OZ-1",
                        "status": "matched",
                    },
                )
            ]
        )
    )

    checked = repository._check_ozon_mapping_candidate(
        {
            "rowNumber": 1,
            "sourceRowId": "realization-1",
            "productName": "Ozon product",
            "offerId": "OZ-1",
            "productId": "product-1",
            "sku": "12345",
            "barcode": "4600000000000",
        },
        indexes,
    )

    assert checked["statusCounter"] == "matched"
    assert checked["row"]["matchMethod"] == "onec_marketplace_ozon_offer"
    assert checked["row"]["onecItemId"] == "ITEM-1"


def test_ozon_mapping_ignores_generic_onec_marketplace_wb_rows() -> None:
    indexes = repository._ozon_onec_marketplace_mapping_indexes_from_rows(
        [
            SimpleNamespace(
                source_type="onec_marketplace_mapping",
                row_payload={
                    "marketplace": "wb",
                    "offer_id": "OZ-1",
                    "onec_item_id": "ITEM-WB",
                    "onec_name": "WB item",
                },
            ),
            SimpleNamespace(
                source_type="onec_marketplace_mapping",
                row_payload={
                    "marketplace": "ozon",
                    "offer_id": "OZ-2",
                    "onec_item_id": "ITEM-OZON",
                    "onec_name": "Ozon item",
                },
            ),
        ]
    )

    first = repository._check_ozon_mapping_candidate(
        {"offerId": "OZ-1"},
        indexes,
    )
    second = repository._check_ozon_mapping_candidate(
        {"offerId": "OZ-2"},
        indexes,
    )

    assert indexes["onecOzonMappingRows"] == 1
    assert first["statusCounter"] == "missing"
    assert second["statusCounter"] == "matched"
    assert second["row"]["onecItemId"] == "ITEM-OZON"


def test_ozon_rows_matching_period_keeps_only_requested_month_pages() -> None:
    collections = [
        SimpleNamespace(
            source_type="ozon_realization",
            payload={
                "results": [
                    {
                        "sellerAccountId": "OZON-1",
                        "pageIndex": 1,
                        "rowCount": 1,
                        "outputFile": "OZON-1_ozon_realization_2026-04.raw.json",
                    },
                    {
                        "sellerAccountId": "OZON-1",
                        "pageIndex": 1,
                        "rowCount": 1,
                        "outputFile": (
                            "OZON-1_ozon_realization_2026-05_page_0001.raw.json"
                        ),
                    },
                ]
            },
        )
    ]
    april = SimpleNamespace(
        source_row_id="OZ-APRIL",
        row_number=1,
        row_payload={"offer_id": "APRIL"},
    )
    may = SimpleNamespace(
        source_row_id="OZ-MAY",
        row_number=2,
        row_payload={"offer_id": "MAY"},
    )
    outside_manifest = SimpleNamespace(
        source_row_id="OZ-OUTSIDE",
        row_number=4,
        row_payload={"offer_id": "OUTSIDE"},
    )

    matched = repository._ozon_rows_matching_period(
        [april, may, outside_manifest],
        collections=collections,
        source_type="ozon_realization",
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )

    assert matched == [april]


def test_client_ozon_diagnostics_filters_ozon_rows_by_cabinet(
    tmp_path: Path,
) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
    )
    login(client)
    with client.app.state.session_factory() as db:
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="ozon-cabinet-filter-test",
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            user=user,
            reason="Ozon cabinet filter test",
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="running",
            started_at=repository.security.utcnow(),
            root_dir="data/source_refresh/ozon-cabinet-filter-test",
        )
        ozon_realization = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="ozon_realization",
            source_label="Ozon realization report",
            required=False,
            status="loaded",
            row_count=2,
            payload={
                "marketplace": "ozon",
                "results": [
                    {
                        "sellerAccountId": "OZON-1",
                        "wbCabinetId": "ozon-cabinet-1",
                        "pageIndex": 1,
                        "rowCount": 1,
                        "outputFile": "OZON-1_ozon_realization_2026-05.raw.json",
                    },
                    {
                        "sellerAccountId": "OZON-2",
                        "wbCabinetId": "ozon-cabinet-2",
                        "pageIndex": 2,
                        "rowCount": 1,
                        "outputFile": "OZON-2_ozon_realization_2026-05.raw.json",
                    },
                ],
            },
        )
        repository.add_source_snapshot_row(
            db,
            ozon_realization,
            row_number=1,
            raw_payload_hash="ozon-realization-cabinet-1",
            source_row_id="ozon_realization:1:1",
            wb_cabinet_id="ozon-cabinet-1",
            row_payload={
                "seller_account_id": "OZON-1",
                "offer_id": "OZ-1",
                "sku": "111",
                "sale_qty": "2",
                "sale_amount": "1000",
            },
        )
        repository.add_source_snapshot_row(
            db,
            ozon_realization,
            row_number=2,
            raw_payload_hash="ozon-realization-cabinet-2",
            source_row_id="ozon_realization:2:1",
            wb_cabinet_id="ozon-cabinet-2",
            row_payload={
                "seller_account_id": "OZON-2",
                "offer_id": "OZ-2",
                "sku": "222",
                "sale_qty": "3",
                "sale_amount": "2000",
            },
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="source_loaded",
            finished_at=repository.security.utcnow(),
        )
        db.commit()

    response = client.get(
        "/api/clients/shumeyko/ozon-diagnostics?limit=10&wb_cabinet_id=ozon-cabinet-1"
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["sourceSummary"]["ozonRealization"]["rowCount"] == 1
    assert payload["pnl"]["realizationRows"] == 1
    assert payload["unitRows"]["rowCount"] == 1
    assert payload["unitRows"]["rows"][0]["offerId"] == "OZ-1"
    assert payload["unitRows"]["rows"][0]["sku"] == "111"
    assert not any(
        item["sourceType"] == "ozon_realization" and item["rowCount"] != 1
        for item in payload["collections"]
    )


def test_client_ozon_diagnostics_empty_state_without_ozon_run(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    response = client.get("/api/clients/shumeyko/ozon-diagnostics")

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "not_started"
    assert payload["latestRun"] is None
    assert payload["finance"]["rowCount"] == 0
    assert payload["financeRows"] == []
    assert payload["ozonMapping"]["status"] == "not_started"
    assert payload["ozonMapping"]["rows"] == []
    assert payload["pnl"]["status"] == "not_started"
    assert payload["pnl"]["totals"]["onecCogs"] is None
    assert payload["ozonMart"]["status"] == "not_started"
    assert payload["ozonMart"]["rows"] == []
    assert payload["issues"]["blockingCount"] == 2
    assert payload["issues"]["reviewCount"] == 1
    assert [item["code"] for item in payload["issues"]["items"]] == [
        "ozon_realization_missing",
        "ozon_mapping_source_missing",
        "ozon_onec_missing",
    ]


def test_ozon_diagnostics_keeps_last_calculable_run_when_new_attempt_fails(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)
    with client.app.state.session_factory() as db:
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        good = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="ozon-good",
            period_start=date(2026, 4, 1),
            period_end=date(2026, 4, 30),
            user=user,
            reason="good snapshot",
        )
        repository.update_source_refresh_run(
            db,
            good,
            status="source_loaded",
            started_at=repository.security.utcnow(),
            finished_at=repository.security.utcnow(),
        )
        failed = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="ozon-failed",
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            user=user,
            reason="failed attempt",
        )
        repository.update_source_refresh_run(
            db,
            failed,
            status="failed",
            started_at=repository.security.utcnow(),
            finished_at=repository.security.utcnow(),
        )
        db.commit()

        selected = repository.latest_calculable_ozon_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
        )
        assert selected is not None
        assert selected.id == good.id

    response = client.get("/api/clients/shumeyko/ozon-diagnostics")
    assert response.status_code == 200
    payload = response.json()
    assert payload["latestRun"]["id"] == good.id
    assert payload["latestAttempt"]["id"] == failed.id
    assert payload["latestAttempt"]["status"] == "failed"


def test_onec_sales_cost_index_accepts_split_quantity_and_cost_rows() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "0",
                    },
                    {
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "0",
                        "Себестоимость": "250",
                    },
                ]
            }
        )
    ]

    assert repository._onec_sales_cost_index(rows)["ITEM-1"] == 25


def test_onec_sales_cost_index_nets_mixed_direct_and_split_same_document() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "Recorder": "B18",
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-20",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "-2",
                        "Себестоимость": "0",
                    },
                    {
                        "Period": "2026-04-20",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "0",
                        "Себестоимость": "-400",
                    },
                ],
            }
        )
    ]

    assert repository._onec_sales_cost_index(rows) == {
        "ITEM-1": Decimal("75")
    }


def test_onec_sales_cost_index_does_not_attach_cost_only_other_document() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Документ": "SALE-1",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-30",
                        "Документ": "CLOSING-1",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "0",
                        "Себестоимость": "-400",
                    },
                ]
            }
        )
    ]

    assert repository._onec_sales_cost_index(rows) == {
        "ITEM-1": Decimal("100")
    }


def test_onec_sales_cost_index_nets_returns_within_period() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-20",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "-2",
                        "Себестоимость": "-400",
                    },
                    {
                        "Period": "2026-05-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "5",
                        "Себестоимость": "2500",
                    },
                ]
            }
        )
    ]

    april = repository._onec_sales_cost_index(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
    )

    assert april == {"ITEM-1": 75}


def test_onec_sales_cost_index_does_not_hide_negative_cost_with_abs() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-20",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "2",
                        "Себестоимость": "-400",
                    }
                ]
            }
        )
    ]

    april = repository._onec_sales_cost_index(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
    )

    assert april == {"ITEM-1": Decimal("-200")}


def test_onec_previous_closed_month_costs_use_three_recent_closed_months() -> None:
    cost_rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-03-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-20",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "-2",
                        "Себестоимость": "-400",
                    },
                    {
                        "Period": "2026-05-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "3000",
                    },
                ]
            }
        )
    ]
    commissioner_rows = [
        SimpleNamespace(
            row_payload={
                "Date": f"2026-{month:02d}-28",
                "Комментарий": "ОЗОН Отчет комиссионера",
                "Запасы": [
                    {
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "1",
                        "Всего": "1",
                    }
                ],
            }
        )
        for month in (3, 4, 5)
    ]

    history = repository._onec_previous_closed_month_costs(
        cost_rows,
        commissioner_rows=commissioner_rows,
        before_month=date(2026, 6, 1),
        organization_id="ORG-1",
    )

    assert history == {"ITEM-1": (Decimal("300"), Decimal("75"), Decimal("100"))}


def test_onec_direct_cost_control_preserves_signed_returns_and_scope() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Документ": "OZON-DOC",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-20",
                        "Документ": "OZON-DOC",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Количество": "-2",
                        "Себестоимость": "-400",
                    },
                    {
                        "Period": "2026-04-21",
                        "Документ": "OZON-CLOSING",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Количество": "0",
                        "Себестоимость": "999",
                    },
                    {
                        "Period": "2026-04-21",
                        "Документ": "OTHER-DOC",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OTHER-CP",
                        "Количество": "100",
                        "Себестоимость": "100000",
                    },
                ]
            }
        )
    ]

    control = repository._onec_direct_cost_control(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
        counterparty_ids=("OZON-CP",),
    )

    assert control == {"quantity": Decimal("8"), "cogs": Decimal("600")}


def test_onec_direct_sales_control_returns_register_reconciliation_totals() -> None:
    rows = [
        SimpleNamespace(
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Документ": "OZON-COMMISSIONER",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Количество": "10",
                        "Сумма": "1000",
                        "Себестоимость": "600",
                    },
                    {
                        "Period": "2026-04-12",
                        "Документ": "OZON-BUYOUT",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Количество": "2",
                        "Сумма": "180",
                        "Себестоимость": "80",
                    },
                    {
                        "Period": "2026-04-12",
                        "Документ": "OTHER-DOC",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OTHER-CP",
                        "Количество": "5",
                        "Сумма": "500",
                        "Себестоимость": "300",
                    },
                ]
            }
        )
    ]

    control = repository._onec_direct_sales_control(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
        counterparty_ids=("OZON-CP",),
    )
    totals = repository._ozon_mart_reconciliation_totals_payload(
        {"totals": {"onecRevenue": 1000, "cogs": 600}},
        control,
    )

    assert control == {
        "quantity": Decimal("12"),
        "revenue": Decimal("1180"),
        "cogs": Decimal("680"),
    }
    assert totals == {
        "basis": "onec_sales_register",
        "quantity": 12.0,
        "onecRevenue": 1180.0,
        "cogs": 680.0,
        "revenueStatus": "available",
        "cogsStatus": "available",
        "revenueDeltaVsSku": 180.0,
        "cogsDeltaVsSku": 80.0,
        "profitDeltaVsSku": None,
    }

    mart = {
        "totals": {
            "onecRevenue": 1000,
            "cogs": 600,
            "profit": 200,
            "profitBeforeTax": 200,
        },
        "articleRows": [
            {"articleId": "revenue", "amount": 1000, "effectAmount": 1000},
            {"articleId": "cogs", "amount": 600, "effectAmount": -600},
            {"articleId": "profit", "amount": 200, "effectAmount": 200},
        ],
    }
    repository._apply_direct_onec_totals_to_ozon_mart(mart, control)

    assert mart["pnlScope"] == "onec_sales_register_including_additional_documents"
    assert mart["totals"]["onecRevenue"] == 1180.0
    assert mart["totals"]["cogs"] == 680.0
    assert mart["totals"]["profitBeforeTax"] == 300.0
    assert [item["effectAmount"] for item in mart["articleRows"]] == [
        1180.0,
        -680.0,
        300.0,
    ]


def test_onec_direct_cost_control_pairs_split_cost_for_single_counterparty() -> None:
    rows = [
        SimpleNamespace(
            source_row_id="single-counterparty-row",
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "0",
                    },
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "0",
                        "Себестоимость": "1000",
                    },
                ]
            },
        )
    ]

    control = repository._onec_direct_cost_control(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
        counterparty_ids=("OZON-CP",),
    )

    assert control == {"quantity": Decimal("10"), "cogs": Decimal("1000")}


def test_onec_direct_cost_control_scopes_documentless_rows_by_item_counterparty(
) -> None:
    rows = [
        SimpleNamespace(
            source_row_id="ambiguous-counterparty-row",
            row_payload={
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OTHER-CP",
                        "Количество": "5",
                        "Себестоимость": "500",
                    },
                ]
            },
        )
    ]

    control = repository._onec_direct_cost_control(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
        counterparty_ids=("OZON-CP",),
    )

    assert control == {"quantity": Decimal("10"), "cogs": Decimal("1000")}


def test_onec_direct_cost_control_separates_counterparties_in_same_document(
) -> None:
    rows = [
        SimpleNamespace(
            source_row_id="shared-recorder-row",
            row_payload={
                "Recorder": "DOC-1",
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OTHER-CP",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "5",
                        "Себестоимость": "500",
                    },
                ],
            },
        )
    ]

    control = repository._onec_direct_cost_control(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
        counterparty_ids=("OZON-CP",),
    )

    assert control == {"quantity": Decimal("10"), "cogs": Decimal("1000")}


def test_onec_direct_cost_control_rejects_unlabeled_cost_in_shared_document(
) -> None:
    rows = [
        SimpleNamespace(
            source_row_id="shared-recorder-unlabeled-cost",
            row_payload={
                "Recorder": "DOC-1",
                "RecordSet": [
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OZON-CP",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "10",
                        "Себестоимость": "1000",
                    },
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Контрагент_Key": "OTHER-CP",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "5",
                        "Себестоимость": "500",
                    },
                    {
                        "Period": "2026-04-10",
                        "Организация_Key": "ORG-1",
                        "Номенклатура_Key": "ITEM-1",
                        "Количество": "0",
                        "Себестоимость": "300",
                    },
                ],
            },
        )
    ]

    control = repository._onec_direct_cost_control(
        rows,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        organization_id="ORG-1",
        counterparty_ids=("OZON-CP",),
    )

    assert control == {"quantity": None, "cogs": None}


def test_ozon_mart_mapping_resolver_prefers_project_current_mapping() -> None:
    project_mapping = {
        "status": "matched",
        "matchMethod": "mapping_service:imported_mapping_file",
        "offerId": "OZ-1",
        "onecItemId": "ITEM-CONFIRMED",
        "onecName": "Подтвержденный товар",
        "onecArticle": "OZ-1",
    }
    resolver = repository._ozon_mart_mapping_resolver(
        onec_indexes={
            "byArticle": {
                repository._mapping_lookup_key("OZ-1"): [
                    {"id": "ITEM-A", "name": "A", "article": "OZ-1"},
                    {"id": "ITEM-B", "name": "B", "article": "OZ-1"},
                ]
            }
        },
        ozon_mapping={
            "rows": [
                {
                    "offerId": "OZ-1",
                    "status": "ambiguous",
                    "onecItemId": "ITEM-A",
                }
            ]
        },
        project_mapping_preview_index={
            ("offerId", repository._mapping_lookup_key("OZ-1")): project_mapping
        },
    )

    assert resolver({"offerId": "OZ-1"}) == project_mapping


def test_ozon_realization_items_include_nested_item_and_quantity() -> None:
    rows = repository._iter_ozon_realization_items(
        {
            "item": {
                "name": "Ozon product",
                "offer_id": "OZ-1",
                "sku": "12345",
                "barcode": "4600000000000",
            },
            "delivery_commission": {"quantity": 3},
        }
    )

    assert rows == [
        {
            "item": {
                "name": "Ozon product",
                "offer_id": "OZ-1",
                "sku": "12345",
                "barcode": "4600000000000",
            },
            "delivery_commission": {"quantity": 3},
            "name": "Ozon product",
            "offer_id": "OZ-1",
            "sku": "12345",
            "barcode": "4600000000000",
            "quantity": 3,
        }
    ]


def test_client_source_refresh_controls_are_staff_only(tmp_path: Path) -> None:
    mapping_dir = tmp_path / "client_mapping_uploads"
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_mapping_dir": str(mapping_dir)},
    )
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "refresh-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "refresh-client@example.com", created["temporaryPassword"])

    upload = client.post(
        "/api/clients/shumeyko/mapping-file",
        files={"file": ("mapping.txt", b"a\tb\n", "text/plain")},
    )
    latest = client.get("/api/clients/shumeyko/source-refresh/latest")
    ozon_diagnostics = client.get("/api/clients/shumeyko/ozon-diagnostics")
    ozon_export = client.get("/api/clients/shumeyko/ozon-diagnostics/export.xlsx")
    run = client.post(
        "/api/clients/shumeyko/source-refresh",
        json={"mode": "full", "dry_run": True},
    )

    assert upload.status_code == 403
    assert latest.status_code == 403
    assert ozon_diagnostics.status_code == 403
    assert ozon_export.status_code == 403
    assert run.status_code == 403
    assert not mapping_dir.exists()


def test_source_refresh_latest_prefers_active_full_over_blocked_daily_attempt(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)
    with client.app.state.session_factory() as db:
        active = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="full-active",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 7, 10),
            reason="active full refresh",
        )
        repository.update_source_refresh_run(db, active, status="running")
        blocked = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="daily",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="daily-blocked",
            period_start=date(2026, 6, 27),
            period_end=date(2026, 7, 10),
            blocked_by_run=active,
            reason="blocked daily refresh",
            enforce_active_check=False,
        )
        repository.update_source_refresh_run(
            db,
            blocked,
            status="blocked_active_refresh",
            finished_at=repository.security.utcnow(),
        )
        db.commit()

    response = client.get("/api/clients/shumeyko/source-refresh/latest")

    assert response.status_code == 200
    payload = response.json()
    assert payload["latest"]["id"] == active.id
    assert payload["activeRun"]["id"] == active.id
    assert payload["activeRun"]["status"] == "running"
    assert payload["activeRun"]["mode"] == "full"
    assert payload["latestAttempt"]["id"] == blocked.id
    assert payload["latestAttempt"]["blockedByRunId"] == active.id


@pytest.mark.parametrize(
    ("source_state", "message"),
    [
        ("confirmed_empty", "Заявок за доступное окно нет"),
        ("access_denied", "Источник заявок недоступен"),
    ],
)
def test_source_refresh_latest_exposes_safe_return_claims_marker(
    tmp_path: Path,
    source_state: str,
    message: str,
) -> None:
    client = make_client(tmp_path)
    login(client)
    with client.app.state.session_factory() as db:
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id=f"claims-{source_state}",
            period_start=date(2026, 7, 1),
            period_end=date(2026, 7, 23),
            reason="claims marker test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_return_claims",
            source_label="WB buyer return claims",
            required=False,
            status=("loaded" if source_state == "confirmed_empty" else "needs_review"),
            row_count=0,
            raw_path="/protected/raw/path",
            error_message="must stay hidden",
            payload={
                "results": [
                    {
                        "status": source_state,
                        "user_comment": "must stay hidden",
                        "id": "must-stay-hidden",
                    }
                ]
            },
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="completed",
            finished_at=repository.security.utcnow(),
        )
        db.commit()

    response = client.get(
        "/api/clients/shumeyko/source-refresh/latest",
        params={"mode": "full"},
    )

    assert response.status_code == 200
    collection = next(
        item
        for item in response.json()["latest"]["collections"]
        if item["sourceType"] == "wb_return_claims"
    )
    assert collection["sourceState"] == source_state
    assert collection["sourceMessage"] == message
    assert collection["payload"] == {}
    assert collection["rawPath"] == ""
    assert collection["errorMessage"] == ""
    assert "must stay hidden" not in response.text


def test_tax_profile_review_is_not_counted_as_missing_cost(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        row = (
            db.query(repository.ReportUnitRow)
            .filter_by(report_run_id="report-1", status="Нет себестоимости 1С")
            .one()
        )
        row.status = "Себестоимость 1С требует сверки"
        row.status_reason = (
            "Для организации 1С не найден налоговый профиль на период строки"
        )
        row.loss_driver = "Себестоимость 1С требует сверки"
        row.tax_method = "Налоговый профиль не найден"
        row.tax_profile_source = "missing"
        row.tax_completeness = "missing_tax_profile"
        db.commit()
    login(client)

    summary = client.get("/api/reports/report-1/summary").json()

    assert summary["quality"]["missingCostRows"] == 0
    assert "cogs_reconciliation_failed" not in {
        reason["code"] for reason in summary["readiness"]["reviewReasons"]
    }


def test_informational_payout_status_does_not_fail_document_reconciliation(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        rows = db.query(repository.ReportDocumentReconciliationRow).all()
        assert rows
        for row in rows:
            row.status = "OK"
            row.payout_status = "Нужен источник выплаты 1С"
            row.period_status = "полный период"
            row.onec_documents = "Документ 1С № 1"
            for field in repository.DOCUMENT_RECONCILIATION_DELTA_FIELDS:
                setattr(row, field, Decimal("0"))
        db.commit()
    login(client)

    summary = client.get("/api/reports/report-1/summary").json()

    assert summary["quality"]["documentReconciliationIssues"] == 0
    assert "onec_reconciliation_review" not in {
        reason["code"] for reason in summary["readiness"]["reviewReasons"]
    }


@pytest.mark.parametrize(
    "adjustment_type",
    ["Корректировка 1С", "Корректировка себестоимости 1С"],
)
def test_onec_adjustment_is_informational_for_document_readiness(
    tmp_path: Path,
    adjustment_type: str,
) -> None:
    payload = deepcopy(sample_payload())
    payload["documentReconciliation"].append(
        {
            **payload["documentReconciliation"][0],
            "id": f"doc-recon-{adjustment_type}",
            "status": adjustment_type,
            "documentType": adjustment_type,
            "documentReport": adjustment_type,
            "periodStatus": "период 1С",
            "wbReportIds": "",
            "onecDocuments": "Документ корректировки 1С",
            "quantityDelta": None,
            "amountDelta": None,
            "settlementDelta": None,
        }
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    summary = client.get("/api/reports/report-1/summary").json()
    delta_only = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"delta_only": "true"},
    ).json()

    assert summary["quality"]["documentReconciliationIssues"] == 0
    assert "onec_reconciliation_review" not in {
        reason["code"] for reason in summary["readiness"]["reviewReasons"]
    }
    assert delta_only["total"] == 0


def test_out_of_period_unmatched_onec_document_is_hidden_fail_closed(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    base = payload["documentReconciliation"][0]
    payload["documentReconciliation"].extend(
        [
            {
                **base,
                "id": "doc-recon-outside-period",
                "status": "Лишний документ в 1С",
                "periodStatus": "период 1С",
                "salesPeriod": "2026-02-16 - 2026-02-22",
                "salesPeriodStart": "2026-02-16",
                "salesPeriodEnd": "2026-02-22",
                "expectedDocumentDate": "",
                "summaryReportId": "",
                "weeklySalesReportId": "",
                "weeklyBuyoutReportId": "",
                "wbReportIds": "",
                "onecDocuments": "OUTSIDE-PERIOD",
                "onecDocumentDates": "2026-02-22",
                "quantityDelta": None,
                "amountDelta": None,
                "settlementDelta": None,
            },
            {
                **base,
                "id": "doc-recon-inside-period",
                "status": "Лишний документ в 1С",
                "periodStatus": "период 1С",
                "salesPeriod": "2026-05-11 - 2026-05-17",
                "salesPeriodStart": "2026-05-11",
                "salesPeriodEnd": "2026-05-17",
                "expectedDocumentDate": "",
                "summaryReportId": "",
                "weeklySalesReportId": "",
                "weeklyBuyoutReportId": "",
                "wbReportIds": "",
                "onecDocuments": "INSIDE-PERIOD",
                "onecDocumentDates": "2026-05-17",
                "quantityDelta": None,
                "amountDelta": None,
                "settlementDelta": None,
            },
        ]
    )
    client = make_client(tmp_path, payload=payload)
    login(client)

    summary = client.get("/api/reports/report-1/summary").json()
    rows = client.get("/api/reports/report-1/document-reconciliation").json()
    delta_only = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"delta_only": "true"},
    ).json()

    assert summary["quality"]["documentReconciliationRows"] == 2
    assert summary["quality"]["documentReconciliationIssues"] == 1
    assert "OUTSIDE-PERIOD" not in {
        item["onecDocuments"] for item in summary["documentReconciliation"]
    }
    assert rows["total"] == 2
    assert "OUTSIDE-PERIOD" not in {
        item["onecDocuments"] for item in rows["items"]
    }
    assert delta_only["total"] == 1
    assert delta_only["items"][0]["onecDocuments"] == "INSIDE-PERIOD"


def test_buyout_amount_and_return_deltas_are_informational(tmp_path: Path) -> None:
    payload = deepcopy(sample_payload())
    payload["documentReconciliation"] = [
        {
            **payload["documentReconciliation"][0],
            "id": "doc-recon-buyout",
            "status": "Документ найден",
            "documentType": "Уведомление о выкупе",
            "wbSalesQuantity": 42,
            "wbReturnQuantity": 2,
            "wbNetQuantity": 40,
            "onecSalesQuantity": 42,
            "onecReturnQuantity": 0,
            "onecNetQuantity": 42,
            "salesQuantityDelta": 0,
            "returnQuantityDelta": 2,
            "netQuantityDelta": -2,
            "wbQuantity": 42,
            "onecQuantity": 42,
            "quantityDelta": 0,
            "wbAmount": 66003.74,
            "onecAmount": 39464.41,
            "amountDelta": 26539.33,
            "buyoutRetailAmountSum": 66003.74,
            "onecExpenseInvoiceAmount": 39464.41,
            "buyoutRetailDelta": 26539.33,
            "periodStatus": "полный период",
            "onecDocuments": "Расходная накладная 1С № 132",
        }
    ]
    client = make_client(tmp_path, payload=payload)
    login(client)

    body = client.get("/api/reports/report-1/document-reconciliation").json()
    delta_only = client.get(
        "/api/reports/report-1/document-reconciliation",
        params={"delta_only": "true"},
    ).json()

    assert body["kpis"]["issueRows"] == 0
    assert body["kpis"]["amountDelta"] == 0
    assert body["kpis"]["buyoutRetailWb"] == 66003.74
    assert body["kpis"]["buyoutNetOnec"] == 39464.41
    assert delta_only["total"] == 0


def test_buyout_reconciliation_lists_missing_and_quantity_issues(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    base = payload["documentReconciliation"][0]
    payload["documentReconciliation"] = [
        {
            **base,
            "id": "buyout-matched",
            "documentType": "Уведомление о выкупе",
            "status": "Документ найден",
            "wbQuantity": 42,
            "onecQuantity": 42,
            "quantityDelta": 0,
            "wbAmount": 66003.74,
            "buyoutRetailAmountSum": 66003.74,
            "onecAmount": 39464.41,
            "onecExpenseInvoiceAmount": 39464.41,
            "onecDocuments": "Расходная накладная 1С № 132",
        },
        {
            **base,
            "id": "buyout-missing",
            "documentType": "Уведомление о выкупе",
            "status": "Не найден в 1С",
            "wbAmount": 12000,
            "buyoutRetailAmountSum": 12000,
            "onecAmount": None,
            "onecExpenseInvoiceAmount": None,
            "onecDocuments": "",
        },
        {
            **base,
            "id": "buyout-quantity-issue",
            "documentType": "Уведомление о выкупе",
            "status": "Нужна проверка",
            "wbQuantity": 10,
            "onecQuantity": 8,
            "quantityDelta": 2,
            "wbAmount": 10000,
            "buyoutRetailAmountSum": 10000,
            "onecAmount": 7000,
            "onecExpenseInvoiceAmount": 7000,
            "onecDocuments": "Расходная накладная 1С № 133",
        },
    ]
    client = make_client(tmp_path, payload=payload)
    login(client)

    response = client.get(
        "/api/reports/report-1/buyout-reconciliation",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["summary"] == {
        "wbRetailAmount": 88003.74,
        "onecNetAmount": 46464.41,
        "informationalDelta": -41539.33,
        "nonComparableDifference": -41539.33,
        "primaryDocumentAmount": None,
        "primaryDocumentDelta": None,
        "primaryDocumentStatus": "not_loaded",
        "unverifiedPrimaryRows": 3,
        "documentCount": 3,
        "missingOnecRows": 1,
        "quantityIssueRows": 1,
        "matchedRows": 1,
    }
    assert [item["quantityStatus"] for item in body["items"]] == [
        "Нет накладной 1С",
        "Проверить количество",
        "Сверено по количеству",
    ]
    assert all(item["primaryDocumentStatus"] == "not_loaded" for item in body["items"])
    assert all(item["primaryDocumentDelta"] is None for item in body["items"])
    assert "Найдите или загрузите" in body["items"][0]["reason"]


def test_buyout_reconciliation_uses_persisted_wb_primary_document(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    base = payload["documentReconciliation"][0]
    payload["documentReconciliation"] = [
        {
            **base,
            "id": "buyout-primary-verified",
            "documentType": "Уведомление о выкупе",
            "status": "Документ найден",
            "weeklyBuyoutReportId": "685214500",
            "summaryReportId": "685214500",
            "wbSalesQuantity": 62,
            "onecSalesQuantity": 62,
            "wbQuantity": 62,
            "onecQuantity": 62,
            "quantityDelta": 0,
            "wbAmount": 85079.99,
            "buyoutRetailAmountSum": 85079.99,
            "onecAmount": 51532.81,
            "onecExpenseInvoiceAmount": 51532.81,
            "onecDocuments": "Расходная накладная 1С № 54",
        }
    ]
    client = make_client(tmp_path, payload=payload)
    login(client)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="buyout-primary-test",
            period_start=report.period_start,
            period_end=report.period_end,
            reason="buyout primary document test",
        )
        collection = repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_redeem_notifications",
            source_label="WB primary redeem notifications",
            required=False,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            collection,
            row_number=1,
            raw_payload_hash="buyout-primary-685214500",
            source_row_id="685214500",
            row_payload={
                "reportId": "685214500",
                "quantity": "62",
                "purchaseAmount": "51532.81",
                "vatAmount": "9292.80",
            },
        )
        scope = repository.apply_wb_buyout_primary_documents(
            db,
            report,
            refresh_run,
        )
        db.commit()

    assert scope == {
        "sourceRows": 1,
        "reportRows": 1,
        "verifiedRows": 1,
        "notLoadedRows": 0,
    }
    body = client.get("/api/reports/report-1/buyout-reconciliation").json()
    assert body["summary"]["primaryDocumentStatus"] == "verified"
    assert body["summary"]["primaryDocumentAmount"] == 51532.81
    assert body["summary"]["primaryDocumentDelta"] == 0.0
    assert body["summary"]["unverifiedPrimaryRows"] == 0
    assert body["items"][0]["primaryDocumentId"] == "685214500"
    assert body["items"][0]["primaryDocumentQuantity"] == 62.0
    assert body["items"][0]["primaryDocumentDelta"] == 0.0
    assert "совпадают" in body["items"][0]["reason"]

    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        base_refresh = db.query(SourceRefreshRun).filter_by(
            snapshot_set_id="buyout-primary-test"
        ).one()
        overlay = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="incremental",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="buyout-primary-overlay",
            period_start=report.period_start,
            period_end=report.period_end,
            source_window_start=report.period_start,
            source_window_end=report.period_end,
            reason="new overlay without stale primary document",
            enforce_active_check=False,
        )
        repository.add_source_refresh_collection(
            db,
            overlay,
            source_type="wb_redeem_notifications",
            source_label="WB primary redeem notifications",
            required=False,
            status="empty_expected",
            row_count=0,
        )
        daily_without_primary = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="daily",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="daily-without-primary-source",
            period_start=report.period_start,
            period_end=report.period_end,
            reason="daily facts only",
            enforce_active_check=False,
        )
        missing_scope = repository.apply_wb_buyout_primary_documents(
            db,
            report,
            overlay,
            source_runs=[base_refresh, daily_without_primary],
        )
        db.commit()

    assert missing_scope["verifiedRows"] == 0
    assert missing_scope["notLoadedRows"] == 1
    updated = client.get("/api/reports/report-1/buyout-reconciliation").json()
    assert updated["summary"]["primaryDocumentStatus"] == "not_loaded"


def test_login_report_filters_and_export(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    me = client.get("/api/me")
    assert me.status_code == 200
    assert me.json()["tenants"][0]["id"] == "shumeyko"

    reports = client.get("/api/reports").json()["items"]
    assert [item["id"] for item in reports] == ["report-1"]

    summary = client.get("/api/reports/report-1/summary").json()
    assert summary["meta"]["periodStatus"] == "предварительный: июнь неполный"
    assert "unitRows" not in summary
    assert summary["kpis"]["rowCount"] == 2
    assert summary["kpis"]["revenue"] == 119000
    assert summary["kpis"]["cogs"] == 65000
    assert summary["kpis"]["costIssueRows"] == 1
    assert summary["kpis"]["profit"] is None
    assert summary["kpis"]["profitManagement"] == 7000
    assert summary["kpis"]["profitBeforeTax"] == 7000
    assert summary["kpis"]["lossRows"] == 1
    assert summary["kpis"]["lostSalesRevenue"] is None
    assert summary["kpis"]["lostSalesProfit"] is None
    assert summary["kpis"]["lostSalesUnits"] is None
    assert summary["taxContext"]["status"] == "missing"
    assert summary["taxContext"]["calculated"] is False
    assert summary["lostSalesCoverage"]["calculated"] is False
    assert summary["monthly"]
    assert [item["monthStart"] for item in summary["monthly"]] == [
        "2026-04-01",
        "2026-06-01",
    ]
    assert summary["monthly"][-1]["isPartial"] is True
    assert summary["monthly"][-1]["daysElapsed"] == 17
    assert summary["monthly"][-1]["daysInMonth"] == 30
    assert summary["expenses"]
    assert {item["expense"] for item in summary["expenses"]}.isdisjoint(
        {"НДС к уплате", "Налог с выручки/НДФЛ"}
    )
    assert summary["lostSales"] == []
    assert summary["liquidityRows"]
    assert "md1Markup" in summary["liquidityRows"][0]
    assert "md6BeforeTax" in summary["liquidityRows"][0]
    assert summary["quality"]["okRows"] == 1
    assert summary["quality"]["missingCostRows"] == 1
    assert summary["quality"]["documentReconciliationRows"] == 1
    assert summary["quality"]["documentReconciliationIssues"] == 0
    assert summary["readiness"]["status"] == "partial_period"
    assert summary["readiness"]["label"] == "Неполный период"
    assert summary["readiness"]["score"] == 70
    assert summary["options"]["periodStart"] == "2026-04-12"
    assert summary["options"]["periodEnd"] == "2026-06-08"
    assert len(summary["liquidityRows"]) == 2
    assert {row["liquidityStatus"] for row in summary["liquidityRows"]} == {
        "Убыточный: логистика и приемка WB",
        "Нужна проверка данных",
    }
    assert summary["options"]["liquidityStatuses"] == [
        "Нужна проверка данных",
        "Убыточный: логистика и приемка WB",
    ]
    assert {reason["code"] for reason in summary["readiness"]["reviewReasons"]} == {
        "partial_period",
        "cogs_reconciliation_failed",
        "client_draft_missing",
    }

    rows = client.get(
        "/api/reports/report-1/rows",
        params={"preset": "losses", "query": "BAR-LOSS"},
    ).json()
    assert rows["total"] == 1
    assert rows["kpis"]["rowCount"] == 1
    assert rows["kpis"]["revenue"] == 99000
    assert rows["kpis"]["cogs"] == 65000
    assert rows["kpis"]["costIssueRows"] == 0
    assert rows["kpis"]["profit"] is None
    assert rows["kpis"]["profitBeforeTax"] == -9000
    assert rows["kpis"]["profitManagement"] == -9000
    assert rows["kpis"]["lossRows"] == 1
    assert rows["kpis"]["lostSalesRevenue"] is None
    assert rows["items"][0]["product"] == "Убыточный товар"
    assert rows["analytics"]["kpis"]["revenue"] == 99000
    assert rows["analytics"]["kpis"]["profitManagement"] == -9000
    assert {item["expense"] for item in rows["analytics"]["expenses"]}.isdisjoint(
        {"НДС к уплате", "Налог с выручки/НДФЛ"}
    )
    assert rows["analytics"]["monthly"][0]["month"] == "Апрель 2026"
    assert rows["analytics"]["liquidityRows"][0]["product"] == "Убыточный товар"
    assert rows["analytics"]["lostSales"] == []

    filtered_rows = client.get(
        "/api/reports/report-1/rows",
        params={"status_filter": "Нет себестоимости 1С", "limit": 50},
    ).json()
    assert filtered_rows["total"] == 1
    assert filtered_rows["kpis"]["rowCount"] == 1
    assert filtered_rows["kpis"]["revenue"] == 20000
    assert filtered_rows["kpis"]["cogs"] == 0
    assert filtered_rows["kpis"]["costIssueRows"] == 1
    assert filtered_rows["kpis"]["profit"] is None
    assert filtered_rows["kpis"]["lossRows"] == 0
    assert filtered_rows["items"][0]["barcode"] == "BAR-NOCOST"

    return_rows = client.get(
        "/api/reports/report-1/rows",
        params={"preset": "returns", "limit": 50},
    ).json()
    assert return_rows["total"] == 1
    assert return_rows["kpis"]["rowCount"] == 1
    assert return_rows["kpis"]["returns"] == 8
    assert return_rows["items"][0]["barcode"] == "BAR-LOSS"

    document_rows = client.get(
        "/api/reports/report-1/rows",
        params={
            "document_report": (
                "Отчет комиссионера · 06.04.2026-12.04.2026 · закрытие 12.04.2026"
            ),
            "limit": 50,
        },
    ).json()
    assert document_rows["total"] == 1
    assert document_rows["items"][0]["barcode"] == "BAR-LOSS"

    period_rows = client.get(
        "/api/reports/report-1/rows",
        params={"period_start": "2026-06-01", "period_end": "2026-06-30"},
    ).json()
    assert period_rows["total"] == 1
    assert period_rows["kpis"]["rowCount"] == 1
    assert period_rows["kpis"]["revenue"] == 20000
    assert period_rows["kpis"]["profit"] is None
    assert period_rows["items"][0]["barcode"] == "BAR-NOCOST"

    sku = client.get("/api/reports/report-1/sku/BAR-NOCOST").json()
    assert sku["status"] == "Нет себестоимости 1С"

    export = client.get("/api/reports/report-1/export.xlsx")
    assert export.status_code == 200
    assert export.content == b"xlsx"

    freshness = client.get("/api/reports/report-1/freshness")
    assert freshness.status_code == 200
    assert freshness.json()["rowCount"] == 2
    assert freshness.json()["sourceLoads"][0]["status"] == "loaded"
    assert freshness.json()["readiness"]["status"] == "partial_period"

    management = client.get("/api/reports/report-1/management-report")
    assert management.status_code == 200
    assert "Убыточных строк" in management.json()["markdown"]


def test_osno_summary_pnl_uses_without_vat_revenue_and_expenses(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    payload = deepcopy(sample_payload())
    osno_row = deepcopy(payload["unitRows"][0])
    osno_row.update(
        {
            "id": "unit-osno",
            "revenue": 1220,
            "revenueWithoutVat": 1000,
            "vatOutput": 220,
            "vatInput": 22,
            "vatPayable": 198,
            "cost": 500,
            "commission": 122,
            "logistics": 0,
            "storage": 0,
            "acceptance": 0,
            "promotion": 0,
            "penalties": 50,
            "acquiring": 0,
            "usn": 0,
            "profitBeforeTax": 350,
            "profit": 350,
            "pnlVatMode": "without_vat_for_osno",
        }
    )
    payload["unitRows"] = [osno_row]
    payload["lostSales"] = []
    payload["documentReconciliation"] = []

    with session_factory() as db:
        import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-osno",
            publication_status="draft",
            publish=False,
        )
        db.commit()
        report = db.get(repository.ReportRun, "report-osno")
        assert report is not None
        summary = repository.report_summary_payload(db, report)

    assert summary["kpis"]["revenue"] == 1000
    assert summary["kpis"]["revenueWithoutVat"] == 1000
    assert summary["kpis"]["revenueWithVat"] == 1220
    assert summary["kpis"]["pnlWithoutVat"] is True
    assert summary["kpis"]["profit"] is None
    assert summary["kpis"]["margin"] is None
    assert summary["kpis"]["profitBeforeTax"] == 350
    expenses = {item["expense"]: item for item in summary["expenses"]}
    assert expenses["Себестоимость 1С"]["amount"] == 500
    assert expenses["Комиссия WB"]["amount"] == 100
    assert expenses["Комиссия WB"]["share"] == 0.1
    assert expenses["Штрафы/доплаты WB"]["amount"] == 50
    assert expenses["Штрафы/доплаты WB"]["share"] == 0.05
    assert summary["monthly"][0]["revenue"] == 1000
    assert summary["monthly"][0]["profit"] == 350


def test_summary_kpis_exposes_signed_tax_bridge() -> None:
    payload = repository._summary_kpis_payload(
        {
            "revenue": Decimal("62702581.93"),
            "revenue_with_vat": Decimal("62702581.93"),
            "revenue_without_vat": Decimal("59716744.72"),
            "profit": Decimal("12615188.45"),
            "profit_before_tax": Decimal("16228051.66"),
            "pnl_tax_deduction": Decimal("3612863.21"),
            "vat_output": Decimal("2985837.21"),
            "vat_input": Decimal("0"),
            "vat_payable": Decimal("2985837.21"),
            "revenue_tax": Decimal("627026.00"),
            "income_tax": Decimal("0"),
            "income_tax_included_rows": 0,
            "pnl_without_vat_rows": 0,
            "sales": 1,
            "returns": 0,
            "loss_rows": 0,
            "penalty_only_rows": 0,
            "row_count": 1,
        },
        tax_context={"calculated": True},
        lost_sales_coverage={"calculated": False},
    )

    assert payload["revenueWithVat"] == 62702581.93
    assert payload["revenueWithoutVat"] == 59716744.72
    assert payload["revenueTax"] == 627026.0
    assert payload["totalTax"] == 3612863.21
    assert payload["profitBeforeTax"] == 16228051.66
    assert payload["profitAfterTax"] == 12615188.45
    assert payload["taxBridgeCalculated"] is True
    assert payload["marginAfterTax"] == pytest.approx(
        12615188.45 / 62702581.93
    )


def test_summary_kpis_osno_keeps_vat_outside_product_pnl() -> None:
    payload = repository._summary_kpis_payload(
        {
            "revenue": Decimal("900"),
            "revenue_with_vat": Decimal("1100"),
            "revenue_without_vat": Decimal("900"),
            "profit": Decimal("300"),
            "profit_before_tax": Decimal("300"),
            "pnl_tax_deduction": Decimal("0"),
            "vat_output": Decimal("150"),
            "vat_input": Decimal("50"),
            "vat_payable": Decimal("100"),
            "revenue_tax": Decimal("0"),
            "income_tax": Decimal("0"),
            "income_tax_included_rows": 0,
            "pnl_without_vat_rows": 1,
            "sales": 1,
            "returns": 0,
            "loss_rows": 0,
            "penalty_only_rows": 0,
            "row_count": 1,
        },
        tax_context={"calculated": True, "taxSystem": "ОСНО"},
    )

    assert payload["totalTax"] == 100
    assert payload["profitBeforeTax"] == 300
    assert payload["profitAfterTax"] == 300
    assert payload["incomeTaxIncluded"] is False
    assert payload["taxBridgeCalculated"] is True
    assert payload["marginAfterTax"] == pytest.approx(1 / 3)


@pytest.mark.parametrize(
    ("tax_context", "profit_before_tax", "profit", "pnl_tax_deduction"),
    [
        ({"calculated": False}, Decimal("500"), Decimal("400"), Decimal("100")),
        ({"calculated": True}, Decimal("500"), Decimal("450"), Decimal("100")),
    ],
)
def test_summary_kpis_does_not_expose_after_tax_margin_without_valid_bridge(
    tax_context: dict[str, object],
    profit_before_tax: Decimal,
    profit: Decimal,
    pnl_tax_deduction: Decimal,
) -> None:
    payload = repository._summary_kpis_payload(
        {
            "revenue": Decimal("1000"),
            "revenue_with_vat": Decimal("1000"),
            "revenue_without_vat": Decimal("1000"),
            "profit": profit,
            "profit_before_tax": profit_before_tax,
            "pnl_tax_deduction": pnl_tax_deduction,
            "vat_payable": Decimal("100"),
            "revenue_tax": Decimal("0"),
            "income_tax": Decimal("0"),
            "income_tax_included_rows": 0,
            "pnl_without_vat_rows": 0,
            "sales": 1,
            "returns": 0,
            "loss_rows": 0,
            "penalty_only_rows": 0,
            "row_count": 1,
        },
        tax_context=tax_context,
    )

    assert payload["taxBridgeCalculated"] is False
    assert payload["marginAfterTax"] is None
    if tax_context["calculated"] is False:
        assert payload["profitAfterTax"] is None


def test_summary_kpis_zero_revenue_keeps_profit_but_not_margin() -> None:
    payload = repository._summary_kpis_payload(
        {
            "revenue": Decimal("0"),
            "revenue_with_vat": Decimal("0"),
            "revenue_without_vat": Decimal("0"),
            "profit": Decimal("100"),
            "profit_before_tax": Decimal("100"),
            "pnl_tax_deduction": Decimal("0"),
            "vat_payable": Decimal("0"),
            "revenue_tax": Decimal("0"),
            "income_tax": Decimal("0"),
            "income_tax_included_rows": 0,
            "pnl_without_vat_rows": 1,
            "sales": 0,
            "returns": 0,
            "loss_rows": 0,
            "penalty_only_rows": 0,
            "row_count": 1,
        },
        tax_context={"calculated": True, "taxSystem": "ОСНО"},
    )

    assert payload["taxBridgeCalculated"] is True
    assert payload["profitAfterTax"] == 100
    assert payload["marginAfterTax"] is None


def test_osno_legacy_draft_pnl_fallback_uses_tax_method_without_vat(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    payload = deepcopy(sample_payload())
    osno_row = deepcopy(payload["unitRows"][0])
    osno_row.update(
        {
            "id": "unit-osno-legacy-draft",
            "revenue": 1220,
            "revenueWithoutVat": 1000,
            "vatOutput": 220,
            "vatInput": 22,
            "vatPayable": 198,
            "cost": 500,
            "commission": 122,
            "logistics": 0,
            "storage": 0,
            "acceptance": 0,
            "promotion": 0,
            "penalties": 0,
            "acquiring": 0,
            "usn": 0,
            "profitBeforeTax": 598,
            "profit": 400,
            "taxMethod": "ОСНО; НДС 22% внутри цены",
            "pnlVatMode": "",
        }
    )
    payload["unitRows"] = [osno_row]
    payload["lostSales"] = []
    payload["documentReconciliation"] = []

    with session_factory() as db:
        import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-osno-legacy-draft",
            publication_status="draft",
            publish=False,
        )
        db.commit()
        report = db.get(repository.ReportRun, "report-osno-legacy-draft")
        assert report is not None
        summary = repository.report_summary_payload(db, report)

    assert summary["kpis"]["revenue"] == 1000
    assert summary["kpis"]["revenueWithoutVat"] == 1000
    assert summary["kpis"]["revenueWithVat"] == 1220
    assert summary["kpis"]["pnlWithoutVat"] is True
    assert summary["kpis"]["profit"] is None
    assert summary["kpis"]["profitBeforeTax"] == 598
    expenses = {item["expense"]: item for item in summary["expenses"]}
    assert expenses["Себестоимость 1С"]["amount"] == 500
    assert expenses["Комиссия WB"]["amount"] == 100
    assert summary["monthly"][0]["revenue"] == 1000
    assert summary["monthly"][0]["profit"] == 400


def test_report_export_uses_current_published_report_for_stale_link(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)
    current_export = tmp_path / "reports" / "current.xlsx"
    current_export.write_bytes(b"current-xlsx")
    with client.app.state.session_factory() as db:
        payload = deepcopy(sample_payload())
        payload["meta"] = {
            **payload["meta"],
            "period": "01.03.2026 - 08.07.2026",
            "reportPeriod": "01.03.2026 - 08.07.2026",
            "sourceCoverage": "01.03.2026 - 08.07.2026",
            "sourceCoverageEnd": "2026-07-08",
        }
        import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-current",
            source_workbook_path=str(current_export),
        )
        db.commit()

    export = client.get("/api/reports/report-1/export.xlsx")

    assert export.status_code == 200
    assert export.content == b"current-xlsx"


def test_report_rows_period_filter_uses_month_when_week_is_missing(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    payload["unitRows"][0]["week"] = ""
    payload["unitRows"][0]["wbReportDate"] = ""
    client = make_client(tmp_path, payload=payload)
    login(client)

    april_may_rows = client.get(
        "/api/reports/report-1/rows",
        params={"period_start": "2026-04-01", "period_end": "2026-05-31"},
    ).json()
    assert april_may_rows["total"] == 1
    assert april_may_rows["kpis"]["rowCount"] == 1
    assert april_may_rows["kpis"]["revenue"] == 99000
    assert april_may_rows["kpis"]["profit"] is None
    assert april_may_rows["kpis"]["profitBeforeTax"] == -9000
    assert april_may_rows["items"][0]["barcode"] == "BAR-LOSS"

    may_rows = client.get(
        "/api/reports/report-1/rows",
        params={"period_start": "2026-05-01", "period_end": "2026-05-31"},
    ).json()
    assert may_rows["total"] == 0
    assert may_rows["kpis"]["rowCount"] == 0
    assert may_rows["kpis"]["revenue"] == 0


def test_report_rows_assign_cross_month_week_by_accounting_period_date(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    march_week = {
        **payload["unitRows"][0],
        "id": "unit-closes-april",
        "week": "2026-03-30",
        "accountingPeriodDate": "2026-04-05",
        "accountingPeriodSource": "onec_document_date",
        "month": "Март 2026",
        "documentReport": (
            "Отчет комиссионера · 30.03.2026-05.04.2026 · закрытие 31.03.2026"
        ),
        "revenue": 100,
        "penalties": 10,
    }
    april_week = {
        **payload["unitRows"][0],
        "id": "unit-closes-may",
        "week": "2026-04-27",
        "accountingPeriodDate": "2026-04-30",
        "accountingPeriodSource": "onec_document_date",
        "month": "Апрель 2026",
        "documentReport": (
            "Отчет комиссионера · 27.04.2026-03.05.2026 · закрытие 30.04.2026"
        ),
        "revenue": 200,
        "penalties": 20,
    }
    payload["unitRows"] = [march_week, april_week]
    client = make_client(tmp_path, payload=payload)
    login(client)

    april = client.get(
        "/api/reports/report-1/rows",
        params={"period_start": "2026-04-01", "period_end": "2026-04-30"},
    ).json()
    assert april["total"] == 2
    assert april["kpis"]["revenueWithVat"] == 300.0
    assert {row["accountingPeriodDate"] for row in april["items"]} == {
        "2026-04-05",
        "2026-04-30",
    }
    assert all(row["month"] == "Апрель 2026" for row in april["items"])

    may = client.get(
        "/api/reports/report-1/rows",
        params={"month": "Май 2026"},
    ).json()
    assert may["total"] == 0
    assert may["kpis"]["revenueWithVat"] == 0.0

    summary = client.get("/api/reports/report-1/summary").json()
    monthly = {row["month"]: row for row in summary["monthly"]}
    assert monthly["Апрель 2026"]["revenue"] == 300.0
    assert "Май 2026" not in monthly
    assert summary["options"]["months"] == ["Апрель 2026"]


def test_report_summary_is_lightweight_for_large_reports(tmp_path: Path) -> None:
    payload = deepcopy(sample_payload())
    rows = []
    for index in range(1200):
        base = payload["unitRows"][index % 2]
        rows.append(
            {
                **base,
                "id": f"unit-large-{index}",
                "product": f"{base['product']} {index}",
                "nmId": f"{base['nmId']}-{index}",
                "articleWb": "" if index < 2 else f"WB-{index:04d}",
                "barcode": f"{base['barcode']}-{index}",
                "sales": index,
            }
        )
    payload["unitRows"] = rows
    client = make_client(tmp_path, payload=payload)
    login(client)

    response = client.get("/api/reports/report-1/summary")
    assert response.status_code == 200
    summary = response.json()
    assert "unitRows" not in summary
    assert summary["kpis"]["rowCount"] == 1200
    assert summary["quality"]["missingCostRows"] == 600
    assert len(summary["liquidityRows"]) <= 100
    assert len(response.content) < 250_000

    rows_response = client.get(
        "/api/reports/report-1/rows",
        params={"limit": 250},
    )
    assert rows_response.status_code == 200
    rows_payload = rows_response.json()
    assert rows_payload["total"] == 1200
    assert rows_payload["kpis"]["rowCount"] == 1200
    assert len(rows_payload["items"]) == 250

    second_page_response = client.get(
        "/api/reports/report-1/rows",
        params={"limit": 100, "offset": 100},
    )
    assert second_page_response.status_code == 200
    second_page_payload = second_page_response.json()
    assert second_page_payload["total"] == 1200
    assert len(second_page_payload["items"]) == 100
    assert second_page_payload["items"][0]["id"] != rows_payload["items"][0]["id"]

    sorted_first_page = client.get(
        "/api/reports/report-1/rows",
        params={
            "limit": 100,
            "offset": 0,
            "sort_by": "sales",
            "sort_direction": "desc",
        },
    ).json()
    sorted_second_page = client.get(
        "/api/reports/report-1/rows",
        params={
            "limit": 100,
            "offset": 100,
            "sort_by": "sales",
            "sort_direction": "desc",
        },
    ).json()
    assert [row["sales"] for row in sorted_first_page["items"]] == list(
        range(1199, 1099, -1)
    )
    assert [row["sales"] for row in sorted_second_page["items"]] == list(
        range(1099, 999, -1)
    )

    ascending_articles = client.get(
        "/api/reports/report-1/rows",
        params={
            "limit": 2,
            "offset": 1198,
            "sort_by": "articleWb",
            "sort_direction": "asc",
        },
    ).json()
    descending_articles = client.get(
        "/api/reports/report-1/rows",
        params={
            "limit": 2,
            "offset": 1198,
            "sort_by": "articleWb",
            "sort_direction": "desc",
        },
    ).json()
    assert [row["articleWb"] for row in ascending_articles["items"]] == ["", ""]
    assert [row["articleWb"] for row in descending_articles["items"]] == ["", ""]

    assert (
        client.get(
            "/api/reports/report-1/rows",
            params={"sort_by": "notAColumn", "sort_direction": "asc"},
        ).status_code
        == 400
    )
    assert (
        client.get(
            "/api/reports/report-1/rows",
            params={"sort_by": "sales", "sort_direction": "sideways"},
        ).status_code
        == 400
    )

    capped_rows_response = client.get(
        "/api/reports/report-1/rows",
        params={"limit": 5000},
    )
    assert capped_rows_response.status_code == 200
    capped_rows_payload = capped_rows_response.json()
    assert capped_rows_payload["total"] == 1200
    assert len(capped_rows_payload["items"]) == 1000


def test_report_rows_accept_every_whitelisted_sort_column(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    for sort_by in sorted(repository.REPORT_ROW_SORT_KEYS):
        response = client.get(
            "/api/reports/report-1/rows",
            params={
                "limit": 1,
                "sort_by": sort_by,
                "sort_direction": "asc",
            },
        )
        assert response.status_code == 200, (sort_by, response.text)


def test_missing_cost_drilldown_separates_review_and_absent_cost(
    tmp_path: Path,
) -> None:
    payload = deepcopy(sample_payload())
    cost_review_row = {
        **payload["unitRows"][1],
        "id": "unit-cost-review",
        "product": "Товар с временной себестоимостью",
        "nmId": "1004",
        "articleWb": "WB-COST-REVIEW",
        "article1c": "A-COST-REVIEW",
        "barcode": "BAR-COST-REVIEW",
        "cost": 5000,
        "status": "Себестоимость 1С требует сверки",
        "statusReason": (
            "Себестоимость взята из ближайшей доступной недели 1С; "
            "нужна сверка после закрытия месяца"
        ),
        "lossDriver": "Себестоимость 1С требует сверки",
    }
    payload["unitRows"].append(cost_review_row)
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="report-cost-drilldown",
            publication_status="draft",
            publish=False,
        )
        db.flush()
        result = repository.query_report_rows(
            db,
            report,
            preset="missingCost",
            limit=50,
        )

    assert result["total"] == 2
    assert {item["status"] for item in result["items"]} == {
        "Нет себестоимости 1С",
        "Себестоимость 1С требует сверки",
    }
    assert result["costIssueBreakdown"] == {
        "totalRows": 2,
        "requiresReviewRows": 1,
        "absentRows": 1,
        "affectedRevenue": 40000.0,
        "byReason": [
            {
                "reason": "Нет действующей себестоимости 1С",
                "rows": 1,
                "affectedRevenue": 20000.0,
            },
            {
                "reason": (
                    "Себестоимость взята из ближайшей доступной недели 1С; "
                    "нужна сверка после закрытия месяца"
                ),
                "rows": 1,
                "affectedRevenue": 20000.0,
            },
        ],
    }


def test_large_report_summary_and_freshness_bound_unit_row_selects(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)
    engine = client.app.state.session_factory.kw["bind"]
    unit_row_selects: list[str] = []

    def record_unit_row_select(
        _connection,
        _cursor,
        statement,
        _parameters,
        _context,
        _executemany,
    ) -> None:
        normalized = statement.strip().casefold()
        if normalized.startswith("select") and "report_unit_rows" in normalized:
            unit_row_selects.append(statement)

    event.listen(engine, "before_cursor_execute", record_unit_row_select)
    try:
        summary = client.get("/api/reports/report-1/summary")
        summary_select_count = len(unit_row_selects)
        summary_statements = list(unit_row_selects)
        unit_row_selects.clear()
        freshness = client.get("/api/reports/report-1/freshness")
        freshness_select_count = len(unit_row_selects)
    finally:
        event.remove(engine, "before_cursor_execute", record_unit_row_select)

    assert summary.status_code == 200
    assert freshness.status_code == 200
    assert summary_select_count <= 8, "\n---\n".join(summary_statements)
    assert freshness_select_count <= 3


def test_compact_tax_context_and_vat_reconciliation_match_row_payload(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        rows = list(
            db.scalars(
                select(repository.ReportUnitRow).where(
                    repository.ReportUnitRow.report_run_id == report.id
                )
            )
        )
        row_tax_context = repository._tax_context_payload(db, report, rows)
        compact_tax_context = repository._report_tax_context_payload(
            db,
            report,
            row_count=len(rows),
        )
        row_reconciliation = (
            repository._tax_input_reconciliation_payload_from_unit_rows(
                rows,
                tax_context=row_tax_context,
            )
        )
        compact_reconciliation = (
            repository._summary_tax_input_reconciliation_payload(
                db,
                report,
                tax_context=compact_tax_context,
            )
        )

    assert compact_tax_context == row_tax_context
    assert compact_reconciliation == row_reconciliation


def test_top_liquidity_query_matches_full_report_aggregation(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        report = db.get(repository.ReportRun, "report-1")
        optimized = repository._summary_liquidity_rows(db, report)
        full = repository._liquidity_rows_for_conditions(
            db,
            repository.ReportUnitRow.report_run_id == report.id,
        )[:100]

    def without_ids(rows: list[dict]) -> list[dict]:
        return [
            {key: value for key, value in row.items() if key != "id"}
            for row in rows
        ]

    assert without_ids(optimized) == without_ids(full)


def test_multi_client_report_access_requires_explicit_client(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        db.get(repository.Client, "shumeyko").name = "Реальный клиент"
        upsert_user(
            db,
            email="admin@example.com",
            password="secret",
            tenant_id="other",
            role="admin",
        )
        db.commit()

    login(client)

    me = client.get("/api/me")
    assert me.status_code == 200
    me_clients = me.json()["clients"]
    shumeyko_client = next(
        item for item in me_clients if item["clientId"] == "shumeyko"
    )
    assert shumeyko_client["name"] == "Реальный клиент"
    assert shumeyko_client["companies"]
    assert shumeyko_client["cabinets"]

    clients = client.get("/api/clients")
    assert clients.status_code == 200
    assert {item["clientId"] for item in clients.json()["items"]} == {
        "shumeyko",
        "other",
    }

    latest_without_client = client.get("/api/reports/latest/summary")
    assert latest_without_client.status_code == 400

    latest_shumeyko = client.get(
        "/api/reports/latest/summary",
        params={"client_id": "shumeyko"},
    )
    assert latest_shumeyko.status_code == 200
    summary = latest_shumeyko.json()
    assert summary["meta"]["clientId"] == "shumeyko"
    assert "unitRows" not in summary
    assert summary["options"]["cabinets"][0]["id"]
    assert summary["options"]["organizations"][0]["id"]

    latest_other = client.get(
        "/api/reports/latest/summary",
        params={"client_id": "other"},
    )
    assert latest_other.status_code == 200
    assert latest_other.json()["meta"]["clientId"] == "other"

    shumeyko_reports = client.get("/api/clients/shumeyko/reports")
    other_reports = client.get("/api/clients/other/reports")
    assert [item["id"] for item in shumeyko_reports.json()["items"]] == ["report-1"]
    assert [item["id"] for item in other_reports.json()["items"]] == ["other-report"]
    assert client.get("/api/clients/missing/reports").status_code == 404

    cabinet_a = next(
        item for item in summary["options"]["cabinets"] if item["label"] == "Кабинет A"
    )
    company_a = next(
        item
        for item in summary["options"]["organizations"]
        if item["label"] == "Организация A"
    )
    rows = client.get(
        "/api/reports/report-1/rows",
        params={
            "wb_cabinet_id": cabinet_a["id"],
            "client_company_id": company_a["id"],
        },
    )
    assert rows.status_code == 200
    cabinet_a_rows = rows.json()
    assert cabinet_a_rows["total"] == 1
    assert cabinet_a_rows["items"][0]["barcode"] == "BAR-LOSS"
    assert cabinet_a_rows["analytics"]["kpis"]["revenue"] == 99000
    assert cabinet_a_rows["analytics"]["monthly"][0]["month"] == "Апрель 2026"
    assert cabinet_a_rows["analytics"]["lostSales"] == []

    cabinet_b = next(
        item for item in summary["options"]["cabinets"] if item["label"] == "Кабинет B"
    )
    cabinet_b_rows = client.get(
        "/api/reports/report-1/rows",
        params={"wb_cabinet_id": cabinet_b["id"]},
    )
    assert cabinet_b_rows.status_code == 200
    cabinet_b_payload = cabinet_b_rows.json()
    assert cabinet_b_payload["total"] == 1
    assert cabinet_b_payload["items"][0]["barcode"] == "BAR-NOCOST"
    assert cabinet_b_payload["analytics"]["kpis"]["revenue"] == 20000
    assert cabinet_b_payload["analytics"]["monthly"][0]["month"] == (
        "Июнь 2026 (неполный месяц)"
    )
    assert cabinet_b_payload["analytics"]["liquidityRows"][0]["product"] == (
        "Товар без себестоимости"
    )
    assert cabinet_b_payload["analytics"]["lostSales"] == []

    legacy_rows = client.get(
        "/api/reports/report-1/rows",
        params={"wb_cabinet_id": "Кабинет A"},
    )
    assert legacy_rows.status_code == 200
    assert legacy_rows.json()["total"] == 1

    with client.app.state.session_factory() as db:
        upsert_user(
            db,
            email="client-only@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()

    login_as(client, "client-only@example.com", "secret")
    own_clients = client.get("/api/clients")
    assert own_clients.status_code == 200
    assert {item["clientId"] for item in own_clients.json()["items"]} == {"shumeyko"}
    assert client.get("/api/clients/other/reports").status_code == 404
    assert (
        client.get("/api/reports/report-1/document-reconciliation").status_code == 200
    )
    assert (
        client.get("/api/reports/other-report/document-reconciliation").status_code
        == 404
    )
    other_summary = client.get(
        "/api/reports/latest/summary",
        params={"client_id": "other"},
    )
    assert other_summary.status_code == 404


def test_consultant_can_create_client_workspace(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        upsert_user(
            db,
            email="consultant@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="consultant",
        )
        upsert_user(
            db,
            email="client-only@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()

    login_as(client, "consultant@example.com", "secret")
    created = client.post(
        "/api/clients",
        json={
            "name": "Новый клиент",
            "tenant_id": "new-tenant",
            "client_id": "new-client",
            "companies": ["ООО Новый"],
            "cabinets": ["ООО Новый::WB Новый"],
        },
    )

    assert created.status_code == 200
    payload = created.json()["client"]
    assert payload["clientId"] == "new-client"
    assert payload["tenantId"] == "new-tenant"
    assert payload["name"] == "Новый клиент"
    assert payload["role"] == "consultant"
    assert payload["companies"][0]["label"] == "ООО Новый"
    assert payload["cabinets"][0]["label"] == "WB Новый"
    assert payload["cabinets"][0]["clientCompanyId"] == payload["companies"][0]["id"]

    clients = client.get("/api/clients").json()["items"]
    assert "new-client" in {item["clientId"] for item in clients}
    assert client.get("/api/clients/new-client/reports").json()["items"] == []
    assert client.get("/api/clients/new-client/integrations").status_code == 200
    cabinet_created = client.post(
        "/api/clients/new-client/cabinets",
        json={"label": "WB Второй", "organization_name": "ООО Новый"},
    )
    assert cabinet_created.status_code == 200
    created_cabinets = cabinet_created.json()["client"]["cabinets"]
    second_cabinet = next(
        item for item in created_cabinets if item["label"] == "WB Второй"
    )
    assert second_cabinet["clientCompanyId"] == payload["companies"][0]["id"]

    cabinet_updated = client.patch(
        f"/api/clients/new-client/cabinets/{second_cabinet['id']}",
        json={"label": "WB Второй / переименован", "organization_name": "ООО Новый"},
    )
    assert cabinet_updated.status_code == 200
    assert "WB Второй / переименован" in {
        item["label"] for item in cabinet_updated.json()["client"]["cabinets"]
    }

    duplicate = client.post(
        "/api/clients",
        json={"name": "Новый клиент", "tenant_id": "new-tenant"},
    )
    assert duplicate.status_code == 400

    client.post("/api/auth/logout")
    login_as(client, "client-only@example.com", "secret")
    forbidden = client.post("/api/clients", json={"name": "Запрещено"})
    assert forbidden.status_code == 403
    forbidden_cabinet = client.post(
        "/api/clients/shumeyko/cabinets",
        json={"label": "Запрещенный кабинет"},
    )
    assert forbidden_cabinet.status_code == 403


def test_staff_can_link_onec_organization_and_manage_tax_override(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        company = (
            db.query(repository.ClientCompany)
            .filter_by(client_id="shumeyko", status="active")
            .order_by(repository.ClientCompany.id)
            .first()
        )
        assert company is not None
        run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="tax-profile-test",
            period_start=date(2026, 1, 1),
            period_end=date(2026, 7, 10),
            reason="tax profile API test",
        )
        collection = repository.add_source_refresh_collection(
            db,
            run,
            source_type="onec_organizations",
            source_label="Catalog_Организации",
            required=True,
            status="loaded",
            row_count=1,
        )
        repository.add_source_snapshot_row(
            db,
            collection,
            row_number=1,
            raw_payload_hash="org-tax-profile-hash",
            source_row_id="ORG-1",
            row_payload={
                "Ref_Key": "ORG-1",
                "Description": company.display_name,
                "СистемаНалогообложения": "УСН Доходы",
                "СтавкаНДС": "0",
                "РежимНДС": "none",
                "РежимВычетаНДС": "not_applicable",
                "СтавкаНалогаСВыручки": "0.06",
                "ДатаНачала": "2026-01-01",
            },
        )
        repository.update_source_refresh_run(db, run, status="source_loaded")
        company_id = company.id
        db.commit()

    login(client)
    organizations = client.get("/api/clients/shumeyko/onec-organizations")
    assert organizations.status_code == 200
    assert organizations.json()["items"] == [{"id": "ORG-1", "name": "Организация A"}]
    linked = client.patch(
        f"/api/clients/shumeyko/companies/{company_id}/onec-organization",
        json={"onec_organization_id": "ORG-1"},
    )
    assert linked.status_code == 200
    linked_company = next(
        item
        for item in linked.json()["client"]["companies"]
        if item["id"] == company_id
    )
    assert linked_company["onecOrganizationId"] == "ORG-1"
    assert linked_company["taxProfileStatus"] == "missing"

    override = client.post(
        f"/api/clients/shumeyko/companies/{company_id}/tax-profile-overrides",
        json={
            "tax_system": "ОСНО",
            "vat_rate": "22",
            "vat_mode": "included",
            "vat_deduction_mode": "allowed",
            "revenue_tax_rate": "0",
            "income_tax_kind": "ip_ndfl_progressive",
            "valid_from": "2026-01-01",
            "valid_to": "2026-12-31",
            "reason": "Временное подтверждение бухгалтера",
        },
    )
    assert override.status_code == 200
    override_id = override.json()["overrideId"]
    override_company = next(
        item
        for item in override.json()["client"]["companies"]
        if item["id"] == company_id
    )
    assert override_company["taxProfileStatus"] == "override"
    assert override_company["taxProfileSource"] == "manual_override"

    overlapping = client.post(
        f"/api/clients/shumeyko/companies/{company_id}/tax-profile-overrides",
        json={
            "tax_system": "УСН Доходы",
            "vat_rate": "0",
            "vat_mode": "none",
            "vat_deduction_mode": "not_applicable",
            "revenue_tax_rate": "0.06",
            "valid_from": "2026-06-01",
            "reason": "Конфликтующий период",
        },
    )
    assert overlapping.status_code == 400

    with client.app.state.session_factory() as db:
        run = db.get(SourceRefreshRun, run.id)
        assert run is not None
        tax_collection = repository.sync_organization_tax_profiles(db, run)
        assert tax_collection.payload["message"] == (
            "Справочник организаций загружен из 1С, но для части организаций "
            "OData не опубликовала полный налоговый профиль."
        )
        company_diagnostic = next(
            item
            for item in tax_collection.payload["companyDiagnostics"]
            if item["clientCompanyId"] == company_id
        )
        assert company_diagnostic["status"] == "ready"
        db.commit()

    clients_payload = client.get("/api/clients").json()["items"]
    current_company = next(
        company
        for item in clients_payload
        if item["clientId"] == "shumeyko"
        for company in item["companies"]
        if company["id"] == company_id
    )
    assert current_company["taxProfileStatus"] == "ready"
    assert current_company["taxProfileSource"] == "Catalog_Организации"
    assert current_company["taxSystem"] == "УСН Доходы"

    disabled = client.patch(
        f"/api/clients/shumeyko/companies/{company_id}/"
        f"tax-profile-overrides/{override_id}/disable"
    )
    assert disabled.status_code == 200

    with client.app.state.session_factory() as db:
        upsert_user(
            db,
            email="tax-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()
    login_as(client, "tax-client@example.com", "secret")
    forbidden = client.patch(
        f"/api/clients/shumeyko/companies/{company_id}/onec-organization",
        json={"onec_organization_id": "ORG-1"},
    )
    assert forbidden.status_code == 403
    assert client.get("/api/clients/shumeyko/onec-organizations").status_code == 403


def test_tax_profile_resolution_detects_conflict_and_keeps_organizations_separate(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        companies = list(
            db.query(repository.ClientCompany)
            .filter_by(client_id="shumeyko", status="active")
            .order_by(repository.ClientCompany.id)
        )
        company_a = companies[0]
        company_b = repository.ensure_client_company(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            display_name="Организация B",
        )
        assert company_b is not None
        company_a.onec_organization_id = "ORG-A"
        company_b.onec_organization_id = "ORG-B"
        run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="tax-profile-conflict",
            period_start=date(2026, 1, 1),
            period_end=date(2026, 12, 31),
            reason="tax conflict test",
        )
        common = {
            "tenant_id": "shumeyko",
            "client_id": "shumeyko",
            "source_refresh_run_id": run.id,
            "source_snapshot_hash": "hash",
            "methodology_version": "ozon-tax-profile-v2",
            "status": "active",
            "created_at": repository.security.utcnow(),
        }
        db.add_all(
            [
                OrganizationTaxProfile(
                    id="profile-a-1",
                    client_company_id=company_a.id,
                    organization_id="ORG-A",
                    tax_system="УСН Доходы",
                    vat_rate=Decimal("0"),
                    vat_mode="none",
                    vat_deduction_mode="not_applicable",
                    revenue_tax_rate=Decimal("0.06"),
                    income_tax_kind="",
                    valid_from=date(2026, 1, 1),
                    valid_to=None,
                    source="Catalog_Организации",
                    **common,
                ),
                OrganizationTaxProfile(
                    id="profile-a-2",
                    client_company_id=company_a.id,
                    organization_id="ORG-A",
                    tax_system="ОСНО",
                    vat_rate=Decimal("22"),
                    vat_mode="included",
                    vat_deduction_mode="allowed",
                    revenue_tax_rate=Decimal("0"),
                    income_tax_kind="ip_ndfl_progressive",
                    valid_from=date(2026, 1, 1),
                    valid_to=None,
                    source="Document_УведомлениеОСпецрежимахНалогообложения",
                    **common,
                ),
                OrganizationTaxProfile(
                    id="profile-b-1",
                    client_company_id=company_b.id,
                    organization_id="ORG-B",
                    tax_system="УСН Доходы",
                    vat_rate=Decimal("0"),
                    vat_mode="none",
                    vat_deduction_mode="not_applicable",
                    revenue_tax_rate=Decimal("0.06"),
                    income_tax_kind="",
                    valid_from=date(2026, 1, 1),
                    valid_to=date(2026, 12, 31),
                    source="Catalog_Организации",
                    **common,
                ),
            ]
        )
        db.flush()

        profile_a, status_a = repository.resolve_company_tax_profile(
            db,
            company=company_a,
            calculation_date=date(2026, 7, 10),
            refresh_run=run,
        )
        profile_b, status_b = repository.resolve_company_tax_profile(
            db,
            company=company_b,
            calculation_date=date(2026, 7, 10),
            refresh_run=run,
        )
        expired_b, expired_status_b = repository.resolve_company_tax_profile(
            db,
            company=company_b,
            calculation_date=date(2027, 1, 1),
            refresh_run=run,
        )

    assert profile_a is None
    assert status_a["status"] == "conflict"
    assert profile_b is not None
    assert profile_b.organization_id == "ORG-B"
    assert status_b["status"] == "ready"
    assert expired_b is None
    assert expired_status_b["status"] == "missing"


def test_tax_profile_resolution_without_run_prefers_latest_run(
    tmp_path: Path,
) -> None:
    # Старый прогон вывел ОСНО с более поздним valid_from, новый прогон после
    # обновления 1С вывел УСН с более ранним valid_from. Старые строки не
    # деактивируются, поэтому при refresh_run=None (карточка компании, налоговый
    # контекст отчёта) должен выбираться профиль самого свежего прогона (УСН),
    # а не профиль с максимальным valid_from (устаревший ОСНО).
    client = make_client(tmp_path)
    with client.app.state.session_factory() as db:
        company = (
            db.query(repository.ClientCompany)
            .filter_by(client_id="shumeyko", status="active")
            .order_by(repository.ClientCompany.id)
            .first()
        )
        assert company is not None
        company.onec_organization_id = "ORG-A"
        old_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="tax-old-run",
            period_start=date(2026, 1, 1),
            period_end=date(2026, 12, 31),
            reason="old run",
        )
        new_run = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id="shumeyko",
            mode="ozon-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="tax-new-run",
            period_start=date(2026, 1, 1),
            period_end=date(2026, 12, 31),
            reason="new run",
            enforce_active_check=False,
        )
        db.add_all(
            [
                OrganizationTaxProfile(
                    id="profile-old-osno",
                    tenant_id="shumeyko",
                    client_id="shumeyko",
                    source_refresh_run_id=old_run.id,
                    client_company_id=company.id,
                    organization_id="ORG-A",
                    tax_system="ОСНО",
                    vat_rate=Decimal("22"),
                    vat_mode="included",
                    vat_deduction_mode="allowed",
                    revenue_tax_rate=Decimal("0"),
                    income_tax_kind="ip_ndfl_progressive",
                    valid_from=date(2026, 3, 1),
                    valid_to=None,
                    source="Document_УведомлениеОСпецрежимахНалогообложения",
                    source_snapshot_hash="hash-old",
                    methodology_version="ozon-tax-profile-v2",
                    status="active",
                    created_at=datetime(2026, 7, 1, 10, 0, 0),
                ),
                OrganizationTaxProfile(
                    id="profile-new-usn",
                    tenant_id="shumeyko",
                    client_id="shumeyko",
                    source_refresh_run_id=new_run.id,
                    client_company_id=company.id,
                    organization_id="ORG-A",
                    tax_system="УСН Доходы",
                    vat_rate=Decimal("0"),
                    vat_mode="none",
                    vat_deduction_mode="not_applicable",
                    revenue_tax_rate=Decimal("0.06"),
                    income_tax_kind="",
                    valid_from=date(2026, 1, 1),
                    valid_to=None,
                    source="Catalog_Организации",
                    source_snapshot_hash="hash-new",
                    methodology_version="ozon-tax-profile-v2",
                    status="active",
                    created_at=datetime(2026, 7, 2, 10, 0, 0),
                ),
            ]
        )
        db.flush()

        profile, status = repository.resolve_company_tax_profile(
            db,
            company=company,
            calculation_date=date(2026, 7, 10),
            refresh_run=None,
        )

    assert profile is not None
    assert profile.tax_system == "УСН Доходы"
    assert status["status"] == "ready"
    assert status["profileId"] == "profile-new-usn"


def test_wb_cabinet_save_reuses_existing_provider_key_cabinet(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)

    saved_integration = client.post(
        "/api/integrations",
        json={
            "provider": "wb_api",
            "label": "WB дополнительный",
            "cabinet_name": "WB без дубля",
            "secret": "wb-token-no-duplicate",
        },
    )
    assert saved_integration.status_code == 200
    wb_cabinet_id = saved_integration.json()["wbCabinetId"]

    saved_cabinet = client.post(
        "/api/clients/shumeyko/cabinets",
        json={"label": "WB без дубля"},
    )
    assert saved_cabinet.status_code == 200
    cabinets = [
        item
        for item in saved_cabinet.json()["client"]["cabinets"]
        if item["label"] == "WB без дубля"
    ]
    assert len(cabinets) == 1
    assert cabinets[0]["id"] == wb_cabinet_id


def test_report_readiness_ready_after_clean_data_and_final_draft(
    tmp_path: Path,
) -> None:
    payload = ready_payload()
    client = make_client(tmp_path, payload=payload)
    login(client)

    saved = client.put(
        "/api/reports/report-1/client-draft",
        json={
            "content": client_ready_draft_text(),
            "instruction": "Готовый клиентский текст",
        },
    )
    assert saved.status_code == 200
    finalized = client.post(
        "/api/reports/report-1/client-draft/finalize",
        json={"revision": 1},
    )
    assert finalized.status_code == 200

    summary = client.get("/api/reports/report-1/summary").json()
    assert summary["readiness"] == {
        "status": "ready",
        "score": 100,
        "label": "Готов к отправке",
        "blockingReasons": [],
        "reviewReasons": [],
        "nextAction": "Можно отправлять клиенту.",
        "checkedBy": "system",
    }


def test_report_readiness_blocks_empty_report(tmp_path: Path) -> None:
    payload = ready_payload()
    payload["unitRows"] = []
    client = make_client(tmp_path, payload=payload, publish_report=False)
    login(client)

    summary = client.get("/api/reports/report-1/summary").json()
    assert summary["readiness"]["status"] == "failed"
    assert summary["readiness"]["score"] == 0
    assert summary["readiness"]["blockingReasons"][0]["code"] == "no_rows"


def test_financial_publication_gate_keeps_current_report_on_blocker(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    flawed = deepcopy(ready_payload())
    flawed_row = {
        **flawed["unitRows"][0],
        "id": "flawed-osno",
        "taxMethod": "ОСНО; НДС 22% внутри цены",
        "pnlVatMode": "",
        "vatInputCompleteness": "partial",
        "profit": 400,
        "profitBeforeTax": 598,
    }
    flawed["unitRows"] = [flawed_row]

    with session_factory() as db:
        current = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="current-report",
        )
        draft = import_dashboard_payload(
            db,
            flawed,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="blocked-draft",
            publication_status="draft",
            publish=False,
        )

        with pytest.raises(repository.ReportPublicationBlocked) as error:
            repository.publish_report(db, draft)

        assert {item["code"] for item in error.value.blockers} >= {
            "pnl_method_mismatch",
            "profit_semantics_mismatch",
            "vat_input_unconfirmed",
        }
        assert current.is_current is True
        assert current.publication_status == "published"
        assert draft.is_current is False
    assert draft.publication_status == "draft"


def test_staff_can_publish_blocked_report_as_audited_kanban_tasks(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, payload=ready_payload())
    with client.app.state.session_factory() as db:
        draft = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="publish-with-tasks-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="publish-with-tasks-snapshot",
        )
        draft_id = draft.id
        db.commit()

    login(client)
    missing_confirmation = client.post(
        f"/api/reports/{draft_id}/publish-with-tasks",
        json={
            "reason": "Публикуем с прозрачными задачами канбана",
            "confirm_blocking_tasks": False,
        },
    )
    assert missing_confirmation.status_code == 400

    published = client.post(
        f"/api/reports/{draft_id}/publish-with-tasks",
        json={
            "reason": "Публикуем с прозрачными задачами канбана",
            "confirm_blocking_tasks": True,
        },
    )
    assert published.status_code == 200
    payload = published.json()
    assert payload["publicationStatus"] == "published"
    assert payload["isCurrent"] is True
    assert payload["blockingTasks"]
    assert payload["readiness"]["blockingReasons"] == payload["blockingTasks"]

    with client.app.state.session_factory() as db:
        current = db.scalar(
            select(repository.ReportRun).where(
                repository.ReportRun.client_id == "shumeyko",
                repository.ReportRun.is_current.is_(True),
            )
        )
        assert current is not None
        assert current.id == draft_id
        event = db.scalar(
            select(repository.AuditEvent)
            .where(
                repository.AuditEvent.action == "report_published_with_tasks",
                repository.AuditEvent.entity_id == draft_id,
            )
            .order_by(repository.AuditEvent.id.desc())
        )
        assert event is not None
        assert event.payload["reason"] == (
            "Публикуем с прозрачными задачами канбана"
        )
        assert event.payload["blockingTasks"]
        upsert_user(
            db,
            email="publish-tasks-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()

    client.post("/api/auth/logout")
    login_as(client, "publish-tasks-client@example.com", "secret")
    forbidden = client.post(
        f"/api/reports/{draft_id}/publish-with-tasks",
        json={
            "reason": "Клиент не может публиковать",
            "confirm_blocking_tasks": True,
        },
    )
    assert forbidden.status_code == 403


def test_input_vat_policy_is_staff_only_periodic_and_audited(tmp_path: Path) -> None:
    client = make_client(tmp_path, payload=ready_payload())
    with client.app.state.session_factory() as db:
        company = db.scalar(
            select(repository.ClientCompany).where(
                repository.ClientCompany.client_id == "shumeyko"
            )
        )
        assert company is not None
        company.onec_organization_id = "GALUSTOV-1C"
        company_id = company.id
        upsert_user(
            db,
            email="input-vat-client@example.com",
            password="secret",
            tenant_id="shumeyko",
            role="client",
        )
        db.commit()

    login(client)
    url = f"/api/clients/shumeyko/companies/{company_id}/input-vat-policies"
    created = client.post(
        url,
        json={
            "mode": "management_assumption",
            "valid_from": "2026-03-01",
            "valid_to": None,
            "reason": "Импортный НДС для управленческой юнит-экономики",
        },
    )
    assert created.status_code == 200
    item = created.json()["item"]
    assert item["organizationId"] == "GALUSTOV-1C"
    assert item["mode"] == "management_assumption"
    assert item["validFrom"] == "2026-03-01"
    assert item["createdByUserId"]

    overlap = client.post(
        url,
        json={
            "mode": "management_assumption",
            "valid_from": "2026-04-01",
            "reason": "Пересекающийся период",
        },
    )
    assert overlap.status_code == 400

    with client.app.state.session_factory() as db:
        event = db.scalar(
            select(repository.AuditEvent)
            .where(
                repository.AuditEvent.action == "input_vat_policy_created",
                repository.AuditEvent.entity_id == item["id"],
            )
            .order_by(repository.AuditEvent.id.desc())
        )
        assert event is not None
        assert event.payload["reason"] == (
            "Импортный НДС для управленческой юнит-экономики"
        )

    client.post("/api/auth/logout")
    login_as(client, "input-vat-client@example.com", "secret")
    assert client.get(url).status_code == 403


def test_management_input_vat_is_review_task_not_publication_blocker(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="management-input-vat",
            publication_status="draft",
            publish=False,
        )
        row = db.scalar(
            select(repository.ReportUnitRow).where(
                repository.ReportUnitRow.report_run_id == report.id
            )
        )
        assert row is not None
        row.tax_method = "ОСНО; НДС 22% внутри цены"
        row.pnl_vat_mode = "without_vat_for_osno"
        row.tax_profile_source = "Catalog_Организации"
        row.vat_input_completeness = "management_assumption"
        row.input_vat_mode = "management_assumption"
        row.vat_input_confirmed = False
        row.vat_input = Decimal("125")
        row.vat_input_from_import_scenario = Decimal("100")
        row.vat_input_from_wb_scenario = Decimal("25")
        row.profit = row.profit_before_tax
        db.flush()

        readiness = repository.report_readiness_payload(db, report)

    blocker_codes = {
        item["code"] for item in readiness["blockingReasons"]
    }
    review_codes = {item["code"] for item in readiness["reviewReasons"]}
    assert "vat_input_unconfirmed" not in blocker_codes
    assert "vat_input_management_assumption" in review_codes


def test_monthly_reconciliation_difference_is_visible_review_not_blocker(
    tmp_path: Path,
) -> None:
    payload = ready_payload()
    payload["reconciliationMonthly"][0]["status"] = "Расхождение"
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="monthly-reconciliation-review-only",
            publication_status="draft",
            publish=False,
        )
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="monthly-reconciliation-source",
            period_start=date(2026, 4, 1),
            period_end=date(2026, 4, 30),
            reason="monthly reconciliation readiness test",
        )
        repository.update_source_refresh_run(
            db,
            refresh,
            status="completed",
            finished_at=repository.security.utcnow(),
        )
        db.add(
            SourceLoad(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="",
                report_run_id=report.id,
                source_refresh_run_id=refresh.id,
                required=False,
                publication_required=False,
                source_type="onec_income_expense_register",
                source_label="Помесячная сверка 1С",
                status="loaded",
                snapshot_hash="monthly-reconciliation-hash",
                row_count=1,
                loaded_at=repository.security.utcnow(),
            )
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        publication_codes = {
            item["code"] for item in repository.report_publication_blockers(db, report)
        }
        reconciliation = repository.report_full_payload(db, report)[
            "reconciliationMonthly"
        ][0]

    blocker_codes = {item["code"] for item in readiness["blockingReasons"]}
    monthly_review = next(
        item
        for item in readiness["reviewReasons"]
        if item["code"] == "monthly_reconciliation_unresolved"
    )
    assert "monthly_reconciliation_unresolved" not in blocker_codes
    assert "monthly_reconciliation_unresolved" not in publication_codes
    assert monthly_review["count"] == 1
    assert reconciliation["status"] == "Расхождение"
    assert reconciliation["wb_quantity"] == 90.0
    assert reconciliation["onec_quantity"] == 91.0
    assert reconciliation["quantity_delta"] == -1.0
    assert reconciliation["cogs_delta"] == -1000.0
    assert reconciliation["mp_expenses_delta"] == 2000.0


def test_source_backed_missing_cost_is_review_only_not_publication_blocker(
    tmp_path: Path,
) -> None:
    payload = ready_payload()
    payload["lostSales"] = []
    payload["unitRows"][0] = {
        **payload["unitRows"][0],
        "status": "Себестоимость 1С требует сверки",
        "statusReason": "Себестоимость взята из ближайшей доступной недели 1С",
        "lossDriver": "Себестоимость 1С требует сверки",
    }
    payload["unitRows"][1] = {
        **payload["unitRows"][1],
        "status": "Нет себестоимости 1С",
        "statusReason": "Нет действующей себестоимости 1С",
        "lossDriver": "Нет себестоимости 1С",
        "cost": 0,
    }
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="cost-review-only",
            publication_status="draft",
            publish=False,
        )
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="cost-review-only-source",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 6, 30),
            reason="cost review readiness test",
        )
        repository.update_source_refresh_run(
            db,
            refresh,
            status="needs_review",
            finished_at=repository.security.utcnow(),
        )
        db.add(
            SourceLoad(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="",
                report_run_id=report.id,
                source_refresh_run_id=refresh.id,
                required=False,
                publication_required=False,
                source_type="sku_mapping",
                source_label="Сопоставление WB ↔ 1С",
                status="loaded",
                snapshot_hash="cost-review-only-hash",
                row_count=2,
                loaded_at=repository.security.utcnow(),
            )
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)
        publication_codes = {
            item["code"] for item in repository.report_publication_blockers(db, report)
        }

    blocker_codes = {item["code"] for item in readiness["blockingReasons"]}
    cost_review = next(
        item
        for item in readiness["reviewReasons"]
        if item["code"] == "cogs_reconciliation_failed"
    )
    assert "cogs_reconciliation_failed" not in blocker_codes
    assert "cogs_reconciliation_failed" not in publication_codes
    assert cost_review["count"] == 2
    assert cost_review["costRequiresReviewRows"] == 1
    assert cost_review["costAbsentRows"] == 1


def test_wb_finance_lineage_must_cover_first_closing_week() -> None:
    report = SimpleNamespace(
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )
    load = SimpleNamespace(
        source_refresh_run_id="refresh-1",
        source_type="wb_finance_detail",
    )
    incomplete_run = SimpleNamespace(
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )
    complete_run = SimpleNamespace(
        period_start=date(2026, 3, 30),
        period_end=date(2026, 4, 30),
    )

    assert not repository._source_load_covers_report(
        SimpleNamespace(get=lambda _model, _id: incomplete_run),
        load,
        report,
    )
    assert repository._source_load_covers_report(
        SimpleNamespace(get=lambda _model, _id: complete_run),
        load,
        report,
    )
    same_run = SimpleNamespace(
        id="refresh-1",
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
    )
    collection = SimpleNamespace(
        payload={
            "sourceCoverageStart": "2026-03-30",
            "sourceCoverageEnd": "2026-04-30",
        }
    )
    assert repository._source_load_covers_report(
        SimpleNamespace(
            get=lambda _model, _id: same_run,
            scalar=lambda _query: collection,
        ),
        load,
        report,
    )


def test_missing_tax_profile_blocks_without_inheriting_current_osno(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        current = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="osno-current",
        )
        current_row = db.scalar(
            select(repository.ReportUnitRow).where(
                repository.ReportUnitRow.report_run_id == current.id
            )
        )
        assert current_row is not None
        current_row.tax_method = "ОСНО; НДС 22% внутри цены"
        current_row.pnl_vat_mode = "without_vat_for_osno"
        current_row.vat_input_completeness = "confirmed"

        draft_payload = ready_payload()
        draft_payload["unitRows"] = [
            {
                **draft_payload["unitRows"][0],
                "id": "missing-tax-profile-row",
                "taxMethod": "Налоговый профиль не найден",
                "pnlVatMode": "",
                "vatInputCompleteness": "missing",
            }
        ]
        draft = import_dashboard_payload(
            db,
            draft_payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="missing-tax-profile-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="missing-tax-profile-snapshot",
        )
        db.flush()

        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }
        with pytest.raises(repository.ReportPublicationBlocked):
            repository.publish_report(db, draft)
        assert current.is_current is True
        assert current.publication_status == "published"
        assert draft.is_current is False

    assert "tax_profile_unconfirmed" in codes
    assert "pnl_method_mismatch" not in codes
    assert "vat_input_unconfirmed" not in codes


def test_confirmed_usn_draft_does_not_inherit_current_osno_requirements(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        current = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="osno-current",
        )
        for row in db.scalars(
            select(repository.ReportUnitRow).where(
                repository.ReportUnitRow.report_run_id == current.id
            )
        ):
            row.tax_method = "ОСНО; НДС 22% внутри цены"
            row.pnl_vat_mode = "without_vat_for_osno"
            row.vat_input_completeness = "confirmed"

        draft_payload = ready_payload()
        draft_payload["unitRows"] = [
            {
                **row,
                "taxMethod": "УСН Доходы; без НДС; налог с выручки 6%",
                "taxProfileSource": "manual_override",
                "taxCompleteness": "profile_complete",
                "pnlVatMode": "legacy_tax_layer",
                "vatInputCompleteness": "missing",
            }
            for row in draft_payload["unitRows"]
        ]
        draft = import_dashboard_payload(
            db,
            draft_payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="confirmed-usn-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="confirmed-usn-snapshot",
        )
        db.flush()
        company_ids = set(
            db.scalars(
                select(repository.ReportUnitRow.client_company_id)
                .where(repository.ReportUnitRow.report_run_id == draft.id)
                .distinct()
            )
        )
        for index, company_id in enumerate(sorted(company_ids), start=1):
            company = db.get(repository.ClientCompany, company_id)
            assert company is not None
            company.onec_organization_id = f"ORG-USN-{index}"
            db.add(
                OrganizationTaxProfileOverride(
                    id=f"usn-override-{index}",
                    tenant_id=draft.tenant_id,
                    client_id=draft.client_id,
                    client_company_id=company.id,
                    organization_id=company.onec_organization_id,
                    tax_system="УСН Доходы",
                    vat_rate=Decimal("0"),
                    vat_mode="none",
                    vat_deduction_mode="not_applicable",
                    revenue_tax_rate=Decimal("0.06"),
                    income_tax_kind="",
                    valid_from=date(2026, 1, 1),
                    valid_to=None,
                    status="active",
                    reason="Подтверждено для теста",
                    created_by_user_id=None,
                    created_at=repository.security.utcnow(),
                    updated_at=repository.security.utcnow(),
                )
            )
        db.flush()

        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }

    assert "tax_profile_unconfirmed" not in codes
    assert "tax_rate_basis_unconfirmed" not in codes
    assert "pnl_method_mismatch" not in codes
    assert "vat_input_unconfirmed" not in codes


def test_non_osno_report_still_checks_common_financial_blockers() -> None:
    report = SimpleNamespace(
        id="usn-report",
        period_start=date(2026, 3, 1),
        period_end=date(2026, 7, 10),
        source_snapshot_set_id="snapshot",
    )
    source_load = SimpleNamespace(
        source_type="sku_mapping",
        required=True,
        publication_required=False,
        status="needs_review",
        row_count=1,
        source_refresh_run_id="refresh-1",
    )
    db = SimpleNamespace(scalar=lambda _query: 0)

    blockers = repository._financial_integrity_blockers(
        db,
        report,
        source_loads=[source_load],
        stats={
            "tax_profile_issue_rows": 0,
            "osno_rows": 0,
            "missing_cost_affected_revenue": 0,
            "report_type_fallback_rows": 0,
            "report_type_fallback_revenue": 0,
            "sales": 0,
            "storage_and_acceptance": 0,
        },
        tax_context={"profiles": [{"status": "ready"}]},
        document_reconciliation_issue_count=0,
    )
    codes = {item["code"] for item in blockers}

    assert "cogs_reconciliation_failed" not in codes
    assert "source_lineage_failed" not in codes
    assert "pnl_method_mismatch" not in codes
    assert "vat_input_unconfirmed" not in codes


def test_unknown_vat_deduction_keeps_tax_context_unconfirmed(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        payload = ready_payload()
        payload["unitRows"] = [
            {
                **row,
                "taxMethod": "УСН Доходы; без НДС; налог с выручки 6%",
                "taxProfileSource": "manual_override",
                "taxCompleteness": "profile_complete",
            }
            for row in payload["unitRows"]
        ]
        draft = import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="unknown-deduction-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="unknown-deduction-snapshot",
        )
        db.flush()
        rows = list(
            db.scalars(
                select(repository.ReportUnitRow).where(
                    repository.ReportUnitRow.report_run_id == draft.id
                )
            )
        )
        for index, company_id in enumerate(
            sorted({row.client_company_id for row in rows}),
            start=1,
        ):
            company = db.get(repository.ClientCompany, company_id)
            assert company is not None
            company.onec_organization_id = f"ORG-UNKNOWN-{index}"
            db.add(
                OrganizationTaxProfileOverride(
                    id=f"unknown-override-{index}",
                    tenant_id=draft.tenant_id,
                    client_id=draft.client_id,
                    client_company_id=company.id,
                    organization_id=company.onec_organization_id,
                    tax_system="УСН Доходы",
                    vat_rate=Decimal("0"),
                    vat_mode="none",
                    vat_deduction_mode="unknown",
                    revenue_tax_rate=Decimal("0.06"),
                    income_tax_kind="",
                    valid_from=date(2026, 1, 1),
                    valid_to=None,
                    status="active",
                    reason="Неполное подтверждение для теста",
                    created_by_user_id=None,
                    created_at=repository.security.utcnow(),
                    updated_at=repository.security.utcnow(),
                )
            )
        db.flush()

        tax_context = repository._tax_context_payload(db, draft, rows)
        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }

    assert tax_context["status"] == "missing"
    assert tax_context["calculated"] is False
    assert tax_context["vatDeductionMode"] == "unknown"
    assert "tax_profile_unconfirmed" in codes


def test_mixed_organizations_apply_osno_checks_only_to_osno_rows(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        payload = ready_payload()
        osno_row = {
            **payload["unitRows"][0],
            "profit": payload["unitRows"][0]["profitBeforeTax"],
            "taxMethod": "ОСНО; НДС 22% внутри цены",
            "taxProfileSource": "manual_override",
            "taxCompleteness": "vat_confirmed_ndfl_not_allocated",
            "pnlVatMode": "without_vat_for_osno",
            "vatInputCompleteness": "confirmed",
        }
        usn_row = {
            **payload["unitRows"][1],
            "taxMethod": "УСН Доходы; без НДС; налог с выручки 6%",
            "taxProfileSource": "manual_override",
            "taxCompleteness": "profile_complete",
            "pnlVatMode": "legacy_tax_layer",
            "vatInputCompleteness": "missing",
        }
        payload["unitRows"] = [osno_row, usn_row]
        draft = import_dashboard_payload(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="mixed-tax-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="mixed-tax-snapshot",
        )
        db.flush()
        rows = list(
            db.scalars(
                select(repository.ReportUnitRow)
                .where(repository.ReportUnitRow.report_run_id == draft.id)
                .order_by(repository.ReportUnitRow.row_uid)
            )
        )
        assert len(rows) == 2
        for index, row in enumerate(rows, start=1):
            company = db.get(repository.ClientCompany, row.client_company_id)
            assert company is not None
            company.onec_organization_id = f"ORG-MIXED-{index}"
            is_osno = row.row_uid == osno_row["id"]
            db.add(
                OrganizationTaxProfileOverride(
                    id=f"mixed-override-{index}",
                    tenant_id=draft.tenant_id,
                    client_id=draft.client_id,
                    client_company_id=company.id,
                    organization_id=company.onec_organization_id,
                    tax_system="ОСНО" if is_osno else "УСН Доходы",
                    vat_rate=Decimal("22" if is_osno else "0"),
                    vat_mode="included" if is_osno else "none",
                    vat_deduction_mode="allowed" if is_osno else "not_applicable",
                    revenue_tax_rate=Decimal("0" if is_osno else "0.06"),
                    income_tax_kind="ip_ndfl_progressive" if is_osno else "",
                    valid_from=date(2026, 1, 1),
                    valid_to=None,
                    status="active",
                    reason="Смешанный профиль для теста",
                    created_by_user_id=None,
                    created_at=repository.security.utcnow(),
                    updated_at=repository.security.utcnow(),
                )
            )
        db.flush()

        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }
        tax_context = repository._tax_context_payload(db, draft, rows)

    assert tax_context["status"] == "mixed"
    assert tax_context["calculated"] is True
    assert "tax_profile_unconfirmed" not in codes
    assert "vat_input_unconfirmed" not in codes


def test_publication_required_partial_source_blocks_publish(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        draft = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="commissioner-partial-draft",
            publication_status="draft",
            publish=False,
        )
        db.add(
            SourceLoad(
                tenant_id="shumeyko",
                client_id=draft.client_id,
                wb_cabinet_id="",
                report_run_id=draft.id,
                source_refresh_run_id=None,
                required=False,
                publication_required=True,
                source_type="onec_commissioner_reports",
                source_label="Document_ОтчетКомиссионера",
                status="partial_source",
                snapshot_hash="partial-hash",
                row_count=500,
                loaded_at=datetime(2026, 7, 10, 12, 0),
            )
        )
        db.flush()

        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }

    assert "source_load_failed" in codes


def test_mapping_source_review_is_review_only(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="mapping-review-report",
            publication_status="draft",
            publish=False,
        )
        db.add(
            SourceLoad(
                tenant_id="shumeyko",
                client_id=report.client_id,
                wb_cabinet_id="",
                report_run_id=report.id,
                source_refresh_run_id=None,
                required=True,
                publication_required=False,
                source_type="sku_mapping",
                source_label="Сопоставление WB ↔ 1С",
                status="needs_review",
                snapshot_hash="mapping-review-hash",
                row_count=25,
                loaded_at=datetime(2026, 7, 11, 3, 0),
            )
        )
        db.flush()

        readiness = repository.report_readiness_payload(db, report)

    blocking_reasons = {
        item["code"]: item["message"] for item in readiness["blockingReasons"]
    }
    review_reasons = {
        item["code"]: item["message"] for item in readiness["reviewReasons"]
    }
    assert "source_load_review_required" not in blocking_reasons
    assert "source_load_failed" not in blocking_reasons
    assert "source_lineage_failed" not in blocking_reasons
    assert review_reasons["source_load_incomplete"] == (
        "Есть неполные или требующие проверки загрузки источников."
    )


def test_stock_provider_window_is_calculated_without_full_report_coverage(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="provider-window-report",
            publication_status="draft",
            publish=False,
        )
        report.period_start = date(2026, 3, 1)
        report.period_end = date(2026, 7, 10)
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id=report.client_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="provider-window-refresh",
            period_start=report.period_start,
            period_end=report.period_end,
            reason="provider window test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh,
            source_type="wb_stock_history_daily",
            source_label="История остатков WB",
            required=False,
            status="loaded",
            row_count=1,
            payload={
                "calculated": True,
                "providerWindowCalculated": True,
                "fullCoverage": False,
                "calculationPeriodStart": "2026-04-10",
                "calculationPeriodEnd": "2026-07-10",
                "accounts": [
                    {
                        "sellerAccountId": "account-1",
                        "status": "partial_provider_window",
                        "coveredDays": 92,
                        "totalDays": 132,
                        "calculated": True,
                        "providerWindowCalculated": True,
                        "fullCoverage": False,
                    }
                ],
            },
        )
        db.add(
            SourceLoad(
                tenant_id="shumeyko",
                client_id=report.client_id,
                wb_cabinet_id="",
                report_run_id=report.id,
                source_refresh_run_id=refresh.id,
                required=False,
                publication_required=False,
                source_type="wb_stock_history_daily",
                source_label="История остатков WB",
                status="loaded",
                snapshot_hash="provider-window-hash",
                row_count=1,
                loaded_at=repository.security.utcnow(),
            )
        )
        db.flush()

        coverage = repository._lost_sales_coverage_payload(db, report)

    assert coverage["status"] == "partial_provider_window"
    assert coverage["calculated"] is True
    assert coverage["providerWindowCalculated"] is True
    assert coverage["fullCoverage"] is False
    assert coverage["coveredDays"] == 92
    assert coverage["totalDays"] == 132
    assert coverage["calculationPeriodStart"] == "2026-04-10"
    assert coverage["calculationPeriodEnd"] == "2026-07-10"
    assert coverage["extrapolated"] is False
    assert "Рассчитано за доступный период" in coverage["message"]


def test_lost_sales_follow_selected_period_inside_provider_window(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    provider_start = date(2026, 4, 10)
    provider_end = date(2026, 7, 10)
    stock_by_date = {}
    current = provider_start
    while current <= provider_end:
        stock_by_date[current.isoformat()] = (
            "0"
            if current in {date(2026, 4, 10), date(2026, 5, 15)}
            else "1"
        )
        current += timedelta(days=1)
    report_payload = ready_payload()
    report_payload["lostSales"] = []
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            report_payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="filterable-provider-window-report",
            publication_status="draft",
            publish=False,
        )
        report.period_start = date(2026, 3, 1)
        report.period_end = date(2026, 7, 10)
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id=report.client_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="filterable-provider-window-refresh",
            period_start=report.period_start,
            period_end=report.period_end,
            reason="filterable provider window test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh,
            source_type="wb_stock_history_daily",
            source_label="История остатков WB",
            required=False,
            status="loaded",
            row_count=1,
            payload={
                "calculated": True,
                "providerWindowCalculated": True,
                "fullCoverage": False,
                "calculationPeriodStart": provider_start.isoformat(),
                "calculationPeriodEnd": provider_end.isoformat(),
                "calculationContextVersion": "lost-sales-filter-v1",
                "accounts": [
                    {
                        "sellerAccountId": "account-filter-1",
                        "wbCabinetId": "cabinet-filter-1",
                        "cabinet": "Кабинет фильтра",
                        "status": "partial_provider_window",
                        "coveredDays": 92,
                        "totalDays": 132,
                        "calculated": True,
                        "providerWindowCalculated": True,
                        "fullCoverage": False,
                    },
                    {
                        "sellerAccountId": "account-filter-2",
                        "wbCabinetId": "cabinet-filter-2",
                        "cabinet": "Второй кабинет фильтра",
                        "status": "partial_provider_window",
                        "coveredDays": 67,
                        "totalDays": 132,
                        "calculated": True,
                        "providerWindowCalculated": True,
                        "fullCoverage": False,
                        "calculationPeriodStart": "2026-05-05",
                        "calculationPeriodEnd": "2026-07-10",
                    }
                ],
            },
        )
        db.add_all(
            [
                SourceLoad(
                    tenant_id="shumeyko",
                    client_id=report.client_id,
                    wb_cabinet_id="",
                    report_run_id=report.id,
                    source_refresh_run_id=refresh.id,
                    required=False,
                    publication_required=False,
                    source_type="wb_stock_history_daily",
                    source_label="История остатков WB",
                    status="loaded",
                    snapshot_hash="filterable-provider-window-hash",
                    row_count=1,
                    loaded_at=repository.security.utcnow(),
                ),
                ReportLostSalesRow(
                    report_run_id=report.id,
                    client_id=report.client_id,
                    wb_cabinet_id="cabinet-filter-1",
                    row_uid="lost-filter-1",
                    cabinet="Кабинет фильтра",
                    product="Товар фильтра",
                    article_1c="ARTICLE-1",
                    barcode="BARCODE-1",
                    zero_stock_days=Decimal("1"),
                    onec_stock_quantity=Decimal("10"),
                    onec_warehouses="Основной: 10",
                    sales=Decimal("31"),
                    lost_units=Decimal("1"),
                    lost_revenue=Decimal("100"),
                    lost_profit=Decimal("50"),
                    note="Предварительная оценка",
                    calculation_context={
                        "version": "lost-sales-filter-v1",
                        "providerPeriodStart": provider_start.isoformat(),
                        "providerPeriodEnd": provider_end.isoformat(),
                        "stockByDate": stock_by_date,
                        "financePeriods": [
                            {
                                "periodStart": "2026-04-08",
                                "periodEnd": "2026-04-14",
                                "salesQuantity": "7",
                                "netRevenue": "700",
                                "contributionMargin": "350",
                            },
                            {
                                "periodStart": "2026-05-01",
                                "periodEnd": "2026-05-31",
                                "salesQuantity": "31",
                                "netRevenue": "3100",
                                "contributionMargin": "1550",
                            }
                        ],
                    },
                ),
            ]
        )
        db.flush()

        may = repository.query_report_rows(
            db,
            report,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            wb_cabinet_id="cabinet-filter-1",
        )
        may_common = repository.query_report_rows(
            db,
            report,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
        )
        march = repository.query_report_rows(
            db,
            report,
            period_start=date(2026, 3, 1),
            period_end=date(2026, 3, 31),
            wb_cabinet_id="cabinet-filter-1",
        )
        april = repository.query_report_rows(
            db,
            report,
            period_start=date(2026, 4, 1),
            period_end=date(2026, 4, 30),
            wb_cabinet_id="cabinet-filter-1",
        )

    may_coverage = may["analytics"]["lostSalesCoverage"]
    assert may_coverage["requestedPeriodStart"] == "2026-05-01"
    assert may_coverage["requestedPeriodEnd"] == "2026-05-31"
    assert may_coverage["calculationPeriodStart"] == "2026-05-01"
    assert may_coverage["calculationPeriodEnd"] == "2026-05-31"
    assert may_coverage["coveredDays"] == 31
    assert may_coverage["totalDays"] == 31
    assert may_coverage["fullCoverage"] is True
    assert may_coverage["extrapolated"] is False
    assert may["analytics"]["kpis"]["lostContributionMargin"] == pytest.approx(
        51.6666666667
    )
    assert may["analytics"]["lostSales"][0]["zeroStockDays"] == 1.0

    common_coverage = may_common["analytics"]["lostSalesCoverage"]
    assert common_coverage["calculationPeriodStart"] == "2026-05-05"
    assert common_coverage["calculationPeriodEnd"] == "2026-05-31"
    assert common_coverage["coveredDays"] == 27
    assert common_coverage["totalDays"] == 31

    march_coverage = march["analytics"]["lostSalesCoverage"]
    assert march_coverage["calculated"] is False
    assert march_coverage["coveredDays"] == 0
    assert march["analytics"]["kpis"]["lostContributionMargin"] is None

    april_coverage = april["analytics"]["lostSalesCoverage"]
    assert april_coverage["calculationPeriodStart"] == "2026-04-10"
    assert april_coverage["calculationPeriodEnd"] == "2026-04-30"
    assert april_coverage["coveredDays"] == 21
    assert april_coverage["totalDays"] == 30
    assert april_coverage["extrapolated"] is False
    assert april["analytics"]["kpis"]["lostContributionMargin"] == pytest.approx(
        12.5
    )


def test_save_report_marts_flushes_and_persists_lost_sales_context(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    payload = ready_payload()
    payload["lostSales"] = [
        {
            "id": "lost-context-save-1",
            "cabinet": "Кабинет контекста",
            "product": "Товар контекста",
            "article1c": "ARTICLE-CONTEXT",
            "barcode": "BARCODE-CONTEXT",
            "zeroStockDays": 1,
            "onecStock": 2,
            "onecWarehouses": "Основной: 2",
            "sales": 3,
            "lostUnits": 1,
            "lostRevenue": 100,
            "lostProfit": 50,
            "note": "Контекст сохранён",
            "calculationContext": {
                "version": "lost-sales-filter-v1",
                "providerPeriodStart": "2026-04-10",
                "providerPeriodEnd": "2026-04-30",
                "stockByDate": {"2026-04-10": "0"},
                "financePeriods": [],
            },
        }
    ]
    with session_factory() as db:
        report = repository.save_report_marts(
            db,
            payload,
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="saved-context-report",
            publication_status="draft",
            publish=False,
        )
        row = db.scalar(
            select(ReportLostSalesRow).where(
                ReportLostSalesRow.report_run_id == report.id
            )
        )

    assert row is not None
    assert row.calculation_context["version"] == "lost-sales-filter-v1"
    assert row.calculation_context["stockByDate"] == {"2026-04-10": "0"}


def test_filtered_lost_sales_do_not_infer_legacy_calculation_context(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="legacy-provider-window-report",
            publication_status="draft",
            publish=False,
        )
        report.period_start = date(2026, 3, 1)
        report.period_end = date(2026, 7, 10)
        refresh = repository.create_source_refresh_run(
            db,
            tenant_id="shumeyko",
            client_id=report.client_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="legacy-provider-window-refresh",
            period_start=report.period_start,
            period_end=report.period_end,
            reason="legacy provider window test",
        )
        repository.add_source_refresh_collection(
            db,
            refresh,
            source_type="wb_stock_history_daily",
            source_label="История остатков WB",
            required=False,
            status="loaded",
            row_count=1,
            payload={
                "calculated": True,
                "providerWindowCalculated": True,
                "calculationPeriodStart": "2026-04-10",
                "calculationPeriodEnd": "2026-07-10",
                "accounts": [
                    {
                        "sellerAccountId": "legacy-account",
                        "wbCabinetId": "legacy-cabinet",
                        "status": "partial_provider_window",
                        "coveredDays": 92,
                        "totalDays": 132,
                        "calculated": True,
                        "providerWindowCalculated": True,
                    }
                ],
            },
        )
        db.add(
            SourceLoad(
                tenant_id="shumeyko",
                client_id=report.client_id,
                wb_cabinet_id="",
                report_run_id=report.id,
                source_refresh_run_id=refresh.id,
                required=False,
                publication_required=False,
                source_type="wb_stock_history_daily",
                source_label="История остатков WB",
                status="loaded",
                snapshot_hash="legacy-provider-window-hash",
                row_count=1,
                loaded_at=repository.security.utcnow(),
            )
        )
        db.flush()

        payload = repository.query_report_rows(
            db,
            report,
            period_start=date(2026, 5, 1),
            period_end=date(2026, 5, 31),
            wb_cabinet_id="legacy-cabinet",
        )

    coverage = payload["analytics"]["lostSalesCoverage"]
    assert coverage["calculated"] is False
    assert coverage["calculationContextVersion"] is None
    assert payload["analytics"]["lostSales"] == []
    assert payload["analytics"]["kpis"]["lostContributionMargin"] is None


def test_optional_stock_movement_failure_is_review_only(tmp_path: Path) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        draft = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="optional-stock-failed-draft",
            publication_status="draft",
            publish=False,
        )
        db.add(
            SourceLoad(
                tenant_id="shumeyko",
                client_id=draft.client_id,
                wb_cabinet_id="",
                report_run_id=draft.id,
                source_refresh_run_id=None,
                required=False,
                publication_required=False,
                source_type="onec_stock_movements",
                source_label="AccumulationRegister_Запасы",
                status="failed",
                snapshot_hash="failed-manifest-hash",
                row_count=0,
                loaded_at=datetime(2026, 7, 10, 12, 0),
            )
        )
        db.flush()

        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }

    assert "source_load_failed" not in codes
    assert "source_lineage_failed" not in codes


def test_source_backed_report_blocks_unlinked_snapshot_and_stock_history(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        draft = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="unlinked-source-backed-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="unlinked-composite-snapshot",
        )
        db.flush()

        codes = {
            item["code"] for item in repository.report_publication_blockers(db, draft)
        }

    assert "source_lineage_missing" in codes
    assert "stock_history_lineage_missing" in codes


def test_report_readiness_hides_staff_draft_state_from_client_role(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, payload=ready_payload())
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "readiness-client@example.com", "role": "client"},
    ).json()

    client.post("/api/auth/logout")
    login_as(client, "readiness-client@example.com", created["temporaryPassword"])

    summary = client.get("/api/reports/report-1/summary").json()
    reason_codes = {reason["code"] for reason in summary["readiness"]["reviewReasons"]}
    assert "client_draft_missing" not in reason_codes
    assert summary["readiness"]["status"] == "ready"


def test_report_summary_includes_latest_source_refresh_safely(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)
    with client.app.state.session_factory() as db:
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        refresh_run = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            mode="full",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="full-test",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 6, 17),
            user=user,
            source_report=report,
            reason="test refresh",
        )
        repository.add_source_refresh_collection(
            db,
            refresh_run,
            source_type="wb_finance_detail",
            source_label="WB Finance sales report details",
            required=True,
            status="failed",
            row_count=0,
            raw_path="data/source_refresh/full-test/wb_finance",
            error_message="HTTP 401 token expired",
        )
        repository.update_source_refresh_run(
            db,
            refresh_run,
            status="failed",
            error_message="HTTP 401 token expired",
            finished_at=repository.security.utcnow(),
        )
        db.commit()

    staff_summary = client.get("/api/reports/report-1/summary").json()
    staff_refresh = staff_summary["latestSourceRefresh"]
    assert staff_refresh["status"] == "failed"
    assert staff_refresh["errorMessage"] == "HTTP 401 token expired"
    assert staff_refresh["collections"][0]["rawPath"].endswith("wb_finance")

    created = client.post(
        "/api/admin/users",
        json={"email": "refresh-view-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "refresh-view-client@example.com", created["temporaryPassword"])

    client_summary = client.get("/api/reports/report-1/summary").json()
    client_refresh = client_summary["latestSourceRefresh"]
    assert client_refresh["status"] == "failed"
    assert client_refresh["errorMessage"].startswith("Последнее обновление данных")
    assert client_refresh["collections"][0]["rawPath"] == ""
    assert client_refresh["collections"][0]["payload"] == {}
    assert "HTTP 401" not in str(client_refresh)


def test_analytical_report_artifact_requires_auth_and_downloads(
    tmp_path: Path,
    monkeypatch,
) -> None:
    client = make_client(tmp_path)
    received: dict[str, object] = {}

    assert client.post("/api/reports/report-1/analytical-report").status_code == 401
    login(client)

    def fake_build_client_analytical_report(**kwargs):
        received.update(kwargs)
        output_dir = kwargs["output_dir"]
        output_dir.mkdir(parents=True, exist_ok=True)
        markdown_path = output_dir / "report.md"
        docx_path = output_dir / "report.docx"
        markdown_path.write_text("# Отчет", encoding="utf-8")
        docx_path.write_bytes(b"docx")
        return SimpleNamespace(
            markdown_path=markdown_path,
            docx_path=docx_path,
            pdf_path=None,
            pdf_status="unavailable",
            pdf_message="PDF converter is unavailable.",
        )

    monkeypatch.setattr(
        "wb_unit_economics.web.app.build_client_analytical_report",
        fake_build_client_analytical_report,
    )

    generated = client.post(
        "/api/reports/report-1/analytical-report",
        json={"branded": True},
    )
    assert generated.status_code == 200
    payload = generated.json()
    assert "workbook_path" not in received
    assert received["summary"]["meta"]["reportId"] == "report-1"
    assert received["summary"]["meta"]["reportPeriod"] == (
        "08.06.2026 - 14.06.2026"
    )
    assert payload["files"]["docx"]["url"].endswith("/analytical-report.docx")
    assert payload["contractVersion"] == "client-analytical-report.v3"
    assert payload["scope"] == "last_closed_week"
    assert payload["periodStart"] == "2026-06-08"
    assert payload["periodEnd"] == "2026-06-14"
    assert payload["files"]["pdf"]["status"] == "unavailable"

    invalid_custom = client.post(
        "/api/reports/report-1/analytical-report",
        json={"scope": "custom"},
    )
    assert invalid_custom.status_code == 400

    docx = client.get("/api/reports/report-1/analytical-report.docx")
    assert docx.status_code == 200
    assert docx.content == b"docx"

    pdf = client.get("/api/reports/report-1/analytical-report.pdf")
    assert pdf.status_code == 404


def test_tenant_isolation(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    assert client.get("/api/reports/other-report/summary").status_code == 404
    assert client.get("/api/reports/other-report/client-draft").status_code == 404


def test_client_draft_is_staff_only(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    created = client.post(
        "/api/admin/users",
        json={"email": "draft-client@example.com", "role": "client"},
    ).json()
    staff_view = client.get("/api/reports/report-1/client-draft")
    assert staff_view.status_code == 200
    assert staff_view.json()["latest"] is None

    client.post("/api/auth/logout")
    login_as(
        client,
        "draft-client@example.com",
        created["temporaryPassword"],
    )

    assert client.get("/api/reports/report-1/client-draft").status_code == 403


def test_client_cannot_open_staff_draft_or_financial_recommendations(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path, payload=ready_payload())
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "financial-gate-client@example.com", "role": "client"},
    ).json()
    with client.app.state.session_factory() as db:
        import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="staff-only-draft",
            publication_status="draft",
            publish=False,
        )
        current_row = db.scalar(
            select(repository.ReportUnitRow).where(
                repository.ReportUnitRow.report_run_id == "report-1"
            )
        )
        assert current_row is not None
        current_row.tax_method = "ОСНО; НДС 22% внутри цены"
        current_row.pnl_vat_mode = ""
        current_row.vat_input_completeness = "partial"
        db.commit()

    client.post("/api/auth/logout")
    login_as(
        client,
        "financial-gate-client@example.com",
        created["temporaryPassword"],
    )

    reports = client.get("/api/clients/shumeyko/reports")
    assert [item["id"] for item in reports.json()["items"]] == ["report-1"]
    assert client.get("/api/reports/staff-only-draft/summary").status_code == 404
    blocked = client.get("/api/reports/report-1/management-report")
    assert blocked.status_code == 409
    assert blocked.json()["detail"].startswith("Финансовая проверка не пройдена")


def test_client_draft_revisions_finalize_and_audit(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.delenv("OPENAI_API_KEY", raising=False)
    client = make_client(tmp_path)
    login(client)

    first = client.post(
        "/api/reports/report-1/client-draft/refine",
        json={"action": "assemble", "instruction": "Собери клиентский текст"},
    )
    assert first.status_code == 200
    first_payload = first.json()
    assert first_payload["changed"] is True
    assert first_payload["latest"]["revision"] == 1
    assert first_payload["latest"]["source"] == "deterministic_base"
    assert "Ключевой вывод" in first_payload["latest"]["content"]
    assert "Ограничения" in first_payload["latest"]["content"]
    assert "draft_management_report" not in first_payload["latest"]["content"]
    assert "tool_completed" not in first_payload["latest"]["content"]

    unavailable = client.post(
        "/api/reports/report-1/client-draft/refine",
        json={"action": "shorten", "instruction": "Сократи"},
    )
    assert unavailable.status_code == 200
    unavailable_payload = unavailable.json()
    assert unavailable_payload["changed"] is False
    assert unavailable_payload["aiAvailable"] is False
    assert "не изменен" in unavailable_payload["message"]
    assert len(unavailable_payload["revisions"]) == 1

    manual_text = (
        "Ключевой вывод\n"
        "Клиенту нужно проверить убыточность и себестоимость.\n\n"
        "Факты\n"
        "- В отчете есть убыточная строка.\n\n"
        "Что требует проверки\n"
        "- Себестоимость 1С по товару без себестоимости.\n\n"
        "Ограничения\n"
        "- Причины возврата не передаются текущими источниками.\n\n"
        "Следующий шаг\n"
        "Согласовать проверку с аналитиком."
    )
    saved = client.put(
        "/api/reports/report-1/client-draft",
        json={"content": manual_text, "instruction": "Ручная правка"},
    )
    assert saved.status_code == 200
    saved_payload = saved.json()
    assert saved_payload["latest"]["revision"] == 2
    assert saved_payload["latest"]["source"] == "manual"

    finalized = client.post(
        "/api/reports/report-1/client-draft/finalize",
        json={"revision": 2},
    )
    assert finalized.status_code == 200
    assert finalized.json()["latest"]["status"] == "ready"

    audit = client.get("/api/admin/audit")
    actions = {item["action"] for item in audit.json()["items"]}
    assert "ai_client_draft_created" in actions
    assert "ai_client_draft_saved" in actions
    assert "ai_client_draft_finalized" in actions


def test_client_draft_rejects_internal_labels(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    response = client.put(
        "/api/reports/report-1/client-draft",
        json={
            "content": "Ключевой вывод\nСработал draft_management_report.",
            "instruction": "Ручная правка",
        },
    )
    assert response.status_code == 400


def test_admin_user_management_and_audit(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    created = client.post(
        "/api/admin/users",
        json={
            "email": "client@example.com",
            "name": "Client",
            "role": "client",
        },
    )
    assert created.status_code == 200
    client_user = created.json()["user"]
    assert client_user["tenants"][0]["role"] == "client"
    assert created.json()["temporaryPassword"]

    users = client.get("/api/admin/users")
    assert users.status_code == 200
    assert {item["email"] for item in users.json()["items"]} == {
        "admin@example.com",
        "client@example.com",
    }

    reset = client.post(f"/api/admin/users/{client_user['id']}/reset-password", json={})
    assert reset.status_code == 200
    assert reset.json()["temporaryPassword"]

    disabled = client.patch(
        f"/api/admin/users/{client_user['id']}",
        json={"is_active": False},
    )
    assert disabled.status_code == 200
    assert disabled.json()["user"]["isActive"] is False

    audit = client.get("/api/admin/audit")
    assert audit.status_code == 200
    assert any(
        item["action"] == "user_password_reset" for item in audit.json()["items"]
    )


def test_client_role_cannot_manage_users(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    client.post(
        "/api/auth/login",
        json={
            "email": "client@example.com",
            "password": created["temporaryPassword"],
        },
    )

    assert client.get("/api/admin/users").status_code == 403
    assert client.get("/api/admin/audit").status_code == 403


def test_tenant_integrations_are_staff_only_and_mask_secrets(
    tmp_path: Path, monkeypatch
) -> None:
    integration_key = Fernet.generate_key().decode("ascii")
    checked_secrets: list[tuple[str, str]] = []

    def fake_check(
        _settings: WebSettings, *, provider: str, secret: str
    ) -> integrations.IntegrationCheckResult:
        checked_secrets.append((provider, secret))
        return integrations.IntegrationCheckResult(
            status="check_ok",
            message=f"{provider} проверен read-only",
            payload={
                "provider": provider,
                "checkedAt": "2026-06-21T09:00:00+03:00",
                "checkMode": "live_read_only",
                "endpointCategory": "test_ping",
                "httpStatus": 200,
            },
        )

    monkeypatch.setattr(integrations, "run_provider_check", fake_check)
    client = make_client(
        tmp_path,
        settings_overrides={"integration_secret_key": integration_key},
    )
    login(client)

    empty = client.get("/api/integrations")
    assert empty.status_code == 200
    empty_payload = empty.json()
    assert {item["provider"] for item in empty_payload["items"]} == {
        "wb_api",
        "ozon_api",
        "onec_readonly",
    }
    provider_metadata = {
        item["providerBase"]: item for item in empty_payload["providers"]
    }
    assert provider_metadata["wb_api"] == {
        "providerBase": "wb_api",
        "label": "API Wildberries",
        "readOnly": True,
        "supportsMultiple": True,
        "primaryProviderId": "wb_api",
        "roles": [
            {
                "id": "finance_reports",
                "label": "Финансовые отчеты",
                "default": True,
            },
            {
                "id": "analytics_stocks",
                "label": "Аналитика и остатки",
                "default": False,
            },
            {
                "id": "content_cards",
                "label": "Карточки товаров",
                "default": False,
            },
            {
                "id": "full_readonly",
                "label": "Полный доступ только для чтения",
                "default": False,
            },
        ],
    }
    assert provider_metadata["ozon_api"]["label"] == "API кабинета продавца Ozon"
    assert provider_metadata["ozon_api"]["roles"][0] == {
        "id": "finance_reports",
        "label": "Финансовые отчеты",
        "default": True,
    }

    saved = client.put(
        "/api/integrations/wb_api",
        json={"label": "WB кабинет", "secret": "wb-token-secret-123456"},
    )
    assert saved.status_code == 200
    payload = saved.json()
    assert payload["status"] == "configured"
    assert payload["configured"] is True
    assert payload["providerBase"] == "wb_api"
    assert payload["connectionKey"] == "primary"
    assert payload["connectionRole"] == "finance_reports"
    assert payload["isPrimary"] is True
    assert payload["secretHint"] == "***3456"
    assert payload["storageMode"] == "encrypted"
    assert "wb-token-secret" not in str(payload)

    checked = client.post("/api/integrations/wb_api/check", json={})
    assert checked.status_code == 200
    assert checked.json()["status"] == "check_ok"
    assert checked.json()["lastCheck"]["message"] == "wb_api проверен read-only"
    assert checked.json()["lastCheck"]["httpStatus"] == 200
    assert checked_secrets == [("wb_api", "wb-token-secret-123456")]

    extra = client.post(
        "/api/integrations",
        json={
            "provider": "wb_api",
            "label": "WB кабинет маркетплейс 2",
            "connection_role": "analytics_stocks",
            "cabinet_name": "Кабинет 2",
            "organization_name": "ООО Тест",
            "secret": "wb-extra-token-654321",
        },
    )
    assert extra.status_code == 200
    extra_payload = extra.json()
    assert extra_payload["provider"].startswith("wb_api:")
    assert extra_payload["providerBase"] == "wb_api"
    assert extra_payload["connectionKey"] != "primary"
    assert extra_payload["connectionRole"] == "analytics_stocks"
    assert extra_payload["cabinetName"] == "Кабинет 2"
    assert extra_payload["organizationName"] == "ООО Тест"
    assert extra_payload["isPrimary"] is False
    assert extra_payload["secretHint"] == "***4321"
    assert "wb-extra-token" not in str(extra_payload)

    listed = client.get("/api/integrations").json()["items"]
    assert {"wb_api", "onec_readonly", extra_payload["provider"]} <= {
        item["provider"] for item in listed
    }

    checked_extra = client.post(
        f"/api/integrations/{quote(extra_payload['provider'], safe='')}/check",
        json={},
    )
    assert checked_extra.status_code == 200
    assert checked_extra.json()["status"] == "check_ok"
    assert checked_secrets[-1] == ("wb_api", "wb-extra-token-654321")

    saved_onec = client.put(
        "/api/integrations/onec_readonly",
        json={
            "label": "1С тест",
            "secret": (
                "baseUrl=https://onec.example.test/odata/standard.odata;"
                "username=reader;password=onec-secret;verifySsl=true"
            ),
        },
    )
    assert saved_onec.status_code == 200
    assert saved_onec.json()["storageMode"] == "encrypted"
    checked_onec = client.post("/api/integrations/onec_readonly/check", json={})
    assert checked_onec.status_code == 200
    assert checked_onec.json()["status"] == "check_ok"
    assert checked_secrets[-1] == (
        "onec_readonly",
        "baseUrl=https://onec.example.test/odata/standard.odata;"
        "username=reader;password=onec-secret;verifySsl=true",
    )

    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        integration = (
            db.query(TenantIntegration)
            .filter_by(tenant_id="shumeyko", provider="wb_api")
            .one()
        )
        assert integration.config_payload["storage"] == "encrypted"
        assert "secretCiphertext" in integration.config_payload
        assert "wb-token-secret" not in str(integration.config_payload)
        assert (
            db.query(WbCabinet)
            .filter_by(client_id="shumeyko", display_name="WB кабинет")
            .one_or_none()
            is None
        )
        cabinet = (
            db.query(WbCabinet)
            .filter_by(client_id="shumeyko", display_name="Кабинет 2")
            .one()
        )
        assert cabinet.provider == extra_payload["provider"]

    disabled = client.post("/api/integrations/wb_api/disable", json={})
    assert disabled.status_code == 200
    assert disabled.json()["status"] == "disabled"
    assert "wb-token-secret" not in str(disabled.json())

    audit = client.get("/api/admin/audit").json()["items"]
    audit_text = str(audit)
    assert "tenant_integration_saved" in {item["action"] for item in audit}
    assert "tenant_integration_checked" in {item["action"] for item in audit}
    assert "tenant_integration_disabled" in {item["action"] for item in audit}
    assert "wb-token-secret" not in audit_text

    created = client.post(
        "/api/admin/users",
        json={"email": "integrations-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "integrations-client@example.com", created["temporaryPassword"])

    assert client.get("/api/integrations").status_code == 403
    assert (
        client.put(
            "/api/integrations/wb_api",
            json={"secret": "another-secret"},
        ).status_code
        == 403
    )
    assert (
        client.post(
            "/api/integrations",
            json={"provider": "wb_api", "secret": "another-secret"},
        ).status_code
        == 403
    )


def test_tenant_integration_hash_only_storage_cannot_live_check(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)

    saved = client.put(
        "/api/integrations/wb_api",
        json={"label": "WB без ключа шифрования", "secret": "legacy-token-123"},
    )
    assert saved.status_code == 200
    assert saved.json()["status"] == "configured"
    assert saved.json()["storageMode"] == "hash_only"

    checked = client.post("/api/integrations/wb_api/check", json={})
    assert checked.status_code == 200
    payload = checked.json()
    assert payload["status"] == "check_failed"
    assert payload["lastCheck"]["checkMode"] == "configuration"
    assert "SHUMEYKO_INTEGRATION_SECRET_KEY" in payload["lastCheck"]["message"]
    assert "legacy-token" not in str(payload)


def test_ozon_live_check_accepts_any_supported_readonly_endpoint(
    monkeypatch,
) -> None:
    calls: list[tuple[str, dict]] = []

    class FakeResponse:
        def __init__(self, status_code: int) -> None:
            self.status_code = status_code

    class FakeClient:
        def __init__(self, **_kwargs) -> None:
            self._statuses = iter([403, 200])

        def __enter__(self) -> FakeClient:
            return self

        def __exit__(self, *_args: object) -> None:
            return None

        def post(self, url: str, json: dict) -> FakeResponse:
            calls.append((url, json))
            return FakeResponse(next(self._statuses))

    monkeypatch.setattr(integrations.httpx, "Client", FakeClient)

    result = integrations.run_provider_check(
        WebSettings(_env_file=None),
        provider="ozon_api",
        secret='{"clientId":"12345","apiKey":"ozon-secret-key"}',
    )

    assert result.status == "check_ok"
    assert result.payload["endpointCategory"] == "stock_on_warehouses"
    assert result.payload["checkedEndpoints"] == [
        {"endpointCategory": "finance_cash_flow", "httpStatus": 403},
        {"endpointCategory": "stock_on_warehouses", "httpStatus": 200},
    ]
    assert calls[0][0].endswith("/v1/finance/cash-flow-statement/list")
    assert calls[1][0].endswith("/v2/analytics/stock_on_warehouses")
    assert "seller/info" not in str(calls)
    assert "ozon-secret-key" not in str(result.payload)


def test_live_checks_are_read_only_and_disabled_by_default(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    response = client.post(
        "/api/reports/report-1/live-checks/onec-cost",
        json={"lookup": "BAR-NOCOST"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "disabled"
    assert payload["reviewStatus"] == "needs_review"
    assert "не опрашивались" in payload["message"]


def test_onec_auto_refresh_is_staff_only_flagged_and_creates_new_report(
    tmp_path: Path,
) -> None:
    fake_service = FakeAutoRefreshService(tmp_path / "reports" / "auto-refresh.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=fake_service,
    )
    login(client)

    response = client.post(
        "/api/reports/report-1/refresh/onec-auto",
        json={"reason": "Нужно дозагрузить себестоимость 1С"},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "report_created"
    assert payload["jobType"] == "source_refresh"
    assert payload["sourceReportRunId"] == "report-1"
    assert payload["newReportRunId"] == "report-1-refresh"
    assert "row_payload" not in str(payload).lower()

    old_summary = client.get("/api/reports/report-1/summary").json()
    new_summary = client.get("/api/reports/report-1-refresh/summary").json()
    assert old_summary["quality"]["missingCostRows"] == 1
    assert new_summary["quality"]["missingCostRows"] == 0
    new_rows = client.get(
        "/api/reports/report-1-refresh/rows",
        params={"status_filter": "ОК", "limit": 10},
    ).json()
    assert new_rows["total"] == 2
    assert [item["id"] for item in client.get("/api/reports").json()["items"]][0] == (
        "report-1-refresh"
    )

    job = client.get(f"/api/reports/report-1/refresh-jobs/{payload['id']}")
    assert job.status_code == 200
    assert job.json()["newReportRunId"] == "report-1-refresh"
    assert job.json()["mode"] == "onec-only"

    audit = client.get("/api/admin/audit")
    actions = {item["action"] for item in audit.json()["items"]}
    assert "source_refresh_requested" in actions
    assert "source_refresh_report_created" in actions
    assert "onec_auto_refresh_started" not in actions


def test_onec_auto_refresh_wrapper_calls_source_refresh_onec_only(
    tmp_path: Path,
) -> None:
    class FakeSourceRefreshService:
        def __init__(self) -> None:
            self.kwargs = {}

        def run(self, db, **kwargs):
            raise AssertionError("non-dry auto refresh must not run in web")

        def enqueue(self, db, **kwargs):
            self.kwargs = kwargs
            refresh_run = repository.create_source_refresh_run(
                db,
                tenant_id=kwargs["tenant_id"],
                client_id=kwargs["source_report"].client_id,
                mode=kwargs["mode"],
                credential_source=kwargs["credential_source"],
                dry_run=False,
                snapshot_set_id="onec-auto-queued",
                period_start=date(2026, 3, 1),
                period_end=date(2026, 6, 30),
                user=kwargs["user"],
                source_report=kwargs["source_report"],
                reason=kwargs["reason"],
            )
            return repository.source_refresh_run_payload(refresh_run)

    client = make_client(tmp_path)
    fake_source_refresh = FakeSourceRefreshService()
    launched: list[str] = []

    class _Launcher:
        def launch(self, refresh_run_id: str) -> str:
            launched.append(refresh_run_id)
            return f"shumeiko-source-refresh-worker@{refresh_run_id}.service"

    service = OnecAutoRefreshService(
        WebSettings(
            database_url=f"sqlite:///{tmp_path / 'web.sqlite3'}",
            cookie_secure=False,
            allowed_export_root=str(tmp_path / "reports"),
            source_refresh_enabled=True,
        ),
        source_refresh_service=fake_source_refresh,
        worker_launcher=_Launcher(),
    )
    with client.app.state.session_factory() as db:
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        payload = service.run(db, user=user, report=report, reason="refresh 1c")

    assert fake_source_refresh.kwargs["tenant_id"] == "shumeyko"
    assert fake_source_refresh.kwargs["mode"] == "onec-only"
    assert fake_source_refresh.kwargs["credential_source"] == "tenant"
    assert "dry_run" not in fake_source_refresh.kwargs
    assert fake_source_refresh.kwargs["source_report"].id == "report-1"
    assert payload["status"] == "queued"
    assert payload["jobType"] == "source_refresh"
    assert payload["sourceRefreshRunId"] == payload["id"]
    assert launched == [payload["id"]]


def test_onec_auto_refresh_disabled_and_client_forbidden(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    disabled = client.post(
        "/api/reports/report-1/refresh/onec-auto",
        json={"reason": "Дозагрузить 1С"},
    )
    assert disabled.status_code == 409
    assert "Обновление источников выключено" in disabled.json()["detail"]

    created = client.post(
        "/api/admin/users",
        json={"email": "refresh-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "refresh-client@example.com", created["temporaryPassword"])
    forbidden = client.post(
        "/api/reports/report-1/refresh/onec-auto",
        json={"reason": "Дозагрузить 1С"},
    )
    assert forbidden.status_code == 403


def test_onec_auto_refresh_rejects_active_job(tmp_path: Path) -> None:
    fake_service = FakeAutoRefreshService(tmp_path / "reports" / "auto-refresh.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=fake_service,
    )
    login(client)
    with client.app.state.session_factory() as db:
        user = db.query(repository.User).filter_by(email="admin@example.com").one()
        report = db.get(repository.ReportRun, "report-1")
        assert report is not None
        repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            mode="onec-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="onec-only-active",
            period_start=date(2026, 3, 1),
            period_end=date(2026, 6, 17),
            user=user,
            source_report=report,
            reason="already running",
        )
        db.commit()

    response = client.post(
        "/api/reports/report-1/refresh/onec-auto",
        json={"reason": "Дозагрузить 1С"},
    )
    assert response.status_code == 409


def test_onec_auto_refresh_returns_503_when_worker_is_unavailable(
    tmp_path: Path,
) -> None:
    class _UnavailableAutoRefresh:
        def run(self, *_args, **_kwargs):
            raise AutoRefreshUnavailableError("Не удалось запустить обновление.")

    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=_UnavailableAutoRefresh(),
    )
    login(client)

    response = client.post(
        "/api/reports/report-1/refresh/onec-auto",
        json={"reason": "Дозагрузить 1С"},
    )

    assert response.status_code == 503
    assert "Не удалось запустить" in response.json()["detail"]


def test_ai_fallback_uses_report_facts(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Что самое важное по убыточности?"},
    ).json()
    assistant_messages = [
        item["content"] for item in answer["messages"] if item["role"] == "assistant"
    ]
    assert assistant_messages
    assert "Убыточных строк" in assistant_messages[-1]
    assert "не меняю данные" in assistant_messages[-1]
    assistant_payloads = [
        item for item in answer["messages"] if item["role"] == "assistant"
    ]
    assert assistant_payloads[-1]["citations"][0]["reportId"] == "report-1"
    assert assistant_payloads[-1]["citations"][0]["scopeHash"]
    assert any(item["type"] == "tool_completed" for item in answer["events"])
    done_events = [
        item for item in answer["events"] if item["type"] == "assistant_done"
    ]
    assert done_events[-1]["payload"]["answerSource"] == "fallback"
    assert done_events[-1]["payload"]["fallbackReason"] == "no_openai_key"


def test_ai_openai_source_is_visible_when_model_answers(
    tmp_path: Path,
    monkeypatch,
) -> None:
    def fake_openai_answer(self, db, user, thread, report, question):
        return "OpenAI: главный риск — себестоимость и убыточные SKU.", ""

    monkeypatch.setattr(AiAnalyst, "_openai_answer", fake_openai_answer)
    client = make_client(tmp_path, settings_overrides={"openai_api_key": "test-key"})
    login(client)

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Что главное?"},
    ).json()

    assistant_messages = [
        item["content"] for item in answer["messages"] if item["role"] == "assistant"
    ]
    assert assistant_messages[-1].startswith("OpenAI:")
    done_events = [
        item for item in answer["events"] if item["type"] == "assistant_done"
    ]
    assert done_events[-1]["payload"]["answerSource"] == "openai"
    assert done_events[-1]["payload"]["model"]
    assert any(item["type"] == "answer_source" for item in answer["events"])


def test_ai_prompts_are_versioned_package_resources() -> None:
    analyst_prompt = load_prompt("ai_analyst")
    client_draft_prompt = load_prompt("client_draft")

    assert "{{LIMITATIONS}}" in analyst_prompt
    assert "короткое приветствие" in analyst_prompt
    assert "Обязательные разделы" in client_draft_prompt
    rendered = render_prompt("ai_analyst", LIMITATIONS="- Только тест")
    assert "{{LIMITATIONS}}" not in rendered
    assert "- Только тест" in rendered
    with pytest.raises(RuntimeError, match="Unresolved placeholders"):
        render_prompt("ai_analyst")

    analyst = AiAnalyst(WebSettings())
    assert analyst._is_conversational_message("Привет!") is True
    assert analyst._is_conversational_message("Привет! Что главное?") is False


def test_ai_short_greeting_does_not_call_report_tools(
    tmp_path: Path,
    monkeypatch,
) -> None:
    import openai

    class FakeResponse:
        output = []
        output_text = "Здравствуйте! Могу помочь разобрать показатели отчёта."

    requests = []

    class FakeResponses:
        def create(self, **kwargs):
            requests.append(kwargs)
            return FakeResponse()

    class FakeOpenAI:
        def __init__(self, **_kwargs):
            self.responses = FakeResponses()

    monkeypatch.setattr(openai, "OpenAI", FakeOpenAI)
    client = make_client(tmp_path, settings_overrides={"openai_api_key": "test-key"})
    login(client)
    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()

    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Привет!"},
    ).json()

    assert requests[0]["tool_choice"] == "none"
    assert not any(
        item["type"] in {"tool_started", "tool_completed"}
        for item in answer["events"]
    )
    assistant = [
        item for item in answer["messages"] if item["role"] == "assistant"
    ][-1]
    assert assistant["content"].startswith("Здравствуйте!")
    assert assistant["citations"] == []
    source_event = [
        item for item in answer["events"] if item["type"] == "answer_source"
    ][-1]
    assert source_event["payload"]["toolNames"] == []
    assert "без обращения к данным отчёта" in source_event["message"]


def test_ai_responses_tool_loop_is_stateless_typed_and_runs_tool_once(
    tmp_path: Path,
    monkeypatch,
) -> None:
    import openai

    class FakeCall:
        type = "function_call"
        name = "get_report_summary"
        call_id = "call-summary"
        arguments = "{}"
        status = "completed"

    class FakeResponse:
        def __init__(self, output, output_text=""):
            self.output = output
            self.output_text = output_text

    requests = []
    responses = iter(
        [
            FakeResponse([FakeCall()]),
            FakeResponse([], "Главный вывод собран по расчетной витрине."),
            FakeResponse([FakeCall()]),
            FakeResponse([], "Продолжение учитывает историю диалога."),
        ]
    )

    class FakeResponses:
        def create(self, **kwargs):
            requests.append(kwargs)
            return next(responses)

    class FakeOpenAI:
        def __init__(self, **_kwargs):
            self.responses = FakeResponses()

    monkeypatch.setattr(openai, "OpenAI", FakeOpenAI)
    client = make_client(tmp_path, settings_overrides={"openai_api_key": "test-key"})
    login(client)
    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()

    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Что главное?"},
    ).json()

    assert len(requests) == 2
    assert requests[0]["tool_choice"] == "required"
    assert all(request["store"] is False for request in requests)
    assert all(
        request["include"] == ["reasoning.encrypted_content"]
        for request in requests
    )
    assert isinstance(requests[1]["input"][2], FakeCall)
    completed = [
        item
        for item in answer["events"]
        if item["type"] == "tool_completed"
        and item["toolName"] == "get_report_summary"
    ]
    assert len(completed) == 1

    client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "А что было в прошлом ответе?"},
    )
    second_input = requests[2]["input"]
    assert any(
        item.get("role") == "assistant"
        and item.get("content") == "Главный вывод собран по расчетной витрине."
        for item in second_input
        if isinstance(item, dict)
    )
    assert sum(
        1
        for item in second_input
        if isinstance(item, dict)
        and item.get("role") == "user"
        and item.get("content") == "А что было в прошлом ответе?"
    ) == 1


def test_ai_thread_requires_report_and_is_private_to_owner(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)
    assert client.post("/api/ai/threads", json={}).status_code == 409
    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    created = client.post(
        "/api/admin/users",
        json={"email": "other-analyst@example.com", "role": "consultant"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "other-analyst@example.com", created["temporaryPassword"])

    assert client.get(f"/api/ai/threads/{thread['id']}").status_code == 404


def test_ai_thread_history_lists_latest_owner_thread_for_report(
    tmp_path: Path,
) -> None:
    client = make_client(tmp_path)
    login(client)
    first = client.post(
        "/api/ai/threads", json={"report_id": "report-1"}
    ).json()
    client.post(
        f"/api/ai/threads/{first['id']}/messages",
        json={"content": "Что главное?"},
    )
    latest = client.post(
        "/api/ai/threads", json={"report_id": "report-1"}
    ).json()
    client.post(
        f"/api/ai/threads/{latest['id']}/messages",
        json={"content": "Где нет себестоимости?"},
    )

    response = client.get("/api/ai/threads?report_id=report-1&limit=1")

    assert response.status_code == 200
    items = response.json()["items"]
    assert [item["id"] for item in items] == [latest["id"]]
    assert [message["role"] for message in items[0]["messages"]] == [
        "user",
        "assistant",
    ]
    assert items[0]["events"]

    created = client.post(
        "/api/admin/users",
        json={"email": "history-other@example.com", "role": "consultant"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "history-other@example.com", created["temporaryPassword"])
    assert client.get("/api/ai/threads?report_id=report-1&limit=1").json() == {
        "items": []
    }


def test_ai_thread_rejects_report_client_scope_mismatch(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    response = client.post(
        "/api/ai/threads",
        json={"report_id": "report-1", "client_id": "another-client"},
    )

    assert response.status_code == 409


def test_chatkit_custom_server_uses_existing_private_ai_store(tmp_path: Path) -> None:
    client = make_client(
        tmp_path,
        settings_overrides={"chatkit_enabled": True},
    )
    login(client)
    config = client.get("/api/ai/config").json()
    assert config["transport"] == "chatkit"
    assert config["chatkitEnabled"] is True
    assert "chatkitDomainKey" not in config
    response = client.post(
        "/api/chatkit",
        json={
            "type": "threads.create",
            "metadata": {
                "reportId": "report-1",
                "scope": {"preset": "losses"},
            },
            "params": {
                "input": {
                    "content": [
                        {"type": "input_text", "text": "Что главное по отчету?"}
                    ],
                    "attachments": [],
                    "inference_options": {},
                }
            },
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/event-stream")
    assert "thread.created" in response.text
    assert "assistant_message" in response.text
    assert "Убыточных строк" in response.text

    listed = client.post(
        "/api/chatkit",
        json={
            "type": "threads.list",
            "metadata": {},
            "params": {"limit": 10, "order": "desc"},
        },
    )
    assert listed.status_code == 200
    threads = listed.json()["data"]
    assert len(threads) == 1
    stored = client.get(f"/api/ai/threads/{threads[0]['id']}").json()
    assert stored["reportId"] == "report-1"
    assert stored["scope"] == {"preset": "losses"}
    assert stored["scopeHash"]
    assistant_messages = [
        item for item in stored["messages"] if item["role"] == "assistant"
    ]
    assert assistant_messages[-1]["citations"][0]["reportId"] == "report-1"


def test_chatkit_protocol_is_disabled_by_default(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    response = client.post(
        "/api/chatkit",
        json={"type": "threads.list", "metadata": {}, "params": {}},
    )

    assert response.status_code == 404


def test_ai_fallback_reason_is_hidden_from_client_role(
    tmp_path: Path,
    monkeypatch,
) -> None:
    def fake_openai_answer(self, db, user, thread, report, question):
        return None, "BadRequestError"

    monkeypatch.setattr(AiAnalyst, "_openai_answer", fake_openai_answer)
    client = make_client(tmp_path, settings_overrides={"openai_api_key": "test-key"})
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "ai-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "ai-client@example.com", created["temporaryPassword"])

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Что главное?"},
    ).json()

    done_events = [
        item for item in answer["events"] if item["type"] == "assistant_done"
    ]
    assert done_events[-1]["payload"]["answerSource"] == "fallback"
    assert "fallbackReason" not in done_events[-1]["payload"]
    assert "BadRequestError" not in str(answer)


def test_ai_explicit_onec_refresh_creates_new_report(tmp_path: Path) -> None:
    fake_service = FakeAutoRefreshService(tmp_path / "reports" / "auto-refresh.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=fake_service,
    )
    login(client)

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Дозагрузи 1С себестоимость и пересобери отчет"},
    ).json()

    assistant_messages = [
        item["content"] for item in answer["messages"] if item["role"] == "assistant"
    ]
    assert "report-1-refresh" in assistant_messages[-1]
    assert client.get("/api/reports/report-1-refresh/summary").status_code == 200
    titles = {item["title"] for item in answer["events"]}
    assert "Нашел нехватку 1С-данных" in titles
    assert "Дозагружаю 1С без изменения данных" in titles
    assert "Пересчитываю отчет" in titles
    assert "Создан новый отчет" in titles

    audit = client.get("/api/admin/audit").json()["items"]
    assert any(item["action"] == "ai_onec_auto_refresh_completed" for item in audit)


def test_ai_openai_failure_does_not_repeat_completed_refresh(
    tmp_path: Path,
    monkeypatch,
) -> None:
    import openai

    class CountingRefresh(FakeAutoRefreshService):
        calls = 0

        def run(self, *args, **kwargs):
            self.calls += 1
            return super().run(*args, **kwargs)

    class RefreshCall:
        type = "function_call"
        name = "refresh_onec_and_rebuild_report"
        call_id = "call-refresh"
        arguments = '{"reason":"Дозагрузить себестоимость"}'

    class FirstResponse:
        output = [RefreshCall()]
        output_text = ""

    class FakeResponses:
        calls = 0

        def create(self, **_kwargs):
            self.calls += 1
            if self.calls == 1:
                return FirstResponse()
            raise RuntimeError("model unavailable after tool")

    class FakeOpenAI:
        def __init__(self, **_kwargs):
            self.responses = FakeResponses()

    monkeypatch.setattr(openai, "OpenAI", FakeOpenAI)
    refresh = CountingRefresh(tmp_path / "reports" / "once.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={
            "openai_api_key": "test-key",
            "source_refresh_enabled": True,
        },
        auto_refresh_service=refresh,
    )
    login(client)
    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()

    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Дозагрузи 1С себестоимость и пересобери отчет"},
    )

    assert answer.status_code == 200
    assert refresh.calls == 1
    assert client.get("/api/reports/report-1-refresh/summary").status_code == 200


def test_ai_reports_worker_launch_failure_without_changing_report(
    tmp_path: Path,
) -> None:
    class _UnavailableAutoRefresh:
        def run(self, *_args, **_kwargs):
            raise AutoRefreshUnavailableError("Не удалось запустить обновление.")

    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=_UnavailableAutoRefresh(),
    )
    login(client)

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Дозагрузи 1С себестоимость и пересобери отчет"},
    ).json()

    assert "Не удалось запустить обновление" in str(answer)
    assert client.get("/api/reports/report-1-refresh/summary").status_code == 404


def test_ai_does_not_refresh_for_general_question(tmp_path: Path) -> None:
    fake_service = FakeAutoRefreshService(tmp_path / "reports" / "auto-refresh.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=fake_service,
    )
    login(client)

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Что главное по отчету?"},
    )

    assert client.get("/api/reports/report-1-refresh/summary").status_code == 404


def test_ai_client_role_cannot_trigger_onec_refresh(tmp_path: Path) -> None:
    fake_service = FakeAutoRefreshService(tmp_path / "reports" / "auto-refresh.xlsx")
    client = make_client(
        tmp_path,
        settings_overrides={"source_refresh_enabled": True},
        auto_refresh_service=fake_service,
    )
    login(client)
    created = client.post(
        "/api/admin/users",
        json={"email": "ai-refresh-client@example.com", "role": "client"},
    ).json()
    client.post("/api/auth/logout")
    login_as(client, "ai-refresh-client@example.com", created["temporaryPassword"])

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    answer = client.post(
        f"/api/ai/threads/{thread['id']}/messages",
        json={"content": "Дозагрузи 1С себестоимость и пересобери отчет"},
    ).json()

    assistant_messages = [
        item["content"] for item in answer["messages"] if item["role"] == "assistant"
    ]
    assert "нужна проверка консультанта" in assistant_messages[-1].lower()
    assert client.get("/api/reports/report-1-refresh/summary").status_code == 404


def test_ai_stream_returns_safe_events_and_final_answer(tmp_path: Path) -> None:
    client = make_client(tmp_path)
    login(client)

    thread = client.post("/api/ai/threads", json={"report_id": "report-1"}).json()
    with client.stream(
        "POST",
        f"/api/ai/threads/{thread['id']}/messages/stream",
        json={"content": "Покажи убыточные SKU"},
    ) as response:
        assert response.status_code == 200
        body = "".join(response.iter_text())

    assert "event: status" in body
    assert "event: tool_completed" in body
    assert "event: answer_source" in body
    assert "event: final" in body
    assert "answerSource" in body
    assert "Убыточных строк" in body

    events = client.get(f"/api/ai/threads/{thread['id']}/events").json()["items"]
    assert any(item["title"] == "Разбираю убыточность" for item in events)
    assert not any("input_payload" in item.get("payload", {}) for item in events)

    with client.stream(
        "POST",
        f"/api/ai/threads/{thread['id']}/messages/stream",
        json={"content": "Повтори главный вывод"},
    ) as response:
        second_body = "".join(response.iter_text())

    assert second_body.count('"title": "Ответ готов"') == 1


def test_client_company_alias_merge_repairs_report_scope_and_is_idempotent(
    tmp_path: Path,
) -> None:
    engine = make_engine(f"sqlite:///{tmp_path / 'web.sqlite3'}")
    init_db(engine)
    session_factory = make_session_factory(engine)
    with session_factory() as db:
        report = import_dashboard_payload(
            db,
            ready_payload(),
            tenant_id="shumeyko",
            tenant_name="Шумейко и Партнеры",
            report_id="company-alias-draft",
            publication_status="draft",
            publish=False,
            source_snapshot_set_id="company-alias-snapshot",
        )
        report_rows = list(
            db.scalars(
                select(repository.ReportUnitRow).where(
                repository.ReportUnitRow.report_run_id == report.id
                )
            )
        )
        row = report_rows[0] if report_rows else None
        assert row is not None
        cabinet = db.get(WbCabinet, row.wb_cabinet_id)
        assert cabinet is not None and cabinet.client_company_id
        canonical = db.get(repository.ClientCompany, cabinet.client_company_id)
        assert canonical is not None
        duplicate = repository.ensure_client_company(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            display_name="Галустов Рафаэль Рудольфович",
        )
        assert duplicate is not None and duplicate.id != canonical.id

        db.execute(text("DROP INDEX uq_client_companies_active_onec_organization"))
        canonical.display_name = "Галустов"
        canonical.onec_organization_id = "ORG-GALUSTOV"
        duplicate.onec_organization_id = "ORG-GALUSTOV"
        resolved_ids = repository._row_entity_ids(
            db,
            report,
            {
                "clientId": report.client_id,
                "clientCompanyId": duplicate.id,
                "wbCabinetId": cabinet.id,
                "organization": duplicate.display_name,
                "cabinet": cabinet.display_name,
            },
        )
        assert resolved_ids["client_company_id"] == canonical.id
        different_company = repository.ensure_client_company(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            display_name="Другая организация",
        )
        assert different_company is not None
        different_company.onec_organization_id = "ORG-OTHER"
        with pytest.raises(ValueError, match="does not match"):
            repository._row_entity_ids(
                db,
                report,
                {
                    "clientId": report.client_id,
                    "clientCompanyId": different_company.id,
                    "wbCabinetId": cabinet.id,
                    "organization": different_company.display_name,
                    "cabinet": cabinet.display_name,
                },
            )
        affected_rows = db.query(repository.ReportUnitRow).filter(
            repository.ReportUnitRow.report_run_id == report.id
        ).update(
            {
                repository.ReportUnitRow.client_company_id: duplicate.id,
                repository.ReportUnitRow.wb_cabinet_id: cabinet.id,
                repository.ReportUnitRow.tax_profile_source: "1C:test",
            },
            synchronize_session=False,
        )

        refresh = repository.create_source_refresh_run(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
            mode="onec-only",
            credential_source="tenant",
            dry_run=False,
            snapshot_set_id="company-alias-tax",
            period_start=report.period_start,
            period_end=report.period_end,
            reason="company alias test",
        )
        tax_collection = repository.add_source_refresh_collection(
            db,
            refresh,
            source_type="onec_tax_profiles",
            source_label="Налоговые профили",
            required=False,
            status="loaded",
            snapshot_hash="company-alias-tax-hash",
            row_count=1,
            payload={
                "profileCount": 1,
                "missingProfileCount": 0,
                "unconfirmedProfileCount": 0,
            },
        )
        db.add(
            OrganizationTaxProfile(
                id="company-alias-tax-profile",
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                client_company_id=canonical.id,
                organization_id="ORG-GALUSTOV",
                tax_system="ОСНО",
                vat_rate=Decimal("22"),
                vat_mode="included",
                vat_deduction_mode="allowed",
                revenue_tax_rate=Decimal("0"),
                income_tax_kind="ip_ndfl_progressive",
                valid_from=date(2026, 1, 1),
                valid_to=None,
                source="1C:test",
                source_refresh_run_id=refresh.id,
                source_snapshot_hash="company-alias-tax-hash",
                methodology_version="marketplace-tax-profile-v3",
                status="active",
                created_at=repository.security.utcnow(),
            )
        )
        user = upsert_user(
            db,
            email="company-alias-admin@example.com",
            password="secret",
            tenant_id=report.tenant_id,
            role="admin",
        )
        db.flush()
        revenue_before = sum(
            db.scalars(
                select(repository.ReportUnitRow.revenue).where(
                    repository.ReportUnitRow.report_run_id == report.id
                )
            )
        )

        readiness = repository.report_readiness_payload(db, report)
        assert "company_cabinet_mismatch" in {
            item["code"] for item in readiness["blockingReasons"]
        }
        before_sync = repository.tax_profile_sync_payload(
            db,
            report,
            include_staff_details=True,
        )
        assert before_sync["scopeStatus"] == "scope_mismatch"
        assert before_sync["reportStatus"] == "scope_mismatch"
        with pytest.raises(repository.ReportPublicationBlocked):
            repository.publish_report_with_tasks(
                db,
                report,
                user=user,
                reason="structural mismatch cannot be overridden",
            )

        counts = repository.dedupe_client_companies(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
        )
        repository.ensure_client_company_identity_index(db)
        db.flush()

        assert counts["duplicate_groups"] == 1
        assert counts["merged_companies"] == 1
        assert counts["report_unit_rows"] == affected_rows
        assert db.get(repository.ClientCompany, duplicate.id) is None
        assert canonical.display_name == "Галустов Рафаэль Рудольфович"
        repaired_row = db.scalar(
            select(repository.ReportUnitRow).where(
                repository.ReportUnitRow.id == row.id
            )
            .execution_options(populate_existing=True)
        )
        assert repaired_row is not None
        assert repaired_row.client_company_id == canonical.id
        repaired_revenue = sum(
            db.scalars(
                select(repository.ReportUnitRow.revenue).where(
                    repository.ReportUnitRow.report_run_id == report.id
                )
            )
        )
        assert repaired_revenue == revenue_before
        aliases = set(
            db.scalars(
                select(repository.ClientCompanyAlias.display_name).where(
                    repository.ClientCompanyAlias.client_company_id == canonical.id
                )
            )
        )
        assert {"Галустов", "Галустов Рафаэль Рудольфович"}.issubset(aliases)
        assert repository._report_company_cabinet_mismatch_count(db, report) == 0

        after_sync = repository.tax_profile_sync_payload(
            db,
            report,
            include_staff_details=True,
        )
        assert after_sync["scopeStatus"] == "ready"
        assert after_sync["reportStatus"] == "stale"

        db.add(
            SourceLoad(
                tenant_id=report.tenant_id,
                client_id=report.client_id,
                wb_cabinet_id="",
                report_run_id=report.id,
                source_refresh_run_id=refresh.id,
                required=False,
                publication_required=False,
                source_type="onec_tax_profiles",
                source_label="Налоговые профили",
                status="loaded",
                snapshot_hash=tax_collection.snapshot_hash,
                row_count=1,
                loaded_at=repository.security.utcnow(),
            )
        )
        db.flush()
        applied = repository.tax_profile_sync_payload(
            db,
            report,
            include_staff_details=True,
        )
        assert applied["reportStatus"] == "applied"
        assert repository.dedupe_client_companies(
            db,
            tenant_id=report.tenant_id,
            client_id=report.client_id,
        )["duplicate_groups"] == 0


def test_ozon_revenue_control_detects_missing_primary_document_in_onec() -> None:
    control = repository._ozon_revenue_document_control_payload(
        pnl={
            "onecOzon": {"documentRows": []},
            "periodFilter": {
                "periodStart": "2026-05-01",
                "periodEnd": "2026-05-31",
            },
        },
        ozon_buyouts={"rows": []},
        ozon_commissioner_amount=Decimal("1000"),
        onec_commissioner_amount=None,
        commissioner_delta=None,
        buyout_amount=Decimal("0"),
        onec_buyout_amount=Decimal("0"),
        buyout_delta=Decimal("0"),
    )

    assert control["status"] == "review"
    assert control["issueCount"] == 1
    assert control["missingPrimaryCount"] == 1
    assert control["rows"][0]["status"] == "missing_in_1c"
    assert "первичный документ 1C не найден" in control["rows"][0]["problem"]


def test_ozon_commissioner_control_detects_wrong_onec_document_date() -> None:
    rows = repository._ozon_commissioner_document_rows(
        [
            {
                "Date": "2026-06-01T00:00:00",
                "Number": "НФНФ-000033",
                "Posted": True,
                "Комментарий": (
                    "ОЗОН Отчет комиссионера № 16 567 305 "
                    "от 01.05.2026 0:00:00 по 31.05.2026 0:00:00"
                ),
                "Запасы": [{"Всего": "1000"}],
                "ЗапасыВозвраты": [],
            }
        ],
        period_start=date(2026, 5, 1),
        period_end=date(2026, 5, 31),
    )

    assert rows == [
        {
            "documentNumber": "НФНФ-000033",
            "documentDate": "2026-06-01",
            "reportNumber": "16567305",
            "periodFrom": "2026-05-01",
            "periodTo": "2026-05-31",
            "amount": 1000.0,
            "posted": True,
            "status": "wrong_date",
        }
    ]
    assert (
        repository._ozon_commissioner_control_status(
            ozon_amount=Decimal("1000"),
            onec_amount=None,
            delta=None,
            documents=rows,
        )
        == "wrong_date"
    )
